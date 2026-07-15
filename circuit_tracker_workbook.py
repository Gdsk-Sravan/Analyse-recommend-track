"""Build `circuit_tracker_workbook.xlsx` — dedicated 2-sheet workbook.

Split out from `tracking_workbook_job.py` on 2026-07-15 per user directive:
"Send the tracking workbook AND a dedicated circuit workbook as 2 separate
Telegram attachments" (was: 3 sends total, one of them a duplicate).

Sheets:
    1. CIRCUIT_TRACKER         — milestone view (D+1, D+2, D+3, D+5, D+7,
                                  D+10, D+15, D+20, D+25, D+30) + summary strip
    2. CIRCUIT_TRACKER_DETAIL  — full daily view (D+1 through D+30)

Reads: results/circuit_tracker.json  (via circuit_tracker.load_all_tracks())
Writes: circuit_tracker_workbook.xlsx (path configurable via CLI --path)

FRESH_START behaviour: the JSON store, its backups, and this workbook are
all listed in main.py's `_fresh_delete_files` so a `FRESH_START=true` run
wipes them cleanly and rebuilds an empty workbook (just headers + summary).

Usage:
    python circuit_tracker_workbook.py
    python circuit_tracker_workbook.py --path /custom/out.xlsx
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ─── Styling (mirror the main workbook look-and-feel) ──────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
CENTER      = Alignment(horizontal="center", vertical="center")

_CIRCUIT_MILESTONES: Tuple[int, ...] = (1, 2, 3, 5, 7, 10, 15, 20, 25, 30)
_CIRCUIT_DETAIL_DAYS: Tuple[int, ...] = tuple(range(1, 31))

_CIRCUIT_POS_FILL       = PatternFill("solid", fgColor="D4EDDA")   # light green
_CIRCUIT_NEG_FILL       = PatternFill("solid", fgColor="F8CBAD")   # light red
_CIRCUIT_ACTIVE_FILL    = PatternFill("solid", fgColor="FFF2CC")   # pale yellow
_CIRCUIT_RETIRED_FILL   = PatternFill("solid", fgColor="E7E6E6")   # grey
_CIRCUIT_SEPARATOR_FILL = PatternFill("solid", fgColor="D9D9D9")

# Where to write by default (next to main.py).
_HERE = Path(__file__).resolve().parent
DEFAULT_OUTPUT_PATH = _HERE / "circuit_tracker_workbook.xlsx"


def _circuit_ret_fill(val: Optional[float]) -> Optional[PatternFill]:
    if val is None:
        return None
    try:
        v = float(val)
    except (TypeError, ValueError):
        return None
    if v > 0:
        return _CIRCUIT_POS_FILL
    if v < 0:
        return _CIRCUIT_NEG_FILL
    return None


def _write_summary(ws: Worksheet,
                   stats: Optional[Dict[str, Any]],
                   tracks: List[Dict[str, Any]],
                   day_offsets: Tuple[int, ...]) -> None:
    """Render one circuit sheet — same layout for milestone + detail views."""
    # ── Row 1: Title ────────────────────────────────────────────────
    view_name = "milestone view" if len(day_offsets) <= 10 else "daily view"
    title = ws.cell(row=1, column=1,
                    value=f"CIRCUIT TRACKER — 30-day forward-return study "
                          f"({view_name})")
    title.font = Font(bold=True, size=13, color="1F4E78")

    # ── Rows 2-4: Summary strip ─────────────────────────────────────
    ws.cell(row=2, column=1, value="Direction").font = Font(bold=True)
    ws.cell(row=2, column=2, value="Active").font = Font(bold=True)
    ws.cell(row=2, column=3, value="Retired").font = Font(bold=True)
    ws.cell(row=2, column=4, value="Avg D+30").font = Font(bold=True)
    ws.cell(row=2, column=5, value="Avg Latest").font = Font(bold=True)
    ws.cell(row=2, column=6, value="% Positive (retired)").font = Font(bold=True)

    if stats:
        pos = stats.get("pos", {})
        neg = stats.get("neg", {})
        ws.cell(row=3, column=1, value="🟢 POS (+15% or more)")
        ws.cell(row=3, column=2, value=pos.get("active"))
        ws.cell(row=3, column=3, value=pos.get("retired"))
        ws.cell(row=3, column=4, value=pos.get("avg_d30"))
        ws.cell(row=3, column=5, value=pos.get("avg_latest"))
        ws.cell(row=3, column=6, value=pos.get("pct_positive"))
        ws.cell(row=4, column=1, value="🔴 NEG (-15% or worse)")
        ws.cell(row=4, column=2, value=neg.get("active"))
        ws.cell(row=4, column=3, value=neg.get("retired"))
        ws.cell(row=4, column=4, value=neg.get("avg_d30"))
        ws.cell(row=4, column=5, value=neg.get("avg_latest"))
        ws.cell(row=4, column=6, value=neg.get("pct_positive"))
        for r in (3, 4):
            for c in (4, 5, 6):
                cell = ws.cell(row=r, column=c)
                if cell.value is not None:
                    cell.number_format = "0.00\"%\""
    else:
        ws.cell(row=3, column=1,
                value="(no data yet — first scheduled run will populate)")

    # ── Row 7: Column headers ───────────────────────────────────────
    header_row = 7
    fixed_cols = [
        "Track ID", "Symbol", "Circuit Date", "Direction", "Circuit Move %",
    ]
    day_cols = [f"D+{d}" for d in day_offsets]
    trailing_cols = [
        "Max Up %", "Max Down %", "Days Tracked", "Status", "Verdict", "Retired Reason",
    ]
    all_headers = fixed_cols + day_cols + trailing_cols
    for col_idx, hdr in enumerate(all_headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    ws.freeze_panes = ws.cell(row=header_row + 1, column=3)  # freeze after Symbol col

    # Column widths.
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 13
    for i in range(len(day_cols)):
        ws.column_dimensions[get_column_letter(6 + i)].width = 8
    trailing_start = 6 + len(day_cols)
    ws.column_dimensions[get_column_letter(trailing_start)].width = 10
    ws.column_dimensions[get_column_letter(trailing_start + 1)].width = 11
    ws.column_dimensions[get_column_letter(trailing_start + 2)].width = 8
    ws.column_dimensions[get_column_letter(trailing_start + 3)].width = 20
    ws.column_dimensions[get_column_letter(trailing_start + 4)].width = 15
    ws.column_dimensions[get_column_letter(trailing_start + 5)].width = 32

    # ── Rows 8+: track rows ─────────────────────────────────────────
    if not tracks:
        ws.cell(row=header_row + 1, column=1,
                value="(no circuit events tracked yet — first scheduled run "
                      "with circuit-limit skips will seed this sheet)")
        return

    # Terminal-status set for retirement classification.
    try:
        import circuit_tracker as _ct  # type: ignore
        _terminal = _ct.TERMINAL_STATUSES
        _active   = _ct.STATUS_ACTIVE
    except Exception:
        _terminal = {"RETIRED_30DAY", "RETIRED_NODATA",
                     "RETIRED_BIG_WIN", "RETIRED_BIG_LOSS"}
        _active   = "ACTIVE"

    active_tracks  = [t for t in tracks if t.get("status") == _active]
    retired_tracks = [t for t in tracks if t.get("status") in _terminal]

    row = header_row + 1

    def _write_track_row(t: Dict[str, Any], row_idx: int, is_retired: bool) -> None:
        base_fill = _CIRCUIT_RETIRED_FILL if is_retired else None

        ws.cell(row=row_idx, column=1, value=t.get("track_id"))
        ws.cell(row=row_idx, column=2, value=t.get("symbol"))
        ws.cell(row=row_idx, column=3, value=t.get("circuit_date"))
        dir_cell = ws.cell(row=row_idx, column=4,
                            value=("🟢 POS" if t.get("direction") == "POS" else "🔴 NEG"))
        dir_cell.fill = (_CIRCUIT_POS_FILL if t.get("direction") == "POS"
                         else _CIRCUIT_NEG_FILL)
        dir_cell.alignment = CENTER
        cm = ws.cell(row=row_idx, column=5, value=t.get("circuit_move"))
        cm.number_format = "+0.00\"%\";-0.00\"%\""

        # Day-offset returns
        returns = t.get("returns") or {}
        for i, day in enumerate(day_offsets, start=6):
            val = returns.get(str(day))
            cell = ws.cell(row=row_idx, column=i, value=val)
            if val is not None:
                cell.number_format = "+0.00\"%\";-0.00\"%\""
                fill = _circuit_ret_fill(val)
                if fill is not None:
                    cell.fill = fill

        # Trailing cols
        col = 6 + len(day_offsets)
        mu = ws.cell(row=row_idx, column=col,     value=t.get("max_up_pct"))
        md = ws.cell(row=row_idx, column=col + 1, value=t.get("max_down_pct"))
        mu.number_format = "+0.00\"%\";-0.00\"%\""
        md.number_format = "+0.00\"%\";-0.00\"%\""
        ws.cell(row=row_idx, column=col + 2, value=t.get("days_tracked"))
        ws.cell(row=row_idx, column=col + 3, value=t.get("status"))
        v_cell = ws.cell(row=row_idx, column=col + 4, value=t.get("verdict"))
        ws.cell(row=row_idx, column=col + 5, value=t.get("retired_reason"))

        verdict = str(t.get("verdict") or "")
        if verdict in ("STILL_UP", "HOLDING", "BOUNCED_HARD", "BOUNCED"):
            v_cell.fill = _CIRCUIT_POS_FILL
        elif verdict in ("CRASHED", "KEPT_FALLING", "FADING", "STILL_DOWN"):
            v_cell.fill = _CIRCUIT_NEG_FILL
        elif verdict == "TOO_EARLY":
            v_cell.fill = _CIRCUIT_ACTIVE_FILL

        if base_fill is not None:
            for c in (1, 2, 3):
                ws.cell(row=row_idx, column=c).fill = base_fill

    # Write ACTIVE tracks.
    for t in active_tracks:
        _write_track_row(t, row, is_retired=False)
        row += 1

    # Separator row before retired tracks.
    if retired_tracks:
        sep_row = row
        for c in range(1, len(all_headers) + 1):
            cell = ws.cell(row=sep_row, column=c, value="")
            cell.fill = _CIRCUIT_SEPARATOR_FILL
        ws.cell(row=sep_row, column=1,
                value="— RETIRED TRACKS —").font = Font(bold=True, italic=True)
        row += 1
        for t in retired_tracks:
            _write_track_row(t, row, is_retired=True)
            row += 1

    # Enable auto-filter.
    last_col_letter = get_column_letter(len(all_headers))
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{row - 1}"


def build_workbook(output_path: Path = DEFAULT_OUTPUT_PATH) -> Path:
    """Build the standalone circuit-tracker workbook.

    Reads state via circuit_tracker.load_all_tracks() + summary_stats().
    Fully non-fatal: if the module is missing or the store is empty, the
    workbook is still created with header rows and an "(empty)" note so
    the file always exists on disk for Telegram delivery.
    """
    output_path = Path(output_path)

    # Ensure circuit_tracker is importable — module lives next to this file.
    _HERE_STR = str(_HERE)
    if _HERE_STR not in sys.path:
        sys.path.insert(0, _HERE_STR)

    tracks: List[Dict[str, Any]] = []
    stats: Optional[Dict[str, Any]] = None
    try:
        import circuit_tracker as _ct  # type: ignore
        tracks = _ct.load_all_tracks()
        stats  = _ct.summary_stats()
        print(f"[circuit_tracker_workbook] loaded {len(tracks)} tracks")
    except Exception as e:
        print(f"[circuit_tracker_workbook] circuit_tracker not available "
              f"({e}) — creating empty workbook")

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    _write_summary(wb.create_sheet("CIRCUIT_TRACKER"),
                   stats, tracks, _CIRCUIT_MILESTONES)
    _write_summary(wb.create_sheet("CIRCUIT_TRACKER_DETAIL"),
                   stats, tracks, _CIRCUIT_DETAIL_DAYS)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"[circuit_tracker_workbook] wrote {output_path.name} "
          f"({len(tracks)} tracks, 2 sheets)")
    return output_path


def _main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="Build the standalone circuit-tracker Excel workbook.")
    parser.add_argument(
        "--path", type=Path, default=DEFAULT_OUTPUT_PATH,
        help=f"Output .xlsx path (default: {DEFAULT_OUTPUT_PATH.name})")
    args = parser.parse_args(argv)
    build_workbook(args.path)
    return 0


if __name__ == "__main__":
    raise SystemExit(_main())
