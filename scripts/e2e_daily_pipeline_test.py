"""End-to-end test of the observation-dataset daily pipeline.

Mirrors the exact sequence that main.py runs in its "Observation-Dataset
redesign" block (right after shadow_master_job.run_scan_and_update).

Purpose:
 1. Confirm every step is wired correctly.
 2. Simulate a 3-day rolling update using MOCK data so we can verify:
    - Day 1 → workbook built, daily_snapshots.jsonl has row per (symbol, bucket).
    - Day 2 → prices refresh; historical Day-1 rows show TODAY's price.
    - Day 3 → hits (T1/T2/STOP) recorded; DONE + WEEKLY_SUMMARY correct.
 3. Verify no shadow_master.xlsx is required at any step.
 4. Verify tracking_workbook.xlsx would be sent to Telegram (DRY-RUN — we
    monkey-patch send_document so no real HTTP call is made).

Usage:
    python scripts/e2e_daily_pipeline_test.py

Exits 0 on success, non-zero + traceback on failure.
"""

from __future__ import annotations

import os
import sys
import shutil
import traceback
from datetime import datetime
from pathlib import Path

# Force UTF-8 on Windows (default cp1252 can't render arrows/checks).
if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

HERE   = Path(__file__).resolve().parent
ROOT   = HERE.parent
sys.path.insert(0, str(ROOT))

# ── Isolate the test to a scratch results dir ────────────────────────────────
SCRATCH_ROOT = ROOT / "results_e2e_test"
if SCRATCH_ROOT.exists():
    shutil.rmtree(SCRATCH_ROOT, ignore_errors=True)
SCRATCH_ROOT.mkdir(parents=True, exist_ok=True)

# Redirect all pipeline artefacts into the scratch dir.
os.environ["TRACKING_STORE_PATH"]     = str(SCRATCH_ROOT / "tracking_store.json")
os.environ["DAILY_SNAPSHOTS_PATH"]    = str(SCRATCH_ROOT / "daily_snapshots.jsonl")
os.environ["TRACKING_EVENTS_PATH"]    = str(SCRATCH_ROOT / "tracking_events.jsonl")
os.environ["WEEKLY_REVIEW_HISTORY_PATH"] = str(SCRATCH_ROOT / "weekly_review_history.jsonl")

WORKBOOK_PATH = SCRATCH_ROOT / "tracking_workbook_e2e.xlsx"

# ── Telegram DRY-RUN ─────────────────────────────────────────────────────────
_TELEGRAM_CALLS: list[dict] = []

def _dryrun_send_document(file_path: str, caption: str = "", **kw) -> int:
    """Monkey-patch replacement for scripts.notify_telegram_document.send_document."""
    _TELEGRAM_CALLS.append({
        "file_path": file_path,
        "file_exists": os.path.exists(file_path),
        "file_size_bytes": os.path.getsize(file_path) if os.path.exists(file_path) else 0,
        "caption": caption,
    })
    return 200

# Install the monkey-patch BEFORE importing the pipeline modules.
import scripts.notify_telegram_document as _ntd
_ntd.send_document = _dryrun_send_document


# ── Import pipeline modules (after env vars set + monkeypatch installed) ────
# NOTE: TrackingStore reads its default path at load-time from a class const,
# not from the env var, so we override its path explicitly per step.
from tracking_store import (
    TrackingStore,
    STATUS_ACTIVE, STATUS_T1_HIT, STATUS_T2_HIT,
    STATUS_STOPPED, STATUS_STOPPED_AFTER_T1,
)
import daily_snapshot_job as dsj
import tracking_workbook_job as twj


# ── Force pipeline modules to use scratch paths ─────────────────────────────
dsj.DAILY_SNAPSHOTS_PATH   = SCRATCH_ROOT / "daily_snapshots.jsonl"
twj.DAILY_SNAPSHOTS_PATH   = SCRATCH_ROOT / "daily_snapshots.jsonl"
twj.DEFAULT_WEEKLY_HISTORY_PATH = SCRATCH_ROOT / "weekly_review_history.jsonl"


# ── Mock universe: 6 stocks across all buckets ──────────────────────────────
MOCK_STOCKS = [
    # (symbol, entry_stage, entry_bucket, setup_type, sector, entry_price, t1, t2, stop, confidence)
    ("MRF",        "BUY",              "BUY",           "BREAKOUT", "AUTO",     100000, 105000, 110000, 97000, 82.5),
    ("TCS",        "BUY",              "BUY",           "MOMENTUM", "IT",         3400,   3510,   3620,   3300, 78.0),
    ("HDFC",       "BUY",              "BUY",           "REVERSAL", "BANK",       1600,   1660,   1720,   1560, 71.0),
    ("BAJFINANCE", "NEAR_MISS_FADING", "NEAR_MISS",     None,       "FINANCE",    7200,   7450,   7700,   7000, 62.0),
    ("LT",         "DEVELOPING",       "DEVELOPING",    None,       "INFRA",      3500,   3625,   3750,   3400, 55.0),
    ("RELIANCE",   "REJECTED",         "REJECTED",      "BREAKOUT", "ENERGY",     2500,   2600,   2700,   2400, 45.0),
]


def _seed_day(store: TrackingStore, run_date: str) -> None:
    """Seed the store with an entry per mock stock on the given run_date."""
    for (sym, stage, bucket, setup, sector, entry, t1, t2, stop, conf) in MOCK_STOCKS:
        entry_snapshot = {
            "stage": stage,
            "bucket": bucket,
            "origin": "e2e_mock",
            "setup_type": setup,
            "sector": sector,
            "regime": "TRENDING",
            "reference_entry_price": float(entry),
            "t1_price": float(t1),
            "t2_price": float(t2),
            "stop_price": float(stop),
            "confidence": conf,
            "tq": conf - 5,
            "rr_ratio": round((t1 - entry) / (entry - stop), 2),
            "opportunity_score": conf - 3,
            "factors": {},
            "fail_reasons": ["CONF_FAIL(low)"] if stage == "REJECTED" else [],
        }
        store.upsert(
            symbol=sym,
            stage=stage,
            entry_snapshot=entry_snapshot,
            origin="e2e_mock",
            as_of=run_date,
            current_snapshot={"stage": stage, "close": float(entry), "price": float(entry)},
        )


def _update_prices(store: TrackingStore, run_date: str, price_map: dict) -> None:
    """Refresh current_price on each active record with the mock price."""
    store.update_prices(
        prices={sym: {"close": float(px), "high": float(px) * 1.001, "low": float(px) * 0.999}
                for sym, px in price_map.items()},
        as_of=run_date,
    )


def _run_daily_cycle(run_date: str, label: str, price_map: dict) -> None:
    """One full simulated main.py daily cycle for the observation pipeline."""
    print(f"\n{'=' * 78}\n{label}  ({run_date})\n{'=' * 78}")

    # ── STEP 1: load / seed the store ───────────────────────────────────────
    store_path = Path(os.environ["TRACKING_STORE_PATH"])
    events_path = SCRATCH_ROOT / "tracking_events.jsonl"
    if store_path.exists():
        store = TrackingStore.load(store_path, events_path)
    else:
        store = TrackingStore(store_path=store_path, events_path=events_path)
        _seed_day(store, run_date)
        store.save()
    print(f"  [1/4] store loaded: {len(store.records)} records")

    # ── STEP 2: refresh prices (mimics tracker_job / update_prices) ────────
    _update_prices(store, run_date, price_map)
    store.save()
    _stats = store.stats()
    print(f"  [2/4] prices refreshed → stats: {_stats}")

    # ── STEP 3: append daily snapshots ─────────────────────────────────────
    written = dsj.append_from_store(store, run_date=run_date,
                                     snapshots_path=SCRATCH_ROOT / "daily_snapshots.jsonl")
    n_appended = len(written) if written else 0
    print(f"  [3/4] daily_snapshot_job appended {n_appended} rows")

    # ── STEP 4: build workbook + Telegram send ─────────────────────────────
    snapshots_path = SCRATCH_ROOT / "daily_snapshots.jsonl"
    weekly_path = SCRATCH_ROOT / "weekly_review_history.jsonl"
    twj.build_workbook(store, WORKBOOK_PATH,
                       weekly_history_path=weekly_path,
                       snapshots_path=snapshots_path)
    print(f"  [4/4] workbook rebuilt at {WORKBOOK_PATH.name}")

    # Simulate the main.py Telegram branch:
    from scripts.notify_telegram_document import send_document
    caption = (
        f"📊 Tracking Workbook — {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
        f"Total: {_stats.get('total',0)} · Active: {_stats.get('active',0)}"
        f" · T1: {_stats.get('t1_hit_active',0)} · T2: {_stats.get('t2_hit',0)}"
        f" · Stop: {_stats.get('stopped',0)+_stats.get('stopped_after_t1',0)}\n"
        f"16 sheets · append-only daily log · live refresh on rebuild"
    )
    rc = send_document(file_path=str(WORKBOOK_PATH), caption=caption)
    print(f"  [tg]  Telegram send_document rc={rc}  (DRY-RUN, no HTTP)")


def _inspect_workbook() -> dict:
    """Return sheet/row counts for the final workbook."""
    from openpyxl import load_workbook
    wb = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    out: dict = {}
    for name in wb.sheetnames:
        ws = wb[name]
        # Count non-empty data rows.
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0] if rows else ()
        data_rows = [r for r in rows[1:] if any(c is not None for c in r)]
        out[name] = {"rows": len(data_rows), "cols": len(header)}
    wb.close()
    return out


def _check_row_keys_unique(sheet_name: str) -> tuple[bool, list[str]]:
    """Verify every Row Key in the given sheet is unique.

    The same stock legitimately fans out to multiple bucket sheets via
    `_route_to_buckets` — e.g. stage=BUY routes to both BUY and BREAKOUT.
    That is by design.  Row Key IS the unique identifier, and includes
    (tracking_id, run_date, bucket) so cross-bucket fanout stays distinct.

    ACTIVE_TRACKING has 'Source Sheet' as column A and 'Row Key' as B;
    every other daily/record sheet has 'Row Key' as column A.  Locate the
    Row Key column dynamically instead of assuming column A.
    """
    from openpyxl import load_workbook
    wb = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return (True, [])
    header = list(rows[0])
    try:
        rk_idx = header.index("Row Key")
    except ValueError:
        wb.close()
        # No Row Key column on this sheet — treat as pass (nothing to check).
        return (True, [])
    seen: dict[str, int] = {}
    dups: list[str] = []
    for r in rows[1:]:
        if not r or rk_idx >= len(r):
            continue
        rk = r[rk_idx]
        if rk is None:
            continue
        seen[rk] = seen.get(rk, 0) + 1
    wb.close()
    for rk, n in seen.items():
        if n > 1:
            dups.append(f"{rk} × {n}")
    return (len(dups) == 0, dups)


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main() -> int:
    try:
        # DAY 1 — Monday: initial seed at entry prices.
        _run_daily_cycle(
            "2026-07-13", "DAY 1  Monday",
            price_map={sym: entry for (sym, _, _, _, _, entry, *_ ) in MOCK_STOCKS},
        )

        # DAY 2 — Tuesday: MRF and TCS moving up; HDFC flat; RELIANCE flat.
        _run_daily_cycle(
            "2026-07-14", "DAY 2  Tuesday",
            price_map={"MRF": 103500, "TCS": 3480, "HDFC": 1600,
                       "BAJFINANCE": 7150, "LT": 3510, "RELIANCE": 2500},
        )

        # DAY 3 — Wednesday: MRF hits T1, TCS hits T2, LT hits stop.
        _run_daily_cycle(
            "2026-07-15", "DAY 3  Wednesday",
            price_map={"MRF": 105500, "TCS": 3650, "HDFC": 1615,
                       "BAJFINANCE": 7000, "LT": 3390, "RELIANCE": 2500},
        )

        # ── Post-run assertions ────────────────────────────────────────────
        print(f"\n{'=' * 78}\nPOST-RUN VERIFICATION\n{'=' * 78}")

        summary = _inspect_workbook()
        print("Sheet inventory:")
        for name, info in summary.items():
            print(f"  {name:<20}  rows={info['rows']:<3}  cols={info['cols']}")

        # Required sheets
        required = ["BUY", "NEAR_MISS", "DEVELOPING", "MONITOR", "REJECTED",
                    "BREAKOUT", "MOMENTUM", "PULLBACK", "REVERSAL",
                    "REJECTED_SETUP", "ACTIVE_TRACKING", "DONE",
                    "WEEKLY_SUMMARY", "WEEKLY_REVIEW", "RESEARCH", "_LEGEND"]
        missing = [s for s in required if s not in summary]
        assert not missing, f"Missing sheets: {missing}"
        print(f"\n[✓] All {len(required)} required sheets present.")

        # Row-key uniqueness for daily sheets.
        for sheet in ["BUY", "MOMENTUM", "REJECTED", "ACTIVE_TRACKING"]:
            ok, dups = _check_row_keys_unique(sheet)
            assert ok, f"{sheet} has duplicate Row Keys: {dups[:5]}"
        print("[✓] Row Key unique in BUY / MOMENTUM / REJECTED / ACTIVE_TRACKING")

        # Telegram calls
        assert len(_TELEGRAM_CALLS) == 3, f"Expected 3 Telegram sends, got {len(_TELEGRAM_CALLS)}"
        for i, call in enumerate(_TELEGRAM_CALLS, 1):
            assert call["file_exists"], f"Day {i} Telegram file missing"
            assert call["file_size_bytes"] > 5000, f"Day {i} file suspiciously small: {call['file_size_bytes']}"
            assert "Tracking Workbook" in call["caption"], f"Day {i} caption malformed"
        print(f"[✓] All 3 daily Telegram sends captured  (avg {_TELEGRAM_CALLS[-1]['file_size_bytes']} bytes)")

        # Verify NO shadow_master.xlsx involvement
        shadow_files_in_scratch = list(SCRATCH_ROOT.glob("shadow*"))
        assert not shadow_files_in_scratch, f"Unexpected shadow files: {shadow_files_in_scratch}"
        print("[✓] No shadow_master.xlsx or shadow_* artefacts written by the pipeline.")

        # WEEKLY_SUMMARY sanity
        assert summary["WEEKLY_SUMMARY"]["rows"] == 12, \
            f"WEEKLY_SUMMARY expected 12 rows, got {summary['WEEKLY_SUMMARY']['rows']}"
        print("[✓] WEEKLY_SUMMARY has 12 rows (5 stage + 5 setup + ACTIVE + DONE).")

        # DONE sheet expectations:
        #   RELIANCE  — seeded as REJECTED on Day 1, terminal from the start.
        #   TCS       — hits T2 on Day 3 (3650 ≥ 3600), terminal.
        #   LT        — hits STOP on Day 3 (3390 < 3400), terminal.
        #   MRF       — hits T1 on Day 3 but stays ACTIVE (T1_HIT_ACTIVE), NOT in DONE.
        # So DONE should have exactly 3 rows.
        done_rows = summary["DONE"]["rows"]
        assert done_rows == 3, (
            f"DONE expected 3 terminal (RELIANCE rejected, TCS T2, LT stop); "
            f"got {done_rows}"
        )
        print("[✓] DONE has 3 terminal records (RELIANCE REJECTED, TCS T2_HIT, LT STOPPED).")

        # Live refresh proof: BUY sheet should have 3 rows for MRF
        # (Monday's, Tuesday's, Wednesday's runs), all showing today's price.
        from openpyxl import load_workbook
        wb = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
        ws = wb["BUY"]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        # Find MRF rows (Symbol column is C = index 2).
        mrf_rows = [r for r in rows[1:] if r[2] == "MRF"]
        assert len(mrf_rows) == 3, f"BUY expected 3 MRF rows (3 daily snapshots); got {len(mrf_rows)}"
        # Find "Current Price" column
        try:
            cp_idx = header.index("Current Price")
            prices = [r[cp_idx] for r in mrf_rows]
            assert all(p == 105500 for p in prices), \
                f"Live refresh broken: MRF current_price varies across daily rows: {prices}"
            print(f"[✓] Live refresh: all 3 MRF rows in BUY show current_price = 105500")
        except ValueError:
            print("[!] Current Price column not found — skipping live-refresh check")
        # Verify days_active is per-snapshot correct (fixes B2 verification).
        # MRF has 3 rows with run_dates 07-13, 07-14, 07-15 and first_seen 07-13
        # → days_active should be {0, 1, 2} (order-independent).
        try:
            da_idx = header.index("Days Active")
            das = [r[da_idx] for r in mrf_rows]
            assert all(isinstance(d, (int, float)) and d >= 0 for d in das), \
                f"Days Active malformed for MRF: {das}"
            assert set(int(d) for d in das) == {0, 1, 2}, \
                f"Days Active expected {{0,1,2}} across 3 daily runs; got {sorted(das)}"
            print(f"[✓] Days Active recomputed per-snapshot: MRF rows show {sorted(das)}")
        except ValueError:
            print("[!] Days Active column not found — skipping B2 verification")
        wb.close()

        print(f"\n{'=' * 78}\n✅ E2E TEST PASSED\n{'=' * 78}")
        print(f"\nArtefacts left in: {SCRATCH_ROOT}")
        print(f"  Workbook: {WORKBOOK_PATH}")
        print(f"  Store:    {SCRATCH_ROOT / 'tracking_store.json'}")
        print(f"  Snapshots: {SCRATCH_ROOT / 'daily_snapshots.jsonl'}")
        return 0

    except AssertionError as e:
        print(f"\n❌ ASSERTION FAILED: {e}")
        traceback.print_exc()
        return 1
    except Exception:
        print(f"\n❌ UNEXPECTED ERROR:")
        traceback.print_exc()
        return 2


if __name__ == "__main__":
    sys.exit(main())
