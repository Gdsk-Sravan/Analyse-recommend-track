"""Inspect tracking_workbook.xlsx — dump sheet names + row counts + first row."""
from pathlib import Path
from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent.parent
wb = load_workbook(HERE / "tracking_workbook.xlsx", read_only=True, data_only=True)

print(f"Workbook: {HERE / 'tracking_workbook.xlsx'}")
print(f"Sheets ({len(wb.sheetnames)}):")
for name in wb.sheetnames:
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    data_rows = len(rows) - 1 if rows else 0
    print(f"  {name:20s} : {data_rows:5d} data rows, {ws.max_column:3d} cols")
    if rows:
        hdr = list(rows[0])[:8]
        print(f"      header[:8] = {hdr}")
    if len(rows) > 1:
        sample = list(rows[1])[:8]
        print(f"      row1[:8]   = {sample}")

print()
print("=== ACTIVE_TRACKING bucket distribution ===")
ws = wb["ACTIVE_TRACKING"]
rows = list(ws.iter_rows(values_only=True))
if len(rows) > 1:
    from collections import Counter
    hdr = list(rows[0])
    src_col = hdr.index("Source Sheet") if "Source Sheet" in hdr else 0
    counts = Counter(r[src_col] for r in rows[1:])
    for b, c in counts.most_common():
        print(f"  {b:15s} : {c}")
