"""3-day mock simulation of the daily-append tracking workbook.

Simulates three business days and shows how every sheet evolves.

Isolated from real data: uses a scratch directory
(../results_mock/) so tracking_store.json and daily_snapshots.jsonl for
the real pipeline are untouched.

Scenario:
  Day 1 (Mon)  — first run. Fresh universe:
                  MRF        (BUY, BREAKOUT setup)
                  TCS        (NEAR_MISS_RISING)
                  INFY       (DEVELOPING, PULLBACK setup)
                  WIPRO      (MONITOR)
                  RELIANCE   (REJECTED, was BREAKOUT idea)

  Day 2 (Tue)  — universe grows + prices move:
                  MRF        gets +6% MFE (still active, closer to T1)
                  TCS        promoted DEVELOPING → BUY, MOMENTUM setup
                  INFY       still DEVELOPING, price ticks up
                  WIPRO      degrades WATCHLIST → REJECTED
                  RELIANCE   still REJECTED
                  NEW: HDFC   (BUY, REVERSAL setup)
                  NEW: LT     (DEVELOPING)

  Day 3 (Wed)  — outcomes crystallise:
                  MRF        hits T1 → tracking_status = T1_HIT_ACTIVE (still shows)
                  TCS        hits T2 → TERMINAL (goes to DONE)
                  INFY       stopped out → TERMINAL (goes to DONE)
                  WIPRO      still REJECTED (price drop)
                  RELIANCE   still REJECTED
                  HDFC       promoted BUY, up 4%
                  LT         still DEVELOPING
                  NEW: BAJFINANCE (NEAR_MISS_FADING)

After each day we rebuild the workbook and print sheet-level summaries.
"""

from __future__ import annotations

import os
import shutil
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(HERE))

# Redirect the pipeline to a scratch results dir BEFORE any imports.
MOCK_ROOT = HERE / "results_mock"
if MOCK_ROOT.exists():
    shutil.rmtree(MOCK_ROOT)
MOCK_ROOT.mkdir()

os.environ["FRESH_START"] = "false"  # we want append semantics

# Force each module's DEFAULT_*_PATH to point at the mock dir.
import tracking_store as ts_mod  # noqa: E402
import daily_snapshot_job as ds_mod  # noqa: E402
import tracking_workbook_job as wb_mod  # noqa: E402
from tracking_store import TrackingStore, STATUS_ACTIVE, STATUS_T1_HIT, STATUS_T2_HIT, STATUS_STOPPED  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

STORE_PATH = MOCK_ROOT / "tracking_store.json"
EVENTS_PATH = MOCK_ROOT / "tracking_events.jsonl"
SNAPSHOTS_PATH = MOCK_ROOT / "daily_snapshots.jsonl"
WEEKLY_PATH = MOCK_ROOT / "weekly_review_history.jsonl"
WORKBOOK_PATH = MOCK_ROOT / "tracking_workbook_mock.xlsx"

# Monkey-patch module defaults so build_workbook picks up the mock jsonl.
ds_mod.DAILY_SNAPSHOTS_PATH = SNAPSHOTS_PATH
wb_mod.DAILY_SNAPSHOTS_PATH = SNAPSHOTS_PATH


# ---------------------------------------------------------------------------
# Mock record factory
# ---------------------------------------------------------------------------

def make_record(*,
                symbol: str,
                first_seen: str,
                entry_stage: str,
                current_stage: str,
                setup: str | None,
                sector: str,
                regime: str = "TRENDING",
                entry_conf: float = 0.7,
                cur_conf: float = 0.7,
                entry_price: float = 100.0,
                current_price: float = 100.0,
                t1: float = 110.0,
                t2: float = 120.0,
                stop: float = 95.0,
                mfe: float = 0.0,
                mae: float = 0.0,
                days: int = 0,
                journey: list[str] | None = None,
                t1_hit: bool = False,
                t2_hit: bool = False,
                stop_hit: bool = False,
                status: str = STATUS_ACTIVE) -> dict:
    tid = f"{symbol}_{first_seen.replace('-', '')}"
    return {
        "tracking_id": tid,
        "symbol": symbol,
        "first_seen_date": first_seen,
        "days_active": days,
        "journey_path": journey or [entry_stage],
        "tracking_status": status,
        "t1_hit": t1_hit,
        "t2_hit": t2_hit,
        "stop_hit": stop_hit,
        "mfe_pct": mfe,
        "mae_pct": mae,
        "entry": {
            "stage": entry_stage,
            "setup_type": setup,
            "sector": sector,
            "regime": regime,
            "confidence": entry_conf,
            "tq": 0.6,
            "rr_ratio": 2.5,
            "opportunity_score": 65.0,
            "reference_entry_price": entry_price,
            "t1_price": t1,
            "t2_price": t2,
            "stop_price": stop,
            "fail_reasons": [],
        },
        "current": {
            "stage": current_stage,
            "confidence": cur_conf,
            "tq": 0.6,
            "rr_ratio": 2.5,
            "opportunity_score": 65.0,
            "close": current_price,
        },
    }


# ---------------------------------------------------------------------------
# Day-by-day fixtures
# ---------------------------------------------------------------------------

def day1_universe() -> list[dict]:
    """Monday 2026-07-13 — fresh universe of 5 stocks."""
    return [
        make_record(symbol="MRF",      first_seen="2026-07-13", entry_stage="BUY",
                    current_stage="BUY", setup="BREAKOUT", sector="AUTO",
                    entry_price=100_000, current_price=100_500, t1=106_000, t2=112_000, stop=97_000,
                    mfe=0.5, mae=0.0),
        make_record(symbol="TCS",      first_seen="2026-07-13", entry_stage="NEAR_MISS_RISING",
                    current_stage="NEAR_MISS_RISING", setup=None, sector="IT",
                    entry_price=3500, current_price=3510, t1=3700, t2=3900, stop=3400,
                    mfe=0.3, mae=0.0),
        make_record(symbol="INFY",     first_seen="2026-07-13", entry_stage="DEVELOPING",
                    current_stage="DEVELOPING", setup="PULLBACK", sector="IT",
                    entry_price=1500, current_price=1505, t1=1600, t2=1700, stop=1450,
                    mfe=0.3, mae=0.0),
        make_record(symbol="WIPRO",    first_seen="2026-07-13", entry_stage="WATCHLIST",
                    current_stage="WATCHLIST", setup=None, sector="IT",
                    entry_price=450, current_price=452, t1=480, t2=510, stop=440,
                    mfe=0.4, mae=0.0),
        make_record(symbol="RELIANCE", first_seen="2026-07-13", entry_stage="REJECTED",
                    current_stage="REJECTED", setup="BREAKOUT", sector="OIL_GAS",
                    entry_price=2500, current_price=2495, t1=None, t2=None, stop=None,
                    mfe=0.0, mae=0.2),
    ]


def day2_updates(records: dict) -> None:
    """Tuesday 2026-07-14 — mutate day-1 records and add 2 new."""
    # MRF up 6% — still active, no T1 yet
    r = records["MRF_20260713"]
    r["current"]["close"] = 106_000 * 0.98  # 103_880 = +3.88% from entry (not T1)
    r["days_active"] = 1
    r["mfe_pct"] = 3.88
    r["mae_pct"] = 0.0
    r["journey_path"] = ["BUY", "BUY"]

    # TCS promoted to BUY, MOMENTUM setup discovered
    r = records["TCS_20260713"]
    r["current"]["stage"] = "BUY"
    r["current"]["close"] = 3600
    r["entry"]["setup_type"] = "MOMENTUM"  # setup identified today
    r["days_active"] = 1
    r["mfe_pct"] = 2.86
    r["journey_path"] = ["NEAR_MISS_RISING", "BUY"]

    # INFY still DEVELOPING, price up
    r = records["INFY_20260713"]
    r["current"]["close"] = 1530
    r["days_active"] = 1
    r["mfe_pct"] = 2.0
    r["journey_path"] = ["DEVELOPING", "DEVELOPING"]

    # WIPRO degrades WATCHLIST → REJECTED
    r = records["WIPRO_20260713"]
    r["current"]["stage"] = "REJECTED"
    r["current"]["close"] = 440
    r["days_active"] = 1
    r["mfe_pct"] = 0.4
    r["mae_pct"] = 2.22
    r["journey_path"] = ["WATCHLIST", "REJECTED"]

    # RELIANCE unchanged
    r = records["RELIANCE_20260713"]
    r["current"]["close"] = 2490
    r["days_active"] = 1
    r["mae_pct"] = 0.4

    # NEW on day 2
    records["HDFC_20260714"] = make_record(
        symbol="HDFC", first_seen="2026-07-14", entry_stage="BUY",
        current_stage="BUY", setup="REVERSAL", sector="BANK",
        entry_price=1600, current_price=1605, t1=1700, t2=1800, stop=1560,
        mfe=0.3)
    records["LT_20260714"] = make_record(
        symbol="LT", first_seen="2026-07-14", entry_stage="DEVELOPING",
        current_stage="DEVELOPING", setup=None, sector="INFRA",
        entry_price=3400, current_price=3410, t1=3600, t2=3800, stop=3300,
        mfe=0.29)


def day3_updates(records: dict) -> None:
    """Wednesday 2026-07-15 — outcomes crystallise."""
    # MRF hits T1 (still active, tracking T2)
    r = records["MRF_20260713"]
    r["current"]["close"] = 106_500  # above t1 106_000
    r["t1_hit"] = True
    r["t1_hit_date"] = "2026-07-15"
    r["tracking_status"] = STATUS_T1_HIT
    r["days_active"] = 2
    r["mfe_pct"] = 6.5
    r["journey_path"] = ["BUY", "BUY", "BUY"]

    # TCS hits T2 → TERMINAL
    r = records["TCS_20260713"]
    r["current"]["close"] = 3920  # above t2 3900
    r["current"]["stage"] = "BUY"
    r["t1_hit"] = True
    r["t1_hit_date"] = "2026-07-14"
    r["t2_hit"] = True
    r["t2_hit_date"] = "2026-07-15"
    r["tracking_status"] = STATUS_T2_HIT
    r["days_active"] = 2
    r["mfe_pct"] = 12.0
    r["journey_path"] = ["NEAR_MISS_RISING", "BUY", "BUY"]

    # INFY stopped out → TERMINAL
    r = records["INFY_20260713"]
    r["current"]["close"] = 1440  # below stop 1450
    r["stop_hit"] = True
    r["stop_hit_date"] = "2026-07-15"
    r["tracking_status"] = STATUS_STOPPED
    r["days_active"] = 2
    r["mfe_pct"] = 2.0
    r["mae_pct"] = 4.0
    r["journey_path"] = ["DEVELOPING", "DEVELOPING", "DEVELOPING"]

    # WIPRO still REJECTED
    r = records["WIPRO_20260713"]
    r["current"]["close"] = 435
    r["days_active"] = 2
    r["mae_pct"] = 3.33

    # RELIANCE unchanged
    r = records["RELIANCE_20260713"]
    r["current"]["close"] = 2485
    r["days_active"] = 2
    r["mae_pct"] = 0.6

    # HDFC promoted (already BUY), still active
    r = records["HDFC_20260714"]
    r["current"]["close"] = 1665
    r["days_active"] = 1
    r["mfe_pct"] = 4.06
    r["journey_path"] = ["BUY", "BUY"]

    # LT still DEVELOPING
    r = records["LT_20260714"]
    r["current"]["close"] = 3430
    r["days_active"] = 1
    r["mfe_pct"] = 0.88

    # NEW on day 3
    records["BAJFINANCE_20260715"] = make_record(
        symbol="BAJFINANCE", first_seen="2026-07-15",
        entry_stage="NEAR_MISS_FADING", current_stage="NEAR_MISS_FADING",
        setup=None, sector="NBFC",
        entry_price=7200, current_price=7180, t1=7500, t2=7800, stop=7050,
        mfe=0.0, mae=0.28)


# ---------------------------------------------------------------------------
# Sheet inspection
# ---------------------------------------------------------------------------

def dump_sheets(day_label: str) -> None:
    print("\n" + "=" * 78)
    print(f"AFTER {day_label} — sheets in {WORKBOOK_PATH.name}")
    print("=" * 78)
    wb = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    for name in wb.sheetnames:
        ws = wb[name]
        # Row 1 is header. Count real data rows.
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0] if rows else ()
        data_rows = [r for r in rows[1:] if any(c is not None for c in r)]
        print(f"\n[{name}]  rows={len(data_rows)}  cols={len(header)}")
        if not data_rows:
            print("  (empty)")
            continue
        # For daily bucket + active sheets, show symbol, run_date, current_stage,
        # current_price, mfe, status per row.
        # NOTE: column 0 is now "Row Key" (added 2026-07-13) so all indices
        # bumped +1 vs the original mock dumper.
        if name == "ACTIVE_TRACKING":
            print(f"  cols shown: RowKey | Source | Run Date | Symbol | Cur Stage | Cur Price | MFE% | Status")
            for r in data_rows[:20]:
                print(f"  {str(r[0]):<38} {str(r[1]):<10} {str(r[2]):<11} {str(r[3]):<12} {str(r[5]):<12} "
                      f"{str(r[25] if len(r) > 25 and r[25] is not None else '-'):>10} "
                      f"{str(r[27] if len(r) > 27 and r[27] is not None else '-'):>7} "
                      f"{r[30] if len(r) > 30 and r[30] is not None else '-'}")
        elif name in ("BUY", "NEAR_MISS", "DEVELOPING", "MONITOR", "REJECTED",
                      "BREAKOUT", "MOMENTUM", "PULLBACK", "REVERSAL", "REJECTED_SETUP"):
            print(f"  cols shown: RowKey | Run Date | Symbol | Cur Stage | Cur Price | MFE% | Status")
            for r in data_rows[:20]:
                print(f"  {str(r[0]):<38} {str(r[1]):<11} {str(r[2]):<12} {str(r[4]):<12} "
                      f"{str(r[24] if len(r) > 24 and r[24] is not None else '-'):>10} "
                      f"{str(r[26] if len(r) > 26 and r[26] is not None else '-'):>7} "
                      f"{r[29] if len(r) > 29 and r[29] is not None else '-'}")
        elif name == "DONE":
            print(f"  cols shown: RowKey | Symbol | Setup | Days | Entry Stage | Status")
            for r in data_rows[:20]:
                print(f"  {str(r[0]):<38} {str(r[2]):<12} {str(r[3]):<10} {str(r[6]):>4} "
                      f"{str(r[7]):<12} {r[16] if len(r) > 16 else '-'}")
        elif name == "WEEKLY_SUMMARY":
            print(f"  cols shown: Sheet | Active | Tracking | T1w | T2w | SLw | T1all | T2all | SLall")
            for r in data_rows[:20]:
                print(f"  {r[0]:<16} act={r[3]:<3} track={r[4]:<3} "
                      f"T1w={r[5]:<3} T2w={r[6]:<3} SLw={r[7]:<3} "
                      f"T1={r[8]:<3} T2={r[9]:<3} SL={r[10]}")
        elif name == "WEEKLY_REVIEW":
            print(f"  cols shown: Week | Cat | New | Active | Resolved")
            for r in data_rows[:20]:
                print(f"  {r[0]:<10} {r[2]:<12} new={r[3]:<3} active={r[4]:<3} resolved={r[5]}")
        elif name == "RESEARCH":
            print(f"  cols shown: RowKey | Symbol | Setup | Status | Cur Stage | Sector")
            for r in data_rows[:20]:
                print(f"  {str(r[0]):<38} {str(r[2]):<12} {str(r[3]):<10} {str(r[5]):<8} {str(r[6]):<12} {r[7]}")
        elif name == "_LEGEND":
            print(f"  (legend — {len(data_rows)} terms defined)")
    wb.close()


# ---------------------------------------------------------------------------
# Simulate one day
# ---------------------------------------------------------------------------

def run_day(store: TrackingStore, run_date: str, label: str) -> None:
    print("\n" + "#" * 78)
    print(f"# {label}  (run_date={run_date})")
    print("#" * 78)
    store.save()
    ds_mod.append_from_store(store, run_date=run_date, snapshots_path=SNAPSHOTS_PATH)
    wb_mod.build_workbook(store, WORKBOOK_PATH,
                          weekly_history_path=WEEKLY_PATH,
                          snapshots_path=SNAPSHOTS_PATH)
    dump_sheets(label)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    # Fresh store pointing at mock dir.
    store = TrackingStore(store_path=STORE_PATH, events_path=EVENTS_PATH)

    # ---- DAY 1 ----
    for rec in day1_universe():
        store.records[rec["tracking_id"]] = rec
    run_day(store, "2026-07-13", "DAY 1  Monday 2026-07-13")

    # ---- DAY 2 ----
    day2_updates(store.records)
    run_day(store, "2026-07-14", "DAY 2  Tuesday 2026-07-14")

    # ---- DAY 3 ----
    day3_updates(store.records)
    run_day(store, "2026-07-15", "DAY 3  Wednesday 2026-07-15")

    # Final row count in the jsonl (audit log).
    lines = SNAPSHOTS_PATH.read_text(encoding="utf-8").splitlines()
    print("\n" + "=" * 78)
    print(f"FINAL AUDIT LOG: {SNAPSHOTS_PATH.name} has {len(lines)} lines total")
    print("=" * 78)


if __name__ == "__main__":
    main()
