"""Dump Tracking ID + Symbol + Run Date from BUY sheet of the mock workbook."""
from pathlib import Path
from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent.parent
wb = load_workbook(HERE / "results_mock" / "tracking_workbook_mock.xlsx", read_only=True, data_only=True)
for sname in ("BUY", "ACTIVE_TRACKING"):
    ws = wb[sname]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    print(f"\n[{sname}]  header cols 1-6: {header[:6]}")
    print(f"           header col 3    : {header[2]!r}")
    for r in rows[1:]:
        if sname == "BUY":
            print(f"  RunDate={r[0]}  Symbol={r[1]}  TrackingID={r[2]}  Stage={r[3]}")
        else:  # ACTIVE_TRACKING has Source col shifting everything +1
            print(f"  Source={r[0]}  RunDate={r[1]}  Symbol={r[2]}  TrackingID={r[3]}  Stage={r[4]}")
wb.close()
