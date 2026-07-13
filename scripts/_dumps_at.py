"""Dump ACTIVE_TRACKING sheet for debugging."""
from openpyxl import load_workbook
wb = load_workbook("results_e2e_test/tracking_workbook_e2e.xlsx", read_only=True)
ws = wb["ACTIVE_TRACKING"]
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
print("HEADERS[:6] =", headers[:6])
seen = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers, r))
    rk = d.get("Row Key")
    seen.setdefault(rk, []).append((d.get("Source Sheet"), d.get("Tracking ID"), d.get("Run Date")))
for k, vs in seen.items():
    if len(vs) > 1:
        print(f"DUP row_key={k!r}: {vs}")
    else:
        print(f"OK  row_key={k!r}: {vs}")
