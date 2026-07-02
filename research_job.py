"""
research_job.py — Research & Analytics (JOB 3)
===============================================
Run weekly (Saturdays) and monthly.
GitHub Actions: schedule 0 4 * * 6 (Saturdays 9:30 AM IST / 4:00 UTC)

Reads:  recommendation_tracker.xlsx
Writes: recommendation_tracker.xlsx — Analysis sheets updated

Usage:
    python research_job.py
"""

import os
import numpy as np
from datetime import datetime

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

try:
    import pandas as pd
    _PD_OK = True
except ImportError:
    _PD_OK = False
    print("[ERROR] pandas not installed. Run: pip install pandas")

try:
    import openpyxl
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False
    print("[ERROR] openpyxl not installed. Run: pip install openpyxl")

TRACKER_XLSX = os.getenv("TRACKER_XLSX", "recommendation_tracker.xlsx")
# Phase C7c: FRESH_START flag — skip research on the reset run (nothing meaningful to analyze yet)
FRESH_START  = os.getenv("FRESH_START", "false").lower() == "true"
# Phase C7e (2026-07-02): min sample size for any bucket-level stat to be
# considered meaningful. Below this the bucket is reported with an
# "(insufficient)" note instead of misleading percentages built on n=1..4.
MIN_SAMPLE_N = int(os.getenv("MIN_SAMPLE_N", "5"))


def run_research():
    print(f"=== RESEARCH JOB: {datetime.now().strftime('%Y-%m-%d')} ===")

    # Phase C7c: FRESH_START safety — skip when state was wiped this run.
    # Historical analysis on 1 day of data is meaningless; wait for real history.
    if FRESH_START:
        print("[FRESH_START] research_job: state was wiped this run — skipping analysis")
        print("[FRESH_START] research_job: will resume on the next scheduled run once history builds")
        return

    if not _PD_OK or not _OPENPYXL_OK:
        print("[ERROR] Missing dependencies — aborting")
        return

    if not os.path.exists(TRACKER_XLSX):
        print(f"[WARN] No tracker file at {TRACKER_XLSX} — nothing to analyze yet")
        return

    try:
        wb = openpyxl.load_workbook(TRACKER_XLSX)
    except Exception as e:
        print(f"[ERROR] Could not open {TRACKER_XLSX}: {e}")
        return

    df_rec   = _sheet_to_df(wb, "Recommendations")
    df_track = _sheet_to_df(wb, "Daily Tracking")

    if df_rec.empty:
        print("[INFO] No recommendation data yet")
        wb.save(TRACKER_XLSX)
        return

    print(f"[INFO] Analyzing {len(df_rec)} recommendations, {len(df_track)} tracking records")

    # Auto-create v2 research sheets if missing (for older tracker files)
    _ensure_sheets(wb, [
        "Weekday Analysis", "Holding Period Analysis", "Category Comparison",
        "Conf x TQ Matrix", "Catalyst Analysis", "Fail Reason Analysis",
        "Regime x Sector", "Confidence Trajectory",
        # Phase 5 (2026-07-01): institutional-flow analytics
        "Delivery Flow", "Sector Rotation",
        # Phase C7e (2026-07-02): portfolio-level metrics + backtest divergence
        "Portfolio Metrics", "Backtest vs Live",
    ])

    # Confidence analysis
    _analyze_by_column(wb, df_rec, df_track, "Confidence", "Confidence Analysis",
                        [70, 75, 80, 83, 85, 88, 90, 95])

    # Phase C7e (2026-07-02): "TQ Analysis" and "Opp Score Analysis" removed —
    # they were single-factor duplicates of what "Conf × TQ Matrix" and the
    # Weekly Factor Summary tercile analysis already cover more precisely.
    # If those sheets exist in an older workbook, they'll be cleared but kept
    # (openpyxl leaves them empty rather than deleting — user can hide/delete).

    # Sector analysis
    _analyze_by_sector(wb, df_rec, df_track)

    # Regime analysis
    _analyze_by_regime(wb, df_rec, df_track)

    # Threshold simulation note
    _add_threshold_note(wb)

    # Monthly report
    _generate_monthly_report(wb, df_rec, df_track)

    # ── v2 research sheets ──
    _analyze_by_weekday(wb, df_rec, df_track)
    _analyze_holding_period(wb, df_rec, df_track)
    _analyze_category_comparison(wb, df_rec, df_track)
    _analyze_conf_tq_matrix(wb, df_rec, df_track)
    _analyze_catalysts(wb, df_rec, df_track)
    _analyze_fail_reasons(wb, df_rec, df_track)
    _analyze_regime_x_sector(wb, df_rec, df_track)
    _analyze_confidence_trajectory(wb, df_rec, df_track)

    # ── Phase 5 (2026-07-01): institutional-flow analytics ──
    _analyze_delivery_flow(wb, df_rec, df_track)
    _analyze_sector_rotation(wb, df_rec, df_track)

    # ── Phase C7e (2026-07-02): portfolio-level metrics + backtest divergence ──
    _analyze_portfolio_metrics(wb, df_rec, df_track)
    _analyze_backtest_vs_live(wb, df_rec, df_track)

    wb.save(TRACKER_XLSX)
    print(f"[INFO] Research job complete — {TRACKER_XLSX} updated")


def _ensure_sheets(wb, names):
    """Create any sheets in the list that don't already exist.
    Also drops orphan sheets from earlier versions that no longer exist in code.
    """
    # Phase C7e (2026-07-02): drop obsolete single-factor sheets from older
    # workbook layouts. "Conf × TQ Matrix" and the Weekly Factor Summary
    # tercile analysis cover their content more precisely. If those old
    # sheets exist, delete them so users don't stare at stale/empty tabs.
    OBSOLETE_SHEETS = ("TQ Analysis", "Opp Score Analysis")
    for _obsolete in OBSOLETE_SHEETS:
        if _obsolete in wb.sheetnames and len(wb.sheetnames) > 1:
            try:
                del wb[_obsolete]
            except Exception:
                pass

    existing = set(wb.sheetnames)
    for name in names:
        if name not in existing:
            wb.create_sheet(name)


def _sheet_to_df(wb, sheet_name):
    try:
        ws = wb[sheet_name]
        data = list(ws.values)
        if len(data) < 2:
            return pd.DataFrame()
        return pd.DataFrame(data[1:], columns=data[0])
    except Exception:
        return pd.DataFrame()


def _analyze_by_column(wb, df_rec, df_track, col, sheet_name, thresholds):
    try:
        ws = wb[sheet_name]
        # Clear existing data (keep header)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.value = None

        ws.append(["Range", "Count", "Avg Return%", "Median Return%",
                    "Win Rate%", "Avg Max Gain%", "T1 Hit%", "T2 Hit%"])

        merged = df_rec.copy()
        if not df_track.empty and "Ticker" in df_track.columns:
            best = (df_track
                    .assign(_day=pd.to_numeric(df_track.get("Day#", pd.Series(0)), errors="coerce"))
                    .sort_values("_day", ascending=False)
                    .drop_duplicates("Ticker"))
            for c in ("Return%", "Max Gain%", "T1 Hit", "T2 Hit"):
                if c in best.columns:
                    merged = merged.merge(best[["Ticker", c]], on="Ticker", how="left",
                                          suffixes=("", "_track"))

        for i in range(len(thresholds)):
            lo = thresholds[i]
            hi = thresholds[i + 1] if i + 1 < len(thresholds) else 9999
            col_series = pd.to_numeric(merged.get(col, pd.Series()), errors="coerce")
            subset = merged[(col_series >= lo) & (col_series < hi)]
            if subset.empty:
                continue

            # Phase C7e: min-N gate. Below MIN_SAMPLE_N the metrics are
            # statistically noise (a single loser can flip win-rate 0→50→100%).
            # Report the bucket but flag it clearly instead of computing stats.
            if len(subset) < MIN_SAMPLE_N:
                ws.append([
                    f"{lo}-{hi}",
                    len(subset),
                    f"(n<{MIN_SAMPLE_N})", "(insufficient)",
                    "(insufficient)", "(insufficient)",
                    "(insufficient)", "(insufficient)",
                ])
                continue

            returns = pd.to_numeric(subset.get("Return%", pd.Series()), errors="coerce").dropna()
            wins    = (returns > 0).sum()
            mg      = pd.to_numeric(subset.get("Max Gain%", pd.Series()), errors="coerce").dropna()
            t1_hits = subset.get("T1 Hit", pd.Series(False)).astype(bool).sum()
            t2_hits = subset.get("T2 Hit", pd.Series(False)).astype(bool).sum()

            ws.append([
                f"{lo}-{hi}",
                len(subset),
                round(float(returns.mean()), 2) if len(returns) > 0 else 0,
                round(float(returns.median()), 2) if len(returns) > 0 else 0,
                round(float(wins / len(returns) * 100), 1) if len(returns) > 0 else 0,
                round(float(mg.mean()), 2) if len(mg) > 0 else 0,
                round(float(t1_hits / len(subset) * 100), 1),
                round(float(t2_hits / len(subset) * 100), 1),
            ])

    except Exception as e:
        print(f"[WARN] Analysis failed for {sheet_name}: {e}")


def _analyze_by_sector(wb, df_rec, df_track):
    try:
        ws = wb["Sector Analysis"]
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.value = None
        ws.append(["Sector", "Count", "Avg Return%", "Win Rate%", "Avg Max Gain%"])

        if df_rec.empty or "Sector" not in df_rec.columns:
            return

        merged = df_rec.copy()
        if not df_track.empty and "Ticker" in df_track.columns:
            best = (df_track
                    .assign(_day=pd.to_numeric(df_track.get("Day#", pd.Series(0)), errors="coerce"))
                    .sort_values("_day", ascending=False)
                    .drop_duplicates("Ticker"))
            merge_cols = ["Ticker"]
            for c in ("Return%", "Max Gain%"):
                if c in best.columns:
                    merge_cols.append(c)
            if len(merge_cols) > 1:
                merged = merged.merge(best[merge_cols], on="Ticker", how="left")

        for sector in df_rec["Sector"].dropna().unique():
            subset  = merged[merged["Sector"] == sector]
            # Phase C7e: min-N gate
            if len(subset) < MIN_SAMPLE_N:
                ws.append([sector, len(subset), f"(n<{MIN_SAMPLE_N})",
                           "(insufficient)", "(insufficient)"])
                continue
            returns = pd.to_numeric(subset.get("Return%", pd.Series()), errors="coerce").dropna()
            max_gain = pd.to_numeric(subset.get("Max Gain%", pd.Series()), errors="coerce").dropna()
            wins    = (returns > 0).sum()
            ws.append([
                sector,
                len(subset),
                round(float(returns.mean()), 2) if len(returns) > 0 else 0,
                round(float(wins / len(returns) * 100), 1) if len(returns) > 0 else 0,
                round(float(max_gain.mean()), 2) if len(max_gain) > 0 else 0,
            ])
    except Exception as e:
        print(f"[WARN] Sector analysis failed: {e}")


def _analyze_by_regime(wb, df_rec, df_track):
    try:
        ws = wb["Regime Analysis"]
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.value = None
        ws.append(["Regime", "Count", "Avg Return%", "Win Rate%", "Best Sector"])

        if df_rec.empty or "Regime" not in df_rec.columns:
            return

        merged = df_rec.copy()
        if not df_track.empty and "Ticker" in df_track.columns:
            best = (df_track
                    .assign(_day=pd.to_numeric(df_track.get("Day#", pd.Series(0)), errors="coerce"))
                    .sort_values("_day", ascending=False)
                    .drop_duplicates("Ticker"))
            if "Return%" in best.columns:
                merged = merged.merge(best[["Ticker", "Return%"]], on="Ticker", how="left")

        for regime in df_rec["Regime"].dropna().unique():
            subset  = merged[merged["Regime"] == regime]
            # Phase C7e: min-N gate
            if len(subset) < MIN_SAMPLE_N:
                ws.append([regime, len(subset), f"(n<{MIN_SAMPLE_N})",
                           "(insufficient)", "—"])
                continue
            returns = pd.to_numeric(subset.get("Return%", pd.Series()), errors="coerce").dropna()
            wins    = (returns > 0).sum()

            # Best sector inside this regime = highest average Return%
            best_sector = "—"
            if "Sector" in subset.columns and not subset.empty:
                try:
                    tmp = subset.copy()
                    tmp["_ret"] = pd.to_numeric(tmp.get("Return%", pd.Series()), errors="coerce")
                    tmp = tmp.dropna(subset=["_ret", "Sector"])
                    if not tmp.empty:
                        sec_means = tmp.groupby("Sector")["_ret"].mean()
                        # only consider sectors with 2+ picks so we don't crown a single outlier
                        sec_counts = tmp.groupby("Sector")["_ret"].count()
                        eligible = sec_means[sec_counts >= 2]
                        if not eligible.empty:
                            best_sector = f"{eligible.idxmax()} ({eligible.max():+.1f}%)"
                        elif not sec_means.empty:
                            best_sector = f"{sec_means.idxmax()} ({sec_means.max():+.1f}%)"
                except Exception:
                    best_sector = "—"

            ws.append([
                regime,
                len(subset),
                round(float(returns.mean()), 2) if len(returns) > 0 else 0,
                round(float(wins / len(returns) * 100), 1) if len(returns) > 0 else 0,
                best_sector,
            ])
    except Exception as e:
        print(f"[WARN] Regime analysis failed: {e}")


def _add_threshold_note(wb):
    """Simulation note in Confidence Analysis sheet."""
    try:
        ws    = wb["Confidence Analysis"]
        note_row = ws.max_row + 2
        ws.cell(row=note_row,     column=1, value="THRESHOLD SIMULATION")
        ws.cell(row=note_row + 1, column=1,
                value="Note: Adjust thresholds in config based on above win rates.")
        ws.cell(row=note_row + 2, column=1,
                value="Never auto-modify — use statistical evidence only (3+ months data).")
    except Exception:
        pass


def _generate_monthly_report(wb, df_rec, df_track):
    try:
        ws = wb["Monthly Report"]
        for row in ws.iter_rows(min_row=1):
            for cell in row:
                cell.value = None

        now = datetime.now()
        ws["A1"] = f"Monthly Research Report — {now.strftime('%B %Y')}"
        ws["A2"] = f"Generated: {now.strftime('%Y-%m-%d %H:%M')}"

        # ── Build month-by-month rollup table ──
        # Merge Recommendations with latest Daily Tracking row per ticker to get Return% + Status
        merged = df_rec.copy()
        if not df_track.empty and "Ticker" in df_track.columns:
            best = (df_track
                    .assign(_day=pd.to_numeric(df_track.get("Day#", pd.Series(0)), errors="coerce"))
                    .sort_values("_day", ascending=False)
                    .drop_duplicates("Ticker"))
            keep = ["Ticker"]
            for c in ("Return%", "Max Gain%", "Status"):
                if c in best.columns:
                    keep.append(c)
            if len(keep) > 1:
                merged = merged.merge(best[keep], on="Ticker", how="left", suffixes=("", "_trk"))

        if merged.empty or "Date" not in merged.columns:
            ws["A4"] = "No data yet — first Recommendations sheet needs data."
            return

        merged["_dt"] = pd.to_datetime(merged["Date"], errors="coerce")
        merged = merged.dropna(subset=["_dt"])
        merged["_month"] = merged["_dt"].dt.strftime("%Y-%m")

        # Table headers at row 4
        headers = ["Month", "Picks", "Closed", "Win Rate%", "Avg Return%",
                   "Annualized%", "Best Trade%", "Worst Trade%", "Best Sector"]
        for i, h in enumerate(headers, start=1):
            ws.cell(row=4, column=i, value=h)

        # Group by month (chronological — oldest first)
        row_idx = 5
        month_stats = []  # keep list so we can compute deltas after
        for month, subset in merged.groupby("_month"):
            returns = pd.to_numeric(subset.get("Return%", pd.Series()), errors="coerce").dropna()

            status_col = subset.get("Status_trk", subset.get("Status", pd.Series()))
            closed_mask = status_col.astype(str).str.upper().isin(
                ["T1_HIT", "T2_HIT", "STOPPED", "EXPIRED"]
            )
            closed_count = int(closed_mask.sum())

            wins = int((returns > 0).sum())
            settled_returns = returns  # settled = has a Return% number
            hit_rate = round(wins / len(settled_returns) * 100, 1) if len(settled_returns) else 0
            avg_ret  = round(float(returns.mean()), 2) if len(returns) else 0
            best_tr  = round(float(returns.max()), 2) if len(returns) else 0
            worst_tr = round(float(returns.min()), 2) if len(returns) else 0

            # Best sector inside this month
            best_sector = "—"
            if "Sector" in subset.columns:
                tmp = subset.copy()
                tmp["_ret"] = pd.to_numeric(tmp.get("Return%", pd.Series()), errors="coerce")
                tmp = tmp.dropna(subset=["_ret", "Sector"])
                if not tmp.empty:
                    sec_means = tmp.groupby("Sector")["_ret"].mean()
                    sec_counts = tmp.groupby("Sector")["_ret"].count()
                    eligible = sec_means[sec_counts >= 2]
                    if not eligible.empty:
                        best_sector = f"{eligible.idxmax()} ({eligible.max():+.1f}%)"
                    elif not sec_means.empty:
                        best_sector = f"{sec_means.idxmax()} ({sec_means.max():+.1f}%)"

            row_data = [month, len(subset), closed_count, hit_rate, avg_ret,
                        None,  # annualized filled below
                        best_tr, worst_tr, best_sector]
            # Phase C7e: annualized % — assumes month's avg return compounded 12x.
            # Formula: ((1 + avg_ret/100) ^ 12 - 1) × 100
            try:
                annualized = ((1 + avg_ret / 100.0) ** 12 - 1) * 100
                row_data[5] = round(annualized, 2)
            except Exception:
                row_data[5] = 0
            for i, v in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=i, value=v)
            month_stats.append({"month": month, "hit_rate": hit_rate, "avg_ret": avg_ret})
            row_idx += 1

        # ── Trend section ──
        row_idx += 1
        ws.cell(row=row_idx, column=1, value="TREND (last month vs previous)")
        row_idx += 1
        if len(month_stats) >= 2:
            last = month_stats[-1]
            prev = month_stats[-2]
            hr_delta  = round(last["hit_rate"] - prev["hit_rate"], 1)
            ret_delta = round(last["avg_ret"]  - prev["avg_ret"],  2)
            trend_dir = "improving" if (hr_delta > 0 and ret_delta > 0) \
                        else ("degrading" if (hr_delta < 0 and ret_delta < 0) \
                              else "mixed")
            ws.cell(row=row_idx,     column=1,
                    value=f"Win Rate: {prev['hit_rate']}% → {last['hit_rate']}% (Δ {hr_delta:+.1f} pts)")
            ws.cell(row=row_idx + 1, column=1,
                    value=f"Avg Return: {prev['avg_ret']}% → {last['avg_ret']}% (Δ {ret_delta:+.2f} pts)")
            ws.cell(row=row_idx + 2, column=1, value=f"Direction: {trend_dir.upper()}")
            row_idx += 4
        else:
            ws.cell(row=row_idx, column=1, value="Need at least 2 months of data for trend.")
            row_idx += 2

        # ── Static guidance (kept from original) ──
        ws.cell(row=row_idx, column=1, value="LESSONS LEARNED")
        row_idx += 1
        for note in [
            "1. Monitor win rates by confidence band (aim for 60%+ at min threshold)",
            "2. Compare NEAR MISS vs BUY outcomes over time",
            "3. Track sector performance vs broad market",
            "4. Validate regime-specific thresholds after 3 months",
            "5. Never lower thresholds based on fewer than 20 closed trades",
        ]:
            ws.cell(row=row_idx, column=1, value=note)
            row_idx += 1

        # ── Formatting ──
        try:
            from openpyxl.styles import Font, PatternFill
            ws["A1"].font = Font(bold=True, size=13)
            ws["A2"].font = Font(italic=True, size=10)
            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            for i in range(1, len(headers) + 1):
                c = ws.cell(row=4, column=i)
                c.font = Font(bold=True)
                c.fill = header_fill
            for i, w in enumerate([10, 8, 8, 10, 12, 12, 12, 12, 22], start=1):
                ws.column_dimensions[chr(64 + i)].width = w
        except Exception:
            pass

    except Exception as e:
        print(f"[WARN] Monthly report failed: {e}")


# ═══════════════════════════════════════════════════════════════════════════
# v2 RESEARCH SHEETS
# ═══════════════════════════════════════════════════════════════════════════

def _cache_freshness_note(cache_path: str, stale_days: int = 2) -> str:
    """Phase C7e: return a warning banner if the cache file's mtime is older
    than stale_days. Empty string if fresh or file missing (handled elsewhere).
    """
    try:
        if not os.path.exists(cache_path):
            return ""
        import time as _time
        mtime = os.path.getmtime(cache_path)
        age_days = (_time.time() - mtime) / 86400.0
        if age_days > stale_days:
            return (f"⚠ CACHE STALE: {cache_path} is {age_days:.1f} days old "
                    f"(threshold {stale_days}d). Signals below may be outdated.")
    except Exception:
        pass
    return ""


def _merge_outcomes(df_rec, df_track):
    """Merge Recommendations with latest Daily Tracking row per ticker.
    Returns merged DataFrame with Return%, Max Gain%, Max DD%, T1 Hit, T2 Hit, Status_trk.
    """
    if df_rec.empty:
        return df_rec.copy()
    merged = df_rec.copy()
    if not df_track.empty and "Ticker" in df_track.columns:
        try:
            best = (df_track
                    .assign(_day=pd.to_numeric(df_track.get("Day#", pd.Series(0)), errors="coerce"))
                    .sort_values("_day", ascending=False)
                    .drop_duplicates("Ticker"))
            keep = ["Ticker"]
            for c in ("Return%", "Max Gain%", "Max DD%", "T1 Hit", "T2 Hit",
                      "Holding Days", "Status"):
                if c in best.columns:
                    keep.append(c)
            if len(keep) > 1:
                merged = merged.merge(best[keep], on="Ticker", how="left",
                                      suffixes=("", "_trk"))
        except Exception:
            pass
    return merged


def _format_header(ws, headers, row=1):
    """Bold + light-blue fill for a header row."""
    try:
        from openpyxl.styles import Font, PatternFill
        fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        for i in range(1, len(headers) + 1):
            c = ws.cell(row=row, column=i)
            c.font = Font(bold=True)
            c.fill = fill
    except Exception:
        pass


# ── 1. Weekday Analysis ──────────────────────────────────────────────────
def _analyze_by_weekday(wb, df_rec, df_track):
    """Do Monday picks outperform Friday picks?"""
    try:
        ws = wb["Weekday Analysis"]
        for row in ws.iter_rows(min_row=1):
            for cell in row:
                cell.value = None

        ws["A1"] = "Weekday Analysis — Does day-of-recommendation matter?"
        headers = ["Weekday", "Count", "Win Rate%", "Avg Return%",
                   "Best Trade%", "Worst Trade%"]
        for i, h in enumerate(headers, start=1):
            ws.cell(row=3, column=i, value=h)
        _format_header(ws, headers, row=3)

        merged = _merge_outcomes(df_rec, df_track)
        if merged.empty or "Date" not in merged.columns:
            ws["A5"] = "No data yet."
            return

        merged["_dt"] = pd.to_datetime(merged["Date"], errors="coerce")
        merged = merged.dropna(subset=["_dt"])
        merged["_weekday"] = merged["_dt"].dt.day_name()

        row_idx = 4
        weekday_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
        for wd in weekday_order:
            subset = merged[merged["_weekday"] == wd]
            if subset.empty:
                continue
            returns = pd.to_numeric(subset.get("Return%", pd.Series()), errors="coerce").dropna()
            wins = int((returns > 0).sum())
            ws.cell(row=row_idx, column=1, value=wd)
            ws.cell(row=row_idx, column=2, value=len(subset))
            ws.cell(row=row_idx, column=3,
                    value=round(wins / len(returns) * 100, 1) if len(returns) else 0)
            ws.cell(row=row_idx, column=4,
                    value=round(float(returns.mean()), 2) if len(returns) else 0)
            ws.cell(row=row_idx, column=5,
                    value=round(float(returns.max()), 2) if len(returns) else 0)
            ws.cell(row=row_idx, column=6,
                    value=round(float(returns.min()), 2) if len(returns) else 0)
            row_idx += 1

        row_idx += 1
        ws.cell(row=row_idx, column=1,
                value="INSIGHT: If Friday's Win Rate is >5pts below the average, "
                      "consider skipping Friday scans (weekend gap risk).")
    except Exception as e:
        print(f"[WARN] Weekday analysis failed: {e}")


# ── 2. Holding Period Analysis ───────────────────────────────────────────
def _analyze_holding_period(wb, df_rec, df_track):
    """When do winners hit target? Buckets by Holding Days."""
    try:
        ws = wb["Holding Period Analysis"]
        for row in ws.iter_rows(min_row=1):
            for cell in row:
                cell.value = None

        ws["A1"] = "Holding Period Analysis — When do trades close out?"
        headers = ["Holding Days", "Count", "% of All Closed",
                   "Win Rate%", "Avg Return%"]
        for i, h in enumerate(headers, start=1):
            ws.cell(row=3, column=i, value=h)
        _format_header(ws, headers, row=3)

        merged = _merge_outcomes(df_rec, df_track)
        if merged.empty or "Holding Days" not in merged.columns:
            ws["A5"] = "No Daily Tracking data yet (need Holding Days column)."
            return

        merged["_hd"] = pd.to_numeric(merged["Holding Days"], errors="coerce")
        settled = merged.dropna(subset=["_hd"])
        if settled.empty:
            ws["A5"] = "No settled trades yet."
            return

        buckets = [
            ("1-2 days",    lambda x: x <= 2),
            ("3-5 days",    lambda x: (x >= 3) & (x <= 5)),
            ("6-10 days",   lambda x: (x >= 6) & (x <= 10)),
            ("11-15 days",  lambda x: (x >= 11) & (x <= 15)),
            ("16+ days",    lambda x: x >= 16),
        ]

        total = len(settled)
        row_idx = 4
        for label, mask_fn in buckets:
            subset = settled[mask_fn(settled["_hd"])]
            if subset.empty:
                continue
            returns = pd.to_numeric(subset.get("Return%", pd.Series()),
                                    errors="coerce").dropna()
            wins = int((returns > 0).sum())
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=len(subset))
            ws.cell(row=row_idx, column=3,
                    value=round(len(subset) / total * 100, 1))
            ws.cell(row=row_idx, column=4,
                    value=round(wins / len(returns) * 100, 1) if len(returns) else 0)
            ws.cell(row=row_idx, column=5,
                    value=round(float(returns.mean()), 2) if len(returns) else 0)
            row_idx += 1

        row_idx += 1
        avg_hd = float(settled["_hd"].mean())
        ws.cell(row=row_idx, column=1,
                value=f"AVG HOLDING PERIOD: {avg_hd:.1f} days")
        ws.cell(row=row_idx + 1, column=1,
                value="INSIGHT: If '11-15 days' bucket has poor win rate, "
                      "TIME_EXIT_DAYS is likely too long — trim to 10.")
    except Exception as e:
        print(f"[WARN] Holding period analysis failed: {e}")


# ── 3. Category Comparison ───────────────────────────────────────────────
def _analyze_category_comparison(wb, df_rec, df_track):
    """BUY vs NEAR_MISS vs WATCHLIST — is the tier system meaningful?"""
    try:
        ws = wb["Category Comparison"]
        for row in ws.iter_rows(min_row=1):
            for cell in row:
                cell.value = None

        ws["A1"] = "Category Comparison — Do BUY, NEAR_MISS, WATCHLIST tiers differ?"
        headers = ["Category", "Count", "Win Rate%", "Avg Return%",
                   "Avg Confidence", "Avg TQ", "Avg Opp Score"]
        for i, h in enumerate(headers, start=1):
            ws.cell(row=3, column=i, value=h)
        _format_header(ws, headers, row=3)

        merged = _merge_outcomes(df_rec, df_track)
        if merged.empty or "Category" not in merged.columns:
            ws["A5"] = "No Category column in Recommendations."
            return

        # Phase C5 (2026-07-02): main.py's classify_watchlist now emits
        # trajectory-aware sub-tiers (NEAR_MISS_RISING, NEAR_MISS_FADING) plus
        # DEVELOPING and MONITOR. Normalize them into their base tier for this
        # analytical comparison so the "BUY vs NEAR_MISS vs WATCHLIST" insight
        # (win rate should be BUY >= NEAR_MISS >= WATCHLIST) stays meaningful.
        def _normalize_tier(v):
            s = str(v or "").upper()
            if s in ("NEAR_MISS_RISING", "NEAR_MISS_FADING"):
                return "NEAR_MISS"
            if s in ("DEVELOPING", "MONITOR"):
                return "WATCHLIST"
            return s
        merged = merged.copy()
        merged["Category"] = merged["Category"].map(_normalize_tier)

        row_idx = 4
        # Preserve a sensible order if categories present
        preferred = ["BUY", "NEAR_MISS", "WATCHLIST"]
        cats_present = [c for c in preferred if c in merged["Category"].dropna().unique()]
        others = [c for c in merged["Category"].dropna().unique() if c not in preferred]
        for cat in cats_present + others:
            subset = merged[merged["Category"] == cat]
            returns = pd.to_numeric(subset.get("Return%", pd.Series()), errors="coerce").dropna()
            wins    = int((returns > 0).sum())
            conf    = pd.to_numeric(subset.get("Confidence", pd.Series()), errors="coerce").dropna()
            tq      = pd.to_numeric(subset.get("TQ", pd.Series()), errors="coerce").dropna()
            opp     = pd.to_numeric(subset.get("Opp Score", pd.Series()), errors="coerce").dropna()
            ws.cell(row=row_idx, column=1, value=cat)
            ws.cell(row=row_idx, column=2, value=len(subset))
            ws.cell(row=row_idx, column=3,
                    value=round(wins / len(returns) * 100, 1) if len(returns) else 0)
            ws.cell(row=row_idx, column=4,
                    value=round(float(returns.mean()), 2) if len(returns) else 0)
            ws.cell(row=row_idx, column=5,
                    value=round(float(conf.mean()), 1) if len(conf) else 0)
            ws.cell(row=row_idx, column=6,
                    value=round(float(tq.mean()), 1) if len(tq) else 0)
            ws.cell(row=row_idx, column=7,
                    value=round(float(opp.mean()), 1) if len(opp) else 0)
            row_idx += 1

        row_idx += 1
        ws.cell(row=row_idx, column=1,
                value="INSIGHT: BUY win rate should be >= NEAR_MISS >= WATCHLIST.")
        ws.cell(row=row_idx + 1, column=1,
                value="If NEAR_MISS matches BUY, the tier boundary is too strict.")
    except Exception as e:
        print(f"[WARN] Category comparison failed: {e}")


# ── 4. Confidence × TQ Matrix ────────────────────────────────────────────
def _analyze_conf_tq_matrix(wb, df_rec, df_track):
    """Cross-tab of Confidence buckets × TQ buckets — sweet-spot detection."""
    try:
        ws = wb["Conf x TQ Matrix"]
        for row in ws.iter_rows(min_row=1):
            for cell in row:
                cell.value = None

        ws["A1"] = "Confidence × TQ Matrix — where do the two factors compound?"
        ws["A2"] = "Each cell shows: Win Rate% (n=count)"

        merged = _merge_outcomes(df_rec, df_track)
        if merged.empty:
            ws["A4"] = "No data yet."
            return

        conf_bins = [(0, 75, "Low"), (75, 83, "Mid"), (83, 100, "High")]
        tq_bins   = [(0, 75, "Low"), (75, 85, "Mid"), (85, 100, "High")]

        merged["_conf"] = pd.to_numeric(merged.get("Confidence", pd.Series()), errors="coerce")
        merged["_tq"]   = pd.to_numeric(merged.get("TQ",         pd.Series()), errors="coerce")
        merged["_ret"]  = pd.to_numeric(merged.get("Return%",    pd.Series()), errors="coerce")

        # Column headers (TQ buckets)
        ws.cell(row=4, column=1, value="Confidence \\ TQ")
        for j, (_, _, tq_lbl) in enumerate(tq_bins, start=2):
            ws.cell(row=4, column=j, value=f"{tq_lbl} TQ")
        _format_header(ws, ["Confidence \\ TQ"] + [b[2] for b in tq_bins], row=4)

        best_cell = ("", 0.0, 0)  # (label, win_rate, count)
        for i, (cf_lo, cf_hi, cf_lbl) in enumerate(conf_bins, start=5):
            ws.cell(row=i, column=1, value=f"{cf_lbl} Conf")
            for j, (tq_lo, tq_hi, tq_lbl) in enumerate(tq_bins, start=2):
                cell_data = merged[
                    (merged["_conf"] >= cf_lo) & (merged["_conf"] < cf_hi) &
                    (merged["_tq"]   >= tq_lo) & (merged["_tq"]   < tq_hi)
                ]
                rets = cell_data["_ret"].dropna()
                if len(rets) == 0:
                    ws.cell(row=i, column=j, value="—")
                    continue
                wr = round(int((rets > 0).sum()) / len(rets) * 100, 1)
                ws.cell(row=i, column=j, value=f"{wr}% (n={len(rets)})")
                # Track best cell (need >= 5 samples to matter)
                if len(rets) >= 5 and wr > best_cell[1]:
                    best_cell = (f"{cf_lbl} Conf × {tq_lbl} TQ", wr, len(rets))

        # Also compute avg Return% matrix below
        ws.cell(row=9,  column=1, value="")
        ws.cell(row=10, column=1, value="Avg Return% matrix")
        _format_header(ws, ["Confidence \\ TQ"] + [b[2] for b in tq_bins], row=11)
        ws.cell(row=11, column=1, value="Confidence \\ TQ")
        for j, (_, _, tq_lbl) in enumerate(tq_bins, start=2):
            ws.cell(row=11, column=j, value=f"{tq_lbl} TQ")

        for i, (cf_lo, cf_hi, cf_lbl) in enumerate(conf_bins, start=12):
            ws.cell(row=i, column=1, value=f"{cf_lbl} Conf")
            for j, (tq_lo, tq_hi, tq_lbl) in enumerate(tq_bins, start=2):
                cell_data = merged[
                    (merged["_conf"] >= cf_lo) & (merged["_conf"] < cf_hi) &
                    (merged["_tq"]   >= tq_lo) & (merged["_tq"]   < tq_hi)
                ]
                rets = cell_data["_ret"].dropna()
                ws.cell(row=i, column=j,
                        value=f"{round(float(rets.mean()), 2)}%" if len(rets) else "—")

        if best_cell[0]:
            ws.cell(row=16, column=1,
                    value=f"SWEET SPOT (n>=5): {best_cell[0]} — "
                          f"{best_cell[1]}% win rate on {best_cell[2]} picks.")
        ws.cell(row=17, column=1,
                value="INSIGHT: If High×High dominates, "
                      "tighten filter to require BOTH factors high.")
    except Exception as e:
        print(f"[WARN] Conf×TQ matrix failed: {e}")


# ── 5. Catalyst Analysis ─────────────────────────────────────────────────
def _analyze_catalysts(wb, df_rec, df_track):
    """Which catalyst tokens (EARNINGS_BEAT, BREAKOUT, ...) drive returns?"""
    try:
        ws = wb["Catalyst Analysis"]
        for row in ws.iter_rows(min_row=1):
            for cell in row:
                cell.value = None

        ws["A1"] = "Catalyst Analysis — which catalysts actually pay off?"
        headers = ["Catalyst", "Count", "Win Rate%", "Avg Return%", "Best Trade%"]
        for i, h in enumerate(headers, start=1):
            ws.cell(row=3, column=i, value=h)
        _format_header(ws, headers, row=3)

        merged = _merge_outcomes(df_rec, df_track)
        if merged.empty or "Catalysts" not in merged.columns:
            ws["A5"] = "No Catalysts column."
            return

        # Explode comma-separated tokens
        from collections import defaultdict
        cat_stats = defaultdict(list)  # token -> [Return%, ...]
        for _, r in merged.iterrows():
            tokens_raw = str(r.get("Catalysts", "") or "")
            if not tokens_raw or tokens_raw.lower() == "nan":
                continue
            ret = pd.to_numeric(r.get("Return%", None), errors="coerce")
            if pd.isna(ret):
                continue
            for tok in [t.strip().upper() for t in tokens_raw.split(",") if t.strip()]:
                cat_stats[tok].append(float(ret))

        if not cat_stats:
            ws["A5"] = "No parseable catalyst data with outcomes yet."
            return

        # Sort by win-rate desc
        rows = []
        for tok, rets in cat_stats.items():
            n = len(rets)
            wins = sum(1 for r in rets if r > 0)
            wr = round(wins / n * 100, 1) if n else 0
            avg = round(sum(rets) / n, 2) if n else 0
            best = round(max(rets), 2) if rets else 0
            # Phase C7e: min-N gate — mark under-sampled tokens so the reader
            # doesn't act on a 100% win-rate from a single trade.
            if n < MIN_SAMPLE_N:
                rows.append((tok, n, f"(n<{MIN_SAMPLE_N})", f"(n<{MIN_SAMPLE_N})", best))
            else:
                rows.append((tok, n, wr, avg, best))
        # sort: gated buckets to the bottom, then by win-rate desc, then count desc
        rows.sort(key=lambda x: (isinstance(x[2], str),
                                  -(x[2] if isinstance(x[2], (int, float)) else 0),
                                  -x[1]))

        for row_idx, r in enumerate(rows, start=4):
            for j, v in enumerate(r, start=1):
                ws.cell(row=row_idx, column=j, value=v)

        row_idx = 4 + len(rows) + 1
        ws.cell(row=row_idx, column=1,
                value="INSIGHT: Catalysts with Win Rate < 50% on 20+ picks "
                      "should be dropped or down-weighted in scoring.")
    except Exception as e:
        print(f"[WARN] Catalyst analysis failed: {e}")


# ── 6. Fail Reason Analysis ──────────────────────────────────────────────
def _analyze_fail_reasons(wb, df_rec, df_track):
    """Fail Reasons frequency — which 'amber warnings' become real losses?"""
    try:
        ws = wb["Fail Reason Analysis"]
        for row in ws.iter_rows(min_row=1):
            for cell in row:
                cell.value = None

        ws["A1"] = "Fail Reason Analysis — which amber flags actually predict losses?"
        headers = ["Fail Reason", "Count in Buys+NearMiss", "Win Rate%",
                   "Avg Return%"]
        for i, h in enumerate(headers, start=1):
            ws.cell(row=3, column=i, value=h)
        _format_header(ws, headers, row=3)

        merged = _merge_outcomes(df_rec, df_track)
        if merged.empty or "Fail Reasons" not in merged.columns:
            ws["A5"] = "No Fail Reasons column."
            return

        # Only look at picks that made it through (BUY / NEAR_MISS) — WATCHLIST
        # is expected to fail. We want to see which fail flags leak into the
        # promoted picks and correlate with loss.
        if "Category" in merged.columns:
            promoted = merged[merged["Category"].isin(["BUY", "NEAR_MISS"])]
        else:
            promoted = merged

        from collections import defaultdict
        fr_stats = defaultdict(list)
        for _, r in promoted.iterrows():
            tokens_raw = str(r.get("Fail Reasons", "") or "")
            if not tokens_raw or tokens_raw.lower() == "nan":
                continue
            ret = pd.to_numeric(r.get("Return%", None), errors="coerce")
            if pd.isna(ret):
                continue
            for tok in [t.strip().upper() for t in tokens_raw.split(",") if t.strip()]:
                fr_stats[tok].append(float(ret))

        if not fr_stats:
            ws["A5"] = ("No fail reasons observed in BUY/NEAR_MISS picks with outcomes. "
                        "Either all promoted picks are clean, or data is thin.")
            return

        rows = []
        for tok, rets in fr_stats.items():
            n = len(rets)
            wins = sum(1 for r in rets if r > 0)
            wr = round(wins / n * 100, 1) if n else 0
            avg = round(sum(rets) / n, 2) if n else 0
            # Phase C7e: min-N gate
            if n < MIN_SAMPLE_N:
                rows.append((tok, n, f"(n<{MIN_SAMPLE_N})", f"(n<{MIN_SAMPLE_N})"))
            else:
                rows.append((tok, n, wr, avg))
        # sort: gated buckets to bottom, then worst win-rate first
        rows.sort(key=lambda x: (isinstance(x[2], str),
                                  (x[2] if isinstance(x[2], (int, float)) else 999),
                                  -x[1]))

        for row_idx, r in enumerate(rows, start=4):
            for j, v in enumerate(r, start=1):
                ws.cell(row=row_idx, column=j, value=v)

        row_idx = 4 + len(rows) + 1
        ws.cell(row=row_idx, column=1,
                value="INSIGHT: Any fail-flag with Win Rate < 45% on 10+ picks "
                      "should be promoted from amber warning to HARD BLOCK.")
    except Exception as e:
        print(f"[WARN] Fail reason analysis failed: {e}")


# ── 7. Regime × Sector Cross-tab ─────────────────────────────────────────
def _analyze_regime_x_sector(wb, df_rec, df_track):
    """Which sectors work in which regimes? (defensive vs cyclical detection)"""
    try:
        ws = wb["Regime x Sector"]
        for row in ws.iter_rows(min_row=1):
            for cell in row:
                cell.value = None

        ws["A1"] = "Regime × Sector — which sectors win in each regime?"
        ws["A2"] = "Cell = Avg Return% (n=count). Blank if <2 picks."

        merged = _merge_outcomes(df_rec, df_track)
        if merged.empty or "Regime" not in merged.columns or "Sector" not in merged.columns:
            ws["A4"] = "Missing Regime or Sector column."
            return

        merged["_ret"] = pd.to_numeric(merged.get("Return%", pd.Series()), errors="coerce")
        settled = merged.dropna(subset=["_ret", "Regime", "Sector"])
        if settled.empty:
            ws["A4"] = "No settled trades with Regime + Sector yet."
            return

        regimes = sorted(settled["Regime"].dropna().unique().tolist())
        sectors = sorted(settled["Sector"].dropna().unique().tolist())

        # Header row (regimes across the top)
        ws.cell(row=4, column=1, value="Sector \\ Regime")
        for j, rg in enumerate(regimes, start=2):
            ws.cell(row=4, column=j, value=str(rg))
        _format_header(ws, ["Sector \\ Regime"] + regimes, row=4)

        for i, sec in enumerate(sectors, start=5):
            ws.cell(row=i, column=1, value=sec)
            for j, rg in enumerate(regimes, start=2):
                cell_data = settled[
                    (settled["Regime"] == rg) & (settled["Sector"] == sec)
                ]
                if len(cell_data) < 2:
                    ws.cell(row=i, column=j, value="")
                    continue
                avg = round(float(cell_data["_ret"].mean()), 2)
                ws.cell(row=i, column=j, value=f"{avg:+.1f}% (n={len(cell_data)})")

        note_row = 5 + len(sectors) + 1
        ws.cell(row=note_row, column=1,
                value="INSIGHT: Green cells = sector works in that regime. "
                      "Use this to overweight/skip sectors when regime flips.")
    except Exception as e:
        print(f"[WARN] Regime×Sector failed: {e}")


# ── 8. Confidence Trajectory ─────────────────────────────────────────────
def _analyze_confidence_trajectory(wb, df_rec, df_track):
    """Do rising-confidence stocks outperform stable/falling ones?
    Reads confidence_history.json — {symbol: {dates:[3], confs:[3]}}
    """
    try:
        ws = wb["Confidence Trajectory"]
        for row in ws.iter_rows(min_row=1):
            for cell in row:
                cell.value = None

        ws["A1"] = "Confidence Trajectory — does rising conf beat stable/falling?"
        headers = ["Trajectory", "Count", "Win Rate%", "Avg Return%", "Definition"]
        for i, h in enumerate(headers, start=1):
            ws.cell(row=3, column=i, value=h)
        _format_header(ws, headers, row=3)

        conf_file = os.getenv("CONF_HISTORY_FILE", "confidence_history.json")
        if not os.path.exists(conf_file):
            ws["A5"] = f"No {conf_file} yet."
            return

        try:
            import json
            with open(conf_file, "r") as f:
                history = json.load(f)
        except Exception as e:
            ws["A5"] = f"Could not read {conf_file}: {e}"
            return

        # Build trajectory per symbol: delta of latest vs first in 3-day window
        traj = {}  # symbol -> "RISING" / "STABLE" / "FALLING"
        for sym, rec in history.items():
            confs = rec.get("confs", []) if isinstance(rec, dict) else []
            if len(confs) < 2:
                continue
            delta = confs[-1] - confs[0]
            if delta >= 5:
                traj[sym] = "RISING"
            elif delta <= -5:
                traj[sym] = "FALLING"
            else:
                traj[sym] = "STABLE"

        if not traj:
            ws["A5"] = ("Confidence history too thin (need 2+ days per symbol). "
                        "Wait ~1 week for meaningful data.")
            return

        merged = _merge_outcomes(df_rec, df_track)
        if merged.empty:
            ws["A5"] = "No recommendations to correlate."
            return

        merged["_traj"] = merged["Ticker"].map(traj).fillna("UNKNOWN")

        definitions = {
            "RISING":   "Confidence up 5+ points in 3-day window",
            "STABLE":   "Confidence changed <5 points",
            "FALLING":  "Confidence down 5+ points",
            "UNKNOWN":  "No 3-day history",
        }

        row_idx = 4
        for label in ["RISING", "STABLE", "FALLING", "UNKNOWN"]:
            subset = merged[merged["_traj"] == label]
            if subset.empty:
                continue
            rets = pd.to_numeric(subset.get("Return%", pd.Series()), errors="coerce").dropna()
            wins = int((rets > 0).sum())
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=len(subset))
            ws.cell(row=row_idx, column=3,
                    value=round(wins / len(rets) * 100, 1) if len(rets) else 0)
            ws.cell(row=row_idx, column=4,
                    value=round(float(rets.mean()), 2) if len(rets) else 0)
            ws.cell(row=row_idx, column=5, value=definitions[label])
            row_idx += 1

        row_idx += 1
        ws.cell(row=row_idx, column=1,
                value="INSIGHT: If RISING clearly beats STABLE, add a "
                      "'momentum-in-confidence' bonus (+2 pts) to the scorer.")
    except Exception as e:
        print(f"[WARN] Confidence trajectory failed: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# Phase 5 (2026-07-01): Delivery-flow + sector-rotation research sheets.
# These probe main.py's live delivery cache and sector rank history to answer
# "which delivery signals correlate with winners?" and "which sectors are
# rotating in vs out this week?".  Best-effort — silently skip if main / nselib
# / files aren't importable in this environment.
# ─────────────────────────────────────────────────────────────────────────────
def _analyze_delivery_flow(wb, df_rec, df_track):
    """Summary of live delivery % signals for stocks that appear in the
    Recommendations sheet, cross-cut by realised return.
    """
    try:
        ws = wb["Delivery Flow"]
        ws.delete_rows(1, ws.max_row)

        # Phase C7e: cache-freshness warning banner
        _cache_note = _cache_freshness_note("delivery_cache.json", stale_days=2)
        if _cache_note:
            ws.append([_cache_note])
            ws.append([])

        ws.append([
            "Symbol", "Rec Date", "Status", "Return%",
            "Deliv %Today", "Deliv 20d Avg", "Ratio", "Signal", "Note",
        ])

        # Import delivery accessor best-effort
        try:
            from main import fetch_delivery_cached, load_delivery_cache  # type: ignore
            deliv_cache = load_delivery_cache()
        except Exception as e:
            ws.cell(row=2, column=1, value=f"[WARN] delivery accessor unavailable: {e}")
            return

        if df_rec is None or df_rec.empty:
            ws.cell(row=2, column=1, value="No recommendation data yet")
            return

        # Merge return% from tracker's Daily Tracking (last row per symbol+date)
        last_ret = {}
        if df_track is not None and not df_track.empty:
            try:
                grp = df_track.sort_values("Day#").groupby(["Ticker", "Rec Date"]).tail(1)
                for _, r in grp.iterrows():
                    last_ret[(str(r.get("Ticker")), str(r.get("Rec Date")))] = float(r.get("Return%") or 0)
            except Exception:
                pass

        # Cap at 200 rows to keep xlsx snappy
        rows_out = 0
        for _, r in df_rec.tail(200).iterrows():
            try:
                sym    = str(r.get("Ticker") or "")
                rdate  = str(r.get("Date") or "")
                status = str(r.get("Status") or "")
                if not sym:
                    continue
                d = fetch_delivery_cached(sym.replace(".NS", ""), deliv_cache)
                if not d or d.get("source") != "nselib":
                    continue
                today   = float(d.get("delivery_pct_today", 0.0) or 0.0)
                avg20   = float(d.get("delivery_pct_20d_avg", 0.0) or 0.0)
                ratio   = float(d.get("delivery_ratio", 1.0) or 1.0)
                signal  = str(d.get("delivery_signal", "NEUTRAL"))
                ret_pct = last_ret.get((sym, rdate), 0.0)

                note = ""
                if signal == "DISTRIBUTION" and ret_pct > 3:
                    note = "trim on distribution"
                elif signal in ("STRONG_ACCUM", "ACCUM") and ret_pct > 0:
                    note = "accumulation, let run"
                elif signal == "WEAK" and ret_pct < 0:
                    note = "weak delivery on loser"
                ws.append([sym, rdate, status, ret_pct, today, avg20, ratio, signal, note])
                rows_out += 1
            except Exception:
                continue

        if rows_out == 0:
            ws.cell(row=2, column=1,
                    value="No nselib delivery data available (cache empty or nselib unavailable)")
    except Exception as e:
        print(f"[WARN] Delivery-flow analysis failed: {e}")


def _analyze_sector_rotation(wb, df_rec, df_track):
    """Read sector_rank_history.json (produced by main.compute_sector_rotation)
    and surface 5-day rank deltas so we can see rotating-in vs rotating-out
    sectors alongside portfolio exposure.
    """
    import json as _json
    from datetime import date as _date

    try:
        ws = wb["Sector Rotation"]
        ws.delete_rows(1, ws.max_row)

        # Phase C7e: cache-freshness warning banner
        _cache_note = _cache_freshness_note("sector_rank_history.json", stale_days=2)
        if _cache_note:
            ws.append([_cache_note])
            ws.append([])

        ws.append(["Sector", "Rank Today", "Rank 5d Ago", "Delta", "Velocity", "Portfolio Exposure"])

        history_path = os.getenv("SECTOR_RANK_HISTORY_FILE", "sector_rank_history.json")
        if not os.path.exists(history_path):
            ws.cell(row=2, column=1, value=f"No history file at {history_path}")
            return

        with open(history_path, "r", encoding="utf-8") as f:
            history = _json.load(f)
        if not isinstance(history, dict) or not history:
            ws.cell(row=2, column=1, value="History file empty")
            return

        # Latest date + closest date ≥5d back
        dates_sorted = sorted(history.keys(), reverse=True)
        today_key = dates_sorted[0]
        today_ranks = history.get(today_key, {})
        prior_ranks, prior_key = {}, None
        try:
            d_today = _date.fromisoformat(today_key)
            for k in dates_sorted[1:]:
                try:
                    d_prior = _date.fromisoformat(k)
                    if (d_today - d_prior).days >= 5:
                        prior_ranks = history[k]
                        prior_key = k
                        break
                except Exception:
                    continue
        except Exception:
            pass

        # Portfolio exposure per sector (count of active recs)
        exposure = {}
        if df_rec is not None and not df_rec.empty:
            try:
                for _, r in df_rec.iterrows():
                    if str(r.get("Status") or "").upper() == "ACTIVE":
                        sec = str(r.get("Sector") or "OTHERS")
                        exposure[sec] = exposure.get(sec, 0) + 1
            except Exception:
                pass

        # Build rows sorted by delta desc (rotating in first)
        rows = []
        for sec, rank in today_ranks.items():
            prior = prior_ranks.get(sec)
            if prior is None:
                delta = None
                velocity = "NEW"
            else:
                delta = int(prior) - int(rank)
                if delta >= 3:
                    velocity = "ROTATING_IN"
                elif delta <= -3:
                    velocity = "ROTATING_OUT"
                else:
                    velocity = "STABLE"
            rows.append((sec, int(rank), prior, delta, velocity, exposure.get(sec, 0)))

        rows.sort(key=lambda r: (r[3] if r[3] is not None else -999), reverse=True)
        for sec, rank, prior, delta, vel, exp in rows:
            ws.append([sec, rank, prior if prior is not None else "—",
                       delta if delta is not None else "—", vel, exp])

        ws.append([])
        ws.append([f"Comparing {today_key} vs {prior_key or 'n/a'} (≥5 trading days back)"])
    except Exception as e:
        print(f"[WARN] Sector rotation analysis failed: {e}")


# ═══════════════════════════════════════════════════════════════════════════
# Phase C7e (2026-07-02): Portfolio Metrics + Backtest-vs-Live Divergence
# ═══════════════════════════════════════════════════════════════════════════

def _analyze_portfolio_metrics(wb, df_rec, df_track):
    """Compute portfolio-level metrics from Daily Tracking:
       - Total return (cumulative)
       - Sharpe ratio (mean/std × √252, using per-trade returns as proxy)
       - Max drawdown from peak equity
       - Expectancy = (Win% × AvgWin) − (Loss% × |AvgLoss|)
       - Profit factor = gross wins / |gross losses|
    """
    ws = wb["Portfolio Metrics"]
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None

    ws.append(["PORTFOLIO METRICS (from Daily Tracking)"])
    ws.append([])
    ws.append(["Metric", "Value", "Interpretation"])

    if df_track is None or df_track.empty:
        ws.append(["No data", "—", "Daily Tracking sheet is empty"])
        return

    # Deduplicate to final outcome per (Symbol, Rec Date) — take last row
    outcome = _merge_outcomes(df_rec, df_track)
    if outcome is None or outcome.empty:
        ws.append(["No merged outcomes", "—", "No closed trades yet"])
        return

    # Only completed trades (Return% is non-null and Status contains SL/T1/T2/EXIT)
    closed = outcome[outcome["Return%"].notna()].copy()
    # Coerce Return% to float
    try:
        closed["Return%"] = closed["Return%"].astype(float)
    except Exception:
        pass

    n = len(closed)
    if n < 3:
        ws.append(["Total closed trades", n, "Insufficient sample (need ≥3)"])
        return

    returns = closed["Return%"].values.tolist()
    wins = [r for r in returns if r > 0]
    losses = [r for r in returns if r < 0]

    total_return = sum(returns)
    avg_return = total_return / n
    win_rate = len(wins) / n * 100.0
    avg_win = (sum(wins) / len(wins)) if wins else 0.0
    avg_loss = (sum(losses) / len(losses)) if losses else 0.0
    gross_win = sum(wins)
    gross_loss = abs(sum(losses))

    # Sharpe (rough — using per-trade returns, annualized assuming ~5 trades/wk)
    import math
    if n >= 2:
        mean_r = avg_return
        var_r = sum((r - mean_r) ** 2 for r in returns) / (n - 1)
        std_r = math.sqrt(var_r) if var_r > 0 else 0.0
        # Assume ~200 trades/yr → annualization factor √200 ≈ 14.1
        # But we don't know true frequency — use √52 (weekly proxy) as conservative
        sharpe = (mean_r / std_r * math.sqrt(52)) if std_r > 0 else 0.0
    else:
        sharpe = 0.0

    # Max drawdown from cumulative equity curve
    cum = 0.0
    peak = 0.0
    max_dd = 0.0
    for r in returns:
        cum += r
        if cum > peak:
            peak = cum
        dd = peak - cum
        if dd > max_dd:
            max_dd = dd

    # Expectancy
    expectancy = (win_rate / 100.0 * avg_win) + ((1 - win_rate / 100.0) * avg_loss)

    # Profit factor
    pf = (gross_win / gross_loss) if gross_loss > 0 else float("inf")

    ws.append(["Total closed trades",       n,                     "Sample size"])
    ws.append(["Total return (cumulative)", f"{total_return:.2f}%", "Sum of per-trade returns"])
    ws.append(["Avg return per trade",      f"{avg_return:.2f}%",   ">0.5% is decent"])
    ws.append(["Win rate",                  f"{win_rate:.1f}%",     ">55% is strong"])
    ws.append(["Avg win",                   f"{avg_win:.2f}%",      ""])
    ws.append(["Avg loss",                  f"{avg_loss:.2f}%",     "should be > -3% (stop discipline)"])
    ws.append(["Expectancy per trade",      f"{expectancy:.3f}%",   ">0 = profitable edge"])
    ws.append(["Profit factor",             f"{pf:.2f}" if pf != float('inf') else "∞",
                                                                    ">1.5 healthy, >2.0 excellent"])
    ws.append(["Sharpe ratio (proxy)",      f"{sharpe:.2f}",        ">1.0 good, >2.0 exceptional"])
    ws.append(["Max drawdown",              f"-{max_dd:.2f}%",      "cumulative peak-to-trough"])
    ws.append([])

    # Health score
    health = 0
    if win_rate >= 55: health += 2
    elif win_rate >= 45: health += 1
    if expectancy > 0.5: health += 2
    elif expectancy > 0: health += 1
    if pf >= 1.5 and pf != float("inf"): health += 2
    elif pf >= 1.0: health += 1
    if sharpe >= 1.0: health += 2
    elif sharpe >= 0.5: health += 1
    if max_dd <= 10.0: health += 2
    elif max_dd <= 20.0: health += 1

    verdict = "EXCELLENT" if health >= 8 else "HEALTHY" if health >= 6 else "OK" if health >= 4 else "WEAK"
    ws.append(["Overall health score", f"{health}/10", verdict])


def _analyze_backtest_vs_live(wb, df_rec, df_track):
    """Compare live pipeline metrics vs backtest_walkforward.py results.

    Reads backtest results from backtest_results.json or backtest_results.xlsx
    (if present). If no backtest artefact is found, notes it as a stub.
    """
    ws = wb["Backtest vs Live"]
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None

    ws.append(["BACKTEST vs LIVE DIVERGENCE"])
    ws.append([])
    ws.append(["Metric", "Backtest", "Live", "Delta", "Note"])

    # Try to load backtest results — check common paths
    bt_data = None
    for candidate in ("backtest_results.json", "backtest_summary.json",
                      "walkforward_results.json"):
        if os.path.exists(candidate):
            try:
                with open(candidate, "r") as f:
                    bt_data = json.load(f)
                break
            except Exception:
                continue

    if bt_data is None:
        ws.append(["(no backtest artefact found)", "—", "—", "—",
                   "Run backtest_walkforward.py and export results.json"])
        return

    # Extract live metrics from df_track / df_rec
    outcome = _merge_outcomes(df_rec, df_track)
    if outcome is None or outcome.empty:
        ws.append(["(no live data)", "—", "—", "—", "Live pipeline empty"])
        return

    closed = outcome[outcome["Return%"].notna()].copy()
    try:
        closed["Return%"] = closed["Return%"].astype(float)
    except Exception:
        pass

    n_live = len(closed)
    if n_live < 5:
        ws.append([f"(n_live={n_live}, need ≥5)", "—", "—", "—", "Insufficient live sample"])
        return

    returns = closed["Return%"].values.tolist()
    wins = [r for r in returns if r > 0]

    live_win_rate = len(wins) / n_live * 100.0
    live_avg = sum(returns) / n_live
    live_sharpe_proxy = 0.0
    import math
    if n_live >= 2:
        mean_r = live_avg
        var_r = sum((r - mean_r) ** 2 for r in returns) / (n_live - 1)
        std_r = math.sqrt(var_r) if var_r > 0 else 0.0
        live_sharpe_proxy = (mean_r / std_r * math.sqrt(52)) if std_r > 0 else 0.0

    bt_win_rate = float(bt_data.get("win_rate", bt_data.get("winrate", 0)) or 0)
    bt_avg = float(bt_data.get("avg_return", bt_data.get("avg_return_pct", 0)) or 0)
    bt_sharpe = float(bt_data.get("sharpe", bt_data.get("sharpe_ratio", 0)) or 0)
    bt_n = int(bt_data.get("n_trades", bt_data.get("trades", 0)) or 0)

    def _fmt_delta(live_v, bt_v):
        d = live_v - bt_v
        sign = "+" if d >= 0 else ""
        return f"{sign}{d:.2f}"

    def _note(live_v, bt_v, threshold=5.0):
        d = live_v - bt_v
        if abs(d) < threshold:
            return "aligned"
        return "LIVE BETTER" if d > 0 else "LIVE WORSE — investigate"

    ws.append(["Trades (n)",  bt_n,                    n_live,               n_live - bt_n, ""])
    ws.append(["Win rate %",  f"{bt_win_rate:.1f}",    f"{live_win_rate:.1f}",
                              _fmt_delta(live_win_rate, bt_win_rate),
                              _note(live_win_rate, bt_win_rate, 5)])
    ws.append(["Avg return %", f"{bt_avg:.2f}",         f"{live_avg:.2f}",
                              _fmt_delta(live_avg, bt_avg),
                              _note(live_avg, bt_avg, 0.3)])
    ws.append(["Sharpe",      f"{bt_sharpe:.2f}",      f"{live_sharpe_proxy:.2f}",
                              _fmt_delta(live_sharpe_proxy, bt_sharpe),
                              _note(live_sharpe_proxy, bt_sharpe, 0.3)])
    ws.append([])
    ws.append(["Interpretation:"])
    ws.append(["• 'LIVE WORSE' on 2+ rows → gate calibration drift; re-run backtest and update thresholds."])
    ws.append(["• 'LIVE BETTER' consistently → gates may be too permissive; sample bias possible."])


if __name__ == "__main__":
    run_research()
