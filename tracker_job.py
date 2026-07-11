"""
tracker_job.py — Recommendation Tracker (JOB 2)
================================================
Run separately every trading day after market close.
GitHub Actions: schedule weekdays at 4:30 PM IST (11:00 UTC).

Reads:  shadow_master.xlsx (created by main.py daily scanner)
Writes: shadow_master.xlsx — Daily Tracking sheet + Performance Summary

Usage:
    python tracker_job.py
"""

import os
import sys
import numpy as np
from datetime import datetime, date

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

try:
    import yfinance as yf
    _YF_OK = True
except ImportError:
    _YF_OK = False
    print("[ERROR] yfinance not installed. Run: pip install yfinance")

try:
    import openpyxl
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False
    print("[ERROR] openpyxl not installed. Run: pip install openpyxl")

TRACKER_XLSX  = os.getenv("TRACKER_XLSX", "shadow_master.xlsx")
TRACKING_DAYS = 60  # track for 60 trading days after recommendation
# Only write new tracking rows when triggered by GitHub Actions cron schedule
IS_SCHEDULED  = os.getenv("SCHEDULED_RUN", "false").lower() == "true"
# Phase C7c: FRESH_START flag — skip if set, main.py owns the reset for that run
FRESH_START   = os.getenv("FRESH_START", "false").lower() == "true"
# Phase C7e: sync stops from main.py's V1 tracker (partial-exit trailed stops).
# V1 tracker.json is the source of truth for partial-exit state — main.py's
# update_tracker() writes 'partial_closed'/'stop' onto entries there. The V2
# tracker (trade_tracker.json) does NOT carry partial-exit fields.
TRACKER_FILE = os.getenv("TRACKER_FILE", "tracker.json")


# ── Phase C7e (2026-07-02): partial-exit stop sync ──
# After T1 hits, main.py's update_tracker() flips 'partial_closed'=True and
# trails the stop up to entry price for the remaining 50%. tracker_job MUST
# use that trailed stop (not the original xlsx stop) or it will misclassify
# healthy T1-then-pullback trades as ACTIVE when the trailed stop was
# actually hit → wrong status → wrong Performance Summary.
def _load_partial_exit_stops() -> dict:
    """Return {symbol: {'stop': trailed_stop, 'partial_closed': True}} from
    the V1 tracker.json for OPEN positions that have hit T1 (partial exit).
    V1 tracker.json is a flat list of entries.
    Empty dict on any failure — non-fatal."""
    if FRESH_START:
        return {}
    if not os.path.exists(TRACKER_FILE):
        return {}
    try:
        import json as _json
        with open(TRACKER_FILE, "r", encoding="utf-8") as f:
            entries = _json.load(f)
    except Exception as e:
        print(f"[WARN] Could not read {TRACKER_FILE}: {e}")
        return {}

    # V1 tracker is a flat list; be tolerant of dict wrappers if ever added
    if isinstance(entries, dict):
        entries = entries.get("entries") or entries.get("open") or []
    if not isinstance(entries, list):
        return {}

    result = {}
    for e in entries:
        if not isinstance(e, dict):
            continue
        if e.get("status") != "OPEN":
            continue
        # Include any post-entry ratchet: partial-closed (T1 booked) OR
        # runner-active (post-T2 riding chandelier). Both raise the trailed
        # stop above the original xlsx stop, and both must be respected here
        # to avoid classifying a healthy pullback as STOPPED.
        if not (e.get("partial_closed") or e.get("runner_active")):
            continue
        sym = str(e.get("symbol", ""))
        if not sym:
            continue
        result[sym] = {
            "stop":           float(e.get("stop", 0) or 0),
            "partial_closed": bool(e.get("partial_closed", False)),
            "runner_active":  bool(e.get("runner_active",  False)),
            "t1_hit_date":    e.get("partial_exit_date") or e.get("t1_hit_date", ""),
            "t2_hit_date":    e.get("t2_hit_date", ""),
        }
    return result


# ── Phase 5 (2026-07-01): live delivery % probe for active positions ──
# Import lazily so a broken nselib does NOT crash the whole tracker; delivery
# insight is a nice-to-have overlay, not a blocker.
def _fetch_live_delivery_pct(symbol: str) -> dict:
    """Return {'today': float, '20d_avg': float, 'signal': str} for a symbol.

    Uses main.fetch_delivery_cached (24h TTL, shared cache) if available.
    Returns empty dict on any failure — caller must handle gracefully.
    """
    try:
        from main import fetch_delivery_cached, load_delivery_cache  # type: ignore
        cache = load_delivery_cache()
        d = fetch_delivery_cached(symbol.replace(".NS", ""), cache)
        if not d or d.get("source") != "nselib":
            return {}
        return {
            "today":   float(d.get("delivery_pct_today", 0.0) or 0.0),
            "20d_avg": float(d.get("delivery_pct_20d_avg", 0.0) or 0.0),
            "ratio":   float(d.get("delivery_ratio", 1.0) or 1.0),
            "signal":  str(d.get("delivery_signal", "NEUTRAL")),
        }
    except Exception:
        return {}


def _flag_position_health(sym: str, cur_return: float, deliv: dict) -> str:
    """Return a short human-readable health tag for an active position.

    Signals to watch:
      - DISTRIBUTION on a winner  → book profit / trim early.
      - WEAK          on a loser  → institutional sellers, cut loss.
      - STRONG_ACCUM  on a winner → let it run.
    """
    if not deliv:
        return ""
    sig = deliv.get("signal", "NEUTRAL")
    if sig == "DISTRIBUTION":
        if cur_return > 5:
            return "⚠ DISTRIBUTION on winner — consider trimming"
        return "⚠ DISTRIBUTION — institutional selling"
    if sig == "WEAK" and cur_return < -3:
        return "⚠ WEAK delivery on loser — cut / tighten stop"
    if sig == "STRONG_ACCUM" and cur_return > 0:
        return "✓ STRONG ACCUM — let it run"
    return ""


def _detect_fresh_start_marker(today_str: str) -> bool:
    """Phase C7f (2026-07-07): auto-detect a wipe main.py did earlier today.

    main.py writes .fresh_start_marker containing today's date when it runs
    with FRESH_START=true. tracker.yml checkout can silently pull the
    pre-wipe xlsx from git (main.yml's persist step doesn't always land
    before tracker.yml runs), so tracker MUST honor this marker or it will
    append fresh rows onto stale rows and undo the reset.

    Returns True if a marker for TODAY is present. Deletes the marker so
    it fires exactly once (the next tracker run will be normal).
    Any read/parse error is treated as "no marker" — non-fatal.
    """
    marker = ".fresh_start_marker"
    if not os.path.exists(marker):
        return False
    try:
        with open(marker, "r", encoding="utf-8") as _fm:
            marker_date = _fm.read().strip()
    except Exception as e:
        print(f"[FRESH_START] Could not read {marker}: {e} — ignoring")
        return False
    if marker_date != today_str:
        # Stale marker from a previous day (should not happen — main.py
        # rewrites it on every FRESH_START run). Remove it and continue.
        print(f"[FRESH_START] Stale marker date={marker_date} (today={today_str}) — removing")
        try:
            os.remove(marker)
        except OSError:
            pass
        return False
    # Marker is for today — consume it so we don't skip tomorrow too.
    try:
        os.remove(marker)
        print(f"[FRESH_START] Consumed .fresh_start_marker for {today_str}")
    except OSError as e:
        print(f"[FRESH_START] Could not delete marker (non-fatal): {e}")
    return True


def _resolve_run_date() -> str:
    """Return the run date as YYYY-MM-DD.

    BUG-D fix: honor the SHADOW_RUN_DATE env var so the 30-day integration
    harness (and any future backfill script) can pin the "today" clock
    to a simulated date. Falls back to wall-clock in production.
    """
    override = os.getenv("SHADOW_RUN_DATE", "").strip()
    if override:
        try:
            datetime.strptime(override, "%Y-%m-%d")
            return override
        except ValueError:
            pass
    return datetime.now().strftime("%Y-%m-%d")


def _resolve_run_dt() -> datetime:
    """Same as _resolve_run_date but returns a full datetime."""
    override = os.getenv("SHADOW_RUN_DATE", "").strip()
    if override:
        try:
            d = datetime.strptime(override, "%Y-%m-%d")
            # Keep current wall-clock time-of-day so distinct runs on the same
            # simulated date still get ordered timestamps.
            now = datetime.now()
            return d.replace(hour=now.hour, minute=now.minute, second=now.second)
        except ValueError:
            pass
    return datetime.now()


def _ts_str() -> str:
    """Formatted 'Last Updated' timestamp respecting SHADOW_RUN_DATE."""
    return _resolve_run_dt().strftime("%Y-%m-%d %H:%M")


def _resolve_shadow_csv_path() -> str | None:
    """Locate the shadow_trades.csv that pairs with TRACKER_XLSX.

    Priority: env SHADOW_CSV_PATH → sibling of TRACKER_XLSX → cwd default.
    Returns None if no candidate is a readable file.
    """
    envp = os.getenv("SHADOW_CSV_PATH", "").strip()
    if envp and os.path.exists(envp):
        return envp
    # Sibling of the tracker xlsx
    try:
        sib = os.path.join(os.path.dirname(os.path.abspath(TRACKER_XLSX)),
                           "shadow_trades.csv")
        if os.path.exists(sib):
            return sib
    except Exception:
        pass
    if os.path.exists("shadow_trades.csv"):
        return "shadow_trades.csv"
    return None


# ─────────────────────────────────────────────────────────────────────────
# BUG-H + BUG-I: close-out CSV writeback
# When the tracker terminates a position (T2_HIT / STOPPED / EXPIRED /
# RUNNER_STOPPED), reflect that outcome back into shadow_trades.csv.
# Trade identity = (symbol, bucket) tuple + earliest PENDING date_added
# — this stays stable even when the same symbol is added multiple times
# to different buckets on different dates (BUG-I).
# ─────────────────────────────────────────────────────────────────────────

# Map tracker terminal statuses → shadow_log status enum
_TRACKER_TO_SHADOW_STATUS = {
    "T2_HIT":         "WIN",
    "RUNNER_STOPPED": "WIN",   # runner is by-definition profitable at stop
    "STOPPED":        "LOSS",
    "EXPIRED":        "TIME_EXIT",
}


def _writeback_closed_to_csv(resolved: list) -> None:
    """Persist tracker-terminated trades into shadow_trades.csv.

    resolved: list of dicts with keys {symbol, rec_date, status, exit_date,
              exit_price, r_multiple, days_held}.
    """
    csv_path = _resolve_shadow_csv_path()
    if not csv_path:
        print("[WARN] BUG-H writeback skipped: no shadow_trades.csv found")
        return
    try:
        # Reuse shadow_log's schema-aware I/O so the CSV stays canonical.
        try:
            from shadow_log import _read_all, _write_all, _CSV_COLS  # type: ignore
        except Exception:
            # Fall back to a local import path (script run from Analyse-recommend-track-main/)
            import importlib.util
            _sp = os.path.join(os.path.dirname(os.path.abspath(__file__)), "shadow_log.py")
            _spec = importlib.util.spec_from_file_location("shadow_log", _sp)
            _mod = importlib.util.module_from_spec(_spec)  # type: ignore
            assert _spec and _spec.loader
            _spec.loader.exec_module(_mod)  # type: ignore
            _read_all = _mod._read_all
            _write_all = _mod._write_all
            _CSV_COLS = _mod._CSV_COLS

        rows = _read_all(csv_path)
        if not rows:
            print(f"[WARN] BUG-H writeback: {csv_path} empty")
            return

        # Index rows by (symbol, bucket) → list of row indices sorted by date_added
        # Only PENDING rows are eligible.
        by_key: dict[tuple, list[int]] = {}
        for i, r in enumerate(rows):
            if str(r.get("status", "")).upper() != "PENDING":
                continue
            k = (str(r.get("symbol", "")).upper(),
                 str(r.get("bucket", "")).upper())
            by_key.setdefault(k, []).append(i)
        for k in by_key:
            by_key[k].sort(key=lambda idx: str(rows[idx].get("date_added", "")))

        # For each resolved trade, find the earliest PENDING match.
        # Tracker doesn't know the bucket → try any bucket for that symbol,
        # preferring the row whose date_added == rec_date. That gives us
        # identity via (symbol, date_added) which is stable across runs.
        n_matched = 0
        for tr in resolved:
            sym    = str(tr["symbol"]).upper()
            rec_dt = str(tr["rec_date"])
            # First, exact-match on (symbol, date_added)
            match_i = None
            for i, r in enumerate(rows):
                if (str(r.get("symbol", "")).upper() == sym and
                        str(r.get("date_added", "")) == rec_dt and
                        str(r.get("status", "")).upper() == "PENDING"):
                    match_i = i
                    break
            # Fallback: earliest PENDING for symbol regardless of bucket
            if match_i is None:
                for k, ilist in by_key.items():
                    if k[0] == sym and ilist:
                        match_i = ilist[0]
                        break
            if match_i is None:
                continue
            new_status = _TRACKER_TO_SHADOW_STATUS.get(str(tr["status"]).upper(), "ERROR")
            rows[match_i]["status"]     = new_status
            rows[match_i]["exit_date"]  = str(tr["exit_date"])
            rows[match_i]["exit_price"] = str(tr["exit_price"])
            rows[match_i]["r_multiple"] = str(tr["r_multiple"])
            rows[match_i]["days_held"]  = str(tr["days_held"])
            rows[match_i]["note"]       = f"tracker:{tr['status']}"
            n_matched += 1

        if n_matched:
            _write_all(csv_path, rows)
            print(f"[INFO] BUG-H: wrote {n_matched} closed outcome(s) back to {csv_path}")
        else:
            print("[INFO] BUG-H: no CSV rows matched resolved trades")
    except Exception as e:
        print(f"[WARN] BUG-H writeback failed: {e}")


def _refresh_summary_bucket_outcomes(wb) -> None:
    """Recompute per-bucket N Closed / Win Rate on the Summary sheet.

    BUG-C: the Summary sheet's bucket-outcome counters were stuck at 0.

    Layout of Summary (product-owned): one row per bucket A/B/C/D with columns
    Bucket, Name, N Total, N Open, N Closed, N WIN_T2, N LOSS, Win Rate %, ...

    Closed trades exist only in bucket A (TAKEN). B/C/D are watch/reject
    lists — they never close, so their counters stay zero.

    Source of truth for closure: Recommendations sheet (Status column).
    Total/Open for each bucket: the bucket-specific sheet row count.
    """
    try:
        if "Summary" not in wb.sheetnames:
            return

        # --- Tally closed trades per bucket from Recommendations sheet -------
        closed_a_win  = 0
        closed_a_loss = 0
        open_a        = 0
        total_a       = 0
        if "Recommendations" in wb.sheetnames:
            ws_rec = wb["Recommendations"]
            rec_headers = [str(c.value or "").strip() for c in
                           next(ws_rec.iter_rows(min_row=1, max_row=1))]
            i_status = next((k for k, h in enumerate(rec_headers)
                            if h.lower() == "status"), -1)
            i_cat    = next((k for k, h in enumerate(rec_headers)
                            if h.lower() == "category"), -1)
            if i_status >= 0:
                for row in ws_rec.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue
                    # Only count BUY (bucket-A) rows
                    if i_cat >= 0 and str(row[i_cat] or "").strip().upper() != "BUY":
                        continue
                    st = str(row[i_status] or "").strip().upper()
                    total_a += 1
                    if st in ("T2_HIT", "RUNNER_STOPPED"):
                        closed_a_win += 1
                    elif st in ("STOPPED", "EXPIRED"):
                        closed_a_loss += 1
                    else:
                        open_a += 1

        # --- Count totals for B/C/D from their bucket sheets -----------------
        totals: dict[str, int] = {"A": total_a, "B": 0, "C": 0, "D": 0}
        for bkt, sn in [("B", "B_WATCH_ME"),
                        ("C", "C_NOT_MY_STYLE"),
                        ("D", "D_SO_CLOSE")]:
            if sn in wb.sheetnames:
                ws_b = wb[sn]
                n = 0
                for row in ws_b.iter_rows(min_row=2, values_only=True):
                    if row and any(c is not None for c in row):
                        n += 1
                totals[bkt] = n

        stats = {
            "A": {"total": total_a, "open": open_a,
                  "closed": closed_a_win + closed_a_loss,
                  "win": closed_a_win, "loss": closed_a_loss},
            "B": {"total": totals["B"], "open": totals["B"],
                  "closed": 0, "win": 0, "loss": 0},
            "C": {"total": totals["C"], "open": totals["C"],
                  "closed": 0, "win": 0, "loss": 0},
            "D": {"total": totals["D"], "open": totals["D"],
                  "closed": 0, "win": 0, "loss": 0},
        }

        # --- Update Summary sheet in place -----------------------------------
        ws_sum = wb["Summary"]
        sum_headers = [str(c.value or "").strip() for c in
                       next(ws_sum.iter_rows(min_row=1, max_row=1))]

        def _col_idx(*aliases: str) -> int:
            for a in aliases:
                for i, h in enumerate(sum_headers):
                    if h.lower() == a.lower():
                        return i + 1  # 1-based
            return -1

        c_bucket = _col_idx("Bucket")
        c_total  = _col_idx("N Total", "Total")
        c_open   = _col_idx("N Open", "Open")
        c_closed = _col_idx("N Closed", "Closed")
        c_win    = _col_idx("N WIN_T2", "N Win", "Wins")
        c_loss   = _col_idx("N LOSS", "N Loss", "Losses")
        c_wr     = _col_idx("Win Rate %", "Win Rate")
        if c_bucket < 0:
            return

        for row in ws_sum.iter_rows(min_row=2):
            bcell = row[c_bucket - 1]
            if bcell.value is None:
                continue
            bkt = str(bcell.value).strip().upper()[:1]  # 'A','B','C','D'
            if bkt not in stats:
                continue
            d = stats[bkt]
            wr = (d["win"] / d["closed"] * 100) if d["closed"] else 0.0
            if c_total  > 0: ws_sum.cell(row=bcell.row, column=c_total,  value=d["total"])
            if c_open   > 0: ws_sum.cell(row=bcell.row, column=c_open,   value=d["open"])
            if c_closed > 0: ws_sum.cell(row=bcell.row, column=c_closed, value=d["closed"])
            if c_win    > 0: ws_sum.cell(row=bcell.row, column=c_win,    value=d["win"])
            if c_loss   > 0: ws_sum.cell(row=bcell.row, column=c_loss,   value=d["loss"])
            if c_wr     > 0: ws_sum.cell(row=bcell.row, column=c_wr,     value=round(wr, 1))
    except Exception as e:
        print(f"[WARN] BUG-C summary refresh failed: {e}")


def run_tracker():
    today_str = _resolve_run_date()  # BUG-D
    print(f"=== TRACKER JOB: {today_str} ===")
    # Phase C7c: FRESH_START safety — if main.py just wiped state, skip tracking
    # today. It'll pick up naturally from tomorrow's run with clean baseline.
    # Phase C7f (2026-07-07): ALSO honor .fresh_start_marker so we detect a
    # wipe even when tracker.yml's own fresh_start input was not set.
    marker_detected = _detect_fresh_start_marker(today_str)
    if FRESH_START or marker_detected:
        reason = "explicit input" if FRESH_START else "detected .fresh_start_marker"
        print(f"[FRESH_START] tracker_job: main.py wiped state this run ({reason}) — skipping tracker update")
        print("[FRESH_START] tracker_job: will resume normal tracking on the next scheduled run")
        return
    print(f"[INFO] Run mode: {'SCHEDULED — will write tracking rows' if IS_SCHEDULED else 'MANUAL — read-only, no rows written'}")

    if not IS_SCHEDULED:
        # Phase W-3 (2026-07-03): LOUD warning when manual-inside-CI.
        # cron-job.org sometimes forgets to send run_mode=scheduled — this
        # produces "green" GitHub Actions runs that silently wrote no rows.
        # We now emit a distinctive banner + set a sentinel file the workflow
        # can check.
        _in_ci = os.getenv("GITHUB_ACTIONS", "").lower() == "true"
        banner_char = "!" if _in_ci else "-"
        print(banner_char * 70)
        print(f"[MANUAL] Tracker running in MANUAL mode ({'INSIDE CI' if _in_ci else 'local'}) — NO rows will be written.")
        print("[MANUAL] If this is a scheduled CI run, the workflow inputs are WRONG.")
        print("[MANUAL] Set inputs.run_mode=scheduled OR env SCHEDULED_RUN=true.")
        print("[MANUAL] Use the scheduled run (4:30 PM IST weekdays) for proper daily tracking.")
        print(banner_char * 70)
        if _in_ci:
            try:
                with open("manual_in_ci.flag", "w", encoding="utf-8") as _f:
                    _f.write("tracker manual-inside-CI at "
                             + datetime.now().isoformat() + "\n")
            except OSError:
                pass
        return

    if not _YF_OK or not _OPENPYXL_OK:
        print("[ERROR] Missing dependencies — aborting")
        return

    # ── Phase W-5 (2026-07-03): yfinance dry-check ──
    # Before touching the xlsx, verify yfinance can actually fetch a known-good
    # symbol. If the API is down, we DO NOT want to write rows with 0-priced
    # positions (that would corrupt Performance Summary + trigger phantom stop
    # hits). Fail loudly and let the workflow's if:failure() notifier fire.
    try:
        _probe = yf.download(
            "RELIANCE.NS", period="5d", progress=False,
            auto_adjust=False, threads=False,
        )
        _probe_ok = _probe is not None and not _probe.empty and "Close" in _probe.columns
    except Exception as e:
        print(f"[YF_PROBE] exception: {e}")
        _probe_ok = False
    if not _probe_ok:
        print("[ERROR] yfinance dry-check FAILED — RELIANCE.NS returned empty/None")
        print("[ERROR] Aborting tracker to protect data integrity (no rows written)")
        print("[ERROR] The shadow_master.xlsx is UNCHANGED — retry when yf is up")
        # Emit a sentinel file the workflow can detect for its Telegram alert
        try:
            with open("yfinance_down.flag", "w", encoding="utf-8") as _f:
                _f.write(f"yfinance dry-check failed at {datetime.now().isoformat()}\n")
        except OSError:
            pass
        return
    print("[INFO] yfinance dry-check OK — RELIANCE.NS reachable")

    if not os.path.exists(TRACKER_XLSX):
        print(f"[WARN] No tracker file found at {TRACKER_XLSX} — run daily scanner first")
        return

    try:
        wb       = openpyxl.load_workbook(TRACKER_XLSX)
        ws_rec   = wb["Recommendations"]
        ws_track = wb["Daily Tracking"]
    except Exception as e:
        print(f"[ERROR] Could not open {TRACKER_XLSX}: {e}")
        return

    # ── Read active recommendations ──
    headers = [cell.value for cell in ws_rec[1]]
    active_recs = []
    for row in ws_rec.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        if d.get("Status") == "ACTIVE":
            active_recs.append(d)

    print(f"[INFO] Tracking {len(active_recs)} active recommendations")
    if not active_recs:
        print("[INFO] No active recommendations to track")
        _update_performance_sheet(wb)
        _update_portfolio_risk_sheet(wb)
        _update_benchmark_overlay(wb)
        _update_equity_curve_sheet(wb)
        wb.save(TRACKER_XLSX)
        return

    # Phase C7e: pull trailed stops for partially-closed positions
    partial_stops = _load_partial_exit_stops()
    if partial_stops:
        print(f"[INFO] {len(partial_stops)} partial-closed position(s) — using trailed stops from trade_tracker.json")

    # ── Batch download prices ──
    symbols = list(set(str(r["Ticker"]) for r in active_recs if r.get("Ticker")))
    prices  = {}

    def _scalar(v):
        """Coerce numpy/pandas scalars, 0-d arrays, or 1-element Series to float."""
        try:
            # pandas Series / DataFrame column → take first element
            if hasattr(v, "iloc"):
                v = v.iloc[0]
            # numpy 0-d array or scalar → use .item()
            if hasattr(v, "item"):
                v = v.item()
            return float(v)
        except Exception:
            return float("nan")

    for sym in symbols:
        try:
            df = yf.download(sym, period="5d", interval="1d",
                             progress=False, auto_adjust=True,
                             multi_level_index=False)
            if df is None or len(df) == 0:
                continue
            # Guard: if MultiIndex slipped through (older/newer yfinance),
            # flatten by picking this symbol's column level.
            if hasattr(df.columns, "nlevels") and df.columns.nlevels > 1:
                try:
                    df = df.xs(sym, axis=1, level=-1)
                except Exception:
                    df.columns = df.columns.get_level_values(0)
            close_val = _scalar(df["Close"].iloc[-1])
            high_val  = _scalar(df["High"].iloc[-1])
            low_val   = _scalar(df["Low"].iloc[-1])
            vol_val   = _scalar(df["Volume"].iloc[-1])
            max_close = _scalar(df["Close"].max())
            min_close = _scalar(df["Close"].min())
            # Skip if any core price came back NaN
            if any(np.isnan(x) for x in (close_val, high_val, low_val, max_close, min_close)):
                print(f"[WARN] Price fetch failed for {sym}: NaN in OHLC")
                continue
            prices[sym] = {
                "close": close_val,
                "high":  high_val,
                "low":   low_val,
                "vol":   vol_val if not np.isnan(vol_val) else 0.0,
                "max_close": max_close,
                "min_close": min_close,
            }
        except Exception as e:
            print(f"[WARN] Price fetch failed for {sym}: {e}")

    # ── Write tracking rows ──
    rows_added = 0
    # BUG-H: collect trades that closed on this run so we can write back to CSV
    resolved_this_run: list[dict] = []
    for rec in active_recs:
        sym = str(rec.get("Ticker", ""))
        px  = prices.get(sym)
        if not px:
            continue

        rec_date = rec.get("Date", "")
        try:
            # BUG-D: use the simulated run date, not wall-clock, so days_held
            # is correct when the harness backfills 30 sim days.
            # BUG-O1 fix: max(0, ...) so a stray future-dated Rec Date (data
            # entry error, clock skew across CI runners) never yields a
            # negative days_held that trips the max-holding logic backwards.
            _today_d = datetime.strptime(today_str, "%Y-%m-%d").date()
            days_held = max(0, (_today_d -
                         datetime.strptime(str(rec_date), "%Y-%m-%d").date()).days)
        except Exception:
            days_held = 0

        entry = float(rec.get("Entry") or 0)
        stop  = float(rec.get("Stop") or 0)
        t1    = float(rec.get("T1") or 0)
        t2    = float(rec.get("T2") or 0)

        # Phase C7e / E1: if this position hit T1 OR is post-T2 running, use
        # the trailed stop (raised by main.py's update_tracker) instead of the
        # original stop from Recommendations. Otherwise stopped_hit will
        # misfire on healthy pullbacks.
        _partial = partial_stops.get(sym)
        if _partial and _partial.get("stop", 0) > stop:
            _new_stop = _partial["stop"]
            _tag = "RUNNER" if _partial.get("runner_active") else "TRAILED"
            print(f"  [{_tag}] {sym}: stop trailed from {stop:.2f} → {_new_stop:.2f}")
            stop = _new_stop

        cur_return  = round((px["close"] - entry) / entry * 100, 2) if entry > 0 else 0
        max_gain    = round((px["max_close"] - entry) / entry * 100, 2) if entry > 0 else 0
        max_dd      = round((px["min_close"] - entry) / entry * 100, 2) if entry > 0 else 0
        remain_up   = round((t2 - px["close"]) / px["close"] * 100, 1) if t2 > px["close"] > 0 else 0

        t1_hit   = px["high"] >= t1   if t1 > 0 else False
        t2_hit   = px["high"] >= t2   if t2 > 0 else False
        stop_hit = px["low"]  <= stop if stop > 0 else False

        # Phase E1a: post-T2 runner mode. If main.py has marked this position
        # as runner_active (i.e., T2 was hit and 50% booked, rest is riding
        # the chandelier trail), the tracker status is RUNNER, not T2_HIT.
        # If the trailed stop is finally hit, the runner is stopped out —
        # but with a positive P&L, so it's still counted as a win by
        # Performance Summary win-rate math (Return% > 0).
        runner_active = bool(_partial and _partial.get("runner_active"))

        # Determine status
        if runner_active and stop_hit:
            status = "RUNNER_STOPPED"
        elif runner_active:
            status = "RUNNER"
        elif t2_hit:
            status = "T2_HIT"
        elif stop_hit:
            status = "STOPPED"
        elif days_held >= TRACKING_DAYS:
            status = "EXPIRED"
        elif t1_hit:
            status = "T1_HIT_ACTIVE"
        else:
            status = "ACTIVE"

        ws_track.append([
            today_str, sym, str(rec_date), days_held,
            px["close"], px["high"], px["low"], px["vol"],
            cur_return, max_gain, max_dd,
            t1_hit, t2_hit, stop_hit, remain_up, days_held, status
        ])
        rows_added += 1

        # Phase 5: probe live delivery % for active positions — surface
        # institutional-flow signal (DISTRIBUTION / STRONG_ACCUM / WEAK) as
        # an early-exit / hold-longer overlay. Never fatal.
        if status in ("ACTIVE", "T1_HIT_ACTIVE", "RUNNER"):
            _deliv = _fetch_live_delivery_pct(sym)
            _health = _flag_position_health(sym, cur_return, _deliv)
            if _deliv:
                print(
                    f"  [DELIV] {sym:<20} ret {cur_return:+.1f}% · "
                    f"deliv {_deliv.get('today', 0):.0f}% "
                    f"(20d {_deliv.get('20d_avg', 0):.0f}%, ratio {_deliv.get('ratio', 1):.2f}) · "
                    f"{_deliv.get('signal', 'NEUTRAL')}"
                    + (f" · {_health}" if _health else "")
                )

        # Update recommendation status if terminal
        if status in ("T2_HIT", "STOPPED", "EXPIRED", "RUNNER_STOPPED"):
            for row in ws_rec.iter_rows(min_row=2):
                if (str(row[1].value) == sym and
                        str(row[0].value) == str(rec_date)):
                    row[-1].value = status
                    break
            # BUG-H: also stage a CSV writeback for this closed trade.
            # r_multiple = (exit - entry) / (entry - stop)   for longs
            _risk = entry - stop
            r_mult = round((px["close"] - entry) / _risk, 2) if _risk > 0 else 0
            resolved_this_run.append({
                "symbol":    sym,
                "rec_date":  str(rec_date),
                "status":    status,
                "exit_date": today_str,
                "exit_price": round(px["close"], 2),
                "r_multiple": r_mult,
                "days_held":  days_held,
            })

    print(f"[INFO] Added {rows_added} tracking rows")
    wb.save(TRACKER_XLSX)
    _update_performance_sheet(wb, run_date=today_str)

    # ── KELLY parallel tracking (2026-07-11) ──
    # Same OHLC-based tracking as the main loop, but reads from
    # 'KELLY Recommendations' and writes to 'KELLY Daily Tracking'.
    # Reuses the already-fetched `prices` dict when possible — KELLY picks
    # are a subset of buys, so most symbols are already cached. Only
    # missing symbols trigger a new yfinance download.
    try:
        _track_kelly_recommendations(wb, prices, today_str)
        _update_kelly_performance_sheet(wb, run_date=today_str)
    except Exception as _kex:
        print(f"[WARN] KELLY parallel tracking failed (non-fatal): {_kex}")

    # BUG-C: refresh the per-bucket Summary sheet so N Closed / Win Rate
    # reflect closed outcomes (not stuck at 0).
    _refresh_summary_bucket_outcomes(wb)
    # BUG-H: write closed outcomes back to shadow_trades.csv so CSV isn't
    # forever full of PENDING rows.
    if resolved_this_run:
        _writeback_closed_to_csv(resolved_this_run)
    # Phase F7a: portfolio-level risk snapshot (env-gated, non-fatal)
    _update_portfolio_risk_sheet(wb)
    # Phase F9: strategy-vs-NIFTY overlay (env-gated, non-fatal)
    _update_benchmark_overlay(wb)
    # Phase G2: equity curve + Sharpe/Sortino (env-gated, non-fatal)
    _update_equity_curve_sheet(wb)
    wb.save(TRACKER_XLSX)
    print(f"[INFO] Tracker job complete — {TRACKER_XLSX} updated")


def _update_performance_sheet(wb, run_date: str | None = None):
    """Recalculates Performance Summary sheet from tracking data.

    BUG-D fix: honor caller-supplied run_date for the Last Updated cell.
    """
    try:
        ws_track = wb["Daily Tracking"]
        ws_perf  = wb["Performance Summary"]

        # Clear existing (keep header)
        for row in ws_perf.iter_rows(min_row=2):
            for cell in row:
                cell.value = None

        # Read all tracking records
        headers = [cell.value for cell in ws_track[1]]
        records = [dict(zip(headers, [cell.value for cell in row]))
                   for row in ws_track.iter_rows(min_row=2)]

        # Group by ticker + rec_date for final outcomes
        outcomes = {}
        for r in records:
            key = f"{r.get('Ticker')}_{r.get('Rec Date')}"
            day = int(r.get("Day#") or 0)
            if key not in outcomes or day > int(outcomes[key].get("Day#") or 0):
                outcomes[key] = r

        closed = [o for o in outcomes.values()
                  if o.get("Status") not in ("ACTIVE", "T1_HIT_ACTIVE", "RUNNER")]
        wins   = [o for o in closed if float(o.get("Return%") or 0) > 0]
        losses = [o for o in closed if float(o.get("Return%") or 0) <= 0]

        # Phase E1: runner-mode metrics
        runners_active = [o for o in outcomes.values() if o.get("Status") == "RUNNER"]
        runners_closed = [o for o in closed if o.get("Status") == "RUNNER_STOPPED"]
        t2_hit_closed  = [o for o in closed if o.get("Status") == "T2_HIT"]

        stats = [
            ("Total Tracked",   len(outcomes)),
            ("Closed",          len(closed)),
            ("Active",          len(outcomes) - len(closed)),
            ("Runners Active",  len(runners_active)),
            ("Win Rate %",      round(len(wins)/len(closed)*100, 1) if closed else 0),
            ("Avg Return %",    round(np.mean([float(o.get("Return%") or 0) for o in closed]), 2) if closed else 0),
            ("Avg Win %",       round(np.mean([float(o.get("Return%") or 0) for o in wins]), 2) if wins else 0),
            ("Avg Loss %",      round(np.mean([float(o.get("Return%") or 0) for o in losses]), 2) if losses else 0),
            ("Avg Runner Ret%", round(np.mean([float(o.get("Return%") or 0) for o in runners_closed]), 2) if runners_closed else 0),
            ("Avg T2-Exit Ret%",round(np.mean([float(o.get("Return%") or 0) for o in t2_hit_closed]), 2) if t2_hit_closed else 0),
            ("Avg Max Gain %",  round(np.mean([float(o.get("Max Gain%") or 0) for o in closed]), 2) if closed else 0),
            ("Avg Max DD %",    round(np.mean([float(o.get("Max DD%") or 0) for o in closed]), 2) if closed else 0),
            ("T1 Hit Rate %",   round(sum(1 for o in closed if o.get("T1 Hit"))/len(closed)*100, 1) if closed else 0),
            ("T2 Hit Rate %",   round(sum(1 for o in closed if o.get("T2 Hit"))/len(closed)*100, 1) if closed else 0),
            ("Stop Hit Rate %", round(sum(1 for o in closed if o.get("Stop Hit"))/len(closed)*100, 1) if closed else 0),
            ("Last Updated",    _ts_str()),
        ]

        ws_perf["A1"] = "Metric"
        ws_perf["B1"] = "Value"
        for i, (metric, value) in enumerate(stats, start=2):
            ws_perf[f"A{i}"] = metric
            ws_perf[f"B{i}"] = value

        # ── Phase F8: P&L Attribution — where does alpha come from? ──────
        # Adds a mini-table below the main stats splitting Avg Return by exit
        # bucket. All env-gated by PNL_ATTRIBUTION_ROWS in main.py.
        try:
            _pnl_attribution_ok = os.getenv("PNL_ATTRIBUTION_ROWS", "true").lower() == "true"
        except Exception:
            _pnl_attribution_ok = True
        if _pnl_attribution_ok and closed:
            buckets = {
                "T2_HIT":         [o for o in closed if o.get("Status") == "T2_HIT"],
                "RUNNER_STOPPED": [o for o in closed if o.get("Status") == "RUNNER_STOPPED"],
                "STOPPED":        [o for o in closed if o.get("Status") == "STOPPED"],
                "EXPIRED":        [o for o in closed if o.get("Status") == "EXPIRED"],
            }
            start_row = len(stats) + 4  # gap after main stats
            ws_perf[f"A{start_row}"]   = "── P&L Attribution ──"
            ws_perf[f"A{start_row+1}"] = "Exit bucket"
            ws_perf[f"B{start_row+1}"] = "Count"
            ws_perf[f"C{start_row+1}"] = "Avg Return %"
            ws_perf[f"D{start_row+1}"] = "% of trades"
            r = start_row + 2
            for name, bucket in buckets.items():
                cnt = len(bucket)
                avg = round(np.mean([float(o.get("Return%") or 0) for o in bucket]), 2) if bucket else 0.0
                pct = round(cnt / len(closed) * 100, 1) if closed else 0.0
                ws_perf[f"A{r}"] = name
                ws_perf[f"B{r}"] = cnt
                ws_perf[f"C{r}"] = avg
                ws_perf[f"D{r}"] = f"{pct}%"
                r += 1

    except Exception as e:
        print(f"[WARN] Performance sheet update failed: {e}")


# ═════════════════════════════════════════════════════════════════════════════
# Phase F7a: Portfolio Risk Sheet
# ═════════════════════════════════════════════════════════════════════════════
def _update_portfolio_risk_sheet(wb):
    """Write a top-down risk snapshot to a 'Portfolio Risk' sheet.

    Columns: Total exposure %, top-sector concentration, rolling drawdown 20d,
    open-position count, active runner count, VaR-95 estimate (parametric
    from daily-return std × sqrt(N holdings)).

    Fully read-only from the tracker perspective — never touches trading logic.
    """
    if os.getenv("PORTFOLIO_RISK_SHEET", "true").lower() != "true":
        return
    try:
        # Non-fatal: nothing to summarise if there's no tracking sheet yet.
        if "Daily Tracking" not in wb.sheetnames:
            return
        ws_track = wb["Daily Tracking"]
        # Create or refresh Portfolio Risk sheet
        if "Portfolio Risk" in wb.sheetnames:
            ws = wb["Portfolio Risk"]
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None
        else:
            ws = wb.create_sheet("Portfolio Risk")

        # Read all rows into per-position latest-day map
        headers = [cell.value for cell in ws_track[1]]
        records = [dict(zip(headers, [cell.value for cell in row]))
                   for row in ws_track.iter_rows(min_row=2)]

        open_map = {}   # {ticker+rec_date: latest_row}
        for r in records:
            if r.get("Status") in ("ACTIVE", "T1_HIT_ACTIVE", "RUNNER"):
                key = f"{r.get('Ticker')}_{r.get('Rec Date')}"
                day = int(r.get("Day#") or 0)
                if key not in open_map or day > int(open_map[key].get("Day#") or 0):
                    open_map[key] = r

        opens = list(open_map.values())
        n_open   = len(opens)
        n_runner = sum(1 for o in opens if o.get("Status") == "RUNNER")
        # daily returns of open positions (Return% column is trade-to-date;
        # daily proxy uses (close - open)/open for the latest bar)
        returns = [float(o.get("Return%") or 0) for o in opens]
        avg_ret = round(np.mean(returns), 2)  if returns else 0.0
        std_ret = round(np.std(returns),  2)  if len(returns) >= 2 else 0.0

        # Total exposure (position_pct field lives in v1/v2 tracker — approximate
        # here from n_open × 5% legacy default; tracker_job doesn't have v2 access)
        total_exposure = round(n_open * 5.0, 1)  # default 5% per position

        # 20d rolling drawdown from Return% distribution (approximation)
        rolling_dd = round(min(returns), 2) if returns else 0.0

        # Parametric VaR-95: 1.65 × std × sqrt(N), rough one-day loss estimate
        var95 = round(1.65 * std_ret * (n_open ** 0.5), 2) if n_open else 0.0

        stats = [
            ("Open Positions",       n_open),
            ("Active Runners",       n_runner),
            ("Total Exposure %",     total_exposure),
            ("Avg Return% (open)",   avg_ret),
            ("Std Return% (open)",   std_ret),
            ("Worst Open Return %",  rolling_dd),
            ("VaR-95 (1-day, %)",    var95),
            ("Last Updated",         _ts_str()),
        ]
        ws["A1"] = "Portfolio-level Risk Snapshot"
        for i, (metric, value) in enumerate(stats, start=3):
            ws[f"A{i}"] = metric
            ws[f"B{i}"] = value

        # Sector concentration (top 5 tickers)
        ws["A13"] = "── Top-5 Positions by symbol ──"
        ws["A14"] = "Ticker"
        ws["B14"] = "Days Held"
        ws["C14"] = "Return %"
        ws["D14"] = "Status"
        opens_sorted = sorted(opens, key=lambda o: float(o.get("Return%") or 0), reverse=True)
        for idx, o in enumerate(opens_sorted[:5], start=15):
            ws[f"A{idx}"] = o.get("Ticker")
            ws[f"B{idx}"] = o.get("Day#")
            ws[f"C{idx}"] = float(o.get("Return%") or 0)
            ws[f"D{idx}"] = o.get("Status")
    except Exception as e:
        print(f"[WARN] Portfolio risk sheet failed: {e}")


# ═════════════════════════════════════════════════════════════════════════════
# Phase F9: Strategy vs NIFTY benchmark overlay
# ═════════════════════════════════════════════════════════════════════════════
def _update_benchmark_overlay(wb):
    """Add a 'Benchmark' sheet with cumulative NIFTY vs strategy return.

    Pulls NIFTY 1-year daily close via yfinance, computes cumulative %,
    and compares to the strategy's cumulative Avg Return% from closed trades.
    Read-only reporting — never touches trading logic.
    """
    if os.getenv("BENCHMARK_OVERLAY", "true").lower() != "true":
        return
    try:
        if not _YF_OK:
            return
        symbol = os.getenv("BENCHMARK_SYMBOL", "^NSEI")
        import yfinance as yf
        # Silence yfinance progress
        df = yf.download(symbol, period="1y", progress=False, auto_adjust=True,
                         multi_level_index=False)
        if df is None or df.empty:
            return

        closes = df["Close"].dropna()
        if len(closes) < 2:
            return

        def _px_scalar(x):
            """Coerce yfinance scalar/0-d array/Series-of-1 → float."""
            try:
                if hasattr(x, "iloc"):
                    x = x.iloc[0] if len(x) > 0 else x
                if hasattr(x, "item"):
                    x = x.item()
                return float(x)
            except Exception:
                return float("nan")

        first = _px_scalar(closes.iloc[0])
        last  = _px_scalar(closes.iloc[-1])
        if not first or np.isnan(first) or np.isnan(last):
            return
        nifty_ytd = round((last - first) / first * 100, 2)

        # Strategy cumulative avg return from closed trades
        if "Daily Tracking" not in wb.sheetnames:
            return
        ws_track = wb["Daily Tracking"]
        headers = [cell.value for cell in ws_track[1]]
        records = [dict(zip(headers, [cell.value for cell in row]))
                   for row in ws_track.iter_rows(min_row=2)]
        outcomes = {}
        for r in records:
            key = f"{r.get('Ticker')}_{r.get('Rec Date')}"
            day = int(r.get("Day#") or 0)
            if key not in outcomes or day > int(outcomes[key].get("Day#") or 0):
                outcomes[key] = r
        closed = [o for o in outcomes.values()
                  if o.get("Status") not in ("ACTIVE", "T1_HIT_ACTIVE", "RUNNER")]
        strategy_ytd = round(np.mean([float(o.get("Return%") or 0) for o in closed]), 2) if closed else 0.0
        alpha = round(strategy_ytd - nifty_ytd, 2)

        if "Benchmark" in wb.sheetnames:
            ws = wb["Benchmark"]
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None
        else:
            ws = wb.create_sheet("Benchmark")

        ws["A1"] = "Strategy vs Benchmark (1Y)"
        ws["A3"] = "Benchmark symbol"
        ws["B3"] = symbol
        ws["A4"] = f"{symbol} 1Y return %"
        ws["B4"] = nifty_ytd
        ws["A5"] = "Strategy avg return % (closed)"
        ws["B5"] = strategy_ytd
        ws["A6"] = "Alpha (strategy − benchmark)"
        ws["B6"] = alpha
        ws["A7"] = "Verdict"
        if alpha > 5:
            ws["B7"] = "✓ Strong alpha"
        elif alpha > 0:
            ws["B7"] = "✓ Positive alpha"
        elif alpha > -5:
            ws["B7"] = "⚠ Near-benchmark"
        else:
            ws["B7"] = "✗ Underperforming"
        ws["A8"] = "Last Updated"
        ws["B8"] = _ts_str()
    except Exception as e:
        print(f"[WARN] Benchmark overlay failed: {e}")


# ═════════════════════════════════════════════════════════════════════════════
# Phase G2: Equity Curve + Sharpe / Sortino in tracker.xlsx
# ═════════════════════════════════════════════════════════════════════════════
def _update_equity_curve_sheet(wb):
    """Write an 'Equity Curve' sheet with:
        - date-ordered closed-trade P&L
        - cumulative equity curve (₹ from a starting cash of EQUITY_START)
        - rolling Sharpe / Sortino / max-drawdown %
        - a line chart plotting the curve

    Read-only reporting; never touches trading logic. Env-gated by
    EQUITY_CURVE_SHEET (default TRUE).
    """
    if os.getenv("EQUITY_CURVE_SHEET", "true").lower() != "true":
        return
    try:
        if "Daily Tracking" not in wb.sheetnames:
            return
        ws_track = wb["Daily Tracking"]
        headers = [cell.value for cell in ws_track[1]]
        records = [dict(zip(headers, [cell.value for cell in row]))
                   for row in ws_track.iter_rows(min_row=2)]

        # Latest row per position → outcome
        outcomes = {}
        for r in records:
            key = f"{r.get('Ticker')}_{r.get('Rec Date')}"
            day = int(r.get("Day#") or 0)
            if key not in outcomes or day > int(outcomes[key].get("Day#") or 0):
                outcomes[key] = r

        closed = [
            o for o in outcomes.values()
            if o.get("Status") not in ("ACTIVE", "T1_HIT_ACTIVE", "RUNNER")
        ]
        # Sort by the tracking date on which the close was observed
        def _close_date(o):
            v = o.get("Date") or o.get("Rec Date") or ""
            try:
                return datetime.strptime(str(v)[:10], "%Y-%m-%d")
            except Exception:
                return datetime.min
        closed.sort(key=_close_date)

        equity_start = float(os.getenv("EQUITY_START", "500000") or 500000)
        # Each trade allocates POSITION_PCT of equity; return applies to that slice
        position_pct = float(os.getenv("EQUITY_POSITION_PCT", "5.0") or 5.0) / 100.0

        # Recompute equity curve
        cash = equity_start
        curve = []   # (date_str, ret_pct, pnl_rupees, cash_after)
        rets = []
        for o in closed:
            r = float(o.get("Return%") or 0)
            pnl = cash * position_pct * (r / 100.0)
            cash += pnl
            ds = str(o.get("Date") or o.get("Rec Date") or "")[:10]
            curve.append((ds, r, round(pnl, 2), round(cash, 2)))
            rets.append(r)

        # Sharpe / Sortino / DD
        n = len(rets)
        if n >= 2:
            mu     = float(np.mean(rets))
            sigma  = float(np.std(rets, ddof=1))
            # daily-risk-free assumed 0; annualise assuming ~ N trades / year
            sharpe = round((mu / sigma) * (n ** 0.5), 2) if sigma > 0 else 0.0
            neg    = [x for x in rets if x < 0]
            downside = float(np.std(neg, ddof=1)) if len(neg) >= 2 else 0.0
            sortino  = round((mu / downside) * (n ** 0.5), 2) if downside > 0 else 0.0
        else:
            mu, sigma, sharpe, sortino = 0.0, 0.0, 0.0, 0.0

        # Max drawdown of the equity curve (%)
        max_dd_pct = 0.0
        peak = equity_start
        for _, _, _, c in curve:
            if c > peak:
                peak = c
            dd = (c - peak) / peak * 100.0 if peak else 0.0
            if dd < max_dd_pct:
                max_dd_pct = dd
        max_dd_pct = round(max_dd_pct, 2)

        # (Re)create sheet
        if "Equity Curve" in wb.sheetnames:
            ws = wb["Equity Curve"]
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None
        else:
            ws = wb.create_sheet("Equity Curve")

        ws["A1"] = "Equity Curve & Risk-Adjusted Return"

        # Stat block
        stats = [
            ("Starting capital (₹)",  round(equity_start, 2)),
            ("Ending capital (₹)",    round(cash, 2)),
            ("Total P&L (₹)",         round(cash - equity_start, 2)),
            ("Total P&L %",           round((cash - equity_start) / equity_start * 100, 2) if equity_start else 0.0),
            ("Trades counted",        n),
            ("Avg trade return %",    round(mu, 2)),
            ("Std trade return %",    round(sigma, 2)),
            ("Sharpe (annualised)",   sharpe),
            ("Sortino (annualised)",  sortino),
            ("Max drawdown %",        max_dd_pct),
            ("Position size % used",  round(position_pct * 100, 2)),
            ("Last Updated",          _ts_str()),
        ]
        for i, (k, v) in enumerate(stats, start=3):
            ws[f"A{i}"] = k
            ws[f"B{i}"] = v

        # Curve table starts at row 17
        header_row = 17
        ws[f"A{header_row}"] = "Date"
        ws[f"B{header_row}"] = "Trade Return %"
        ws[f"C{header_row}"] = "P&L (₹)"
        ws[f"D{header_row}"] = "Equity (₹)"
        for i, (ds, r, pnl, c) in enumerate(curve, start=header_row + 1):
            ws[f"A{i}"] = ds
            ws[f"B{i}"] = round(r, 2)
            ws[f"C{i}"] = pnl
            ws[f"D{i}"] = c

        # Try to attach a chart. openpyxl.chart is optional; if it fails,
        # the numeric table is still there.
        try:
            if n >= 2:
                from openpyxl.chart import LineChart, Reference
                chart = LineChart()
                chart.title  = "Equity Curve (₹)"
                chart.y_axis.title = "Equity"
                chart.x_axis.title = "Trade #"
                data = Reference(ws,
                                 min_col=4, min_row=header_row,
                                 max_col=4, max_row=header_row + n)
                chart.add_data(data, titles_from_data=True)
                chart.height = 10
                chart.width  = 22
                ws.add_chart(chart, "F3")
        except Exception as _ce:
            # BUG-F2 fix: at least log the reason so a missing/broken
            # openpyxl chart module doesn't disappear silently.
            print(f"[WARN] Equity curve chart skipped: {_ce}")
    except Exception as e:
        print(f"[WARN] Equity curve sheet failed: {e}")


if __name__ == "__main__":
    # ── Phase W (2026-07-03): watchdog wrapper ──
    # Record run status to run_health.json so other jobs can detect staleness.
    import subprocess
    _status = "ok"
    _err_msg = ""
    try:
        run_tracker()
    except SystemExit:
        raise  # honor explicit exits
    except Exception as _e:
        _status = "fail"
        _err_msg = str(_e)[:200]
        print(f"[FATAL] tracker_job crashed: {_e}")
        # Re-raise AFTER writing the heartbeat so the workflow marks the run
        # as failed and the Telegram alert fires.
    finally:
        _mode = "scheduled" if IS_SCHEDULED else "manual"
        _extras = [f"fresh_start={str(FRESH_START).lower()}"]
        if _err_msg:
            _extras.append(f"error={_err_msg.replace(' ', '_')[:80]}")
        try:
            # BUG-E: point pipeline_health at the same folder as TRACKER_XLSX
            # so run_health.json lands next to the workbook, not in whichever
            # cwd the caller happened to be in.
            _env = os.environ.copy()
            _env.setdefault(
                "PIPELINE_HEALTH_FILE",
                os.path.join(os.path.dirname(os.path.abspath(TRACKER_XLSX)),
                             "run_health.json"),
            )
            subprocess.run(
                [sys.executable,
                 os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              "scripts", "pipeline_health.py"), "record",
                 "--job", "tracker",
                 "--status", _status,
                 "--mode", _mode,
                 "--extras", *_extras],
                check=False, timeout=15, env=_env,
            )
        except Exception as _pe:
            print(f"[WARN] pipeline_health record failed: {_pe}")

    if _status == "fail":
        # Ensure non-zero exit so the workflow's if:failure() step fires
        sys.exit(1)
