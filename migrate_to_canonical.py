"""One-shot migration from existing state -> results/tracking_store.json.

Reads (never writes) these existing files, in priority order:
  1. shadow_master.xlsx      — Recommendations (BUY/WATCHLIST/MONITOR) + Rejected
  2. trade_tracker.json      — buys[], watchlist[], completed[]
  3. tracker.json            — v1 tracker list
  4. watchlist_persist.json  — for first_seen dates
  5. shadow_trades.csv       — bucket rows (A_TAKEN/B_WATCH_ME/etc)

Backs up ALL relevant files under
    $AgentWorkspace/backups/redesign_pre_migration_YYYYMMDD_HHMMSS/
BEFORE writing the new tracking_store.json.

Rows whose stage cannot be inferred confidently are stamped with
migration_status = NEEDS_REVIEW and appear only in the hidden _NEEDS_REVIEW
sheet, never in user-facing sheets.

Usage:
    python migrate_to_canonical.py --write
    python migrate_to_canonical.py --dry-run   # print summary, no writes
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import shutil
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

_HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(_HERE))

from tracking_store import (  # noqa: E402
    DEFAULT_EVENTS_PATH,
    DEFAULT_STORE_PATH,
    MIGRATION_MIGRATED,
    MIGRATION_NEEDS_REVIEW,
    STATUS_ACTIVE,
    STATUS_STOPPED,
    STATUS_STOPPED_AFTER_T1,
    STATUS_T1_HIT,
    STATUS_T2_HIT,
    TrackingStore,
    make_tracking_id,
)

try:
    from openpyxl import load_workbook  # noqa: E402
    HAVE_OPENPYXL = True
except Exception:  # pragma: no cover
    HAVE_OPENPYXL = False


# ---------------------------------------------------------------------------
# Backup helper
# ---------------------------------------------------------------------------

BACKUP_FILES: Tuple[str, ...] = (
    "shadow_master.xlsx",
    "shadow_report_weekly.xlsx",
    "recommendation_tracker.xlsx",
    "trade_tracker.json",
    "tracker.json",
    "watchlist_persist.json",
    "shadow_trades.csv",
    "confidence_history.json",
    "gate_memory.json",
    "results/watchlist_persist.json",
    "results/trade_tracker.json",
    "results/confidence_history.json",
    "results/gate_memory.json",
    "results/shadow_master.xlsx",
)


def _agent_workspace() -> Path:
    aw = os.environ.get("FORGE_AGENT_WORKSPACE")
    if aw:
        return Path(aw)
    runtime = os.environ.get("FORGE_RUNTIME")
    if runtime:
        return Path(runtime) / "agent_workspace" / (os.environ.get("WORKSPACE_ID") or "default")
    # Fallback — local backups/ folder next to script.
    return _HERE / "_local_backups"


def _backup_dir() -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    root = _agent_workspace() / "backups" / f"redesign_pre_migration_{stamp}"
    root.mkdir(parents=True, exist_ok=True)
    return root


def backup_everything(dry_run: bool = False) -> Path:
    dst = _backup_dir()
    for rel in BACKUP_FILES:
        src = _HERE / rel
        if not src.exists():
            continue
        target = dst / rel
        target.parent.mkdir(parents=True, exist_ok=True)
        if dry_run:
            print(f"[backup] would copy {src} -> {target}")
        else:
            shutil.copy2(src, target)
            print(f"[backup] {src} -> {target}")
    return dst


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _iso(d: Any) -> Optional[str]:
    if d is None or d == "":
        return None
    if isinstance(d, date):
        return d.isoformat()
    s = str(d).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date().isoformat()
        except ValueError:
            continue
    # Fallback: assume already isoformat prefix.
    return s[:10]


def _float(x: Any) -> Optional[float]:
    if x is None or x == "":
        return None
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def _stage_from_category(cat: Any) -> Tuple[str, str]:
    """Map main.py's 'Category' cell to (stage, origin)."""
    c = str(cat or "").strip().upper()
    if c == "BUY":
        return "BUY", "buys_list"
    if c in ("NEAR_MISS", "NEAR_MISS_RISING", "NEAR_MISS_FADING"):
        return "NEAR_MISS", "watchlist_tier"
    if c == "DEVELOPING":
        return "DEVELOPING", "watchlist_tier"
    if c == "MONITOR":
        return "MONITOR", "watchlist_tier"
    if c == "WATCHLIST":
        return "WATCHLIST", "watchlist_tier"
    if c == "REJECTED":
        return "REJECTED", "rejected_list"
    return "", ""


# ---------------------------------------------------------------------------
# Readers
# ---------------------------------------------------------------------------

def read_recommendations_sheet(path: Path) -> List[Dict[str, Any]]:
    if not HAVE_OPENPYXL or not path.exists():
        return []
    rows: List[Dict[str, Any]] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Recommendations" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["Recommendations"]
    hdr = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            hdr = list(row)
            continue
        if not row or all(v is None for v in row):
            continue
        d = {h: v for h, v in zip(hdr or [], row) if h}
        rows.append(d)
    wb.close()
    return rows


def read_rejected_sheet(path: Path) -> List[Dict[str, Any]]:
    if not HAVE_OPENPYXL or not path.exists():
        return []
    rows: List[Dict[str, Any]] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Rejected" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["Rejected"]
    hdr = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            hdr = list(row)
            continue
        if not row or all(v is None for v in row):
            continue
        d = {h: v for h, v in zip(hdr or [], row) if h}
        rows.append(d)
    wb.close()
    return rows


def read_bucket_sheets(path: Path) -> Dict[str, List[Dict[str, Any]]]:
    """Read the 4 bucket sheets. Handles both 21-col and 26-col schemas."""
    if not HAVE_OPENPYXL or not path.exists():
        return {}
    out: Dict[str, List[Dict[str, Any]]] = {}
    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet_name in ("A_TAKEN", "B_WATCH_ME", "C_NOT_MY_STYLE", "D_SO_CLOSE"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        hdr = None
        rows: List[Dict[str, Any]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                hdr = [str(h).strip() if h is not None else "" for h in row]
                continue
            if not row or all(v is None for v in row):
                continue
            d = {h: v for h, v in zip(hdr, row) if h}
            if d:
                rows.append(d)
        if rows:
            out[sheet_name] = rows
    wb.close()
    return out


def read_trade_tracker(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {}
    try:
        with path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError):
        return {}


def read_watchlist_persist(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {}
    try:
        with path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError):
        return {}


def read_shadow_trades(path: Path) -> List[Dict[str, Any]]:
    if not path.exists():
        return []
    rows: List[Dict[str, Any]] = []
    with path.open("r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            rows.append(r)
    return rows


# ---------------------------------------------------------------------------
# Migration core
# ---------------------------------------------------------------------------

def migrate(
    project_root: Path = _HERE,
    output_store: Path = DEFAULT_STORE_PATH,
    events_path: Path = DEFAULT_EVENTS_PATH,
    dry_run: bool = False,
) -> Dict[str, Any]:
    store = TrackingStore(store_path=output_store, events_path=events_path)
    summary = {
        "rows_read": 0,
        "rows_migrated": 0,
        "rows_deduped": 0,
        "needs_review": 0,
        "active_records": 0,
        "resolved_records": 0,
    }

    # 1) Watchlist persist first — gives us authoritative first_seen dates.
    wp = read_watchlist_persist(project_root / "watchlist_persist.json")
    if not wp:
        wp = read_watchlist_persist(project_root / "results" / "watchlist_persist.json")
    first_seen_by_symbol: Dict[str, str] = {}
    for sym, entry in (wp or {}).items():
        fs = _iso((entry or {}).get("first_seen"))
        if fs:
            first_seen_by_symbol[sym] = fs

    def _fs(sym: str, fallback: Any) -> str:
        return first_seen_by_symbol.get(sym) or _iso(fallback) or date.today().isoformat()

    # 2) Recommendations sheet — the primary source for BUY / watchlist tiers.
    rec_rows = read_recommendations_sheet(project_root / "shadow_master.xlsx")
    summary["rows_read"] += len(rec_rows)
    for r in rec_rows:
        sym = r.get("Ticker") or r.get("Symbol")
        if not sym:
            continue
        stage, origin = _stage_from_category(r.get("Category"))
        as_of = _fs(sym, r.get("Date"))
        if not stage:
            # Unknown category — mark for review.
            tid = make_tracking_id(sym, as_of)
            if tid in store.records:
                summary["rows_deduped"] += 1
                continue
            store.records[tid] = _needs_review_record(sym, as_of, r, origin="rec_sheet")
            summary["needs_review"] += 1
            continue
        # Enrich the source dict with fail_reasons list.
        src = dict(r)
        fr = str(r.get("Fail Reasons") or "").strip()
        src["fail_reasons"] = [x.strip() for x in fr.split(",") if x.strip()] if fr else []
        # Give the record a synthetic setup_type when the source has none
        # (the Recommendations sheet does not carry setup_type — it's per-bucket).
        tid = store.upsert(
            symbol=sym,
            stage=stage,
            entry_snapshot=src,
            origin=origin,
            as_of=as_of,
            current_snapshot=src,
        )
        rec = store.records[tid]
        rec["migration_status"] = MIGRATION_MIGRATED
        summary["rows_migrated"] += 1

    # 3) Rejected sheet — every rejected stock (user-authorised).
    rej_rows = read_rejected_sheet(project_root / "shadow_master.xlsx")
    summary["rows_read"] += len(rej_rows)
    for r in rej_rows:
        sym = r.get("Ticker") or r.get("Symbol")
        if not sym:
            continue
        as_of = _fs(sym, r.get("Date"))
        src = dict(r)
        fr = str(r.get("Fail Reasons") or "").strip()
        src["fail_reasons"] = [x.strip() for x in fr.split(",") if x.strip()] if fr else []
        # Rejected stocks are their own "stage" for the REJECTED sheet filter.
        tid = store.upsert(
            symbol=sym,
            stage="REJECTED",
            entry_snapshot=src,
            origin="rejected_list",
            as_of=as_of,
            current_snapshot=src,
        )
        rec = store.records[tid]
        rec["migration_status"] = MIGRATION_MIGRATED
        summary["rows_migrated"] += 1

    # 4) Bucket sheets from shadow_master.xlsx — supplement setup_type/regime
    #    onto existing records; add new records if a bucket row has a symbol
    #    not covered above.
    buckets = read_bucket_sheets(project_root / "shadow_master.xlsx")
    for sheet_name, rows in buckets.items():
        origin = f"shadow_bucket_{sheet_name[0]}"  # A/B/C/D
        for r in rows:
            sym = r.get("symbol") or r.get("Ticker") or r.get("Symbol")
            if not sym:
                continue
            summary["rows_read"] += 1
            existing_tid = store.find_active(sym)
            entry_date = _iso(r.get("entry_date")) or _fs(sym, None)
            if existing_tid:
                # Enrich existing record — set setup_type/regime if blank,
                # and pick up any T1/T2/stop dates.
                rec = store.records[existing_tid]
                entry = rec.setdefault("entry", {})
                if not entry.get("setup_type"):
                    entry["setup_type"] = r.get("setup_type")
                if not entry.get("regime"):
                    entry["regime"] = r.get("regime")
                # Pull hit dates if present and not already set.
                for src_k, dst_k in (
                    ("t1_hit_date", "t1_hit_date"),
                    ("t2_hit_date", "t2_hit_date"),
                    ("stop_hit_date", "stop_hit_date"),
                ):
                    val = _iso(r.get(src_k))
                    if val and not rec.get(dst_k):
                        rec[dst_k] = val
                        rec[dst_k.replace("_date", "")] = True
                # MFE / MAE seed from prior tracking.
                mfe = _float(r.get("MFE") or r.get("max_favorable_pct"))
                mae = _float(r.get("MAE") or r.get("max_adverse_pct"))
                if mfe is not None and mfe > (rec.get("mfe_pct") or 0.0):
                    rec["mfe_pct"] = round(mfe, 4)
                if mae is not None and mae < (rec.get("mae_pct") or 0.0):
                    rec["mae_pct"] = round(mae, 4)
                # Terminal-status from bucket sheet.
                st = str(r.get("status") or "").upper()
                if st == "WIN_T2":
                    rec["tracking_status"] = STATUS_T2_HIT
                    rec["t2_hit"] = True
                elif st == "LOSS":
                    if rec.get("t1_hit"):
                        rec["tracking_status"] = STATUS_STOPPED_AFTER_T1
                    else:
                        rec["tracking_status"] = STATUS_STOPPED
                    rec["stop_hit"] = True
                elif st == "OPEN" and rec.get("t1_hit"):
                    rec["tracking_status"] = STATUS_T1_HIT
                summary["rows_deduped"] += 1
                continue
            # New occurrence from bucket sheet.
            stage_map = {"A": "BUY", "B": "WATCHLIST", "C": "REJECTED", "D": "NEAR_MISS"}
            stage = stage_map.get(sheet_name[0], "WATCHLIST")
            src = dict(r)
            tid = store.upsert(
                symbol=sym,
                stage=stage,
                entry_snapshot=src,
                origin=origin,
                as_of=entry_date,
                current_snapshot=src,
            )
            rec = store.records[tid]
            rec["migration_status"] = MIGRATION_MIGRATED
            # Seed MFE/MAE.
            mfe = _float(r.get("MFE") or r.get("max_favorable_pct"))
            mae = _float(r.get("MAE") or r.get("max_adverse_pct"))
            if mfe is not None:
                rec["mfe_pct"] = round(mfe, 4)
            if mae is not None:
                rec["mae_pct"] = round(mae, 4)
            summary["rows_migrated"] += 1

    # 5) trade_tracker.json — resolves T1/T2/stop dates and status for real BUYs.
    tt_path = project_root / "trade_tracker.json"
    tt = read_trade_tracker(tt_path)
    for lst_name, default_stage in (("buys", "BUY"), ("watchlist", "WATCHLIST"),
                                    ("completed", "BUY")):
        for r in tt.get(lst_name) or []:
            sym = r.get("symbol")
            if not sym:
                continue
            summary["rows_read"] += 1
            as_of = _iso(r.get("rec_date") or r.get("entry_date")) or _fs(sym, None)
            existing_tid = store.find_active(sym)
            src = {
                "reference_entry_price": r.get("entry") or r.get("entry_price"),
                "t1_price": r.get("target1"),
                "t2_price": r.get("target2"),
                "stop_price": r.get("stop"),
                "confidence": r.get("confidence"),
                "tq": r.get("tq"),
                "rr_ratio": r.get("rr_ratio"),
                "setup_type": r.get("setup_type"),
                "regime": r.get("regime"),
                "sector": r.get("sector"),
            }
            if not existing_tid:
                tid = store.upsert(
                    symbol=sym,
                    stage=default_stage,
                    entry_snapshot=src,
                    origin="tracker_v2_buys" if lst_name == "buys" else "tracker_v2_watchlist",
                    as_of=as_of,
                    current_snapshot=src,
                )
                store.records[tid]["migration_status"] = MIGRATION_MIGRATED
                summary["rows_migrated"] += 1
            else:
                tid = existing_tid
                summary["rows_deduped"] += 1
            rec = store.records[tid]
            # Hit dates.
            for k in ("t1_hit_date", "t2_hit_date", "stop_hit_date"):
                val = _iso(r.get(k))
                if val and not rec.get(k):
                    rec[k] = val
                    rec[k.replace("_date", "")] = True
            # Status mapping.
            st = str(r.get("status") or "").upper()
            if st in ("STOPPED_OUT", "STOP", "LOSS"):
                rec["tracking_status"] = (
                    STATUS_STOPPED_AFTER_T1 if rec.get("t1_hit") else STATUS_STOPPED
                )
                rec["stop_hit"] = True
            elif st in ("T2_HIT", "WIN_T2", "COMPLETED_WIN"):
                rec["tracking_status"] = STATUS_T2_HIT
                rec["t2_hit"] = True
            elif st == "T1_HIT":
                rec["tracking_status"] = STATUS_T1_HIT
                rec["t1_hit"] = True

    # 6) shadow_trades.csv — final catch-all.
    sh = read_shadow_trades(project_root / "shadow_trades.csv")
    for r in sh:
        sym = r.get("symbol")
        if not sym:
            continue
        summary["rows_read"] += 1
        if store.find_active(sym):
            summary["rows_deduped"] += 1
            continue
        bucket = str(r.get("bucket") or "").upper()
        stage_map = {"A": "BUY", "B": "WATCHLIST", "C": "REJECTED", "D": "NEAR_MISS"}
        stage = stage_map.get(bucket[:1], "")
        if not stage:
            # Unknown bucket (e.g. B_LEGACY) — needs review.
            as_of = _iso(r.get("date_added")) or date.today().isoformat()
            tid = make_tracking_id(sym, as_of)
            if tid in store.records:
                summary["rows_deduped"] += 1
                continue
            store.records[tid] = _needs_review_record(sym, as_of, r, origin="shadow_csv")
            summary["needs_review"] += 1
            continue
        as_of = _iso(r.get("date_added")) or _fs(sym, None)
        src = {
            "reference_entry_price": _float(r.get("entry")),
            "t1_price": _float(r.get("target_1")),
            "stop_price": _float(r.get("stop_loss")),
            "confidence": _float(r.get("conf")),
            "setup_type": r.get("setup_type"),
        }
        tid = store.upsert(
            symbol=sym, stage=stage, entry_snapshot=src,
            origin=f"shadow_bucket_{bucket[:1]}",
            as_of=as_of, current_snapshot=src,
        )
        store.records[tid]["migration_status"] = MIGRATION_MIGRATED
        summary["rows_migrated"] += 1

    # Final tally.
    stats = store.stats()
    summary["active_records"] = stats["active"] + stats["t1_hit_active"]
    summary["resolved_records"] = (
        stats["t2_hit"] + stats["stopped"] + stats["stopped_after_t1"]
    )
    summary["needs_review"] = stats["needs_review"] + summary["needs_review"]

    if not dry_run:
        store.save()
        print(f"[migrate] wrote {store.store_path}")
    else:
        print(f"[migrate] dry-run: would write {store.store_path}")

    return summary


def _needs_review_record(sym: str, as_of: str, src: Dict[str, Any], *, origin: str) -> Dict[str, Any]:
    tid = make_tracking_id(sym, as_of)
    return {
        "tracking_id": tid,
        "symbol": sym,
        "first_seen_date": as_of,
        "strategy_version": "UNKNOWN_LEGACY",
        "entry": {"stage": "UNKNOWN", "origin": origin, "raw": src},
        "current": {"stage": "UNKNOWN", "as_of_date": as_of},
        "journey_path": ["UNKNOWN"],
        "stage_change_count": 0,
        "previous_stage": None,
        "last_stage_change_date": as_of,
        "tracking_status": STATUS_ACTIVE,
        "t1_hit": False, "t1_hit_date": None, "days_to_t1": None,
        "t2_hit": False, "t2_hit_date": None, "days_to_t2": None,
        "stop_hit": False, "stop_hit_date": None, "days_to_stop": None,
        "mfe_pct": 0.0, "mae_pct": 0.0,
        "max_price_seen": None, "min_price_seen": None,
        "days_active": 0,
        "migration_status": MIGRATION_NEEDS_REVIEW,
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv: Optional[List[str]] = None) -> int:
    p = argparse.ArgumentParser(description="Migrate legacy state -> tracking_store.json")
    p.add_argument("--dry-run", action="store_true", help="Print summary, no writes")
    p.add_argument("--no-backup", action="store_true", help="Skip backup step (dangerous)")
    p.add_argument("--out", default=str(DEFAULT_STORE_PATH))
    args = p.parse_args(argv)

    if not args.no_backup:
        dst = backup_everything(dry_run=args.dry_run)
        print(f"[backup] destination: {dst}")

    summary = migrate(
        project_root=_HERE,
        output_store=Path(args.out),
        dry_run=args.dry_run,
    )
    print("\nMigration Summary")
    for k, v in summary.items():
        print(f"  {k:22} : {v}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
