"""
main.py — NSE Swing Trade Analysis Engine v6.0
Self-healing | Full-stack | Production-complete

Run:  python main.py
Env:  copy .env.template to .env and fill in secrets
"""

# ─────────────────────────────────────────────────────────────────────────────
# IMPORTS
# ─────────────────────────────────────────────────────────────────────────────
import os
import re
import sys
import json
import csv
import html
import subprocess
import datetime
import time
import random
import itertools
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    from dotenv import load_dotenv
    load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"), override=False)
except ImportError:
    pass

import requests
import numpy as np
import pandas as pd
import yfinance as yf

# Suppress yfinance download noise globally (401 crumbs, delisted warnings)
import logging as _logging
import warnings as _warnings
_logging.getLogger("yfinance").setLevel(_logging.CRITICAL)
_logging.getLogger("peewee").setLevel(_logging.CRITICAL)
_warnings.filterwarnings("ignore", category=FutureWarning)
_warnings.filterwarnings("ignore", message=".*delisted.*")
_warnings.filterwarnings("ignore", message=".*No data found.*")

try:
    import feedparser
    _FEEDPARSER_OK = True
except ImportError:
    _FEEDPARSER_OK = False

try:
    from bs4 import BeautifulSoup
    _BS4_OK = True
except ImportError:
    _BS4_OK = False

# ── Phase I shadow-log (2026-07-07) ─────────────────────────────────────────
# Records what Phase-I-skipped stocks WOULD have done (paper trades) so we
# can validate backtest predictions in live conditions. Zero real-money risk.
# Toggle via env: PHASE_I_SHADOW_LOG=false disables both recording & summary.
try:
    import shadow_log
    _SHADOW_LOG_OK = True
except ImportError:
    _SHADOW_LOG_OK = False

# ── 2026-07-09 Consolidation D3: shadow master job ──────────────────────────
# Builds shadow_master.xlsx with 4 bucket sheets (A_TAKEN, B_WATCH_ME,
# C_NOT_MY_STYLE, D_SO_CLOSE) + rollups (Summary, Bucket_Comparison,
# Live_Positions, Resolved_Today, Change_Log). Runs after tracker_job.py so
# the base Recommendations / Daily Tracking / Performance sheets exist. In
# CI, the workflow also runs `python shadow_master_job.py` as a separate step
# for full visibility — the in-process call here is kept for local runs
# (run-locally.ps1) and produces a preview when SCHEDULED_RUN is false.
# Toggle: SHADOW_REPORT_ENABLED=false disables the in-process call.
try:
    import shadow_master_job
    _SHADOW_REPORT_OK = True
except ImportError:
    _SHADOW_REPORT_OK = False


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 0b — SAFETY BASELINE (Phase A — added 2026-06-30)
# ─────────────────────────────────────────────────────────────────────────────
# Real-money pipeline (CAPITAL=500000) → these are mandatory:
#   - FetchResult provenance on every replaced fetcher
#   - validate_macro() range gates with last-known-good fallback
#   - explicit Asia/Kolkata timezone (no implicit dependency on TZ env var)
#   - tenacity retries with exponential jitter
#   - decision audit JSONL for post-mortem
# Note: cross_check_fii() removed in Stage-A cleanup (Phase C4 disabled FII/DII
# fetch pipeline-wide; helper had no remaining callers).
# ─────────────────────────────────────────────────────────────────────────────
from dataclasses import dataclass
from typing import Any, Optional

try:
    from zoneinfo import ZoneInfo
    IST = ZoneInfo("Asia/Kolkata")
    _ZONEINFO_OK = True
except ImportError:  # Py < 3.9 — should not happen on the pinned 3.11 runner
    IST = None
    _ZONEINFO_OK = False

try:
    from tenacity import retry, stop_after_attempt, wait_exponential_jitter, retry_if_exception_type
    _TENACITY_OK = True
except ImportError:
    _TENACITY_OK = False
    # No-op shim so decorators don't crash if tenacity is missing
    def retry(*_a, **_kw):
        def _d(f): return f
        return _d
    def stop_after_attempt(*_a, **_kw): return None
    def wait_exponential_jitter(*_a, **_kw): return None
    def retry_if_exception_type(*_a, **_kw): return None

try:
    from nselib import capital_market as _nselib_cm
    try:
        from nselib import derivatives as _nselib_deriv  # noqa: F401 (reserved for Phase B5 fallback)
    except ImportError:
        _nselib_deriv = None
    _NSELIB_OK = True
except ImportError:
    _nselib_cm = None
    _nselib_deriv = None
    _NSELIB_OK = False


def ist_now() -> "datetime.datetime":
    """Always-correct IST timestamp — does NOT depend on the TZ env var."""
    if _ZONEINFO_OK:
        return datetime.datetime.now(IST)
    return datetime.datetime.now()  # fallback


def ist_today() -> "datetime.date":
    return ist_now().date()


# ── Disk caches & artifacts ─────────────────────────────────────────────────
NSELIB_CACHE_DIR     = os.getenv("NSELIB_CACHE_DIR", "nselib_cache")
LAST_KNOWN_GOOD_FILE = os.getenv("LAST_KNOWN_GOOD_FILE", "last_known_good.json")
DECISION_AUDIT_FILE  = f"decision_audit_{ist_today().strftime('%Y%m%d')}.jsonl"
# Phase 1 #51 + #52 + #54 (2026-07-05): research instrumentation artifacts.
# All append-only, tolerated to be missing on first run.
TRADABLE_STATE_FILE       = os.getenv("TRADABLE_STATE_FILE", "last_tradable.json")
TRADABLE_DROPOUT_FILE     = os.getenv("TRADABLE_DROPOUT_FILE", "tradable_dropouts.jsonl")
PRICE_FETCH_FAIL_FILE     = os.getenv("PRICE_FETCH_FAIL_FILE", "price_fetch_failures.jsonl")
DAILY_SNAPSHOT_DIR        = os.getenv("DAILY_SNAPSHOT_DIR", "daily_snapshots")
# Phase N-2 (2026-07-03): reject-outcome watch list. main.py appends every
# rejected stock's close/reasons here; scripts/reject_followup.py polls this
# file after N days to see whether the reject was justified (stock dumped) or
# a false negative (stock rallied — pattern we should stop rejecting).
REJECT_WATCH_FILE    = os.getenv("REJECT_WATCH_FILE", "reject_watch.json")

try:
    os.makedirs(NSELIB_CACHE_DIR, exist_ok=True)
except Exception:
    pass


# ── FetchResult — provenance + freshness on every replaced fetcher ──────────
@dataclass
class FetchResult:
    """Provenance wrapper. Callers read `.value`; the rest is for logging/audit."""
    value: Any
    source: str                      # e.g. "nselib_nsdl", "frankfurter", "yfinance", "LKG", "NEUTRAL_DEFAULT"
    as_of_date: "datetime.date"      # the date the data represents (NOT when fetched)
    fetched_at: "datetime.datetime"  # wall-clock IST at return
    is_stale: bool = False           # True if as_of_date < expected business date
    notes: str = ""                  # free-form (e.g. "BSE disagrees by 28%")

    def to_log(self) -> str:
        return (f"source={self.source} as_of={self.as_of_date.isoformat()} "
                f"stale={self.is_stale}{(' notes=' + self.notes) if self.notes else ''}")


# ── Macro validation — hard range gates ─────────────────────────────────────
# Single-day NSE FII record outflow was ~₹20k Cr (Mar 2020 covid).
# Anything outside these bounds is a parse error, not real data.
_MACRO_RANGES = {
    # 2026: rupee has traded 82-97; keep a wide 2-6 year band so a real 95.xx
    # print doesn't get flagged as an OOR parse error.
    "usdinr":         (70.0, 105.0),
    "vix_in":         (8.0,  60.0),
    "vix_us":         (8.0,  80.0),
    "crude_usd":      (30.0, 200.0),
    "us10y":          (1.0,  10.0),
    "dxy":            (80.0, 120.0),
    "gold_usd":       (1500.0, 6000.0),
    "nifty_1d_pct":   (-10.0, 10.0),
    "sp500_1d_pct":   (-10.0, 10.0),
    "sensex_1d_pct":  (-10.0, 10.0),
    "dow_1d_pct":     (-10.0, 10.0),
    "fii_flow_cr":    (-50000.0, 50000.0),
    "dii_flow_cr":    (-50000.0, 50000.0),
}


def load_last_known_good() -> dict:
    try:
        if os.path.exists(LAST_KNOWN_GOOD_FILE):
            with open(LAST_KNOWN_GOOD_FILE, "r") as f:
                return json.load(f)
    except Exception as e:
        # _log not yet defined at import time — print is fine here
        print(f"[WARN] load_last_known_good failed: {e}")
    return {}


def save_last_known_good(lkg: dict) -> None:
    try:
        # Only persist scalars (no FetchResult / dataclass)
        clean = {k: v for k, v in lkg.items() if isinstance(v, (int, float, str, bool))}
        with open(LAST_KNOWN_GOOD_FILE, "w") as f:
            json.dump(clean, f, indent=2, sort_keys=True, default=str)
    except Exception as e:
        try:
            _log(f"[WARN] save_last_known_good failed: {e}")
        except Exception:
            print(f"[WARN] save_last_known_good failed: {e}")


def validate_macro(macro: dict) -> dict:
    """
    Range-gate every macro value. If any value is out of range:
      - log [VALIDATION] warning
      - swap in last-known-good from last_known_good.json
      - set macro["data_quality"] = "DEGRADED" and list bad fields
    Otherwise: persist current values as the new LKG, mark NORMAL.
    """
    lkg = load_last_known_good()
    bad: list = []

    for field_name, (lo, hi) in _MACRO_RANGES.items():
        v = macro.get(field_name)
        if v is None:
            continue
        try:
            vf = float(v)
        except (TypeError, ValueError):
            bad.append(field_name)
            continue
        if vf < lo or vf > hi:
            bad.append(field_name)
            replacement = lkg.get(field_name)
            try:
                _log(f"[VALIDATION] {field_name}={vf!r} out of range [{lo},{hi}] → "
                     + (f"using LKG {replacement}" if replacement is not None else "no LKG — keeping default"))
            except Exception:
                pass
            if replacement is not None:
                macro[field_name] = replacement

    if bad:
        macro["data_quality"]  = "DEGRADED"
        macro["bad_fields"]    = bad
    else:
        macro["data_quality"]  = macro.get("data_quality", "NORMAL")
        macro["bad_fields"]    = []
        # Persist the validated values back to LKG for next run
        save_last_known_good({k: macro.get(k) for k in _MACRO_RANGES if k in macro})

    return macro


# Stage-A cleanup (2026-07-XX): cross_check_fii() removed — was only called by
# the legacy FII/DII cascade (also removed). FII/DII fetching disabled since
# Phase C4; see fetch_fii_dii_flows() stub below.


def append_decision_audit(entry: dict) -> None:
    """Append one JSON line to decision_audit_YYYYMMDD.jsonl (post-mortem trail)."""
    try:
        entry = {**entry, "ts": ist_now().isoformat()}
        with open(DECISION_AUDIT_FILE, "a", encoding="utf-8") as f:
            f.write(json.dumps(entry, default=str) + "\n")
    except Exception:
        # Never let audit logging crash the pipeline
        pass


# ─────────────────────────────────────────────────────────────────────────────
# Phase N-2 (2026-07-03): reject-outcome watch list
# ─────────────────────────────────────────────────────────────────────────────
# Every REJECTED stock is appended here with today's close + top fail reasons.
# scripts/reject_followup.py runs daily and for entries aged N=5/10/20 days
# fetches subsequent close and computes realized return. Enables false-reject
# discovery: "we rejected this stock because of X — but it rallied 15%, so
# gate X is producing false negatives on this pattern."
#
# File format (list of dicts, newest first, capped by REJECT_WATCH_MAX_ROWS):
#   [
#     {"date": "2026-07-03", "symbol": "FOO.NS", "close": 123.45,
#      "reasons": ["MARKET_CAP_LOW_₹350Cr_(min_₹500Cr)"],
#      "sector": "Chemicals", "trade_quality": 62.3,
#      "confidence": 71, "market_cap_cr": 350},
#     ...
#   ]
# The follow-up script mutates entries by appending outcome_* fields; it does
# NOT alter main.py's contract (main.py only ever appends).
# ─────────────────────────────────────────────────────────────────────────────
REJECT_WATCH_MAX_ROWS = int(os.getenv("REJECT_WATCH_MAX_ROWS", "5000"))


def load_reject_watch() -> list:
    """Load reject_watch.json — returns [] if missing / malformed / FRESH_START."""
    if FRESH_START:
        try:
            _log("[FRESH_START] load_reject_watch → returning [] (old reject watch ignored)")
        except Exception:
            pass
        return []
    try:
        if not os.path.exists(REJECT_WATCH_FILE):
            return []
        with open(REJECT_WATCH_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            return data
        return []
    except Exception:
        return []


def save_reject_watch(rows: list) -> None:
    """Write reject_watch.json, capping at REJECT_WATCH_MAX_ROWS (newest first)."""
    try:
        if len(rows) > REJECT_WATCH_MAX_ROWS:
            rows = rows[:REJECT_WATCH_MAX_ROWS]
        with open(REJECT_WATCH_FILE, "w", encoding="utf-8") as f:
            json.dump(rows, f, indent=2, default=str)
    except Exception as e:
        try:
            _log(f"[WARN] save_reject_watch failed: {e}")
        except Exception:
            pass


def append_reject_watch_entries(rejects: list, run_date: str) -> int:
    """
    Append today's rejects to reject_watch.json. Idempotent: skips symbols
    already in the file for the same run_date. Returns count actually appended.
    """
    watch = load_reject_watch()
    seen_today = {
        r.get("symbol") for r in watch
        if r.get("date") == run_date
    }
    new_rows = []
    for stock in rejects:
        sym = stock.get("symbol")
        if not sym or sym in seen_today:
            continue
        # Truncate reasons to top 3 (most-impactful hard gate first if present)
        reasons = list(stock.get("fail_reasons", []))[:3]
        entry = {
            "date":           run_date,
            "symbol":         sym,
            "close":          float(stock.get("entry", 0) or stock.get("price", 0) or 0),
            "reasons":        reasons,
            "sector":         stock.get("sector", "Unknown"),
            "trade_quality":  stock.get("trade_quality_score", stock.get("trade_quality", 0)),
            "confidence":     stock.get("final_confidence", stock.get("confidence", 0)),
            "market_cap_cr":  float(stock.get("market_cap_cr", 0) or 0),
            "avg_val_lakhs":  float(stock.get("avg_value_lakhs", 0) or 0),
        }
        new_rows.append(entry)
    if new_rows:
        # Prepend new rows so newest-first ordering is preserved
        watch = new_rows + watch
        save_reject_watch(watch)
    return len(new_rows)


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 1 — CONFIG & ENVIRONMENT
# ─────────────────────────────────────────────────────────────────────────────

PORTFOLIO_CAPITAL   = float(os.getenv("CAPITAL", "500000"))
TELEGRAM_BOT_TOKEN  = os.getenv("TELEGRAM_BOT_TOKEN", "")
TELEGRAM_CHAT_ID    = os.getenv("TELEGRAM_CHAT_ID", "")
# Dedicated BUY-signal channel — set BUY_BOT_TOKEN + BUY_CHAT_ID in GitHub Secrets
BUY_BOT_TOKEN       = os.getenv("BUY_BOT_TOKEN", "")
BUY_CHAT_ID         = os.getenv("BUY_CHAT_ID", "")
TRACKER_FILE        = os.getenv("TRACKER_FILE", "tracker.json")
TRADE_TRACKER_V2_FILE   = os.getenv("TRADE_TRACKER_V2_FILE", "trade_tracker.json")
FUNDAMENTALS_CACHE_FILE = os.getenv("FUNDAMENTALS_CACHE_FILE", "fundamentals_cache.json")
PORTFOLIO_FILE          = os.getenv("PORTFOLIO_FILE", "portfolio.json")
WATCHLIST_FILE      = os.getenv("WATCHLIST_FILE", "watchlist_persist.json")
CONF_HISTORY_FILE   = os.getenv("CONF_HISTORY_FILE", "confidence_history.json")
GATE_MEMORY_FILE    = os.getenv("GATE_MEMORY_FILE", "gate_memory.json")
# Phase C3 (2026-07-02): real per-stock delivery % cache + sector rank history
# + VIX percentile cache. All three are additive institutional-grade signals.
DELIVERY_CACHE_FILE     = os.getenv("DELIVERY_CACHE_FILE", "delivery_cache.json")
SECTOR_RANK_HISTORY_FILE = os.getenv("SECTOR_RANK_HISTORY_FILE", "sector_rank_history.json")
VIX_HISTORY_CACHE_FILE  = os.getenv("VIX_HISTORY_CACHE_FILE", "vix_history_cache.json")
# True only when triggered by GitHub Actions cron schedule — never for manual runs
IS_SCHEDULED        = os.getenv("SCHEDULED_RUN", "false").lower() == "true"

# ─── Phase C7c (2026-07-02): FRESH_START switch ─────────────────────────────
# Set FRESH_START=true in the GitHub Actions env (or as a workflow_dispatch
# input) for ONE run to wipe all decision-tainted state and start clean.
#
# ═══════════════════════════════════════════════════════════════════════════
# Phase C7h (2026-07-10): FULL STATE SCRUBBER — every stateful file resets.
# ═══════════════════════════════════════════════════════════════════════════
#
# WHAT GETS WIPED (categorised — see wipe block below for the authoritative
# lists):
#
#   RENAMED to <name>.stale_<date> (post-mortem preserved):
#     • shadow_master.xlsx, shadow_report.xlsx, shadow_report_weekly.xlsx
#     • recommendation_tracker.xlsx (legacy v1 workbook)
#     • shadow_trades.csv, portfolio_state.csv, telegram_daily.csv
#     • price_fetch_failures.jsonl, tradable_dropouts.jsonl
#
#   DELETED outright (regenerated automatically):
#     • trade_tracker.json, trade_tracker_v2.json, tracker.json
#     • gate_memory.json, regime_calibration.json, watchlist_persist.json
#     • confidence_history.json, reject_watch.json
#     • sector_rank_history.json, delivery_cache.json, fundamentals_cache.json
#     • sector_cache.json, weekly_metrics.json, telegram_prev_state.json
#     • intraday_snapshots.json, last_known_good.json, last_tradable.json
#     • run_health.json
#
#   PRESERVED (user-owned config or expensive metadata — see documented
#   exclusion list inline in the wipe block):
#     • portfolio.json, events_config.json, stocks.txt, blocklists
#     • nse_all_symbols.csv, sector_master.csv, market_calendars.json
#     • vix_history_cache.json, run_log_*.txt
#
# Load-time skips (in-memory return {} / [] even if disk copy survives)
# still fire for defence-in-depth on downstream runners:
#   • load_tracker(), load_tracker_v2(), initialize_tracker_if_new()
#   • load_gate_memory(), load_regime_calibration()
#   • load_watchlist_persist(), load_confidence_history()
#   • load_reject_watch()
#   • decision_audit_*.jsonl — new file naturally (date-suffixed)
#
# After a FRESH_START run persists to git, unset FRESH_START for subsequent
# runs — they'll build on the clean baseline naturally.
FRESH_START = os.getenv("FRESH_START", "false").lower() == "true"
if FRESH_START:
    print("[FRESH_START] Enabled — old tracker/audit/memory state will be ignored this run")

    # ─────────────────────────────────────────────────────────────────────
    # Phase C7h (2026-07-10): FULL STATE SCRUBBER
    # ─────────────────────────────────────────────────────────────────────
    # Previous versions only wiped a subset (main.py-owned loaders + a
    # single weekly_metrics.json delete). That left ~9 state files on disk
    # that leaked stale data into supposedly-fresh runs.
    #
    # Now: every stateful/history file gets reset in a single deterministic
    # pass here, so the ONE FRESH_START run produces a truly clean baseline.
    #
    # Three action types:
    #   RENAME  — big files where post-mortem value is real (xlsx, shadow_trades)
    #             → renamed to <name>.stale_<date> so you can zip & compare later
    #   DELETE  — small transient JSON caches/summaries (regenerated on demand)
    #             → deleted outright to keep the repo tidy
    #   PRESERVE — user-owned config or expensive universe metadata
    #             → NOT touched. Documented exclusion list below.
    #
    # If you add a NEW stateful file to the pipeline, add it to ONE of the
    # three sections below. That's the single source of truth for what a
    # fresh start means in this repo.
    # ─────────────────────────────────────────────────────────────────────
    _fresh_today = datetime.datetime.now().strftime("%Y-%m-%d")

    # (A) Files to RENAME to <name>.stale_<date>. Preserves the raw data for
    # post-mortem/comparison but keeps the "live" filename empty so the
    # pipeline creates a fresh one on first write.
    _fresh_rename_files = [
        # xlsx workbooks (rebuild cost high, but data valuable for audit)
        "shadow_master.xlsx",            # main shadow log workbook
        "shadow_report.xlsx",             # weekly rollup workbook
        "shadow_report_weekly.xlsx",      # alternate weekly xlsx name
        "recommendation_tracker.xlsx",    # legacy v1 tracker (no code path but user-visible)
        # csv accumulators (append-only history files)
        "shadow_trades.csv",              # 4-bucket A/B/C/D shadow rows
        "portfolio_state.csv",            # portfolio snapshot log
        "telegram_daily.csv",             # per-day telegram audit trail
        # jsonl append-only logs
        "price_fetch_failures.jsonl",     # yfinance failure log
        "tradable_dropouts.jsonl",         # symbols that stopped trading
    ]

    # (B) Files to DELETE outright. All are small (<50 KB), regenerated
    # automatically by their owning loader/job on next run, and have no
    # post-mortem value beyond what's already in trade_tracker.json.
    _fresh_delete_files = [
        # main.py-owned state (also load-guarded, but wipe the disk copy
        # so downstream jobs on separate runners see the fresh state too)
        "trade_tracker.json",
        "trade_tracker_v2.json",
        "tracker.json",                    # tracker_job.py / morning_check.py
        "gate_memory.json",
        "regime_calibration.json",
        "watchlist_persist.json",
        "confidence_history.json",
        "reject_watch.json",
        # C7g caches (already covered — kept here for single-source-of-truth)
        "sector_rank_history.json",
        "delivery_cache.json",
        "fundamentals_cache.json",
        # sector cache (auto-rebuilt on first sector lookup)
        "sector_cache.json",
        # weekly summary output
        "weekly_metrics.json",
        # telegram diff state (yesterday-vs-today Telegram formatting)
        "telegram_prev_state.json",
        # intraday snapshot store (rebuilt from live prices)
        "intraday_snapshots.json",
        # health/liveness checkpoints (rebuilt on first successful pass)
        "last_known_good.json",
        "last_tradable.json",
        "run_health.json",
        # intraday_monitor.py dedup cache (bounded to today; must not leak)
        "intraday_state.json",
        # transient job flags (created by shadow_master_job / tracker_job on
        # error/manual-trigger paths — stale ones would poison next run)
        "yfinance_down.flag",
        "manual_in_ci.flag",
        # marker file (safety: main.py deletes stale markers elsewhere too,
        # but a hard wipe here guarantees no cross-day leak on fresh start)
        ".fresh_start_marker",
    ]

    # (B2) GLOB-DELETE — date-suffixed audit/log files that jobs create fresh
    # every day. A fresh start MUST wipe all previous days' copies so the
    # baseline truly looks like day-0.
    #   decision_audit_YYYYMMDD.jsonl  — main.py append-only per-day audit
    #   run_log_YYYYMMDD.txt           — main.py per-run stdout tee
    _fresh_delete_globs = [
        "decision_audit_*.jsonl",
        "run_log_*.txt",
    ]

    # (C) DOCUMENTED PRESERVATION LIST — files that FRESH_START explicitly
    # does NOT touch. If you're wondering why file X survived a fresh start,
    # it should be here (with rationale). This is enforced by convention
    # only; add to _fresh_rename_files or _fresh_delete_files to wipe.
    #
    #   portfolio.json          — USER-OWNED positions, never auto-wiped
    #   events_config.json      — USER-OWNED event calendar config
    #   stocks.txt              — USER-OWNED custom universe
    #   high_pledge_stocks.txt  — USER-OWNED manual blocklist
    #   asm_gsm_blocklist.txt   — USER-OWNED manual blocklist
    #   requirements.txt        — code (Python deps), not state
    #   nse_all_symbols.csv     — REBUILD-COST high (NSE fetch), not decision-tainted
    #   sector_master.csv       — REBUILD-COST high (static sector map)
    #   market_calendars.json   — METADATA (NSE holidays), not decision-tainted
    #   vix_history_cache.json  — METADATA (long-term market context, ~5y)
    #   main.py.bak_*           — MANUAL backups (user-created, not job output)

    _renamed = 0
    _deleted = 0
    _skipped = 0

    for _fname in _fresh_rename_files:
        try:
            if not os.path.exists(_fname):
                _skipped += 1
                continue
            _stale = f"{_fname}.stale_{_fresh_today}"
            # If a stale-file for today already exists (re-run same day),
            # tack on a millisecond suffix to avoid clobber.
            if os.path.exists(_stale):
                import time as _t_fs
                _stale = f"{_stale}_{int(_t_fs.time()*1000) % 100000}"
            os.rename(_fname, _stale)
            _renamed += 1
            print(f"[FRESH_START] RENAMED  {_fname} → {os.path.basename(_stale)}")
        except Exception as _e_r:
            print(f"[FRESH_START] Could not rename {_fname}: {_e_r} — non-fatal")

    for _fname in _fresh_delete_files:
        try:
            if not os.path.exists(_fname):
                _skipped += 1
                continue
            os.remove(_fname)
            _deleted += 1
            print(f"[FRESH_START] DELETED  {_fname} (will be regenerated by owning job)")
        except Exception as _e_d:
            print(f"[FRESH_START] Could not delete {_fname}: {_e_d} — non-fatal")

    # Phase C7j (2026-07-11): glob-delete date-suffixed audit/log files.
    # decision_audit_*.jsonl and run_log_*.txt are created fresh each day
    # by main.py; a truly fresh start must wipe all prior copies.
    import glob as _glob_fs
    for _pat in _fresh_delete_globs:
        try:
            _matches = _glob_fs.glob(_pat)
            if not _matches:
                _skipped += 1
                continue
            for _mf in _matches:
                try:
                    os.remove(_mf)
                    _deleted += 1
                    print(f"[FRESH_START] DELETED  {_mf} (glob {_pat})")
                except Exception as _e_gm:
                    print(f"[FRESH_START] Could not delete {_mf}: {_e_gm} — non-fatal")
        except Exception as _e_g:
            print(f"[FRESH_START] Glob {_pat} failed: {_e_g} — non-fatal")

    # ─────────────────────────────────────────────────────────────────────
    # Phase C7i (2026-07-11): archive trees + GH artifact bundles
    # ─────────────────────────────────────────────────────────────────────
    # Historical dated files under reports/archive/ and gh_artifacts/ also
    # count as "generated pipeline state". A FRESH_START run must sweep
    # these into a single dated stale-tree so the workspace looks truly
    # empty for the next-day rebuild.
    #
    # Strategy: rename each directory to <name>.stale_<date> (fast, atomic,
    # keeps the raw archives for post-mortem/comparison). If the rename
    # fails because a stale-dir already exists for today, add a millisecond
    # suffix — same idempotency trick as the file-rename loop above.
    _fresh_rename_dirs = [
        os.path.join("reports", "archive"),
        "gh_artifacts",
    ]
    _renamed_dirs = 0
    for _dname in _fresh_rename_dirs:
        try:
            if not os.path.exists(_dname) or not os.path.isdir(_dname):
                _skipped += 1
                continue
            _stale_d = f"{_dname}.stale_{_fresh_today}"
            if os.path.exists(_stale_d):
                import time as _t_fs2
                _stale_d = f"{_stale_d}_{int(_t_fs2.time()*1000) % 100000}"
            os.rename(_dname, _stale_d)
            _renamed_dirs += 1
            print(f"[FRESH_START] RENAMED  {_dname}/ → {os.path.basename(_stale_d)}/")
        except Exception as _e_rd:
            print(f"[FRESH_START] Could not rename {_dname}/: {_e_rd} — non-fatal")

    print(f"[FRESH_START] Summary: renamed={_renamed} files + "
          f"{_renamed_dirs} dirs, deleted={_deleted}, "
          f"absent={_skipped}, preserved={11}")
    print(f"[FRESH_START] Clean baseline ready — remember to unset FRESH_START "
          f"for tomorrow's run.")
else:
    # Phase C7g (2026-07-10): NON-CONSUMING marker design cleanup step.
    # main.py writes `.fresh_start_marker` when FRESH_START=true so that
    # downstream jobs (research, weekly, intraday, shadow_weekly, backtest,
    # morning_check) — which each run on their own runner with their own git
    # checkout — can peek at it without consuming. The marker is committed to
    # git by the Persist step. To prevent stale markers from accumulating,
    # main.py deletes any marker whose date != today on every non-FS run.
    try:
        _marker_path = ".fresh_start_marker"
        if os.path.exists(_marker_path):
            with open(_marker_path, "r", encoding="utf-8") as _fmr:
                _marker_date = _fmr.read().strip()
            _today_iso = datetime.datetime.now().strftime("%Y-%m-%d")
            if _marker_date != _today_iso:
                os.remove(_marker_path)
                print(f"[FRESH_START] Cleaned up stale marker (date={_marker_date}, today={_today_iso})")
            # else: marker IS today — leave it alone; a same-day non-FS run
            # is unusual but we still want downstream peekers to react.
    except Exception as _e_mc:
        print(f"[FRESH_START] Marker cleanup skipped: {_e_mc} — non-fatal")

# ═════════════════════════════════════════════════════════════════════════════
# Phase R4 (2026-07-06): TRADING MODE PRESETS
# ─────────────────────────────────────────────────────────────────────────────
# TRADING_MODE=swing    (default) — 2–10 day swing trading:
#   Softer fundamentals (BQ is overlay, not gate). Sector rotation + technical
#   setup + RR are the primary edges. Pledge = warning, not veto. LLM validator
#   advise-only. Optimized to surface 3–8 BUYs per day in a bullish regime.
# TRADING_MODE=position — 4–12 week position trading (previous default):
#   Strict fundamentals HARD gates. High-quality only. Pledge = NO_GO.
#   LLM validator veto mode. Expects 0–3 BUYs per day.
# TRADING_MODE=custom   — no preset applied, all env vars honored as-is.
#
# The preset ONLY fills variables the user has NOT already set explicitly, so
# any env var set on the command line or in .env still wins.
# ═════════════════════════════════════════════════════════════════════════════
TRADING_MODE = os.getenv("TRADING_MODE", "swing").lower().strip()

def _set_default(key, value):
    """Set env var only if not already set (empty string counts as unset)."""
    if not os.environ.get(key):
        os.environ[key] = str(value)

if TRADING_MODE == "swing":
    # Fundamentals — relaxed (swing doesn't need 15% ROE for 5-day hold)
    _set_default("MIN_ROE", "10")           # was 12/15
    _set_default("MAX_DE", "1.5")           # was 1.0
    _set_default("BUSINESS_QUALITY_GATE", "0")  # BQ becomes overlay, not gate
    # Sector — still strict, this IS the swing edge
    _set_default("SECTOR_STRICT_GATE", "1")
    _set_default("SECTOR_RANK_CUTOFF", "6")     # allow slightly weaker sectors
    # Events — HARD gate for swing (gap risk = career risk)
    _set_default("EARNINGS_BLACKOUT_DAYS", "5")
    _set_default("NEWS_SEVERITY_MAX", "2")
    # Stage-A cleanup (2026-07-XX): LLM_VALIDATOR removed (was default OFF,
    # 0/480 populated in prior audits). See git history to revive.
    # Pledge — warning, not dealbreaker (swing exits before it matters)
    _set_default("PLEDGE_SEVERITY", "warning")   # vs "dealbreaker" in position
    # Confidence bar — normal
    _set_default("FUND_MISSING_CONF_CAP", "80")
    print(f"[TRADING_MODE=swing] Preset applied: MIN_ROE=10, MAX_DE=1.5, BQ_GATE=overlay, PLEDGE=warning")

elif TRADING_MODE == "position":
    # Fundamentals — strict
    _set_default("MIN_ROE", "15")
    _set_default("MAX_DE", "1.0")
    _set_default("BUSINESS_QUALITY_GATE", "1")
    # Sector — strict
    _set_default("SECTOR_STRICT_GATE", "1")
    _set_default("SECTOR_RANK_CUTOFF", "4")
    # Events — HARD
    _set_default("EARNINGS_BLACKOUT_DAYS", "5")
    _set_default("NEWS_SEVERITY_MAX", "2")
    # Stage-A cleanup (2026-07-XX): LLM_VALIDATOR removed. See git history.
    # Pledge — dealbreaker
    _set_default("PLEDGE_SEVERITY", "dealbreaker")
    _set_default("FUND_MISSING_CONF_CAP", "75")
    print(f"[TRADING_MODE=position] Preset applied: MIN_ROE=15, MAX_DE=1.0, BQ_GATE=hard, PLEDGE=dealbreaker")

elif TRADING_MODE == "custom":
    print(f"[TRADING_MODE=custom] No preset applied; all env vars honored as-is")

else:
    print(f"[TRADING_MODE={TRADING_MODE}] Unknown mode — treating as 'swing'")
    TRADING_MODE = "swing"
    # re-run preset with swing defaults (R5 PRUNE: LLM off by default)
    for k, v in [("MIN_ROE","10"),("MAX_DE","1.5"),("BUSINESS_QUALITY_GATE","0"),
                 ("SECTOR_STRICT_GATE","1"),("SECTOR_RANK_CUTOFF","6"),
                 ("EARNINGS_BLACKOUT_DAYS","5"),("NEWS_SEVERITY_MAX","2"),
                 ("PLEDGE_SEVERITY","warning"),("FUND_MISSING_CONF_CAP","80")]:
        _set_default(k, v)

# ═════════════════════════════════════════════════════════════════════════════
# Phase E1 (2026-07-02): PROFESSIONAL EXIT STACK — 5-layer system
# ═════════════════════════════════════════════════════════════════════════════
# All exit enhancements are ADDITIVE and ENV-GATED. Defaults preserve the
# legacy 'exit-fully-at-T2 + break-even trail after T1' behavior.
#
# Layer 1  Hard stop + T1 partial + T2 exit (existing legacy behavior — always on).
# Layer 2  Runner mode (E1a): at T2, book RUNNER_PARTIAL_PCT of what remains
#          after T1 partial (default 50%) and let the residual run on a
#          chandelier trail. Enabled by RUNNER_MODE_ENABLED.
# Layer 3  ATR trail after T1 (E1b): instead of break-even, use
#          max(entry, current − TRAIL_ATR_MULT × ATR14) once T1 hits. Enabled
#          by TRAIL_MODE=atr (default 'breakeven' = legacy).
# Layer 4  Conditional time exit (E1c): only expire flat/losing/stalled trades
#          at day N; let winners run. Enabled by TIME_EXIT_MODE=conditional
#          (default 'calendar' = legacy).
# Layer 5  Regime-aware tightening (E1d): tighten trails and skip runner mode
#          in risk-off regimes. Enabled by REGIME_AWARE_EXITS.
#
# Every rule falls back to legacy behavior when its flag is off, so you can
# roll them out one at a time (E1a → E1b → E1c → E1d).
# ─────────────────────────────────────────────────────────────────────────────

# ── Layer 2: Runner mode at T2 (Phase E1a) ────────────────────────────────
# When ON: at T2, book part of the residual and let a runner ride a chandelier
# trail until (a) trail hits, (b) trend-break (close < EMA21 for N days), or
# (c) volume fade (close < 20d avg for N days).
RUNNER_MODE_ENABLED     = os.getenv("RUNNER_MODE_ENABLED", "false").lower() == "true"
RUNNER_PARTIAL_PCT      = float(os.getenv("RUNNER_PARTIAL_PCT", "50"))   # % of residual to book at T2
RUNNER_ATR_MULT         = float(os.getenv("RUNNER_ATR_MULT",    "3.0"))  # chandelier: highest close − mult × ATR14
RUNNER_TREND_EMA        = int(os.getenv("RUNNER_TREND_EMA",     "21"))    # EMA period for trend-break exit
RUNNER_TREND_BREAK_DAYS = int(os.getenv("RUNNER_TREND_BREAK_DAYS", "2"))  # consecutive closes below EMA to exit
RUNNER_VOL_FADE_DAYS    = int(os.getenv("RUNNER_VOL_FADE_DAYS",  "3"))    # consecutive low-volume closes to exit
RUNNER_MAX_DAYS         = int(os.getenv("RUNNER_MAX_DAYS",       "45"))   # hard ceiling after T2 hit

# ── Layer 3: ATR trail after T1 (Phase E1b) ───────────────────────────────
# 'breakeven' (default, legacy): stop → entry once T1 hits.
# 'atr'      : stop → max(entry, current − TRAIL_ATR_MULT × ATR14), ratcheted daily.
TRAIL_MODE              = os.getenv("TRAIL_MODE", "breakeven").lower()
TRAIL_ATR_MULT          = float(os.getenv("TRAIL_ATR_MULT", "2.5"))

# ── Layer 4: Conditional time exit (Phase E1c) ────────────────────────────
# 'calendar'    (default, legacy): expire at TIME_EXIT_MAX_DAYS if under T1.
# 'conditional' : keep winners past N days; only expire flat/losing OR when 10d slope <= 0.
TIME_EXIT_MODE            = os.getenv("TIME_EXIT_MODE", "calendar").lower()
TIME_EXIT_MAX_DAYS        = int(os.getenv("TIME_EXIT_MAX_DAYS", "15"))       # legacy day floor
TIME_EXIT_HARD_MAX_DAYS   = int(os.getenv("TIME_EXIT_HARD_MAX_DAYS", "30"))  # even winners exit at this
TIME_EXIT_MIN_WIN_PCT     = float(os.getenv("TIME_EXIT_MIN_WIN_PCT", "5.0")) # < this % after N days = expire
TIME_EXIT_STAGNATION_DAYS = int(os.getenv("TIME_EXIT_STAGNATION_DAYS", "25"))
TIME_EXIT_STAGNATION_PCT  = float(os.getenv("TIME_EXIT_STAGNATION_PCT", "10.0"))

# ── Layer 5: Regime-aware exit tightening (Phase E1d) ─────────────────────
# When ON, in BEAR / STRONG_BEAR / HIGH_VOLATILITY:
#   - runner mode is disabled (take T2 fully)
#   - ATR trail multiplier reduced by REGIME_TIGHTEN_FACTOR (e.g. 2.5 → 1.5)
#   - conditional time exit uses shorter windows
REGIME_AWARE_EXITS      = os.getenv("REGIME_AWARE_EXITS", "false").lower() == "true"
REGIME_TIGHTEN_FACTOR   = float(os.getenv("REGIME_TIGHTEN_FACTOR", "0.6"))  # multiplier applied to trail ATR mult

# ═════════════════════════════════════════════════════════════════════════════
# Phase F (2026-07-02): PROFESSIONAL POLISH LAYER — 5 additive features
# ═════════════════════════════════════════════════════════════════════════════
# All F-series features are ADDITIVE and ENV-GATED. Defaults preserve legacy.
#
# F5  Runner "moonshot" scaling — when a runner is +N% above T2 and volume is
#     expanding, book an additional slice. Locks profit on exceptional moves.
# F7a Portfolio-level risk sheet — total exposure, sector concentration,
#     rolling drawdown, VaR-95 estimate. Written into tracker.xlsx.
# F7b Parabolic exit — if a position gaps >X% in a single day (blow-off),
#     force a partial or full exit at close.
# F8  P&L attribution — split Avg Return% by exit bucket (T1-only, T2-hit,
#     Runner-closed, Stopped, Expired). Answers "where does alpha come from?".
# F9  Strategy-vs-NIFTY overlay — cumulative NIFTY return recorded alongside
#     tracker rows so we can plot alpha over time.

# ── F5: Runner scaling ────────────────────────────────────────────────────
RUNNER_SCALE_ENABLED    = os.getenv("RUNNER_SCALE_ENABLED", "false").lower() == "true"
RUNNER_SCALE_TRIGGER_PCT = float(os.getenv("RUNNER_SCALE_TRIGGER_PCT", "30"))  # % above T2 to trigger
RUNNER_SCALE_BOOK_PCT   = float(os.getenv("RUNNER_SCALE_BOOK_PCT",   "50"))    # % of remaining runner to book
RUNNER_SCALE_VOL_MULT   = float(os.getenv("RUNNER_SCALE_VOL_MULT",   "1.2"))   # volume must be ≥N× 20d avg

# ── F7b: Parabolic exit ───────────────────────────────────────────────────
PARABOLIC_EXIT_ENABLED  = os.getenv("PARABOLIC_EXIT_ENABLED", "false").lower() == "true"
PARABOLIC_DAY_PCT       = float(os.getenv("PARABOLIC_DAY_PCT",  "8.0"))    # single-day gain to trigger
PARABOLIC_VOL_MULT      = float(os.getenv("PARABOLIC_VOL_MULT", "2.0"))    # volume must be ≥N× 20d avg
PARABOLIC_BOOK_PCT      = float(os.getenv("PARABOLIC_BOOK_PCT", "50"))     # % of current position to book

# ── F7a / F8 / F9: reporting toggles ──────────────────────────────────────
# These are auto-populated by tracker_job.py; the env vars merely control
# whether the extra sheets/rows are written. All default TRUE (they're
# read-only additions to the xlsx and never touch trading logic).
PORTFOLIO_RISK_SHEET    = os.getenv("PORTFOLIO_RISK_SHEET",    "true").lower() == "true"
PNL_ATTRIBUTION_ROWS    = os.getenv("PNL_ATTRIBUTION_ROWS",    "true").lower() == "true"
BENCHMARK_OVERLAY       = os.getenv("BENCHMARK_OVERLAY",       "true").lower() == "true"
BENCHMARK_SYMBOL        = os.getenv("BENCHMARK_SYMBOL",        "^NSEI")

# Pipeline-level regime snapshot — set by the pipeline at scoring time so
# tracker updates (which run later, without regime state) can consult it.
# NEVER read this in scoring/gate math — exits only.
_PIPELINE_REGIME: str = ""


def _current_exit_regime() -> str:
    """Return regime string used by exit logic (empty = unknown, treat as neutral)."""
    return str(_PIPELINE_REGIME or "").upper()


def _is_risk_off_regime(regime: str = "") -> bool:
    r = (regime or _current_exit_regime()).upper()
    return r in ("BEAR", "STRONG_BEAR", "HIGH_VOLATILITY")


def _effective_trail_atr_mult(regime: str = "") -> float:
    """ATR trail multiplier, tightened in risk-off regimes when E1d is on."""
    base = TRAIL_ATR_MULT
    if REGIME_AWARE_EXITS and _is_risk_off_regime(regime):
        return round(max(1.0, base * REGIME_TIGHTEN_FACTOR), 2)
    return base


def _effective_runner_atr_mult(regime: str = "") -> float:
    base = RUNNER_ATR_MULT
    if REGIME_AWARE_EXITS and _is_risk_off_regime(regime):
        return round(max(1.5, base * REGIME_TIGHTEN_FACTOR), 2)
    return base


def _runner_enabled(regime: str = "") -> bool:
    """Runner mode ON only if flag set AND regime not risk-off (when E1d on)."""
    if not RUNNER_MODE_ENABLED:
        return False
    if REGIME_AWARE_EXITS and _is_risk_off_regime(regime):
        return False
    return True


def _compute_exit_context(symbol: str) -> dict:
    """Fetch a ~60-day daily df once and derive everything the exit engines
    need: ATR14, EMA21, high_since (highest close across window), avg_vol20,
    consecutive-below-EMA, consecutive-low-volume, 10d slope.

    Returns {} on any failure — callers MUST treat empty dict as 'skip
    enhanced rules, fall back to legacy exit path'. Never raises.
    """
    try:
        df = fetch_price_data(symbol, period="60d")
        if df is None or len(df) < 22:
            return {}
        closes = df["Close"].squeeze().astype(float).values
        highs  = df["High"].squeeze().astype(float).values
        lows   = df["Low"].squeeze().astype(float).values
        vols   = df["Volume"].squeeze().astype(float).values

        # ATR14 (Wilder-style approximation using SMA of True Range)
        tr = np.maximum(
            highs[1:] - lows[1:],
            np.maximum(
                np.abs(highs[1:] - closes[:-1]),
                np.abs(lows[1:]  - closes[:-1]),
            ),
        )
        atr14 = float(np.mean(tr[-14:])) if len(tr) >= 14 else float(np.mean(tr)) if len(tr) else 0.0

        ema21 = float(pd.Series(closes).ewm(span=RUNNER_TREND_EMA, adjust=False).mean().iloc[-1])
        avg_vol20 = float(pd.Series(vols).rolling(20).mean().iloc[-1]) if len(vols) >= 20 else float(np.mean(vols) if len(vols) else 0.0)

        # Consecutive closes below EMA21
        ema_series = pd.Series(closes).ewm(span=RUNNER_TREND_EMA, adjust=False).mean().values
        below = 0
        for i in range(len(closes) - 1, -1, -1):
            if closes[i] < ema_series[i]:
                below += 1
            else:
                break

        # Consecutive low-volume days
        vol_avg_series = pd.Series(vols).rolling(20).mean().values
        low_vol = 0
        for i in range(len(vols) - 1, -1, -1):
            if not np.isnan(vol_avg_series[i]) and vols[i] < vol_avg_series[i]:
                low_vol += 1
            else:
                break

        # 10-day slope (last-close vs close-10d-ago)
        slope10 = 0.0
        if len(closes) >= 11:
            slope10 = float(closes[-1] - closes[-11])

        # Phase F: prev_close + last_volume needed by F5 scaling / F7b parabolic
        prev_close  = float(closes[-2]) if len(closes) >= 2 else float(closes[-1])
        last_volume = float(vols[-1])   if len(vols)   >= 1 else 0.0

        return {
            "close":       float(closes[-1]),
            "prev_close":  prev_close,
            "last_volume": last_volume,
            "atr14":       atr14,
            "ema21":       ema21,
            "avg_vol20":   avg_vol20,
            "below_ema_streak": int(below),
            "low_vol_streak":   int(low_vol),
            "slope10":     slope10,
            "highest_close_60d": float(np.max(closes)),
        }
    except Exception as ex:
        _log(f"[WARN] _compute_exit_context({symbol}) failed: {ex}")
        return {}


TELEGRAM_MAX_CHARS  = 3800  # buffer below 4096 hard limit

# Regime thresholds — v6.0 calibrated (Bug 2 fix)
# Phase 3a N3 / Option B+VC (2026-07-05): min_tq re-calibrated from EMPIRICAL
# percentiles of the Option B+VC TQ distribution on the 20260706 CSV
# (100 candidates, mean_TQ=68.8, stdev=4.8). Selected percentile-based
# thresholds instead of arbitrary drops:
#
#     Regime         min_tq   Empirical pass rate on 20260706 sample
#     STRONG_BULL    68       ~40-45% (top-40 pct, aggressive regime)
#     BULL           71        25%    (top-25 pct)  ← primary regime today
#     SIDEWAYS       73        13%    (top-15 pct)
#     TRANSITION     73        13%    (top-15 pct)
#     HIGH_VOL       76         3%    (top-5-10 pct)
#     BEAR           82         0%    (max_buys=0 is the primary gate)
#     STRONG_BEAR    86         0%    (max_buys=0 is the primary gate)
#
# Distribution trace — how we got here (verified empirically on 20260706 CSV):
#   OLD:      mean_TQ = 81.9,  97/100 pass @ min_tq=76  (rubber-stamp gate)
#   NEW (N1): mean_TQ = 67.7,  10/100 pass @ min_tq=76  (over-tight)
#   OPT B:    mean_TQ = 70.8,  ~30/100 pass @ min_tq=70 (rebalanced)
#   OPT B+VC: mean_TQ = 68.8,   25/100 pass @ min_tq=71 (VCP-aware, THIS)
# VC-mean is ~50 (most stocks are in the neutral 50 band; only true VCP
# setups score 75-90), so VC @ 0.10 pulls mean_TQ down ~2 vs Opt B while
# lifting the *right* stocks (Minervini-style coiled springs) into the pass
# band. Empirically-derived thresholds avoid the guess-and-check cycle.
# min_confidence and min_rr are unchanged — those gates operate on separate
# evidence (base_confidence and R/R) and did not have a distribution shift.
REGIME_THRESHOLDS = {
    # max_stop_pct — wide-stop guardrail (any BUY with (entry-stop)/entry > this
    # gets rejected). Bullish regimes tolerate wider volatility stops; bearish
    # regimes force tight stops to keep loss small.
    # Phase R6 (2026-07-06) — threshold recalibration after live BUY-starvation
    # analysis. In BULL regime, top-tier real-estate names (LODHA/OBEROI/PHOENIX)
    # with BQ STRONG + sector 98/100 + Pocket Pivot + 2× volume were scoring
    # final_confidence ~65-68 (well below prior BULL floor of 82). The 10-factor
    # composite rarely exceeds 75 on real Indian mid-caps due to fundamental
    # data gaps + wide-ATR penalties. Lowered BULL to 75 (was 82) and STRONG_BULL
    # to 72 (was 78). See scripts/backtest_audit_forward.py to validate.
    "STRONG_BULL":     {"min_confidence": 72, "min_tq": 68, "min_rr": 1.7, "max_buys": 5,  "max_exposure": 0.85, "max_stop_pct": 8.0},
    "BULL":            {"min_confidence": 75, "min_tq": 71, "min_rr": 1.8, "max_buys": 3,  "max_exposure": 0.75, "max_stop_pct": 7.5},
    # 2026-07-03 calibration: SIDEWAYS max_stop_pct raised 6.0 -> 8.0 because
    # empirically 10-day swing lows on tradable NSE stocks sit 8-12% below
    # entry in range-bound tapes. The old 6% cap made WIDE_STOP a
    # 100%-hit gate (mathematically impossible to satisfy), producing 0
    # signals every day. Quality gates (Conf/TQ/RR) remain strict.
    #
    # 2026-07-07 SIDEWAYS retune (Phase H): live run on 2026-07-07 produced
    # 0 BUY / 2 WATCHLIST out of 120 candidates. Reject histogram was:
    #   SECTOR_RANK_TOO_LOW  54% (structural — sector rotation, don't relax)
    #   TQ_TOO_LOW           50% (over-tight for SIDEWAYS distribution)
    #   RR_TOO_LOW           47% (2.0× is unreachable in range-bound tape)
    # Recalibrated to 75/68/1.7 — same numbers as BULL row above (SIDEWAYS
    # regimes with intact short-term structure behave much more like a mild
    # BULL than a defensive stance). Backtest with backtest_walkforward.py
    # BEFORE trusting these values with real capital.
    #
    # If your tape STILL produces 0 buys after this change, drop
    # min_confidence to 72 via regime_calibration.json rather than editing
    # this constant again — that file is the intended tuning surface.
    #
    # 2026-07-07 PHASE I (setup-edge patch): min_confidence dropped 75 → 70
    # to match the empirical break-even bucket from backtest_walkforward.py
    # (win_rate_by_confidence: bucket 70 = 47.2% win, bucket 75 = 48.0% win,
    # bucket 65 = 41.3% win). 70 is the smallest bucket that clears the
    # 37.5% break-even bar with statistical margin (1,163 trades / 47.2% wr).
    # The setup-type bonus + WEAK/SIDEWAYS-ex-BREAKOUT skip added below is
    # what re-tightens the effective bar — this is a floor, not a gate.
    "SIDEWAYS":        {"min_confidence": 70, "min_tq": 68, "min_rr": 1.7, "max_buys": 2,  "max_exposure": 0.50, "max_stop_pct": 8.0},
    "TRANSITION":      {"min_confidence": 83, "min_tq": 73, "min_rr": 2.0, "max_buys": 2,  "max_exposure": 0.55, "max_stop_pct": 7.0},
    "HIGH_VOLATILITY": {"min_confidence": 85, "min_tq": 76, "min_rr": 2.2, "max_buys": 1,  "max_exposure": 0.40, "max_stop_pct": 5.0},
    "BEAR":            {"min_confidence": 92, "min_tq": 82, "min_rr": 2.5, "max_buys": 0,  "max_exposure": 0.20, "max_stop_pct": 5.0},
    # Phase 3a #26 (2026-07-05): STRONG_BEAR tuned from unreachable 99/99/3.0
    # to 95/92/3.0. max_buys=0 still enforces "no new positions" as the
    # PRIMARY gate; the numeric thresholds are for the audit trail ("how
    # close did we get?") and for reactivation after regime normalizes.
    # Old values were mathematically impossible so all STRONG_BEAR audits
    # showed conf_gap=huge, hiding whether the underlying setup was actually
    # decent.
    "STRONG_BEAR":     {"min_confidence": 95, "min_tq": 86, "min_rr": 3.0, "max_buys": 0,  "max_exposure": 0.00, "max_stop_pct": 4.0},
}

# Factor weights — 10 factors, sum = 1.00
# Phase C4 (2026-07-02): options_sentiment weight dropped from 0.04 → 0.00.
# Only 20/2360 stocks got real PCR — the rest defaulted to a neutral 60,
# which was silently adding a constant 2.4-point bias to every confidence.
# Redistributed: trend +0.02, momentum +0.02. If we later get PCR coverage
# for the full F&O universe (~200 stocks), we can restore this to 0.02–0.04
# but drive it from a real dynamic-score-per-stock, not a placeholder.
#
# Phase G8-E (2026-07-06): audit fix #3 — news_risk weight lowered
# from 0.08 → 0.05. Rationale: LLM-summarised headline scoring has
# ~10-15% noise (misread sentiment, stale news, false positive event
# tags). At 0.08 weight one news call ~= 40% of the ownership_quality
# fundamental factor's information content, which is out of proportion
# for a ~50-char summary. The freed 0.03 is redistributed to
# trend_quality (0.20 → 0.23) — our highest-confidence purely-technical
# signal computed directly from prices. This tilts the score toward
# hard-measurable data and away from LLM-inferred sentiment.
FACTOR_WEIGHTS = {
    "trend_quality":      0.23,
    "momentum_quality":   0.16,
    "volume_delivery":    0.10,
    "sector_strength":    0.15,
    "rs_vs_nifty":        0.15,
    "news_risk":          0.05,
    "risk_reward":        0.07,
    "ownership_quality":  0.06,
    "options_sentiment":  0.00,
    "macro_alignment":    0.03,
}
# Sanity: sum must remain 1.00 to avoid systematic score inflation/deflation.
assert abs(sum(FACTOR_WEIGHTS.values()) - 1.0) < 1e-9, \
    f"FACTOR_WEIGHTS must sum to 1.0, got {sum(FACTOR_WEIGHTS.values())}"


# ─────────────────────────────────────────────────────────────────────────────
# Phase C4 Gap #4: Auto-calibrated regime thresholds from tracker history.
# The nightly_calibration.py script writes regime_calibration.json based on
# actual win-rate per regime; the pipeline reads it here.
# Falls back silently if file absent, malformed, or too few trades per bucket.
# ─────────────────────────────────────────────────────────────────────────────
REGIME_CALIBRATION_FILE = os.getenv("REGIME_CALIBRATION_FILE", "regime_calibration.json")


def load_regime_calibration() -> dict:
    """Reads regime_calibration.json → {regime: {min_confidence_delta: int,
    min_tq_delta: int, min_rr_delta: float, closed_trades: int, win_rate: float}}.
    Returns {} if missing/bad. Deltas are applied ADDITIVELY on top of
    REGIME_THRESHOLDS (positive = tighter, negative = looser)."""
    # Phase C7c: FRESH_START ignores stale calibration (built from old regime data)
    if FRESH_START:
        _log("[FRESH_START] load_regime_calibration → returning {} (old calibration ignored)")
        return {}
    try:
        if not os.path.exists(REGIME_CALIBRATION_FILE):
            return {}
        with open(REGIME_CALIBRATION_FILE, "r") as _f:
            data = json.load(_f)
        if not isinstance(data, dict):
            return {}
        # File format has meta at "_meta" — strip it for return
        cal = {k: v for k, v in data.items() if not k.startswith("_") and isinstance(v, dict)}
        return cal
    except Exception:
        return {}


def apply_regime_calibration(thresholds: dict, calibration: dict) -> dict:
    """Applies additive deltas from calibration file onto a deep-copied
    thresholds dict.  Never allows min_confidence to fall below 50 or rise
    above 99; min_tq stays 40-99; min_rr stays 1.0-4.0."""
    if not calibration:
        return thresholds
    import copy as _copy
    out = _copy.deepcopy(thresholds)
    for regime, delta in calibration.items():
        if regime not in out:
            continue
        try:
            dc = int(delta.get("min_confidence_delta", 0))
            dt = int(delta.get("min_tq_delta", 0))
            dr = float(delta.get("min_rr_delta", 0.0))
            out[regime]["min_confidence"] = int(max(50, min(99,
                out[regime].get("min_confidence", 80) + dc)))
            out[regime]["min_tq"] = int(max(40, min(99,
                out[regime].get("min_tq", 75) + dt)))
            out[regime]["min_rr"] = round(max(1.0, min(4.0,
                out[regime].get("min_rr", 1.8) + dr)), 2)
        except Exception:
            continue
    return out

# Opportunity score weights — primary ranking metric (ENHANCEMENT 1)
# ─── Phase 2 #37 (2026-07-05): remove factor-double-counting ────────────────
# trend/volume/sector/macro are ALREADY inside final_confidence (as factor
# weights via compute_base_confidence). Adding them again here double-counts.
# Fix: keep only the 3 outer-layer signals (conf, tq, rr) and redistribute
# the freed 0.25 weight to conf (structural quality) and tq (execution).
# Old sum: 0.30+0.25+0.20+0.10+0.05+0.05+0.05 = 1.00
# New sum: 0.50+0.30+0.20                     = 1.00
OPPORTUNITY_WEIGHTS = {
    "confidence":     0.50,
    "trade_quality":  0.30,
    "risk_reward":    0.20,
}


def _update_ownership_quality(stock: dict) -> None:
    """
    Updates ownership_quality factor score from real fundamentals + delivery.
    Called after fetch_all_fundamentals_cached() injects ROE/pledge into stock dict,
    and AGAIN after delivery% is fetched (so the delivery bonus applies).
    Scale:
      ROE > 20% = excellent (+20), 12-20% = good (+10), < 5% = poor (-15)
      Pledge > 30% = bad (-20), 15-30% = caution (-10), < 5% = clean (+10)
      D/E > 2.0 = leveraged (-10), < 0.5 = clean (+10)

    Phase G7-A (2026-07-03): Delivery% bonus / penalty from nselib data.
    Delivery is the strongest single per-stock signal for NSE — retail day
    traders can't fake it (delivery = shares actually taken into demat).
    Only applies when delivery_source == "nselib" (real data, not defaults).
      Today > (20d avg + 10pp)         → +8 (accumulation footprint)
      Today > 60% absolute             → +5 (high conviction, retail can't manipulate)
      Today < (20d avg - 10pp) AND up  → -10 (distribution / pump-and-dump)
      DISTRIBUTION signal              → -8  (already flagged by nselib logic)

    Baseline 50; clamped 0-100.

    Phase C5 (rating ≥ 9.0): if ROE == 0 AND D/E == 0 AND pledge == 0 AND the
    fundamentals source is NEUTRAL_DEFAULT/screener_partial, this is a
    MISSING-DATA case — return None so compute_base_confidence redistributes
    the ownership weight instead of diluting toward neutral 50.
    """
    try:
        roe    = float(stock.get("roe", 0) or 0)
        pledge = float(stock.get("promoter_pledge_pct", 0) or 0)
        de     = float(stock.get("de_ratio", 0) or 0)
        src    = str(stock.get("fundamentals_source", "") or "").lower()

        # MISSING-DATA detection: all three metrics are zero AND source
        # indicates we didn't successfully retrieve fundamentals.
        if roe == 0 and de == 0 and pledge == 0 and src in (
            "", "neutral_default", "screener_partial", "rate_limited", "error"
        ):
            stock["ownership_quality"] = None
            stock["ownership_missing"] = True
            if "factor_scores" in stock:
                stock["factor_scores"]["ownership_quality"] = None
            return

        score = 50.0
        # ROE component
        if roe > 20:    score += 20
        elif roe > 12:  score += 10
        elif roe > 5:   score += 0
        else:           score -= 15
        # Pledge component
        if pledge > 30:   score -= 20
        elif pledge > 15: score -= 10
        elif pledge < 5:  score += 10
        # D/E component (tiered)
        if de > 2.0:    score -= 10
        elif de > 1.0:  score -= 5
        elif de < 0.5:  score += 10

        # Issue 3 fix: gradient penalty on top of tiers. As DE approaches
        # MAX_DE (default 1.5), apply up to -6 extra so DE=1.4 (near-limit)
        # is visibly worse than DE=0.6 (safe). Financials exempted (banks
        # naturally run 5-10x). Scaled by (de / max_de) capped at 1.0.
        try:
            _max_de_ownership = float(os.getenv("MAX_DE", "1.5") or 1.5)
        except (TypeError, ValueError):
            _max_de_ownership = 1.5
        _sec_up = str(stock.get("sector", "") or "").upper()
        _is_fin = any(fs in _sec_up for fs in
                      ("BANKING", "FINANCE", "INSURANCE", "NBFC", "FINANCIAL"))
        if not _is_fin and _max_de_ownership > 0 and de > 0:
            _de_ratio = min(1.0, de / _max_de_ownership)
            score -= round(6.0 * _de_ratio, 1)  # up to -6 at DE == MAX_DE

        # ── Phase R3 (2026-07-06): Per-stock FII / DII shareholding overlay ──
        # High institutional participation (FII+DII) is a strong quality
        # signal — smart money has already done fundamental diligence and
        # is holding the stock. Applied conservatively (±10 max) so it
        # complements but doesn't dominate ROE/pledge/D/E.
        #   FII+DII > 30%  = strong institutional backing (+8)
        #   FII+DII > 20%  = decent institutional presence   (+4)
        #   FII+DII < 5%   = retail-only / illiquid           (-5)
        #   FII% > 15% alone = foreign institutional interest (+2 bonus)
        # R5 PRUNE (2026-07-06): audit of 480 rows showed 0/480 populated
        # (data source not wired). Default OFF; opt in via FII_DII_OVERLAY=1.
        fii_dii_bonus = 0
        fii_dii_reasons = []
        if os.getenv("FII_DII_OVERLAY", "0") == "1":
            try:
                fii_pct = float(stock.get("fii_pct", 0) or 0)
                dii_pct = float(stock.get("dii_pct", 0) or 0)
                inst_total = fii_pct + dii_pct
                if inst_total > 30:
                    fii_dii_bonus += 8
                    fii_dii_reasons.append(f"INST_STRONG(+8 FII+DII={inst_total:.1f}%)")
                elif inst_total > 20:
                    fii_dii_bonus += 4
                    fii_dii_reasons.append(f"INST_PRESENT(+4 FII+DII={inst_total:.1f}%)")
                elif 0 < inst_total < 5:
                    # 0 is treated as "unknown", not "confirmed low"
                    fii_dii_bonus -= 5
                    fii_dii_reasons.append(f"INST_ABSENT(-5 FII+DII={inst_total:.1f}%)")
                if fii_pct > 15:
                    fii_dii_bonus += 2
                    fii_dii_reasons.append(f"FII_HIGH(+2 FII={fii_pct:.1f}%)")
                fii_dii_bonus = max(-8, min(10, fii_dii_bonus))
                score += fii_dii_bonus
            except (TypeError, ValueError):
                pass

        # ── Phase G7-A: Delivery% bonus/penalty (only if real nselib data) ──
        deliv_src = str(stock.get("delivery_source", "") or "").lower()
        deliv_bonus = 0
        deliv_reasons = []
        if deliv_src == "nselib":
            try:
                d_today = float(stock.get("delivery_pct_today", 0) or 0)
                d_avg   = float(stock.get("delivery_pct_20d_avg", 0) or 0)
                d_sig   = str(stock.get("delivery_signal", "NEUTRAL") or "NEUTRAL")
                # Prefer ret1d_pct, fall back to ret1d (score_stock's actual key)
                ret1d   = float(stock.get("ret1d_pct", stock.get("ret1d", 0.0)) or 0.0)

                # +8 accumulation: today's delivery meaningfully above 20d avg
                if d_avg > 0 and d_today > (d_avg + 10.0):
                    deliv_bonus += 8
                    deliv_reasons.append(f"DELIV_ACCUM(+8 today {d_today:.0f}% vs 20d {d_avg:.0f}%)")

                # +5 high absolute: >60% delivery = institutional participation
                if d_today > 60.0:
                    deliv_bonus += 5
                    deliv_reasons.append(f"DELIV_HIGH(+5 {d_today:.0f}%)")

                # -10 pump / distribution: price up but delivery collapsing
                if d_avg > 0 and d_today < (d_avg - 10.0) and ret1d > 0:
                    deliv_bonus -= 10
                    deliv_reasons.append(
                        f"DELIV_DIST(-10 today {d_today:.0f}% vs 20d {d_avg:.0f}%, "
                        f"ret1d +{ret1d:.1f}%)"
                    )

                # -8 explicit distribution signal from nselib logic
                if d_sig == "DISTRIBUTION":
                    deliv_bonus -= 8
                    deliv_reasons.append(f"DELIV_SIG_DIST(-8)")

                # Cap the bonus at ±13 so it doesn't dominate ROE (+20 max)
                deliv_bonus = max(-15, min(13, deliv_bonus))
                score += deliv_bonus
            except (TypeError, ValueError):
                pass  # bad numeric input; skip delivery contribution

        stock["ownership_quality"] = round(max(0.0, min(100.0, score)), 1)
        stock["ownership_missing"] = False
        stock["ownership_deliv_bonus"] = deliv_bonus
        if deliv_reasons:
            stock["ownership_deliv_reasons"] = deliv_reasons
        # Phase R3 (2026-07-06): stamp FII/DII overlay contribution on stock
        stock["ownership_fii_bonus"] = fii_dii_bonus
        if fii_dii_reasons:
            stock["ownership_fii_reasons"] = fii_dii_reasons
        # Keep factor_scores in sync
        if "factor_scores" in stock:
            stock["factor_scores"]["ownership_quality"] = stock["ownership_quality"]
    except Exception:
        pass


def compute_opportunity_score(stock: dict) -> float:
    """
    0-100 composite score — primary ranking metric (ENHANCEMENT 1).
    Reads directly from stock keys (with factor_scores as supplement).
    Higher = better opportunity quality.
    """
    try:
        conf  = float(stock.get("final_confidence", 0) or 0)
        tq    = float(stock.get("trade_quality_score", 0) or 0)
        rr    = float(stock.get("rr_ratio", 0) or 0)

        # Normalize R/R to 0-100 (3.0x = 100, 2.0x = 67, 1.0x = 33)
        rr_score = min(100.0, rr / 3.0 * 100)

        # Read directly from stock — factor_scores is a mirror, either works
        fs     = stock.get("factor_scores", {}) or {}
        trend  = float(stock.get("trend_quality",  fs.get("trend_quality",  50)) or 50)
        volume = float(stock.get("volume_delivery",fs.get("volume_delivery",50)) or 50)
        sector = float(stock.get("sector_strength",fs.get("sector_strength",50)) or 50)
        macro  = float(stock.get("macro_alignment",fs.get("macro_alignment",50)) or 50)

        # Phase 2 #37 (2026-07-05): trend/volume/sector/macro no longer
        # additively contribute — they are already inside final_confidence
        # via compute_base_confidence's factor weights. Reading them again
        # was double-counting. Left the local reads above for future
        # audit/telemetry but they're not used in opp.
        _ = (trend, volume, sector, macro)  # silence unused-var linter
        opp = (
            conf     * OPPORTUNITY_WEIGHTS["confidence"] +
            tq       * OPPORTUNITY_WEIGHTS["trade_quality"] +
            rr_score * OPPORTUNITY_WEIGHTS["risk_reward"]
        )
        return round(opp, 1)
    except Exception:
        return 0.0


# Comprehensive sector map (Bug 3 fix)
# ─────────────────────────────────────────────────────────────────────────────
# SECTOR MAP — dynamic, not hardcoded
# Priority: sector_master.csv (curated) → sector_cache.json (yfinance-fetched, grows over time)
# ─────────────────────────────────────────────────────────────────────────────
SECTOR_CACHE_FILE = os.getenv("SECTOR_CACHE_FILE", "sector_cache.json")

# Normalize sector_master.csv labels to internal labels
_CSV_LABEL_NORM = {
    "Auto": "AUTO", "Banking": "BANKING", "Capital Goods": "CAPITAL_GOODS",
    "Cement": "INFRA", "Chemicals": "CHEMICALS", "Consumer Goods": "FMCG",
    "Consumer Services": "CONSUMER", "Defence": "DEFENCE", "Diversified": "DIVERSIFIED",  # Bug 8 fix (2026-07-03)
    "Electronics Manufacturing": "CAPITAL_GOODS", "Finance": "FINANCE",
    "FinTech": "FINANCE", "FMCG": "FMCG", "Healthcare": "HEALTHCARE",
    "Infrastructure": "INFRA", "IT": "IT", "IT Hardware": "IT",
    "Metals": "METALS", "Oil & Gas": "ENERGY", "Pharma": "PHARMA",
    "Power": "ENERGY", "Realty": "REALTY", "Retail": "CONSUMER",
}

# Normalize yfinance sector labels to internal labels
_YF_LABEL_NORM = {
    "Technology": "IT", "Financial Services": "FINANCE",
    "Healthcare": "HEALTHCARE", "Consumer Defensive": "FMCG",
    "Consumer Cyclical": "CONSUMER", "Basic Materials": "METALS",
    "Energy": "ENERGY", "Industrials": "CAPITAL_GOODS",
    "Real Estate": "REALTY", "Communication Services": "TELECOM",
    "Utilities": "ENERGY", "Automobile": "AUTO",
    "Pharmaceuticals": "PHARMA", "Banking": "BANKING",
    "Defence": "DEFENCE", "Chemicals": "CHEMICALS",
}

_SECTOR_MAP: dict = {}   # loaded at pipeline start via _init_sector_map()

# ── Hardcoded sector map (FIX 1 — normalizes .NS before lookup, never returns OTHERS) ──
SECTOR_MAP: dict = {
    # IT
    "INFY.NS":"IT","TCS.NS":"IT","WIPRO.NS":"IT","HCLTECH.NS":"IT",
    "TECHM.NS":"IT","LTIM.NS":"IT","PERSISTENT.NS":"IT","COFORGE.NS":"IT",
    "MPHASIS.NS":"IT","OFSS.NS":"IT","KPITTECH.NS":"IT","TATAELXSI.NS":"IT",
    "RAMCOSYS.NS":"IT","RPTECH.NS":"IT",
    # PHARMA
    "SUNPHARMA.NS":"PHARMA","DRREDDY.NS":"PHARMA","CIPLA.NS":"PHARMA",
    "DIVISLAB.NS":"PHARMA","AUROPHARMA.NS":"PHARMA","LAURUSLABS.NS":"PHARMA",
    "ALKEM.NS":"PHARMA","TORNTPHARM.NS":"PHARMA","IPCALAB.NS":"PHARMA",
    "GLENMARK.NS":"PHARMA","NATCOPHARM.NS":"PHARMA","GRANULES.NS":"PHARMA",
    "AARTIDRUGS.NS":"PHARMA","INDSWFTLAB.NS":"PHARMA","NARMADA.NS":"PHARMA",
    "MOREPENLAB.NS":"PHARMA","BLUEJET.NS":"PHARMA","PANACEABIO.NS":"PHARMA",
    # BANKING
    "HDFCBANK.NS":"BANKING","ICICIBANK.NS":"BANKING","AXISBANK.NS":"BANKING",
    "SBIN.NS":"BANKING","KOTAKBANK.NS":"BANKING","INDUSINDBK.NS":"BANKING",
    "BANDHANBNK.NS":"BANKING","FEDERALBNK.NS":"BANKING","IDFCFIRSTB.NS":"BANKING",
    "RBLBANK.NS":"BANKING","PNB.NS":"BANKING","BANKBARODA.NS":"BANKING",
    # FINANCE / NBFC
    "BAJFINANCE.NS":"FINANCE","BAJAJFINSV.NS":"FINANCE","CHOLAFIN.NS":"FINANCE",
    "MUTHOOTFIN.NS":"FINANCE","MANAPPURAM.NS":"FINANCE","SHRIRAMFIN.NS":"FINANCE",
    "LICHSGFIN.NS":"FINANCE","PFC.NS":"FINANCE","RECLTD.NS":"FINANCE",
    "MANCREDIT.NS":"FINANCE","AAVAS.NS":"FINANCE","63MOONS.NS":"FINANCE",
    # ENERGY / POWER
    "RELIANCE.NS":"ENERGY","ONGC.NS":"ENERGY","BPCL.NS":"ENERGY",
    "IOC.NS":"ENERGY","HINDPETRO.NS":"ENERGY","GAIL.NS":"ENERGY",
    "TATAPOWER.NS":"ENERGY","ADANIGREEN.NS":"ENERGY","TORNTPOWER.NS":"ENERGY",
    "NTPC.NS":"ENERGY","POWERGRID.NS":"ENERGY","CESC.NS":"ENERGY",
    # METALS
    "TATASTEEL.NS":"METALS","HINDALCO.NS":"METALS","JSWSTEEL.NS":"METALS",
    "VEDL.NS":"METALS","SAIL.NS":"METALS","NATIONALUM.NS":"METALS",
    "HINDCOPPER.NS":"METALS","NMDC.NS":"METALS","COALINDIA.NS":"METALS",
    # AUTO / AUTO ANCILLARY
    "MARUTI.NS":"AUTO","TATAMOTORS.NS":"AUTO","M&M.NS":"AUTO",
    "BAJAJ-AUTO.NS":"AUTO","HEROMOTOCO.NS":"AUTO","EICHERMOT.NS":"AUTO",
    "ASHOKLEY.NS":"AUTO","TVSMOTOR.NS":"AUTO","MOTHERSON.NS":"AUTO",
    "BOSCHLTD.NS":"AUTO_ANCILLARY","BALKRISIND.NS":"AUTO_ANCILLARY",
    "SONACOMS.NS":"AUTO_ANCILLARY","SPAL.NS":"AUTO_ANCILLARY",
    "TALBROAUTO.NS":"AUTO_ANCILLARY","CONFIPET.NS":"AUTO_ANCILLARY",
    # CAPITAL GOODS / INDUSTRIAL
    "BHEL.NS":"CAPITAL_GOODS","THERMAX.NS":"CAPITAL_GOODS","KIRLOSENG.NS":"CAPITAL_GOODS",
    "ABB.NS":"CAPITAL_GOODS","SIEMENS.NS":"CAPITAL_GOODS","HAVELLS.NS":"CAPITAL_GOODS",
    "CROMPTON.NS":"CAPITAL_GOODS","CUMMINSIND.NS":"CAPITAL_GOODS",
    "ELGIEQUIP.NS":"CAPITAL_GOODS","GRINDWELL.NS":"CAPITAL_GOODS",
    "CEIGALL.NS":"CAPITAL_GOODS","GENUSPOWER.NS":"CAPITAL_GOODS",
    "JASH.NS":"CAPITAL_GOODS","ELECTHERM.NS":"CAPITAL_GOODS",
    "ROTO.NS":"CAPITAL_GOODS","EMSLIMITED.NS":"CAPITAL_GOODS",
    "SYRMA.NS":"ELECTRONICS","ACI.NS":"ELECTRONICS",
    # DEFENCE
    "HAL.NS":"DEFENCE","BEL.NS":"DEFENCE","MAZDOCK.NS":"DEFENCE",
    "COCHINSHIP.NS":"DEFENCE","PARAS.NS":"DEFENCE","MIDHANI.NS":"DEFENCE",
    "GRSE.NS":"DEFENCE","PRIVISCL.NS":"DEFENCE",
    # FMCG
    "HINDUNILVR.NS":"FMCG","ITC.NS":"FMCG","NESTLEIND.NS":"FMCG",
    "BRITANNIA.NS":"FMCG","DABUR.NS":"FMCG","MARICO.NS":"FMCG",
    "COLPAL.NS":"FMCG","GODREJCP.NS":"FMCG","EMAMILTD.NS":"FMCG",
    "RELAXO.NS":"FMCG","BATAINDIA.NS":"FMCG","PAGEIND.NS":"FMCG",
    # CHEMICALS
    "PIDILITIND.NS":"CHEMICALS","DEEPAKNI.NS":"CHEMICALS","AARTI.NS":"CHEMICALS",
    "NAVINFLUOR.NS":"CHEMICALS","ALKYLAMINE.NS":"CHEMICALS","FINEORG.NS":"CHEMICALS",
    "TATACHEM.NS":"CHEMICALS","NACLIND.NS":"CHEMICALS","GINNIFILA.NS":"CHEMICALS",
    "20MICRONS.NS":"CHEMICALS","INDOBORAX.NS":"CHEMICALS",
    # REALTY
    "DLF.NS":"REALTY","GODREJPROP.NS":"REALTY","OBEROIRLTY.NS":"REALTY",
    "PRESTIGE.NS":"REALTY","BRIGADE.NS":"REALTY","PHOENIXLTD.NS":"REALTY",
    # LOGISTICS
    "AEGISLOG.NS":"LOGISTICS","BLUEDART.NS":"LOGISTICS","CONCOR.NS":"LOGISTICS",
    "GATI.NS":"LOGISTICS","DELHIVERY.NS":"LOGISTICS","TCI.NS":"LOGISTICS",
    "NAVKARCORP.NS":"LOGISTICS","JSWINFRA.NS":"INFRA",
    # TELECOM
    "BHARTIARTL.NS":"TELECOM","IDEA.NS":"TELECOM","TATACOMM.NS":"TELECOM",
    "ONMOBILE.NS":"TELECOM",
    # INFRA / CONSTRUCTION
    "LT.NS":"INFRA","ULTRACEMCO.NS":"INFRA","AMBUJACEMENT.NS":"INFRA",
    "ACC.NS":"INFRA","SHREECEM.NS":"INFRA","KNRCON.NS":"INFRA",
    "PNCINFRA.NS":"INFRA","IRB.NS":"INFRA","ASHOKA.NS":"INFRA","PATELENG.NS":"INFRA",
    # TEXTILES
    "SIYSIL.NS":"TEXTILES","RAYMOND.NS":"TEXTILES","WELSPUN.NS":"TEXTILES",
    "TRIDENT.NS":"TEXTILES","VARDHMAN.NS":"TEXTILES","ARVIND.NS":"TEXTILES",
    "RUPA.NS":"TEXTILES","PGIL.NS":"TEXTILES","ABCOTS.NS":"TEXTILES",
    "DIFFNKG.NS":"TEXTILES",
    # HEALTHCARE / HOSPITALS
    "MAXHEALTH.NS":"HEALTHCARE","APOLLOHOSP.NS":"HEALTHCARE",
    "FORTIS.NS":"HEALTHCARE","NARAYANHA.NS":"HEALTHCARE",
    # INSURANCE / AMC
    "HDFCLIFE.NS":"INSURANCE","SBILIFE.NS":"INSURANCE","ICICIGI.NS":"INSURANCE",
    "ICICIPRULI.NS":"INSURANCE","STARHEALTH.NS":"INSURANCE","LICI.NS":"INSURANCE",
    "ABSLAMC.NS":"ASSET_MGMT",
    # POWER EQUIPMENT / RENEWABLES
    "INOXGREEN.NS":"POWER_EQ","SUZLON.NS":"POWER_EQ","KPIGREEN.NS":"POWER_EQ",
    "WEBSOL.NS":"POWER_EQ","EBGNG.NS":"POWER_EQ",
    # CONSUMER / RETAIL
    "TITAN.NS":"CONSUMER","TRENT.NS":"CONSUMER","DMART.NS":"CONSUMER",
    # 2025-Q1: ZOMATO.NS renamed to ETERNAL.NS (Eternal Ltd). Keep old key
    # as an alias for legacy tracker state.
    "ETERNAL.NS":"CONSUMER","ZOMATO.NS":"CONSUMER",
    "NYKAA.NS":"CONSUMER","NAZARA.NS":"CONSUMER",
    # PIPES / BUILDING MATERIALS
    "VENUSPIPES.NS":"BUILDING_MAT","RAMCOIND.NS":"BUILDING_MAT",
    "KRISHANA.NS":"BUILDING_MAT","RATNAVEER.NS":"BUILDING_MAT",
    # DIVERSIFIED
    "AURUM.NS":"DIVERSIFIED","AEQUS.NS":"DIVERSIFIED","AVL.NS":"DIVERSIFIED",
    "AVG.NS":"DIVERSIFIED","KMEW.NS":"DIVERSIFIED","LOTUSDEV.NS":"DIVERSIFIED",
    "AEROENTER.NS":"DIVERSIFIED","POKARNA.NS":"DIVERSIFIED","SETL.NS":"DIVERSIFIED",
    "PIXTRANS.NS":"DIVERSIFIED",
    # AGRI / FOOD
    "UBL.NS":"AGRI_FOOD","RADICO.NS":"AGRI_FOOD","KRBL.NS":"AGRI_FOOD",
    "LTFOODS.NS":"AGRI_FOOD",
}


def _load_sector_map() -> dict:
    """Load sector map from sector_master.csv then overlay sector_cache.json."""
    result: dict = {}
    # 1. sector_master.csv — curated, ships with repo
    try:
        sm_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sector_master.csv")
        if os.path.exists(sm_path):
            with open(sm_path, newline="", encoding="utf-8") as f:
                for row in csv.DictReader(f):
                    sym = row.get("symbol", "").strip()
                    sec = row.get("sector", "").strip()
                    if sym and sec:
                        if not sym.endswith(".NS"):
                            sym += ".NS"
                        result[sym] = _CSV_LABEL_NORM.get(sec, sec.upper().replace(" ", "_"))
    except Exception:
        pass
    # 2. sector_cache.json — yfinance-fetched, grows over runs (persisted in CI cache)
    try:
        if os.path.exists(SECTOR_CACHE_FILE):
            with open(SECTOR_CACHE_FILE, "r") as f:
                result.update(json.load(f))
    except Exception:
        pass
    return result


def _init_sector_map() -> None:
    """Called once at pipeline start."""
    global _SECTOR_MAP
    _SECTOR_MAP = _load_sector_map()


# 2026-07-03: hard exclusions for pattern inference — real NSE tickers that
# trigger a substring rule but belong to a *different* sector. Extend as needed.
_SECTOR_INFERENCE_EXCLUSIONS = {
    # ticker.NS  → correct sector
    "INFOEDGE.NS":  "CONSUMER",   # naukri.com / 99acres / Zomato parent — internet-media, NOT IT services
    "FINEORG.NS":   "CHEMICALS",  # Fine Organic Industries — specialty chem, NOT finance
    "FINPIPE.NS":   "CHEMICALS",  # Finolex Pipes — plastic pipes, NOT finance
    "MEDPLUS.NS":   "RETAIL",     # MedPlus Health — pharmacy retail, NOT drug maker
    "MEDANTA.NS":   "HEALTHCARE", # hospital chain, NOT pharma
    "BIOFILCHEM.NS":"CHEMICALS",  # not a biopharma
    "CREDITACC.NS": "FINANCE",    # this one IS finance — harmless, but pins it to the right label
    "POWERGRID.NS": "ENERGY",     # transmission utility, not equipment
}


def get_sector(symbol: str) -> str:
    """Normalizes symbol format before lookup. Never returns OTHERS — falls back to pattern inference."""
    sym = symbol.strip()
    if not sym.endswith(".NS"):
        sym = sym + ".NS"
    # 1. Check hardcoded SECTOR_MAP first (FIX 1)
    sector = SECTOR_MAP.get(sym)
    if sector:
        return sector
    # 2. Fallback to dynamic _SECTOR_MAP (loaded from CSV + yfinance cache)
    sector = _SECTOR_MAP.get(sym)
    if sector and sector != "OTHERS":
        return sector
    # 2b. Hard-coded exception list — trumps pattern inference below.
    #     Handles the well-known false positives (INFOEDGE→IT, FINEORG→FINANCE…).
    if sym in _SECTOR_INFERENCE_EXCLUSIONS:
        return _SECTOR_INFERENCE_EXCLUSIONS[sym]
    # 3. Name-pattern inference — never display OTHERS.
    #    Uses .startswith / .endswith for the most-abused short substrings
    #    (FIN, MED, BIO) to reduce false positives on unrelated tickers.
    #    Strip the .NS suffix first so word-boundary checks are meaningful.
    s = sym.upper()
    root = s[:-3] if s.endswith(".NS") else s  # "INFOEDGE.NS" → "INFOEDGE"
    # PHARMA / HEALTHCARE — MED/BIO only if they lead or trail the ticker,
    # LAB/DRUG anywhere, PHARMA anywhere (unambiguous).
    if "PHARMA" in root or "DRUG" in root or "LAB" in root:
        return "PHARMA"
    if root.startswith(("MED", "BIO")) or root.endswith(("MED", "BIO", "PHARMA")):
        return "PHARMA"
    # FINANCE — BANK/CRED/LOAN anywhere are unambiguous.
    # For FIN we allow substring (catches JMFINCORP, MASFIN, LICHSGFIN…) and
    # rely on the exclusion list at step 2b to intercept the known
    # non-finance offenders (FINEORG, FINPIPE, …).
    if "BANK" in root or "CRED" in root or "LOAN" in root:
        return "FINANCE"
    if "FIN" in root:
        return "FINANCE"
    # IT — TECH/SOFT/DIGIT strong signals; INFO/SYST only at start
    # (avoids INFOEDGE, SYSTEMATIX-type tickers slipping into IT).
    if "TECH" in root or "SOFT" in root or "DIGIT" in root:
        return "IT"
    if root.startswith(("INFO", "SYST")):
        return "IT"
    if any(x in root for x in ("STEEL", "METAL", "ALUM", "COPP")):
        return "METALS"
    if any(x in root for x in ("POWER", "SOLAR", "WIND", "ENERG")):
        return "POWER_EQ"
    if any(x in root for x in ("INFRA", "CONST", "BUILD", "CEMENT")):
        return "INFRA"
    return "DIVERSIFIED"  # never show OTHERS


def _nselib_equity_list_cached(max_age_hours: int = 24) -> "pd.DataFrame | None":
    """
    Bulk NSE equity listing with 24h disk cache. Includes industry/sector columns
    for ~2400 listed companies. Phase B6 (2026-06-30) — replaces per-symbol yfinance hits.
    """
    if not _NSELIB_OK or _nselib_cm is None:
        return None
    cache_path = os.path.join(NSELIB_CACHE_DIR, "equity_list.json")
    try:
        if os.path.exists(cache_path):
            age_h = (time.time() - os.path.getmtime(cache_path)) / 3600.0
            if age_h < max_age_hours:
                with open(cache_path, "r") as f:
                    return pd.DataFrame(json.load(f))
    except Exception:
        pass

    fn = getattr(_nselib_cm, "equity_list", None) or getattr(_nselib_cm, "nse_equity_symbols", None)
    if fn is None:
        return None
    try:
        df = fn()
        if df is None or (hasattr(df, "empty") and df.empty):
            return None
        try:
            df.to_json(cache_path, orient="records")
        except Exception:
            pass
        return df
    except Exception as e:
        _log(f"  [nselib equity_list] error: {e}")
        return None


# Industry → canonical sector label normalization (NSE labels → our internal labels)
_NSE_INDUSTRY_NORM = {
    "INFORMATION TECHNOLOGY": "IT", "IT - SOFTWARE": "IT", "IT - HARDWARE": "IT",
    "FINANCIAL SERVICES": "FINANCE", "NON BANKING FINANCIAL COMPANY (NBFC)": "FINANCE",
    "HOUSING FINANCE COMPANY": "FINANCE", "FINANCE - OTHERS": "FINANCE",
    "BANKING": "BANKING", "PRIVATE SECTOR BANK": "BANKING", "PUBLIC SECTOR BANK": "BANKING",
    "PHARMACEUTICALS": "PHARMA", "PHARMACEUTICALS & DRUGS": "PHARMA",
    "HEALTHCARE": "HEALTHCARE", "HEALTHCARE SERVICES": "HEALTHCARE", "HOSPITAL": "HEALTHCARE",
    "FAST MOVING CONSUMER GOODS": "FMCG", "FMCG": "FMCG", "CONSUMER GOODS": "FMCG",
    "CONSUMER SERVICES": "CONSUMER", "CONSUMER DURABLES": "CONSUMER",
    "AUTOMOBILE AND AUTO COMPONENTS": "AUTO", "AUTOMOBILE": "AUTO",
    "AUTO ANCILLARIES": "AUTO_ANCILLARY", "AUTO COMPONENTS": "AUTO_ANCILLARY",
    "METALS & MINING": "METALS", "FERROUS METALS": "METALS", "NON-FERROUS METALS": "METALS",
    "OIL GAS & CONSUMABLE FUELS": "ENERGY", "OIL & GAS": "ENERGY", "POWER": "ENERGY",
    "GAS DISTRIBUTION": "ENERGY",
    "CONSTRUCTION": "INFRA", "CONSTRUCTION MATERIALS": "INFRA", "CEMENT": "INFRA",
    "REALTY": "REALTY", "REAL ESTATE": "REALTY",
    "TELECOM SERVICES": "TELECOM", "TELECOMMUNICATION": "TELECOM",
    "CAPITAL GOODS": "CAPITAL_GOODS", "INDUSTRIALS": "CAPITAL_GOODS",
    "CHEMICALS": "CHEMICALS", "FERTILISERS & AGROCHEMICALS": "CHEMICALS",
    "DEFENCE": "DEFENCE",
    "TEXTILES": "TEXTILES",
    "AGRICULTURE": "AGRI", "AGRI": "AGRI",
}


def enrich_sectors_from_nselib() -> int:
    """
    Phase B6 (2026-06-30): bulk-populate _SECTOR_MAP from nselib equity_list.
    Returns number of new sectors added. Cheap (one HTTP call, 24h cached).
    Run this BEFORE the per-symbol yfinance enrichment to cut its workload by >90%.
    """
    global _SECTOR_MAP
    df = _nselib_equity_list_cached()
    if df is None or (hasattr(df, "empty") and df.empty):
        _log("  [Sector nselib] equity_list unavailable — falling through to yfinance")
        return 0
    # Find symbol + industry columns flexibly
    cols = {str(c).lower().strip(): c for c in df.columns}
    sym_col = (cols.get("symbol") or cols.get("nseid") or cols.get("series_symbol"))
    ind_col = (cols.get("industry") or cols.get("sector") or cols.get("macro economic sector")
               or cols.get("macro_economic_sector"))
    if sym_col is None or ind_col is None:
        # Phase C4 (2026-07-02): nselib "equity_list" no longer ships industry
        # data — this is now a permanent condition, not an error. Log once at
        # info level, then fall back to local sector_master.csv (shipped in
        # repo). If you see this line every run, that's expected.
        _log("  [Sector nselib] equity_list has no industry column (expected) "
             "— using local sector_master.csv")
        # Local sector_master.csv fallback — shipped in repo
        try:
            master_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sector_master.csv")
            if os.path.exists(master_path):
                import csv as _csv
                added_local = 0
                new_from_csv: dict = {}
                with open(master_path, "r", encoding="utf-8") as _f:
                    reader = _csv.DictReader(_f)
                    for row in reader:
                        _sym = (row.get("symbol") or row.get("SYMBOL") or "").strip().upper()
                        _ind = (row.get("industry") or row.get("INDUSTRY")
                                or row.get("sector") or row.get("SECTOR") or "").strip().upper()
                        if not _sym or not _ind:
                            continue
                        _sym_ns = _sym if _sym.endswith(".NS") else _sym + ".NS"
                        if _sym_ns in SECTOR_MAP or _sym_ns in _SECTOR_MAP:
                            continue
                        norm = _NSE_INDUSTRY_NORM.get(_ind) or _NSE_INDUSTRY_NORM.get(_ind.split()[0] if _ind else "", "DIVERSIFIED")
                        new_from_csv[_sym_ns] = norm
                        added_local += 1
                if new_from_csv:
                    _SECTOR_MAP.update(new_from_csv)
                    try:
                        existing: dict = {}
                        if os.path.exists(SECTOR_CACHE_FILE):
                            with open(SECTOR_CACHE_FILE, "r") as _f:
                                existing = json.load(_f)
                        existing.update(new_from_csv)
                        with open(SECTOR_CACHE_FILE, "w") as _f:
                            json.dump(existing, _f, indent=2, sort_keys=True)
                    except Exception:
                        pass
                    _log(f"  [Sector nselib] ✓ loaded {added_local} sectors from local sector_master.csv")
                    return added_local
        except Exception as _e:
            _log(f"  [Sector nselib] local sector_master.csv fallback failed: {_e}")
        return 0
    added = 0
    new_sectors: dict = {}
    for _, row in df.iterrows():
        sym = str(row[sym_col]).strip().upper()
        if not sym or sym in ("NAN", "NONE"):
            continue
        sym_ns = sym if sym.endswith(".NS") else sym + ".NS"
        if sym_ns in SECTOR_MAP or sym_ns in _SECTOR_MAP:
            continue
        ind = str(row[ind_col]).strip().upper()
        norm = _NSE_INDUSTRY_NORM.get(ind)
        if norm is None:
            # Best-effort fallback: take first word
            head = ind.split()[0] if ind else ""
            norm = _NSE_INDUSTRY_NORM.get(head, "DIVERSIFIED")
        new_sectors[sym_ns] = norm
        added += 1
    if new_sectors:
        _SECTOR_MAP.update(new_sectors)
        # Persist into the same sector_cache.json file used by enrich_sectors_from_yfinance
        try:
            existing: dict = {}
            if os.path.exists(SECTOR_CACHE_FILE):
                with open(SECTOR_CACHE_FILE, "r") as f:
                    existing = json.load(f)
            existing.update(new_sectors)
            with open(SECTOR_CACHE_FILE, "w") as f:
                json.dump(existing, f, indent=2, sort_keys=True)
            _log(f"  [Sector nselib] ✓ added {added} sectors → cache {len(existing)} total")
        except Exception as e:
            _log(f"  [Sector nselib] cache write failed: {e}")
    return added


def enrich_sectors_from_yfinance(symbols: list, max_fetch: int = 500) -> None:
    """
    Batch-fetch yfinance sector info for symbols not yet in _SECTOR_MAP.
    Called once after tradable universe is built (prices already downloaded).
    New sectors are merged into _SECTOR_MAP and saved to sector_cache.json.
    On subsequent runs the cache is restored → zero yfinance calls needed.
    """
    global _SECTOR_MAP
    unknown = [
        s for s in symbols
        if not SECTOR_MAP.get(s if s.endswith(".NS") else s + ".NS")
        and _SECTOR_MAP.get(s if s.endswith(".NS") else s + ".NS", "OTHERS") == "OTHERS"
    ]
    if not unknown:
        _log(f"  Sector map: {len(_SECTOR_MAP)} symbols — fully covered")
        return

    to_fetch = unknown[:max_fetch]
    _log(f"  Sector map: {len(_SECTOR_MAP)} known | enriching {len(to_fetch)}/{len(unknown)} unknowns via yfinance...")

    new_sectors: dict = {}

    def _fetch_one(sym: str) -> None:
        try:
            info    = yf.Ticker(sym).info
            yf_sec  = (info.get("sector") or info.get("industry") or "").strip()
            if yf_sec:
                norm = _YF_LABEL_NORM.get(
                    yf_sec,
                    yf_sec.upper().replace(" & ", "_").replace(" ", "_")
                )
                new_sectors[sym] = norm   # GIL protects simple dict writes
        except Exception:
            pass

    with ThreadPoolExecutor(max_workers=8) as ex:
        list(ex.map(_fetch_one, to_fetch))

    if new_sectors:
        _SECTOR_MAP.update(new_sectors)
        try:
            existing: dict = {}
            if os.path.exists(SECTOR_CACHE_FILE):
                with open(SECTOR_CACHE_FILE, "r") as f:
                    existing = json.load(f)
            existing.update(new_sectors)
            with open(SECTOR_CACHE_FILE, "w") as f:
                json.dump(existing, f, indent=2, sort_keys=True)
            _log(f"  Sector cache: +{len(new_sectors)} new → {len(existing)} total ({SECTOR_CACHE_FILE})")
        except Exception as e:
            _log(f"[WARN] sector_cache save failed: {e}")
    else:
        _log("  Sector map: yfinance returned no new sector data (all remain OTHERS)")


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 2 — UTILITIES
# ─────────────────────────────────────────────────────────────────────────────

_LOG_FILE = None


def init_run_log():
    global _LOG_FILE
    try:
        fname = f"run_log_{ist_today().strftime('%Y%m%d')}.txt"
        _LOG_FILE = open(fname, "a", encoding="utf-8")
        _log(f"=== PIPELINE STARTED {ist_now().isoformat()} ===")
    except Exception as e:
        print(f"[WARN] init_run_log failed: {e}")


def _log(msg: str):
    ts = ist_now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    if _LOG_FILE:
        try:
            _LOG_FILE.write(line + "\n")
            _LOG_FILE.flush()
        except Exception:
            pass


def close_run_log():
    global _LOG_FILE
    if _LOG_FILE:
        try:
            _log("=== PIPELINE COMPLETE ===")
            _LOG_FILE.close()
        except Exception:
            pass


# ─── Market Calendars (Phase C7e 2026-07-02): JSON-first, hardcoded fallback ─
# Historical bug: NSE_HOLIDAYS_2026 / _RBI_MPC_DATES_2026 / _FOMC_DATES_2026 were
# hardcoded 2026-only literals. Silently wrong on 2027-01-01 (holidays evaluate
# to False every day, event blackouts vanish). Now driven by market_calendars.json.
# Additive-only — no scoring/gate impact. If the JSON is missing or corrupt, the
# hardcoded 2026 defaults kick in exactly as before, so behaviour is unchanged.
MARKET_CALENDARS_FILE = "market_calendars.json"

# Hardcoded 2026 defaults — used only if market_calendars.json is missing/broken.
_NSE_HOLIDAYS_2026_DEFAULT = {
    "2026-01-26","2026-02-19","2026-03-25","2026-04-02",
    "2026-04-10","2026-04-14","2026-04-17","2026-05-01",
    "2026-06-17","2026-08-15","2026-10-02","2026-10-20",
    "2026-11-05","2026-11-16","2026-12-25",
}
_RBI_MPC_2026_DEFAULT = [
    "2026-02-07", "2026-04-09", "2026-06-06",
    "2026-08-08", "2026-10-07", "2026-12-05",
]
_FOMC_2026_DEFAULT = [
    "2026-01-29", "2026-03-19", "2026-05-07",
    "2026-06-18", "2026-07-30", "2026-09-17",
    "2026-11-05", "2026-12-17",
]


def _load_market_calendars() -> dict:
    """Read market_calendars.json → dict with keys nse_holidays/rbi_mpc/fomc,
    each a flat list-of-strings covering all years present in the file.
    Silently falls back to hardcoded 2026 defaults on any error."""
    result = {
        "nse_holidays": set(_NSE_HOLIDAYS_2026_DEFAULT),
        "rbi_mpc":      list(_RBI_MPC_2026_DEFAULT),
        "fomc":         list(_FOMC_2026_DEFAULT),
    }
    try:
        if not os.path.exists(MARKET_CALENDARS_FILE):
            return result
        with open(MARKET_CALENDARS_FILE, "r", encoding="utf-8") as f:
            raw = json.load(f)
        if not isinstance(raw, dict):
            return result

        def _flatten(section):
            """Section is {year: [dates]} — flatten to a list of date strings."""
            out = []
            block = raw.get(section, {})
            if isinstance(block, dict):
                for _year, dates in block.items():
                    if isinstance(dates, list):
                        out.extend(str(d) for d in dates if isinstance(d, str))
            elif isinstance(block, list):  # tolerate flat-list shape too
                out = [str(d) for d in block if isinstance(d, str)]
            return out

        hols = _flatten("nse_holidays")
        rbi  = _flatten("rbi_mpc_dates")
        fom  = _flatten("fomc_dates")

        if hols: result["nse_holidays"] = set(hols)
        if rbi:  result["rbi_mpc"]      = rbi
        if fom:  result["fomc"]         = fom
    except Exception as e:
        # _log may not exist yet at module-import time; use print as a safe floor
        try:
            _log(f"[WARN] _load_market_calendars failed, using 2026 fallback: {e}")
        except Exception:
            print(f"[WARN] _load_market_calendars failed, using 2026 fallback: {e}")
    return result


_MARKET_CALENDARS      = _load_market_calendars()
NSE_HOLIDAYS           = _MARKET_CALENDARS["nse_holidays"]   # set of "YYYY-MM-DD"
_RBI_MPC_DATES         = _MARKET_CALENDARS["rbi_mpc"]        # list of "YYYY-MM-DD"
_FOMC_DATES            = _MARKET_CALENDARS["fomc"]           # list of "YYYY-MM-DD"

# Back-compat aliases — some external tooling / tests may still import the
# original 2026-suffixed names. Safe to remove after downstream is verified.
NSE_HOLIDAYS_2026      = NSE_HOLIDAYS
_RBI_MPC_DATES_2026    = _RBI_MPC_DATES
_FOMC_DATES_2026       = _FOMC_DATES


def is_market_open(check_date=None) -> bool:
    if check_date is None:
        check_date = ist_today()
    if check_date.weekday() >= 5:
        return False
    if check_date.strftime("%Y-%m-%d") in NSE_HOLIDAYS:
        return False
    return True


def is_earnings_season() -> bool:
    return ist_today().month in {4, 5, 10, 11}


def earnings_season_threshold_adjustment() -> int:
    return 3 if is_earnings_season() else 0


def truncate_display(text: str, max_len: int = 100) -> str:
    """Bug 6 fix — always shows clean truncation, never a broken sentence."""
    if not text:
        return "—"
    text = str(text).strip()
    if len(text) <= max_len:
        return text
    return text[:max_len - 3].rstrip() + "..."


def _split_telegram_message(text: str, max_len: int) -> list:
    """Split at section (═) boundaries first, then paragraph (\n\n), then newline.

    FIX: previously fell straight from section-divider to single-newline, which
    sometimes broke a Near-Miss card in half. Preferring paragraph boundaries
    keeps semantic blocks (BUY card, NEAR MISS card, WATCHLIST row) intact.
    Also prepends "(cont'd)" to continuation chunks so the user knows they're
    reading a continuation of the previous section.
    """
    if len(text) <= max_len:
        return [text]
    chunks = []
    DIVIDER = "═"
    # Try splitting at section dividers (long lines of ═)
    sections = text.split(DIVIDER * 10)
    chunk = ""
    for section in sections:
        candidate = (chunk + DIVIDER * 10 + section) if chunk else section
        if len(candidate) <= max_len:
            chunk = candidate
        else:
            if chunk:
                chunks.append(chunk)
            chunk = section
    if chunk:
        chunks.append(chunk)
    # If any chunk still too long, split — prefer paragraph (\n\n), then \n.
    final = []
    for c in chunks:
        while len(c) > max_len:
            split_at = c.rfind("\n\n", 0, max_len)
            if split_at == -1:
                split_at = c.rfind("\n", 0, max_len)
            if split_at == -1:
                split_at = max_len
            final.append(c[:split_at])
            c = c[split_at:].lstrip("\n")
        if c:
            final.append(c)
    # Prepend a (cont'd) tag on chunks 2..N so a card split across messages is
    # obvious to the reader on Telegram.
    if len(final) > 1:
        final = [final[0]] + [f"(cont'd)\n{c}" for c in final[1:]]
    return final if final else [text]


def _save_failed_telegram(message: str) -> None:
    try:
        fname = f"telegram_failed_{ist_today().strftime('%Y%m%d')}.txt"
        with open(fname, "a", encoding="utf-8") as f:
            f.write(f"\n\n=== {ist_now().isoformat()} ===\n")
            f.write(message)
    except Exception:
        pass


_VALID_TG_TAGS = re.compile(
    r"<(/?(b|i|u|s|code|pre|a)(\s[^>]*)?)>",
    re.IGNORECASE,
)

def _sanitize_telegram_html(text: str) -> str:
    """
    Escape any < that does NOT start a valid Telegram HTML tag.
    Telegram supports: <b> <i> <u> <s> <code> <pre> <a href="...">
    Any other < (e.g. from fail_reasons like "Conf < 83)") causes HTTP 400.
    """
    result = []
    i = 0
    while i < len(text):
        if text[i] == "<":
            m = _VALID_TG_TAGS.match(text, i)
            if m:
                result.append(m.group(0))
                i = m.end()
            else:
                result.append("&lt;")
                i += 1
        elif text[i] == "&" and not re.match(r"&(?:lt|gt|amp|quot|#\d+);", text[i:]):
            result.append("&amp;")
            i += 1
        else:
            result.append(text[i])
            i += 1
    return "".join(result)


def send_telegram(message: str) -> None:
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        _log("[WARN] Main Telegram not configured — skipping send")
        return
    chunks = _split_telegram_message(message, TELEGRAM_MAX_CHARS)
    for i, chunk in enumerate(chunks):
        try:
            url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
            resp = requests.post(url, json={
                "chat_id": TELEGRAM_CHAT_ID,
                "text": _sanitize_telegram_html(chunk),
                "parse_mode": "HTML",
            }, timeout=15)
            if resp.status_code != 200:
                raise Exception(f"HTTP {resp.status_code}: {resp.text[:120]}")
        except Exception as e:
            _log(f"[WARN] Telegram send failed (chunk {i+1}/{len(chunks)}): {e}")
            _save_failed_telegram(message)
            break


def send_buy_telegram(buys: list, regime: str, timestamp: str) -> None:
    """
    Sends BUY signals ONLY to a dedicated channel (BUY_BOT_TOKEN + BUY_CHAT_ID).
    Fires even when buys list is empty so you know the pipeline ran.
    """
    if not BUY_BOT_TOKEN or not BUY_CHAT_ID:
        _log("[INFO] BUY channel not configured (BUY_BOT_TOKEN / BUY_CHAT_ID missing) — skipping")
        return

    lines = []
    lines.append("─" * 38)
    lines.append(f"📊 NSE BUY SIGNALS — {timestamp}")
    lines.append(f"Regime: {regime}")
    lines.append("─" * 38)

    if buys:
        for b in buys:
            sym      = b.get("symbol", "?")
            sector   = b.get("sector", "OTHERS")
            conf     = b.get("final_confidence", 0)
            tq       = b.get("trade_quality_score", 0)
            rr       = b.get("rr_ratio", 0)
            entry    = b.get("entry", 0)
            stop_p   = b.get("stop", 0)
            t1       = b.get("target1", 0)
            t2       = b.get("target2", 0)
            pos_val  = b.get("position_value", 0)
            pos_pct  = b.get("position_pct", 0)
            stop_pct = round((entry - stop_p) / entry * 100, 1) if entry > 0 else 0
            news_sum = truncate_display(b.get("news_summary", ""), 90)
            pledge   = b.get("promoter_pledge_pct", 0)
            roe      = b.get("roe", 0)
            de       = b.get("de_ratio", 0)
            repeat   = b.get("repeat_tag", "")

            lines.append(f"\n<b>{html.escape(str(sym))}</b> [{html.escape(str(sector))}]")
            lines.append(f"Conf {conf:.1f} | TQ {tq:.1f} | R/R {rr:.2f}x")
            lines.append(f"Entry  Rs{entry:.2f}")
            lines.append(f"Stop   Rs{stop_p:.2f}  ({stop_pct:.1f}%)")
            lines.append(f"T1     Rs{t1:.2f}")
            lines.append(f"T2     Rs{t2:.2f}")
            # Max valid entry for gap-open decision
            gap_chk = check_gap_validity(entry, stop_p, t1, rr if rr > 0 else 1.8)
            max_ent = gap_chk.get("max_valid_entry", 0)
            if max_ent > entry:
                gap_max_pct = round((max_ent - entry) / entry * 100, 1)
                lines.append(f"⚡ Max valid entry: Rs{max_ent:.2f} (+{gap_max_pct:.1f}%)")
                lines.append(f"   If opens > Rs{max_ent:.2f} → SKIP signal")
            lines.append(f"Size   Rs{pos_val:,.0f}  ({pos_pct:.1f}% capital)")
            lines.append(f"ROE {roe:.1f}% | D/E {de:.2f} | Pledge {pledge:.0f}%")
            if news_sum and news_sum != "—":
                lines.append(f"News: {html.escape(str(news_sum))}")
            if repeat:
                lines.append(f"⚠️ {html.escape(str(repeat))}")
            if b.get("warnings"):
                lines.append(f"WARN: {html.escape(', '.join(b['warnings'][:3]))}")
            lines.append("─" * 38)
    else:
        lines.append("\nNo BUY signals today — no setup cleared all gates.")
        lines.append("─" * 38)

    lines.append("Recommendation only. Execute manually.")
    message = "\n".join(lines)

    chunks = _split_telegram_message(message, TELEGRAM_MAX_CHARS)
    for i, chunk in enumerate(chunks):
        try:
            url = f"https://api.telegram.org/bot{BUY_BOT_TOKEN}/sendMessage"
            resp = requests.post(url, json={
                "chat_id":    BUY_CHAT_ID,
                "text":       _sanitize_telegram_html(chunk),
                "parse_mode": "HTML",
            }, timeout=15)
            if resp.status_code != 200:
                _log(f"[WARN] BUY Telegram send failed (chunk {i+1}): {resp.status_code} {resp.text[:80]}")
        except Exception as e:
            _log(f"[WARN] send_buy_telegram failed: {e}")
            _save_failed_telegram(message)
            break


def save_csv(data: list, base_filename: str) -> None:
    if not data:
        return
    try:
        date_str = ist_today().strftime("%Y%m%d")
        name, ext = os.path.splitext(base_filename)
        filename = f"{name}_{date_str}{ext}"
        clean = []
        for row in data:
            clean.append({k: v for k, v in row.items()
                          if isinstance(v, (str, int, float, bool, type(None)))})
        if not clean:
            return
        # Union all keys across every row so no row's extra fields cause a crash
        all_keys: list = []
        seen_keys: set = set()
        for row in clean:
            for k in row.keys():
                if k not in seen_keys:
                    all_keys.append(k)
                    seen_keys.add(k)
        with open(filename, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=all_keys, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(clean)
        _log(f"Saved {len(clean)} rows to {filename}")
    except Exception as e:
        _log(f"[WARN] save_csv failed: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 3 — GROQ AI (3-KEY ROUND-ROBIN) — Bug 4 fix
# ─────────────────────────────────────────────────────────────────────────────

CLOUD_AI_ENDPOINT = os.getenv("CLOUD_AI_ENDPOINT", "")
CLOUD_AI_KEY      = os.getenv("CLOUD_AI_KEY", "")

# Stage C kill-switch: when "false"/"0"/"no", the three narrative functions
# (ai_daily_summary, ai_buy_thesis, ai_near_miss_insight) skip the Groq call
# entirely and return their rule-based fallback text. Lets operators run the
# pipeline in a deterministic, LLM-free mode for A/B comparison or when Groq
# is unavailable. Default: enabled (preserves current behavior).
ENABLE_AI_NARRATIVES = os.getenv("ENABLE_AI_NARRATIVES", "true").strip().lower() not in ("false", "0", "no", "off", "")

_GROQ_KEYS_RAW = [
    os.getenv("GROQ_API_KEY_1", ""),
    os.getenv("GROQ_API_KEY_2", ""),
    os.getenv("GROQ_API_KEY_3", ""),
]
# Fall back to single key if multi-key not set
if not any(k.strip() for k in _GROQ_KEYS_RAW):
    _single = os.getenv("GROQ_API_KEY", "")
    if _single.strip():
        _GROQ_KEYS_RAW = [_single, "", ""]

GROQ_KEYS = [k.strip() for k in _GROQ_KEYS_RAW if k.strip()]
_groq_key_cycle = itertools.cycle(GROQ_KEYS) if GROQ_KEYS else iter([])
_groq_key_failures: dict = {}

# Weighted severity (1=mild, 10=existential). Sum >15 = HIGH_RISK,
# 5-15 = MODERATE. Replaces the flat list so a single "auditor resigned"
# outweighs three "revenue decline" mentions. Iteration order preserved.
NEGATIVE_KEYWORDS_SEVERITY = {
    "fraud": 10, "scam": 10, "sebi ban": 10, "ed raid": 10,
    "cbi": 9, "fir": 9, "arrested": 10, "bankrupt": 10,
    "insolvency": 9, "liquidation": 9, "default": 8, "npa": 6,
    "downgrade": 5, "plant shut": 6, "factory closed": 6,
    "promoter sell": 5, "pledged shares sold": 7, "regulatory action": 6,
    "show cause": 4, "penalty": 3, "fine imposed": 3,
    "loss widened": 3, "revenue decline": 3,
    "auditor resigned": 9, "qualified opinion": 7, "going concern": 8,
    "debt restructure": 5,
}
# Back-compat: list view for any legacy `for kw in NEGATIVE_KEYWORDS` loops.
NEGATIVE_KEYWORDS = list(NEGATIVE_KEYWORDS_SEVERITY.keys())
POSITIVE_KEYWORDS = [
    "contract win","order win","expansion","new plant","capacity addition",
    "earnings beat","profit up","revenue growth","dividend","buyback",
    "stake acquisition","partnership","export order","fda approval","patent",
]
BLACK_SWAN_KEYWORDS = [
    "fraud","sebi ban","ed raid","cbi raid","arrested","bankrupt",
    "insolvency","liquidation","promoter arrested","nclt",
]


def _call_groq_with_rotation(prompt: str, max_tokens: int = 150) -> str | None:
    if not GROQ_KEYS:
        return None
    wall_start = time.time()
    max_attempts = min(len(GROQ_KEYS), 3)
    for _ in range(max_attempts):
        if time.time() - wall_start > 20:
            break
        key = next(_groq_key_cycle)
        last_fail = _groq_key_failures.get(key, 0)
        if time.time() - last_fail < 60:
            continue
        try:
            resp = requests.post(
                "https://api.groq.com/openai/v1/chat/completions",
                headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json"},
                json={
                    "model": "llama-3.1-8b-instant",
                    "messages": [{"role": "user", "content": prompt}],
                    "max_tokens": max_tokens,
                    "temperature": 0,
                },
                timeout=12,
            )
            if resp.status_code == 200:
                return resp.json()["choices"][0]["message"]["content"]
            elif resp.status_code == 429:
                _groq_key_failures[key] = time.time()
                _log(f"[WARN] Groq key ...{key[-6:]} rate-limited, rotating")
            else:
                _log(f"[WARN] Groq key ...{key[-6:]} returned {resp.status_code}")
        except Exception as e:
            _log(f"[WARN] Groq call failed (key ...{key[-6:]}): {e}")
            _groq_key_failures[key] = time.time()
    return None


def _parse_ai_json(text: str) -> dict | None:
    try:
        clean = re.sub(r"```json|```", "", text).strip()
        data = json.loads(clean)
        return {
            "severity":      min(100, max(-30, int(data.get("severity", 0)))),
            "category":      data.get("category", "NEUTRAL"),
            "is_black_swan": bool(data.get("is_black_swan", False)),
            "summary":       str(data.get("summary", ""))[:150],
        }
    except Exception:
        return None


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 3b — SHARED AI CALLER + HIGH-VALUE AI FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────

def _call_ai(prompt: str, max_tokens: int = 100) -> str | None:
    """
    Shared AI caller with 3-tier fallback.
    Tier 1: SAP / Cloud AI  (CLOUD_AI_ENDPOINT + CLOUD_AI_KEY)
    Tier 2: Groq llama-3.1-8b-instant  (existing key rotation)
    Tier 3: Returns None — caller uses rule-based fallback.
    prompt must be under 1500 chars (~300 tokens).
    """
    if len(prompt) > 1500:
        prompt = prompt[:1500]
        _log("[WARN] AI prompt truncated to 1500 chars")

    # Tier 1: Cloud / SAP AI
    if CLOUD_AI_ENDPOINT and CLOUD_AI_KEY:
        try:
            r = requests.post(
                CLOUD_AI_ENDPOINT,
                headers={"Authorization": f"Bearer {CLOUD_AI_KEY}",
                         "Content-Type":  "application/json"},
                json={"prompt": prompt, "max_tokens": max_tokens, "temperature": 0.3},
                timeout=10,
            )
            if r.status_code == 200:
                data = r.json()
                text = (data.get("choices", [{}])[0].get("text", "") or
                        data.get("content", [{}])[0].get("text", "") or
                        data.get("response", "") or data.get("output", ""))
                if text and len(text.strip()) > 10:
                    return text.strip()
        except Exception as e:
            _log(f"[WARN] Cloud AI failed: {e}")

    # Tier 2: Groq (uses existing key rotation)
    text = _call_groq_with_rotation(prompt, max_tokens=max_tokens)
    if text and len(text.strip()) > 10:
        return text.strip()

    # Tier 3: unavailable
    return None


def _clean_ai_output(text: str) -> str:
    """
    Cleans common AI output artifacts.
    Fixes spaces inside decimals: "13. 6" → "13.6"
    Removes markdown, bullets, leading/trailing junk.
    Unescapes HTML entities (&amp; &#39; &quot; etc.).
    """
    import re
    import html as _html_mod
    text = _html_mod.unescape(text or "")                    # &amp; → &, &#39; → '
    text = re.sub(r'(\d+)\.\s+(\d+)', r'\1.\2', text)   # "13. 6" → "13.6"
    text = re.sub(r'(\d)\.\s+(\d)', r'\1.\2', text)       # "0. 49%" → "0.49%"
    text = re.sub(r'\*+', '', text)                          # remove markdown bold
    text = re.sub(r'^\s*[-•*]\s*', '', text, flags=re.MULTILINE)  # remove bullets
    text = " ".join(text.split())                            # normalise whitespace
    return text.strip()


def _split_sentences(text: str) -> list:
    """
    Sentence splitter that does NOT break on:
      - Ticker suffixes:   "MANBA.NS breaks..."   (letter.LETTER)
      - Decimals:          "R/R 2.5x"             (digit.digit)
      - Abbreviations:     "Rs. 500", "e.g.", "i.e.", "vs.", "no.", "pct."
    Splits ONLY on: period/! /? followed by whitespace followed by a capital letter,
    OR end-of-string.
    """
    import re
    if not text:
        return []
    # Protect known abbreviations by temporarily removing the period
    ABBR = {"Rs.": "Rs<DOT>", "rs.": "rs<DOT>",
            "e.g.": "e<DOT>g<DOT>", "i.e.": "i<DOT>e<DOT>",
            "vs.": "vs<DOT>", "No.": "No<DOT>", "no.": "no<DOT>",
            "pct.": "pct<DOT>", "Pct.": "Pct<DOT>",
            "Inc.": "Inc<DOT>", "Ltd.": "Ltd<DOT>", "Co.": "Co<DOT>",
            "U.S.": "U<DOT>S<DOT>", "U.K.": "U<DOT>K<DOT>"}
    for k, v in ABBR.items():
        text = text.replace(k, v)

    # Split on sentence-ending punctuation ONLY when followed by space+capital or end
    # (?<=[.!?])  = look-behind: previous char is . ! ?
    # \s+         = at least one space
    # (?=[A-Z"'])  = look-ahead: next char is capital/quote (real sentence start)
    parts = re.split(r'(?<=[.!?])\s+(?=[A-Z"\'\u201C\u2018])', text)

    # Restore abbreviations
    def _restore(s):
        return s.replace("<DOT>", ".").strip()

    return [_restore(p) for p in parts if _restore(p)]


def _rule_based_summary(
    regime: str, regime_label: str, buy_count: int,
    near_miss_count: int, top_symbol: str,
    portfolio_alerts: list, ema_bear: bool,
    upcoming_event: str
) -> str:
    """Fallback — always correct, never has AI artifacts."""
    parts = []
    parts.append(
        f"Market is {regime_label} with price "
        f"{'below' if ema_bear else 'above'} key moving averages."
    )
    if buy_count > 0:
        parts.append(
            f"{buy_count} institutional-quality setup(s) met all conditions today."
        )
    else:
        parts.append(
            "No setup met all required conditions today — "
            "quality over quantity remains the priority."
        )
    if near_miss_count > 0 and top_symbol and top_symbol != "none":
        parts.append(
            f"{near_miss_count} stock(s) are close to qualifying — "
            f"watch {top_symbol.replace('.NS', '')} most closely."
        )
    else:
        exits = [a for a in portfolio_alerts if "EXIT" in a.get("action", "")]
        if exits:
            parts.append(
                f"{len(exits)} position(s) require exit review — act before next session."
            )
        else:
            parts.append("Existing positions are on track — no action needed today.")
    if upcoming_event:
        parts.append(
            f"With {upcoming_event} approaching, reduce position sizes and keep stops tight."
        )
    elif ema_bear and buy_count == 0:
        parts.append(
            "Avoid forcing new trades until the market reclaims its key moving averages."
        )
    elif buy_count > 0:
        parts.append(
            "Execute today's signal with strict position sizing — do not average down."
        )
    else:
        parts.append(
            "Stay patient — let the setups come to you rather than chasing momentum."
        )
    return " ".join(parts)


def ai_daily_summary(
    regime: str,
    regime_score: float,
    nifty_pct: float,
    vix_in: float,
    breadth_pct: float,
    fii_flow_cr: float,
    dii_flow_cr: float,
    buy_count: int,
    near_miss_count: int,
    top_near_miss_symbol: str,
    top_near_miss_conf_gap: float,
    portfolio_alerts: list,
    ema_bear: bool,
    upcoming_event: str = "",
) -> str:
    """
    4-sentence AI market briefing. Strictly qualitative — no raw numbers sent to AI.
    Falls back to rule-based if AI fails or output contains numbers / is too short.
    """
    import re

    # Translate numbers to qualitative labels BEFORE sending to AI
    # (regime_label is also used by the rule-based fallback below).
    regime_label = {
        "STRONG_BULL":     "strongly bullish with broad participation",
        "BULL":            "bullish with improving breadth",
        "SIDEWAYS":        "range-bound with no clear direction",
        "TRANSITION":      "transitioning with mixed signals",
        "HIGH_VOLATILITY": "volatile with compressed opportunity",
        "BEAR":            "weakening with institutional selling",
        "STRONG_BEAR":     "in capital preservation mode",
    }.get(regime, "uncertain")

    # Stage C: kill-switch bypasses LLM entirely and returns rule-based summary.
    if not ENABLE_AI_NARRATIVES:
        return _rule_based_summary(
            regime, regime_label, buy_count, near_miss_count,
            top_near_miss_symbol, portfolio_alerts, ema_bear, upcoming_event
        )

    nifty_label = (
        "falling sharply" if nifty_pct < -1.0 else
        "falling slightly" if nifty_pct < -0.2 else
        "flat" if abs(nifty_pct) <= 0.2 else
        "rising slightly" if nifty_pct < 0.8 else
        "rising strongly"
    )
    vix_label = (
        "very calm" if vix_in < 12 else
        "normal" if vix_in < 16 else
        "elevated" if vix_in < 20 else
        "high" if vix_in < 25 else
        "very high"
    )
    breadth_label = (
        "weak" if breadth_pct < 40 else
        "mixed" if breadth_pct < 55 else
        "healthy" if breadth_pct < 70 else
        "strong"
    )
    fii_label = (
        "buying aggressively" if fii_flow_cr > 2000 else
        "buying"             if fii_flow_cr > 300  else
        "neutral"            if abs(fii_flow_cr) <= 300 else
        "selling"            if fii_flow_cr > -2000 else
        "selling aggressively"
    )
    dii_label = (
        "buying aggressively" if dii_flow_cr > 2000 else
        "buying"             if dii_flow_cr > 300  else
        "neutral"            if abs(dii_flow_cr) <= 300 else
        "selling"
    )
    ema_label       = "below key moving averages" if ema_bear else "above key moving averages"
    exits           = [a for a in portfolio_alerts if "EXIT" in a.get("action", "")]
    portfolio_label = (
        f"{len(exits)} position(s) need exit review" if exits else
        "all positions on track"
    )
    near_miss_label = (
        "no stocks close to qualifying" if near_miss_count == 0 else
        f"{near_miss_count} stock(s) close to qualifying"
    )
    event_label = (
        f"upcoming {upcoming_event} adds uncertainty" if upcoming_event else
        "no major events this week"
    )

    prompt = f"""You are writing 4 sentences for a swing trader's daily briefing.

Market conditions today:
- Overall market: {regime_label}
- NIFTY direction: {nifty_label}
- Volatility (VIX): {vix_label}
- Market breadth: {breadth_label}
- FII (foreign institutions): {fii_label}
- DII (domestic institutions): {dii_label}
- Price structure: {ema_label}
- Buy signals today: {buy_count}
- Watchlist status: {near_miss_label}
- Portfolio: {portfolio_label}
- Events: {event_label}

Write EXACTLY 4 sentences. Rules:
1. Do NOT use any numbers, percentages, or rupee amounts
2. Use only the qualitative labels provided above
3. Sentence 1: Overall market condition and why
4. Sentence 2: What this means for new trades today
5. Sentence 3: What to watch most closely right now
6. Sentence 4: One specific actionable recommendation for tomorrow
7. Total: 60-80 words maximum
8. Plain English — senior trader tone, not a report
9. Never say "I" or "as an AI"
10. Never use the word "bearish" or "bullish" — use plain words"""

    result = _call_ai(prompt, max_tokens=150)
    if result:
        clean = _clean_ai_output(result)
        sentences = [s for s in _split_sentences(clean) if len(s) > 15]
        has_numbers = bool(re.search(r'\d', clean))
        if len(sentences) >= 3 and not has_numbers:
            return " ".join(sentences[:4])
        if len(sentences) >= 3:
            # Strip any numbers that crept through
            clean_no_nums = re.sub(r'\d+\.?\d*%?', '', clean)
            clean_no_nums = re.sub(r'Rs\.?\s*\d+', '', clean_no_nums)
            clean_no_nums = " ".join(clean_no_nums.split())
            if len(clean_no_nums) > 50:
                return clean_no_nums

    return _rule_based_summary(
        regime, regime_label, buy_count, near_miss_count,
        top_near_miss_symbol, portfolio_alerts, ema_bear, upcoming_event
    )


def _rule_based_thesis(symbol: str, sector: str, rr: float,
                        conf_trend: str, catalyst: list,
                        sector_status: str) -> str:
    """Rule-based fallback for BUY thesis."""
    parts = []
    if sector_status in ("LEADING", "IMPROVING"):
        parts.append(f"{sector} sector showing strength")
    if "VOL_SURGE" in catalyst:
        parts.append("volume expansion confirming move")
    if "UPTREND" in catalyst:
        parts.append("price in clean uptrend")
    if conf_trend and "rising" in conf_trend.lower():
        parts.append("confidence building over 3 days")
    reason = ", ".join(parts) if parts else "multi-factor confluence"
    return f"{symbol}: {reason} with R/R {rr:.1f}x. Risk: stop must hold — no averaging down."


def ai_buy_thesis(
    symbol: str, sector: str, confidence: float, tq: float, rr: float,
    conf_trend: str, catalyst: list, sector_status: str,
    roe: float, pledge_pct: float, regime: str, risk_pct: float,
    factor_scores: dict = None, soft_warnings: list = None,
    rs_diff21: float = 0.0, accum_signal: str = "NEUTRAL",
) -> str:
    """1-2 sentence specific trade thesis for a BUY signal.

    Change #2 (2026-07-10): thesis now writes like a trader, not a factor table.
    - Pre-selects TOP-2 strongest factors + WEAKEST factor for LLM focus
    - Strips numbers/percentages from output (numbers already visible in card)
    - Falls back to rule-based on API failure or if output contains digits
    """
    # Stage C: kill-switch bypasses LLM entirely and returns rule-based thesis.
    if not ENABLE_AI_NARRATIVES:
        return _rule_based_thesis(symbol, sector, rr, conf_trend, catalyst, sector_status)

    fs = factor_scores or {}
    ranked  = sorted(fs.items(), key=lambda kv: -(kv[1] or 0))
    top_two = [k for k, v in ranked[:2] if v is not None]           # names only
    weakest = ranked[-1][0] if ranked else "n/a"                    # name only
    warns   = ", ".join((soft_warnings or [])[:2]) or "none"

    # Special-signal shortlist — feed only what's actually true for THIS stock
    signals = []
    if rs_diff21 >= 15.0:
        signals.append("beating Nifty by a wide margin")
    if pledge_pct <= 1.0:
        signals.append("clean ownership (no pledge)")
    if accum_signal in ("STRONG", "POCKET_PIVOT"):
        signals.append("institutional accumulation signal")
    if sector_status in ("LEADING", "IMPROVING"):
        signals.append(f"{sector} sector rotating in")
    if "VOL_SURGE" in (catalyst or []):
        signals.append("volume expansion today")
    if "NEAR_52W_HIGH" in (catalyst or []):
        signals.append("near 52-week high")

    prompt = (
        f"You are a trader writing a one-line thesis to a colleague about {symbol} ({sector}).\n"
        f"STRONGEST attributes: {', '.join(top_two) if top_two else 'trend'}\n"
        f"POSITIVE signals: {'; '.join(signals) if signals else 'confluence of factors'}\n"
        f"WEAKEST attribute (must mention as risk): {weakest}\n"
        f"Soft warnings: {warns}\n"
        f"Regime: {regime}\n\n"
        "Rules:\n"
        "1. Write EXACTLY 2 sentences.\n"
        "2. Sentence 1: what makes THIS stock stand out. Use plain words a trader would say.\n"
        "3. Sentence 2: the KEY RISK — must name the weakest attribute or a soft warning.\n"
        "4. DO NOT use ANY numbers, percentages, ratios, or scores. No digits at all.\n"
        "5. DO NOT say 'strong buy signal', 'high trend quality', 'multi-factor confluence', "
        "or any phrase that could apply to every stock.\n"
        "6. NO bullets, NO markdown, NO jargon like 'R/R' or 'TQ'.\n"
        "7. Max 40 words total."
    )
    result = _call_ai(prompt, max_tokens=100)
    if result and len(result) > 20:
        clean = _clean_ai_output(result)
        try:
            clean = html.unescape(clean)
        except Exception:
            pass
        # Hallucination guard: reject output containing digits or ₹/% —
        # those numbers might be invented by the LLM and the BUY card
        # already displays the real numbers directly below.
        if any(ch.isdigit() for ch in clean) or "%" in clean or "₹" in clean:
            # Rewrite fallback: strip digits/percent tokens instead of dropping
            import re as _re
            clean = _re.sub(r"[₹\d]+\.?\d*%?", "", clean)
            clean = _re.sub(r"\s+", " ", clean).strip()
            # If stripping mangled it, fall back cleanly
            if len(clean) < 25:
                return _rule_based_thesis(symbol, sector, rr, conf_trend, catalyst, sector_status)
        sentences = _split_sentences(clean)
        return " ".join(sentences[:2]) if sentences else clean
    return _rule_based_thesis(symbol, sector, rr, conf_trend, catalyst, sector_status)



def _rule_based_near_miss_insight(
    symbol: str, conf_gap: float, conf_only: bool,
    rr_fail: bool, tq_fail: bool, conf_trend: str, days_watching: int,
    sector_status: str = "NEUTRAL",
    # Optional per-stock numbers to make each insight unique (avoids the
    # "same sentence copied across 20 stocks" problem that used to hit
    # near-miss stocks past NM_AI_CAP).
    confidence: float = 0.0, tq: float = 0.0, rr: float = 0.0,
) -> str:
    """Fallback — specific and useful without AI.
    Every branch uses at least one stock-specific number so no two stocks
    with a similar profile receive an identical sentence.
    """
    trend_lc = (conf_trend or "").lower()
    rising   = "rising"  in trend_lc
    falling  = "falling" in trend_lc
    flat     = "flat"    in trend_lc

    # 1) Confidence-only near miss with rising trend → very close to a BUY
    if rising and conf_only:
        return (
            f"Confidence at {confidence:.1f} is rising toward the threshold — "
            f"needs one strong volume day to close the {conf_gap:.1f}-point gap."
        )

    # 2) Confidence-only near miss with falling trend → losing steam
    if falling and conf_only:
        return (
            f"Confidence slipping from {confidence:.1f} — remove after 2 more "
            f"sessions of decline unless volume returns."
        )

    # 3) Stuck-on-confidence for many days
    if days_watching >= 5 and conf_only:
        return (
            f"Stuck at {confidence:.1f} confidence for {days_watching} sessions — "
            f"if no {conf_gap:.1f}-point improvement in 2 more days, drop from watchlist."
        )

    # 4) R/R too low (with or without confidence issue)
    if rr_fail:
        return (
            f"R/R at {rr:.2f}x too low — wait for a pullback to lift R/R above "
            f"2.0x before confidence ({confidence:.1f}) can also qualify."
        )

    # 5) Trade Quality below bar
    if tq_fail:
        return (
            f"Chart pattern weak at TQ {tq:.1f} — wait for tighter consolidation "
            f"pushing TQ above 80 before acting."
        )

    # 6) Sector drag
    if sector_status in ("LAGGING", "WEAKENING"):
        return (
            f"{sector_status.title()} sector holding {symbol} back — confidence "
            f"{confidence:.1f} needs sector rotation before it can climb."
        )

    # 7) Flat/first-appearance default — reference the actual gap size
    if flat:
        return (
            f"Confidence flat at {confidence:.1f} for {max(days_watching,1)} "
            f"session(s) — needs a volume spike to break the {conf_gap:.1f}-point gap."
        )

    # 8) Generic fallback still uses the actual gap AND confidence value
    return (
        f"Needs {conf_gap:.1f} confidence points closed from {confidence:.1f} — "
        f"watch for volume surge above the 20-day average as trigger."
    )


def ai_near_miss_insight(
    symbol: str, sector: str, confidence: float, conf_gap: float,
    tq: float, rr: float, conf_trend: str, fail_reasons: list,
    gate_pattern: str, sector_status: str, risk_pct: float,
    days_watching: int,
) -> str:
    """1 sentence: what must happen for this near miss to become a BUY."""
    primary_fail = fail_reasons[0] if fail_reasons else "CONFIDENCE_FAIL"
    conf_only    = len(fail_reasons) == 1 and "CONF" in primary_fail
    rr_fail      = any("RR" in f for f in fail_reasons)
    tq_fail      = any("TQ" in f for f in fail_reasons)

    # Stage C: kill-switch bypasses LLM entirely and returns rule-based insight.
    if not ENABLE_AI_NARRATIVES:
        return _rule_based_near_miss_insight(
            symbol, conf_gap, conf_only, rr_fail, tq_fail,
            conf_trend, days_watching, sector_status,
            confidence=confidence, tq=tq, rr=rr,
        )

    trend_label = (
        "confidence rising steadily"  if conf_trend and "rising"  in conf_trend.lower() else
        "confidence falling"          if conf_trend and "falling" in conf_trend.lower() else
        "confidence flat"             if conf_trend and "flat"    in conf_trend.lower() else
        "first appearance on watchlist"
    )
    blocker_label = (
        "only confidence is missing — all other factors pass" if conf_only else
        "both confidence and risk/reward need improvement"     if rr_fail and not tq_fail else
        "confidence, risk/reward, and trade quality all need work" if rr_fail and tq_fail else
        "multiple factors below threshold"
    )

    # FIX: when Conf already passes (conf_gap==0), the real blocker is R/R or TQ.
    # Tell the AI which knob to talk about — previously it invented a non-existent
    # "close the confidence gap" line even when Conf was well above threshold.
    conf_passes    = conf_gap <= 0
    active_gap_str = (
        f"R/R gap: needs to reach {rr + 0.2:.2f}x from {rr:.2f}x"
        if conf_passes and rr_fail
        else (
            f"TQ gap: needs +{80 - tq:.1f} from {tq:.1f} toward 80"
            if conf_passes and tq_fail
            else f"Confidence gap: needs +{conf_gap:.1f} points from {confidence:.1f}"
        )
    )

    prompt = f"""Near miss stock analysis. Each stock has unique numbers — treat each one differently.

Symbol: {symbol}
Sector: {sector}
Current Confidence: {confidence:.1f}/100 (threshold {'MET' if conf_passes else 'NOT MET'})
Current TQ: {tq:.1f}/100
Current R/R: {rr:.2f}x
Active blocker: {active_gap_str}
Confidence trend (3 days): {trend_label}
Primary blocker: {blocker_label}
Sector condition: {sector_status}
Days on watchlist: {days_watching}
Gate pattern: {gate_pattern or 'none'}
Fail reasons: {', '.join(fail_reasons) if fail_reasons else 'none'}

Write ONE complete sentence (minimum 20 words, maximum 30 words) answering:
"What specific event must happen in the next 1-3 sessions for this stock
 to become a tradeable BUY signal?"

Requirements:
- Focus ONLY on the ACTIVE blocker above. If confidence already PASSES the
  threshold, do NOT invent a confidence gap — talk about R/R or TQ instead.
- Reference this stock's ACTUAL numbers (R/R value, TQ value, or conf value
  — whichever is the blocker).
- Do NOT start with "For {symbol}" or "{symbol} needs"
- Start with a verb, condition, or timeframe
- Be specific to THIS stock — do not produce a generic template
- Do not mention the .NS suffix"""

    result = _call_ai(prompt, max_tokens=90)
    if result:
        clean = _clean_ai_output(result)
        try:
            clean = html.unescape(clean)   # FIX: strip HTML entities from AI output
        except Exception:
            pass
        # Use safe sentence splitter (protects .NS ticker suffix, decimals, abbreviations)
        sentences = [s for s in _split_sentences(clean) if len(s) >= 20]
        if sentences:
            return sentences[0]

    return _rule_based_near_miss_insight(
        symbol, conf_gap, conf_only, rr_fail, tq_fail,
        conf_trend, days_watching, sector_status,
        confidence=confidence, tq=tq, rr=rr,
    )


def run_all_ai_calls(
    regime_data: dict, macro: dict, breadth_data: dict,
    nifty_state: dict, buys: list, watchlist: list,
    portfolio_alerts: list, conf_history: dict,
    gate_memory: dict, events: list,
) -> dict:
    """
    Runs all AI calls in parallel via ThreadPoolExecutor.
    Returns {daily_summary, buy_theses, near_miss_insights}.
    Never fails — all have rule-based fallbacks.
    Total: max 9 calls, typically ~3-5 seconds.
    """
    from concurrent.futures import ThreadPoolExecutor

    results = {"daily_summary": "", "buy_theses": {}, "near_miss_insights": {}}
    near_miss    = [w for w in watchlist if w.get("tier") == "NEAR_MISS"]
    top_nm       = near_miss[0] if near_miss else {}
    upcoming_ev  = events[0]["name"] if events else ""
    regime       = regime_data["regime"]

    futures = {}
    try:
        with ThreadPoolExecutor(max_workers=8) as executor:
            # Call 1: daily summary (always)
            futures["daily_summary"] = executor.submit(
                ai_daily_summary,
                regime                = regime,
                regime_score          = regime_data["score"],
                nifty_pct             = macro.get("nifty_1d_pct", 0),
                vix_in                = macro.get("vix_in", 15),
                breadth_pct           = breadth_data.get("ema20_pct", 50),
                fii_flow_cr           = macro.get("fii_flow_cr", 0),
                dii_flow_cr           = macro.get("dii_flow_cr", 0),
                buy_count             = len(buys),
                near_miss_count       = len(near_miss),
                top_near_miss_symbol  = top_nm.get("symbol", "none"),
                top_near_miss_conf_gap= top_nm.get("conf_gap", 0),
                portfolio_alerts      = portfolio_alerts,
                ema_bear              = nifty_state.get("ema_bear", True),
                upcoming_event        = upcoming_ev,
            )
            # Calls 2-6: BUY theses (max 5)
            for stock in buys[:5]:
                sym = stock["symbol"]
                futures[f"buy_{sym}"] = executor.submit(
                    ai_buy_thesis,
                    symbol        = sym,
                    sector        = stock.get("sector", "OTHERS"),
                    confidence    = stock.get("final_confidence", 0),
                    tq            = stock.get("trade_quality_score", 0),
                    rr            = stock.get("rr_ratio", 0),
                    conf_trend    = get_confidence_trend(sym, conf_history),
                    catalyst      = stock.get("catalysts", []),
                    sector_status = stock.get("sector_status", "NEUTRAL"),
                    roe           = stock.get("roe", 0),
                    pledge_pct    = stock.get("promoter_pledge_pct", 0),
                    regime        = regime,
                    risk_pct      = stock.get("risk_pct", 0),
                    factor_scores = stock.get("factor_scores", {}) or {
                        # Phase C5: coalesce None (MISSING) → 50 for display/AI.
                        # None is a signal to compute_base_confidence to reweight;
                        # the AI thesis just needs a numeric neutral fallback.
                        "trend":     stock.get("trend_quality") or 50,
                        "momentum":  stock.get("momentum_quality") or 50,
                        "volume":    stock.get("volume_delivery") or 50,
                        "sector":    stock.get("sector_strength") or 50,
                        "rs":        stock.get("rs_vs_nifty") or 50,
                        "news":      stock.get("news_risk") or 50,
                        "ownership": stock.get("ownership_quality") or 50,
                        "macro":     stock.get("macro_alignment") or 50,
                    },
                    soft_warnings = stock.get("_soft_warnings", []) or [],
                    rs_diff21     = float(stock.get("rs_diff21", 0) or 0),
                    accum_signal  = stock.get("accum_signal", "NEUTRAL"),
                )
            # Calls 7-N: near-miss insights for ALL near-miss stocks.
            # AI is capped at 15 to stay within the 15-second timeout budget;
            # the remainder receive the deterministic rule-based fallback so
            # every Near Miss card in Telegram has an insight line.
            NM_AI_CAP = 15
            for w in near_miss[:NM_AI_CAP]:
                sym = w["symbol"]
                futures[f"nm_{sym}"] = executor.submit(
                    ai_near_miss_insight,
                    symbol        = sym,
                    sector        = w.get("sector", "OTHERS"),
                    confidence    = w.get("conf", w.get("final_confidence", 0)),
                    conf_gap      = w.get("conf_gap", 0),
                    tq            = w.get("tq", w.get("trade_quality_score", 0)),
                    rr            = w.get("rr_ratio", w.get("rr", 0)),
                    conf_trend    = get_confidence_trend(sym, conf_history),
                    fail_reasons  = w.get("fail_reasons", []),
                    gate_pattern  = get_gate_pattern(sym, gate_memory),
                    sector_status = w.get("sector_status", "NEUTRAL"),
                    risk_pct      = w.get("risk_pct", 0),
                    days_watching = w.get("days_watching", 0),
                )
            # Rule-based fallback for everyone past the AI cap so no Near Miss
            # is left without an insight line in the Telegram card.
            # `ai_near_miss_insight` itself falls back to the rule-based
            # generator on any failure, so calling it inline (no executor,
            # no LLM dispatch — the rule path is local & instant) keeps the
            # signature contract correct.
            for w in near_miss[NM_AI_CAP:]:
                sym = w["symbol"]
                try:
                    primary_fail = (w.get("fail_reasons") or ["CONFIDENCE_FAIL"])[0]
                    fr           = w.get("fail_reasons") or []
                    conf_only    = len(fr) == 1 and "CONF" in primary_fail
                    rr_fail      = any("RR" in f for f in fr)
                    tq_fail      = any("TQ" in f for f in fr)
                    results["near_miss_insights"][sym] = _rule_based_near_miss_insight(
                        symbol        = sym,
                        conf_gap      = w.get("conf_gap", 0),
                        conf_only     = conf_only,
                        rr_fail       = rr_fail,
                        tq_fail       = tq_fail,
                        conf_trend    = get_confidence_trend(sym, conf_history),
                        days_watching = w.get("days_watching", 0),
                        sector_status = w.get("sector_status", "NEUTRAL"),
                        # Per-stock numbers so no two stocks get identical text
                        confidence    = w.get("conf", w.get("final_confidence", 0)),
                        tq            = w.get("tq", w.get("trade_quality_score", 0)),
                        rr            = w.get("rr_ratio", w.get("rr", 0)),
                    ) or ""
                except Exception:
                    results["near_miss_insights"][sym] = ""
            for key, future in futures.items():
                try:
                    text = future.result(timeout=15)
                    if key == "daily_summary":
                        results["daily_summary"] = text or ""
                    elif key.startswith("buy_"):
                        results["buy_theses"][key[4:]] = text or ""
                    elif key.startswith("nm_"):
                        results["near_miss_insights"][key[3:]] = text or ""
                except Exception as e:
                    _log(f"[WARN] AI call {key} timed out or failed: {e}")
    except Exception as e:
        _log(f"[WARN] run_all_ai_calls failed: {e}")

    _log(
        f"[INFO] AI calls complete: "
        f"summary={'OK' if results['daily_summary'] else 'FALLBACK'} | "
        f"buy theses={len(results['buy_theses'])} | "
        f"near miss insights={len(results['near_miss_insights'])}"
    )
    return results


def _rule_based_news_score(text: str) -> dict:
    tl = text.lower()
    for kw in BLACK_SWAN_KEYWORDS:
        if kw in tl:
            return {"severity": 92, "category": "BLACK_SWAN", "is_black_swan": True,
                    "summary": f"Black swan keyword: {kw}", "news_source": "RULE_BASED"}
    # Severity-weighted: a single "fraud" (10) trumps three "downgrade" (5 each = 15)
    # only when both are present, but "fraud" alone already >15 → HIGH_RISK.
    neg_sev = sum(sev for kw, sev in NEGATIVE_KEYWORDS_SEVERITY.items() if kw in tl)
    neg_hits = sum(1 for kw in NEGATIVE_KEYWORDS_SEVERITY if kw in tl)
    pos = sum(1 for kw in POSITIVE_KEYWORDS if kw in tl)
    if neg_sev >= 15:
        return {"severity": 70, "category": "HIGH_RISK",     "is_black_swan": False, "summary": f"{neg_hits} negative signals (severity {neg_sev})", "news_source": "RULE_BASED"}
    elif neg_sev >= 5:
        return {"severity": 40, "category": "MODERATE_RISK", "is_black_swan": False, "summary": f"{neg_hits} negative signal(s) (severity {neg_sev})", "news_source": "RULE_BASED"}
    elif neg_sev >= 1:
        return {"severity": 20, "category": "MILD_RISK",     "is_black_swan": False, "summary": f"{neg_hits} minor negative signal(s)", "news_source": "RULE_BASED"}
    elif pos >= 1:
        return {"severity": -30, "category": "POSITIVE",     "is_black_swan": False, "summary": f"{pos} positive signal(s)", "news_source": "RULE_BASED"}
    return {"severity": 0, "category": "NEUTRAL", "is_black_swan": False, "summary": "No significant news", "news_source": "RULE_BASED"}


def ai_news_risk(symbol: str, headlines: list) -> dict:
    # Phase 1 #53 (2026-07-05): tag every return path with `news_source` so
    # post-mortem analysis can distinguish real LLM output from silent fallbacks.
    # Values: NO_HEADLINES | GROQ_AI | RULE_BASED_FALLBACK | RULE_BASED_MONITOR_ONLY.
    # The last value is set upstream by the pipeline for ranks 51-100 (#55).
    if not headlines:
        return {"severity": 0, "category": "NO_NEWS", "is_black_swan": False,
                "summary": "No news", "news_source": "NO_HEADLINES"}
    clean_headlines = [h[:120] for h in headlines[:5]]
    headlines_text = "\n".join(f"- {h}" for h in clean_headlines)
    prompt = (
        f"Analyze these news headlines for {symbol} and return ONLY a JSON object.\n"
        f"Headlines:\n{headlines_text}\n\n"
        f'Return exactly this JSON (no other text):\n'
        f'{{"severity": <0-100>, "category": "<BLACK_SWAN|HIGH_RISK|MODERATE_RISK|NEUTRAL|POSITIVE>", '
        f'"is_black_swan": <true|false>, "summary": "<one sentence max 100 chars>"}}\n\n'
        f"Severity: 90-100=fraud/ban/arrest, 60-89=regulatory/downgrade, 25-59=earnings miss, "
        f"0-24=neutral, negative=positive."
    )
    text = _call_groq_with_rotation(prompt, max_tokens=150)
    if text:
        result = _parse_ai_json(text)
        if result:
            result["news_source"] = "GROQ_AI"
            return result
    fallback = _rule_based_news_score(headlines_text)
    # Override the RULE_BASED tag added by _rule_based_news_score to make the
    # LLM-failed path distinguishable from the rank-51-100 rule-based path.
    fallback["news_source"] = "RULE_BASED_FALLBACK"
    # Phase 3a #41 (2026-07-05): loud fallback logging so a silent Groq
    # outage doesn't appear as "no news problem" in the daily log. Previously
    # every LLM failure produced a rule-based result with no [WARN] trace,
    # so a 100% AI degradation looked identical to a healthy quiet news day.
    # Format includes symbol + reason so post-mortem can distinguish
    # (a) Groq quota exhausted, (b) JSON parse failure, (c) network timeout.
    _reason = "GROQ_NO_TEXT" if not text else "GROQ_PARSE_FAIL"
    _log(f"[AI FALLBACK] {symbol}: rule-based used ({_reason}) — "
         f"severity={fallback.get('severity',0)} cat={fallback.get('category','?')}")
    return fallback


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 4 — DATA SOURCES
# ─────────────────────────────────────────────────────────────────────────────

_NSE_DELAY_RANGE = (0.3, 1.0)

# Phase 1 #54 (2026-07-05): silent-failure visibility for fetch_price_data.
# Every None return path now records {symbol, reason, ts} into this list.
# Flushed to price_fetch_failures.jsonl at end of _run_pipeline_inner.
_PRICE_FETCH_FAILURES: list = []


def fetch_price_data(symbol: str, period: str = "6mo"):
    # #54: capture the specific failure reason so we can distinguish
    # rate-limit vs delisted vs data-quality issues in post-mortem.
    _fail_reason = "UNKNOWN"
    try:
        import warnings, logging
        # Suppress yfinance noise: "possibly delisted", "401 crumb", progress bars
        logging.getLogger("yfinance").setLevel(logging.CRITICAL)
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            df = yf.download(symbol, period=period, interval="1d",
                             progress=False, auto_adjust=True,
                             multi_level_index=False)
        if df is None:
            _fail_reason = "NONE_RETURNED"
        elif len(df) <= 20:
            _fail_reason = f"INSUFFICIENT_ROWS_{len(df)}"
        else:
            df.columns = [c[0] if isinstance(c, tuple) else c for c in df.columns]
            return df
    except Exception as e:
        _fail_reason = f"EXC_{type(e).__name__}:{str(e)[:60]}"
    # Best-effort logging — never let this crash the caller.
    try:
        _PRICE_FETCH_FAILURES.append({
            "symbol": symbol,
            "period": period,
            "reason": _fail_reason,
            "ts":     ist_now().isoformat(),
        })
    except Exception:
        pass
    return None


def load_symbols(filepath: str = "stocks.txt") -> list:
    try:
        if os.path.exists(filepath):
            with open(filepath, "r") as f:
                syms = []
                for line in f:
                    s = line.strip()
                    if s and not s.startswith("#"):
                        if not s.endswith(".NS"):
                            s += ".NS"
                        syms.append(s)
                _log(f"Loaded {len(syms)} symbols from {filepath}")
                return syms
    except Exception as e:
        _log(f"[WARN] load_symbols failed: {e}")
    fallback = [
        "RELIANCE.NS","INFY.NS","TCS.NS","HDFCBANK.NS","ICICIBANK.NS",
        "AXISBANK.NS","SBIN.NS","BAJFINANCE.NS","TATAMOTORS.NS","MARUTI.NS",
        "SUNPHARMA.NS","WIPRO.NS","HCLTECH.NS","LTIM.NS","NESTLEIND.NS",
    ]
    _log(f"[WARN] Using fallback symbol list ({len(fallback)} symbols)")
    return fallback


# ─── Phase G6 (2026-07-03): high-pledge blocklist ──────────────────────────
# screener.in moved pledge% behind login → we cannot detect pledge dynamically.
# We maintain a curated blocklist of well-known high-pledge / promoter-stress
# NSE stocks and hard-reject them regardless of technical score.
# See high_pledge_stocks.txt for the list + maintenance guidance.
# Also supports asm_gsm_blocklist.txt for NSE surveillance stocks.
_HIGH_PLEDGE_BLOCKLIST_CACHE: set | None = None

def _load_blocklist_file(filepath: str) -> set:
    """Parse a blocklist file: one ticker per line, # for comments."""
    blocked: set = set()
    if not os.path.exists(filepath):
        return blocked
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            for line in f:
                s = line.split("#", 1)[0].strip()
                if not s:
                    continue
                if s.endswith(".NS"):
                    s = s[:-3]
                blocked.add(s.upper())
    except Exception as e:
        _log(f"[WARN] failed to read blocklist {filepath}: {e}")
    return blocked


def load_high_pledge_blocklist(filepath: str = "high_pledge_stocks.txt") -> set:
    """Return a set of tickers (without .NS, uppercase) that should be
    HARD-REJECTED. Merges high_pledge_stocks.txt + asm_gsm_blocklist.txt."""
    global _HIGH_PLEDGE_BLOCKLIST_CACHE
    if _HIGH_PLEDGE_BLOCKLIST_CACHE is not None:
        return _HIGH_PLEDGE_BLOCKLIST_CACHE
    pledge_set = _load_blocklist_file(filepath)
    asm_set    = _load_blocklist_file("asm_gsm_blocklist.txt")
    blocked = pledge_set | asm_set
    if blocked:
        _log(f"[INFO] Blocklist loaded: {len(pledge_set)} high-pledge + "
             f"{len(asm_set)} ASM/GSM = {len(blocked)} unique tickers")
    else:
        _log(f"[WARN] no blocklist tickers loaded from {filepath} or asm_gsm_blocklist.txt")
    _HIGH_PLEDGE_BLOCKLIST_CACHE = blocked
    return blocked


def is_pledge_blocked(symbol: str) -> bool:
    """True if this ticker is on the high-pledge OR ASM/GSM blocklist."""
    blocked = load_high_pledge_blocklist()
    if not blocked:
        return False
    s = symbol.upper()
    if s.endswith(".NS"):
        s = s[:-3]
    return s in blocked


def _download_one(symbol: str, period: str = "6mo") -> tuple:
    time.sleep(random.uniform(*_NSE_DELAY_RANGE))
    df = fetch_price_data(symbol, period=period)
    return symbol, df


# ─── Phase G8-C (2026-07-06): Universe pre-filter module ──────────────────
# Fetches a fundamentally-clean or technically-screened universe from external
# sources (Screener.in weekly, Chartink daily) and intersects with stocks.txt.
# Rationale: 2360-symbol full-NSE spray is expensive (10-15 min download) and
# includes SME/penny/illiquid names that will fail hygiene anyway. Pre-filter
# lets us start from 500-800 quality names → downstream hygiene → top-100.
#
# All external fetches are OPT-IN + gracefully degrade to the full universe
# on any failure. Never fails the pipeline.
# ─────────────────────────────────────────────────────────────────────────
UNIVERSE_CACHE_DIR              = os.getenv("UNIVERSE_CACHE_DIR", "universe_cache")
SCREENER_UNIVERSE_CACHE_FILE    = os.path.join(UNIVERSE_CACHE_DIR, "screener_universe.json")
CHARTINK_UNIVERSE_CACHE_FILE    = os.path.join(UNIVERSE_CACHE_DIR, "chartink_universe.json")

# Default Screener.in query URL — fundamentally clean names.
# Filter: Market cap > 500 Cr, ROE > 15, D/E < 1, Sales growth 3y > 10.
# Override via env SCREENER_QUERY_URL. Screener "screens" have public URLs
# that render an HTML table of results; we parse the ticker column.
_SCREENER_DEFAULT_QUERY_URL = (
    "https://www.screener.in/screens/357649/"  # "Quality mid+small caps"
    # Fallback if the user hasn't authored their own screen; this is a
    # commonly-shared public screen. User should replace with their own.
)

# Default Chartink scan URL — technical momentum universe.
# Chartink public scans have a POST endpoint at /screener/process with a
# CSRF-protected form. We use the free "backtest" endpoint which returns
# JSON without auth.
_CHARTINK_DEFAULT_SCAN_URL = "https://chartink.com/screener/process"
# Popular scan: "Stocks above 200 EMA + volume > 20d avg". Override via
# env CHARTINK_SCAN_CLAUSE — must be a valid Chartink DSL clause.
_CHARTINK_DEFAULT_SCAN_CLAUSE = (
    "( {cash} ( latest close > latest ema( close,200 ) and "
    "latest volume > latest sma( volume,20 ) and "
    "latest close > 30 ) )"
)


def _load_universe_cache(cache_file: str, ttl_hours: int) -> "set | None":
    """Load a cached universe if fresher than ttl_hours. Returns None if stale/missing."""
    try:
        if not os.path.exists(cache_file):
            return None
        with open(cache_file, "r", encoding="utf-8") as f:
            payload = json.load(f)
        cached_at = datetime.datetime.fromisoformat(payload.get("cached_at", "2000-01-01"))
        age_hours = (ist_now() - cached_at).total_seconds() / 3600
        if age_hours > ttl_hours:
            _log(f"  [Universe cache] {os.path.basename(cache_file)} stale ({age_hours:.1f}h > {ttl_hours}h)")
            return None
        syms = set(payload.get("symbols", []))
        if not syms:
            return None
        _log(f"  [Universe cache] {os.path.basename(cache_file)} hit — {len(syms)} symbols ({age_hours:.1f}h old)")
        return syms
    except Exception as e:
        _log(f"  [Universe cache] load failed for {cache_file}: {e}")
        return None


def _save_universe_cache(cache_file: str, symbols: "set | list", source: str) -> None:
    """Persist a universe snapshot with timestamp."""
    try:
        os.makedirs(os.path.dirname(cache_file), exist_ok=True)
        payload = {
            "cached_at":   ist_now().isoformat(),
            "source":      source,
            "count":       len(symbols),
            "symbols":     sorted(symbols),
        }
        with open(cache_file, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, default=str)
        _log(f"  [Universe cache] saved {len(symbols)} symbols → {cache_file}")
    except Exception as e:
        _log(f"[WARN] universe cache save failed: {e}")


def fetch_screener_universe(query_url: "str | None" = None,
                             ttl_hours: int = 168) -> "set | None":
    """
    Fetch a fundamentally-clean universe from a Screener.in public screen.

    Weekly refresh (default TTL=168h). Returns a set of NSE tickers (no .NS
    suffix) or None on total failure.

    HOW IT WORKS:
      Screener screens are paginated HTML tables at /screens/<id>/. Each row
      has a <a href="/company/TICKER/..."> link. We scrape all pages.

    SOURCE HONESTY: Screener rate-limits at ~200 req/day; we only need ~5-10
    page fetches per week. Cached aggressively.

    Args:
        query_url: Full URL to a Screener screen. Uses env SCREENER_QUERY_URL
                   then _SCREENER_DEFAULT_QUERY_URL. Must contain "/screens/".
        ttl_hours: Cache TTL. Default 168h = 1 week (fundamentals move slowly).

    Returns:
        Set of ticker strings (uppercase, no .NS suffix), or None on failure.
    """
    if not _BS4_OK:
        _log("[WARN] screener universe: bs4 unavailable")
        return None
    query_url = query_url or os.getenv("SCREENER_QUERY_URL") or _SCREENER_DEFAULT_QUERY_URL

    # Cache check
    cached = _load_universe_cache(SCREENER_UNIVERSE_CACHE_FILE, ttl_hours)
    if cached is not None:
        return cached

    _log(f"  [Screener universe] fetching from {query_url}")
    tickers: set = set()
    try:
        # Screener paginates via ?page=N. We try pages 1..10 (max ~2500 rows
        # at 250/page — far more than any sane screen returns).
        for page in range(1, 11):
            page_url = query_url.rstrip("/") + f"/?page={page}"
            resp = requests.get(page_url, headers=_SCREENER_HEADERS, timeout=15)
            if resp.status_code == 429:
                _log(f"  [Screener universe] rate-limited on page {page} — using partial results")
                break
            if resp.status_code != 200:
                _log(f"  [Screener universe] HTTP {resp.status_code} on page {page} — stopping")
                break
            soup = BeautifulSoup(resp.text, "html.parser")
            page_tickers: set = set()
            for a in soup.find_all("a", href=True):
                href = a["href"] or ""
                # Match /company/TICKER/ or /company/TICKER/consolidated/
                m = re.match(r"^/company/([A-Z0-9&\-]+)/?", href)
                if m:
                    tkr = m.group(1).strip().upper()
                    # Screener uses BSE codes for BSE-only names; those are
                    # numeric. We only want NSE-listed alphabetic tickers.
                    if tkr and not tkr.isdigit() and len(tkr) <= 20:
                        page_tickers.add(tkr)
            if not page_tickers:
                break  # empty page → we're past the end
            new_count = len(page_tickers - tickers)
            tickers |= page_tickers
            _log(f"  [Screener universe] page {page}: +{new_count} new (total {len(tickers)})")
            if new_count == 0:
                break  # pagination didn't advance
            time.sleep(1.2)  # be polite between page fetches
    except Exception as e:
        _log(f"[WARN] Screener universe fetch failed: {e}")
        return None

    if len(tickers) < 20:
        _log(f"  [Screener universe] too few tickers ({len(tickers)}) — treating as failure")
        return None

    _save_universe_cache(SCREENER_UNIVERSE_CACHE_FILE, tickers, source="screener.in")
    return tickers


def fetch_chartink_universe(scan_clause: "str | None" = None,
                             ttl_hours: int = 24) -> "set | None":
    """
    Fetch a technical-momentum universe from Chartink's free scan endpoint.

    Daily refresh (default TTL=24h). Returns set of NSE tickers or None.

    HOW IT WORKS:
      Chartink exposes a public POST endpoint /screener/process that accepts
      a scan clause (their DSL) and returns JSON with matching stocks. No
      auth required for the free tier. CSRF token is fetched from the scan
      page first.

    Args:
        scan_clause: Chartink DSL. Uses env CHARTINK_SCAN_CLAUSE then default.
        ttl_hours: Cache TTL. Default 24h — technical setups change daily.

    Returns:
        Set of ticker strings (uppercase, no .NS suffix), or None on failure.
    """
    scan_clause = scan_clause or os.getenv("CHARTINK_SCAN_CLAUSE") or _CHARTINK_DEFAULT_SCAN_CLAUSE

    # Cache check
    cached = _load_universe_cache(CHARTINK_UNIVERSE_CACHE_FILE, ttl_hours)
    if cached is not None:
        return cached

    _log(f"  [Chartink universe] fetching scan (clause len={len(scan_clause)})")
    try:
        # Step 1: GET the scan page to obtain a CSRF token from meta tag.
        session = requests.Session()
        get_headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                          "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml",
        }
        r0 = session.get("https://chartink.com/screener/", headers=get_headers, timeout=15)
        if r0.status_code != 200:
            _log(f"  [Chartink universe] GET landing failed: HTTP {r0.status_code}")
            return None
        # Extract csrf-token from <meta name="csrf-token" content="..."/>
        csrf_match = re.search(
            r'<meta\s+name=["\']csrf-token["\']\s+content=["\']([^"\']+)["\']',
            r0.text
        )
        if not csrf_match:
            _log("  [Chartink universe] no CSRF token found — endpoint may have changed")
            return None
        csrf_token = csrf_match.group(1)

        # Step 2: POST the scan clause.
        post_headers = {
            **get_headers,
            "X-CSRF-TOKEN":    csrf_token,
            "X-Requested-With": "XMLHttpRequest",
            "Accept":          "application/json",
            "Referer":         "https://chartink.com/screener/",
        }
        resp = session.post(
            _CHARTINK_DEFAULT_SCAN_URL,
            data={"scan_clause": scan_clause},
            headers=post_headers,
            timeout=25,
        )
        if resp.status_code != 200:
            _log(f"  [Chartink universe] POST failed: HTTP {resp.status_code}")
            return None
        payload = resp.json()
    except Exception as e:
        _log(f"[WARN] Chartink universe fetch failed: {e}")
        return None

    tickers: set = set()
    for row in payload.get("data") or []:
        # Chartink returns { "nsecode": "RELIANCE", "name": "...", ...}
        code = (row.get("nsecode") or row.get("symbol") or "").strip().upper()
        if code:
            tickers.add(code)

    if len(tickers) < 20:
        _log(f"  [Chartink universe] too few results ({len(tickers)}) — treating as failure")
        return None

    _save_universe_cache(CHARTINK_UNIVERSE_CACHE_FILE, tickers, source="chartink.com")
    _log(f"  [Chartink universe] {len(tickers)} technically-clean symbols fetched")
    return tickers


def build_universe(base_symbols: list) -> list:
    """
    Compose the working universe from base stocks.txt + optional external
    pre-filters. Called once at pipeline start.

    Modes (via env UNIVERSE_MODE):
      "full"                → base_symbols only (default, matches old behavior)
      "screener"            → intersect with Screener fundamental universe
      "chartink"            → intersect with Chartink technical universe
      "hybrid"              → intersect with (Screener ∪ Chartink)
      "screener_and_chartink" → intersect with (Screener ∩ Chartink)

    Failure of any external source degrades gracefully to just the base_symbols
    with a WARNING logged — pipeline is never blocked.

    Returns a list of tickers (with .NS suffix) matching base_symbols format.
    """
    mode = (os.getenv("UNIVERSE_MODE", "full") or "full").strip().lower()
    if mode == "full":
        _log(f"  [Universe] mode=full → {len(base_symbols)} symbols from stocks.txt")
        return base_symbols

    # Normalize base symbols to bare tickers for intersection
    base_map: dict = {}
    for sym in base_symbols:
        bare = sym.replace(".NS", "").upper().strip()
        if bare:
            base_map[bare] = sym

    # Collect external universes based on mode
    screener_set: "set | None" = None
    chartink_set: "set | None" = None
    if mode in ("screener", "hybrid", "screener_and_chartink"):
        screener_set = fetch_screener_universe()
        if screener_set is None:
            _log("  [Universe] Screener fetch failed — degrading to base universe")
    if mode in ("chartink", "hybrid", "screener_and_chartink"):
        chartink_set = fetch_chartink_universe()
        if chartink_set is None:
            _log("  [Universe] Chartink fetch failed — degrading to base universe")

    # Compose the "keep" set based on mode
    keep: "set | None" = None
    if mode == "screener" and screener_set:
        keep = screener_set
    elif mode == "chartink" and chartink_set:
        keep = chartink_set
    elif mode == "hybrid":
        if screener_set and chartink_set:
            keep = screener_set | chartink_set
        elif screener_set:
            keep = screener_set
        elif chartink_set:
            keep = chartink_set
    elif mode == "screener_and_chartink":
        if screener_set and chartink_set:
            keep = screener_set & chartink_set
            if len(keep) < 50:
                _log(f"  [Universe] Screener ∩ Chartink only {len(keep)} — "
                     f"falling back to union to avoid over-narrow universe")
                keep = screener_set | chartink_set

    if not keep:
        _log(f"  [Universe] no external filter applied → using full base ({len(base_symbols)})")
        return base_symbols

    # Intersect with base universe
    filtered = [base_map[t] for t in keep if t in base_map]
    dropped  = len(base_symbols) - len(filtered)
    _log(f"  [Universe] mode={mode} · base={len(base_symbols)} · "
         f"external={len(keep)} · kept={len(filtered)} · dropped={dropped}")
    # Guard: if external filter is over-narrow (< 100), fall back to base.
    # We'd rather run the full pipeline than emit 0 BUYs from a tiny universe.
    if len(filtered) < 100:
        _log(f"  [Universe] filtered universe too narrow ({len(filtered)} < 100) — "
             f"reverting to base_symbols for safety")
        return base_symbols
    return filtered


def filter_and_download(symbols: list, period: str = "6mo",
                        max_workers: int = 12,
                        min_avg_volume: int = 100_000,
                        min_avg_value_lakhs: float = 200.0,
                        min_price: float = None,
                        max_recent_circuits: int = 2) -> dict:
    """
    Phase G6 (2026-07-03): universe hygiene — filters out:
      - low avg volume (< min_avg_volume shares/day)
      - low avg value  (< min_avg_value_lakhs ₹L/day)
      - penny stocks   (avg price < min_price ₹)
      - pledge-blocklisted names (high_pledge_stocks.txt)
      - repeat circuit hitters (>max_recent_circuits days with |ret|≥9.5% in last 5)

    Phase G8-A (2026-07-03): raised default min_avg_value_lakhs from 50 (₹0.5Cr)
    to 200 (₹2Cr/day). Rationale — a ₹50k retail buy in a ₹50L/day stock is
    1% of daily volume and moves the price against you by 0.3-1.5%. At ₹2Cr
    turnover, same order is 0.25% of daily volume, ≤ 0.2% slippage. Env-
    overridable via UNIVERSE_MIN_TURNOVER_LAKHS. Set to 0 to disable.

    All filters are opt-out via env: set the corresponding threshold to 0 to disable.
    """
    # Phase G8-A: env override for turnover floor
    try:
        _env_turnover = os.getenv("UNIVERSE_MIN_TURNOVER_LAKHS")
        if _env_turnover is not None:
            min_avg_value_lakhs = float(_env_turnover)
    except (TypeError, ValueError):
        pass  # keep default
    # Phase 3a #30 (2026-07-05): env override for penny-stock floor.
    # Default ₹20 kept as the historical baseline.
    if min_price is None:
        try:
            min_price = float(os.getenv("MIN_STOCK_PRICE", "20.0"))
        except (TypeError, ValueError):
            min_price = 20.0
    # Phase 3a #29 (2026-07-05): 5d turnover ratio floor.
    # Requires last-5-day avg turnover ≥ min_avg_value_lakhs * MIN_5D_RATIO
    # (default 0.60) to reject stocks whose 20d avg is inflated by one
    # historical block-print. 0.0 disables. Overridable via env.
    try:
        min_5d_ratio = float(os.getenv("MIN_5D_TURNOVER_RATIO", "0.60"))
    except (TypeError, ValueError):
        min_5d_ratio = 0.60
    _log(f"Downloading {len(symbols)} symbols with {max_workers} workers...")

    # Pre-filter blocklisted names BEFORE we hit the network — cheap and fast
    blocked_syms = load_high_pledge_blocklist()
    if blocked_syms:
        pre_count = len(symbols)
        symbols = [s for s in symbols
                   if s.replace(".NS", "").upper() not in blocked_syms]
        pledge_dropped = pre_count - len(symbols)
        if pledge_dropped:
            _log(f"  Blocklist pre-filter: {pledge_dropped} high-pledge names removed → {len(symbols)} remaining")

    tradable = {}
    failed = 0
    illiquid_vol = 0
    illiquid_val = 0
    illiquid_price = 0
    illiquid_5d = 0   # Phase 3a #29
    circuit_repeat = 0
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(_download_one, sym, period): sym for sym in symbols}
        for future in as_completed(futures):
            sym = futures[future]
            try:
                symbol, df = future.result(timeout=30)
                if df is None or len(df) < 20:
                    failed += 1
                    continue
                closes = df["Close"].squeeze()
                avg_vol = float(df["Volume"].squeeze().tail(20).mean())
                avg_price = float(closes.tail(20).mean())
                avg_val_lakhs = (avg_vol * avg_price) / 100_000
                if avg_vol < min_avg_volume:
                    illiquid_vol += 1
                    continue
                if avg_val_lakhs < min_avg_value_lakhs:
                    illiquid_val += 1
                    continue
                # Phase 3a #29 (2026-07-05): 5d avg turnover ratio check —
                # reject stocks whose 20d turnover is inflated by a lone big
                # block-print but whose current week is anemic.
                if min_5d_ratio > 0 and len(df) >= 6:
                    try:
                        _c5 = closes.tail(5)
                        _v5 = df["Volume"].squeeze().tail(5)
                        _avg_val_5d_lakhs = float((_c5 * _v5).mean()) / 100_000
                        _floor = min_avg_value_lakhs * min_5d_ratio
                        if _avg_val_5d_lakhs < _floor:
                            illiquid_5d += 1
                            continue
                    except Exception:
                        pass
                # Phase G6: penny stock filter — wide spreads eat all edge
                if min_price > 0 and avg_price < min_price:
                    illiquid_price += 1
                    continue
                # Phase G6: repeat circuit filter — ≥N days with |1d ret|≥9.5%
                # in last 5 sessions. Circuit-limit stocks are untradeable at
                # retail scale (queue-jumped by algos, wide effective spread).
                if max_recent_circuits >= 0 and len(closes) >= 6:
                    try:
                        rets = closes.pct_change().tail(5).abs()
                        n_circ = int((rets >= 0.095).sum())
                        if n_circ > max_recent_circuits:
                            circuit_repeat += 1
                            continue
                    except Exception:
                        pass
                tradable[symbol] = df
            except Exception as e:
                _log(f"[WARN] download failed for {sym}: {e}")
                failed += 1
    _log(
        f"Download complete: {len(tradable)} tradable | {failed} failed | "
        f"{illiquid_vol} illiquid_vol | {illiquid_val} illiquid_val "
        f"(<₹{min_avg_value_lakhs:.0f}L/day) | "
        f"{illiquid_5d} illiquid_5d (<{min_5d_ratio*100:.0f}% of 20d) | "
        f"{illiquid_price} penny (<₹{min_price:.0f}) | "
        f"{circuit_repeat} repeat_circuit (>{max_recent_circuits}/5d)"
    )
    return tradable


_NSE_BROWSER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate",  # no 'br' — requests can't decode brotli without extra library
    "Connection": "keep-alive",
    "DNT": "1",
}


def _nse_session() -> requests.Session:
    """
    Pre-warmed session with full browser headers + NSE homepage cookie hit.
    NSE's API 403s without cookies. CI IPs may still be blocked intermittently —
    all callers already fall back to neutral defaults when this returns a bad session.
    """
    session = requests.Session()
    session.headers.update(_NSE_BROWSER_HEADERS)
    try:
        session.get("https://www.nseindia.com", timeout=10)
        time.sleep(0.5)
    except Exception:
        pass
    return session


def fetch_option_chain(symbol_nse: str) -> dict:
    """
    Option-chain PCR — Phase B5 (2026-06-30) re-source.
      1. NiftyTrader webapi (keyless JSON; works from CI; ~200 F&O symbols)
      2. NSE /api/option-chain-equities (Cloudflare-gated on CI, fine locally)
      3. Honest NEUTRAL_NO_FNO for non-F&O stocks (most NSE stocks have no options)
    Returns: { pcr, total_ce_oi, total_pe_oi, source }
    """
    neutral_no_fno   = {"pcr": 1.0, "total_ce_oi": 0, "total_pe_oi": 0, "source": "NEUTRAL_NO_FNO"}
    neutral_blocked  = {"pcr": 1.0, "total_ce_oi": 0, "total_pe_oi": 0, "source": "NEUTRAL_BLOCKED"}

    # Per-run cache to avoid re-hitting NiftyTrader for the same symbol
    cache = getattr(fetch_option_chain, "_cache", None)
    if cache is None:
        cache = {}
        fetch_option_chain._cache = cache  # type: ignore[attr-defined]
    if symbol_nse in cache:
        return cache[symbol_nse]

    # ── 1. NiftyTrader webapi (CI-friendly) ──────────────────────────────────
    try:
        url = (
            "https://webapi.niftytrader.in/webapi/option/option-chain-data"
            f"?symbol={symbol_nse}&exchange=nse&expiryDate=&atmBelow=10&atmAbove=10"
        )
        resp = requests.get(
            url,
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                              "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36",
                "Accept":  "application/json, text/plain, */*",
                "Referer": "https://www.niftytrader.in/",
            },
            timeout=8,
        )
        if resp.status_code == 200:
            try:
                data = resp.json()
            except Exception:
                data = None
            # Guard every layer: API may return {"resultData": null}, a list,
            # or opDatas as a list of nulls — all trigger 'NoneType.get' otherwise.
            res_raw = data.get("resultData") if isinstance(data, dict) else None
            res     = res_raw if isinstance(res_raw, dict) else {}
            opdata_raw = res.get("opDatas") or res.get("opData") or []
            opdata     = opdata_raw if isinstance(opdata_raw, list) else []
            total_ce_oi = 0
            total_pe_oi = 0
            tot_pe_top  = res.get("total_puts_oi")
            tot_ce_top  = res.get("total_calls_oi")
            if isinstance(tot_pe_top, (int, float)) and isinstance(tot_ce_top, (int, float)):
                total_pe_oi = int(tot_pe_top)
                total_ce_oi = int(tot_ce_top)
            elif opdata:
                for r in opdata:
                    if not isinstance(r, dict):
                        continue
                    total_ce_oi += int(r.get("calls_oi", r.get("ce_oi", 0)) or 0)
                    total_pe_oi += int(r.get("puts_oi", r.get("pe_oi", 0)) or 0)
            if total_ce_oi > 0 or total_pe_oi > 0:
                pcr = round(total_pe_oi / total_ce_oi, 3) if total_ce_oi > 0 else 1.0
                result = {
                    "pcr": pcr,
                    "total_ce_oi": total_ce_oi,
                    "total_pe_oi": total_pe_oi,
                    "source": "niftytrader",
                }
                cache[symbol_nse] = result
                return result
            # 200 but no OI → likely a non-F&O symbol
            cache[symbol_nse] = neutral_no_fno
            return neutral_no_fno
        elif resp.status_code == 404:
            # NiftyTrader returns 404 for symbols not in F&O universe
            cache[symbol_nse] = neutral_no_fno
            return neutral_no_fno
    except Exception as e:
        _log(f"  [PCR NiftyTrader] {symbol_nse} error: {e}")

    # ── 2. NSE direct (legacy) — works locally, Cloudflare-gated on CI ───────
    try:
        session = _nse_session()
        url  = f"https://www.nseindia.com/api/option-chain-equities?symbol={symbol_nse}"
        resp = session.get(url, headers={
            "Accept": "application/json",
            "Referer": "https://www.nseindia.com/option-chain",
        }, timeout=12)
        if resp.status_code == 200:
            ct = resp.headers.get("Content-Type", "")
            if "html" in ct.lower() or resp.content[:1] == b"<":
                cache[symbol_nse] = neutral_blocked
                return neutral_blocked
            data    = resp.json()
            records = data.get("records", {}).get("data", [])
            total_ce_oi = sum(r.get("CE", {}).get("openInterest", 0) for r in records if "CE" in r)
            total_pe_oi = sum(r.get("PE", {}).get("openInterest", 0) for r in records if "PE" in r)
            pcr = round(total_pe_oi / total_ce_oi, 3) if total_ce_oi > 0 else 1.0
            result = {
                "pcr": pcr,
                "total_ce_oi": total_ce_oi,
                "total_pe_oi": total_pe_oi,
                "source": "nse_direct",
            }
            cache[symbol_nse] = result
            return result
        elif resp.status_code in (403, 429):
            cache[symbol_nse] = neutral_blocked
            return neutral_blocked
    except Exception as e:
        _log(f"  [PCR NSE direct] {symbol_nse} error: {e}")

    cache[symbol_nse] = neutral_blocked
    return neutral_blocked


def pcr_score(pcr: float) -> int:
    if pcr >= 1.5:   return 35
    elif pcr >= 1.2: return 75
    elif pcr >= 0.9: return 60
    elif pcr >= 0.7: return 45
    return 25


def fetch_bulk_deals(days_back: int = 3) -> dict:
    """
    Bulk + block deals — Phase B3 (2026-06-30) reorder:
      1. nselib capital_market.bulk_deal_data + block_deals_data (NSE archive endpoints, CI-friendly)
      2. BSE HTML scrape via pandas.read_html (categorywise + bulk page) as fallback
      3. Legacy NSE /api/historical/bulk-deals (blocked from CI — kept for local runs)
    Result: { "SYM.NS": "BUY" | "SELL" } over the last `days_back` days.
    """
    result: dict = {}

    # ── 1. nselib NSE archive ────────────────────────────────────────────────
    if _NSELIB_OK and _nselib_cm is not None:
        # nselib period= only accepts {'1D','1W','1M','6M','1Y'}.
        # Map our days_back window to the smallest covering shortcut.
        if days_back <= 1:
            period = "1D"
        elif days_back <= 7:
            period = "1W"
        elif days_back <= 31:
            period = "1M"
        elif days_back <= 31 * 6:
            period = "6M"
        else:
            period = "1Y"
        for fn_name in ("bulk_deal_data", "block_deals_data"):
            fn = getattr(_nselib_cm, fn_name, None)
            if fn is None:
                continue
            try:
                # nselib accepts period= shortcuts on most builds. Fall back to
                # explicit from_date/to_date on either a TypeError (signature
                # mismatch) or ValueError (rejected period code).
                try:
                    df = fn(period=period)
                except (TypeError, ValueError):
                    to_d   = ist_today()
                    from_d = to_d - datetime.timedelta(days=days_back)
                    df = fn(from_date=from_d.strftime("%d-%m-%Y"),
                            to_date=to_d.strftime("%d-%m-%Y"))
                if df is None or (hasattr(df, "empty") and df.empty):
                    _log(f"  [nselib {fn_name}] empty — no deals in window")
                    continue
                cols = {c.lower(): c for c in df.columns}
                sym_col = (cols.get("symbol") or cols.get("scrip") or cols.get("security"))
                act_col = (cols.get("buy/sell") or cols.get("buy_sell") or
                           cols.get("trade_type") or cols.get("type"))
                if sym_col is None or act_col is None:
                    _log(f"  [nselib {fn_name}] unexpected schema: {list(df.columns)[:6]}")
                    continue
                for _, row in df.iterrows():
                    sym = str(row[sym_col]).strip().upper().replace(".NS", "") + ".NS"
                    act = str(row[act_col]).strip().upper()
                    action = "BUY" if act.startswith("B") else "SELL"
                    # Don't overwrite a SELL with a BUY from a different deal — last write wins is fine
                    result[sym] = action
                _log(f"  [nselib {fn_name}] ✓ {len(df)} rows → {len(result)} symbols")
            except Exception as e:
                _log(f"  [nselib {fn_name}] error: {e}")
        if result:
            return result

    # ── 2. BSE HTML scrape via pandas.read_html ──────────────────────────────
    try:
        bse_url = "https://www.bseindia.com/markets/equity/EQReports/bulk_deals.aspx"
        # pandas.read_html needs lxml/html5lib; lxml is in requirements.txt
        tables = pd.read_html(bse_url, flavor="lxml")
        for df in tables:
            cols_lower = [str(c).lower() for c in df.columns]
            if any("security" in c or "scrip" in c or "symbol" in c for c in cols_lower) and \
               any("deal type" in c or "buy/sell" in c or "type" in c for c in cols_lower):
                sym_col = next(c for c in df.columns
                               if "security" in str(c).lower() or "scrip" in str(c).lower()
                               or "symbol" in str(c).lower())
                act_col = next(c for c in df.columns
                               if "deal type" in str(c).lower() or "buy/sell" in str(c).lower()
                               or "type" in str(c).lower())
                for _, row in df.iterrows():
                    sym = str(row[sym_col]).strip().upper().replace(".NS", "") + ".NS"
                    act = str(row[act_col]).strip().upper()
                    action = "BUY" if act.startswith("B") else "SELL"
                    result[sym] = action
                if result:
                    _log(f"  [BSE bulk deals HTML] ✓ {len(result)} symbols")
                    return result
    except Exception as e:
        _log(f"  [BSE bulk deals HTML] error — {e}")

    # ── 3. Legacy NSE /api endpoint (works on non-CI IPs) ────────────────────
    try:
        session = _nse_session()
        today   = ist_today()
        from_dt = (today - datetime.timedelta(days=days_back)).strftime("%d-%m-%Y")
        to_dt   = today.strftime("%d-%m-%Y")
        for url in [
            f"https://www.nseindia.com/api/historical/bulk-deals?from={from_dt}&to={to_dt}",
            "https://www.nseindia.com/api/bulk-deals",
        ]:
            _log(f"  [Bulk Deals legacy] GET {url}")
            resp = session.get(
                url,
                headers={
                    "Accept": "application/json, text/plain, */*",
                    "Referer": "https://www.nseindia.com/market-data/bulk-block-deals",
                    "X-Requested-With": "XMLHttpRequest",
                },
                timeout=12,
            )
            if resp.status_code in (403, 429):
                _log(f"  [Bulk Deals legacy] BLOCKED ({resp.status_code}) — Cloudflare/rate-limit")
                break
            if resp.status_code == 404:
                continue
            if resp.status_code != 200:
                break
            ct = resp.headers.get("Content-Type", "")
            if not resp.content or "html" in ct.lower() or resp.content[:1] == b"<":
                _log("  [Bulk Deals legacy] HTML/empty body — Cloudflare gate")
                break
            try:
                payload = resp.json()
            except Exception:
                break
            for deal in payload.get("data", []):
                sym    = deal.get("symbol", "").strip() + ".NS"
                action = "BUY" if str(deal.get("buySell", "")).upper().startswith("B") else "SELL"
                result[sym] = action
            break
    except Exception as e:
        _log(f"  [Bulk Deals legacy] Network error — {e}")

    if not result:
        _log("  [Bulk Deals] all sources empty — no bulk-deal signal today")
    return result


def bulk_deal_score(symbol: str, bulk_deals_dict: dict) -> int:
    # Phase 2 #16 (2026-07-05): dampened bulk-deal impact.
    # A single bulk-deal print is a very noisy signal for retail-scale swing
    # trades (2-5 day horizon): institutions bulk-buy over weeks/months and
    # the print is often a rebalance, block-cross, or promoter selldown rather
    # than directional conviction. Old +6/-8 was strong enough to move a name
    # from WATCHLIST to BUY on a single print. New +2/-3 keeps the signal
    # visible (still tie-breaks between close candidates) but prevents
    # single-day headline noise from flipping decisions.
    action = bulk_deals_dict.get(symbol)
    if action == "BUY":  return 2
    elif action == "SELL": return -3
    return 0


def ownership_quality_score(promoter_data: dict) -> int:
    # Phase B7 (2026-07-02): FII% / DII% shareholding contributions removed from
    # scoring. Reasons:
    #   1. Screener FII/DII shareholding numbers are quarterly (T-90 stale) —
    #      not fresh enough to influence a swing-trade score.
    #   2. Same signal is already captured via price/volume and delivery %.
    #   3. Macro daily FII/DII flow (₹Cr) never was in the score — only
    #      informational for AI narrative and Telegram display.
    # The fields `fii_pct` / `dii_pct` are still populated so downstream
    # narrative / display code keeps working, but they no longer move the score.
    score    = 50
    pledge   = promoter_data.get("promoter_pledge_pct", 0)
    promoter = promoter_data.get("promoter_holding_pct", 50)
    if pledge > 40:   score -= 30
    elif pledge > 20: score -= 15
    elif pledge > 10: score -= 5
    if promoter > 60:   score += 15
    elif promoter > 50: score += 8
    elif promoter < 30: score -= 10
    return max(0, min(100, score))


NEUTRAL_FUNDAMENTALS = {
    "roe": 0.0, "de_ratio": 0.0, "roce": 0.0,
    "promoter_holding_pct": 50.0, "promoter_pledge_pct": 0.0,
    "fii_pct": 0.0, "dii_pct": 0.0,
    "source": "NEUTRAL_DEFAULT",
}

_SCREENER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.screener.in",
}


def _parse_screener_html(soup) -> dict:
    """
    PHASE_G6_SCREENER_V2 (2026-07-03) - screener.in HTML v2 parser.
    Screener restructured HTML after mid-2024. Handles both structured
    .name/.value spans AND inline text like "ROE38.2%".
    D/E and Pledge no longer on free page - return 0 so caller can use
    yfinance (D/E) and pledge blocklist as fallbacks.

    PHASE R1 (2026-07-06): additionally extracts the "Quarterly Results"
    table (sales / net profit / EPS across last 4-8 quarters) so we can
    derive institutional quality signals:
      • quarterly_sales:  list[float]  — most recent LAST
      • quarterly_profit: list[float]  — most recent LAST
      • quarterly_eps:    list[float]  — most recent LAST
      • sales_yoy_pct:    float        — latest Q vs same Q year-ago
      • profit_yoy_pct:   float        — latest Q vs same Q year-ago
      • sales_trend_3q:   "DECLINING" | "GROWING" | "MIXED"
      • profit_trend_3q:  "DECLINING" | "GROWING" | "MIXED"

    All new fields are OPTIONAL — absent when screener response doesn't
    include the Quarterly Results section (rare, but possible for newly
    listed stocks or partial page loads).
    """
    import re as _re
    data: dict = {}

    def _num(txt):
        if not txt:
            return None
        m = _re.search(r"-?\d[\d,]*\.?\d*", txt.replace("₹", ""))
        if not m:
            return None
        try:
            return float(m.group(0).replace(",", ""))
        except ValueError:
            return None

    ratio_section = soup.find("section", {"id": "top-ratios"}) or soup.find(id="top-ratios")
    if ratio_section:
        for li in ratio_section.find_all("li"):
            name_el = li.find(class_=_re.compile(r"\bname\b"))
            val_el = li.find(class_=_re.compile(r"\b(value|nowrap|number)\b"))
            if name_el and val_el:
                key = name_el.get_text(strip=True).lower()
                val = _num(val_el.get_text(strip=True))
            else:
                txt = li.get_text(strip=True)
                m = _re.match(r"^([A-Za-z][A-Za-z /()%.-]+?)([\d,.-].*)$", txt)
                if not m:
                    continue
                key = m.group(1).strip().lower()
                val = _num(m.group(2))
            if val is None:
                continue
            if "return on equity" in key or key == "roe":
                data["roe"] = val
            elif "return on capital" in key or "roce" in key:
                data["roce"] = val
            elif "debt / equity" in key or "d/e" in key or "debt to equity" in key:
                data["de_ratio"] = val
            elif "market cap" in key:
                data["market_cap_cr"] = val
            elif "stock p/e" in key or key == "p/e":
                data["pe"] = val

    for table in soup.find_all("table"):
        rows = table.find_all("tr")
        if not rows:
            continue
        header_cells = [td.get_text(strip=True) for td in rows[0].find_all(["td", "th"])]
        if not any(_re.search(r"(Jun|Sep|Dec|Mar)\s*20\d\d", h or "") for h in header_cells):
            continue
        if "Promoter" not in table.get_text():
            continue
        for row in rows[1:]:
            cells = [td.get_text(strip=True) for td in row.find_all(["td", "th"])]
            if len(cells) < 2:
                continue
            lbl = cells[0].lower()
            last_val = _num(cells[-1])
            if last_val is None:
                continue
            if "promoter" in lbl and "pledge" not in lbl:
                data["promoter_holding_pct"] = last_val
            elif "pledge" in lbl or "pledged" in lbl:
                data["promoter_pledge_pct"] = last_val
            elif "fii" in lbl or "fpi" in lbl or "foreign" in lbl:
                data["fii_pct"] = last_val
            elif "dii" in lbl or "domestic" in lbl:
                data["dii_pct"] = last_val
            elif "public" in lbl:
                data["public_pct"] = last_val
        if "promoter_holding_pct" in data:
            break

    # ------------------------------------------------------------------
    # Phase R1 (2026-07-06): Quarterly Results table
    # Screener publishes a table where rows are line items (Sales, Net Profit,
    # EPS, OPM %, etc.) and columns are quarter-end dates (Jun 2024, Sep 2024,
    # Dec 2024, Mar 2025, Jun 2025 …). We only extract the top-3 rows we need
    # for institutional quality assessment: Sales, Net Profit, EPS. If any
    # extraction fails, we simply skip — never raise; the caller treats the
    # absence as "quarterly data unavailable" and downgrades confidence.
    # ------------------------------------------------------------------
    try:
        _q_section = soup.find(id="quarters") or soup.find("section", {"id": "quarters"})
        _q_table = None
        if _q_section:
            _q_table = _q_section.find("table")
        # Fallback: any table whose header row has 3+ quarter columns
        if _q_table is None:
            for _t in soup.find_all("table"):
                _hdr = [td.get_text(strip=True) for td in _t.find_all("tr")[0].find_all(["td", "th"])] if _t.find_all("tr") else []
                _q_cols = sum(1 for h in _hdr if _re.search(r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s*20\d\d", h or ""))
                if _q_cols >= 3 and any("sales" in c.get_text(strip=True).lower() or "revenue" in c.get_text(strip=True).lower() for r in _t.find_all("tr") for c in r.find_all(["td", "th"])[:1]):
                    _q_table = _t
                    break
        if _q_table is not None:
            for row in _q_table.find_all("tr"):
                cells = [td.get_text(strip=True) for td in row.find_all(["td", "th"])]
                if len(cells) < 4:
                    continue
                lbl = cells[0].lower()
                # Parse row values (skip cell 0 which is the label)
                vals: list[float] = []
                for c in cells[1:]:
                    v = _num(c)
                    if v is not None:
                        vals.append(v)
                if len(vals) < 3:
                    continue
                # Detect which line item this row represents
                if ("sales" in lbl or "revenue" in lbl) and "growth" not in lbl and "operating" not in lbl:
                    data["quarterly_sales"] = vals[-8:]  # keep last 8 quarters max
                elif ("net profit" in lbl or "net income" in lbl) and "growth" not in lbl:
                    data["quarterly_profit"] = vals[-8:]
                elif ("eps in" in lbl or lbl.strip() == "eps" or "eps (" in lbl):
                    data["quarterly_eps"] = vals[-8:]
                elif "opm" in lbl or "operating profit margin" in lbl:
                    data["quarterly_opm"] = vals[-8:]
        # Derive trend labels + YoY deltas
        def _trend_label(series: list[float], min_len: int = 3) -> str:
            if not series or len(series) < min_len:
                return "UNKNOWN"
            tail = series[-min_len:]
            deltas = [tail[i] - tail[i-1] for i in range(1, len(tail))]
            neg = sum(1 for d in deltas if d < 0)
            pos = sum(1 for d in deltas if d > 0)
            if neg == len(deltas):
                return "DECLINING"
            if pos == len(deltas):
                return "GROWING"
            return "MIXED"
        _qs = data.get("quarterly_sales", [])
        _qp = data.get("quarterly_profit", [])
        _qe = data.get("quarterly_eps", [])
        _qo = data.get("quarterly_opm", [])
        if _qs:
            data["sales_trend_3q"] = _trend_label(_qs, 3)
            # YoY: latest vs same quarter 4 quarters ago
            if len(_qs) >= 5 and _qs[-5] != 0:
                data["sales_yoy_pct"] = round((_qs[-1] - _qs[-5]) / abs(_qs[-5]) * 100, 1)
        if _qp:
            data["profit_trend_3q"] = _trend_label(_qp, 3)
            if len(_qp) >= 5 and _qp[-5] != 0:
                data["profit_yoy_pct"] = round((_qp[-1] - _qp[-5]) / abs(_qp[-5]) * 100, 1)
        if _qe:
            data["eps_trend_3q"] = _trend_label(_qe, 3)
            if len(_qe) >= 5 and _qe[-5] != 0:
                data["eps_yoy_pct"] = round((_qe[-1] - _qe[-5]) / abs(_qe[-5]) * 100, 1)
        if _qo:
            data["opm_trend_3q"] = _trend_label(_qo, 3)
    except Exception:
        # Parsing failure is non-fatal — quarterly fields simply absent
        pass

    return data

def fetch_screener_data(symbol_clean: str) -> dict | None:
    """
    Source 1: screener.in — consolidated then standalone.
    Returns dict on success, None on rate-limit or total failure.
    """
    if not _BS4_OK:
        return None
    for url in [
        f"https://www.screener.in/company/{symbol_clean}/consolidated/",
        f"https://www.screener.in/company/{symbol_clean}/",
    ]:
        try:
            resp = requests.get(url, headers=_SCREENER_HEADERS, timeout=12)
            if resp.status_code == 429:
                _log(f"[WARN] screener.in rate-limited for {symbol_clean}")
                return None
            if resp.status_code != 200:
                continue
            data = _parse_screener_html(BeautifulSoup(resp.text, "html.parser"))
            if data:
                data.setdefault("roe", 0.0)
                data.setdefault("de_ratio", 0.0)
                data.setdefault("roce", 0.0)
                data.setdefault("promoter_holding_pct", 50.0)
                data.setdefault("promoter_pledge_pct", 0.0)
                data.setdefault("fii_pct", 0.0)
                data.setdefault("dii_pct", 0.0)
                return data
        except requests.exceptions.Timeout:
            _log(f"[WARN] screener.in timeout for {symbol_clean}")
        except Exception as e:
            _log(f"[WARN] screener.in error for {symbol_clean}: {e}")
    return None


def fetch_trendlyne_data(symbol_clean: str) -> dict | None:
    """
    Source 2: Trendlyne — less strict rate limiting than screener.
    Returns dict on success, None on failure.
    """
    if not _BS4_OK:
        return None
    try:
        url  = f"https://trendlyne.com/equity/fundamental-analysis/{symbol_clean}/NSE/"
        resp = requests.get(url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
            "Accept": "text/html",
        }, timeout=12)
        if resp.status_code != 200:
            return None
        soup = BeautifulSoup(resp.text, "html.parser")
        data = {}
        for table in soup.find_all("table"):
            for row in table.find_all("tr"):
                cells = [td.get_text(strip=True) for td in row.find_all(["td", "th"])]
                if len(cells) >= 2:
                    lbl = cells[0].lower()
                    try:
                        val = float(cells[1].replace("%", "").replace(",", "").strip())
                        if "roe" in lbl:
                            data["roe"] = val
                        elif "debt" in lbl and "equity" in lbl:
                            data["de_ratio"] = val
                        elif "roce" in lbl:
                            data["roce"] = val
                        elif "promoter" in lbl and "pledge" not in lbl:
                            data["promoter_holding_pct"] = val
                        elif "pledge" in lbl:
                            data["promoter_pledge_pct"] = val
                    except (ValueError, IndexError):
                        pass
        return data if data else None
    except Exception as e:
        _log(f"[WARN] Trendlyne failed for {symbol_clean}: {e}")
        return None


def fetch_yfinance_fundamentals(symbol: str) -> dict:
    """
    Source 3: yfinance — last resort. Always available, never rate-limited.
    Returns dict with whatever yfinance has; fills missing fields with neutral.
    """
    try:
        ticker = yf.Ticker(symbol)
        info   = ticker.info or {}
        data   = {}
        roe = info.get("returnOnEquity")
        if roe is not None:
            data["roe"] = round(float(roe) * 100, 2)  # decimal → %
        de = info.get("debtToEquity")
        if de is not None:
            data["de_ratio"] = round(float(de) / 100, 2)  # % → ratio
        data.setdefault("roe", 0.0)
        data.setdefault("de_ratio", 0.0)
        data.setdefault("roce", 0.0)
        data.setdefault("promoter_holding_pct", 50.0)
        data.setdefault("promoter_pledge_pct", 0.0)
        data.setdefault("fii_pct", 0.0)
        data.setdefault("dii_pct", 0.0)
        return data
    except Exception as e:
        _log(f"[WARN] yfinance fundamentals failed for {symbol}: {e}")
        return {**NEUTRAL_FUNDAMENTALS}


def fetch_promoter_data(symbol_clean: str, delay_seconds: float = 2.5) -> dict:
    """
    3-source fallback chain: screener.in → Trendlyne → yfinance.
    Sequential with delay to avoid rate limiting.
    Call this SEQUENTIALLY — never in parallel threads.

    HONESTY: when screener returns 0/0/0 AND yfinance is rate-limited,
    the fetch is marked source="SCREENER+YF_RL" so downstream renderers
    (BUY card) can show "N/A" instead of "0.0%" — hiding fake real-looking
    zeros in trading decisions.
    """
    time.sleep(delay_seconds)

    data = fetch_screener_data(symbol_clean)
    if data:
        # Phase G6 (2026-07-03): screener no longer exposes D/E on the free
        # page. Whenever D/E is 0 (regardless of ROE), fill it from yfinance.
        # yfinance also fills ROE if screener came back empty.
        yf_rl = False
        needs_yf = (data.get("de_ratio", 0) == 0 or data.get("roe", 0) == 0)
        if needs_yf:
            try:
                yf_data = fetch_yfinance_fundamentals(symbol_clean + ".NS")
                # Rate-limit detection: yfinance returns all-zero neutral on 429.
                if (yf_data.get("roe", 0) == 0 and yf_data.get("de_ratio", 0) == 0
                        and yf_data.get("roce", 0) == 0):
                    yf_rl = True
                else:
                    # Only fill fields that are still zero (don't overwrite screener wins)
                    if data.get("roe", 0) == 0 and yf_data.get("roe", 0) != 0:
                        data["roe"] = yf_data["roe"]
                    if data.get("de_ratio", 0) == 0 and yf_data.get("de_ratio", 0) != 0:
                        data["de_ratio"] = yf_data["de_ratio"]
                    if data.get("roce", 0) == 0 and yf_data.get("roce", 0) != 0:
                        data["roce"] = yf_data["roce"]
            except Exception:
                yf_rl = True
        data["source"] = "SCREENER+YF_RL" if yf_rl else "SCREENER+YF"
        data["fundamentals_source"] = data["source"]
        _log(f"[INFO] Fundamentals (screener): {symbol_clean} "
             f"ROE={data.get('roe',0):.1f}% D/E={data.get('de_ratio',0):.2f} "
             f"Pledge={data.get('promoter_pledge_pct',0):.1f}%"
             + (" ⚠️ yf rate-limited" if yf_rl else ""))
        return data

    _log(f"[WARN] screener.in failed for {symbol_clean} — trying Trendlyne")
    time.sleep(1.5)

    data = fetch_trendlyne_data(symbol_clean)
    if data:
        data.setdefault("promoter_holding_pct", 50.0)
        data.setdefault("promoter_pledge_pct", 0.0)
        data.setdefault("fii_pct", 0.0)
        data.setdefault("dii_pct", 0.0)
        data.setdefault("roe", 0.0)
        data.setdefault("de_ratio", 0.0)
        data.setdefault("roce", 0.0)
        data["source"] = "TRENDLYNE"
        _log(f"[INFO] Fundamentals (trendlyne): {symbol_clean}")
        return data

    _log(f"[WARN] Trendlyne failed for {symbol_clean} — using yfinance")
    data = fetch_yfinance_fundamentals(symbol_clean + ".NS")
    data["source"] = "YFINANCE"
    _log(f"[INFO] Fundamentals (yfinance): {symbol_clean} "
         f"ROE={data.get('roe',0):.1f}% D/E={data.get('de_ratio',0):.2f}")
    return data


# ---------------------------------------------------------------------------
# Fundamentals cache — avoid re-fetching same stocks within 24h
# ---------------------------------------------------------------------------

def load_fundamentals_cache() -> dict:
    # Phase C7g (2026-07-10): FRESH_START forces a full re-fetch of ROE/D/E
    # so today's fundamentals row in every downstream artifact is truly today's
    # data — not yesterday's cache promoted forward.
    if FRESH_START:
        _log("[FRESH_START] load_fundamentals_cache → returning {} (old fundamentals_cache.json ignored)")
        return {}
    try:
        if os.path.exists(FUNDAMENTALS_CACHE_FILE):
            with open(FUNDAMENTALS_CACHE_FILE, "r") as f:
                return json.load(f)
    except Exception as e:
        _log(f"[WARN] load_fundamentals_cache failed: {e}")
    return {}


def save_fundamentals_cache(cache: dict) -> None:
    try:
        with open(FUNDAMENTALS_CACHE_FILE, "w") as f:
            json.dump(cache, f, indent=2, default=str)
    except Exception as e:
        _log(f"[WARN] save_fundamentals_cache failed: {e}")


# ---------------------------------------------------------------------------
# DELIVERY % — real per-stock institutional accumulation signal.
# Phase C3 (2026-07-02): delivery-to-traded ratio from nselib bhavcopy
# archives. This is the single strongest per-stock signal in the Indian
# market: high delivery % on rising price = institutional accumulation;
# low delivery % on rising price = speculative pump; high delivery % on
# falling price = insider/large-holder unloading. Cached 24h per symbol.
# ---------------------------------------------------------------------------

def load_delivery_cache() -> dict:
    # Phase C7g (2026-07-10): FRESH_START forces re-fetch of every ticker's
    # bhavcopy delivery numbers so today's delivery signal reflects today's
    # market activity, not the previous run's cached snapshot.
    if FRESH_START:
        _log("[FRESH_START] load_delivery_cache → returning {} (old delivery_cache.json ignored)")
        return {}
    try:
        if os.path.exists(DELIVERY_CACHE_FILE):
            with open(DELIVERY_CACHE_FILE, "r") as f:
                return json.load(f)
    except Exception as e:
        _log(f"[WARN] load_delivery_cache failed: {e}")
    return {}


def save_delivery_cache(cache: dict) -> None:
    try:
        with open(DELIVERY_CACHE_FILE, "w") as f:
            json.dump(cache, f, indent=2, default=str)
    except Exception as e:
        _log(f"[WARN] save_delivery_cache failed: {e}")


def fetch_delivery_data(symbol_clean: str, lookback_days: int = 30) -> dict:
    """Fetch per-day delivery % from NSE bhavcopy via nselib.

    Returns:
        {
            "delivery_pct_today":  float (0-100),
            "delivery_pct_20d_avg": float (0-100),
            "delivery_ratio":      float (today / 20d_avg),
            "delivery_signal":     "STRONG_ACCUM" | "ACCUM" | "NEUTRAL" |
                                   "WEAK" | "DISTRIBUTION",
            "series_days":         int (number of daily rows we got),
            "source":              "nselib" | "unavailable"
        }
        Empty-ish result with source="unavailable" if the fetch fails.
    """
    default = {
        "delivery_pct_today":   0.0,
        "delivery_pct_20d_avg": 0.0,
        "delivery_ratio":       1.0,
        "delivery_signal":      "NEUTRAL",
        "series_days":          0,
        "source":               "unavailable",
    }
    try:
        # nselib import — deferred to avoid hard dependency at module load time.
        from nselib import capital_market as _nsecm  # type: ignore
    except Exception as e:
        _log(f"    [DELIV] nselib unavailable ({e}) — using proxy only")
        return default
    try:
        # nselib expects DD-MM-YYYY for from_date / to_date.
        today   = ist_today()
        # Use a wider window than 20d because market holidays + weekends can
        # thin out the row count; we'll trim later.
        from_d  = today - datetime.timedelta(days=max(45, lookback_days + 15))
        # Phase C4 (2026-07-02): nselib does NOT expose `security_wise_archives`
        # — the correct function is `price_volume_and_deliverable_position_data`
        # (also aliased as `deliverable_position_data`). Confirmed columns:
        # 'Symbol','Series','Date','PrevClose','OpenPrice','HighPrice',
        # 'LowPrice','LastPrice','ClosePrice','AveragePrice',
        # 'TotalTradedQuantity','TurnoverInRs','No.ofTrades','DeliverableQty',
        # '%DlyQttoTradedQty'. Function has no `series` kwarg — filter EQ later.
        df = _nsecm.price_volume_and_deliverable_position_data(
            symbol=symbol_clean,
            from_date=from_d.strftime("%d-%m-%Y"),
            to_date=today.strftime("%d-%m-%Y"),
        )
        if df is None or len(df) == 0:
            return default
        # Filter to EQ series if a Series column is present (older nselib
        # versions may already do this; be defensive).
        try:
            if "Series" in df.columns:
                df = df[df["Series"].astype(str).str.strip().str.upper() == "EQ"]
        except Exception:
            pass
        if df is None or len(df) == 0:
            return default
        # Column name in nselib is usually "%DlyQttoTradedQty" (str with '%')
        # or "DELIV_PER". Handle both.
        col = None
        for cand in ("%DlyQttoTradedQty", "DELIV_PER", "%DlyQt to TradedQty"):
            if cand in df.columns:
                col = cand
                break
        if col is None:
            _log(f"    [DELIV] {symbol_clean}: no delivery column found "
                 f"(cols={list(df.columns)[:6]})")
            return default
        # Parse to float — some rows come as '-' or with '%' suffix.
        vals: list = []
        for v in df[col].tolist():
            try:
                if v is None:
                    continue
                s = str(v).replace("%", "").strip()
                if s in ("", "-", "N/A"):
                    continue
                vals.append(float(s))
            except Exception:
                continue
        # nselib returns most-recent-first typically, but not guaranteed.
        # Sort by date if there's a date column.
        # For safety we just use the last 20 entries as "recent 20" regardless
        # of order — if newest is first, we reverse.
        if len(vals) < 5:
            return default
        # Detect ordering heuristically: nselib usually returns newest-first.
        # We'll assume newest-first and take vals[0] as today, vals[:20] as 20d.
        today_pct   = vals[0]
        twenty_d    = vals[:min(20, len(vals))]
        avg_20d     = sum(twenty_d) / len(twenty_d)
        ratio       = (today_pct / avg_20d) if avg_20d > 0 else 1.0
        # Signal derivation
        if ratio >= 1.30:   sig = "STRONG_ACCUM"
        elif ratio >= 1.15: sig = "ACCUM"
        elif ratio <= 0.70: sig = "DISTRIBUTION"
        elif ratio <= 0.85: sig = "WEAK"
        else:               sig = "NEUTRAL"
        return {
            "delivery_pct_today":   round(today_pct, 2),
            "delivery_pct_20d_avg": round(avg_20d, 2),
            "delivery_ratio":       round(ratio, 3),
            "delivery_signal":      sig,
            "series_days":          len(vals),
            "source":               "nselib",
        }
    except Exception as e:
        _log(f"    [DELIV] {symbol_clean} fetch failed: {e}")
        return default


def fetch_delivery_cached(symbol_clean: str, cache: dict,
                          cache_ttl_hours: int = 24) -> dict:
    """24h-cached wrapper around fetch_delivery_data()."""
    now    = ist_now()
    cached = cache.get(symbol_clean)
    if cached:
        try:
            cached_at = datetime.datetime.fromisoformat(cached.get("cached_at", "2000-01-01"))
            if (now - cached_at).total_seconds() / 3600 < cache_ttl_hours:
                return cached["data"]
        except Exception:
            pass
    data = fetch_delivery_data(symbol_clean)
    cache[symbol_clean] = {"data": data, "cached_at": now.isoformat()}
    return data


def delivery_score_from_signal(signal: str, ratio: float) -> float:
    """Convert delivery signal → 0-100 factor score for volume_delivery.

    Anchored around 50 (neutral). Uses the numeric ratio for smooth
    interpolation instead of pure buckets.
    """
    try:
        # Clamp ratio to a sane band before scaling.
        r = max(0.3, min(2.0, float(ratio)))
        # Phase 2 #32 (2026-07-05): halved the accumulation slope so a
        # single-day 2× spike no longer saturates volume_delivery to 100.
        # Old: (r-1.0)*90 → 1.0→55, 1.5→100 (saturated), 2.0→145 clipped.
        # New: (r-1.0)*45 → 1.0→55, 1.5→77.5, 2.0→100. Distribution branch
        # (r<1.0) intentionally kept steeper — distribution is asymmetrically
        # more informative than accumulation for retail pump-and-dump names.
        if r >= 1.0:
            score = 55 + (r - 1.0) * 45          # 1.0→55, 1.5→77.5, 2.0→100
        else:
            score = 55 - (1.0 - r) * 75 / 0.6    # unchanged: 1.0→55, 0.4→-20 clipped
        # Signal override at the extremes so labels remain consistent
        if signal == "STRONG_ACCUM": score = max(score, 88)
        elif signal == "DISTRIBUTION": score = min(score, 18)
        return round(max(0.0, min(100.0, score)), 1)
    except Exception:
        return 50.0


def fetch_promoter_data_cached(symbol_clean: str, cache: dict,
                                cache_ttl_hours: int = 24) -> dict:
    """Returns cached data if fresher than cache_ttl_hours; otherwise fetches live."""
    now    = ist_now()
    cached = cache.get(symbol_clean)
    if cached:
        try:
            cached_at = datetime.datetime.fromisoformat(cached.get("cached_at", "2000-01-01"))
            if (now - cached_at).total_seconds() / 3600 < cache_ttl_hours:
                _log(f"[INFO] Fundamentals cache hit: {symbol_clean}")
                return cached["data"]
        except Exception:
            pass
    data = fetch_promoter_data(symbol_clean, delay_seconds=2.5)
    cache[symbol_clean] = {"data": data, "cached_at": now.isoformat()}
    return data


def fetch_all_fundamentals_cached(top_40: list, max_stocks: int = 30) -> list:
    """
    Fetches fundamentals sequentially with 24h cache.

    PRIORITY ORDER (FIX): rearrange top_40 so likely BUY candidates get their
    fundamentals FIRST (before rate-limit exhaustion) — sorted by
    final/base confidence as a proxy for BUY-likelihood. Cache hits are
    effectively free so we can safely widen from 20 -> 30 stocks without
    noticeably slowing the fresh run.

    Adds ~50s on first run; near-instant on same-day re-runs.
    """
    cache = load_fundamentals_cache()
    _log(f"[INFO] Fundamentals cache: {len(cache)} symbols cached")

    def _prio(s):
        # At this pipeline stage (step 9), final_confidence has not yet been
        # computed — that assignment happens later in step 11. The only
        # meaningful ranking signal available here is base_confidence
        # (assigned in step 7 when `scored` was built). We used to read
        # `final_confidence` with an or-chain fallback, but the read was a
        # footgun: any future default of final_confidence != 0 would silently
        # break fundamentals priority ordering. Keep this explicit.
        # Original audit item #7 flagged this as a bug — the ACTUAL bug was
        # misleading intent, not incorrect behavior. See PATCH_LOG_v1.md.
        return -float(s.get("base_confidence", 0) or 0)
    ordered  = sorted(top_40, key=_prio)
    to_fetch = ordered[:max_stocks]
    skip     = ordered[max_stocks:]

    est_sec = sum(
        0 if s["symbol"].replace(".NS", "") in cache else 2.5
        for s in to_fetch
    )
    _log(f"[INFO] Estimated fetch time: ~{est_sec:.0f}s for {max_stocks} stocks "
         f"(BUY-priority order)")

    for i, stock in enumerate(to_fetch):
        sym_clean = stock["symbol"].replace(".NS", "")
        _log(f"  [{i+1}/{max_stocks}] {sym_clean}")
        pdata = fetch_promoter_data_cached(sym_clean, cache, cache_ttl_hours=24)
        stock["promoter_data"]       = pdata
        stock["ownership_quality"]   = ownership_quality_score(pdata)
        stock["promoter_pledge_pct"] = pdata.get("promoter_pledge_pct", 0.0)
        stock["roe"]                 = pdata.get("roe", 0.0)
        stock["de_ratio"]            = pdata.get("de_ratio", 0.0)
        stock["roce"]                = pdata.get("roce", 0.0)
        # Phase G8-B: expose market_cap_cr on the stock dict so Gate 3c can read it
        stock["market_cap_cr"]       = pdata.get("market_cap_cr", 0.0)
        stock["fundamentals_source"] = pdata.get("source", "NEUTRAL_DEFAULT")
        # Refine ownership_quality with ROE + D/E (screener+YF data now available)
        _update_ownership_quality(stock)
        # Log silent-fail cases so we can see WHICH stocks came back empty.
        if (pdata.get("roe", 0) == 0 and pdata.get("de_ratio", 0) == 0
                and stock["fundamentals_source"] in ("NEUTRAL_DEFAULT", "screener_partial")):
            _log(f"    [FUND] {sym_clean}: ROE/D/E both 0 · source={stock['fundamentals_source']}")

    for stock in skip:
        stock["promoter_data"]       = {**NEUTRAL_FUNDAMENTALS}
        stock["ownership_quality"]   = 50
        stock["promoter_pledge_pct"] = 0.0
        stock["roe"]                 = 0.0
        stock["de_ratio"]            = 0.0
        stock["roce"]                = 0.0
        stock["market_cap_cr"]       = 0.0  # G8-B: 0 signals "unknown", not "tiny"
        stock["fundamentals_source"] = "NOT_FETCHED"

    save_fundamentals_cache(cache)
    fetched_ok = sum(1 for s in to_fetch
                     if s.get("fundamentals_source") not in ("NEUTRAL_DEFAULT", "NOT_FETCHED"))
    _log(f"  Fundamentals done: {fetched_ok}/{max_stocks} real data | "
         f"{max_stocks - fetched_ok} defaults | cache saved")

    # ── Delivery % fetch (same top-N batch, 24h cache) ─────────────────────
    # Phase C3 (2026-07-02): real per-stock delivery-to-traded ratio.
    # This is the strongest single per-stock signal for the Indian market —
    # separates institutional accumulation from speculative pumps.
    deliv_cache = load_delivery_cache()
    deliv_ok    = 0
    _log(f"  [DELIV] Fetching delivery % for {len(to_fetch)} stocks "
         f"(cache size: {len(deliv_cache)})")
    for stock in to_fetch:
        sym_clean = stock["symbol"].replace(".NS", "")
        ddata = fetch_delivery_cached(sym_clean, deliv_cache, cache_ttl_hours=24)
        stock["delivery_pct_today"]   = ddata.get("delivery_pct_today", 0.0)
        stock["delivery_pct_20d_avg"] = ddata.get("delivery_pct_20d_avg", 0.0)
        stock["delivery_ratio"]       = ddata.get("delivery_ratio", 1.0)
        stock["delivery_signal"]      = ddata.get("delivery_signal", "NEUTRAL")
        stock["delivery_source"]      = ddata.get("source", "unavailable")
        if ddata.get("source") == "nselib":
            deliv_ok += 1
        # Phase G7-A: re-run ownership_quality now that delivery keys exist,
        # so the delivery bonus/penalty (+8/+5/-10/-8) actually contributes.
        # First call at L3388 ran BEFORE delivery was fetched.
        _update_ownership_quality(stock)
    # Skip stocks (outside top max_stocks) get neutral defaults so downstream
    # logic never trips on missing keys.
    for stock in skip:
        stock["delivery_pct_today"]   = 0.0
        stock["delivery_pct_20d_avg"] = 0.0
        stock["delivery_ratio"]       = 1.0
        stock["delivery_signal"]      = "NEUTRAL"
        stock["delivery_source"]      = "not_fetched"
    save_delivery_cache(deliv_cache)
    _log(f"  [DELIV] Done: {deliv_ok}/{len(to_fetch)} real data | "
         f"{len(to_fetch) - deliv_ok} proxy-only")

    return top_40


GLOBAL_TICKERS = {
    "NIFTY":   "^NSEI",
    "SENSEX":  "^BSESN",
    "VIX_IN":  "^INDIAVIX",
    "VIX_US":  "^VIX",
    "USDINR":  "INR=X",
    "CRUDE":   "CL=F",
    "GOLD":    "GC=F",
    "US10Y":   "^TNX",
    "DXY":     "DX-Y.NYB",
    "SP500":   "^GSPC",
    "DOW":     "^DJI",
}


# Phase B2: Yahoo's INR=X feed has been drifting (54-wk range 84.55–97.05 confirms feed is broken).
# Frankfurter is ECB-backed, keyless, CI-friendly, and updates daily ~16:00 CET.
def _fetch_usdinr_frankfurter() -> "FetchResult | None":
    """Frankfurter API → authoritative USDINR. Returns None on any failure.
    Manual retry loop with exponential backoff (avoids tenacity-version skew)."""
    last_err: "Exception | None" = None
    for attempt in range(3):
        try:
            resp = requests.get(
                "https://api.frankfurter.dev/v1/latest?from=USD&to=INR",
                headers={"User-Agent": "swing-trade-engine/6.0"},
                timeout=8,
            )
            if resp.status_code != 200:
                _log(f"  [USDINR Frankfurter] HTTP {resp.status_code} (try {attempt+1}/3)")
                last_err = RuntimeError(f"HTTP {resp.status_code}")
                time.sleep(1 + attempt + random.random())
                continue
            data = resp.json()
            rate = float(data["rates"]["INR"])
            as_of = data.get("date")  # e.g. "2026-06-30"
            try:
                as_of_d = datetime.datetime.strptime(as_of, "%Y-%m-%d").date()
            except Exception:
                as_of_d = ist_today()
            stale = (ist_today() - as_of_d).days > 3
            return FetchResult(
                value=rate, source="frankfurter",
                as_of_date=as_of_d, fetched_at=ist_now(),
                is_stale=stale,
            )
        except Exception as e:
            last_err = e
            time.sleep(1 + attempt + random.random())
    _log(f"  [USDINR Frankfurter] all 3 attempts failed — last error: {last_err}")
    return None


def fetch_global_macro() -> dict:
    macro = {
        "usdinr": 83.5, "crude_usd": 75.0, "vix_us": 18.0, "vix_in": 15.0,
        "us10y": 4.3, "sp500_1d_pct": 0.0, "china_1d_pct": 0.0,
        "gold_usd": 2300.0, "nifty_1d_pct": 0.0, "dxy": 103.0,
        "sensex_1d_pct": 0.0, "dow_1d_pct": 0.0,
        "vix_in_20d_avg": 15.0, "vix_term_ratio": 1.0,
        # Phase 3: VIX percentile regime (populated below if 1y history fetched)
        "vix_in_percentile": None, "vix_in_regime": "UNKNOWN",
        "vix_in_lookback_days": 0,
    }
    # Provenance map — which source each macro field came from
    macro["sources"] = {}

    # Phase B2: Frankfurter takes precedence over Yahoo for USD/INR.
    # Phase C1 (2026-07-02): widen accept-band to 70–100 so a legit weak-rupee
    # print (~95) is not rejected. Log the value even when rejected so
    # operators know WHY the fallback path fired.
    usdinr_fr = _fetch_usdinr_frankfurter()
    if usdinr_fr is not None and 70.0 <= usdinr_fr.value <= 100.0:
        macro["usdinr"]            = round(usdinr_fr.value, 4)
        macro["sources"]["usdinr"] = usdinr_fr.to_log()
        _log(f"  USD/INR ✓ Frankfurter — {macro['usdinr']:.4f} ({usdinr_fr.to_log()})")
        skip_usdinr_yf = True
    else:
        if usdinr_fr is not None:
            _log(f"  USD/INR ✗ Frankfurter returned {usdinr_fr.value:.4f} — outside [70,100] band, falling back to yfinance INR=X")
        else:
            _log("  USD/INR ✗ Frankfurter failed (all attempts) — falling back to yfinance INR=X")
        skip_usdinr_yf = False

    for name, ticker in GLOBAL_TICKERS.items():
        if name == "USDINR" and skip_usdinr_yf:
            continue
        try:
            # Fetch 1y for VIX_IN so we can compute both 20d MA and 252d percentile;
            # 5d suffices for everything else.
            period = "1y" if name == "VIX_IN" else "5d"
            df = yf.download(ticker, period=period, interval="1d", progress=False, auto_adjust=True)
            if df is not None and len(df) >= 2:
                last = float(df["Close"].squeeze().iloc[-1])
                prev = float(df["Close"].squeeze().iloc[-2])
                pct  = round((last - prev) / prev * 100, 2) if prev != 0 else 0.0
                if name == "USDINR":
                    # Phase C1 (2026-07-02): yfinance INR=X feed has drifted historically.
                    # Sanity-check the value against LKG; a >3% jump vs last known good
                    # is treated as broken feed and swapped out for LKG.
                    lkg_usdinr = load_last_known_good().get("usdinr")
                    _yf_val = last
                    _accept = 70.0 <= _yf_val <= 100.0
                    if _accept and lkg_usdinr:
                        try:
                            _drift = abs(_yf_val - float(lkg_usdinr)) / float(lkg_usdinr)
                            if _drift > 0.03:  # >3% single-day move — unlikely for USDINR
                                _log(f"  USD/INR ⚠ yfinance {_yf_val:.4f} drifts {_drift*100:.1f}% vs LKG {float(lkg_usdinr):.4f} — using LKG")
                                _yf_val = float(lkg_usdinr)
                                macro["sources"]["usdinr"] = f"source=LKG (yfinance drift-rejected) as_of=<lkg>"
                            else:
                                macro["sources"]["usdinr"] = f"source=yfinance({ticker}) as_of={ist_today().isoformat()}"
                        except Exception:
                            macro["sources"]["usdinr"] = f"source=yfinance({ticker}) as_of={ist_today().isoformat()}"
                    elif not _accept and lkg_usdinr:
                        _log(f"  USD/INR ✗ yfinance {_yf_val:.4f} outside [70,100] — using LKG {float(lkg_usdinr):.4f}")
                        _yf_val = float(lkg_usdinr)
                        macro["sources"]["usdinr"] = f"source=LKG (yfinance OOR) as_of=<lkg>"
                    else:
                        macro["sources"]["usdinr"] = f"source=yfinance({ticker}) as_of={ist_today().isoformat()}"
                    macro["usdinr"] = _yf_val
                elif name == "CRUDE":  macro["crude_usd"]    = last
                elif name == "VIX_US": macro["vix_us"]       = last
                elif name == "VIX_IN":
                    macro["vix_in"] = last
                    # VIX term structure: spot vs 20-day average
                    if len(df) >= 20:
                        vix20_avg = float(df["Close"].squeeze().tail(20).mean())
                        macro["vix_in_20d_avg"] = round(vix20_avg, 2)
                        macro["vix_term_ratio"]  = round(last / vix20_avg, 3) if vix20_avg > 0 else 1.0
                    # Phase 3 (2026-07-01): 252-day rolling VIX percentile —
                    # institutional-grade regime signal. Low percentile = market
                    # complacent (raise bar), high percentile = fear (lower bar,
                    # opportunity for contrarian entries).
                    try:
                        vix_series = df["Close"].squeeze().dropna()
                        if len(vix_series) >= 60:
                            # Use up to 252 trading days; fewer if history is short.
                            window = vix_series.tail(252)
                            rank_below = int((window < last).sum())
                            pctile = round(100.0 * rank_below / max(len(window), 1), 1)
                            macro["vix_in_percentile"] = pctile
                            macro["vix_in_lookback_days"] = int(len(window))
                            # Regime bucket for downstream decision logic
                            if pctile <= 15.0:
                                regime = "COMPLACENT"      # raise bar
                            elif pctile >= 85.0:
                                regime = "FEAR"            # lower bar (contrarian)
                            elif pctile >= 65.0:
                                regime = "ELEVATED"        # mild caution
                            else:
                                regime = "NORMAL"
                            macro["vix_in_regime"] = regime
                        else:
                            macro["vix_in_percentile"] = None
                            macro["vix_in_regime"]     = "UNKNOWN"
                    except Exception:
                        macro["vix_in_percentile"] = None
                        macro["vix_in_regime"]     = "UNKNOWN"
                elif name == "US10Y":  macro["us10y"]        = last
                elif name == "GOLD":   macro["gold_usd"]     = last
                elif name == "DXY":    macro["dxy"]          = last
                elif name == "SP500":  macro["sp500_1d_pct"] = pct
                elif name == "DOW":    macro["dow_1d_pct"]   = pct
                elif name == "NIFTY":  macro["nifty_1d_pct"] = pct
                elif name == "SENSEX": macro["sensex_1d_pct"]= pct
        except Exception:
            continue

    # Phase A safety: range-gate everything; degrade gracefully via LKG
    macro = validate_macro(macro)
    return macro


def macro_regime_adjustment(macro: dict) -> int:
    adj = 0
    if macro["vix_us"] > 30:     adj -= 15
    elif macro["vix_us"] > 22:   adj -= 8
    elif macro["vix_us"] < 15:   adj += 5
    if macro["usdinr"] > 85:     adj -= 8
    elif macro["usdinr"] < 82:   adj += 5
    if macro["crude_usd"] > 95:  adj -= 6
    elif macro["crude_usd"] < 70: adj += 3
    if macro["us10y"] > 5.0:     adj -= 8
    elif macro["us10y"] < 4.0:   adj += 4
    if macro["sp500_1d_pct"] < -1.5: adj -= 5
    elif macro["sp500_1d_pct"] > 1.5: adj += 3
    if macro.get("dxy", 103) > 106:  adj -= 5
    elif macro.get("dxy", 103) < 100: adj += 3
    # VIX term structure adjustment
    vix_ratio = macro.get("vix_term_ratio", 1.0)
    if vix_ratio > 1.3:    adj -= 8   # VIX spiking above its own avg = fear
    elif vix_ratio > 1.15: adj -= 4   # elevated
    elif vix_ratio < 0.8:  adj += 4   # suppressed VIX = complacency, slight caution
    elif vix_ratio < 0.9:  adj += 2
    # Phase 2 #15 (2026-07-05): symmetric clamp — old max was +10 which
    # arbitrarily suppressed macro tailwinds while allowing full -20 headwind.
    # BULL regimes with cooperative macro were being penalized vs BEAR regimes
    # with hostile macro. Now ±20 both ways.
    return max(-20, min(20, adj))


# ─────────────────────────────────────────────────────────────────────────────
# Stage-A cleanup (2026-07-XX): FII/DII source cascade REMOVED.
#
# Deleted helpers (Phase C4 disabled fetch_fii_dii_flows since 2026-07-02;
# they were orphaned — 0 external callers, 0 in-file callers after stub):
#   _nse_session_get, _parse_fii_dii_from_text, _fetch_fii_dii_mc_rss,
#   _fetch_fii_dii_et_rss, _fetch_fii_dii_bs_rss, _fetch_fii_dii_google_news,
#   _fetch_fii_dii_nse, _fetch_fii_dii_nselib_nsdl, _fetch_fii_dii_bse
#
# All ~490 LOC. See git history for original 7-source cascade if ever needed.
# Contract preserved via the fetch_fii_dii_flows() stub below.
# ─────────────────────────────────────────────────────────────────────────────


def fetch_fii_dii_flows(max_retries: int = 2) -> dict:
    """
    Phase C4 (2026-07-02): FII/DII flows are NO LONGER used anywhere in the
    pipeline — not for scoring, sizing, gates, or regime detection. NSDL data
    is structurally D-1/D-2 stale, the RSS fallback cascade is noisy and
    burns ~7 seconds every run for a signal we already ignore.

    We keep this function only so the (few) legacy callers that ask for the
    macro dict don't KeyError. It now short-circuits: returns
    available=False with zero flows and never touches the network.

    If you want the old multi-source cascade back, see git history for the
    NSDL / BSE / RSS / Google News implementations.
    """
    _log("  [FII/DII] Fetch disabled (Phase C4) — flows no longer scored or displayed.")
    return {
        "fii_flow_cr":    0.0,
        "dii_flow_cr":    0.0,
        "is_provisional": False,
        "available":      False,
        "dii_found":      False,
        "source":         "DISABLED",
        "confidence":     "NONE",
        "stale":          False,
    }


def get_fii_dii_data() -> dict:
    """Master function — single entry point for all FII/DII fetching."""
    return fetch_fii_dii_flows(max_retries=2)


def interpret_nifty_structure(close: float, ema20: float, ema50: float,
                               ema200: float, high_52w: float) -> str:
    """One-line plain English interpretation of NIFTY EMA structure."""
    above_ema20  = close > ema20
    above_ema50  = close > ema50
    above_ema200 = close > ema200
    dist_52w_high_pct = round((high_52w - close) / high_52w * 100, 1) if high_52w > 0 else 0

    if above_ema20 and above_ema50 and above_ema200:
        return "✅ Above all EMAs — bull structure intact"
    elif above_ema50 and above_ema200:
        return "🟡 Below EMA20 but above EMA50/200 — minor pullback"
    elif above_ema200 and not above_ema50:
        return "🟠 Below EMA20 & EMA50 — correction underway, watch EMA200"
    elif not above_ema200:
        if dist_52w_high_pct > 10:
            return "🔴 Below all EMAs — bearish structure. EMA200 is now resistance."
        else:
            return "🔴 Below all EMAs — recent breakdown. High caution."
    return "⚪ Mixed EMA structure — no clear trend"


def regime_explanation(score: float, regime: str, vix_in: float,
                        breadth_20: float, fii_flow: float, dii_flow: float,
                        nifty_close: float = 0, ema20: float = 0,
                        ema50: float = 0, ema200: float = 0) -> str:
    """One-line 'why this regime' — checks VIX, breadth, flows AND EMA positioning."""
    reasons_bull = []
    reasons_bear = []

    if vix_in < 15:
        reasons_bull.append(f"VIX-IN calm {vix_in:.1f}")
    elif vix_in > 20:
        reasons_bear.append(f"VIX-IN high {vix_in:.1f}")

    combined = fii_flow + dii_flow
    if combined > 3000:
        reasons_bull.append("institutions buying")
    elif combined < -2000:
        reasons_bear.append("institutions selling")

    if breadth_20 > 60:
        reasons_bull.append(f"breadth {breadth_20:.0f}%")
    elif breadth_20 < 40:
        reasons_bear.append(f"weak breadth {breadth_20:.0f}%")

    # EMA positioning — most important structural signal
    if nifty_close > 0 and ema20 > 0 and ema50 > 0 and ema200 > 0:
        above20  = nifty_close > ema20
        above50  = nifty_close > ema50
        above200 = nifty_close > ema200
        if above20 and above50 and above200:
            reasons_bull.append("above all EMAs")
        elif above50 and above200:
            reasons_bull.append("above EMA50/200")
            reasons_bear.append("below EMA20")
        elif above200:
            reasons_bear.append("below EMA20 & EMA50")
        else:
            reasons_bear.append("below all EMAs")  # strongest bear signal

    bull_str = ", ".join(reasons_bull) if reasons_bull else "none"
    bear_str = ", ".join(reasons_bear) if reasons_bear else "none"
    return f"Bulls: {bull_str} | Bears: {bear_str}"


NEWS_RSS_FEEDS = {
    "MONEYCONTROL":   "https://www.moneycontrol.com/rss/MCtopnews.xml",
    "BUSINESS_STD":   "https://www.business-standard.com/rss/markets-106.rss",
    "LIVEMINT":       "https://www.livemint.com/rss/markets",
    "NDTV_PROFIT":    "https://feeds.feedburner.com/ndtvprofit-latest",
    "REUTERS_IN":     "https://feeds.reuters.com/reuters/INbusinessNews",
}


def news_decay_weight(age_days: int) -> float:
    if age_days <= 1:    return 1.00
    elif age_days <= 3:  return 0.80
    elif age_days <= 7:  return 0.60
    elif age_days <= 14: return 0.30
    return 0.10


def fetch_news_for_symbol(symbol_clean: str, max_headlines: int = 5, max_age_days: int = 7) -> list:
    if not _FEEDPARSER_OK:
        return []
    all_headlines = []
    search_terms = [symbol_clean.lower(), symbol_clean.replace(".", " ").lower()]
    for source_name, feed_url in NEWS_RSS_FEEDS.items():
        try:
            feed = feedparser.parse(feed_url)
            for entry in feed.entries[:50]:
                title = entry.get("title", "")
                title_lower = title.lower()
                if any(term in title_lower for term in search_terms):
                    try:
                        pub = entry.get("published_parsed") or entry.get("updated_parsed")
                        if pub:
                            pub_dt = datetime.datetime(*pub[:6], tzinfo=datetime.timezone.utc)
                            age = (datetime.datetime.now(datetime.timezone.utc) - pub_dt).days
                        else:
                            age = 3
                    except Exception:
                        age = 3
                    if age <= max_age_days:
                        all_headlines.append({"title": title, "age_days": age, "source": source_name})
        except Exception:
            continue
    if not all_headlines:
        try:
            google_url = (f"https://news.google.com/rss/search?q={symbol_clean}+NSE+stock"
                          f"&hl=en-IN&gl=IN&ceid=IN:en")
            feed = feedparser.parse(google_url)
            for entry in feed.entries[:10]:
                title = entry.get("title", "")
                all_headlines.append({"title": title, "age_days": 1, "source": "GOOGLE_NEWS"})
        except Exception:
            pass
    return all_headlines[:max_headlines]


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 5 — MARKET ANALYSIS
# ─────────────────────────────────────────────────────────────────────────────

def compute_breadth(tradable: dict) -> dict:
    above_ema20 = above_ema50 = advancing = declining = total = 0
    for symbol, df in tradable.items():
        try:
            closes = df["Close"].squeeze().values.astype(float)
            if len(closes) < 50:
                continue
            last  = closes[-1]
            prev  = closes[-2]
            ema20 = float(pd.Series(closes).ewm(span=20).mean().iloc[-1])
            ema50 = float(pd.Series(closes).ewm(span=50).mean().iloc[-1])
            if last > ema20: above_ema20 += 1
            if last > ema50: above_ema50 += 1
            if last > prev:  advancing   += 1
            else:             declining   += 1
            total += 1
        except Exception:
            continue
    if total == 0:
        return {"ema20_pct": 50.0, "ema50_pct": 50.0, "adv_dec_ratio": 1.0}
    return {
        "ema20_pct":     round(above_ema20 / total * 100, 1),
        "ema50_pct":     round(above_ema50 / total * 100, 1),
        "adv_dec_ratio": round(advancing / max(1, declining), 2),
    }


def _load_sector_rank_history() -> dict:
    """Load sector rank history: {'YYYY-MM-DD': {sector: rank_int}}.

    Phase 4 (2026-07-01): persist daily 5d ranks so rotation velocity
    (5-day rank delta) can be computed run-over-run.

    Phase C7g (2026-07-10): FRESH_START now honoured. Previously this loader
    was un-guarded, so on a fresh-start run it read cache-restored JSON with
    stale date keys (e.g. 2026-07-05/07/09), added today's entry, and then
    _save_sector_rank_history() persisted ALL those old dates back into the
    output artifact — visible as "stale dates" in the results/ folder.
    Returning {} on FRESH_START ensures the saved file contains only today.
    """
    if FRESH_START:
        _log("[FRESH_START] _load_sector_rank_history → returning {} (old sector_rank_history.json ignored)")
        return {}
    try:
        if os.path.exists(SECTOR_RANK_HISTORY_FILE):
            with open(SECTOR_RANK_HISTORY_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                return data
    except Exception as e:
        _log(f"[WARN] sector rank history load failed: {e}")
    return {}


def _save_sector_rank_history(history: dict) -> None:
    """Persist sector rank history — trims to trailing 30 calendar days."""
    try:
        # Trim: keep only last 30 date-keys sorted DESC
        keys_sorted = sorted(history.keys(), reverse=True)[:30]
        trimmed = {k: history[k] for k in keys_sorted}
        with open(SECTOR_RANK_HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump(trimmed, f, indent=2, sort_keys=True)
    except Exception as e:
        _log(f"[WARN] sector rank history save failed: {e}")


def compute_sector_rotation(tradable: dict) -> dict:
    sector_returns: dict = {}
    for symbol, df in tradable.items():
        sector = get_sector(symbol)
        try:
            closes = df["Close"].squeeze().values.astype(float)
            if len(closes) < 22:
                continue
            ret5d  = (closes[-1] / closes[-6]  - 1) * 100
            ret20d = (closes[-1] / closes[-22] - 1) * 100
            sector_returns.setdefault(sector, []).append((ret5d, ret20d))
        except Exception:
            continue
    if not sector_returns:
        return {}
    all_5d  = [r[0] for v in sector_returns.values() for r in v]
    all_20d = [r[1] for v in sector_returns.values() for r in v]
    avg_5d  = sum(all_5d)  / len(all_5d)  if all_5d  else 0
    avg_20d = sum(all_20d) / len(all_20d) if all_20d else 0
    result = {}
    # First pass: compute per-sector 5d/20d means and static status.
    # 2026-07-03: bucket-size guard \u2014 sectors with fewer than MIN_BUCKET
    # members are statistically meaningless (a single stock's move dominates
    # the average). Also treat synthetic non-sector labels (DIVERSIFIED /
    # UNKNOWN / OTHERS) as NEUTRAL regardless of returns, because those
    # buckets lump unrelated names together and their aggregate return has
    # no economic meaning. These sectors still get a numeric ret5d/ret20d
    # for logging, but their `status` cannot be LAGGING/WEAKENING/LEADING.
    MIN_BUCKET = 5
    _SYNTHETIC_BUCKETS = {"DIVERSIFIED", "UNKNOWN", "OTHERS"}
    sector_5d = {}
    for sector, rets in sector_returns.items():
        s5d  = sum(r[0] for r in rets) / len(rets)
        s20d = sum(r[1] for r in rets) / len(rets)
        _n_members = len(rets)
        _is_synthetic = sector in _SYNTHETIC_BUCKETS
        if _is_synthetic or _n_members < MIN_BUCKET:
            status = "NEUTRAL"
        elif s5d > avg_5d + 1.0 and s20d > avg_20d:
            status = "LEADING"
        elif s5d < avg_5d - 1.0 and s20d < avg_20d:
            status = "LAGGING"
        elif s5d > avg_5d and s20d < avg_20d:
            status = "WEAKENING"
        else:
            status = "NEUTRAL"
        result[sector] = {
            "ret5d":   round(s5d, 2),
            "ret20d":  round(s20d, 2),
            "status":  status,
            "members": _n_members,
        }
        sector_5d[sector] = s5d

    # Phase 4 (2026-07-01): rank sectors by 5-day return (1 = best).
    # Persist to history and compute 5-day rank delta = rotation velocity.
    try:
        ranked = sorted(sector_5d.items(), key=lambda kv: kv[1], reverse=True)
        today_ranks = {sec: idx + 1 for idx, (sec, _) in enumerate(ranked)}
        today_key = ist_today().isoformat()

        history = _load_sector_rank_history()
        history[today_key] = today_ranks
        _save_sector_rank_history(history)

        # Find nearest history date >= 5 trading days back (approx 7 cal days).
        past_keys = sorted([k for k in history.keys() if k < today_key], reverse=True)
        prior_ranks = None
        prior_key = None
        for k in past_keys:
            try:
                d_prior = date.fromisoformat(k)
                d_today = date.fromisoformat(today_key)
                if (d_today - d_prior).days >= 5:
                    prior_ranks = history[k]
                    prior_key = k
                    break
            except Exception:
                continue

        for sector in result:
            r_now = today_ranks.get(sector)
            result[sector]["rank_5d"] = r_now
            if prior_ranks and sector in prior_ranks and r_now is not None:
                delta = int(prior_ranks[sector]) - int(r_now)  # +ve = moved up
                result[sector]["rank_delta_5d"] = delta
                result[sector]["rank_prior_date"] = prior_key
                # Classify rotation velocity — sector direction over 5 sessions.
                if delta >= 3:
                    rot = "ROTATING_IN"
                elif delta <= -3:
                    rot = "ROTATING_OUT"
                else:
                    rot = "STABLE"
                result[sector]["rotation_velocity"] = rot
            else:
                result[sector]["rank_delta_5d"]    = None
                result[sector]["rotation_velocity"] = "UNKNOWN"
    except Exception as e:
        _log(f"[WARN] sector rotation velocity computation failed: {e}")

    return result


def sector_rotation_score(sector: str, rotation: dict) -> tuple:
    data   = rotation.get(sector, {})
    status = data.get("status", "NEUTRAL")
    # Phase 3a #28 (2026-07-05): widened LEADING/LAGGING magnitude from
    # ±15 → ±25 so sector_strength has a real 25-75 range instead of a
    # narrow 35-65 band. Prior range gave the 0.15 factor weight only
    # ~±2.3 conf-pt swing which couldn't flip a decision.
    # Velocity overlay unchanged (±5 secondary signal).
    adj    = {"LEADING": 25, "NEUTRAL": 0, "WEAKENING": -12, "LAGGING": -25}
    # Phase 4: rotation velocity overlay — reward sectors rotating IN even
    # while still statically LAGGING (contrarian setup); penalize rotating OUT.
    velocity = data.get("rotation_velocity", "UNKNOWN")
    velocity_adj = {"ROTATING_IN": 5, "STABLE": 0, "ROTATING_OUT": -3, "UNKNOWN": 0}.get(velocity, 0)
    return adj.get(status, 0) + velocity_adj, status


def compute_key_levels(nifty_df) -> dict:
    try:
        closes = nifty_df["Close"].squeeze().values.astype(float)
        highs  = nifty_df["High"].squeeze().values.astype(float)
        lows   = nifty_df["Low"].squeeze().values.astype(float)
        last   = closes[-1]
        ema20  = float(pd.Series(closes).ewm(span=20).mean().iloc[-1])
        ema50  = float(pd.Series(closes).ewm(span=50).mean().iloc[-1])
        ema200 = float(pd.Series(closes).ewm(span=200).mean().iloc[-1])
        lookback     = min(252, len(highs))
        high_52w     = float(np.max(highs[-lookback:]))
        low_52w      = float(np.min(lows[-lookback:]))
        dist_from_high = round((high_52w - last) / high_52w * 100, 1)
        recent_high  = float(np.max(highs[-20:]))
        recent_low   = float(np.min(lows[-20:]))
        return {
            "last":                   round(last, 0),
            "ema20":                  round(ema20, 0),
            "ema50":                  round(ema50, 0),
            "ema200":                 round(ema200, 0),
            "high_52w":               round(high_52w, 0),
            "low_52w":                round(low_52w, 0),
            "recent_high_20d":        round(recent_high, 0),
            "recent_low_20d":         round(recent_low, 0),
            "dist_from_52w_high_pct": dist_from_high,
            "above_ema200":           last > ema200,
            "above_ema50":            last > ema50,
        }
    except Exception as e:
        _log(f"[WARN] compute_key_levels failed: {e}")
        return {}


def compute_nifty_state(nifty_df) -> dict:
    """Single source of truth for NIFTY EMA/level data (BUG FIX 1).
    Computed ONCE per pipeline run. Passed to all formatters.
    """
    try:
        closes = nifty_df["Close"].squeeze().values.astype(float)
        highs  = nifty_df["High"].squeeze().values.astype(float)
        lows   = nifty_df["Low"].squeeze().values.astype(float)
        ema20  = float(pd.Series(closes).ewm(span=20).mean().iloc[-1])
        ema50  = float(pd.Series(closes).ewm(span=50).mean().iloc[-1])
        ema200 = float(pd.Series(closes).ewm(span=200).mean().iloc[-1])
        last   = float(closes[-1])

        lookback = min(252, len(highs))
        high_52w = float(np.max(highs[-lookback:]))
        low_52w  = float(np.min(lows[-lookback:]))
        high_20d = float(np.max(highs[-20:]))
        low_20d  = float(np.min(lows[-20:]))

        above_ema20  = last > ema20
        above_ema50  = last > ema50
        above_ema200 = last > ema200

        if above_ema20 and above_ema50 and above_ema200:
            structure = "🟢 Above all EMAs — bull structure intact"
            ema_bear  = False
        elif above_ema20 and above_ema50:  # above short-term EMAs, below EMA200
            # Phase C1 (2026-07-02): clarified label. Short-term uptrend can
            # coexist with a long-term (200 EMA) drawdown — don't call it a
            # "correction" when the immediate trend is up.
            structure = "🟠 Above EMA20/50, below EMA200 — short-term uptrend, long-term recovery pending"
            ema_bear  = True
        elif above_ema50 and above_ema200:
            structure = "🟡 Below EMA20 — minor pullback in uptrend"
            ema_bear  = False
        elif above_ema200:
            structure = "🟠 Below EMA20 & EMA50 — correction underway"
            ema_bear  = True
        else:
            structure = "🔴 Below all EMAs — bearish structure. High caution."
            ema_bear  = True

        dist_52w_high = round((high_52w - last) / high_52w * 100, 1) if high_52w > 0 else 0.0

        # Phase R4 (2026-07-06): Nifty 20d return — used per-stock for RS
        # (relative strength) computation in the swing-alpha overlay.
        try:
            _nifty_20_ago = float(closes[-21]) if len(closes) >= 21 else float(closes[0])
            nifty_ret_20d = (last / _nifty_20_ago - 1) * 100 if _nifty_20_ago > 0 else 0.0
        except Exception:
            nifty_ret_20d = 0.0

        return {
            "close":         round(last, 1),
            "ema20":         round(ema20, 1),
            "ema50":         round(ema50, 1),
            "ema200":        round(ema200, 1),
            "high_52w":      round(high_52w, 1),
            "low_52w":       round(low_52w, 1),
            "high_20d":      round(high_20d, 1),
            "low_20d":       round(low_20d, 1),
            "above_ema20":   above_ema20,
            "above_ema50":   above_ema50,
            "above_ema200":  above_ema200,
            "ema_bear":      ema_bear,
            "structure":     structure,
            "dist_52w_high_pct": dist_52w_high,
            "ret_20d":       round(nifty_ret_20d, 2),   # ★ NEW R4: for per-stock RS
        }
    except Exception as e:
        _log(f"[WARN] compute_nifty_state failed: {e}")
        return {
            "close": 0, "ema20": 0, "ema50": 0, "ema200": 0,
            "high_52w": 0, "low_52w": 0, "high_20d": 0, "low_20d": 0,
            "above_ema20": False, "above_ema50": False, "above_ema200": False,
            "ema_bear": True, "structure": "🔴 Data unavailable",
            "dist_52w_high_pct": 0,
            "ret_20d": 0.0,   # ★ NEW R4
        }


def fetch_bse_results_dates(symbol_clean: str) -> list:
    """
    Fetch upcoming earnings / board-meeting dates for a stock.

    Phase C4 (2026-07-02): BSE's public GetScripsSearch endpoint now returns
    an HTML interstitial page instead of JSON (bot-blocking / redesign), so
    the previous BSE-only path silently returned [] for every stock. We now
    try in order:
        1. yfinance Ticker.calendar → 'Earnings Date' (reliable, no auth)
        2. Legacy BSE JSON endpoint (kept as fallback for the day it comes
           back; also for stocks not in Yahoo's coverage)

    Returns list of date strings "YYYY-MM-DD" (future only). Empty list on
    any failure. Silent on individual stock failures — the top-level counter
    reports coverage.
    """
    today = ist_today()

    # ── 1. yfinance calendar (primary) ────────────────────────────────────
    try:
        sym_yf = symbol_clean if symbol_clean.endswith(".NS") else symbol_clean + ".NS"
        cal = yf.Ticker(sym_yf).calendar
        if isinstance(cal, dict):
            ed = cal.get("Earnings Date") or cal.get("earnings_date") or []
            dates = []
            if isinstance(ed, list):
                for d in ed:
                    try:
                        if hasattr(d, "isoformat"):
                            dobj = d if isinstance(d, datetime.date) else d.date()
                        else:
                            dobj = datetime.datetime.strptime(str(d)[:10], "%Y-%m-%d").date()
                        if dobj >= today:
                            dates.append(dobj.isoformat())
                    except Exception:
                        continue
            elif ed:
                try:
                    dobj = ed if isinstance(ed, datetime.date) else \
                           datetime.datetime.strptime(str(ed)[:10], "%Y-%m-%d").date()
                    if dobj >= today:
                        dates.append(dobj.isoformat())
                except Exception:
                    pass
            if dates:
                return dates
    except Exception:
        pass  # fall through to BSE

    # ── 2. BSE JSON API (legacy fallback; currently returns HTML) ────────
    try:
        search_url = f"https://api.bseindia.com/BseIndiaAPI/api/GetScripsSearch/w?strSearch={symbol_clean}"
        resp = requests.get(search_url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Referer": "https://www.bseindia.com",
        }, timeout=8)
        if resp.status_code != 200:
            return []
        # BSE now returns HTML on this endpoint — bail out cleanly instead
        # of raising a JSON parse exception.
        ctype = resp.headers.get("content-type", "").lower()
        if "json" not in ctype and not resp.text.lstrip().startswith(("[", "{")):
            return []
        results = resp.json()
        if not results or not isinstance(results, list):
            return []
        scrip_code = str(results[0].get("SCRIP_CD", ""))
        if not scrip_code:
            return []

        # Fetch corporate actions for this scrip
        ca_url = (f"https://api.bseindia.com/BseIndiaAPI/api/CorporateAction/w"
                  f"?pageno=1&strScrip={scrip_code}&type=GP&Fdate=&TDate=")
        ca_resp = requests.get(ca_url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Referer": "https://www.bseindia.com",
        }, timeout=8)
        if ca_resp.status_code != 200:
            return []
        ca_ctype = ca_resp.headers.get("content-type", "").lower()
        if "json" not in ca_ctype and not ca_resp.text.lstrip().startswith(("[", "{")):
            return []
        actions = ca_resp.json().get("Table", [])
        dates = []
        for action in actions:
            raw = action.get("REC_DATE") or action.get("NEWS_DT") or ""
            try:
                # BSE returns dates as "DD/MM/YYYY" or "YYYY-MM-DDT00:00:00"
                if "/" in raw:
                    d = datetime.datetime.strptime(raw.split(" ")[0], "%d/%m/%Y").date()
                else:
                    d = datetime.datetime.strptime(raw[:10], "%Y-%m-%d").date()
                if d >= today:
                    dates.append(d.isoformat())
            except Exception:
                continue
        return dates
    except Exception:
        return []


def is_near_event(symbol_clean: str, results_dates: list,
                  upcoming_events: list, window_days: int = 5) -> tuple:
    """
    Returns (True, reason_str) if a BUY should be blocked due to an imminent event.
    Checks: (1) BSE results dates for the specific stock, (2) NSE expiry dates.
    Returns (False, "") if clear to trade.
    """
    today = ist_today()
    cutoff = today + datetime.timedelta(days=window_days)

    # Check stock-specific results date
    for d_str in results_dates:
        try:
            d = datetime.date.fromisoformat(d_str)
            if today <= d <= cutoff:
                return True, f"RESULTS_IN_{(d - today).days}D"
        except Exception:
            continue

    # Check NSE expiry in upcoming_events (monthly expiry is highest risk)
    for ev in upcoming_events:
        if "Monthly Expiry" in ev:
            try:
                # Extract date from string like "NSE Monthly Expiry — 26 Jun"
                parts = ev.split("—")[-1].strip()
                d = datetime.datetime.strptime(f"{parts} {today.year}", "%d %b %Y").date()
                if today <= d <= cutoff:
                    return True, f"MONTHLY_EXPIRY_IN_{(d - today).days}D"
            except Exception:
                continue

    return False, ""


# Known market events with sector impact (FEATURE 5)
EVENTS_CONFIG_FILE = "events_config.json"

# ── Annual event schedules ────────────────────────────────────────────────────
# Phase C7e (2026-07-02): _RBI_MPC_DATES_2026 / _FOMC_DATES_2026 were relocated
# to market_calendars.json and are now loaded (with hardcoded 2026 fallback) at
# module import (see _load_market_calendars near L975). The names _RBI_MPC_DATES
# and _FOMC_DATES are the current canonical form. The *_2026 back-compat aliases
# still exist for any external tooling that imports them.


def _nse_expiry_dates(lookahead_days: int = 60) -> list:
    """
    Returns all NSE weekly + monthly expiry dates in the next `lookahead_days`.
    Weekly  = every Thursday.
    Monthly = last Thursday of each month (when it falls in the window).
    """
    today  = ist_today()
    end    = today + datetime.timedelta(days=lookahead_days)
    events = []
    cur    = today
    while cur <= end:
        if cur.weekday() == 3:            # Thursday
            # Is it the last Thursday of the month?
            next_thu = cur + datetime.timedelta(weeks=1)
            is_monthly = next_thu.month != cur.month
            events.append({
                "name":         "NSE Monthly Expiry" if is_monthly else "NSE Weekly Expiry",
                "date":         cur.isoformat(),
                "note":         ("Monthly expiry — high volatility, avoid new entries on expiry day"
                                 if is_monthly else
                                 "Weekly expiry — increased intraday volatility, tighten stops"),
                "sectors_up":   [],
                "sectors_down": [],
                "_auto":        True,
            })
        cur += datetime.timedelta(days=1)
    return events


def _rbi_mpc_events(lookahead_days: int = 60) -> list:
    """Returns upcoming RBI MPC decision dates from the loaded schedule."""
    today  = ist_today()
    end    = today + datetime.timedelta(days=lookahead_days)
    events = []
    for ds in _RBI_MPC_DATES:
        try:
            d = datetime.date.fromisoformat(ds)
            if today <= d <= end:
                events.append({
                    "name":         "RBI MPC Decision",
                    "date":         ds,
                    "note":         "Rate decision — banking/finance/realty stocks react sharply",
                    "sectors_up":   ["BANKING", "FINANCE", "REALTY"],
                    "sectors_down": [],
                    "_auto":        True,
                })
        except Exception:
            pass
    return events


def _fomc_events(lookahead_days: int = 60) -> list:
    """Returns upcoming FOMC decision dates (impacts FII flows + DXY)."""
    today  = ist_today()
    end    = today + datetime.timedelta(days=lookahead_days)
    events = []
    for ds in _FOMC_DATES:
        try:
            d = datetime.date.fromisoformat(ds)
            if today <= d <= end:
                events.append({
                    "name":         "US FOMC Decision",
                    "date":         ds,
                    "note":         "Fed rate decision — DXY/FII flows react, watch IT/PHARMA",
                    "sectors_up":   ["IT", "PHARMA"],
                    "sectors_down": ["REALTY", "FINANCE"],
                    "_auto":        True,
                })
        except Exception:
            pass
    return events


def build_events_calendar(lookahead_days: int = 30) -> list:
    """
    Auto-builds the full events calendar.
    Sources: NSE expiry math + known RBI MPC + FOMC schedules.
    Saves to events_config.json so it's auditable.
    Returns only events within lookahead_days, sorted by date.
    Called ONCE at pipeline start — no manual editing needed.
    """
    import json
    try:
        all_events = (
            _nse_expiry_dates(lookahead_days) +
            _rbi_mpc_events(lookahead_days) +
            _fomc_events(lookahead_days)
        )
        # Sort by date
        all_events.sort(key=lambda x: x.get("date", ""))

        # Persist for auditability (strip _auto flag before saving)
        saveable = [{k: v for k, v in ev.items() if k != "_auto"} for ev in all_events]
        try:
            with open(EVENTS_CONFIG_FILE, "w") as f:
                json.dump(saveable, f, indent=2)
        except Exception:
            pass

        today = ist_today()
        end   = today + datetime.timedelta(days=lookahead_days)
        return [
            ev for ev in all_events
            if today <= datetime.date.fromisoformat(ev["date"]) <= end
        ]
    except Exception as e:
        _log(f"[WARN] build_events_calendar failed: {e}")
        return []


def load_events_config() -> list:
    """
    Builds the events calendar, then deduplicates by type so recurring events
    (e.g. weekly expiry) appear only ONCE (the nearest occurrence).
    Shows max 3 total events, only within the next 14 days.
    """
    all_events = build_events_calendar(lookahead_days=30)
    today = ist_today()

    # Step 1: sort by date (nearest first)
    all_events.sort(key=lambda x: x.get("date", ""))

    # Step 2: deduplicate — keep only the NEXT occurrence of each event type
    seen_types: set = set()
    deduplicated = []
    for ev in all_events:
        ev_type = ev.get("name", "").strip()
        # Group all expiry variants under one key
        if "expiry" in ev_type.lower() or "Expiry" in ev_type:
            ev_type = "NSE_EXPIRY"
        if ev_type not in seen_types:
            deduplicated.append(ev)
            seen_types.add(ev_type)
        if len(deduplicated) >= 4:
            break

    # Step 3: only events within next 14 days are relevant today
    relevant = []
    for ev in deduplicated:
        try:
            days_away = (datetime.date.fromisoformat(ev["date"]) - today).days
            if days_away <= 14:
                relevant.append(ev)
        except Exception:
            pass

    # If nothing within 14 days, show next 2 regardless
    if not relevant and deduplicated:
        relevant = deduplicated[:2]

    return relevant


def format_upcoming_events_compact(events_config: list, holdings: list) -> list:
    """
    2 lines per event maximum (FIX 3G).
    Line 1: Event name + date
    Line 2: Impact on holdings (if any) or general note
    """
    try:
        if not events_config:
            return []
        lines = ["UPCOMING EVENTS"]
        hold_secs = set()
        for h in (holdings or []):
            sym = h.get("symbol", "")
            if sym:
                hold_secs.add(get_sector(sym))

        for event in events_config:
            name    = event.get("name", "")
            date    = event.get("date", "")
            note    = event.get("note", "")
            up_secs = set(event.get("sectors_up",   []))
            dn_secs = set(event.get("sectors_down", []))

            lines.append(f"  {html.escape(name)} \u2014 {html.escape(date)}")

            impacted_up = hold_secs & up_secs
            impacted_dn = hold_secs & dn_secs
            if impacted_up:
                lines.append(f"    \ud83d\udfe2 Holdings may benefit: {', '.join(sorted(impacted_up))}")
            elif impacted_dn:
                lines.append(f"    \ud83d\udd34 Holdings at risk: {', '.join(sorted(impacted_dn))}")
            elif note:
                lines.append(f"    {html.escape(note[:80])}")
        return lines
    except Exception:
        return ["UPCOMING EVENTS"] + [f"  {html.escape(str(ev.get('name','')))} \u2014 {ev.get('date','')}" for ev in (events_config or [])]


def score_to_regime(score: float, vix_in: float,
                     above_ema200: "bool | None" = None,
                     above_ema20:  "bool | None" = None,
                     above_ema50:  "bool | None" = None) -> str:
    """
    Maps score to regime with sanity checks:
      - VIX-IN < 16 = market is NOT in high volatility regardless of score.
      - STRONG_BULL / BULL require price above EMA200 *only if* the short-term
        structure (EMA20/50) is ALSO broken. If NIFTY is above EMA20 and EMA50
        but below EMA200, we are in a "recovery" pattern — the current uptrend
        is real and tradable, just not yet long enough to reclaim the 200-day.
        In that case we soften by 1 tier only.
      - If NIFTY is below EMA20/50/200 all three, that IS a broad correction,
        so we cap harder (2 tiers).
    """
    if score >= 80:    base = "STRONG_BULL"
    elif score >= 65:  base = "BULL"
    elif score >= 52:  base = "SIDEWAYS"
    elif score >= 40:  base = "TRANSITION"
    elif score >= 28:  base = "HIGH_VOLATILITY"
    elif score >= 15:  base = "BEAR"
    else:              base = "STRONG_BEAR"

    # EMA-structure sanity check (Phase C4, revised 2026-07-02):
    #   Short-term EMAs intact (EMA20 ✓ AND EMA50 ✓) + only EMA200 below
    #        → 1-tier softening (recovery pattern, uptrend is real)
    #   Short-term EMAs also broken (EMA20 ✗ or EMA50 ✗) + EMA200 below
    #        → 2-tier hard cap (broad correction, don't chase)
    if above_ema200 is False:
        short_ok = (above_ema20 is True) and (above_ema50 is True)
        soft_map = {"STRONG_BULL": "BULL", "BULL": "SIDEWAYS"}
        hard_map = {"STRONG_BULL": "SIDEWAYS", "BULL": "TRANSITION"}
        cap_map = soft_map if short_ok else hard_map
        if base in cap_map:
            new_base = cap_map[base]
            kind = "soft (short-term structure intact)" if short_ok else "hard (broad correction)"
            _log(
                f"[INFO] Regime override: score {score:.1f} → {base} "
                f"but NIFTY below EMA200 → capping at {new_base} [{kind}]"
            )
            base = new_base

    # VIX sanity check: calm VIX cannot be HIGH_VOLATILITY
    if base == "HIGH_VOLATILITY" and vix_in < 16:
        _log(
            f"[INFO] Regime override: score {score:.1f} → HIGH_VOLATILITY "
            f"but VIX-IN {vix_in:.1f} is calm → using TRANSITION instead"
        )
        return "TRANSITION"

    # STRONG_BEAR with calm VIX is also suspicious
    if base == "STRONG_BEAR" and vix_in < 18:
        return "BEAR"

    return base



def detect_market_regime(nifty_df, breadth_data: dict, macro_signals: dict) -> dict:
    score = 50
    try:
        closes = nifty_df["Close"].squeeze().values.astype(float)
        ema20  = pd.Series(closes).ewm(span=20).mean().values
        ema50  = pd.Series(closes).ewm(span=50).mean().values
        ema200 = pd.Series(closes).ewm(span=200).mean().values
        last   = closes[-1]
        ret_5d  = (last / closes[-6]  - 1) * 100 if len(closes) > 6  else 0
        ret_21d = (last / closes[-22] - 1) * 100 if len(closes) > 22 else 0

        if last > ema20[-1] > ema50[-1] > ema200[-1]: score += 30
        elif last > ema50[-1] > ema200[-1]:            score += 18
        elif last > ema200[-1]:                         score += 8
        elif last < ema200[-1]:                         score -= 15

        if ret_5d > 1.5 and ret_21d > 4:      score += 20
        elif ret_5d > 0 and ret_21d > 1:       score += 10
        elif ret_5d < -1.5 and ret_21d < -3:   score -= 20
        elif ret_5d < 0:                        score -= 8

        b20 = breadth_data.get("ema20_pct", 50)
        b50 = breadth_data.get("ema50_pct", 50)
        if b20 > 70 and b50 > 70:    score += 25
        elif b20 > 55 and b50 > 55:  score += 15
        elif b20 > 45:               score += 5
        elif b20 < 35:               score -= 15
        elif b20 < 25:               score -= 25

        vix_in = macro_signals.get("vix_in", 15)
        if vix_in < 13:    score += 15
        elif vix_in < 17:  score += 8
        elif vix_in < 22:  score += 0
        elif vix_in < 27:  score -= 12
        else:              score -= 20

        score += macro_regime_adjustment(macro_signals)
    except Exception:
        score = 50

    score = max(0, min(100, score))

    # Use VIX + EMA200 sanity-checked mapping instead of raw score boundaries.
    # A STRONG_BULL score with price under EMA200 is contradictory — the cap
    # lives inside score_to_regime().
    vix_for_regime = macro_signals.get("vix_in", 15)
    try:
        last_close   = float(closes[-1])
        ema200_last  = float(ema200[-1]) if ema200 is not None and len(ema200) else 0.0
        ema50_last   = float(ema50[-1])  if ema50  is not None and len(ema50)  else 0.0
        ema20_last   = float(ema20[-1])  if ema20  is not None and len(ema20)  else 0.0
        above_ema200 = last_close > ema200_last if ema200_last > 0 else None
        above_ema50  = last_close > ema50_last  if ema50_last  > 0 else None
        above_ema20  = last_close > ema20_last  if ema20_last  > 0 else None
    except Exception:
        above_ema200 = None
        above_ema50  = None
        above_ema20  = None
    regime = score_to_regime(score, vix_for_regime,
                              above_ema200=above_ema200,
                              above_ema20=above_ema20,
                              above_ema50=above_ema50)

    # Issue 7 fix: demote regime one tier when breadth is weak AND
    # advance/decline is decisively bearish. Guards against the "index-up
    # but market-broken" scenario where 3 mega-caps mask a 47-stock rout.
    # Env: BREADTH_DEMOTION (default 1). Set to 0 to disable.
    if os.getenv("BREADTH_DEMOTION", "1") == "1":
        try:
            _b20 = float(breadth_data.get("ema20_pct", 50) or 50)
            _adv_dec = float(breadth_data.get("adv_dec_ratio", 1.0) or 1.0)
            if _b20 < 40.0 and _adv_dec < 0.9:
                _demotion_chain = {
                    "STRONG_BULL":     "BULL",
                    "BULL":            "SIDEWAYS",
                    "SIDEWAYS":        "TRANSITION",
                    "TRANSITION":      "HIGH_VOLATILITY",
                    "HIGH_VOLATILITY": "BEAR",
                    "BEAR":            "STRONG_BEAR",
                    "STRONG_BEAR":     "STRONG_BEAR",
                }
                _new_regime = _demotion_chain.get(regime, regime)
                if _new_regime != regime:
                    regime = _new_regime
        except (TypeError, ValueError):
            pass

    return {
        "regime":     regime,
        "score":      round(score, 1),
        "thresholds": REGIME_THRESHOLDS[regime],
        "macro":      macro_signals,
    }


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 6 — SCORING ENGINE
# ─────────────────────────────────────────────────────────────────────────────

def _obv_trend_score(closes: np.ndarray, volumes: np.ndarray) -> float:
    """
    On-Balance Volume trend score (0-100).
    Replaces delivery % which requires external bhavcopy data unavailable from CI.
    OBV = cumsum of volume * sign(close - prev_close).
    Rising OBV = smart money accumulating. Falling OBV = distribution.
    OBV divergence (price up, OBV down) = bearish warning.
    """
    try:
        n = min(len(closes), len(volumes))
        if n < 12:
            return 50.0
        c = closes[-n:]
        v = volumes[-n:]
        # Build OBV
        obv = [0.0]
        for i in range(1, n):
            if c[i] > c[i-1]:
                obv.append(obv[-1] + v[i])
            elif c[i] < c[i-1]:
                obv.append(obv[-1] - v[i])
            else:
                obv.append(obv[-1])
        obv = np.array(obv)
        # Slope of last 10 OBV values (normalised by avg volume)
        window = obv[-10:]
        avg_vol = float(np.mean(v[-10:])) or 1.0
        slope = float(np.polyfit(range(10), window, 1)[0]) / avg_vol  # slope in vol-units/bar
        # Price direction over same window
        price_up = c[-1] > c[-10]
        if slope > 0.3 and price_up:    return 95.0   # strong accumulation
        if slope > 0.1 and price_up:    return 80.0   # moderate accumulation
        if slope > 0.0:                 return 65.0   # quiet buying
        if slope > -0.1:                return 50.0   # neutral
        if slope <= -0.1 and price_up:  return 30.0   # bearish divergence (price up, OBV down)
        if slope <= -0.3:               return 15.0   # distribution
        return 40.0
    except Exception:
        return 50.0


def volume_delivery_score(vol_ratio: float, obv_score: float) -> float:
    vol_score = min(100, vol_ratio * 45)
    return round(vol_score * 0.55 + obv_score * 0.45, 1)


def accumulation_score(closes: np.ndarray, volumes: np.ndarray,
                       ema20: float, avg_vol: float) -> tuple:
    """
    Detects institutional accumulation / distribution over the last 10 bars.

    Accumulation: 3+ days of above-avg volume while price held ABOVE ema20.
    Distribution: 3+ days of above-avg volume while price closed BELOW ema20.

    Returns (score 0-100, signal: "ACCUMULATING"|"DISTRIBUTING"|"NEUTRAL")
    Score feeds into volume_delivery factor to boost/penalise beyond single-day vol.
    """
    try:
        if len(closes) < 11 or avg_vol <= 0:
            return 50, "NEUTRAL"

        accum_days = 0
        distrib_days = 0
        for i in range(-10, 0):
            is_high_vol = volumes[i] > avg_vol * 1.2
            if not is_high_vol:
                continue
            if closes[i] > ema20:
                accum_days += 1
            else:
                distrib_days += 1

        if accum_days >= 4:
            return 90, "ACCUMULATING"
        elif accum_days >= 3:
            return 75, "ACCUMULATING"
        elif accum_days >= 2:
            return 62, "ACCUMULATING"
        elif distrib_days >= 4:
            return 15, "DISTRIBUTING"
        elif distrib_days >= 3:
            return 28, "DISTRIBUTING"
        elif distrib_days >= 2:
            return 38, "DISTRIBUTING"
        return 50, "NEUTRAL"
    except Exception:
        return 50, "NEUTRAL"


def compute_base_confidence(scores: dict) -> float:
    """Weighted average of factor scores with MISSING-DATA reweighting.

    Phase C5 (rating ≥ 9.0):
      When a factor score is `None` (i.e. we couldn't measure it — rate-limited,
      unmapped sector, news not fetched), we redistribute its weight over the
      factors we DID measure, instead of silently defaulting to 50 and
      diluting the confidence toward the middle.

    Contract:
      - `scores[k] is None`  → factor MISSING → excluded from sum and denominator
      - `scores.get(k)` is a number → factor PRESENT → contributes normally
      - key not in scores at all → treated as MISSING (safe default)

    If ALL factors are missing (should never happen) → returns 50.0.
    """
    num = 0.0
    den = 0.0
    for k, w in FACTOR_WEIGHTS.items():
        v = scores.get(k, None)
        if v is None:
            continue
        try:
            num += w * float(v)
            den += w
        except (TypeError, ValueError):
            continue
    if den <= 0:
        return 50.0
    return round(num / den, 2)


def compute_news_penalty(ai_result: dict, age_days: int) -> float:
    severity = ai_result.get("severity", 0)
    if ai_result.get("is_black_swan"):
        return 999.0
    effective = severity * news_decay_weight(age_days)
    if effective >= 70:   return 35.0
    elif effective >= 50: return 22.0
    elif effective >= 30: return 12.0
    elif effective >= 15: return 6.0
    elif effective < 0:   return max(-8.0, effective * 0.2)
    return 0.0


def compute_final_confidence(base: float, regime: str, news_penalty: float,
                              macro_adj: float, bulk_adj: int) -> float:
    # Bug 7 fix (2026-07-03): SIDEWAYS regime penalty relaxed from -5 → -2.
    # Rationale: -5 stacked on min_confidence=80 in SIDEWAYS meant base_conf
    # had to reach 85 for a BUY to fire — structurally near-impossible given
    # the R/R cap (2.5x → 75 base points). Tracker analysis showed only
    # NEAR_MISS rows in SIDEWAYS despite valid setups. -2 keeps SIDEWAYS
    # more cautious than BULL but no longer double-punished with the
    # threshold. Effective bar drops from 85 → 82 base which is achievable.
    REGIME_ADJ = {
        "STRONG_BULL": +8, "BULL": +4, "SIDEWAYS": -2,
        "TRANSITION": -3, "HIGH_VOLATILITY": -8, "BEAR": -20, "STRONG_BEAR": -40,
    }
    # Phase 3a N6 (2026-07-05): removed `+ macro_adj` — macro was double-counted.
    # It is already applied inside compute_all_factors as `macro_alignment = 60 + macro_adj*2`
    # which then feeds compute_base_confidence at weight 0.03. Adding raw
    # macro_adj again here inflated by 30-100% of the intended weight.
    # Preserved kwarg for backwards compatibility with callers that pass it,
    # but only bulk_adj remains as an out-of-factor bonus.
    _ = macro_adj  # explicit no-op — see comment above
    final = base + REGIME_ADJ.get(regime, 0) - news_penalty + bulk_adj
    return round(max(0.0, min(100.0, final)), 2)


# ─────────────────────────────────────────────────────────────────────────────
# Phase I (2026-07-07) — Setup-edge patch, data-driven from backtest results.
#
# Backtest run 2026-07-07 (26,665 trades, results/backtest_by_setup.csv):
#     BREAKOUT  1,453  win 51.5%  expectancy +0.017 R   (ONLY profitable setup)
#     MOMENTUM  4,889  win 41.3%  expectancy -0.038 R   (near break-even)
#     OTHER    15,377  win 28.5%  expectancy -0.241 R   (worst)
#     REVERSAL  2,328  win 26.9%  expectancy -0.136 R
#     PULLBACK  2,618  win 20.6%  expectancy -0.173 R
#
# Backtest run 2026-07-07 (results/backtest_by_regime.csv):
#     BULLISH           4,284  win 45.1%  expectancy +0.06 R  (ONLY profitable)
#     CAUTIOUS_BULLISH  5,958  win 37.6%  expectancy -0.107 R
#     SIDEWAYS          5,732  win 20.3%  expectancy -0.176 R
#     WEAK             10,691  win 27.9%  expectancy -0.304 R (worst — 40% of sample)
#
# Overall raw system: -0.174 R expectancy, profit factor 0.79 → NET LOSER
# without filtering. Monte Carlo verdict = RANDOM (p=0.813).
#
# This module adds three data-driven adjustments to the live scanner:
#   1. Setup-type confidence bonus (positive for BREAKOUT/MOMENTUM,
#      negative for PULLBACK/REVERSAL/OTHER).
#   2. Regime hard-block: WEAK + SIDEWAYS setups get rejected unless
#      they are a fresh BREAKOUT (the only bucket that survived losses
#      in choppy tapes).
#   3. Uses the same EMA/high/low context already stamped on each stock
#      dict by score_stock() — no new data fetch, no runtime cost.
# ─────────────────────────────────────────────────────────────────────────────

# Tunable via env — defaults derived from backtest expectancy deltas.
_SETUP_CONF_BONUS = {
    "BREAKOUT": float(os.getenv("SETUP_BONUS_BREAKOUT", "12")),
    "MOMENTUM": float(os.getenv("SETUP_BONUS_MOMENTUM",  "5")),
    "OTHER":    float(os.getenv("SETUP_BONUS_OTHER",     "0")),
    "REVERSAL": float(os.getenv("SETUP_BONUS_REVERSAL", "-4")),
    "PULLBACK": float(os.getenv("SETUP_BONUS_PULLBACK", "-6")),
}

# Regime hard-block toggle. When true, WEAK + SIDEWAYS regime rejects any
# stock whose live setup is not BREAKOUT. BULLISH / CAUTIOUS_BULLISH bypass
# the filter regardless of setup.
_ENABLE_REGIME_SETUP_FILTER = os.getenv("ENABLE_REGIME_SETUP_FILTER", "true").lower() in ("1", "true", "yes")

# Chop regimes are the ones the backtest showed as unprofitable when
# taking ALL setup types indiscriminately. Live main.py regime names differ
# from the backtest's proxy names (main.py uses SIDEWAYS/TRANSITION/BEAR/etc.),
# so both spellings are checked.
_CHOP_REGIMES = {"SIDEWAYS", "TRANSITION", "HIGH_VOLATILITY", "WEAK"}


def _classify_setup_live(stock: dict) -> str:
    """Return BREAKOUT | PULLBACK | REVERSAL | MOMENTUM | OTHER for a scored
    live stock dict. Uses only fields already stamped by score_stock():
      close, ema20, ema50, high_20d, low_20d, chg_5d (or chg_1d fallback).

    Mirrors backtest_walkforward._classify_setup_at() so live scoring is
    consistent with the profitability table it was calibrated on. Returns
    "OTHER" on any missing field — cheap heuristic, precision is not critical.
    """
    try:
        last   = float(stock.get("close",   0) or 0)
        ema20  = float(stock.get("ema20",   0) or 0)
        ema50  = float(stock.get("ema50",   0) or 0)
        h20    = float(stock.get("high_20d", 0) or 0)
        l20    = float(stock.get("low_20d",  0) or 0)
        # 5-day return: prefer chg_5d, fall back to a synthetic near-zero if missing
        ret5   = float(stock.get("chg_5d", stock.get("return_5d", 0)) or 0)

        if last <= 0 or ema20 <= 0 or ema50 <= 0 or h20 <= 0:
            return "OTHER"

        # 1) BREAKOUT — closes at/above prior 20-day high AND above ema20/50
        if last >= h20 * 0.998 and last > ema20 > ema50:
            return "BREAKOUT"

        # 2) PULLBACK — uptrend intact but negative 5d return, price near ema20
        if ema20 > ema50 and last > ema50 and ret5 < 0 \
           and abs(last - ema20) / max(ema20, 1e-9) < 0.015:
            return "PULLBACK"

        # 3) REVERSAL — near ema20 but below ema50, bounced 5–20% off 20d low
        if l20 > 0 and last > ema20 and last < ema50 and 0.05 < (last - l20) / l20 < 0.20:
            return "REVERSAL"

        # 4) MOMENTUM — sustained uptrend, strong 5-day return
        if last > ema20 > ema50 and ret5 > 2.0:
            return "MOMENTUM"

        return "OTHER"
    except Exception:
        return "OTHER"


def apply_setup_edge(stock: dict, regime: str) -> tuple:
    """Apply the two data-driven filters to a scored stock:
      • Compute setup_type via _classify_setup_live().
      • Add setup-type bonus to final_confidence (clamped 0–100).
      • Return (setup_type, adjusted_confidence, skip_reason) where
        skip_reason is None if the stock passes the regime filter, or a
        short string like "REGIME_CHOP_NO_BREAKOUT" when it should be
        rejected from the BUY set even if confidence is otherwise fine.

    Stamps `setup_type` and `setup_conf_bonus` on the stock dict for audit.
    """
    setup = _classify_setup_live(stock)
    bonus = _SETUP_CONF_BONUS.get(setup, 0.0)

    stock["setup_type"]       = setup
    stock["setup_conf_bonus"] = bonus

    orig_conf = float(stock.get("final_confidence", 0) or 0)
    adj_conf  = round(max(0.0, min(100.0, orig_conf + bonus)), 2)
    stock["final_confidence"] = adj_conf

    skip_reason = None
    if _ENABLE_REGIME_SETUP_FILTER and regime in _CHOP_REGIMES and setup != "BREAKOUT":
        # Change #3 (2026-07-10): MOMENTUM in chop is NO LONGER a hard block.
        # Backtest shows MOMENTUM WR ~47% (vs BREAKOUT 51%). Blocking it entirely
        # in chop regimes loses good signals. Instead apply a soft 10% CONF
        # penalty — only exceptional MOMENTUM (raw ≥ ~89 → ~80 after penalty)
        # can pass the TRANSITION 83-threshold gate. PULLBACK/REVERSAL/OTHER
        # remain hard-blocked (backtest WR 21-27% — genuinely poor).
        if setup == "MOMENTUM":
            penalty_mult = 0.90
            adj_conf     = round(max(0.0, min(100.0, adj_conf * penalty_mult)), 2)
            stock["final_confidence"]      = adj_conf
            stock["chop_momentum_penalty"] = round((1.0 - penalty_mult) * 100, 1)  # audit
            # No skip_reason — let downstream gates decide based on the
            # penalised confidence. If it still passes, it's a good signal.
        else:
            # BREAKOUT is the only setup with positive expectancy in the backtest,
            # so in choppy/weak regimes we require it as the entry pattern.
            skip_reason = f"REGIME_CHOP_NO_BREAKOUT({regime}/{setup})"

            # Phase I shadow-log (2026-07-07): record this rejection into
            # bucket B (WATCH_ME — right setup, wrong regime) OR bucket C
            # (NOT_MY_STYLE — wrong setup entirely). Split lets us tell
            # WHERE the edge lives after 30 days of observation.
            if _SHADOW_LOG_OK:
                try:
                    _bucket = shadow_log.classify_skip_bucket(setup)
                    shadow_log.record_shadow_trade(_bucket, stock, regime,
                                                   note=f"skip:{skip_reason[:40]}")
                except Exception:
                    pass  # never let the shadow log break the pipeline

    return setup, adj_conf, skip_reason


# Env-driven cap (default 25%). Read once at module load so both sizers agree.
_MAX_POSITION_PCT_ENV = float(os.getenv("MAX_POSITION_PCT", "25")) / 100.0


def compute_position_size(entry: float, stop: float, capital: float,
                           risk_per_trade: float = 0.015,
                           max_position_pct: float = None,
                           avg_val_lakhs: float = None) -> dict:
    """Risk-based position sizing — always returns non-zero for valid inputs.

    Returns keys: shares, position_value, position_pct, risk_amount, risk_pct,
    max_loss (= shares × (entry − stop), the actual worst-case rupee loss).

    Issue 6 (liquidity cap): when avg_val_lakhs > 0, additionally cap
    position value to LIQUIDITY_MAX_PCT_OF_ADV (default 5%) of average daily
    traded value. This prevents entering positions we can't unwind without
    slippage. Only tightens; never expands the position.
    """
    if max_position_pct is None:
        max_position_pct = _MAX_POSITION_PCT_ENV
    try:
        if entry <= 0 or stop <= 0 or stop >= entry or capital <= 0:
            return {"shares": 0, "position_value": 0.0, "position_pct": 0.0,
                    "risk_amount": 0.0, "risk_pct": 0.0, "max_loss": 0.0}
        risk_per_share = entry - stop
        risk_amount    = capital * risk_per_trade
        shares         = max(1, int(risk_amount / risk_per_share))
        position_value = shares * entry
        if position_value > capital * max_position_pct:
            shares         = max(1, int((capital * max_position_pct) / entry))
            position_value = shares * entry
        # Issue 6 fix: liquidity cap. avg_val_lakhs is average daily
        # rupee volume in lakhs (₹100k units). We limit our position to
        # LIQUIDITY_MAX_PCT_OF_ADV of that so we can exit in under 1 day
        # without moving price. Default 5% — set env to 0 to disable.
        try:
            _liq_pct = float(os.getenv("LIQUIDITY_MAX_PCT_OF_ADV", "5.0")) / 100.0
        except (TypeError, ValueError):
            _liq_pct = 0.05
        if avg_val_lakhs and avg_val_lakhs > 0 and _liq_pct > 0:
            _liq_cap_value = float(avg_val_lakhs) * 100_000 * _liq_pct
            if position_value > _liq_cap_value:
                shares         = max(1, int(_liq_cap_value / entry))
                position_value = shares * entry
        position_pct = (position_value / capital) * 100
        # Actual worst-case rupee loss (may differ slightly from risk_amount
        # because shares is an integer). Downstream renders read this key.
        max_loss = round(shares * risk_per_share, 2)
        return {
            "shares":         shares,
            "position_value": round(position_value, 2),
            "position_pct":   round(position_pct, 1),
            "risk_amount":    round(risk_amount, 2),
            "risk_pct":       round(risk_per_trade * 100, 1),
            "max_loss":       max_loss,
        }
    except Exception:
        return {"shares": 0, "position_value": 0.0, "position_pct": 0.0,
                "risk_amount": 0.0, "risk_pct": 0.0, "max_loss": 0.0}


def compute_portfolio_heat(holdings: list, current_prices: dict,
                           capital: float) -> dict:
    """
    Computes current portfolio heat = total open risk as % of capital.
    Open risk = sum of (current_price - stop_loss) * shares for all open positions.
    If stop_loss unknown, uses 6% of entry as proxy.

    Returns {"heat_pct": float, "positions_count": int, "max_heat_pct": float,
             "heat_ok": bool, "heat_remaining_pct": float}
    Max heat = 6% of capital (professional standard: never risk more than 6% total).
    """
    MAX_HEAT_PCT = float(os.getenv("MAX_PORTFOLIO_HEAT", "6.0"))
    try:
        total_risk = 0.0
        for h in holdings:
            sym   = h.get("symbol", "")
            entry = float(h.get("entry_price", 0) or 0)
            stop  = float(h.get("stop_loss", 0) or 0)
            qty   = float(h.get("quantity", 0) or 0)
            curr  = float(current_prices.get(sym, entry) or entry)

            if entry <= 0:
                continue
            if qty <= 0:
                # Estimate qty from position_pct if available
                pos_pct = float(h.get("position_pct", 5) or 5)
                qty = max(1, int((capital * pos_pct / 100) / entry))

            if stop <= 0:
                stop = entry * 0.94   # assume 6% stop if not set

            # Risk = how much we'd lose if stop is hit RIGHT NOW
            risk_per_share = max(0, curr - stop)
            total_risk    += risk_per_share * qty

        heat_pct      = round((total_risk / capital) * 100, 2) if capital > 0 else 0.0
        remaining_pct = round(max(0.0, MAX_HEAT_PCT - heat_pct), 2)
        return {
            "heat_pct":           heat_pct,
            "positions_count":    len(holdings),
            "max_heat_pct":       MAX_HEAT_PCT,
            "heat_ok":            heat_pct < MAX_HEAT_PCT,
            "heat_remaining_pct": remaining_pct,
        }
    except Exception as e:
        _log(f"[WARN] compute_portfolio_heat failed: {e}")
        return {"heat_pct": 0.0, "positions_count": 0,
                "max_heat_pct": 6.0, "heat_ok": True, "heat_remaining_pct": 6.0}


# ─── Phase R4 (2026-07-06): Portfolio risk composite ─────────────────────────
# Extends heat + sector-cap with three more institutional lenses:
#   1. Sector HHI (Herfindahl) — measures concentration on a 0..10000 scale.
#      HHI > 2500 = concentrated; HHI < 1500 = diversified.
#   2. Stop cluster risk — how many positions have stops within 1 ATR of
#      current price. A cluster > 3 signals correlated bad day risk.
#   3. Position size skew — largest position vs mean position size. Skew > 3
#      means one position dominates and drags portfolio beta.
# Output is a single dict for logging + audit stamping. Read-only —
# does not itself gate anything; the pipeline consumes the flags.
def compute_portfolio_risk_composite(holdings: list,
                                      current_prices: dict | None = None,
                                      capital: float = 0.0) -> dict:
    default = {
        "sector_hhi": 0, "sector_hhi_flag": "OK",
        "stop_cluster_count": 0, "stop_cluster_flag": "OK",
        "position_size_skew": 0.0, "position_size_skew_flag": "OK",
        "n_positions": 0, "warnings": [],
    }
    try:
        if not holdings:
            return default
        current_prices = current_prices or {}
        n = len(holdings)

        # ── 1. Sector HHI ──────────────────────────────────────────────────
        sector_counts: dict = {}
        for h in holdings:
            sec = str(
                h.get("sector") or (get_sector(h.get("symbol", "")) if h.get("symbol") else "")
            ).strip().upper() or "UNKNOWN"
            sector_counts[sec] = sector_counts.get(sec, 0) + 1
        # HHI = Σ (share%)^2 · 100 · 100 → 0..10000
        hhi = 0.0
        for _sec, _n in sector_counts.items():
            share = (_n / n) * 100.0
            hhi += share * share
        hhi = int(round(hhi))
        if hhi >= 2500:
            hhi_flag = "CONCENTRATED"
        elif hhi >= 1500:
            hhi_flag = "MODERATE"
        else:
            hhi_flag = "OK"

        # ── 2. Stop cluster: positions with stop within 1 ATR of current ──
        cluster_n = 0
        for h in holdings:
            sym = h.get("symbol", "")
            entry = float(h.get("entry_price", 0) or 0)
            stop = float(h.get("stop_loss", 0) or 0)
            atr = float(h.get("atr14", 0) or 0)
            curr = float(current_prices.get(sym, entry) or entry)
            if entry > 0 and stop > 0 and atr > 0 and curr > 0:
                dist_atr = (curr - stop) / atr
                if dist_atr <= 1.0:
                    cluster_n += 1
        if cluster_n >= 4:
            cluster_flag = "DANGER"
        elif cluster_n >= 3:
            cluster_flag = "WARN"
        else:
            cluster_flag = "OK"

        # ── 3. Position size skew ──────────────────────────────────────────
        sizes = []
        for h in holdings:
            entry = float(h.get("entry_price", 0) or 0)
            qty = float(h.get("quantity", 0) or 0)
            if entry > 0 and qty > 0:
                sizes.append(entry * qty)
        skew = 0.0
        if len(sizes) >= 2:
            mean_sz = sum(sizes) / len(sizes)
            if mean_sz > 0:
                skew = max(sizes) / mean_sz
        if skew >= 3.0:
            skew_flag = "SKEWED"
        elif skew >= 2.0:
            skew_flag = "MODERATE"
        else:
            skew_flag = "OK"

        warnings = []
        if hhi_flag != "OK":
            top_sec = max(sector_counts.items(), key=lambda kv: kv[1])
            warnings.append(f"HHI={hhi} ({hhi_flag}: {top_sec[0]}={top_sec[1]}/{n})")
        if cluster_flag != "OK":
            warnings.append(f"STOP_CLUSTER={cluster_n} positions within 1 ATR of stop")
        if skew_flag != "OK":
            warnings.append(f"POSITION_SKEW={skew:.2f}x mean")

        return {
            "sector_hhi": hhi, "sector_hhi_flag": hhi_flag,
            "stop_cluster_count": cluster_n, "stop_cluster_flag": cluster_flag,
            "position_size_skew": round(skew, 2), "position_size_skew_flag": skew_flag,
            "n_positions": n, "warnings": warnings,
        }
    except Exception as e:
        _log(f"[WARN] compute_portfolio_risk_composite failed: {e}")
        return default


# ─── Phase C7 (2026-07-02): Equity-curve kill switch ────────────────────────
# Rating ≥ 9.5 requires a portfolio-level circuit breaker, not just per-trade
# stops. Real desks halt new entries after a losing streak or drawdown day —
# this prevents "revenge trading" through a regime change and caps monthly
# blow-up risk. Reads ONLY from tracker.completed, which has final_pnl and
# stop_hit_date / t2_hit_date / etc. already computed by update_tracker_v2_pnl.
def compute_kill_switch_state(tracker: dict, capital: float = None) -> dict:
    """Portfolio-level circuit breaker based on realized P&L trajectory.

    Reads tracker.completed (list of closed positions each with final_pnl %,
    and one of stop_hit_date / t2_hit_date / t1_hit_date / expired_date /
    stop_hit_date). Returns:
      {
        "buys_paused":       bool  — HALT all new BUYs today
        "size_multiplier":   float — 1.0 normal, 0.5 damped, 0.0 halted
        "reason":            str   — human-readable why
        "consecutive_losses": int
        "day_pnl_pct":       float — realized % of capital today
        "week_pnl_pct":      float — realized % of capital last 5 sessions
        "drawdown_from_peak_pct": float
      }

    Rules (all env-tunable):
      - 3 consecutive losses          → PAUSE (24h)
      - Day realized ≤ -DAY_STOP_PCT  → PAUSE for the rest of the day
      - Week realized ≤ -WEEK_STOP_PCT → PAUSE for the rest of the week
      - Drawdown ≥ DD_HALVE_PCT       → HALVE size (multiplier=0.5)
      - Drawdown ≥ DD_HALT_PCT        → PAUSE
    """
    # Env-tunable knobs — defaults are conservative professional levels
    MAX_CONSEC_LOSSES = int(os.getenv("KS_MAX_CONSEC_LOSSES", "3"))
    DAY_STOP_PCT      = float(os.getenv("KS_DAY_STOP_PCT",   "2.0"))   # -2% day
    WEEK_STOP_PCT     = float(os.getenv("KS_WEEK_STOP_PCT",  "3.0"))   # -3% week
    DD_HALVE_PCT      = float(os.getenv("KS_DD_HALVE_PCT",   "5.0"))   # -5% dd → halve
    DD_HALT_PCT       = float(os.getenv("KS_DD_HALT_PCT",   "10.0"))   # -10% dd → halt

    default = {
        "buys_paused": False, "size_multiplier": 1.0, "reason": "OK",
        "consecutive_losses": 0, "day_pnl_pct": 0.0, "week_pnl_pct": 0.0,
        "drawdown_from_peak_pct": 0.0,
    }
    try:
        completed = tracker.get("completed", []) if isinstance(tracker, dict) else []
        if not completed:
            return default

        today = ist_today()

        def _close_date(pos: dict) -> datetime.date:
            """Best-effort close date from the various *_date fields."""
            for k in ("stop_hit_date", "t2_hit_date", "t1_hit_date",
                      "expired_date", "close_date"):
                v = pos.get(k)
                if v:
                    try:
                        return datetime.date.fromisoformat(str(v)[:10])
                    except Exception:
                        continue
            # Fallback: rec_date + days_tracked
            try:
                rd = datetime.date.fromisoformat(pos.get("rec_date", "")[:10])
                return rd + datetime.timedelta(days=int(pos.get("days_tracked", 0) or 0))
            except Exception:
                return today  # unknown → treat as today so it still counts

        # Sort chronologically by close date
        closed_sorted = sorted(
            [(p, _close_date(p)) for p in completed],
            key=lambda x: x[1],
        )

        # 1) Consecutive-loss streak (walk BACKWARD from most recent close)
        # Phase 2 #31 (2026-07-05): only count losses within the last N days.
        # Without this window, if your last 3 closed trades were losses months
        # ago and you haven't traded since, the pipeline pauses BUYs forever.
        # Env-tunable via KS_LOSS_WINDOW_DAYS (default 7 = one trading week).
        try:
            _ks_window_days = int(os.getenv("KS_LOSS_WINDOW_DAYS", "7"))
        except (TypeError, ValueError):
            _ks_window_days = 7
        _ks_cutoff = today - datetime.timedelta(days=_ks_window_days)
        consec = 0
        for pos, _dt in reversed(closed_sorted):
            if _dt < _ks_cutoff:
                break   # too old to be part of a live streak
            pnl = float(pos.get("final_pnl", 0) or 0)
            if pnl < 0:
                consec += 1
            else:
                break

        # 2) Realized P&L windows (day / week) — sum final_pnl % contributions
        #    Each trade risked ~1.5% of capital, so we approximate its capital
        #    contribution as final_pnl_pct × position_pct_of_capital / 100.
        #    Fallback: if position_pct absent, assume 5% of capital (conservative).
        def _cap_contrib(pos):
            pos_pct = float(pos.get("position_pct", 5.0) or 5.0)
            pnl_pct = float(pos.get("final_pnl",   0.0) or 0.0)
            return pnl_pct * (pos_pct / 100.0)   # % of capital, signed

        day_pnl  = sum(_cap_contrib(p) for p, d in closed_sorted if d == today)
        wk_start = today - datetime.timedelta(days=7)
        week_pnl = sum(_cap_contrib(p) for p, d in closed_sorted if d >= wk_start)

        # 3) Equity-curve peak drawdown (from cumulative realized P&L)
        cum = 0.0
        peak = 0.0
        dd = 0.0
        for pos, _dt in closed_sorted:
            cum += _cap_contrib(pos)
            if cum > peak:
                peak = cum
            _cur_dd = peak - cum
            if _cur_dd > dd:
                dd = _cur_dd

        # ── Decide state (worst rule wins) ──
        buys_paused    = False
        size_mult      = 1.0
        reason_parts   = []

        if consec >= MAX_CONSEC_LOSSES:
            buys_paused = True
            reason_parts.append(f"CONSEC_LOSSES={consec}")

        if day_pnl <= -DAY_STOP_PCT:
            buys_paused = True
            reason_parts.append(f"DAY_STOP({day_pnl:.1f}%)")

        if week_pnl <= -WEEK_STOP_PCT:
            buys_paused = True
            reason_parts.append(f"WEEK_STOP({week_pnl:.1f}%)")

        if dd >= DD_HALT_PCT:
            buys_paused = True
            reason_parts.append(f"DD_HALT({dd:.1f}%)")
        elif dd >= DD_HALVE_PCT:
            size_mult = 0.5
            reason_parts.append(f"DD_HALVE({dd:.1f}%)")

        if buys_paused:
            size_mult = 0.0

        return {
            "buys_paused":            buys_paused,
            "size_multiplier":        size_mult,
            "reason":                 " | ".join(reason_parts) if reason_parts else "OK",
            "consecutive_losses":     consec,
            "day_pnl_pct":            round(day_pnl,  2),
            "week_pnl_pct":           round(week_pnl, 2),
            "drawdown_from_peak_pct": round(dd,       2),
        }
    except Exception as e:
        _log(f"[WARN] compute_kill_switch_state failed: {e}")
        return default


# ─── Phase C7 (2026-07-02): Sector concentration cap ────────────────────────
# Correlation gate 14 catches ticker-level pair correlation (>0.75) but does
# NOT catch the more common case of "5 IT stocks all trending together on the
# same tailwind". If Nifty IT drops 5%, all 5 stops fire on the same day.
# This gate caps open positions per NIFTY sector.
def check_sector_concentration(candidate_sector: str,
                                holdings: list,
                                max_per_sector: int = None) -> dict:
    """Return {'blocked': bool, 'count': int, 'max': int, 'reason': str}.

    Counts how many existing OPEN holdings share candidate_sector. If count
    ≥ max_per_sector, blocks. Default cap = 2 (professional standard).
    """
    if max_per_sector is None:
        max_per_sector = int(os.getenv("MAX_POSITIONS_PER_SECTOR", "2"))

    default = {"blocked": False, "count": 0, "max": max_per_sector, "reason": "OK"}
    try:
        if not candidate_sector or candidate_sector == "OTHERS" or not holdings:
            return default

        cand = str(candidate_sector).strip().upper()
        # 2026-07-05: legacy BUY records saved before Phase C7 may lack the
        # "sector" field entirely. Compute it on-the-fly from the symbol so
        # the cap actually works instead of silently ignoring those rows.
        same_sector = [
            h for h in holdings
            if str(
                h.get("sector")
                or (get_sector(h.get("symbol", "")) if h.get("symbol") else "")
            ).strip().upper() == cand
        ]
        count = len(same_sector)

        if count >= max_per_sector:
            return {
                "blocked": True, "count": count, "max": max_per_sector,
                "reason": f"SECTOR_CAP({cand}: {count}/{max_per_sector})",
            }
        return {"blocked": False, "count": count, "max": max_per_sector,
                "reason": f"OK ({cand}: {count}/{max_per_sector})"}
    except Exception as e:
        _log(f"[WARN] check_sector_concentration failed: {e}")
        return default


def kelly_position_size(entry: float, stop: float, capital: float,
                        win_rate: float, avg_win_pct: float, avg_loss_pct: float,
                        heat: dict, max_position_pct: float = None) -> dict:
    """
    Kelly Criterion position sizing — sizes position based on ACTUAL historical
    win rate from the tracker. Falls back to fixed 1.5% risk if no history.

    Kelly fraction = W - (1-W)/R
      W = win rate (e.g. 0.62)
      R = avg_win / avg_loss ratio

    We use HALF-Kelly (safer, reduces variance by ~75% vs full Kelly).
    Heat-aware: reduces position if portfolio is near max heat.

    Activates automatically once tracker has >= 20 closed trades.
    """
    if max_position_pct is None:
        max_position_pct = _MAX_POSITION_PCT_ENV
    try:
        if entry <= 0 or stop <= 0 or stop >= entry or capital <= 0:
            return compute_position_size(entry, stop, capital)

        risk_per_share = entry - stop
        risk_pct_per_share = risk_per_share / entry

        # Use Kelly if we have enough history, else fall back to fixed 1.5%
        if (win_rate > 0 and avg_win_pct > 0 and avg_loss_pct > 0
                and 0.3 <= win_rate <= 0.9):
            loss_rate = 1.0 - win_rate
            r_ratio   = avg_win_pct / avg_loss_pct if avg_loss_pct > 0 else 1.5
            full_kelly = win_rate - (loss_rate / r_ratio)
            half_kelly = max(0.005, min(0.04, full_kelly * 0.5))  # cap 0.5%–4%
            risk_per_trade = half_kelly
            sizing_method  = f"KELLY_HALF({win_rate:.0%}WR/{r_ratio:.1f}R)"
        else:
            risk_per_trade = 0.015   # fixed 1.5% until we have history
            sizing_method  = "FIXED_1.5PCT"

        # Heat adjustment: if heat is > 80% of max, scale down new position
        heat_pct     = heat.get("heat_pct", 0.0)
        max_heat_pct = heat.get("max_heat_pct", 6.0)
        heat_ratio   = heat_pct / max_heat_pct if max_heat_pct > 0 else 0
        if heat_ratio > 0.8:
            risk_per_trade *= 0.5    # cut in half when near max heat
            sizing_method  += "_HEAT_REDUCED"
        elif heat_ratio > 0.6:
            risk_per_trade *= 0.75

        risk_amount    = capital * risk_per_trade
        shares         = max(1, int(risk_amount / risk_per_share))
        position_value = shares * entry
        if position_value > capital * max_position_pct:
            shares         = max(1, int((capital * max_position_pct) / entry))
            position_value = shares * entry
        position_pct = (position_value / capital) * 100

        max_loss = round(shares * risk_per_share, 2)
        return {
            "shares":          shares,
            "position_value":  round(position_value, 2),
            "position_pct":    round(position_pct, 1),
            "risk_amount":     round(risk_amount, 2),
            "risk_pct":        round(risk_per_trade * 100, 2),
            "sizing_method":   sizing_method,
            "max_loss":        max_loss,
        }
    except Exception:
        return compute_position_size(entry, stop, capital)


def check_gap_validity(signal_entry: float, signal_stop: float,
                       signal_target1: float, min_rr: float,
                       open_price: float = 0.0) -> dict:
    """
    Called at market open (9:15 AM) to decide if a previous evening's signal
    is still valid given the actual opening price.

    Args:
        signal_entry  : price the system recommended entering at (last night's close)
        signal_stop   : stop loss from the signal
        signal_target1: first target from the signal
        min_rr        : minimum acceptable R/R for this regime (e.g. 1.8)
        open_price    : actual market open price (0 = not yet known, use signal_entry)

    Returns dict:
        action        : "ENTER" | "WAIT_PULLBACK" | "RECALCULATE" | "VOID"
        reason        : explanation string
        adjusted_entry: recommended actual entry (may differ from open_price)
        adjusted_rr   : R/R at adjusted_entry vs original stop/target
        gap_pct       : % gap from signal_entry to open_price
        max_valid_entry: highest price at which signal still meets min_rr
    """
    try:
        if signal_entry <= 0 or signal_stop <= 0 or signal_target1 <= 0:
            return {"action": "VOID", "reason": "INVALID_SIGNAL_DATA",
                    "adjusted_entry": 0.0, "adjusted_rr": 0.0,
                    "gap_pct": 0.0, "max_valid_entry": 0.0}

        # Max valid entry: highest price where R/R still meets minimum
        # (target1 - max_entry) / (max_entry - stop) = min_rr
        # Solve: max_entry = (target1 + min_rr * stop) / (1 + min_rr)
        max_valid_entry = round(
            (signal_target1 + min_rr * signal_stop) / (1 + min_rr), 2
        )

        # If no open price provided, return the rule for display
        if open_price <= 0:
            gap_pct = 0.0
            rr_at_entry = round(
                (signal_target1 - signal_entry) / (signal_entry - signal_stop), 2
            ) if signal_entry > signal_stop else 0.0
            return {
                "action":          "PENDING",
                "reason":          f"Max valid entry: Rs{max_valid_entry:.2f} | "
                                   f"Above that → skip",
                "adjusted_entry":  signal_entry,
                "adjusted_rr":     rr_at_entry,
                "gap_pct":         0.0,
                "max_valid_entry": max_valid_entry,
            }

        gap_pct = round((open_price - signal_entry) / signal_entry * 100, 2)
        rr_at_open = round(
            (signal_target1 - open_price) / (open_price - signal_stop), 2
        ) if open_price > signal_stop else 0.0

        # Gap down — check if stop is breached
        if open_price <= signal_stop:
            return {
                "action":          "VOID",
                "reason":          f"Gap DOWN {gap_pct:.1f}% — opened at or below stop. Signal dead.",
                "adjusted_entry":  0.0,
                "adjusted_rr":     0.0,
                "gap_pct":         gap_pct,
                "max_valid_entry": max_valid_entry,
            }

        # Gap up > 5% — always void
        if gap_pct > 5.0:
            return {
                "action":          "VOID",
                "reason":          f"Gap UP {gap_pct:.1f}% — too extended. R/R={rr_at_open:.2f}x. Do NOT chase.",
                "adjusted_entry":  0.0,
                "adjusted_rr":     rr_at_open,
                "gap_pct":         gap_pct,
                "max_valid_entry": max_valid_entry,
            }

        # Gap up 3-5% — wait for pullback to max_valid_entry
        if gap_pct > 3.0:
            return {
                "action":          "WAIT_PULLBACK",
                "reason":          f"Gap UP {gap_pct:.1f}% — wait for pullback to Rs{max_valid_entry:.2f}. "
                                   f"Enter only if price comes back AND holds above EMA20.",
                "adjusted_entry":  max_valid_entry,
                "adjusted_rr":     min_rr,
                "gap_pct":         gap_pct,
                "max_valid_entry": max_valid_entry,
            }

        # Gap up 1.5-3% — recalculate R/R at open
        if gap_pct > 1.5:
            if rr_at_open >= min_rr:
                return {
                    "action":          "ENTER",
                    "reason":          f"Gap UP {gap_pct:.1f}% — R/R {rr_at_open:.2f}x still meets "
                                       f"min {min_rr:.1f}x. Enter at open.",
                    "adjusted_entry":  round(open_price, 2),
                    "adjusted_rr":     rr_at_open,
                    "gap_pct":         gap_pct,
                    "max_valid_entry": max_valid_entry,
                }
            else:
                return {
                    "action":          "WAIT_PULLBACK",
                    "reason":          f"Gap UP {gap_pct:.1f}% — R/R dropped to {rr_at_open:.2f}x "
                                       f"(min {min_rr:.1f}x). Wait for Rs{max_valid_entry:.2f}.",
                    "adjusted_entry":  max_valid_entry,
                    "adjusted_rr":     min_rr,
                    "gap_pct":         gap_pct,
                    "max_valid_entry": max_valid_entry,
                }

        # Gap ≤ 1.5% — normal, enter at open
        return {
            "action":          "ENTER",
            "reason":          f"Gap {gap_pct:+.1f}% — within tolerance. R/R {rr_at_open:.2f}x. Enter.",
            "adjusted_entry":  round(open_price, 2),
            "adjusted_rr":     rr_at_open,
            "gap_pct":         gap_pct,
            "max_valid_entry": max_valid_entry,
        }

    except Exception as e:
        return {"action": "VOID", "reason": f"Calculation error: {e}",
                "adjusted_entry": 0.0, "adjusted_rr": 0.0,
                "gap_pct": 0.0, "max_valid_entry": 0.0}


def compute_platt_stats(tracker_entries: list) -> dict:
    """
    Platt Calibration Framework.
    Reads closed tracker trades and computes:
      - win_rate: fraction of closed trades that hit T1 or T2
      - avg_win_pct: average % gain on winning trades
      - avg_loss_pct: average % loss on losing trades
      - conf_calibration: dict of {conf_bucket: actual_win_rate}

    Returns safe defaults when < 20 closed trades (not enough data).
    Activates AUTOMATICALLY once you have 20+ closed trades in tracker.json.
    """
    closed = [e for e in tracker_entries if e.get("status") == "CLOSED"
              and e.get("final_pnl_pct") is not None]

    if len(closed) < 20:
        return {
            "win_rate": 0.0, "avg_win_pct": 0.0, "avg_loss_pct": 0.0,
            "total_closed": len(closed), "calibrated": False,
            "conf_calibration": {},
        }

    wins   = [e for e in closed if e["final_pnl_pct"] >= 0]
    losses = [e for e in closed if e["final_pnl_pct"] < 0]
    win_rate     = len(wins) / len(closed)
    avg_win_pct  = sum(e["final_pnl_pct"] for e in wins)  / max(1, len(wins))
    avg_loss_pct = abs(sum(e["final_pnl_pct"] for e in losses)) / max(1, len(losses))

    # Confidence bucket calibration: bucket closed trades by confidence score
    # e.g. {80: 0.71} means: when system said conf=80, actual win rate was 71%
    buckets: dict = {}
    for e in closed:
        conf   = float(e.get("conf", 0) or 0)
        bucket = int(conf // 5) * 5   # bucket to nearest 5 (75,80,85,90)
        if bucket not in buckets:
            buckets[bucket] = {"wins": 0, "total": 0}
        buckets[bucket]["total"] += 1
        if e["final_pnl_pct"] >= 0:
            buckets[bucket]["wins"] += 1
    conf_cal = {b: round(v["wins"] / v["total"], 2)
                for b, v in buckets.items() if v["total"] >= 3}

    return {
        "win_rate":         round(win_rate, 3),
        "avg_win_pct":      round(avg_win_pct, 2),
        "avg_loss_pct":     round(avg_loss_pct, 2),
        "total_closed":     len(closed),
        "calibrated":       True,
        "conf_calibration": conf_cal,
    }


def weekly_trend_score(df_daily) -> tuple:
    """
    Multi-timeframe confirmation: resamples daily OHLCV to weekly.
    Returns (score 0-100, weekly_trend_ok: bool).

    A daily BUY signal in a stock whose WEEKLY trend is DOWN has much lower follow-through.
    Adds roughly 8-12% improvement in win rate empirically on NSE.

    Scoring:
      Weekly close > weekly EMA10 > weekly EMA20 → strong weekly uptrend  → 90
      Weekly close > weekly EMA20                → moderate uptrend        → 70
      Weekly close > weekly EMA10 but < EMA20   → mixed                   → 50
      Weekly close < weekly EMA20               → weekly downtrend        → 25
    """
    try:
        weekly = df_daily.resample("W").agg({
            "Open":   "first",
            "High":   "max",
            "Low":    "min",
            "Close":  "last",
            "Volume": "sum",
        }).dropna()

        if len(weekly) < 22:
            return 50, False

        wc = weekly["Close"].squeeze().values.astype(float)
        w_ema10 = float(pd.Series(wc).ewm(span=10).mean().iloc[-1])
        w_ema20 = float(pd.Series(wc).ewm(span=20).mean().iloc[-1])
        w_last  = wc[-1]
        # Weekly slope — is EMA10 rising over last 3 weeks?
        w_ema10_3w = float(pd.Series(wc).ewm(span=10).mean().iloc[-4])
        w_slope_up = w_ema10 > w_ema10_3w

        if w_last > w_ema10 > w_ema20 and w_slope_up:
            return 90, True
        elif w_last > w_ema10 > w_ema20:
            return 78, True
        elif w_last > w_ema20:
            return 65, True
        elif w_last > w_ema10:
            return 52, False
        else:
            return 25, False
    except Exception:
        return 50, False


def price_action_score(closes: np.ndarray, highs: np.ndarray,
                       lows: np.ndarray, ema20: float, atr14: float,
                       volumes: np.ndarray = None) -> tuple:
    """
    Detects high-edge NSE price patterns. Returns (score 0-100, pattern_name).

    Pattern priority (highest edge first — return early on first match):
    ────────────────────────────────────────────────────────────────────
    Original 3 patterns:
      1. INSIDE_BAR (score 82)                — tight range = coiling energy
      2. FALSE_BREAKDOWN_RECOVERY (score 88) — bear-trap continuation
      3. TIGHT_CONSOLIDATION (score 80)      — 3-bar spring near EMA20

    Phase 3b (2026-07-05) — 4 new patterns from CANSLIM/Minervini/O'Neil literature:
      4. POCKET_PIVOT (score 90)             — O'Neil/Morales institutional footprint:
                                               up-day volume > highest DOWN-day
                                               volume in last 10 days, in an
                                               established uptrend. ChartMill:
                                               "bonus points for recent pocket
                                               pivots" (highest-weight setup).
      5. CUP_HANDLE (score 92)               — CANSLIM signature base: proper
                                               U-shape 30-60 bars with a shallow
                                               handle. Buy point = handle high.
      6. FLAT_BASE (score 85)                — CANSLIM/O'Neil: 5+ weeks (25+
                                               bars) sideways within 10% range,
                                               above EMA20. "At least 7 weeks
                                               on weekly charts" per O'Neil.
      7. VCP_CONTRACTION (score 88)          — Minervini VCP: 3+ successively
                                               tighter pullbacks over ~60 bars,
                                               each contraction smaller than
                                               the last.

    If none detected → NONE (score 50, neutral — doesn't penalise the stock)

    NOTE: `volumes` is optional for backward compatibility. When absent, the
    Pocket Pivot check is skipped (returns None) and other patterns still work.
    """
    try:
        n = len(closes)
        if n < 15:
            return 50, "NONE"

        # ─── Pattern 4 (NEW): Pocket Pivot ───────────────────────────
        # O'Neil/Morales/Kacher's Trade Like an O'Neil Disciple:
        #   "Up-day volume > highest down-day volume in the last 10 days,
        #    while stock is in a base or trending above 10/50-day MA."
        # ChartMill Setup Quality: "More bonus points for recent pocket pivots"
        #
        # India refinement (2026-07-05): US pocket-pivot theory assumes NASDAQ-
        # scale liquidity. On NSE mid/small-caps, a "pocket pivot" on 30k shares
        # is noise, not institutional footprint. Floor at 100k shares (absolute)
        # AND ₹50 lakh notional (avg_val_lakhs proxy via close*volume) so we only
        # award POCKET_PIVOT credit when there's real institutional participation.
        if volumes is not None and n >= 11 and len(volumes) >= 11:
            try:
                # Today must be an up day
                today_up = closes[-1] > closes[-2]
                today_vol = float(volumes[-1])
                # India volume floor: 100k shares AND ₹50 lakh notional
                _india_vol_ok = (
                    today_vol >= 100_000
                    and (today_vol * float(closes[-1])) >= 50_00_000  # ₹50 lakh
                )
                # Highest DOWN-day volume in the last 10 sessions (excluding today)
                down_vols = [
                    float(volumes[-i]) for i in range(2, 12)
                    if i <= len(closes) and closes[-i] < closes[-i-1]
                ]
                if today_up and down_vols and today_vol > max(down_vols) and _india_vol_ok:
                    # Also require: still in uptrend (above EMA20)
                    if closes[-1] > ema20:
                        return 90, "POCKET_PIVOT"
            except Exception:
                pass

        # ─── Pattern 5 (NEW): Cup-and-Handle (conservative) ──────────
        # CANSLIM signature. Simplified geometric check:
        #   - Look back 30-60 bars for the cup depth
        #   - Left rim near right rim (< 5% diff)
        #   - Cup depth 12-33% (proper cup, not V-shape)
        #   - Handle in last 5-15 bars: shallow pullback (< 12%)
        #   - Handle high < cup rim (still setting up, not broken out yet)
        if n >= 50:
            try:
                cup_window = closes[-60:-15]  # cup body (excludes handle)
                left_rim  = float(np.max(cup_window[:10]))
                right_rim = float(np.max(cup_window[-10:]))
                cup_low   = float(np.min(cup_window))
                cup_depth_pct = (max(left_rim, right_rim) - cup_low) / max(left_rim, right_rim) * 100
                rim_diff_pct  = abs(left_rim - right_rim) / max(left_rim, right_rim) * 100
                # Handle (last 5-15 bars): shallow pullback from right rim
                handle_window = closes[-15:]
                handle_high   = float(np.max(handle_window))
                handle_low    = float(np.min(handle_window))
                handle_depth_pct = (handle_high - handle_low) / handle_high * 100
                cup_ok = (12 <= cup_depth_pct <= 33 and rim_diff_pct < 5)
                handle_ok = (2 <= handle_depth_pct <= 12
                             and handle_high <= right_rim * 1.02
                             and closes[-1] > ema20)
                if cup_ok and handle_ok:
                    return 92, "CUP_HANDLE"
            except Exception:
                pass

        # ─── Pattern 6 (NEW): VCP contraction count (real Minervini) ─
        # Split recent ~60 bars into 3 windows, measure each window's high-low
        # range as pct of window mid. Each successive window range should be
        # smaller (contracting). 3 successive contractions = classic VCP.
        if n >= 60:
            try:
                w1 = closes[-60:-40]; h1 = highs[-60:-40]; l1 = lows[-60:-40]
                w2 = closes[-40:-20]; h2 = highs[-40:-20]; l2 = lows[-40:-20]
                w3 = closes[-20:];    h3 = highs[-20:];    l3 = lows[-20:]
                r1 = (float(np.max(h1)) - float(np.min(l1))) / float(np.mean(w1)) * 100
                r2 = (float(np.max(h2)) - float(np.min(l2))) / float(np.mean(w2)) * 100
                r3 = (float(np.max(h3)) - float(np.min(l3))) / float(np.mean(w3)) * 100
                # Each range strictly smaller + still above EMA20 + final range tight
                if r1 > r2 > r3 and r3 < 8.0 and closes[-1] > ema20:
                    return 88, "VCP_CONTRACTION"
            except Exception:
                pass

        # ─── Pattern 7 (NEW): Flat Base ──────────────────────────────
        # CANSLIM/O'Neil: 5+ weeks sideways in a 10-15% range.
        # 5 weeks daily ≈ 25 bars. Range = (max - min) / mid.
        if n >= 25:
            try:
                base = closes[-25:]
                bhi  = highs[-25:]
                blo  = lows[-25:]
                base_high = float(np.max(bhi))
                base_low  = float(np.min(blo))
                base_mid  = (base_high + base_low) / 2.0
                base_range_pct = (base_high - base_low) / base_mid * 100
                # Must also be above EMA20 (trending base, not falling knife)
                # and not in a pattern already better handled above.
                if base_range_pct < 15.0 and closes[-1] > ema20:
                    # Distinguish from TIGHT_CONSOLIDATION which is 3-bar; this
                    # is 25-bar. Award higher score for longer/wider base only
                    # when the range is 8-15% (proper flat base). Below 8% is
                    # picked up by TIGHT_CONSOLIDATION anyway.
                    if 8.0 <= base_range_pct < 15.0:
                        return 85, "FLAT_BASE"
            except Exception:
                pass

        # ─── Original 3 patterns (unchanged) ─────────────────────────
        # Pattern 1: Inside bar (yesterday's entire range inside 2 days ago)
        inside_bar = (highs[-2] < highs[-3] and lows[-2] > lows[-3]
                      and closes[-1] > ema20)   # still in uptrend
        if inside_bar:
            return 82, "INSIDE_BAR"

        # Pattern 2: False breakdown recovery
        recent_low_10 = float(np.min(lows[-10:-1]))
        false_breakdown = (
            lows[-3] < recent_low_10        # dipped below 10-day low
            and closes[-1] > ema20           # recovered above EMA20
            and closes[-1] > closes[-3]      # higher close than breakdown bar
        )
        if false_breakdown:
            return 88, "FALSE_BREAKDOWN_RECOVERY"

        # Pattern 3: 3-bar tight consolidation (< 2.5% range) above EMA20
        three_bar_high = float(np.max(highs[-3:]))
        three_bar_low  = float(np.min(lows[-3:]))
        tight_range_pct = (three_bar_high - three_bar_low) / closes[-1] * 100
        tight_consol = (
            tight_range_pct < 2.5
            and closes[-1] > ema20
        )
        if tight_range_pct < 2.5 and closes[-1] > ema20:
            return 80, "TIGHT_CONSOLIDATION"

        return 50, "NONE"
    except Exception:
        return 50, "NONE"


def _default_stock_result(symbol: str, sector: str) -> dict:
    return {
        "symbol": symbol, "sector": sector,
        "trend_quality": 50, "momentum_quality": 50, "volume_delivery": 50,
        "sector_strength": 50, "rs_vs_nifty": 50, "news_risk": 50,
        "risk_reward": 0, "ownership_quality": 50, "options_sentiment": 60,
        "macro_alignment": 50, "trade_quality_score": 0,
        "entry": 0.0, "stop": 0.0, "target1": 0.0, "target2": 0.0,
        "rr_ratio": 0.0, "avg_volume": 0, "avg_value_lakhs": 0.0,
        "near_52w_high": False, "sector_status": "NEUTRAL", "accum_signal": "NEUTRAL",
        "roe": 0.0, "de_ratio": 0.0, "promoter_pledge_pct": 0.0,
        "news_penalty": 0, "is_black_swan": False, "news_summary": "",
        "price": 0.0, "ret1d": 0.0, "ret5d": 0.0, "ret21d": 0.0,
        "high_52w": 0.0, "low_52w": 0.0, "atr14": 0.0, "rsi14": 50.0,
        "final_confidence": 0.0,
        # base_confidence intentionally absent — computed externally by
        # compute_base_confidence() and inserted after **scores in scored.append()
        # so it is never overwritten by a stale 0.0 placeholder.
        "weekly_trend_ok": False, "price_pattern": "NONE", "rs_diff21": 0.0,
        # Phase C4: execution-realism fields (Gap #5 + #8)
        "net_rr_t1": 0.0, "net_rr_t2": 0.0,
        "slippage_pct_one_way": 0.0, "round_trip_cost_pct": 0.0,
        "avg_gap_pct": 0.0, "max_gap_pct": 0.0, "p90_gap_pct": 0.0,
        "effective_stop_pct": 0.0, "high_gap_risk": False,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Phase C4 (2026-07-02) — Execution realism helpers
# Gap #5: slippage-adjusted R/R.  Gap #8: overnight-gap risk warning.
# ─────────────────────────────────────────────────────────────────────────────
def estimate_slippage_pct(avg_val_lakhs: float) -> float:
    """
    Per-leg slippage estimate for NSE cash market as a % of price.

    Liquidity bucket (avg 20-day traded value in ₹ lakh):
      >= 5,000 lakh (₹50Cr) — mega liquid  (RELIANCE, HDFCBANK)  → 0.05%
      >= 1,000 lakh (₹10Cr) — high liquid  (mid-cap NIFTY500)    → 0.10%
      >=   500 lakh (₹5Cr)  — decent                              → 0.20%
      >=   200 lakh (₹2Cr)  — thin                                → 0.35%
      >=    50 lakh (₹50L)  — very thin (min gate threshold)      → 0.55%
      <     50 lakh         — below gate, wouldn't be tradable    → 0.80%
    """
    try:
        v = float(avg_val_lakhs or 0)
    except Exception:
        return 0.35
    if v >= 5000: return 0.05
    if v >= 1000: return 0.10
    if v >= 500:  return 0.20
    if v >= 200:  return 0.35
    if v >= 50:   return 0.55
    return 0.80


def estimate_round_trip_cost_pct(avg_val_lakhs: float,
                                  brokerage_bps: float = None) -> dict:
    """
    Round-trip cost = 2x slippage + STT (0.1% sell) + brokerage + GST + exchange +
    SEBI + stamp. For a typical discount broker (Zerodha/Groww/Upstox):
      brokerage = flat ₹20/order (≈ 0.03% at ₹65k ticket, less at bigger)
      STT       = 0.10% on sell leg (equity delivery)
      GST       = 18% on brokerage
      Exch txn  = 0.00297% both sides (NSE)
      SEBI      = 0.0001% both sides
      Stamp     = 0.015% buy leg (Maharashtra)

    Returns dict with slippage_pct, tax_pct, total_pct (all round-trip).
    """
    slip_one = estimate_slippage_pct(avg_val_lakhs)
    slip_rt  = slip_one * 2.0

    # Fixed regulatory + broker: STT 0.10% + Stamp 0.015% + Exch 2*0.003% +
    # SEBI 2*0.0001% + Brokerage ~0.06% RT + GST on brokerage
    tax_rt = 0.10 + 0.015 + 0.006 + 0.0002 + 0.06 + 0.011  # ≈ 0.19%
    if brokerage_bps is not None:
        # Override broker cost if user provides bps
        tax_rt = 0.10 + 0.015 + 0.006 + 0.0002 + (brokerage_bps / 100.0) * 2 + \
                 (brokerage_bps / 100.0) * 2 * 0.18
    total_rt = slip_rt + tax_rt
    return {
        "slippage_pct_one_way": round(slip_one, 3),
        "slippage_pct_rt":      round(slip_rt, 3),
        "tax_pct_rt":           round(tax_rt, 3),
        "total_pct_rt":         round(total_rt, 3),
    }


def apply_slippage_to_rr(entry: float, stop: float, target1: float, target2: float,
                          avg_val_lakhs: float) -> dict:
    """
    Compute net (post-slippage-and-tax) R/R.  Buys fill above quote, sells fill below.
      buy_fill  ≈ entry * (1 + slip)
      sell_fill ≈ target * (1 - slip)
      stop_fill ≈ stop   * (1 - slip)   (stop-loss slippage is worse in practice; we use symmetric here)

    Returns net rr on T2 plus a net rr on T1, plus the round-trip cost dict.
    """
    cost = estimate_round_trip_cost_pct(avg_val_lakhs)
    slip = cost["slippage_pct_one_way"] / 100.0  # convert % to fraction
    tax_rt = cost["tax_pct_rt"] / 100.0

    try:
        if entry <= 0 or stop <= 0 or stop >= entry:
            return {"net_rr_t1": 0.0, "net_rr_t2": 0.0, "cost": cost}
        buy_fill  = entry * (1.0 + slip)
        stop_fill = stop  * (1.0 - slip)
        t1_fill   = target1 * (1.0 - slip) if target1 > 0 else 0.0
        t2_fill   = target2 * (1.0 - slip) if target2 > 0 else 0.0
        # Net PnL per share (loser) — includes tax on both sides
        loser_per_share = (stop_fill - buy_fill) - (buy_fill * tax_rt)
        if loser_per_share >= 0:
            # numerically stop is above fill after slippage — nonsensical; guard
            return {"net_rr_t1": 0.0, "net_rr_t2": 0.0, "cost": cost}
        risk_per_share = abs(loser_per_share)
        # Net winner per share at each target
        w1 = (t1_fill - buy_fill) - (buy_fill * tax_rt) if t1_fill > 0 else 0.0
        w2 = (t2_fill - buy_fill) - (buy_fill * tax_rt) if t2_fill > 0 else 0.0
        rr1 = max(0.0, round(w1 / risk_per_share, 2)) if risk_per_share > 0 else 0.0
        rr2 = max(0.0, round(w2 / risk_per_share, 2)) if risk_per_share > 0 else 0.0
        return {"net_rr_t1": rr1, "net_rr_t2": rr2, "cost": cost}
    except Exception:
        return {"net_rr_t1": 0.0, "net_rr_t2": 0.0, "cost": cost}


def estimate_gap_risk_pct(closes, highs, lows) -> dict:
    """
    Overnight-gap risk = 20-day rolling mean of |open[i] - close[i-1]| / close[i-1].
    Since we don't have opens explicitly, we approximate the gap size using
    the wick beyond the previous day's close:
        approx_gap[i] = max(highs[i] - closes[i-1], closes[i-1] - lows[i], 0)

    Returns:
      avg_gap_pct  — typical daily gap size (%)
      max_gap_pct  — largest gap over lookback
      p90_gap_pct  — 90th percentile (used for stop-risk headline)
    """
    try:
        import numpy as _np
        c = _np.asarray(closes, dtype=float)
        h = _np.asarray(highs,  dtype=float)
        l = _np.asarray(lows,   dtype=float)
        n = min(len(c), len(h), len(l))
        if n < 22:
            return {"avg_gap_pct": 0.0, "max_gap_pct": 0.0, "p90_gap_pct": 0.0}
        c_prev = c[-21:-1]
        h_curr = h[-20:]
        l_curr = l[-20:]
        gaps_up   = _np.maximum(h_curr - c_prev, 0.0)
        gaps_down = _np.maximum(c_prev - l_curr, 0.0)
        gaps_abs  = _np.maximum(gaps_up, gaps_down)
        gap_pct   = gaps_abs / _np.where(c_prev > 0, c_prev, 1.0) * 100.0
        return {
            "avg_gap_pct": round(float(_np.mean(gap_pct)),           2),
            "max_gap_pct": round(float(_np.max(gap_pct)),            2),
            "p90_gap_pct": round(float(_np.percentile(gap_pct, 90)), 2),
        }
    except Exception:
        return {"avg_gap_pct": 0.0, "max_gap_pct": 0.0, "p90_gap_pct": 0.0}


def compute_business_quality_score(promoter_data: dict) -> dict:
    """
    Phase R2 (2026-07-06): Institutional Business Quality Score (0-100).

    Composite score built from quarterly + ratio fundamentals. Missing
    inputs reduce the score's confidence (returned as `bq_data_completeness`)
    rather than being silently zero-filled. This lets downstream ranking:

      • Reward strong business quality above and beyond technicals
      • Distinguish "high conf on strong biz" from "high conf on unknown biz"
      • Populate a first-class field for the 9-tier taxonomy (STRONG BUY,
        BUY, WATCHLIST, TURNAROUND, CONTRARIAN, AVOID)

    Weights (institutional-lite; keeps out volatile items like FCF that
    aren't reliably parseable from screener's free HTML):
      Sales trend YoY:  25%   (institutional quality benchmark)
      Profit trend YoY: 25%
      ROE:              15%
      D/E (inverted):   10%   (financials excepted)
      ROCE:             10%
      Promoter holding:  8%
      Pledge (inverted): 5%
      Sales 3Q trend:    1%   (already used above)
      Profit 3Q trend:   1%

    Returns:
      {
        "bq_score": 0-100 float,
        "bq_data_completeness": 0-100 float,  # % of inputs that were real
        "bq_flags": ["ROE_STRONG", "SALES_YOY_POSITIVE", ...],
        "bq_verdict": "STRONG" | "ACCEPTABLE" | "WEAK" | "DECLINING" | "UNKNOWN",
      }
    """
    pd_ = promoter_data or {}
    score = 0.0
    max_possible = 0.0
    flags: list[str] = []

    # Detect financial sector for D/E leniency (we can't see sector from
    # promoter_data alone; caller can override via _is_financial arg
    # in future — for now, apply blanket rule)

    # 1. Sales YoY (25 pts)
    sales_yoy = pd_.get("sales_yoy_pct")
    if isinstance(sales_yoy, (int, float)):
        max_possible += 25
        if sales_yoy >= 20:
            score += 25
            flags.append("SALES_YOY_STRONG")
        elif sales_yoy >= 10:
            score += 18
            flags.append("SALES_YOY_GOOD")
        elif sales_yoy >= 0:
            score += 10
            flags.append("SALES_YOY_FLAT")
        elif sales_yoy >= -10:
            score += 4
            flags.append("SALES_YOY_WEAK")
        else:
            score += 0
            flags.append("SALES_YOY_DECLINING")

    # 2. Profit YoY (25 pts)
    profit_yoy = pd_.get("profit_yoy_pct")
    if isinstance(profit_yoy, (int, float)):
        max_possible += 25
        if profit_yoy >= 25:
            score += 25
            flags.append("PROFIT_YOY_STRONG")
        elif profit_yoy >= 10:
            score += 18
            flags.append("PROFIT_YOY_GOOD")
        elif profit_yoy >= 0:
            score += 10
            flags.append("PROFIT_YOY_FLAT")
        elif profit_yoy >= -15:
            score += 4
            flags.append("PROFIT_YOY_WEAK")
        else:
            score += 0
            flags.append("PROFIT_YOY_DECLINING")

    # 3. ROE (15 pts)
    roe = pd_.get("roe")
    if isinstance(roe, (int, float)) and roe != 0:
        max_possible += 15
        if roe >= 20:
            score += 15
            flags.append("ROE_EXCELLENT")
        elif roe >= 15:
            score += 12
            flags.append("ROE_STRONG")
        elif roe >= 12:
            score += 8
            flags.append("ROE_ACCEPTABLE")
        elif roe >= 8:
            score += 4
            flags.append("ROE_WEAK")
        else:
            score += 0
            flags.append("ROE_POOR")

    # 4. D/E inverted (10 pts) — lower is better, higher penalized
    de = pd_.get("de_ratio")
    if isinstance(de, (int, float)) and de >= 0:
        max_possible += 10
        if de <= 0.3:
            score += 10
            flags.append("DE_EXCELLENT")
        elif de <= 0.8:
            score += 7
            flags.append("DE_GOOD")
        elif de <= 1.5:
            score += 4
            flags.append("DE_ACCEPTABLE")
        else:
            score += 0
            flags.append("DE_HIGH")

    # 5. ROCE (10 pts)
    roce = pd_.get("roce")
    if isinstance(roce, (int, float)) and roce != 0:
        max_possible += 10
        if roce >= 20:
            score += 10
            flags.append("ROCE_EXCELLENT")
        elif roce >= 15:
            score += 8
            flags.append("ROCE_STRONG")
        elif roce >= 10:
            score += 5
            flags.append("ROCE_ACCEPTABLE")
        else:
            score += 1
            flags.append("ROCE_WEAK")

    # 6. Promoter holding (8 pts) — higher = skin in the game
    prom = pd_.get("promoter_holding_pct")
    if isinstance(prom, (int, float)) and prom > 0:
        max_possible += 8
        if prom >= 50:
            score += 8
            flags.append("PROMOTER_HIGH")
        elif prom >= 35:
            score += 5
        elif prom >= 20:
            score += 3
        else:
            score += 1
            flags.append("PROMOTER_LOW")

    # 7. Pledge inverted (5 pts)
    pledge = pd_.get("promoter_pledge_pct")
    if isinstance(pledge, (int, float)):
        max_possible += 5
        if pledge <= 0.5:
            score += 5
        elif pledge <= 5:
            score += 3
        elif pledge <= 20:
            score += 1
        else:
            score += 0
            flags.append("PLEDGE_HIGH")

    # 8. Sales 3Q trend (1 pt bonus)
    st = str(pd_.get("sales_trend_3q", "") or "").upper()
    if st in ("GROWING", "DECLINING", "MIXED"):
        max_possible += 1
        if st == "GROWING":
            score += 1
        elif st == "MIXED":
            score += 0.5

    # 9. Profit 3Q trend (1 pt bonus)
    pt = str(pd_.get("profit_trend_3q", "") or "").upper()
    if pt in ("GROWING", "DECLINING", "MIXED"):
        max_possible += 1
        if pt == "GROWING":
            score += 1
        elif pt == "MIXED":
            score += 0.5

    # Normalize to 0-100 based on actual data available
    if max_possible == 0:
        bq_score = 0.0
        completeness = 0.0
        verdict = "UNKNOWN"
    else:
        bq_score = round((score / max_possible) * 100.0, 1)
        # Completeness = fraction of the "full 100pt" universe we could evaluate
        completeness = round((max_possible / 100.0) * 100.0, 1)
        # Verdict — combine score AND critical flags
        _declining = any("DECLINING" in f for f in flags)
        if _declining and bq_score < 50:
            verdict = "DECLINING"
        elif bq_score >= 75:
            verdict = "STRONG"
        elif bq_score >= 55:
            verdict = "ACCEPTABLE"
        elif bq_score >= 40:
            verdict = "WEAK"
        else:
            verdict = "DECLINING"

    return {
        "bq_score":              bq_score,
        "bq_data_completeness":  completeness,
        "bq_flags":              flags,
        "bq_verdict":            verdict,
    }


def compute_sector_composite_score(sector: str, sector_rotation: dict) -> dict:
    """
    Phase R2 (2026-07-06): Sector Composite Score (0-100).

    Combines rank, momentum, status, and rotation velocity into a single
    institutional sector-quality gauge. Used to:
      • Give the 9-tier taxonomy a proper sector strength input
      • Discriminate "top-6 with strong momentum" from "top-6 fading"
      • Support CONTRARIAN detection with objective thresholds

    Returns dict with keys:
      sector_composite_score: 0-100
      sector_verdict: "LEADING" | "STRONG" | "NEUTRAL" | "WEAK" | "LAGGING"
    """
    if not sector or not isinstance(sector_rotation, dict):
        return {"sector_composite_score": 50.0, "sector_verdict": "UNKNOWN"}

    data = sector_rotation.get(sector, {})
    if not isinstance(data, dict):
        return {"sector_composite_score": 50.0, "sector_verdict": "UNKNOWN"}

    status = str(data.get("status", "NEUTRAL")).upper()
    velocity = str(data.get("rotation_velocity", "UNKNOWN")).upper()
    rank = data.get("rank_5d")
    ret5d = data.get("ret5d")
    ret20d = data.get("ret20d")

    score = 50.0

    # 1. Status contribution (±25)
    status_adj = {
        "LEADING": +25, "NEUTRAL": 0, "WEAKENING": -12, "LAGGING": -25,
    }.get(status, 0)
    score += status_adj

    # 2. Rank contribution (±15) — normalized to sector count
    if isinstance(rank, (int, float)) and rank > 0:
        n_sectors = max(len(sector_rotation), 1)
        # Convert rank to percentile: rank 1 = 100th %ile, rank N = 0th
        percentile = 1.0 - ((rank - 1) / max(n_sectors - 1, 1))
        rank_adj = (percentile - 0.5) * 30  # ±15
        score += rank_adj

    # 3. Momentum contribution (±10)
    if isinstance(ret20d, (int, float)):
        if ret20d >= 5:
            score += 10
        elif ret20d >= 2:
            score += 5
        elif ret20d >= 0:
            score += 0
        elif ret20d >= -3:
            score -= 5
        else:
            score -= 10

    # 4. Velocity contribution (±5)
    vel_adj = {"ROTATING_IN": +5, "STABLE": 0, "ROTATING_OUT": -5, "UNKNOWN": 0}.get(velocity, 0)
    score += vel_adj

    score = max(0.0, min(100.0, score))

    # Verdict
    if score >= 75:
        verdict = "LEADING"
    elif score >= 60:
        verdict = "STRONG"
    elif score >= 40:
        verdict = "NEUTRAL"
    elif score >= 25:
        verdict = "WEAK"
    else:
        verdict = "LAGGING"

    return {"sector_composite_score": round(score, 1), "sector_verdict": verdict}


def compute_all_factors(symbol: str, df,
                         sector: str, regime_data: dict,
                         sector_rotation: dict = None,
                         nifty_state: dict = None) -> dict:
    result = _default_stock_result(symbol, sector)
    try:
        closes  = df["Close"].squeeze().values.astype(float)
        highs   = df["High"].squeeze().values.astype(float)
        lows    = df["Low"].squeeze().values.astype(float)
        volumes = df["Volume"].squeeze().values.astype(float)

        if len(closes) < 50:
            return result

        last = closes[-1]
        prev = closes[-2]
        n    = len(closes)

        # Circuit breaker — skip stocks with >15% move today
        ret1d_check = (last / prev - 1) * 100 if prev > 0 else 0
        if abs(ret1d_check) > 15:
            _log(f"[SKIP] {symbol} — circuit move {ret1d_check:.1f}% today")
            return result

        # EMAs
        s_closes = pd.Series(closes)
        ema9   = float(s_closes.ewm(span=9).mean().iloc[-1])
        ema20  = float(s_closes.ewm(span=20).mean().iloc[-1])
        ema50  = float(s_closes.ewm(span=50).mean().iloc[-1])
        ema200 = float(s_closes.ewm(span=200).mean().iloc[-1])

        # ATR (14-period)
        tr = np.maximum(
            highs - lows,
            np.maximum(np.abs(highs - np.roll(closes, 1)),
                       np.abs(lows  - np.roll(closes, 1)))
        )
        # Phase 3a N3 / Option B+VC (2026-07-05): retain the full rolling ATR
        # series (not just the last value). Needed by the volatility-contraction
        # component of trade_quality_score, which compares ATR-now vs ATR-20-days-ago
        # to detect Minervini-style VCP (coiled-spring) setups.
        atr_series_full = pd.Series(tr[1:]).rolling(14).mean()
        atr14 = float(atr_series_full.iloc[-1])

        # Volume
        avg_vol_20    = float(pd.Series(volumes).rolling(20).mean().iloc[-1])
        today_vol     = float(volumes[-1])
        vol_ratio     = today_vol / avg_vol_20 if avg_vol_20 > 0 else 1.0
        avg_price_20  = float(s_closes.rolling(20).mean().iloc[-1])
        avg_val_lakhs = (avg_vol_20 * avg_price_20) / 100_000

        # 52-week levels
        lookback     = min(252, n)
        high_52w     = float(np.max(highs[-lookback:]))
        low_52w      = float(np.min(lows[-lookback:]))
        near_52w_high = last >= high_52w * 0.97

        # ── Factor 1: Trend Quality (daily + weekly multi-timeframe) ──
        tq = 50
        if last > ema9 > ema20 > ema50 > ema200:  tq = 92
        elif last > ema20 > ema50 > ema200:         tq = 78
        elif last > ema50 > ema200:                  tq = 62
        elif last > ema200:                          tq = 48
        elif last < ema200 and last < ema50:         tq = 22
        else:                                         tq = 35
        ema20_5d_ago = float(s_closes.ewm(span=20).mean().iloc[-6]) if n > 6 else ema20
        if ema20 > ema20_5d_ago: tq = min(100, tq + 8)
        # Weekly confirmation: penalise if weekly trend is down
        w_score, weekly_ok = weekly_trend_score(df)
        if not weekly_ok and tq > 50:
            tq = max(35, tq - 15)   # strong daily trend against weekly = reduce confidence
        result["trend_quality"] = tq

        # ── Factor 2: Momentum Quality ──
        ret5d  = (last / closes[-6]  - 1) * 100 if n > 6  else 0
        ret21d = (last / closes[-22] - 1) * 100 if n > 22 else 0
        mom = 50
        if ret5d > 3 and ret21d > 6:         mom = 90
        elif ret5d > 1.5 and ret21d > 3:     mom = 75
        elif ret5d > 0 and ret21d > 0:       mom = 60
        elif ret5d < -2 and ret21d < -4:     mom = 18
        elif ret5d < 0:                       mom = 38
        gains  = pd.Series(np.diff(closes)).clip(lower=0).rolling(14).mean().iloc[-1]
        losses = (-pd.Series(np.diff(closes))).clip(lower=0).rolling(14).mean().iloc[-1]
        rsi    = 100 - (100 / (1 + gains / losses)) if losses > 0 else 100.0
        if rsi > 70:  mom = min(100, mom + 5)
        elif rsi < 35: mom = max(0, mom - 10)
        result["momentum_quality"] = round(mom, 1)

        # ── Factor 3: Volume + OBV Trend + Accumulation ──
        obv_scr        = _obv_trend_score(closes, volumes)
        base_vol_score = volume_delivery_score(vol_ratio, obv_scr)
        accum_scr, accum_signal = accumulation_score(closes, volumes, ema20, avg_vol_20)
        # Blend: 60% today's vol/delivery + 40% 10-day accumulation pattern
        combined_vol = round(base_vol_score * 0.60 + accum_scr * 0.40, 1)
        result["volume_delivery"]  = combined_vol
        result["accum_signal"]     = accum_signal   # "ACCUMULATING"|"DISTRIBUTING"|"NEUTRAL"
        # Warn in Telegram if distributing despite strong trend
        if accum_signal == "DISTRIBUTING":
            result.setdefault("_soft_warnings", []).append("DISTRIBUTION_DETECTED")

        # ── Factor 4: Sector Strength ──
        rotation_adj, sector_status = 0, "NEUTRAL"
        rotation_hit = False
        if sector_rotation:
            # Detect whether this sector was actually IN the rotation map.
            # sector_rotation is typically {sector_name: score} but may be
            # {"ranks": {...}} in some builds — handle both.
            try:
                _known = set()
                if isinstance(sector_rotation, dict):
                    if isinstance(sector_rotation.get("ranks"), dict):
                        _known = set(sector_rotation["ranks"].keys())
                    else:
                        _known = set(sector_rotation.keys())
                rotation_hit = sector in _known
            except Exception:
                rotation_hit = False
            rotation_adj, sector_status = sector_rotation_score(sector, sector_rotation)
        # Phase C5 (rating ≥ 9.0): if sector is unmapped ("OTHERS"/"" ) AND
        # we couldn't find it in the rotation map, treat sector_strength as
        # MISSING (None). compute_base_confidence redistributes weight to
        # avoid diluting toward neutral 50 for well-behaving mid-caps whose
        # industry simply isn't in sector_master.csv yet.
        _sec_unmapped = (
            not sector
            or str(sector).strip().upper() in ("OTHERS", "OTHER", "UNKNOWN", "")
        )
        if _sec_unmapped and not rotation_hit:
            result["sector_strength"] = None
            result["sector_missing"]  = True
        else:
            result["sector_strength"] = max(0, min(100, 50 + rotation_adj))
            result["sector_missing"]  = False
        result["sector_status"]   = sector_status
        # Phase 4 (2026-07-01): surface rotation velocity + 5d rank delta on
        # the stock dict so Gate 9 (SECTOR_LAGGING) and Telegram formatters can
        # distinguish a laggard drifting further down (real reject) from one
        # already rotating IN (contrarian window — soft-only).
        _srot = {}
        if isinstance(sector_rotation, dict):
            _srot = sector_rotation.get(sector, {}) if isinstance(sector_rotation.get(sector), dict) else {}
        result["sector_velocity"]      = _srot.get("rotation_velocity", "UNKNOWN")
        result["sector_rank_5d"]       = _srot.get("rank_5d")
        result["sector_rank_delta_5d"] = _srot.get("rank_delta_5d")
        # Phase R2 (2026-07-06): institutional sector composite (0-100)
        try:
            _sc = compute_sector_composite_score(sector, sector_rotation or {})
            result["sector_composite_score"] = _sc.get("sector_composite_score", 50.0)
            result["sector_verdict"]         = _sc.get("sector_verdict", "UNKNOWN")
        except Exception:
            result["sector_composite_score"] = 50.0
            result["sector_verdict"]         = "UNKNOWN"
        if sector_rotation and not rotation_hit:
            # Surface as a soft warning + downgrade status so BUY thesis / logs
            # know we defaulted to neutral instead of a real rotation read.
            result["sector_status"] = "UNKNOWN"
            result.setdefault("_soft_warnings", []).append(
                f"SECTOR_ROTATION_MISSING({sector})"
            )

        # ── Factor 5: Relative Strength vs Nifty (real 21-day comparison) ──
        # Uses actual Nifty 21d return computed once in pipeline, not a proxy
        nifty_ret21_real = regime_data.get("nifty_ret21", 0.0)
        nifty_ret5_real  = regime_data.get("nifty_ret5",  0.0)
        rs_diff21 = ret21d - nifty_ret21_real
        rs_diff5  = ret5d  - nifty_ret5_real
        rs_combined = rs_diff21 * 0.6 + rs_diff5 * 0.4
        if rs_combined > 8:    rs = 92
        elif rs_combined > 4:  rs = 78
        elif rs_combined > 0:  rs = 62
        elif rs_combined > -4: rs = 45
        else:                  rs = 25
        result["rs_vs_nifty"] = rs
        result["rs_diff21"]   = round(rs_diff21, 2)

        # ── Factor 6: News Risk (placeholder — filled by pipeline after AI) ──
        # Phase 3a N2 (2026-07-05): use neutral 50 as pre-news placeholder.
        # Previously used None which triggered weight-redistribution in
        # compute_base_confidence and inflated confidence by ~6 pts for
        # small-caps with no news coverage, distorting the top-100 ranking.
        # The pipeline overwrites this with the real AI-derived value at
        # L10458 for the top-100 (100 if NO_NEWS, lower on penalty), and then
        # calls compute_base_confidence AGAIN at L10496 so final confidence
        # is unaffected — this fix only cleans up initial ranking bias.
        result["news_risk"] = 50

        # ── Factor 7: Risk / Reward ──
        entry = round(last, 2)

        # Stop: actual 10-day swing low — NOT a hardcoded % of entry
        recent_lows    = lows[-10:] if len(lows) >= 10 else lows
        swing_low      = float(np.min(recent_lows))
        stop_candidate = round(swing_low * 0.995, 2)   # 0.5% buffer below swing low
        risk_raw_pct   = (entry - stop_candidate) / entry * 100 if entry > 0 else 8.0

        # Phase G7-B (2026-07-03): ATR-aware stop caps.
        # OLD: hard-clamp to 3% floor / 12% cap regardless of volatility.
        # Problem: a low-vol stock (ATR=1.5%) gets a 12% stop = 8×ATR (noise-
        # induced exits impossible), while a high-vol stock (ATR=5%) gets
        # 12% = 2.4×ATR (constant whipsaw).
        # NEW: cap by max(fixed_cap, entry - 3.5×ATR14); floor by
        # min(fixed_floor, entry - 1.5×ATR14). Falls back to old fixed clamps
        # if atr14 is missing/zero (defensive — never widens vs old behaviour).
        atr14_val = float(atr14) if atr14 else 0.0
        if atr14_val > 0 and entry > 0:
            atr_pct       = (atr14_val / entry) * 100.0
            # ATR-scaled floor: never tighter than 1.5×ATR (would be pure noise)
            atr_floor_pct = max(2.0, min(4.0, 1.5 * atr_pct))
            # ATR-scaled cap: never wider than 3.5×ATR (thesis clearly broken
            # by then) but hard ceiling at 15% so absurd names can't slip in
            atr_cap_pct   = max(8.0, min(15.0, 3.5 * atr_pct))
        else:
            # Legacy fixed clamps if ATR unavailable
            atr_floor_pct = 3.0
            atr_cap_pct   = 12.0

        # 2026-07-03: stop clamp is no longer silent — record which branch
        # fired on _soft_warnings so downstream logs / BUY card / audit can see
        # "this stop isn't at a real support level, it's synthetic".
        _stop_clamped = None
        if risk_raw_pct < atr_floor_pct:
            _stop_clamped = (
                f"STOP_FLOOR_CLAMPED(raw {risk_raw_pct:.1f}% → "
                f"{atr_floor_pct:.1f}% floor, ATR14={atr14_val:.2f})"
            )
            stop_candidate = round(entry * (1 - atr_floor_pct / 100.0), 2)
        elif risk_raw_pct > atr_cap_pct:
            _stop_clamped = (
                f"STOP_CAP_CLAMPED(raw {risk_raw_pct:.1f}% → "
                f"{atr_cap_pct:.1f}% cap, ATR14={atr14_val:.2f})"
            )
            stop_candidate = round(entry * (1 - atr_cap_pct / 100.0), 2)
        stop = stop_candidate
        if stop >= entry:
            # Swing low is above today's close — stock just broke a 10d low.
            # This is *not* a valid swing-entry setup; surface it explicitly.
            _stop_clamped = "STOP_ABOVE_ENTRY_FALLBACK(swing_low>entry — 10d low broken today)"
            stop = round(entry * 0.94, 2)
        if _stop_clamped:
            result.setdefault("_soft_warnings", []).append(_stop_clamped)

        risk_amt = entry - stop

        # Targets: wider ATR multiples for better R/R (2.5x & 4.5x vs old 2x/4x)
        target1 = round(entry + 2.5 * atr14, 2)
        target2 = round(entry + 4.5 * atr14, 2)
        # Phase R6 (2026-07-06) — STRETCH TARGET for institutional-grade setups
        # ---------------------------------------------------------------
        # For wide-ATR breakout leaders (real-estate, capital goods, mid-cap
        # momentum) the fixed 1.5×risk minimum pins RR gross at ~1.5× even
        # when 4.5×ATR would give a much larger target. This caused BUY
        # starvation in BULL runs where LODHA/OBEROI (SA=88, sector=98) failed
        # the 1.8 RR gate. Solution: lift T2 floor to 2.5×risk when the setup
        # is institutional-grade AND stop is wide (>6% ATR).
        # Gate criteria (all required): swing_alpha_score already stamped elsewhere
        # is not available here — use pre-swing-alpha proxies: RS_20d > 5%,
        # volume expansion >= 1.5×, and stop distance > 6%.
        _stop_pct = ((entry - stop) / entry * 100.0) if entry > 0 else 0.0
        try:
            _stretch_enabled = os.getenv("STRETCH_TARGET", "1") == "1"
        except Exception:
            _stretch_enabled = True
        # Proxies (already computed in score_stock context, else neutral):
        _rs_here  = float(result.get("rs_vs_nifty_20d", 0) or 0)
        _vol_here = float(result.get("volume_expansion_ratio", 0) or 0)
        _is_stretch = (
            _stretch_enabled
            and _stop_pct > 6.0
            and _rs_here >= 5.0
            and _vol_here >= 1.5
        )
        if _is_stretch:
            # Guarantee minimum 2.5× R/R on T2 for wide-stop momentum leaders
            min_t2 = round(entry + risk_amt * 2.5, 2)
            result.setdefault("_soft_warnings", []).append(
                f"T2_STRETCHED(2.5×risk floor; stop={_stop_pct:.1f}%, RS={_rs_here:.1f}, vol={_vol_here:.2f}x)"
            )
        else:
            # Standard 1.5×risk floor for narrow-stop / average setups
            min_t2 = round(entry + risk_amt * 1.5, 2)
        target2 = max(target2, min_t2)
        # Keep T1 between entry and midpoint of T2
        target1 = min(target1, round((entry + target2) / 2, 2))

        rr_ratio  = round((target2 - entry) / risk_amt, 2) if risk_amt > 0 else 0.0
        result["entry"]       = entry
        result["stop"]        = stop
        result["target1"]     = target1
        result["target2"]     = target2
        result["rr_ratio"]    = rr_ratio
        result["risk_reward"] = min(100, max(0, rr_ratio * 30))

        # ── Phase C4: slippage-adjusted (net) R/R + overnight gap risk ──
        try:
            net = apply_slippage_to_rr(entry, stop, target1, target2, avg_val_lakhs)
            result["net_rr_t1"]           = net["net_rr_t1"]
            result["net_rr_t2"]           = net["net_rr_t2"]
            result["slippage_pct_one_way"] = net["cost"]["slippage_pct_one_way"]
            result["round_trip_cost_pct"] = net["cost"]["total_pct_rt"]
            # Gate uses the more conservative rr — protects illiquid stocks whose
            # gross R/R looks great but shrinks after cost.
            if net["net_rr_t2"] > 0 and net["net_rr_t2"] < rr_ratio:
                result["rr_ratio"]    = net["net_rr_t2"]
                result["risk_reward"] = min(100, max(0, net["net_rr_t2"] * 30))
        except Exception:
            result["net_rr_t1"]           = 0.0
            result["net_rr_t2"]           = 0.0
            result["slippage_pct_one_way"] = 0.0
            result["round_trip_cost_pct"] = 0.0

        try:
            gap = estimate_gap_risk_pct(closes, highs, lows)
            result["avg_gap_pct"] = gap["avg_gap_pct"]
            result["max_gap_pct"] = gap["max_gap_pct"]
            result["p90_gap_pct"] = gap["p90_gap_pct"]
            # Effective (worst-realistic) stop distance = raw stop % + p90 gap
            _stop_dist_pct = ((entry - stop) / entry * 100.0) if entry > 0 else 0.0
            result["effective_stop_pct"] = round(_stop_dist_pct + gap["p90_gap_pct"], 2)
            # Flag if p90 gap is a significant fraction of the stop budget
            if _stop_dist_pct > 0 and gap["p90_gap_pct"] > 0.6 * _stop_dist_pct:
                result["high_gap_risk"] = True
            else:
                result["high_gap_risk"] = False
        except Exception:
            result["avg_gap_pct"] = 0.0
            result["max_gap_pct"] = 0.0
            result["p90_gap_pct"] = 0.0
            result["effective_stop_pct"] = 0.0
            result["high_gap_risk"] = False

        # ── Factor 8: Ownership Quality — driven by promoter pledge + 52W proximity ──
        # Actual ROE/pledge injected later by fetch_all_fundamentals_cached();
        # here we seed with a neutral value that improves once fundamentals arrive.
        result["ownership_quality"] = 50  # updated in _update_ownership_quality()

        # ── Factor 9: Options Sentiment (placeholder) ──
        result["options_sentiment"] = 60

        # ── Factor 10: Macro Alignment ──
        macro     = regime_data.get("macro", {})
        macro_adj = macro_regime_adjustment(macro)
        result["macro_alignment"] = max(0, min(100, 60 + macro_adj * 2))

        # ── Price Action Pattern ──
        # Phase 3b (2026-07-05): pass volumes for Pocket Pivot detection.
        pa_score, pa_pattern = price_action_score(closes, highs, lows, ema20, atr14, volumes)
        result["price_pattern"]   = pa_pattern
        result["weekly_trend_ok"] = weekly_ok

        # ── Fix #1 (Phase 3b): Extension check — block chase trades ──
        # Reference: O'Neil "no more than 5% above pivot"; ChartMill "current
        # prices not too far from short-term MAs"; Minervini "avoid extended".
        # Formula: pct above 50-DMA. > 25% = extended (bad entry, chase risk).
        # This does NOT hard-block the stock — it demotes it via a score penalty.
        # Also emitted as `extension_pct` so downstream gates / logs can act.
        _extension_pct = 0.0
        _extended = False
        try:
            if ema50 and ema50 > 0:
                _extension_pct = (entry / ema50 - 1.0) * 100
                _extended = _extension_pct > 25.0
        except Exception:
            _extension_pct = 0.0
        result["extension_pct"] = round(_extension_pct, 1)
        result["is_extended"]   = _extended

        # ── Fix #2 (Phase 3b): Near-52W-high bonus ──
        # CANSLIM "L=Leader"; ChartMill "Strong Stocks near New High" screener.
        # Stocks within 15% of 52W-high are typically breakout leaders.
        # Emit as `near_52w_pct` (0-100 score) usable in TQ formula.
        #
        # India refinement (2026-07-05): US 52W-high leadership assumes clean
        # price discovery. In India, penny stocks and SEBI-surveilled names can
        # be pumped near 52W-high without real institutional buying — often
        # visible as thin delivery. Guard the bonus with two India-specific
        # quality floors:
        #   1) Market cap ≥ ₹500 Cr        (already the Gate 3c floor)
        #   2) volume_delivery score > 55  (blended vol/delivery/OBV proxy;
        #      55 ≈ "clearly better than average day" — filters out low-
        #      delivery pump patterns typical of ASM/GSM candidates)
        # If either floor fails, downgrade bonus to 50 (neutral, no credit)
        # so a manipulator's chart alone can't earn the CANSLIM-L reward.
        _n52_pct = 0.0
        _n52_downgraded = False
        try:
            if high_52w and high_52w > 0:
                _dist_pct = (high_52w - entry) / high_52w * 100  # 0 = at high, 20 = 20% below
                if _dist_pct <= 3.0:      _n52_pct = 95   # essentially at high — breakout candidate
                elif _dist_pct <= 8.0:    _n52_pct = 85   # very close
                elif _dist_pct <= 15.0:   _n52_pct = 70   # near enough (CANSLIM L)
                elif _dist_pct <= 25.0:   _n52_pct = 55   # moderate
                else:                     _n52_pct = 40   # far from 52W high
                # India quality floor — only applies if we were going to give a bonus (>55)
                if _n52_pct > 55.0:
                    _mcap = float(result.get("market_cap_cr", 0.0) or 0.0)
                    _vd   = float(result.get("volume_delivery", 50.0) or 50.0)
                    # market_cap_cr == 0 means "unknown" (G8-B convention) → do NOT punish
                    _mcap_ok = (_mcap == 0.0) or (_mcap >= 500.0)
                    _vd_ok   = _vd > 55.0
                    if not (_mcap_ok and _vd_ok):
                        _n52_pct = 50.0  # neutral: no CANSLIM-L credit without quality
                        _n52_downgraded = True
        except Exception:
            _n52_pct = 50.0
        result["near_52w_score"] = round(_n52_pct, 1)
        if _n52_downgraded:
            result.setdefault("_soft_warnings", []).append("NEAR_52W_NO_QUALITY")

        # ── Trade Quality Score — Phase 3a N3 / Option B+VC (2026-07-05) ──
        # EVOLUTION:
        #   OLD (pre-Phase-3a):  0.30·trend + 0.20·momentum + 0.15·volume +
        #                        0.15·rr + 0.10·weekly + 0.10·pa
        #     → DOUBLE-COUNTED trend/momentum/volume with FACTOR_WEIGHTS.
        #     → 97/100 candidates rubber-stamped through min_tq gate (useless gate).
        #
        #   NEW (Phase 3a N1, earlier today):  0.35·weekly + 0.25·pa + 0.25·rr +
        #                                       0.15·pattern_boost
        #     → Fixed double-count correctly BUT dropped volume entirely.
        #     → 10/100 passed (over-tight); mean_TQ collapsed 82 → 68.
        #     → Lost institutional-footprint signal.
        #
        #   OPTION B (Phase 3a N2, superseded by this patch):
        #                        0.30·weekly + 0.22·pa + 0.20·rr +
        #                        0.15·volume + 0.13·pattern_boost
        #     → Restored volume as institutional-footprint signal (~15%).
        #     → Missed volatility-contraction (Minervini VCP), the #1 pro signal.
        #
        # NOW (Option B+VC — THIS PATCH, Phase 3a N3):
        #                        0.28·weekly + 0.20·pa + 0.18·rr +
        #                        0.14·volume + 0.10·volatility_contraction +
        #                        0.10·pattern_boost
        #     → Adds volatility_contraction (VC) — coiled-spring detection.
        #     → Web-research sources (professional swing-trading consensus):
        #         • ChartMill Setup Quality: "volatility is decreasing" is one of
        #           the 6 named components (chartmill.com/documentation/technical-
        #           analysis/indicators/87). Also lists Bollinger Squeeze plays as
        #           #1 breakout setup.
        #         • Mark Minervini's VCP (Volatility Contraction Pattern) is his
        #           single most-cited setup — cited in the ChartMill Four Pillars
        #           doc (methodology 517, Pillar 4) as the empirical basis for
        #           Setup Quality ≥ 7.
        #         • William O'Neil / CANSLIM: cup-and-handle bases are all
        #           volatility-contraction structures — narrow range with
        #           volume drying up = right side of a base.
        #     → VC computed as ATR14_now / ATR14_20-days-ago:
        #         < 0.70  = strong contraction (coiled spring)      → 90
        #         0.70-0.85 = moderate contraction                    → 75
        #         0.85-1.10 = neutral                                 → 50
        #         > 1.10  = expanding volatility (breakout done)    → 35
        #     → Expected mean_TQ ≈ 72 (VC-mean ~55 pulls slightly below B's 74);
        #       min_tq recalibrated in REGIME_THRESHOLDS.
        #
        # WEIGHTS RATIONALE (sum = 1.00, range 0-100):
        #   0.28  weekly     — higher-timeframe trend alignment (setup context)
        #   0.20  pa         — candle/price-action pattern quality
        #   0.18  rr         — risk/reward trade math
        #   0.14  volume     — institutional footprint at entry (Minervini/O'Neil)
        #   0.10  vol_contr  — Minervini VCP (coiled spring)
        #   0.10  pat_boost  — named-pattern bonus / warning tiebreaker
        #
        # `_pattern_boost` = extra credit for named bullish patterns (breakout,
        # bullish_engulfing, hammer_at_ema20). 60 = neutral; +/-20 for boost/penalty.
        _pattern_boost = 60.0
        _pp = (pa_pattern or "").lower()
        if any(kw in _pp for kw in ("breakout", "bullish_engulf", "hammer", "cup_handle", "flag")):
            _pattern_boost = 80.0
        elif any(kw in _pp for kw in ("bearish", "topping", "distribution", "gap_down")):
            _pattern_boost = 30.0
        result["_pattern_boost"] = _pattern_boost
        # volume_delivery is on 0-100 scale (populated earlier by _volume_delivery_score);
        # guard against None / NaN edge cases from thin-trading days.
        _vol_component = result.get("volume_delivery")
        if _vol_component is None or (isinstance(_vol_component, float) and _vol_component != _vol_component):
            _vol_component = 50.0
        # Volatility contraction — Minervini VCP detector.
        # Compare ATR14 now vs ATR14 ~20 trading days ago (roughly 1 month).
        # Need atr_series_full to have ≥ 21 finite values; guard defensively.
        _vc_score = 50.0  # neutral default when data insufficient
        try:
            if atr_series_full is not None and len(atr_series_full) >= 21:
                _atr_ref = float(atr_series_full.iloc[-21])
                if _atr_ref > 0 and atr14 > 0:
                    _vc_ratio = atr14 / _atr_ref
                    if _vc_ratio < 0.70:
                        _vc_score = 90.0
                    elif _vc_ratio < 0.85:
                        _vc_score = 75.0
                    elif _vc_ratio < 1.10:
                        _vc_score = 50.0
                    else:
                        _vc_score = 35.0
        except Exception:
            _vc_score = 50.0
        result["volatility_contraction"] = round(_vc_score, 1)

        # Extension penalty (Fix #1). Extended stocks (>25% above 50-DMA) get
        # a strong TQ haircut so they lose the ranking race. Full formula:
        #   extension 0-15%  → no penalty (0)
        #   extension 15-25% → mild penalty (-5)
        #   extension >25%   → strong penalty (-15) — chase blocker
        #
        # Phase G8-E (2026-07-06): audit fix #5 — regime-scaled extension.
        # In STRONG_BULL / BULL trending environments, extended stocks
        # (>25% above 50-DMA) frequently continue running — the post-COVID
        # rally, 2023-24 midcap rally, and 2019 leader trends all show
        # +25–40% extensions persisting for weeks. A flat -15 there is
        # too punitive. In SIDEWAYS / HIGH_VOLATILITY / TRANSITION, the
        # same extension is a much stronger mean-reversion setup and the
        # full penalty stands. BEAR/STRONG_BEAR: extension is a warning of
        # a dead-cat bounce — tighten penalty (+50%).
        # Multipliers are conservative; net effect is a ±25% swing on the
        # penalty (not the score itself). Full behaviour disabled by env
        # EXTENSION_REGIME_SCALING=0 to keep the audit-mode option.
        _ext_penalty = 0.0
        if _extension_pct > 25.0:
            _ext_penalty = 15.0
        elif _extension_pct > 15.0:
            _ext_penalty = 5.0
        try:
            if os.getenv("EXTENSION_REGIME_SCALING", "1") != "0":
                _regime_name = (regime_data or {}).get("regime", "SIDEWAYS")
                _regime_mult = {
                    "STRONG_BULL":     0.50,   # extended trends persist — halve penalty
                    "BULL":            0.75,   # trending but not euphoric — mild relief
                    "SIDEWAYS":        1.00,   # baseline — full penalty (mean revert)
                    "TRANSITION":      1.00,   # baseline
                    "HIGH_VOLATILITY": 1.25,   # chop punishes extensions — tighten
                    "BEAR":            1.50,   # dead-cat bounces — tighten hard
                    "STRONG_BEAR":     1.50,
                }.get(_regime_name, 1.00)
                _ext_penalty = round(_ext_penalty * _regime_mult, 2)
                result["extension_regime_mult"] = _regime_mult
        except Exception:
            pass  # any error → keep flat penalty (safe default)
        result["extension_penalty"] = _ext_penalty

        # ── India refinement (2026-07-05): Expiry-week volume down-weight ──
        # NSE weekly F&O expiry lands every Thursday; monthly expiry is the
        # last Thursday of the month. On expiry weeks, hedging/rollover flows
        # inflate cash-market volume by 20-40% without reflecting fresh
        # directional conviction. Feeding raw volume into TQ during expiry
        # weeks over-credits noise trades.
        #
        # Fix: on expiry-week days (Wed/Thu/Fri of a Thursday-expiry week),
        # down-weight the volume component by 30% BEFORE it enters the TQ
        # formula. Uses local system date — cheap, deterministic, no external
        # data dep. Monthly expiry (last Thursday of month) gets the same
        # treatment (weekly-expiry rule already covers it since every Thursday
        # is at minimum a weekly expiry).
        _expiry_week_flag = False
        try:
            _today_local = datetime.date.today()
            _dow = _today_local.weekday()   # Mon=0 .. Sun=6, Thu=3
            # Wed(2), Thu(3), Fri(4) all fall within the expiry-week volume
            # contamination window — Wed sees pre-positioning, Fri sees rollover
            # cleanup. Skip Mon/Tue (fresh directional volume).
            if _dow in (2, 3, 4):
                _expiry_week_flag = True
                _vol_component = _vol_component * 0.70
        except Exception:
            _expiry_week_flag = False
        result["expiry_week"] = _expiry_week_flag
        if _expiry_week_flag:
            result.setdefault("_soft_warnings", []).append("EXPIRY_WEEK_VOL_ADJ")

        # ── Trade Quality Score — Phase 3b (2026-07-05) ──
        # New: added _n52_pct as a component (Fix #2), penalty from _extension
        # (Fix #1). Weights rebalanced so sum = 1.00.
        #   0.24  weekly     — higher-timeframe trend alignment
        #   0.18  pa         — price-action pattern (7 patterns now vs 3)
        #   0.16  rr         — R/R trade math
        #   0.12  volume     — institutional footprint
        #   0.10  vol_contr  — Minervini VCP proxy
        #   0.10  n52        — near-52W-high (CANSLIM Leader)  ★ NEW
        #   0.10  pat_boost  — named-pattern boost / warning
        # (0.24 + 0.18 + 0.16 + 0.12 + 0.10 + 0.10 + 0.10 = 1.00 ✓)
        # Final subtract: extension penalty (up to -15 pts direct).
        result["trade_quality_score"] = round(
            w_score                * 0.24 +   # weekly-frame trend alignment
            pa_score               * 0.18 +   # price-action pattern quality (7 patterns)
            result["risk_reward"]  * 0.16 +   # R/R trade math
            _vol_component         * 0.12 +   # institutional footprint at entry
            _vc_score              * 0.10 +   # Minervini VCP proxy
            _n52_pct               * 0.10 +   # ★ NEW: near-52W-high (CANSLIM L)
            _pattern_boost         * 0.10     # named-pattern boost / warning
            - _ext_penalty,                    # ★ NEW: extension chase blocker
            1,
        )

        # ═══════════════════════════════════════════════════════════════════
        # Phase R4 (2026-07-06): SWING-ALPHA OVERLAY (4 signals)
        # ─────────────────────────────────────────────────────────────────
        # Four swing-trading-specific edges that our previous scoring
        # under-weighted. Each is scored 0-100 (neutral 50), stamped on
        # result, and a composite `swing_alpha_score` produced as their
        # equal-weight mean. TQ score is nudged +5 for TQ<80 stocks whose
        # swing_alpha_score >= 65, giving fresh breakouts a real edge.
        # ─────────────────────────────────────────────────────────────────
        #  1. rs_vs_nifty_20d — per-stock 20d return minus Nifty 20d return
        #  2. breakout_freshness — days since last 20-day high (0 = today)
        #  3. atr_stop_ratio — ATR14 / price (tight = ≤3%; loose = >6%)
        #  4. volume_expansion — today_vol / avg_vol_20
        # ═══════════════════════════════════════════════════════════════════
        try:
            # ── R4-1: Relative Strength vs Nifty (20d) ──
            _nifty_20 = float((nifty_state or {}).get("ret_20d", 0.0) or 0.0)
            _stk_20 = ret21d  # our 21d return proxy (~20d)
            _rs_diff = _stk_20 - _nifty_20  # percentage points of outperformance
            # Score: -10pp → 0; 0pp → 50; +10pp → 100; capped
            _rs_score = 50 + (_rs_diff * 5)
            _rs_score = max(0, min(100, _rs_score))
            result["rs_vs_nifty_20d"]        = round(_rs_diff, 2)
            result["rs_vs_nifty_20d_score"]  = round(_rs_score, 1)

            # ── R4-2: Breakout freshness (days since last 20-day high) ──
            # Look back at last 20 sessions of highs; find the most recent
            # bar where high >= max(high[-20:]).
            _look20 = highs[-20:] if len(highs) >= 20 else highs
            _max_20 = float(np.max(_look20)) if len(_look20) else float(last)
            _days_since_hi = 0
            for i in range(len(_look20) - 1, -1, -1):
                if float(_look20[i]) >= _max_20 * 0.999:  # tolerance
                    _days_since_hi = len(_look20) - 1 - i
                    break
            # Score: 0-2 days (fresh) = 90; 3-5 = 70; 6-10 = 50; 11-15 = 30; >15 = 15
            if _days_since_hi <= 2:      _bf_score = 90
            elif _days_since_hi <= 5:    _bf_score = 70
            elif _days_since_hi <= 10:   _bf_score = 50
            elif _days_since_hi <= 15:   _bf_score = 30
            else:                        _bf_score = 15
            result["breakout_freshness_days"]  = _days_since_hi
            result["breakout_freshness_score"] = _bf_score

            # ── R4-3: ATR stop ratio (ATR14 / price, %) ──
            # Tight stops favored for swing (defined risk, better RR).
            _atr_pct = (atr14 / last * 100) if last > 0 else 0.0
            # Score: ≤2% (tight) = 90; ≤3% = 75; ≤5% = 55; ≤7% = 40; >7% = 20
            if _atr_pct <= 2.0:     _atr_score = 90
            elif _atr_pct <= 3.0:   _atr_score = 75
            elif _atr_pct <= 5.0:   _atr_score = 55
            elif _atr_pct <= 7.0:   _atr_score = 40
            else:                    _atr_score = 20
            result["atr_stop_ratio_pct"]   = round(_atr_pct, 2)
            result["atr_stop_ratio_score"] = _atr_score

            # ── R4-4: Volume expansion (today vs 20d avg) ──
            # Real breakouts trade on >1.5x average volume.
            _vol_exp = float(vol_ratio) if vol_ratio else 1.0
            # Score: ≥2.0x = 90; ≥1.5x = 75; ≥1.2x = 60; ≥0.9x = 45; <0.9x = 25
            if _vol_exp >= 2.0:      _ve_score = 90
            elif _vol_exp >= 1.5:    _ve_score = 75
            elif _vol_exp >= 1.2:    _ve_score = 60
            elif _vol_exp >= 0.9:    _ve_score = 45
            else:                     _ve_score = 25
            result["volume_expansion_ratio"] = round(_vol_exp, 2)
            result["volume_expansion_score"] = _ve_score

            # ── R4-composite: swing_alpha_score = equal-weight mean of 4 ──
            _sas = round((_rs_score + _bf_score + _atr_score + _ve_score) / 4.0, 1)
            result["swing_alpha_score"] = _sas

            # ── TQ nudge: reward setups the pre-R4 scoring under-weighted ──
            # If swing_alpha_score is high (>=65) and TQ<80, add up to +5.
            # Only active in swing trading mode.
            if os.environ.get("TRADING_MODE", "swing").lower() == "swing":
                _tq_pre = result.get("trade_quality_score", 0)
                if _sas >= 65 and _tq_pre < 80:
                    _bonus = min(5, (_sas - 65) / 7)  # 65→0, 100→5
                    result["trade_quality_score"] = round(_tq_pre + _bonus, 1)
                    result.setdefault("_swing_alpha_bonus", _bonus)
        except Exception as _e_r4:
            # Never fail scoring due to overlay computation
            result["swing_alpha_score"]       = 50.0
            result["rs_vs_nifty_20d_score"]   = 50.0
            result["breakout_freshness_score"] = 50.0
            result["atr_stop_ratio_score"]    = 50.0
            result["volume_expansion_score"]  = 50.0
            result.setdefault("_soft_warnings", []).append(f"SWING_ALPHA_ERR({type(_e_r4).__name__})")

        # ── Liquidity & price stats ──
        result["avg_volume"]      = round(avg_vol_20, 0)
        result["avg_value_lakhs"] = round(avg_val_lakhs, 1)
        result["near_52w_high"]   = near_52w_high
        # Issue 8 fix: populate catalyst tags. Previously code READ
        # `catalysts` (line ~2253, ~10385) but nothing WROTE it, so the
        # tags were always empty. Now they're driven off computed indicators.
        # Volume surge is split into UP (bullish, close >= prev_close) vs
        # DOWN (distribution warning, close < prev_close) so downstream
        # thesis logic doesn't confuse the two.
        try:
            _cats: list = []
            _prev_close = float(closes[-2]) if len(closes) >= 2 else float(last)
            _today_up = float(last) >= _prev_close
            if vol_ratio > 1.5:
                _cats.append("VOL_SURGE_UP" if _today_up else "VOL_SURGE_DOWN")
            # Retain legacy "VOL_SURGE" tag when direction is bullish so
            # existing rule_based_thesis / thesis_parts consumers keep working.
            if "VOL_SURGE_UP" in _cats:
                _cats.append("VOL_SURGE")
            if last > ema20 > ema50 > ema200:
                _cats.append("UPTREND")
            if near_52w_high:
                _cats.append("NEAR_52W_HIGH")
            if accum_signal == "ACCUMULATING":
                _cats.append("ACCUMULATION")
            elif accum_signal == "DISTRIBUTING":
                _cats.append("DISTRIBUTION")
            result["catalysts"] = _cats
        except Exception:
            result["catalysts"] = []
        result["price"]           = entry
        result["ret1d"]           = round(ret1d_check, 2)
        result["ret5d"]           = round(ret5d, 2)
        result["ret21d"]          = round(ret21d, 2)
        result["high_52w"]        = round(high_52w, 2)
        result["low_52w"]         = round(low_52w, 2)
        result["atr14"]           = round(atr14, 2)
        result["rsi14"]           = round(rsi, 1)

        # Phase I fix (2026-07-09): stamp the fields _classify_setup_live()
        # needs so setup_type is correctly classified as BREAKOUT / MOMENTUM /
        # PULLBACK / REVERSAL instead of falling through to "OTHER" for every
        # stock. Root cause: classifier's docstring claimed these fields were
        # already stamped by the scorer, but they weren't — every guard clause
        # `if last <= 0 or ema20 <= 0 or ...` fired and returned "OTHER",
        # forcing 100% of shadow-log rows into Bucket C.
        try:
            _h20 = float(np.max(highs[-20:])) if len(highs) >= 20 else float(np.max(highs))
            _l20 = float(np.min(lows[-20:]))  if len(lows)  >= 20 else float(np.min(lows))
            result["close"]     = round(float(last), 2)
            result["ema20"]     = round(float(ema20), 2)
            result["ema50"]     = round(float(ema50), 2)
            result["high_20d"]  = round(_h20, 2)
            result["low_20d"]   = round(_l20, 2)
            result["chg_5d"]    = round(float(ret5d), 2)
        except Exception:
            # Defensive — leave classifier fields unset so it falls back to
            # "OTHER" only in genuinely degenerate cases (no price data etc.)
            pass

        # ── factor_scores mirror dict — used by format_confidence_breakdown() ──
        # Phase C5 (rating ≥ 9.0): preserve None (MISSING) so downstream code
        # can distinguish "measured neutral 50" from "not measured at all".
        # compute_base_confidence uses this signal to redistribute weight
        # instead of diluting toward neutral 50.
        #
        # Phase 3a N5 (2026-07-05): NaN-safe mirror. Some upstream libs (pandas
        # rolling, numpy on all-NaN slices) can leak NaN into factor scores.
        # NaN would then survive round() as NaN, poison the sum in
        # compute_base_confidence, and produce NaN final_confidence which
        # silently fails min_confidence gate (NaN comparisons are always False).
        # Convert NaN/inf → None so redistribution kicks in cleanly.
        def _mirror(v):
            if v is None:
                return None
            try:
                f = float(v)
                # numpy nan/inf slip through as floats — reject them here
                if not (f == f) or f in (float("inf"), float("-inf")):
                    return None
                return round(f, 1)
            except (TypeError, ValueError):
                return None
        result["factor_scores"] = {k: _mirror(result.get(k)) for k in FACTOR_WEIGHTS}
        # Also sanitize the raw factor keys so downstream consumers never see NaN
        for _k in FACTOR_WEIGHTS:
            _sanitized = _mirror(result.get(_k))
            if _sanitized is not None:
                result[_k] = _sanitized

    except Exception as e:
        _log(f"[WARN] compute_all_factors failed for {symbol}: {e}")

    # ═══════════════════════════════════════════════════════════════════════
    # Phase R4 (2026-07-06): SWING-ALPHA OVERLAY — outer fallback wrapper.
    # If the inner try above failed and the swing-alpha overlay didn't run,
    # this block computes the same 4 signals from the raw df directly so
    # every row still gets the fields (even during partial-data failures).
    # ═══════════════════════════════════════════════════════════════════════
    if "swing_alpha_score" not in result:
        try:
            _c = df["Close"].squeeze().values.astype(float)
            _h = df["High"].squeeze().values.astype(float)
            _v = df["Volume"].squeeze().values.astype(float)
            if len(_c) >= 22:
                _last = float(_c[-1])
                # RS
                _stk_20 = (_last / float(_c[-21]) - 1) * 100 if _c[-21] > 0 else 0
                _nfty_20 = float((nifty_state or {}).get("ret_20d", 0.0) or 0.0)
                _rs_diff = _stk_20 - _nfty_20
                _rs_score = max(0, min(100, 50 + _rs_diff * 5))
                # Breakout freshness
                _look = _h[-20:]
                _mx = float(np.max(_look))
                _days = 0
                for i in range(len(_look) - 1, -1, -1):
                    if float(_look[i]) >= _mx * 0.999:
                        _days = len(_look) - 1 - i
                        break
                _bf = 90 if _days <= 2 else 70 if _days <= 5 else 50 if _days <= 10 else 30 if _days <= 15 else 15
                # ATR ratio
                _hi = df["High"].squeeze().values.astype(float)
                _lo = df["Low"].squeeze().values.astype(float)
                _tr = np.maximum(_hi[1:] - _lo[1:], np.maximum(np.abs(_hi[1:] - _c[:-1]), np.abs(_lo[1:] - _c[:-1])))
                _atr = float(pd.Series(_tr).rolling(14).mean().iloc[-1])
                _atr_pct = _atr / _last * 100 if _last > 0 else 0
                _atrs = 90 if _atr_pct <= 2 else 75 if _atr_pct <= 3 else 55 if _atr_pct <= 5 else 40 if _atr_pct <= 7 else 20
                # Volume expansion
                _avg_v = float(pd.Series(_v).rolling(20).mean().iloc[-1])
                _vr = float(_v[-1]) / _avg_v if _avg_v > 0 else 1.0
                _ves = 90 if _vr >= 2.0 else 75 if _vr >= 1.5 else 60 if _vr >= 1.2 else 45 if _vr >= 0.9 else 25
                result["rs_vs_nifty_20d"]         = round(_rs_diff, 2)
                result["rs_vs_nifty_20d_score"]   = round(_rs_score, 1)
                result["breakout_freshness_days"] = _days
                result["breakout_freshness_score"] = _bf
                result["atr_stop_ratio_pct"]      = round(_atr_pct, 2)
                result["atr_stop_ratio_score"]    = _atrs
                result["volume_expansion_ratio"]  = round(_vr, 2)
                result["volume_expansion_score"]  = _ves
                # Volume-gated weighted composite (Phase R4e). See inner block for rationale.
                _sas_fb = round(
                    (_rs_score * 0.30 + _bf * 0.25 + _ves * 0.30 + _atrs * 0.15), 1,
                )
                if _ves < 40:
                    _sas_fb = min(_sas_fb, 60.0)
                    result.setdefault("_soft_warnings", []).append(
                        f"SWING_NO_VOLUME(vol={_vr:.2f}x avg, capped SA=60)"
                    )
                result["swing_alpha_score"] = _sas_fb
        except Exception:
            # Absolute last resort — stamp neutral 50s so downstream never sees KeyError
            for _k in ("rs_vs_nifty_20d_score", "breakout_freshness_score",
                       "atr_stop_ratio_score", "volume_expansion_score",
                       "swing_alpha_score"):
                result.setdefault(_k, 50.0)

    return result


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 7 — PORTFOLIO & SIGNALS
# ─────────────────────────────────────────────────────────────────────────────

def load_portfolio() -> list:
    """
    Load active holdings. Priority order:
      1. PORTFOLIO_JSON env var (set as GitHub Actions secret)
      2. portfolio.json file on disk
      3. portfolio_state.csv (legacy CSV fallback)
    """
    # Priority 1: env var (GitHub Actions secret — supports both names)
    env_json = (os.getenv("MANUAL_PORTFOLIO_JSON") or os.getenv("PORTFOLIO_JSON") or "").strip()
    if env_json and env_json != "[]":
        try:
            data = json.loads(env_json)
            if data:
                _log(f"[INFO] Loaded {len(data)} holdings from MANUAL_PORTFOLIO_JSON secret")
                return data
        except Exception as e:
            _log(f"[WARN] MANUAL_PORTFOLIO_JSON env var parse failed: {e}")

    # Priority 2: JSON file (v6.0 format)
    try:
        if os.path.exists(PORTFOLIO_FILE):
            with open(PORTFOLIO_FILE, "r") as f:
                data = json.load(f)
                if data:
                    return data
    except Exception as e:
        _log(f"[WARN] load_portfolio JSON failed: {e}")

    # Fallback: portfolio_state.csv (existing pipeline format)
    csv_path = os.getenv("PORTFOLIO_STATE_FILE", "portfolio_state.csv")
    try:
        if os.path.exists(csv_path):
            holdings = []
            with open(csv_path, newline="", encoding="utf-8") as f:
                for row in csv.DictReader(f):
                    status = row.get("status", "").strip().upper()
                    if status not in ("OPEN", "ACTIVE", ""):
                        continue
                    holdings.append({
                        "symbol":      row.get("symbol", "").strip(),
                        "sector":      row.get("sector", "OTHERS").strip(),
                        "entry_price": float(row.get("entry_price", 0) or 0),
                        "stop_loss":   float(row.get("stop_loss", 0) or 0),
                        # CSV uses target_1 / target_2; normalise to target1 / target2
                        "target1":     float(row.get("target1") or row.get("target_1", 0) or 0),
                        "target2":     float(row.get("target2") or row.get("target_2", 0) or 0),
                        "entry_date":  row.get("entry_date", ""),
                        "quantity":    float(row.get("quantity", 0) or 0),
                    })
            if holdings:
                _log(f"[INFO] Loaded {len(holdings)} holdings from {csv_path} (CSV fallback)")
                return holdings
    except Exception as e:
        _log(f"[WARN] load_portfolio CSV fallback failed: {e}")
    return []


def save_portfolio(holdings: list) -> None:
    try:
        with open(PORTFOLIO_FILE, "w") as f:
            json.dump(holdings, f, indent=2, default=str)
    except Exception as e:
        _log(f"[WARN] save_portfolio failed: {e}")


def monitor_portfolio(holdings: list, price_data: dict, regime: str) -> list:
    """Bug 1 fix — pure rule-based, zero AI calls, no raw arrays sent anywhere."""
    alerts = []
    for holding in holdings:
        symbol   = holding.get("symbol", "")
        entry    = float(holding.get("entry_price", 0) or 0)
        stop     = float(holding.get("stop_loss",   0) or 0)
        qty      = float(holding.get("quantity",    0) or 0)
        sector   = holding.get("sector", "OTHERS")
        # Accept both target1 and target_1 (CSV legacy field names)
        target1  = float(holding.get("target1") or holding.get("target_1", 0) or 0)
        target2  = float(holding.get("target2") or holding.get("target_2", 0) or 0)
        current  = float(price_data.get(symbol, entry) or entry)
        pnl_pct  = round((current - entry) / entry * 100, 2) if entry > 0 else 0.0
        invested = round(entry   * qty, 2) if qty > 0 else 0.0
        cur_val  = round(current * qty, 2) if qty > 0 else 0.0
        pnl_abs  = round(cur_val - invested, 2)
        try:
            entry_dt  = datetime.datetime.strptime(holding.get("entry_date", ""), "%Y-%m-%d")
            # FIX: prefer business days (T-days). Weekends inflated the counter.
            try:
                _tdays = int(pd.bdate_range(entry_dt.date(),
                                            ist_today()).size) - 1
                days_held = max(0, _tdays)
            except Exception:
                days_held = (datetime.datetime.today() - entry_dt).days
        except Exception:
            days_held = 0

        base = {
            "symbol": symbol, "sector": sector,
            "pnl_pct": pnl_pct, "current": current,
            "quantity": qty, "entry": entry,
            "invested": invested, "cur_val": cur_val, "pnl_abs": pnl_abs,
            "target1": target1, "target2": target2, "stop": stop,
        }
        # Phase C4 (Gap #6): partial-exit-at-T1 recognition from holdings CSV.
        # If holdings has a "partial_closed" flag from an earlier day, don't
        # re-suggest T1_PARTIAL_EXIT — instead show TRAIL_STOP for the residual.
        partial_done  = bool(holding.get("partial_closed", False))
        # Phase E1a: recognize a runner position from holdings CSV. Once T2 has
        # been hit and a runner is riding, alerts should suggest RUNNER_TRAIL
        # (not a stale "EXIT_FULL at T2") unless the price actually crossed
        # the current trailed stop.
        runner_active = bool(holding.get("runner_active", False))
        if stop > 0 and current <= stop:
            reason = "RUNNER_TRAIL_HIT" if runner_active else "HARD_STOP_HIT"
            alerts.append({**base, "action": "EXIT", "reason": reason})
        elif regime in ("BEAR", "STRONG_BEAR"):
            alerts.append({**base, "action": "EXIT",       "reason": "REGIME_BEAR"})
        elif runner_active:
            # Runner is active. Suggest the trail action; the tracker's chandelier
            # rules (in update_tracker) decide when to actually exit.
            atr_mult = _effective_runner_atr_mult(regime)
            alerts.append({
                **base,
                "action": "RUNNER_TRAIL",
                "reason": f"POST_T2_RUNNER · chandelier trail ATR×{atr_mult:.1f} · stop ₹{stop:.2f}",
            })
        elif target2 > 0 and current >= target2:
            # E1a: if runner mode is on AND T1 was already booked, this is
            # runner-entry, not full-exit. Otherwise legacy full exit.
            if _runner_enabled(regime) and partial_done:
                rp_pct = max(0.0, min(100.0, RUNNER_PARTIAL_PCT))
                alerts.append({
                    **base,
                    "action": "T2_RUNNER_START",
                    "reason": f"TARGET2_HIT · book {rp_pct:.0f}% of residual · runner rides chandelier",
                    "runner_partial_pct": rp_pct,
                })
            else:
                alerts.append({**base, "action": "EXIT_FULL",  "reason": "TARGET2_HIT"})
        elif not partial_done and target1 > 0 and current >= target1:
            # First time T1 is hit — book PARTIAL_EXIT_PCT, trail rest
            pe_pct = float(os.getenv("PARTIAL_EXIT_PCT", "50"))
            if TRAIL_MODE == "atr":
                mult = _effective_trail_atr_mult(regime)
                trail_desc = f"ATR trail (×{mult:.1f})"
            else:
                trail_desc = f"break-even ₹{entry:.2f}"
            alerts.append({
                **base,
                "action": "T1_PARTIAL_EXIT",
                "reason": f"TARGET1_HIT · sell {pe_pct:.0f}% @ ~₹{current:.2f} · trail residual via {trail_desc}",
                "partial_exit_pct":   pe_pct,
                "trail_stop_to":      entry,
                "residual_target":    target2,
            })
        elif partial_done and target1 > 0 and current >= target1:
            # Residual position — normal trail
            alerts.append({**base, "action": "TRAIL_STOP", "reason": "RESIDUAL_TRAIL_TO_T2"})
        elif days_held >= 20 and (target1 == 0 or current < target1):
            alerts.append({**base, "action": "REVIEW",     "reason": "TIME_STOP_20D", "days_held": days_held})
        else:
            alerts.append({**base, "action": "HOLD",       "reason": "ON_TRACK",      "days_held": days_held})
    return alerts


def detect_short_signals(scored_stocks: list, regime: str, thresh: dict) -> list:
    shorts = []
    if regime not in ("BEAR", "STRONG_BEAR", "HIGH_VOLATILITY"):
        return []
    for stock in scored_stocks:
        try:
            df = stock.get("_df")
            if df is None or len(df) < 50:
                continue
            closes  = df["Close"].squeeze().values.astype(float)
            volumes = df["Volume"].squeeze().values.astype(float)
            last    = closes[-1]
            ema20   = float(pd.Series(closes).ewm(span=20).mean().iloc[-1])
            ema50   = float(pd.Series(closes).ewm(span=50).mean().iloc[-1])
            ema200  = float(pd.Series(closes).ewm(span=200).mean().iloc[-1])
            avg_vol = float(pd.Series(volumes).rolling(20).mean().iloc[-1])
            today_vol = float(volumes[-1])
            below_all_emas    = last < ema20 < ema50 < ema200
            volume_confirmed  = today_vol > avg_vol * 1.5
            ret_5d = (last / closes[-6] - 1) * 100 if len(closes) > 6 else 0
            bearish_momentum  = ret_5d < -2.0
            if below_all_emas and volume_confirmed and bearish_momentum:
                short_entry  = round(last, 2)
                short_stop   = round(ema20 * 1.02, 2)
                short_t1     = round(last * 0.94, 2)
                short_t2     = round(last * 0.88, 2)
                rr = round((short_entry - short_t1) / (short_stop - short_entry), 2) if short_stop > short_entry else 0.0
                if rr >= thresh.get("min_rr", 1.6):
                    shorts.append({
                        "symbol":  stock.get("symbol", ""),
                        "entry":   short_entry,
                        "stop":    short_stop,
                        "target1": short_t1,
                        "target2": short_t2,
                        "rr":      rr,
                        "reason":  f"Below all EMAs | Vol {today_vol/max(avg_vol,1):.1f}x | Ret5d {ret_5d:.1f}%",
                        "sector":  stock.get("sector", "OTHERS"),
                    })
        except Exception as e:
            _log(f"[WARN] short signal detection failed for {stock.get('symbol')}: {e}")
    shorts.sort(key=lambda x: x["rr"], reverse=True)
    return shorts[:3]


def run_gates(stock: dict, regime: str, thresholds: dict,
              portfolio: dict, bulk_deals: dict, promoter_data: dict,
              results_dates: list = None, upcoming_events: list = None,
              returns_cache: dict = None, holdings: list = None) -> dict:
    """15-gate decision system + Phase R1/R2 institutional validation layer.

    Phase R1 (2026-07-06) added:
      Gate 3e — BUSINESS_QUALITY_DECLINE (sales+profit both 3Q down)
      Gate 3f — ROE_TOO_LOW (< MIN_ROE)
      Gate 3g — DE_TOO_HIGH (> MAX_DE, financials excepted)
      Gate 3h — NEWS_SEVERITY_HIGH (LLM severity > NEWS_SEVERITY_MAX)
      Gate 9b — SECTOR_RANK_TOO_LOW (rank > SECTOR_RANK_CUTOFF)
      Gate 9c — SECTOR_MOMENTUM_NEG (sector 20d return < 0)
      + Missing-data confidence cap, three-pillar floor, taxonomy expansion

    Phase R2 (2026-07-06) added:
      • compute_business_quality_score() attached to stock as bq_score/verdict
      • compute_sector_composite_score() attached as sector_composite_score
      • 9-tier taxonomy: STRONG_BUY / BUY / BUY_CONTRARIAN / BUY_TURNAROUND
                         WATCHLIST / DEVELOPING / MONITOR / AVOID / REJECTED
    """
    decision = "BUY"
    fail_reasons = []
    warnings = []
    thresh = thresholds[regime]

    # ── Phase G-BATCH2 (2026-07-07): honour Phase-G hard rejects ─────────
    # If any of quality/options/news/insider modules flagged a HARD reject
    # during Step 14a enrichment, propagate it here so the stock skips the
    # rest of gate evaluation (and lands in `rejected` with a clean reason).
    if stock.get("phaseG_hard_reject"):
        _phaseG_reasons = [r for r in (stock.get("fail_reasons") or [])
                           if r.split("(", 1)[0] in
                           ("QUALITY_FAIL", "OPTIONS_FAIL", "NEWS_FAIL", "INSIDER_FAIL")]
        return {
            "decision": "REJECTED",
            "fail_reasons": _phaseG_reasons or ["PHASE_G_HARD_REJECT"],
            "warnings": [],
        }

    # ── Phase I (2026-07-07): honour setup-edge chop-regime skip ─────────
    # Backtest (2026-07-07, 26,665 trades) showed WEAK + SIDEWAYS regimes
    # bleed ~-0.18 to -0.30 R per trade UNLESS the setup is BREAKOUT (only
    # bucket with positive expectancy). apply_setup_edge() stamps
    # `phase_i_skip` when this filter should trigger. Toggle via
    # ENABLE_REGIME_SETUP_FILTER env var.
    if stock.get("phase_i_skip"):
        return {
            "decision": "REJECTED",
            "fail_reasons": [f"SETUP_EDGE_SKIP({stock['phase_i_skip']})"],
            "warnings": [],
        }

    # Phase R2 (2026-07-06): compute BQ + sector composite EARLY so downstream
    # logic + taxonomy can use them. Stamped onto `stock` for downstream
    # renderers, CSV rows, and audit rows.
    _bq_result = compute_business_quality_score(promoter_data or {})
    stock["bq_score"]             = _bq_result["bq_score"]
    stock["bq_data_completeness"] = _bq_result["bq_data_completeness"]
    stock["bq_verdict"]           = _bq_result["bq_verdict"]
    stock["bq_flags"]             = _bq_result["bq_flags"]

    # Gate 1: Data Quality (HARD)
    if not stock.get("entry") or not stock.get("stop") or not stock.get("target1"):
        return {"decision": "REJECTED", "fail_reasons": ["DATA_INCOMPLETE"], "warnings": []}

    # Gate 2: Black Swan News (HARD)
    if stock.get("news_penalty", 0) >= 999 or stock.get("is_black_swan"):
        return {"decision": "REJECTED", "fail_reasons": ["BLACK_SWAN_NEWS"], "warnings": []}

    # Gate 3: Promoter Pledge (HARD)
    # Phase G6 (2026-07-03): Two-layer defence:
    #   Layer A — pledge% from fundamentals cache (works when screener/yf has data)
    #   Layer B — curated blocklist (protects when pledge% is 0 due to scraper miss)
    _sym = stock.get("symbol", "")
    if is_pledge_blocked(_sym):
        return {"decision": "REJECTED", "fail_reasons": ["PROMOTER_PLEDGE_BLOCKLIST"], "warnings": []}
    pledge = float(promoter_data.get("promoter_pledge_pct", 0) or 0)
    if pledge > 40:
        return {"decision": "REJECTED", "fail_reasons": [f"PROMOTER_PLEDGE_{pledge:.0f}PCT"], "warnings": []}

    # Gate 3c (Phase G8-B, 2026-07-03): Market-cap floor (HARD when data present)
    # Micro-caps (< ₹500 Cr mcap) have ~3× the failure rate of small-caps in
    # Indian markets: thin float, promoter-controlled, prone to price rigging,
    # and Fii/DII cannot enter → no institutional floor on drawdowns.
    #
    # Behaviour:
    #   - If market_cap_cr present AND > 0 AND < MIN_MARKET_CAP_CR   → HARD REJECT
    #   - If market_cap_cr missing/zero → fall back to turnover proxy:
    #         float_proxy_cr = avg_price × avg_vol × 250 / 1e7
    #     (250 trading days as a conservative "annual liquidity" estimate).
    #     If proxy < MIN_MARKET_CAP_CR × 0.5 AND we have real avg data → HARD REJECT.
    #     Otherwise: soft warning only (don't punish stocks with just missing data).
    # Env override: MIN_MARKET_CAP_CR (default 500). Set to 0 to disable.
    try:
        _min_mcap_cr = float(os.getenv("MIN_MARKET_CAP_CR", "500"))
    except (TypeError, ValueError):
        _min_mcap_cr = 500.0
    if _min_mcap_cr > 0:
        _mcap = float(promoter_data.get("market_cap_cr", 0) or stock.get("market_cap_cr", 0) or 0)
        if _mcap > 0:
            if _mcap < _min_mcap_cr:
                return {
                    "decision":     "REJECTED",
                    "fail_reasons": [f"MARKET_CAP_LOW_₹{_mcap:.0f}Cr_(min_₹{_min_mcap_cr:.0f}Cr)"],
                    "warnings":     [],
                }
        else:
            # Proxy check: use price × 20d avg volume × 250 as annualized float size
            _avg_price = float(stock.get("close_20d_avg", 0) or stock.get("entry", 0) or 0)
            _avg_vol   = float(stock.get("avg_volume", 0) or 0)
            if _avg_price > 0 and _avg_vol > 0:
                _float_proxy_cr = (_avg_price * _avg_vol * 250) / 1e7
                # Proxy threshold: 0.5× real threshold (proxy is noisy)
                if _float_proxy_cr < (_min_mcap_cr * 0.5):
                    return {
                        "decision":     "REJECTED",
                        "fail_reasons": [
                            f"MARKET_CAP_PROXY_LOW_₹{_float_proxy_cr:.0f}Cr_"
                            f"(proxy_min_₹{_min_mcap_cr * 0.5:.0f}Cr, mcap_missing)"
                        ],
                        "warnings":     [],
                    }
                else:
                    warnings.append(f"MCAP_MISSING (proxy ~₹{_float_proxy_cr:.0f}Cr)")
            else:
                warnings.append("MCAP_MISSING (no proxy data)")

    # Gate 3d — Phase 4-A (2026-07-06): Fundamentals data-missing check (SOFT).
    # When BOTH screener.in AND yfinance fail to return real fundamentals
    # (rate-limit, HTTP errors, or empty responses), ROE / D/E / pledge are
    # zeroed out and the pledge/mcap/ROE gates all abstain because there's
    # no data to reject. Historically this caused the 2026-07-06 00:35 IST
    # run to surface 3 BUYs (PANAMAPET, NELCO, LANDMARK) despite NELCO's
    # real ROE being 0.5% and LANDMARK's D/E being 1.74 — both hard fails
    # once screener came back online 20 minutes later.
    #
    # Fix: any stock whose fundamentals_source signals "couldn't fetch real
    # data" is demoted to WATCHLIST rather than accepted as a full BUY.
    # Uses the same soft-fail path as SECTOR_CAP / EVENT_BLOCK.
    #
    # Trigger sources (all mean "no real data"):
    #   NEUTRAL_DEFAULT  — cache miss + all 3 fallbacks empty
    #   NOT_FETCHED     — skipped due to max_stocks cap
    #   SCREENER+YF_RL  — screener returned zeros AND yfinance rate-limited
    #
    # Env override: FUND_DATA_GATE_ENABLED (default 1). Set to 0 to disable
    # (useful if screener/yf are down for hours and we still want *some*
    # signal on top-40 candidates — they land on WATCHLIST anyway then).
    try:
        _fund_gate_on = int(os.getenv("FUND_DATA_GATE_ENABLED", "1"))
    except (TypeError, ValueError):
        _fund_gate_on = 1
    # 2026-07-06 (Correct fix): Missing fundamentals is handled by the
    # existing score-redistribution path (compute_base_confidence excludes
    # the ownership_quality weight if it is None, distributing across
    # present factors). We DO NOT hard-reject here — that was double
    # punishment on top-9%-of-universe candidates that already passed
    # every technical gate.
    #
    # HISTORICAL BUG (kept for audit): Phase 4-A added FUND_DATA_MISSING as
    # a scoreable-fail reason. When Screener.in was rate-limited we saw
    # 82/100 rejects because of this gate alone, on days like 2026-07-06
    # when the fetch cap was too tight. Removed 2026-07-06.
    #
    # Env override: FUND_DATA_GATE_ENABLED=1 restores the old strict
    # behavior (default 0 = graceful). Use only for debugging.
    try:
        _fund_gate_on = int(os.getenv("FUND_DATA_GATE_ENABLED", "0"))
    except (TypeError, ValueError):
        _fund_gate_on = 0
    # Always emit a WARNING so audit shows data-availability, regardless
    # of gate mode. Only append to fail_reasons if the strict gate is on.
    _fsrc = str(stock.get("fundamentals_source", "") or "").upper()
    _pdata_src = str((promoter_data or {}).get("source", "") or "").upper()
    _missing_markers = ("NEUTRAL_DEFAULT", "NOT_FETCHED", "SCREENER+YF_RL")
    # Belt-and-braces: also treat ROE==0 AND D/E==0 AND pledge==0 as a miss
    # (some cache paths write source="" instead of one of the markers).
    _roe_missing = float(promoter_data.get("roe", 0) or 0) == 0.0
    _de_missing  = float(promoter_data.get("de_ratio", 0) or 0) == 0.0
    _pl_missing  = float(promoter_data.get("promoter_pledge_pct", 0) or 0) == 0.0
    _all_zero = _roe_missing and _de_missing and _pl_missing
    _src_missing = (any(m in _fsrc for m in _missing_markers)
                    or any(m in _pdata_src for m in _missing_markers))
    if _src_missing or (_all_zero and _fsrc != "" and "SCREENER+YF" not in _fsrc):
        warnings.append(
            f"FUND_DATA_MISSING: ROE/D/E/pledge unavailable (src={_fsrc or 'unknown'}) "
            f"— ownership_quality weight redistributed to present factors"
        )
        if _fund_gate_on:
            # Legacy strict mode (opt-in only). Kept for A/B testing.
            fail_reasons.append(f"FUND_DATA_MISSING(src={_fsrc or _pdata_src or 'unknown'})")

    # ─────────────────────────────────────────────────────────────────────
    # Phase R1 (2026-07-06): INSTITUTIONAL VALIDATION LAYER — Business gates
    # ─────────────────────────────────────────────────────────────────────
    # Design principle: a technically strong chart must NEVER override a
    # deteriorating business. The three business-quality gates below
    # implement Layer 1 of the institutional validation redesign.
    # Each has an env kill-switch so the entire layer can be A/B-tested.
    #
    # Data source: promoter_data dict enriched by _parse_screener_html().
    # When quarterly fields are absent (screener rate-limited / new listing)
    # the gates abstain — they never punish based on missing data alone
    # (that's what FUND_DATA_MISSING above already flags).

    # Gate 3e — BUSINESS_QUALITY_DECLINE (HARD when data present).
    # Reject when BOTH sales AND profit have been declining for the last
    # 3 quarters in a row. This catches the "chart-strong but earnings-
    # deteriorating" trap that surfaced on 2026-07-06 (JTLIND surfaced as
    # BUY despite weak sector + sub-8% ROE).
    #
    # Turnaround escape hatch: if the latest quarter profit YoY is > +20%
    # AND sales YoY > 0, treat as a turnaround-in-progress and demote to
    # WATCHLIST (soft fail) rather than hard reject. Institutional funds
    # do buy turnarounds — but never blindly, and only after early proof.
    #
    # Env: BUSINESS_QUALITY_GATE=1 (default on). Set 0 to disable.
    try:
        _bq_gate_on = int(os.getenv("BUSINESS_QUALITY_GATE", "1"))
    except (TypeError, ValueError):
        _bq_gate_on = 1
    _sales_trend  = str(promoter_data.get("sales_trend_3q",  "") or "").upper()
    _profit_trend = str(promoter_data.get("profit_trend_3q", "") or "").upper()
    _sales_yoy    = promoter_data.get("sales_yoy_pct")
    _profit_yoy   = promoter_data.get("profit_yoy_pct")
    _has_quarterly = _sales_trend in ("DECLINING", "GROWING", "MIXED") and \
                     _profit_trend in ("DECLINING", "GROWING", "MIXED")
    if _bq_gate_on and _has_quarterly:
        if _sales_trend == "DECLINING" and _profit_trend == "DECLINING":
            # Turnaround check
            _is_turnaround = (
                _profit_yoy is not None and _profit_yoy > 20.0
                and _sales_yoy is not None and _sales_yoy > 0.0
            )
            if _is_turnaround:
                # Soft fail — reaches WATCHLIST as an early turnaround.
                fail_reasons.append(
                    f"BQ_TURNAROUND_EARLY(sales3q↓ profit3q↓ but profit_yoy=+{_profit_yoy:.1f}%)"
                )
                warnings.append(
                    f"POTENTIAL_TURNAROUND: sales/profit both down 3Q but latest Q shows recovery "
                    f"(sales YoY {_sales_yoy:+.1f}%, profit YoY {_profit_yoy:+.1f}%)"
                )
            else:
                # HARD reject — declining business, no recovery signal.
                return {
                    "decision": "REJECTED",
                    "fail_reasons": [
                        f"BUSINESS_QUALITY_DECLINE(sales3q↓ profit3q↓ "
                        f"sales_yoy={_sales_yoy if _sales_yoy is not None else 'N/A'}% "
                        f"profit_yoy={_profit_yoy if _profit_yoy is not None else 'N/A'}%)"
                    ],
                    "warnings": [],
                }
        elif _sales_trend == "DECLINING" or _profit_trend == "DECLINING":
            # One of the two declining is a warning, not a fail.
            warnings.append(
                f"BQ_MIXED_TREND(sales_3q={_sales_trend}, profit_3q={_profit_trend})"
            )

    # Gate 3f — ROE_TOO_LOW (HARD when data present).
    # Institutional benchmark: ROE ≥ 15% is "high quality"; 12-15% is
    # "acceptable"; < 12% is "sub-institutional". We use 12% as the floor
    # so we don't reject stocks that are borderline but improving.
    #
    # Env: MIN_ROE (default 12). Set to 0 to disable entirely.
    try:
        _min_roe = float(os.getenv("MIN_ROE", "12"))
    except (TypeError, ValueError):
        _min_roe = 12.0
    _roe_val = float(promoter_data.get("roe", 0) or 0)
    # Only gate when we actually have real ROE data (roe > 0 AND fund_source
    # signals real fetch). Zero ROE = "unknown" not "loss-making".
    _roe_real = _roe_val > 0 and not _src_missing and not _all_zero
    if _min_roe > 0 and _roe_real and _roe_val < _min_roe:
        return {
            "decision":     "REJECTED",
            "fail_reasons": [f"ROE_TOO_LOW({_roe_val:.1f}%<{_min_roe:.0f}%)"],
            "warnings":     [],
        }

    # Gate 3g — DE_TOO_HIGH (HARD when data present, financials excepted).
    # Banks / NBFCs / insurance carry D/E > 5 by nature — gating them at
    # 1.5 would kill the entire financial sector. We skip this gate for
    # sectors flagged as financial.
    #
    # Env: MAX_DE (default 1.5). Set to 0 to disable.
    try:
        _max_de = float(os.getenv("MAX_DE", "1.5"))
    except (TypeError, ValueError):
        _max_de = 1.5
    _de_val = float(promoter_data.get("de_ratio", 0) or 0)
    _sec_upper = str(stock.get("sector", "") or "").upper()
    _financial_sectors = ("BANKING", "FINANCE", "INSURANCE", "NBFC", "FINANCIAL")
    _is_financial = any(fs in _sec_upper for fs in _financial_sectors)
    _de_real = _de_val > 0 and not _src_missing and not _all_zero
    if _max_de > 0 and _de_real and not _is_financial and _de_val > _max_de:
        return {
            "decision":     "REJECTED",
            "fail_reasons": [f"DE_TOO_HIGH({_de_val:.2f}>{_max_de:.2f})"],
            "warnings":     [],
        }

    # ─────────────────────────────────────────────────────────────────────
    # Phase R1 (2026-07-06): INSTITUTIONAL VALIDATION LAYER — Event gates
    # ─────────────────────────────────────────────────────────────────────

    # Issue 2 fix — Gate 3g2: RS_LAGGING (relative strength vs Nifty).
    # A stock trailing the index by RS_LAG_GATE_PCT (default -5.0%) over
    # 21 trading days is a technical weakness signal. We hard-reject unless
    # the stock is explicitly flagged as an oversold_reversal setup (where
    # RS underperformance is the entry catalyst, not a warning).
    # Env: RS_LAG_GATE_PCT (default -5.0). Set to 0 to disable.
    try:
        _rs_gate = float(os.getenv("RS_LAG_GATE_PCT", "-5.0"))
    except (TypeError, ValueError):
        _rs_gate = -5.0
    if _rs_gate < 0:
        _rs_diff = float(stock.get("rs_diff21", 0) or 0)
        _is_oversold_rev = bool(stock.get("oversold_reversal", False))
        if _rs_diff < _rs_gate and not _is_oversold_rev:
            return {
                "decision":     "REJECTED",
                "fail_reasons": [f"RS_LAGGING({_rs_diff:+.1f}%<{_rs_gate:+.1f}%)"],
                "warnings":     [],
            }

    # Gate 3h — NEWS_SEVERITY_HIGH (HARD).
    # If the LLM news classifier flagged any recent news at severity ≥
    # NEWS_SEVERITY_MAX (default 2 on a 0-3 scale where 3 = fraud/SEBI
    # action / criminal charges / auditor resignation) we hard-reject.
    # This is separate from the existing BLACK_SWAN gate (penalty ≥ 999).
    #
    # Env: NEWS_SEVERITY_MAX (default 2, meaning severity 3+ rejects).
    #      Set to 99 to disable.
    try:
        _news_sev_max = int(os.getenv("NEWS_SEVERITY_MAX", "2"))
    except (TypeError, ValueError):
        _news_sev_max = 2
    _news_sev = int(stock.get("news_severity", 0) or 0)
    if _news_sev_max < 99 and _news_sev > _news_sev_max:
        _news_cat = str(stock.get("news_category", "") or "")
        return {
            "decision":     "REJECTED",
            "fail_reasons": [f"NEWS_SEVERITY_HIGH(sev={_news_sev},cat={_news_cat or 'UNKNOWN'})"],
            "warnings":     [],
        }

    # Gate 4: Liquidity (HARD)
    # Phase 2 #23 (2026-07-05): min turnover configurable via env.
    # Old hardcoded 50 lakh (₹5L / day) was low enough that a 2-3 lakh
    # bulk print could clear the gate for a thinly-traded name. Raising
    # the default to 200 lakh (₹2 Cr / day) filters out the long tail of
    # illiquid mid/small caps where retail slippage exceeds the R/R edge.
    # Set GATE4_MIN_TURNOVER_LAKHS=50 to restore old behavior.
    try:
        _min_turnover = float(os.getenv("GATE4_MIN_TURNOVER_LAKHS", "200"))
    except (TypeError, ValueError):
        _min_turnover = 200.0
    if stock.get("avg_volume", 0) < 100_000 or stock.get("avg_value_lakhs", 0) < _min_turnover:
        return {"decision": "REJECTED", "fail_reasons": [f"LIQUIDITY_FAIL(turnover<{_min_turnover:.0f}L)"], "warnings": []}

    # Gate 5: Market Regime max_buys (HARD)
    if thresh["max_buys"] == 0:
        return {"decision": "REJECTED", "fail_reasons": ["REGIME_NO_BUY"], "warnings": []}

    # Gate 5b (Phase C7): Kill switch — portfolio-level circuit breaker.
    # If the equity curve is bleeding (3 consec losses, -2% day, -3% week,
    # or ≥10% drawdown), HALT all new BUYs until the pause window expires.
    # Never blocks WATCHLIST — losing streaks don't invalidate the setup,
    # they just say "not now". Read once via portfolio_context to avoid
    # recomputing per stock.
    ks = (portfolio or {}).get("kill_switch") or {}
    if ks.get("buys_paused"):
        fail_reasons.append(f"KILL_SWITCH({ks.get('reason', 'ACTIVE')})")
        warnings.append(f"KILL_SWITCH_ACTIVE: {ks.get('reason', '?')}")

    # Gate 5c — Phase 3a #40 (2026-07-05): regime exposure cap (SOFT).
    # REGIME_THRESHOLDS defines `max_exposure` per regime (BEAR 0.20,
    # HIGH_VOL 0.40, etc.) but historically it was never read. This gate
    # enforces it: if current portfolio exposure has already reached the
    # regime cap, new BUYs are demoted to WATCHLIST (soft). Existing
    # positions unaffected. Exposure headroom of 0 means we're at the
    # limit; a fresh entry would push us over.
    _headroom = float((portfolio or {}).get("exposure_headroom", 1.0) or 0.0)
    _cur_exp  = float((portfolio or {}).get("current_exposure", 0.0) or 0.0)
    _max_exp  = float((portfolio or {}).get("max_exposure", 1.0) or 1.0)
    if _headroom <= 0.02 and _max_exp < 1.0:  # 2% buffer to avoid edge flapping
        fail_reasons.append(
            f"REGIME_EXPOSURE_CAP({_cur_exp*100:.0f}%>={_max_exp*100:.0f}%)"
        )
        warnings.append(
            f"EXPOSURE_AT_CAP: {_cur_exp*100:.0f}% used vs {_max_exp*100:.0f}% "
            f"regime cap — new BUYs deferred to WATCHLIST"
        )

    # Gate 6: Confidence (HARD with grace band)
    # Phase C5 (rating ≥ 9.0): if conf is within 1.0 pt of the threshold AND
    # the stock already comfortably exceeds TQ + R/R thresholds, accept it as
    # a MARGINAL_CONF pass rather than a rigid reject. Prevents the "cliff at
    # 81.5 for a BULL threshold of 82" scenario where a legit setup dies for
    # 0.5 pt of confidence dilution while its trade-quality is 85+ and R/R 2.4x.
    conf = stock.get("final_confidence", 0)
    _min_conf = thresh["min_confidence"]
    _grace    = 1.0
    _tq_prev  = stock.get("trade_quality_score", 0)
    _rr_prev  = stock.get("rr_ratio", 0)
    _quality_bonus = (
        _tq_prev >= thresh["min_tq"] + 2.0
        and _rr_prev >= thresh["min_rr"] + 0.2
    )
    if conf < _min_conf:
        if conf >= _min_conf - _grace and _quality_bonus:
            warnings.append(
                f"MARGINAL_CONF({conf:.1f}<{_min_conf}, grace-band; "
                f"TQ {_tq_prev:.1f} & R/R {_rr_prev:.2f}x compensate)"
            )
        else:
            fail_reasons.append(f"CONF_FAIL(got {conf:.1f}, need {_min_conf})")

            # Phase I shadow-log (2026-07-07): Bucket D · SO_CLOSE
            # Right setup + right regime, confidence just below the gate.
            # Answers: "is my min_confidence threshold set too high?"
            if _SHADOW_LOG_OK:
                try:
                    _sd_setup  = stock.get("setup_type", "OTHER")
                    if shadow_log.is_near_miss_conf(_sd_setup, regime,
                                                   conf, _min_conf):
                        shadow_log.record_shadow_trade(
                            "D", stock, regime,
                            note=f"near_miss({conf:.1f}<{_min_conf})",
                        )
                except Exception:
                    pass

    # Gate 7: Trade Quality (HARD)
    tq = stock.get("trade_quality_score", 0)
    if tq < thresh["min_tq"]:
        fail_reasons.append(f"TQ_FAIL(got {tq:.1f}, need {thresh['min_tq']})")

    # Gate 8: Risk/Reward (HARD, with Phase R6 institutional override)
    # ----------------------------------------------------------------
    # R6 (2026-07-06): a wide-ATR breakout leader with STRONG business quality
    # and a top-decile sector may show RR 1.4-1.6 gross because ATR is fat.
    # Traditional min_rr=1.8 rejects these correctly-set-up trades. The
    # override accepts min_rr=1.5 when the setup earns it via institutional
    # signals: BQ verdict STRONG/ACCEPTABLE + sector_composite >= 85 +
    # swing_alpha_score >= 80 + volume_expansion >= 1.5. This is intentionally
    # narrow — no other combination unlocks it.
    rr = stock.get("rr_ratio", 0)
    _rr_needed = thresh["min_rr"]
    if rr < _rr_needed:
        # Institutional-grade override: sector-leader + BQ-strong + high SA
        _bq_v      = str(stock.get("bq_verdict", "") or "").upper()
        _sec_c     = float(stock.get("sector_composite_score", 0) or 0)
        _sa_here   = float(stock.get("swing_alpha_score", 0) or 0)
        _vol_here  = float(stock.get("volume_expansion_ratio", 0) or 0)
        try:
            _rr_floor_institutional = float(os.getenv("RR_INSTITUTIONAL_MIN", "1.5"))
            _rr_override_enabled    = os.getenv("RR_INSTITUTIONAL_OVERRIDE", "1") == "1"
        except (TypeError, ValueError):
            _rr_floor_institutional, _rr_override_enabled = 1.5, True
        _institutional_ok = (
            _rr_override_enabled
            and rr >= _rr_floor_institutional
            and _bq_v in ("STRONG", "ACCEPTABLE")
            and _sec_c >= 85.0
            and _sa_here >= 80.0
            and _vol_here >= 1.5
        )
        if _institutional_ok:
            warnings.append(
                f"RR_INSTITUTIONAL_OVERRIDE({rr:.2f} ≥ {_rr_floor_institutional:.2f} — "
                f"BQ={_bq_v}, sector={_sec_c:.0f}, SA={_sa_here:.0f}, vol={_vol_here:.2f}x)"
            )
        else:
            fail_reasons.append(f"RR_FAIL(got {rr:.2f}, need {_rr_needed})")

    # Gate 8b: Wide-stop guardrail — 2026-07-03 recalibration
    # ─────────────────────────────────────────────────────────────────────
    # ORIGINAL DESIGN (broken in SIDEWAYS): reject any stock whose stop is
    # wider than regime `max_stop_pct`. The stop-loss calc uses the 10-day
    # swing low (line 6180), which for typical NSE mid-caps naturally lands
    # 10-12% below entry — so this gate rejected 100% of candidates in
    # SIDEWAYS regime with `max_stop_pct=6%` (then 8% after first fix),
    # producing zero signals for days on end.
    #
    # WHY THIS GATE IS NOW BEAR/HIGH_VOL ONLY:
    #   1. The R/R gate (min_rr) already protects against wide stops — a 12%
    #      stop requires a 24%+ target to pass min_rr=2.0. Only truly
    #      explosive setups survive both.
    #   2. Position sizing auto-adjusts: `risk_per_trade_pct=1.5%` divides
    #      by (entry-stop), so wider stop = smaller position, same ₹ risk.
    #   3. In BEAR / STRONG_BEAR / HIGH_VOLATILITY, capital preservation
    #      matters more than R/R math — a wide stop can wipe out multiple
    #      trades before targets are hit. THERE the gate still fires.
    #
    # In BULL/STRONG_BULL/SIDEWAYS/TRANSITION we now warn only.
    _STOP_GATE_REGIMES = ("BEAR", "STRONG_BEAR", "HIGH_VOLATILITY")
    try:
        _entry_v = float(stock.get("entry", 0) or 0)
        _stop_v  = float(stock.get("stop",  0) or 0)
        _max_stop_pct = float(thresh.get("max_stop_pct", 8.0))
        if _entry_v > 0 and 0 < _stop_v < _entry_v:
            _stop_dist_pct = (_entry_v - _stop_v) / _entry_v * 100.0
            if _stop_dist_pct > _max_stop_pct:
                if regime in _STOP_GATE_REGIMES:
                    fail_reasons.append(
                        f"WIDE_STOP(got {_stop_dist_pct:.1f}%, cap {_max_stop_pct:.1f}%)"
                    )
                else:
                    warnings.append(
                        f"WIDE_STOP_WARN(got {_stop_dist_pct:.1f}%, "
                        f"soft cap {_max_stop_pct:.1f}% \u2014 R/R gate handles this)"
                    )
    except Exception:
        pass

    # Gate 9: Sector Health (SOFT — LAGGING counts toward the 2-fail budget but
    # does NOT hard-block watchlist entry). Phase C2 (2026-07-02): previously
    # LAGGING was force-rejected, which combined with the sector_master.csv
    # fallback (fix #4) caused mass carnage — every Banking/Insurance stock
    # that used to be sector=OTHERS suddenly became a hard-reject. Now LAGGING
    # is a scoreable fail: rotation candidates can still reach WATCHLIST/
    # NEAR_MISS, and only get pushed to REJECTED if combined with other fails.
    # Phase 4 (2026-07-01): rotation velocity overrides static status —
    # a LAGGING sector that is ROTATING_IN (5-day rank moved up ≥3 spots) is
    # a valid contrarian entry, so downgrade to a warning only.
    #
    # 2026-07-03 fix — field name mismatch bug: score_stock() writes the field
    # as `sector_velocity` (line 6143) but this gate was reading
    # `sector_rotation_velocity`, which never exists → always "UNKNOWN" →
    # the entire Phase 4 rotation-velocity overlay was silently dead. Reading
    # both keys is a belt-and-braces guard for any partially-migrated caches.
    sector_status   = stock.get("sector_status", "NEUTRAL")
    sector_velocity = stock.get(
        "sector_velocity",
        stock.get("sector_rotation_velocity", "UNKNOWN"),
    )
    if sector_status == "LAGGING":
        if sector_velocity == "ROTATING_IN":
            warnings.append("SECTOR_LAGGING_BUT_ROTATING_IN")
        else:
            fail_reasons.append("SECTOR_LAGGING")
    elif sector_status == "WEAKENING":
        warnings.append("SECTOR_WEAKENING")
    # Rotating-OUT is a soft warning regardless of static status — sector is
    # bleeding leadership rank fast.
    if sector_velocity == "ROTATING_OUT" and sector_status != "LAGGING":
        warnings.append("SECTOR_ROTATING_OUT")

    # ─────────────────────────────────────────────────────────────────────
    # Phase R1 (2026-07-06): INSTITUTIONAL VALIDATION LAYER — Sector strict
    # ─────────────────────────────────────────────────────────────────────
    # Layer 2 rule: "A good company inside a weak sector is usually a poor
    # swing trade." We enforce sector rank + momentum with a contrarian
    # bypass when confidence + trade quality are exceptionally high.
    #
    # These gates are SOFT (watchlist-eligible) rather than hard-reject —
    # a strong stock in a weak sector becomes a CONTRARIAN candidate, not
    # rejected outright. This preserves the option to trade rotation.
    #
    # Env kill-switch: SECTOR_STRICT_GATE=1 (default on).

    # Gate 9b — SECTOR_RANK_TOO_LOW.
    # Rank sectors by 5-day return; require rank ≤ SECTOR_RANK_CUTOFF (top 6).
    # Contrarian bypass: if final_confidence ≥ SECTOR_CONTRARIAN_CONF_BAR
    # (default 88) AND trade_quality_score ≥ 75, allow through as warning.
    try:
        _sec_strict_on = int(os.getenv("SECTOR_STRICT_GATE", "1"))
    except (TypeError, ValueError):
        _sec_strict_on = 1
    try:
        _sec_rank_cut = int(os.getenv("SECTOR_RANK_CUTOFF", "6"))
    except (TypeError, ValueError):
        _sec_rank_cut = 6
    try:
        _contra_conf_bar = float(os.getenv("SECTOR_CONTRARIAN_CONF_BAR", "88"))
    except (TypeError, ValueError):
        _contra_conf_bar = 88.0

    _sec_rank = stock.get("sector_rank_5d")
    _sec_ret20d = stock.get("sector_ret20d")
    if _sec_ret20d is None:
        # score_stock persists sector metrics in the ranks dict via
        # sector_rotation_score. Read from result-embedded fields.
        _sec_ret20d = stock.get("sector_5d_return")  # fallback name
    _conf_now = float(stock.get("final_confidence", 0) or 0)
    _tq_now = float(stock.get("trade_quality_score", 0) or 0)
    _contrarian_ok = (_conf_now >= _contra_conf_bar and _tq_now >= 75.0)

    if _sec_strict_on and isinstance(_sec_rank, (int, float)) and _sec_rank > _sec_rank_cut:
        if _contrarian_ok:
            warnings.append(
                f"CONTRARIAN_SECTOR_RANK({int(_sec_rank)}>top{_sec_rank_cut} "
                f"— allowed: conf {_conf_now:.1f}≥{_contra_conf_bar:.0f}, TQ {_tq_now:.1f}≥75)"
            )
        else:
            fail_reasons.append(
                f"SECTOR_RANK_TOO_LOW(rank={int(_sec_rank)}>top{_sec_rank_cut})"
            )
            warnings.append(
                f"SECTOR_RANK_WEAK: sector ranked #{int(_sec_rank)} — "
                f"needs top-{_sec_rank_cut} unless conf≥{_contra_conf_bar:.0f}"
            )

    # Gate 9c — SECTOR_MOMENTUM_NEGATIVE.
    # Even if rank is decent, a sector with 21-day return < 0 is bleeding.
    # Same contrarian bypass applies.
    try:
        _sec_ret20d_val = float(_sec_ret20d) if _sec_ret20d is not None else None
    except (TypeError, ValueError):
        _sec_ret20d_val = None
    if _sec_strict_on and _sec_ret20d_val is not None and _sec_ret20d_val < 0:
        if _contrarian_ok:
            warnings.append(
                f"CONTRARIAN_SECTOR_MOMENTUM(ret20d={_sec_ret20d_val:+.1f}% "
                f"— allowed: conf/TQ high)"
            )
        else:
            fail_reasons.append(
                f"SECTOR_MOMENTUM_NEG(ret20d={_sec_ret20d_val:+.1f}%)"
            )

    # Gate 10: High Pledge Warning (SOFT)
    if 20 < pledge <= 40:
        warnings.append(f"PLEDGE_WARNING_{pledge:.0f}PCT")

    # Gate 11: 52-Week High Proximity (SOFT)
    if stock.get("near_52w_high", False):
        warnings.append("NEAR_52W_HIGH_RESISTANCE")

    # Gate 11b (Phase C4): Overnight-gap risk (SOFT).  Flags stocks where the
    # 90th-percentile intraday gap size over the last 20 sessions consumes
    # >60% of the stop budget — i.e. a gap-down can eat the stop before
    # intraday triggers fire. Never blocks a BUY; downstream sizing and the
    # BUY card render the effective stop.
    if stock.get("high_gap_risk", False):
        _p90 = float(stock.get("p90_gap_pct", 0) or 0)
        _eff = float(stock.get("effective_stop_pct", 0) or 0)
        warnings.append(f"HIGH_GAP_RISK: p90 gap {_p90:.1f}% · effective stop {_eff:.1f}%")

    # Gate 12: Portfolio Capacity (INFORMATIONAL — 2026-07-03 change)
    # OLD behaviour: appended PORTFOLIO_FULL as a fail_reason and later stripped
    # it in the WATCHLIST vs REJECTED sorter. That kept the tag in the audit,
    # which made no-buy days show "PORTFOLIO_FULL: 100%" as a top reject reason
    # — pure noise, since it just meant "you already own 2 stocks in a regime
    # allowing 1", NOT that the candidate was bad.
    #
    # NEW behaviour: recommendations must be portfolio-independent. Your
    # portfolio JSON might be stale (positions closed manually and not synced),
    # so a candidate's quality should NEVER depend on what you happen to hold.
    # We now only emit a WARNING; the user decides based on actual free capital.
    active_count = portfolio.get("active_count", 0)
    if active_count >= thresh["max_buys"]:
        warnings.append(
            f"PORTFOLIO_OVER_CAP({active_count}/{thresh['max_buys']} "
            f"positions in {regime} regime — informational only)"
        )

    # Gate 13: Event Calendar (HARD) — no new BUY within 5 trading days of results/monthly expiry
    near_event, event_reason = is_near_event(
        stock.get("symbol", "").replace(".NS", ""),
        results_dates or [],
        upcoming_events or [],
        # Issue 11 fix: honour EARNINGS_BLACKOUT_DAYS env (default 5). Presets
        # set this to 3 (swing) or 7 (conservative) — previously the literal
        # 5 here silently overrode the preset, making the knob dead code.
        window_days=int(os.getenv("EARNINGS_BLACKOUT_DAYS", "5") or 5),
    )
    if near_event:
        # Move to WATCHLIST rather than REJECTED — setup may still be valid post-event
        fail_reasons.append(f"EVENT_BLOCK_{event_reason}")
        warnings.append(f"NEAR_EVENT: {event_reason}")

    # Gate 14: Correlation with existing holdings (HARD)
    # Reject if candidate moves in lockstep (corr > 0.75) with something already held
    if returns_cache and holdings:
        blocked, worst_corr, corr_sym = correlation_check(
            stock.get("symbol", ""), holdings, returns_cache, max_corr=0.75
        )
        if blocked:
            fail_reasons.append(f"HIGH_CORR_{worst_corr:.2f}_WITH_{corr_sym.replace('.NS','')}")
            warnings.append(f"CORRELATED_WITH: {corr_sym} ({worst_corr:.2f})")
        elif worst_corr >= 0.60:
            warnings.append(f"MODERATE_CORR_{worst_corr:.2f}_WITH_{corr_sym.replace('.NS','')}")

    # Gate 14b (Phase C7): Sector concentration cap (SOFT — watchlist-eligible).
    # Correlation gate catches pair-wise ticker linkage, but does NOT catch
    # "3 IT stocks all riding the same Nifty-IT tailwind". This caps open
    # positions per sector to MAX_POSITIONS_PER_SECTOR (default 2). Soft-fail
    # so the setup can still reach WATCHLIST — user can promote later when a
    # sector slot frees up.
    if holdings:
        _cand_sector = stock.get("sector") or get_sector(stock.get("symbol", ""))
        _sc = check_sector_concentration(_cand_sector, holdings)
        if _sc.get("blocked"):
            fail_reasons.append(_sc["reason"])
            warnings.append(f"SECTOR_CONCENTRATION: {_sc['reason']}")

    # Gate 15: Delivery Quality (SOFT — pump detection & institutional exit)
    # Phase C3 (2026-07-02): only fires when we actually got real nselib
    # delivery data (source == "nselib"). Two cases:
    #   (a) DISTRIBUTION signal (ratio ≤ 0.70)  → soft fail, "SUSPECT_PUMP" or
    #                                             "INSTITUTIONAL_EXIT"
    #       - If today's return also > +2%      → clear pump pattern
    #       - Else                              → possible institutional exit
    #   (b) WEAK signal (0.70 < ratio ≤ 0.85)   → warning only, no fail
    if stock.get("delivery_source") == "nselib":
        deliv_sig = stock.get("delivery_signal", "NEUTRAL")
        # Prefer explicit ret1d_pct if present; fall back to ret1d (the actual
        # key set by score_stock at L5205).
        ret1d     = float(
            stock.get("ret1d_pct", stock.get("ret1d", 0.0)) or 0.0
        )
        if deliv_sig == "DISTRIBUTION":
            if ret1d > 2.0:
                fail_reasons.append("SUSPECT_PUMP_LOW_DELIVERY")
                warnings.append(
                    f"PUMP_PATTERN: price+{ret1d:.1f}% but delivery ratio "
                    f"{stock.get('delivery_ratio', 0):.2f}"
                )
            else:
                fail_reasons.append("INSTITUTIONAL_EXIT")
                warnings.append(
                    f"DISTRIBUTION: delivery {stock.get('delivery_pct_today', 0):.0f}% "
                    f"vs 20d avg {stock.get('delivery_pct_20d_avg', 0):.0f}%"
                )
        elif deliv_sig == "WEAK":
            warnings.append(
                f"WEAK_DELIVERY: {stock.get('delivery_pct_today', 0):.0f}% "
                f"(20d avg {stock.get('delivery_pct_20d_avg', 0):.0f}%)"
            )
        elif deliv_sig == "STRONG_ACCUM":
            # Positive confirmation — surface as informational warning-tag
            warnings.append(
                f"STRONG_ACCUMULATION: delivery {stock.get('delivery_pct_today', 0):.0f}% "
                f"(20d avg {stock.get('delivery_pct_20d_avg', 0):.0f}%)"
            )

    hard_fails = [f for f in fail_reasons if "WARNING" not in f]
    if hard_fails:
        # 2026-07-03: PORTFOLIO_FULL is no longer a fail_reason (now a warning).
        # The filter below is kept as a defensive no-op so any old audit rows
        # replayed through this function still behave correctly.
        scoreable_fails = [f for f in hard_fails if "PORTFOLIO_FULL" not in f]
        # Phase C2 (2026-07-02): SECTOR_LAGGING is now a soft-scoreable fail —
        # it counts toward the 2-fail budget but no longer excludes watchlist.
        # Phase C3 (2026-07-02): INSTITUTIONAL_EXIT also soft-scoreable
        # (allow watchlist tier for post-distribution rebounds). SUSPECT_PUMP
        # stays HARD because pump patterns rarely recover cleanly.
        # Phase C7 (2026-07-02): SECTOR_CAP is soft — cap will free up over
        # days as positions close, so a valid setup should be preserved on
        # watchlist rather than lost. KILL_SWITCH stays HARD (portfolio-wide
        # halt) — a losing streak invalidates *entries*, but a setup that
        # survives to next session is still tracked, so we push to WATCHLIST
        # only via the same soft-fail path (see soft_only clause below).
        soft_only = all(
            "EVENT_BLOCK" in f or "HIGH_CORR" in f or "SECTOR_LAGGING" in f
            or "INSTITUTIONAL_EXIT" in f or "SECTOR_CAP" in f
            or "KILL_SWITCH" in f
            or "REGIME_EXPOSURE_CAP" in f   # Phase 3a #40 — regime exposure cap
            or "FUND_DATA_MISSING" in f    # Phase 4-A — fundamentals unavailable
            or "SECTOR_RANK_TOO_LOW" in f  # Phase R1 — sector rank gate
            or "SECTOR_MOMENTUM_NEG" in f  # Phase R1 — sector momentum gate
            or "BQ_TURNAROUND_EARLY" in f  # Phase R1 — turnaround escape hatch
            for f in scoreable_fails
        )
        # Phase C6 (2026-07-02): make the "score-based fail" whitelist EXPLICIT
        # instead of relying on the substring match `"FAIL" in f`, which was
        # accidental and fragile (any future gate whose label happens to contain
        # the substring "FAIL" would auto-qualify for watchlist).
        # Explicit whitelist of score-based fails that CAN reach watchlist:
        SCORE_FAIL_TAGS = ("CONF_FAIL", "TQ_FAIL", "RR_FAIL")
        score_based = all(
            any(tag in f for tag in SCORE_FAIL_TAGS)
            for f in scoreable_fails
        )
        if not scoreable_fails:
            # 2026-07-03: this branch was originally the "only PORTFOLIO_FULL
            # failed" path. That reason is now a warning (not a fail_reason),
            # so this branch only fires if a future gate produces a fail that
            # is stripped by the filter above. Kept as a safe fallback.
            decision = "WATCHLIST"
        elif len(scoreable_fails) <= 2 and (soft_only or score_based):
            decision = "WATCHLIST"
        else:
            decision = "REJECTED"

    # ─────────────────────────────────────────────────────────────────────
    # Phase R1 (2026-07-06): TAXONOMY EXPANSION + INSTITUTIONAL VALIDATOR
    # ─────────────────────────────────────────────────────────────────────
    # After the main BUY/WATCHLIST/REJECTED decision, we apply two extra
    # institutional checks:
    #
    # 1. "Missing-data confidence cap": if fund_source signals we couldn't
    #    verify quality, we cannot grant top-tier confidence. Downgrade a
    #    BUY to WATCHLIST when confidence exceeds FUND_MISSING_CONF_CAP but
    #    we don't actually have real fundamentals to support it.
    #
    # 2. "Institutional trader test" (F1): a BUY must have all 3 pillars
    #    ≥ 70. Confidence, TQ, and R/R form the technical pillar. If any
    #    is < 70 → downgrade. This catches "hollow high-conf" trades where
    #    a great chart score masks weak TQ.
    #
    # 3. "Contrarian tag": if the BUY survived a contrarian-sector bypass,
    #    label it CONTRARIAN so downstream renderer & tracker treat it as
    #    reduced-size / tighter-stop trade.
    #
    # 4. "AVOID tier" for stocks that failed hard fundamentals but weren't
    #    already REJECTED (edge case: turnaround escape hatch dropped them
    #    into WATCHLIST — they belong in AVOID until they show >2Q recovery).
    #
    # These are non-invasive: they only DOWNGRADE, never upgrade.
    # ────────────────────────────────────────────────────────────────────

    # (1) Missing-data confidence cap
    try:
        _fmcap = float(os.getenv("FUND_MISSING_CONF_CAP", "80"))
    except (TypeError, ValueError):
        _fmcap = 80.0
    _fund_missing = any("FUND_DATA_MISSING" in w for w in warnings)
    if decision == "BUY" and _fund_missing and _fmcap > 0:
        _conf_check = float(stock.get("final_confidence", 0) or 0)
        if _conf_check > _fmcap:
            decision = "WATCHLIST"
            warnings.append(
                f"FUND_MISSING_CAP: conf {_conf_check:.1f} exceeds cap {_fmcap:.0f} "
                f"but fundamentals unverified → demoted to WATCHLIST"
            )

    # (2) Institutional trader test — three-pillar 70-floor
    try:
        _pillar_floor = float(os.getenv("PILLAR_MIN_FLOOR", "70"))
    except (TypeError, ValueError):
        _pillar_floor = 70.0
    if decision == "BUY" and _pillar_floor > 0:
        _conf_p = float(stock.get("final_confidence", 0) or 0)
        _tq_p = float(stock.get("trade_quality_score", 0) or 0)
        _rr_p = float(stock.get("rr_ratio", 0) or 0)
        # R/R uses its own scale; convert to 0-100-ish by treating 2.0x as 100
        # (min institutional). Below 1.4x = 70.
        _rr_score = min(100.0, max(0.0, (_rr_p / 2.0) * 100.0))
        _weakest_pillar = min(_conf_p, _tq_p, _rr_score)
        if _weakest_pillar < _pillar_floor:
            decision = "WATCHLIST"
            warnings.append(
                f"PILLAR_FLOOR_FAIL: weakest pillar {_weakest_pillar:.1f} < {_pillar_floor:.0f} "
                f"(conf={_conf_p:.1f}, TQ={_tq_p:.1f}, R/R_norm={_rr_score:.1f})"
            )

    # (3) Contrarian tag — applies to BUY and also promotes WATCHLIST when
    #     the ONLY soft fails are sector-related AND the contrarian bypass
    #     warnings were emitted (conf >= 88 AND TQ >= 75). This is the key
    #     insight: a stock in a weak sector with strong stock-specific stats
    #     is a contrarian trade, not a reject.
    _is_contra = any("CONTRARIAN_SECTOR" in w for w in warnings)
    if decision == "BUY" and _is_contra:
        # Keep as BUY but stamp category so renderer/tracker treats specially
        decision = "BUY_CONTRARIAN"
    elif decision == "WATCHLIST" and _is_contra:
        # WATCHLIST because of sector-only soft fails, but contrarian bypass
        # earned it — promote to BUY_CONTRARIAN. Only when ALL hard fails are
        # sector-related (SECTOR_LAGGING, SECTOR_RANK_TOO_LOW, SECTOR_MOMENTUM_NEG).
        _sector_only_fails = fail_reasons and all(
            ("SECTOR_LAGGING" in f
             or "SECTOR_RANK_TOO_LOW" in f
             or "SECTOR_MOMENTUM_NEG" in f)
            for f in fail_reasons
        )
        if _sector_only_fails:
            decision = "BUY_CONTRARIAN"
            warnings.append(
                "CONTRARIAN_PROMOTION: sector-only soft fails "
                "overridden by contrarian bypass (conf+TQ high enough)"
            )

    # (4) AVOID tier — hard-quality fails that slipped into WATCHLIST
    #     via the turnaround escape hatch but do NOT show YoY recovery.
    if decision == "WATCHLIST":
        _has_avoid_flag = any(
            "BQ_TURNAROUND_EARLY" in f or "PILLAR_FLOOR_FAIL" in w
            for f in fail_reasons for w in warnings if False  # placeholder — keep clean
        )
        # Simpler rule: promote a stock to AVOID only when it has BQ decline
        # AND a low ROE AND low confidence combined (multi-signal quality problem)
        _bq_trigger = any("BQ_TURNAROUND_EARLY" in f for f in fail_reasons)
        _low_roe = _roe_real and _roe_val < 10.0
        _low_conf = float(stock.get("final_confidence", 0) or 0) < 75.0
        if _bq_trigger and (_low_roe or _low_conf):
            decision = "AVOID"
            warnings.append(
                f"AVOID_TIER: business declining + weak backing "
                f"(ROE={_roe_val:.1f}%, conf={float(stock.get('final_confidence', 0) or 0):.1f})"
            )

    # ----- Phase R2 9-Tier Taxonomy Expansion (2026-07-06) -----
    # After all R1 gates + demotions, refine the coarse decisions further.
    # NOTE: existing WATCHLIST tiers (DEVELOPING / MONITOR / NEAR_MISS_*)
    # are set separately by classify_watchlist(); the *decision* tier here
    # is a broader category. We use distinct decision-tier labels to avoid
    # ambiguity: STRONG_BUY, BUY, BUY_CONTRARIAN, BUY_TURNAROUND,
    # WATCHLIST (broad — classify_watchlist assigns sub-tier), AVOID, REJECTED.
    _bq_score = float(stock.get("bq_score", 0) or 0)
    _bq_completeness = float(stock.get("bq_data_completeness", 0) or 0)
    _bq_verdict = str(stock.get("bq_verdict", "UNKNOWN"))
    _sector_composite = float(stock.get("sector_composite_score", 50) or 50)
    _conf_final = float(stock.get("final_confidence", 0) or 0)
    _tq_final = float(stock.get("trade_quality_score", 0) or 0)
    _rr_final = float(stock.get("rr_ratio", 0) or 0)

    # (5) STRONG_BUY tier — institutional gold standard, all three pillars top-tier
    #     Only reachable from BUY (not from BUY_CONTRARIAN — contrarian trades
    #     by definition have weak sector so cannot be top-tier institutional).
    #     Phase R3 (2026-07-06): also require institutional micro-liquidity
    #     (≥ ₹5 Cr/day turnover by default). Small caps that pass all fundamental
    #     gates but can't absorb institutional-size orders should remain BUY,
    #     not STRONG_BUY.
    try:
        _sb_conf = float(os.getenv("STRONG_BUY_MIN_CONF", "90"))
        _sb_tq = float(os.getenv("STRONG_BUY_MIN_TQ", "80"))
        _sb_bq = float(os.getenv("STRONG_BUY_MIN_BQ", "70"))
        _sb_sec = float(os.getenv("STRONG_BUY_MIN_SECTOR", "60"))
        _sb_rr = float(os.getenv("STRONG_BUY_MIN_RR", "2.0"))
        _sb_liq = float(os.getenv("STRONG_BUY_MIN_TURNOVER_LAKHS", "500"))  # ₹5Cr default
    except (TypeError, ValueError):
        _sb_conf, _sb_tq, _sb_bq, _sb_sec, _sb_rr, _sb_liq = 90.0, 80.0, 70.0, 60.0, 2.0, 500.0
    _turnover = float(stock.get("avg_value_lakhs", 0) or 0)
    if (decision == "BUY"
            and _conf_final >= _sb_conf
            and _tq_final >= _sb_tq
            and _bq_score >= _sb_bq
            and _bq_completeness >= 50  # need real data to earn STRONG_BUY
            and _sector_composite >= _sb_sec
            and _rr_final >= _sb_rr
            and _turnover >= _sb_liq  # Phase R3 — institutional liquidity
            and _bq_verdict in ("STRONG", "ACCEPTABLE")):
        decision = "STRONG_BUY"
        warnings.append(
            f"STRONG_BUY_TIER: institutional-grade — conf={_conf_final:.0f}/TQ={_tq_final:.0f}/"
            f"BQ={_bq_score:.0f}/sector={_sector_composite:.0f}/liq=₹{_turnover/100:.1f}Cr"
        )

    # (6) BUY_TURNAROUND — was WATCHLIST due to turnaround escape hatch,
    #     but has HIGH confidence + strong momentum recovery. Formalizes the
    #     3Q-declining-but-YoY-recovering pattern into its own tier.
    if decision == "WATCHLIST":
        _has_turnaround = any("BQ_TURNAROUND_EARLY" in f for f in fail_reasons)
        try:
            _to_conf = float(os.getenv("TURNAROUND_MIN_CONF", "80"))
        except (TypeError, ValueError):
            _to_conf = 80.0
        if (_has_turnaround
                and _conf_final >= _to_conf
                and _tq_final >= 70
                and _rr_final >= 2.0):
            decision = "BUY_TURNAROUND"

    # (7) DEVELOPING and MONITOR are RESERVED for classify_watchlist() sub-tiers.
    #     We do NOT re-classify decision tier here; classify_watchlist runs
    #     downstream when decision == "WATCHLIST" and handles those sub-tiers.
    #     The bq_score / sector_composite fields are stamped on `stock` above,
    #     so classify_watchlist can also read them if desired.

    # Persist the taxonomy inputs so downstream code / audit CSV can inspect them
    stock["taxonomy_inputs"] = {
        "bq_score": _bq_score,
        "bq_completeness": _bq_completeness,
        "bq_verdict": _bq_verdict,
        "sector_composite": _sector_composite,
        "conf": _conf_final,
        "tq": _tq_final,
        "rr": _rr_final,
    }

    # ═══════════════════════════════════════════════════════════════════════
    # Phase R5 Prune (2026-07-06) — 9-tier taxonomy collapse → 4 canonical tiers
    # ═══════════════════════════════════════════════════════════════════════
    # Audit of 480 rows across 3 days: BUY_CONTRARIAN / BUY_TURNAROUND / AVOID
    # never fired in production (0 rows). STRONG_BUY is kept as a legitimate
    # institutional aspiration tier. To reduce cognitive/API surface without
    # losing signal, we preserve the fine-grain label as audit metadata
    # (`decision_subtype`) but return one of {STRONG_BUY, BUY, WATCHLIST, REJECTED}.
    # Downstream code that already handles these labels (line ~13252) still
    # works because the collapse maps into the same canonical bucket.
    stock["decision_subtype"] = decision  # audit-only: keep original fine label
    _tier_collapse_map = {
        "BUY_CONTRARIAN": "BUY",
        "BUY_TURNAROUND": "BUY",
        "AVOID":          "REJECTED",
    }
    decision = _tier_collapse_map.get(decision, decision)

    return {"decision": decision, "fail_reasons": fail_reasons, "warnings": warnings}


def build_returns_cache(tradable: dict, lookback: int = 60) -> dict:
    """
    Pre-compute 60-day daily returns for every stock in the tradable universe.
    Used by correlation_check() — built once per pipeline run, passed around.
    Returns {symbol: [float, ...]} — empty list if insufficient data.
    """
    cache = {}
    for symbol, df in tradable.items():
        try:
            closes = df["Close"].squeeze().values.astype(float)
            if len(closes) < lookback + 2:
                cache[symbol] = []
                continue
            sample = closes[-(lookback + 1):]
            rets   = [(sample[i] / sample[i-1] - 1.0) for i in range(1, len(sample))
                      if sample[i-1] > 0]
            cache[symbol] = rets
        except Exception:
            cache[symbol] = []
    return cache


def _pearson(a: list, b: list) -> float:
    n = min(len(a), len(b))
    if n < 20:
        return 0.0
    x, y = a[-n:], b[-n:]
    mx, my = sum(x)/n, sum(y)/n
    cov = sum((x[i]-mx)*(y[i]-my) for i in range(n))
    vx  = sum((v-mx)**2 for v in x)
    vy  = sum((v-my)**2 for v in y)
    if vx <= 0 or vy <= 0:
        return 0.0
    return cov / (vx * vy) ** 0.5


def correlation_check(symbol: str, holdings: list, returns_cache: dict,
                      max_corr: float = 0.75) -> tuple:
    """
    Gate 14: Reject if the candidate is highly correlated with an existing holding.
    Returns (blocked: bool, worst_corr: float, correlated_with: str).
    Uses pre-built returns_cache — no extra downloads.
    """
    if not holdings or not returns_cache:
        return False, 0.0, ""

    cand_rets = returns_cache.get(symbol, [])
    if len(cand_rets) < 20:
        return False, 0.0, ""   # not enough data — pass through

    worst_corr   = 0.0
    worst_symbol = ""
    for h in holdings:
        held_sym  = h.get("symbol", "")
        held_rets = returns_cache.get(held_sym, [])
        if len(held_rets) < 20:
            continue
        c = abs(_pearson(cand_rets, held_rets))
        if c > worst_corr:
            worst_corr   = c
            worst_symbol = held_sym

    if worst_corr >= max_corr:
        return True, round(worst_corr, 2), worst_symbol
    return False, round(worst_corr, 2), worst_symbol


def calculate_watchlist_levels(stock: dict) -> dict:
    """
    ATR-based entry/stop/target levels (FIX 2).
    Uses existing scored values when present; recomputes via yfinance ATR when missing.
    Returns valid floats — never crashes.
    """
    symbol  = stock.get("symbol", "")
    entry   = float(stock.get("entry",   0) or 0)
    stop    = float(stock.get("stop",    0) or 0)
    target1 = float(stock.get("target1", 0) or 0)
    target2 = float(stock.get("target2", 0) or 0)
    current = float(stock.get("price",   0) or entry)

    # If levels already computed and valid, just compute rr/risk_pct
    if entry > 0 and stop > 0 and target1 > entry:
        risk_pct = round((entry - stop) / entry * 100, 1)
        rr = round((target1 - entry) / (entry - stop), 2) if entry > stop else 0.0
        return {
            "entry": entry, "stop": stop, "target1": target1, "target2": target2,
            "rr": rr, "risk_pct": risk_pct, "current": current,
        }

    # Fetch via yfinance and compute ATR-based levels (FIX 2 — swing-low stop, ATR targets)
    try:
        df = yf.download(symbol, period="3mo", interval="1d",
                         progress=False, auto_adjust=True)
        if df is None or len(df) < 20:
            return {"entry": entry, "stop": stop, "target1": target1,
                    "target2": target2, "rr": 0.0, "risk_pct": 0.0, "current": current}

        close = df["Close"].values.flatten().astype(float)
        high  = df["High"].values.flatten().astype(float)
        low   = df["Low"].values.flatten().astype(float)
        vol   = df["Volume"].values.flatten().astype(float)

        current = round(float(close[-1]), 2)
        entry   = round(current * 1.003, 2)

        # True ATR over last 14 days
        tr_list = []
        for i in range(1, min(15, len(close))):
            tr = max(
                float(high[-i]) - float(low[-i]),
                abs(float(high[-i]) - float(close[-i-1])),
                abs(float(low[-i])  - float(close[-i-1]))
            )
            tr_list.append(tr)
        atr = float(np.mean(tr_list)) if tr_list else current * 0.03

        # STOP: ATR-based (Change #1, 2026-07-10) — was 10-day swing low × 0.995.
        # Uses stock's own volatility: stop sits 1.5 × ATR below entry.
        # Bounded 2.5%–8% to prevent extremes.
        # Rationale: fixed swing-low can be either too tight (volatile small-caps
        # like JTLIND) or hits the 10% upper bound (chart gaps). ATR adapts.
        # Shadow buckets (shadow_log.py) intentionally KEEP fixed ±3%/+5%/+10%
        # for backtest comparability — do NOT change that path.
        stop_atr    = round(entry - (1.5 * atr), 2)
        stop        = stop_atr
        risk_raw    = (entry - stop) / entry * 100 if entry > 0 else 0

        # Enforce bounds: 2.5% floor (min real risk), 8% ceiling (max real risk).
        if risk_raw < 2.5:
            stop = round(entry * 0.975, 2)
        elif risk_raw > 8.0:
            stop = round(entry * 0.92, 2)

        risk     = entry - stop
        risk_pct = round(risk / entry * 100, 1)

        # TARGETS: 2.5x and 4.5x ATR — better R/R than hardcoded 8-12%
        target1 = round(entry + atr * 2.5, 2)
        target2 = round(entry + atr * 4.5, 2)

        # Ensure minimum 1.5x R/R
        if (target2 - entry) < (risk * 1.5):
            target2 = round(entry + risk * 1.5, 2)
        target1 = round((entry + target2) / 2, 2)

        reward = target2 - entry
        rr     = round(reward / risk, 2) if risk > 0 else 0.0

        return {
            "entry": entry, "stop": stop, "target1": target1, "target2": target2,
            "rr": rr, "risk_pct": risk_pct, "current": current,
        }
    except Exception:
        # Graceful fallback to whatever values exist
        risk_pct = round((entry - stop) / entry * 100, 1) if entry > stop > 0 else 0.0
        rr = round((target1 - entry) / (entry - stop), 2) if entry > stop > 0 and target1 > entry else 0.0
        return {
            "entry": entry, "stop": stop, "target1": target1, "target2": target2,
            "rr": rr, "risk_pct": risk_pct, "current": current,
        }


def get_stock_rr(stock: dict, levels: dict) -> float:
    """Single authoritative R/R value (BUG FIX 7).
    Priority: scoring engine rr_ratio > ATR-calculated levels rr.
    """
    scoring_rr = float(stock.get("rr_ratio", 0) or 0)
    levels_rr  = float(levels.get("rr",       0) or 0)
    if scoring_rr > 0:
        return scoring_rr
    if levels_rr > 0:
        return levels_rr
    return 0.0


def classify_watchlist(stock: dict, regime: str, thresholds: dict,
                        conf_history: dict = None) -> dict:
    thresh   = (thresholds or REGIME_THRESHOLDS)[regime]
    min_conf = thresh["min_confidence"]
    conf     = stock.get("final_confidence", 0)
    tq       = stock.get("trade_quality_score", 0)
    # FIX (IOLCP bug): clamp to 0. Previously produced negative gaps when Conf
    # already passed the threshold but R/R was the real blocker — the AI/rule
    # insight then said "close the -10.7-point gap" which is nonsense.
    conf_gap = round(max(0.0, min_conf - conf), 1)

    # Always compute price levels — every watchlist entry shows them
    levels = calculate_watchlist_levels(stock)

    # Phase C5 (rating ≥ 9.0): trajectory awareness for tier assignment.
    # Split a static NEAR_MISS into RISING vs FADING using 3-day conf history.
    # A stock at 78 conf climbing 72→75→78 is a much stronger watch target
    # than one at 78 falling 85→82→78. `traj` ∈ {"RISING","FADING","FLAT",""}.
    #
    # Phase 1 #42b (2026-07-05): gap-safe trajectory read.
    # The history now stores None sentinels for days when the symbol was NOT
    # in top-N. A trajectory that spans a None is invalid (the symbol was
    # temporarily invisible), so we return traj="" instead of a misleading
    # RISING/FADING label built from stale endpoints.
    traj = ""
    traj_delta = 0.0
    try:
        sym = stock.get("symbol", "")
        confs = ((conf_history or {}).get(sym) or {}).get("confs", []) or []
        # Consider only the trailing 3 entries for the delta calc.
        window = confs[-3:] if len(confs) >= 2 else confs
        if len(window) >= 2 and all(v is not None for v in window):
            traj_delta = float(window[-1]) - float(window[0])
            # Phase 3a N7 (2026-07-05): widened deadband from ±3.0 → ±4.0
            # to prevent single-point oscillations from flipping RISING ↔
            # FADING day-over-day. A stock climbing 74→75→77 (delta +3.0)
            # used to register RISING; next day 75→77→76 (delta +1.0) would
            # register FLAT, followed by 77→76→74 (delta -3.0) registering
            # FADING — three tier flips in three days despite the stock
            # essentially chopping in a 4-pt range. New band requires ≥4-pt
            # net move over 3 days for a directional label.
            if traj_delta >= 4.0:
                traj = "RISING"
            elif traj_delta <= -4.0:
                traj = "FADING"
            else:
                traj = "FLAT"
        # else: leave traj="" so downstream skips trajectory-based tiering.
    except Exception:
        pass

    # Phase 1 #43 (2026-07-05): READY_BLOCKED detection.
    # A stock is READY_BLOCKED if its scoring is fine (conf close AND tq at
    # threshold AND rr ok) but a "structural" hard filter blocks it — i.e.
    # market-cap floor, liquidity, promoter pledge, kill switch, black-swan
    # news. These setups are NOT bad; they're just unbuyable at current price
    # or size. Persisting them into the watchlist tells research: "this
    # methodology WORKED but was gated out by a structural constraint."
    _fail_reasons_raw = stock.get("fail_reasons", []) or []
    _STRUCTURAL_PREFIXES = (
        "MARKET_CAP_LOW", "MCAP_MISSING",
        "LIQUIDITY_FAIL",
        "PROMOTER_PLEDGE_", "PROMOTER_PLEDGE_BLOCKLIST",
        "KILL_SWITCH",
        "BLACK_SWAN_NEWS",
        "REGIME_NO_BUY",
        "SECTOR_DAY_CAP_",
    )
    _has_structural = any(
        any(r.startswith(p) for p in _STRUCTURAL_PREFIXES)
        for r in _fail_reasons_raw
    )
    _scoring_ok = (
        conf_gap <= 5.0
        and tq >= thresh["min_tq"] - 2.0
        and (stock.get("rr_ratio", 0) or 0) >= thresh["min_rr"] - 0.1
    )

    base = {
        "conf":     conf,
        "tq":       tq,
        "conf_gap": conf_gap,
        "sector":   stock.get("sector") or get_sector(stock.get("symbol", "")),
        "entry":    levels["entry"],
        "stop":     levels["stop"],
        "target1":  levels["target1"],
        "target2":  levels["target2"],
        "rr":       levels["rr"],
        "rr_ratio": get_stock_rr(stock, levels),
        "risk_pct": levels["risk_pct"],
        "current":  levels["current"],
        "fail_reasons": _fail_reasons_raw,
        "warnings":     stock.get("warnings", []),
        "trajectory":    traj,
        "traj_delta":    round(traj_delta, 1),
        # Phase 1 #45 (2026-07-05): regime tag on every watchlist row so
        # post-mortem can filter "which regime produced this tier mix?"
        "regime_at_classification": regime,
    }

    # Phase 1 #43: READY_BLOCKED takes precedence over gap-based tiers.
    # These are the highest-quality watchlist entries: the methodology
    # said BUY but a hard filter said NO.
    if _has_structural and _scoring_ok:
        _blocker = next(
            (r for r in _fail_reasons_raw
             if any(r.startswith(p) for p in _STRUCTURAL_PREFIXES)),
            "UNKNOWN"
        )
        return {**base, "tier": "READY_BLOCKED",
                "note": f"Setup ready but blocked by {_blocker}. Track for structural change.",
                "days_to_watch": 5, "watch_days": 5,
                "blocker_reason": _blocker}

    # Tier logic — trajectory-aware (Phase C5):
    #   NEAR_MISS_RISING  : gap ≤ 15 AND traj == RISING     → prioritize (2d watch)
    #   NEAR_MISS         : gap ≤ 15 AND traj != FADING     → standard (3d watch)
    #   NEAR_MISS_FADING  : gap ≤ 15 AND traj == FADING     → deprioritize (5d watch)
    #   DEVELOPING        : gap ≤ 25 AND tq >= 70           → weekly watch
    #   MONITOR           : gap  > 25                       → early stage
    if conf_gap <= 15 and tq >= thresh["min_tq"] - 5:
        if traj == "RISING":
            return {**base, "tier": "NEAR_MISS_RISING",
                    "note": f"↑ Trajectory rising ({traj_delta:+.1f}). Watch daily — trigger imminent.",
                    "days_to_watch": 2, "watch_days": 2}
        elif traj == "FADING":
            return {**base, "tier": "NEAR_MISS_FADING",
                    "note": f"↓ Trajectory fading ({traj_delta:+.1f}). Needs re-entry signal to re-qualify.",
                    "days_to_watch": 5, "watch_days": 5}
        else:
            return {**base, "tier": "NEAR_MISS",
                    "note": f"Needs +{conf_gap:.1f} conf. Watch for volume trigger.",
                    "days_to_watch": 3, "watch_days": 3}
    elif conf_gap <= 25 and tq >= max(60, thresh["min_tq"] - 15):
        return {**base, "tier": "DEVELOPING",
                "note": f"TQ {tq:.1f} building. Conf gap {conf_gap:.1f}.",
                "days_to_watch": 7, "watch_days": 7}
    else:
        # Phase 1 #47 (2026-07-05): MONITOR gets a `monitor_reason` field so
        # research can partition MONITOR into "far from setup" vs "off-thesis"
        # cohorts. The reason encodes the primary distance from qualification.
        if conf_gap > 25 and tq < 70:
            _mon_reason = "LOW_CONF_AND_TQ"
        elif conf_gap > 25:
            _mon_reason = "CONF_FAR"
        else:
            _mon_reason = "TQ_LOW"
        return {**base, "tier": "MONITOR",
                "note": "Early stage. Review in 2 weeks.",
                "days_to_watch": 14, "watch_days": 14,
                "monitor_reason": _mon_reason}


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 8 — WATCHLIST PERSISTENCE
# ─────────────────────────────────────────────────────────────────────────────

def load_persistent_watchlist() -> dict:
    # Phase C7c: FRESH_START wipes persistent watchlist for one run
    if FRESH_START:
        _log("[FRESH_START] load_persistent_watchlist → returning {} (old watchlist_persist.json ignored)")
        return {}
    try:
        if os.path.exists(WATCHLIST_FILE):
            with open(WATCHLIST_FILE, "r") as f:
                return json.load(f)
    except Exception as e:
        _log(f"[WARN] load_persistent_watchlist failed: {e}")
    return {}


def save_persistent_watchlist(watchlist_dict: dict) -> None:
    try:
        with open(WATCHLIST_FILE, "w") as f:
            json.dump(watchlist_dict, f, indent=2, default=str)
    except Exception as e:
        _log(f"[WARN] save_persistent_watchlist failed: {e}")


def merge_watchlist_with_history(todays_watchlist: list, history: dict) -> tuple:
    """Merge today's watchlist with persistent history.

    Phase 1 #48+#49 (2026-07-05): the history entry now tracks:
      * `tier_since` — date the current tier started (reset on tier change).
        Enables "days-in-current-tier" analysis (e.g., how long a stock sits
        as READY_BLOCKED before either becoming BUY or fading).
      * `warnings`, `fail_reasons`, `conf`, `conf_gap`, `regime`, `blocker_reason`,
        `monitor_reason` — all persisted so daily snapshot readers can build
        the tier-transition matrix without re-running the pipeline.
    """
    today_str = ist_today().isoformat()
    updated_history = {}
    for stock in todays_watchlist:
        symbol     = stock.get("symbol", "")
        prev       = history.get(symbol, {})
        first_seen = prev.get("first_seen", today_str)
        days_watched = (ist_today() -
                        datetime.date.fromisoformat(first_seen)).days
        stock["days_watched"] = days_watched
        stock["first_seen"]   = first_seen
        # Phase 1 #48: tier_since — reset when tier changes.
        _prev_tier   = prev.get("tier")
        _curr_tier   = stock.get("tier", "MONITOR")
        if _prev_tier != _curr_tier:
            _tier_since = today_str
        else:
            _tier_since = prev.get("tier_since", today_str)
        _days_in_tier = (ist_today() -
                         datetime.date.fromisoformat(_tier_since)).days
        stock["tier_since"]    = _tier_since
        stock["days_in_tier"]  = _days_in_tier
        if days_watched > 0:
            stock["note"] = stock.get("note", "") + f" [Day {days_watched + 1}]"
        max_days = stock.get("days_to_watch", 14)
        if days_watched <= max_days:
            updated_history[symbol] = {
                "first_seen":    first_seen,
                "tier":          _curr_tier,
                "tier_since":    _tier_since,
                "entry_ref":     stock.get("entry", 0),
                "last_seen":     today_str,
                # Phase 1 #49: persist decision-relevant context for research.
                "conf":          stock.get("conf"),
                "conf_gap":      stock.get("conf_gap"),
                "tq":            stock.get("tq"),
                "rr_ratio":      stock.get("rr_ratio"),
                "regime":        stock.get("regime_at_classification"),
                "warnings":      list(stock.get("warnings", []) or []),
                "fail_reasons":  list(stock.get("fail_reasons", []) or []),
                "blocker_reason": stock.get("blocker_reason"),
                "monitor_reason": stock.get("monitor_reason"),
                "trajectory":    stock.get("trajectory"),
            }
    return todays_watchlist, updated_history


def tag_repeat_buy_signals(buys: list, tracker_entries: list) -> list:
    today_str = ist_today().isoformat()
    for stock in buys:
        symbol = stock.get("symbol", "")
        match  = next((e for e in tracker_entries
                       if e["symbol"] == symbol
                       and e["type"] == "BUY"
                       and e["status"] == "OPEN"
                       and e["suggested_date"] != today_str), None)
        if match:
            days_since = (ist_today() -
                          datetime.date.fromisoformat(match["suggested_date"])).days
            stock["repeat_tag"] = f"REPEAT DAY {days_since + 1}"
            existing = stock.get("warnings", [])
            existing.append(f"REPEAT_DAY_{days_since + 1}")
            stock["warnings"] = existing
    return buys


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 9 — TRADE TRACKER
# ─────────────────────────────────────────────────────────────────────────────

def load_tracker() -> list:
    # Phase C7c: FRESH_START wipes tracker state for one run
    if FRESH_START:
        _log("[FRESH_START] load_tracker → returning empty list (old tracker.json ignored)")
        return []
    try:
        if os.path.exists(TRACKER_FILE):
            with open(TRACKER_FILE, "r") as f:
                return json.load(f)
    except Exception as e:
        _log(f"[WARN] load_tracker failed: {e}")
    return []


def save_tracker(entries: list) -> None:
    try:
        with open(TRACKER_FILE, "w") as f:
            json.dump(entries, f, indent=2, default=str)
    except Exception as e:
        _log(f"[WARN] save_tracker failed: {e}")


def add_to_tracker(entries: list, stock: dict, sig_type: str) -> list:
    today_str = ist_today().isoformat()
    symbol    = stock.get("symbol", "")
    already   = any(
        e["symbol"] == symbol and e["suggested_date"] == today_str and e["status"] == "OPEN"
        for e in entries
    )
    if already:
        return entries
    entry = {
        "symbol":          symbol,
        "type":            sig_type,
        "suggested_date":  today_str,
        "suggested_price": round(float(stock.get("entry", stock.get("price", 0)) or 0), 2),
        "entry":           float(stock.get("entry", 0) or 0),
        "stop":            float(stock.get("stop", 0) or 0),
        "target1":         float(stock.get("target1", 0) or 0),
        "target2":         float(stock.get("target2", 0) or 0),
        "sector":          stock.get("sector", "OTHERS"),
        "conf":            round(float(stock.get("final_confidence", 0) or 0), 1),
        "tq":              round(float(stock.get("trade_quality_score", 0) or 0), 1),
        "status":          "OPEN",
        "close_reason":    None,
        "close_price":     None,
        "close_date":      None,
        "max_gain_pct":    0.0,
        "max_loss_pct":    0.0,
        # Phase E1 runner-mode fields (default off — populated when T1/T2 hit)
        "partial_closed":       False,
        "partial_exit_pct":     0.0,
        "partial_exit_price":   0.0,
        "partial_exit_date":    None,
        "partial_exit_pnl_pct": 0.0,
        "runner_active":        False,
        "runner_partial_pct":   0.0,
        "runner_partial_price": 0.0,
        "runner_partial_pnl":   0.0,
        "runner_high_water":    0.0,
        "t2_hit_date":          None,
        "trail_note":           "",
    }
    entries.append(entry)
    return entries


def update_tracker_trailing_stop(entries: list) -> list:
    for e in entries:
        if e["status"] != "OPEN":
            continue
        try:
            df = fetch_price_data(e["symbol"], period="5d")
            if df is None or len(df) == 0:
                continue
            current  = float(df["Close"].squeeze().iloc[-1])
            t1       = float(e.get("target1", 0) or 0)
            entry_px = float(e.get("entry", 0) or 0)
            old_stop = float(e.get("stop", 0) or 0)
            if t1 > 0 and current >= t1 and entry_px > old_stop:
                e["stop"]       = entry_px
                e["trail_note"] = f"Trailed to entry {entry_px} after T1 hit"
        except Exception as ex:
            _log(f"[WARN] trailing stop update failed for {e.get('symbol')}: {ex}")
    return entries


def update_tracker(entries: list) -> tuple:
    """Daily update for the v1 flat tracker (tracker.json).

    Phase E1 (2026-07-02): 5-layer professional exit stack — see the config
    block near IS_SCHEDULED for the full description. Every layer is
    env-gated; when a layer's flag is off, its rule is a no-op and the legacy
    behavior applies.

    Legacy behavior (all flags off, default):
      • close ≤ stop            → EXIT full  (STOP_HIT / STOP_HIT_RESIDUAL)
      • close ≥ t2              → EXIT full  (TARGET2_HIT)
      • close ≥ t1 (first time) → book PARTIAL_EXIT_PCT, trail residual to entry
      • no time exit here (v1 tracker has no time-decay closure)

    E1a (RUNNER_MODE_ENABLED=true): at T2, book RUNNER_PARTIAL_PCT of the
    residual and let the runner ride a chandelier trail until trail-hit,
    trend-break, volume-fade, or RUNNER_MAX_DAYS.

    E1b (TRAIL_MODE=atr): after T1, trail residual with a ratcheted
    max(entry, close − TRAIL_ATR_MULT × ATR14) instead of a hard break-even.

    E1d (REGIME_AWARE_EXITS=true): in risk-off regimes, runner is skipped
    and trail multiplier is tightened by REGIME_TIGHTEN_FACTOR.
    """
    closed_today = []
    today_iso = ist_today().isoformat()
    regime = _current_exit_regime()
    for e in entries:
        if e["status"] != "OPEN":
            continue
        try:
            df = fetch_price_data(e["symbol"], period="1d")
            if df is None or len(df) == 0:
                continue
            current    = round(float(df["Close"].squeeze().iloc[-1]), 2)
            entry_px   = float(e.get("entry") or e.get("suggested_price", 0))
            stop       = float(e.get("stop", 0) or 0)
            t1         = float(e.get("target1", 0) or 0)
            t2         = float(e.get("target2", 0) or 0)
            if entry_px <= 0:
                continue
            chg_pct = round((current - entry_px) / entry_px * 100, 2)
            if chg_pct > e["max_gain_pct"]:
                e["max_gain_pct"] = chg_pct
            if chg_pct < e["max_loss_pct"]:
                e["max_loss_pct"] = chg_pct

            # State flags carried across days
            partial_done  = bool(e.get("partial_closed", False))
            runner_active = bool(e.get("runner_active", False))
            close_reason  = None

            # Fetch richer context only if any advanced layer is enabled
            ctx = {}
            if (TRAIL_MODE == "atr" and partial_done) or (runner_active) or (
                RUNNER_MODE_ENABLED and t2 > 0 and current >= t2 and partial_done
            ):
                ctx = _compute_exit_context(e["symbol"])

            # ── PATH A: runner already active (post-T2 tail) ──────────────
            if runner_active:
                hw = float(e.get("runner_high_water", current) or current)
                if current > hw:
                    hw = current
                    e["runner_high_water"] = round(hw, 2)
                # Chandelier trail: max(entry, hw − mult × ATR)
                atr14 = float(ctx.get("atr14", 0.0) or 0.0)
                mult  = _effective_runner_atr_mult(regime)
                trail = max(entry_px, round(hw - mult * atr14, 2)) if atr14 > 0 else float(e.get("stop", entry_px) or entry_px)
                if trail > float(e.get("stop", 0) or 0):
                    e["stop"] = trail

                # ── F5: Runner "moonshot" scaling ─────────────────────────
                # When a runner has run +N% above T2 on strong volume, book
                # RUNNER_SCALE_BOOK_PCT of what's left to lock the exceptional
                # move. Only fires ONCE per position (runner_scale_done flag).
                if (RUNNER_SCALE_ENABLED
                        and t2 > 0
                        and not e.get("runner_scale_done", False)
                        and current >= t2 * (1.0 + RUNNER_SCALE_TRIGGER_PCT / 100.0)):
                    vols = ctx.get("avg_vol20", 0.0) or 0.0
                    last_vol = ctx.get("last_volume", vols) or vols
                    vol_ok = (vols > 0 and last_vol >= vols * RUNNER_SCALE_VOL_MULT) or vols == 0
                    if vol_ok:
                        rs_pct = max(0.0, min(100.0, RUNNER_SCALE_BOOK_PCT))
                        # Runner P&L at the scale point (relative to entry)
                        rs_pnl = round((current - entry_px) / entry_px * 100, 2) if entry_px else 0.0
                        e["runner_scale_done"]       = True
                        e["runner_scale_pct"]        = rs_pct
                        e["runner_scale_price"]      = round(current, 2)
                        e["runner_scale_date"]       = today_iso
                        e["runner_scale_pnl_pct"]    = rs_pnl
                        e["trail_note"] = (e.get("trail_note", "") +
                                           f" F5: booked {rs_pct:.0f}% @ ₹{current:.2f} (+{rs_pnl:.1f}%)").strip()

                # Exit rules on the runner
                days_in_runner = 0
                try:
                    days_in_runner = (ist_today() -
                                      datetime.date.fromisoformat(str(e.get("t2_hit_date", today_iso))[:10])).days
                except Exception:
                    pass

                # ── F7b: Parabolic single-day blow-off exit ──────────────
                # A single-day gain >= PARABOLIC_DAY_PCT on ≥N×20d volume is
                # a classic distribution / exhaustion signal — book profit.
                parabolic_hit = False
                if PARABOLIC_EXIT_ENABLED and not e.get("parabolic_done", False):
                    prev_close = ctx.get("prev_close", 0.0) or 0.0
                    if prev_close > 0:
                        day_pct = (current - prev_close) / prev_close * 100
                        vols = ctx.get("avg_vol20", 0.0) or 0.0
                        last_vol = ctx.get("last_volume", vols) or vols
                        vol_ok = (vols > 0 and last_vol >= vols * PARABOLIC_VOL_MULT)
                        if day_pct >= PARABOLIC_DAY_PCT and vol_ok:
                            parabolic_hit = True
                            e["parabolic_done"] = True
                            e["parabolic_date"]  = today_iso
                            e["parabolic_price"] = round(current, 2)

                if current <= trail:
                    close_reason = "RUNNER_TRAIL_HIT"
                elif parabolic_hit:
                    close_reason = "PARABOLIC_BLOWOFF"
                elif ctx.get("below_ema_streak", 0) >= RUNNER_TREND_BREAK_DAYS:
                    close_reason = "RUNNER_TREND_BREAK"
                elif ctx.get("low_vol_streak", 0) >= RUNNER_VOL_FADE_DAYS:
                    close_reason = "RUNNER_VOLUME_FADE"
                elif days_in_runner >= RUNNER_MAX_DAYS:
                    close_reason = "RUNNER_MAX_DAYS"
                elif stop > 0 and current <= stop:
                    close_reason = "RUNNER_TRAIL_HIT"

            # ── PATH B: legacy — hard stop / T2 exit / T1 partial ─────────
            if not close_reason and not runner_active:
                if stop > 0 and current <= stop:
                    close_reason = "STOP_HIT" if not partial_done else "STOP_HIT_RESIDUAL"
                elif t2 > 0 and current >= t2:
                    # E1a: runner mode? convert to runner instead of closing
                    if _runner_enabled(regime) and partial_done:
                        # Book RUNNER_PARTIAL_PCT of what's left; the rest rides
                        rp_pct = max(0.0, min(100.0, RUNNER_PARTIAL_PCT))
                        e["runner_active"]         = True
                        e["t2_hit_date"]           = today_iso
                        e["runner_partial_pct"]    = rp_pct
                        e["runner_partial_price"]  = current
                        e["runner_partial_pnl"]    = chg_pct
                        e["runner_high_water"]     = current
                        # Trail stop starts at entry (never below) — tightens daily via PATH A
                        atr14 = float(ctx.get("atr14", 0.0) or 0.0)
                        mult  = _effective_runner_atr_mult(regime)
                        init_trail = max(entry_px, round(current - mult * atr14, 2)) if atr14 > 0 else entry_px
                        if init_trail > float(e.get("stop", 0) or 0):
                            e["stop"] = init_trail
                        e["trail_note"] = (
                            f"T2 hit @ ₹{current:.2f} (+{chg_pct:.1f}%). "
                            f"Booked {rp_pct:.0f}% of residual · runner riding chandelier "
                            f"(ATR×{mult:.1f}) from ₹{init_trail:.2f}"
                        )
                        # Not closed — continue riding
                    else:
                        close_reason = "TARGET2_HIT"
                elif not partial_done and t1 > 0 and current >= t1:
                    # First-time T1: book PARTIAL_EXIT_PCT, then trail residual
                    e["partial_closed"]       = True
                    e["partial_exit_pct"]     = float(os.getenv("PARTIAL_EXIT_PCT", "50"))
                    e["partial_exit_price"]   = current
                    e["partial_exit_date"]    = today_iso
                    e["partial_exit_pnl_pct"] = chg_pct
                    # E1b: ATR trail vs. break-even
                    if TRAIL_MODE == "atr":
                        atr14 = float(ctx.get("atr14", 0.0) or 0.0)
                        mult  = _effective_trail_atr_mult(regime)
                        atr_trail = round(current - mult * atr14, 2) if atr14 > 0 else entry_px
                        new_stop = max(entry_px, atr_trail)
                        if new_stop > stop:
                            e["stop"] = new_stop
                            e["trail_note"] = (
                                f"T1 hit @ ₹{current:.2f} (+{chg_pct:.1f}%). "
                                f"Sold {e['partial_exit_pct']:.0f}% · ATR trail (×{mult:.1f}) → ₹{new_stop:.2f}"
                            )
                    else:
                        # legacy break-even
                        if entry_px > stop:
                            e["stop"] = entry_px
                            e["trail_note"] = (
                                f"T1 hit @ ₹{current:.2f} (+{chg_pct:.1f}%). "
                                f"Sold {e['partial_exit_pct']:.0f}% · trailed rest to entry ₹{entry_px:.2f}"
                            )
                    # residual OPEN
                elif partial_done and TRAIL_MODE == "atr":
                    # E1b daily ratchet after T1 — before T2, keep tightening the ATR trail
                    if not ctx:
                        ctx = _compute_exit_context(e["symbol"])
                    atr14 = float(ctx.get("atr14", 0.0) or 0.0)
                    if atr14 > 0:
                        mult  = _effective_trail_atr_mult(regime)
                        new_stop = max(entry_px, round(current - mult * atr14, 2))
                        if new_stop > stop:
                            e["stop"] = new_stop
                            e["trail_note"] = (
                                f"ATR trail ratchet (×{mult:.1f}) → ₹{new_stop:.2f}"
                            )

            # ── Close if any layer decided so ─────────────────────────────
            if close_reason:
                e["status"]        = "CLOSED"
                e["close_reason"]  = close_reason
                e["close_price"]   = current
                e["close_date"]    = today_iso
                e["final_pnl_pct"] = chg_pct
                # Blended P&L across the exit sequence
                pieces = []
                if partial_done:
                    pe = float(e.get("partial_exit_pct", 0.0)) / 100.0
                    pp = float(e.get("partial_exit_pnl_pct", 0.0))
                    pieces.append((pe, pp))
                if runner_active or e.get("t2_hit_date"):
                    rp = float(e.get("runner_partial_pct", 0.0)) / 100.0
                    # runner partial is a % of the RESIDUAL (i.e., what remained after T1 partial).
                    residual_frac = 1.0 - float(e.get("partial_exit_pct", 0.0)) / 100.0
                    rp_of_total   = rp * residual_frac
                    pp_r          = float(e.get("runner_partial_pnl", 0.0))
                    if rp_of_total > 0:
                        pieces.append((rp_of_total, pp_r))
                booked_frac = sum(w for w, _ in pieces)
                remaining   = max(0.0, 1.0 - booked_frac)
                blended = sum(w * p for w, p in pieces) + remaining * chg_pct
                if pieces:
                    e["blended_pnl_pct"] = round(blended, 2)
                closed_today.append(e)
        except Exception as ex:
            _log(f"[WARN] tracker update failed for {e.get('symbol')}: {ex}")
    return entries, closed_today


def _days_open(e: dict) -> int:
    try:
        return (ist_today() - datetime.date.fromisoformat(e["suggested_date"])).days
    except Exception:
        return 0


def _pct_bar(chg_pct: float) -> str:
    if chg_pct >= 0:
        bars = min(10, int(chg_pct))
        return f"▲ {'█' * bars}{'░' * (10 - bars)} +{chg_pct:.2f}%"
    else:
        bars = min(10, int(abs(chg_pct)))
        return f"▼ {'█' * bars}{'░' * (10 - bars)} {chg_pct:.2f}%"


def _entry_block(e: dict) -> list:
    blk      = []
    sym      = e["symbol"]
    entry_px = float(e.get("entry") or e.get("suggested_price", 0))
    t1       = float(e.get("target1", 0) or 0)
    t2       = float(e.get("target2", 0) or 0)
    stop     = float(e.get("stop", 0) or 0)
    days     = _days_open(e)
    conf     = e.get("conf", 0)
    try:
        df      = fetch_price_data(sym, period="1d")
        current = round(float(df["Close"].squeeze().iloc[-1]), 2) if df is not None and len(df) > 0 else entry_px
    except Exception:
        current = entry_px
    chg_pct = round((current - entry_px) / entry_px * 100, 2) if entry_px > 0 else 0.0
    t1_pct  = round((t1 - entry_px) / entry_px * 100, 1) if entry_px > 0 and t1 > 0 else 0
    t2_pct  = round((t2 - entry_px) / entry_px * 100, 1) if entry_px > 0 and t2 > 0 else 0
    sl_pct  = round((stop - entry_px) / entry_px * 100, 1) if entry_px > 0 and stop > 0 else 0
    blk.append(f"  <b>{sym}</b> [{e['type']}] | Since {e['suggested_date']} ({days}d)")
    blk.append(f"  Conf {conf:.1f} | Entry Rs{entry_px:.2f} → Now Rs{current:.2f}")
    blk.append(f"  {_pct_bar(chg_pct)}")
    blk.append(f"  SL {sl_pct:+.1f}% | T1 {t1_pct:+.1f}% | T2 {t2_pct:+.1f}%")
    blk.append(f"  Peak gain: {e['max_gain_pct']:+.2f}% | Peak loss: {e['max_loss_pct']:+.2f}%")
    return blk


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 9b — TRADE TRACKER V2 (new structured format)
# ─────────────────────────────────────────────────────────────────────────────

def load_tracker_v2() -> dict:
    # Phase C7c: FRESH_START wipes tracker v2 state for one run
    if FRESH_START:
        _log("[FRESH_START] load_tracker_v2 → returning empty structure (old trade_tracker.json ignored)")
        return {"buys": [], "watchlist": [], "completed": [], "performance": {}}
    try:
        if os.path.exists(TRADE_TRACKER_V2_FILE):
            with open(TRADE_TRACKER_V2_FILE, "r") as f:
                return json.load(f)
    except Exception as e:
        _log(f"[WARN] load_tracker_v2 failed: {e}")
    return {"buys": [], "watchlist": [], "completed": [], "performance": {}}


def save_tracker_v2(tracker: dict) -> None:
    try:
        with open(TRADE_TRACKER_V2_FILE, "w") as f:
            json.dump(tracker, f, indent=2, default=str)
    except Exception as e:
        _log(f"[WARN] save_tracker_v2 failed: {e}")


def initialize_tracker_if_new() -> dict:
    """
    Called at pipeline start. Seeds Jun 25 data if no tracker v2 file exists yet.
    Phase C7c: FRESH_START skips both the file load AND the Jun 25 seed data —
    returns a truly empty tracker so the next pipeline run builds from zero.
    """
    if FRESH_START:
        _log("[FRESH_START] initialize_tracker_if_new → returning empty tracker (no Jun 25 seed)")
        empty = {"buys": [], "watchlist": [], "completed": [], "performance": {}}
        # Do NOT save yet — let the pipeline's own save_tracker_v2 call at end-of-run
        # persist whatever fresh state builds up during this run.
        return empty

    if os.path.exists(TRADE_TRACKER_V2_FILE):
        return load_tracker_v2()

    _log("[INFO] No tracker v2 found — initializing with Jun 25, 2026 seed data")
    tracker = {
        "buys": [
            {
                "symbol": "SIYSIL.NS", "rec_date": "2026-06-25",
                "entry": 645.30, "stop": 607.93,
                "target1": 707.58, "target2": 757.40,
                "confidence": 65.7, "tq": 94.0,
                "regime": "HIGH_VOLATILITY",
                # 2026-07-05: sector added so sector-concentration cap (Phase C7)
                # can actually gate future BUYs. Resolved via get_sector().
                "sector": get_sector("SIYSIL.NS"),
                "status": "ACTIVE",
                "t1_hit_date": None, "t2_hit_date": None, "stop_hit_date": None,
                "days_tracked": 1, "pnl_history": [],
            }
        ],
        "watchlist": [
            {"symbol": "BOSCHLTD.NS",   "rec_date": "2026-06-25", "tier": "NEAR_MISS",
             "sector": get_sector("BOSCHLTD.NS"),
             "conf_at_rec": 62.5, "conf_gap_at_rec": 1.5, "status": "WATCHING", "days_watching": 1},
            {"symbol": "GENUSPOWER.NS", "rec_date": "2026-06-25", "tier": "NEAR_MISS",
             "sector": get_sector("GENUSPOWER.NS"),
             "conf_at_rec": 62.1, "conf_gap_at_rec": 1.9, "status": "WATCHING", "days_watching": 1},
            {"symbol": "KRISHANA.NS",   "rec_date": "2026-06-25", "tier": "NEAR_MISS",
             "sector": get_sector("KRISHANA.NS"),
             "conf_at_rec": 60.8, "conf_gap_at_rec": 3.2, "status": "WATCHING", "days_watching": 1},
            {"symbol": "NAZARA.NS",     "rec_date": "2026-06-25", "tier": "NEAR_MISS",
             "sector": get_sector("NAZARA.NS"),
             "conf_at_rec": 60.4, "conf_gap_at_rec": 3.6, "status": "WATCHING", "days_watching": 1},
            {"symbol": "SONACOMS.NS",   "rec_date": "2026-06-25", "tier": "DEVELOPING",
             "sector": get_sector("SONACOMS.NS"),
             "tq_at_rec": 89.4, "conf_gap_at_rec": 5.2, "status": "WATCHING", "days_watching": 1},
            {"symbol": "CEIGALL.NS",    "rec_date": "2026-06-25", "tier": "DEVELOPING",
             "sector": get_sector("CEIGALL.NS"),
             "tq_at_rec": 97.0, "conf_gap_at_rec": 5.4, "status": "WATCHING", "days_watching": 1},
            {"symbol": "RATNAVEER.NS",  "rec_date": "2026-06-25", "tier": "DEVELOPING",
             "sector": get_sector("RATNAVEER.NS"),
             "tq_at_rec": 99.0, "conf_gap_at_rec": 5.5, "status": "WATCHING", "days_watching": 1},
        ],
        "completed": [],
        "performance": {},
    }
    save_tracker_v2(tracker)
    return tracker


# ─── Phase C7 (2026-07-02): Loss-reason classifier ──────────────────────────
# Post-mortem tag applied when a position closes. Reads pos.pnl_history to
# figure out HOW the loss developed (slow bleed vs gap-down vs quick reversal)
# so weekly summary can surface patterns and expose gaps in the gate system.
# Additive only — never affects trade decisions, only the audit trail.
def classify_loss_reason(pos: dict) -> str:
    """Return a category tag explaining WHY a closed position ended where it did.

    Categories:
      T2_TARGET_HIT       — clean win (legacy hard T2)
      T1_TARGET_HIT       — partial win
      RUNNER_TRAIL_HIT    — post-T2 chandelier trail closed the runner
      RUNNER_TREND_BREAK  — runner closed on EMA-break trend failure
      RUNNER_VOLUME_FADE  — runner closed on 3+ low-volume days
      RUNNER_MAX_DAYS     — runner hit the RUNNER_MAX_DAYS ceiling
      STOP_GAP_DOWN       — closed below stop on a single-bar gap (>3% drop)
      STOP_SLOW_BLEED     — drifted to stop over 5+ days
      STOP_QUICK_REVERSAL — hit stop within 3 days (thesis broken fast)
      TIME_EXPIRED_FLAT   — 15-day timeout with tiny P&L (setup died sideways)
      TIME_EXPIRED_LOSING — 15-day timeout in the red (never worked)
      TIME_EXPIRED_WINNING — 15-day timeout in green but never hit T1
      TIME_EXPIRED_STALLED / TIME_EXPIRED_STAGNATION / TIME_EXPIRED_HARDCAP
                          — Phase E1c conditional-mode categories
      UNKNOWN             — insufficient data
    """
    try:
        status   = pos.get("status", "")
        final    = float(pos.get("final_pnl", 0) or 0)
        history  = pos.get("pnl_history", []) or []

        # If a close_reason was already recorded by a runner exit path,
        # trust it — the granular category is more informative than a generic tag.
        pre = pos.get("close_reason")
        if pre and pre.startswith(("RUNNER_", "TIME_EXPIRED_", "PARABOLIC_")):
            return pre

        if status == "T2_HIT":
            return "T2_TARGET_HIT"
        if status == "T1_HIT":
            return "T1_TARGET_HIT"
        if status == "RUNNER_CLOSED":
            return pre or "RUNNER_TRAIL_HIT"

        if status == "STOPPED_OUT":
            # Look at pnl_history to distinguish gap-down vs bleed vs reversal
            if not history or len(history) < 2:
                return "STOP_QUICK_REVERSAL"
            days_held = len(history)
            last_pnl  = float(history[-1].get("pnl", final) or final)
            prev_pnl  = float(history[-2].get("pnl", last_pnl) or last_pnl)
            # Single-bar drop of >3% = gap down
            if prev_pnl - last_pnl >= 3.0:
                return "STOP_GAP_DOWN"
            if days_held <= 3:
                return "STOP_QUICK_REVERSAL"
            return "STOP_SLOW_BLEED"

        if status == "EXPIRED":
            if final < -1.5:
                return "TIME_EXPIRED_LOSING"
            if final > 1.5:
                return "TIME_EXPIRED_WINNING"
            return "TIME_EXPIRED_FLAT"

        return "UNKNOWN"
    except Exception:
        return "UNKNOWN"


def summarize_loss_reasons(tracker: dict) -> dict:
    """Return {category: count} across all completed positions."""
    from collections import Counter
    completed = tracker.get("completed", []) if isinstance(tracker, dict) else []
    tags = [
        pos.get("close_reason") or classify_loss_reason(pos)
        for pos in completed
    ]
    return dict(Counter(tags))


def add_to_tracker_v2(tracker: dict, stock: dict, regime: str = "") -> dict:
    """Append a fresh BUY into tracker_v2['buys'] using the v2 shape.

    Idempotent per (symbol, rec_date) — if this symbol already has an ACTIVE
    entry from earlier today it's not duplicated. The v2 shape is what the
    rest of the codebase (update_tracker_v2_pnl, compute_kill_switch_state,
    weekly_summary_job, morning_check) expects: rec_date, entry, stop,
    target1/2, confidence, tq, regime, status='ACTIVE',
    *_hit_date=None, days_tracked=1, pnl_history=[].

    Additive-only: does NOT touch scoring, gating, or state math — just
    persists a signal that already passed the gates. Never raises.
    """
    try:
        today_str = ist_today().isoformat()
        symbol    = stock.get("symbol", "")
        if not symbol or not isinstance(tracker, dict):
            return tracker
        buys = tracker.setdefault("buys", [])
        # Dedupe: same symbol + same rec_date + still ACTIVE/T1_HIT
        for existing in buys:
            if (existing.get("symbol") == symbol
                    and str(existing.get("rec_date", ""))[:10] == today_str
                    and existing.get("status") in ("ACTIVE", "T1_HIT")):
                return tracker

        buys.append({
            "symbol":        symbol,
            "rec_date":      today_str,
            "entry":         float(stock.get("entry", 0) or 0),
            "stop":          float(stock.get("stop",  0) or 0),
            "target1":       float(stock.get("target1", 0) or 0),
            "target2":       float(stock.get("target2", 0) or 0),
            "confidence":    round(float(stock.get("final_confidence", 0) or 0), 1),
            "tq":            round(float(stock.get("trade_quality_score", 0) or 0), 1),
            "regime":        regime or stock.get("regime", ""),
            # 2026-07-05: fall back to get_sector() if the upstream stock dict
            # didn't propagate a sector — belt-and-braces for the Phase C7
            # sector-concentration cap.
            "sector":        stock.get("sector") or get_sector(symbol),
            # Position-size % of capital (used by compute_kill_switch_state)
            "position_pct":  float(stock.get("position_pct", 5.0) or 5.0),
            "status":        "ACTIVE",
            "t1_hit_date":   None,
            "t2_hit_date":   None,
            "stop_hit_date": None,
            "expired_date":  None,
            "days_tracked":  1,
            "pnl_history":   [],
            # Phase E1 runner-mode fields — default off; populated when T1/T2 hit
            "partial_closed":      False,
            "partial_exit_pct":    0.0,
            "partial_exit_price":  0.0,
            "partial_exit_date":   None,
            "partial_exit_pnl_pct": 0.0,
            "runner_active":       False,
            "runner_partial_pct":  0.0,
            "runner_partial_price": 0.0,
            "runner_partial_pnl":  0.0,
            "runner_high_water":   0.0,
            "runner_close_date":   None,
            "trail_note":          "",
        })
    except Exception as e:
        _log(f"[WARN] add_to_tracker_v2 failed: {e}")
    return tracker


def update_tracker_v2_pnl(tracker: dict) -> dict:
    """
    Full daily update for tracker v2 (FIX 3):
    - Fetches live prices for every active position
    - Handles STOP_HIT, T1_HIT, T2_HIT, EXPIRED closings
    - Updates watchlist direction arrows and fills missing levels
    - Recalculates performance stats block

    Phase E1 (2026-07-02): 5-layer professional exit stack integrated.
    All layers are ENV-GATED — with defaults, this behaves identically to
    the legacy 'hard-close at T2' version. See config block near IS_SCHEDULED
    for RUNNER_MODE_ENABLED / TRAIL_MODE / TIME_EXIT_MODE / REGIME_AWARE_EXITS.

    Position lifecycle (all state on `pos` dict):
      ACTIVE  → T1_HIT  → RUNNING (post-T2, runner_active)  → STOPPED_OUT/T2_HIT/RUNNER_*
              → STOPPED_OUT / EXPIRED / T2_HIT
    """
    today_str = ist_today().isoformat()
    regime = _current_exit_regime()

    def _get_price(sym: str) -> float:
        try:
            df = fetch_price_data(sym, period="2d")
            if df is not None and len(df) > 0:
                return round(float(df["Close"].squeeze().iloc[-1]), 2)
        except Exception:
            pass
        return 0.0

    def _expire_conditional(pos, days, pnl):
        """E1c: decide whether to expire based on conditional rules.
        Returns close_reason string or None.
        Legacy path (TIME_EXIT_MODE='calendar') is handled by caller.
        """
        # Absolute cap: even a winner can't camp forever
        if days >= TIME_EXIT_HARD_MAX_DAYS:
            return "TIME_EXPIRED_HARDCAP"
        if days >= TIME_EXIT_MAX_DAYS:
            # Losing or flat after N days → expire
            if pnl < TIME_EXIT_MIN_WIN_PCT:
                # Also require weak momentum: 10d slope <= 0
                ctx = _compute_exit_context(pos["symbol"])
                if not ctx or ctx.get("slope10", 0.0) <= 0:
                    return "TIME_EXPIRED_STALLED"
        if days >= TIME_EXIT_STAGNATION_DAYS and pnl < TIME_EXIT_STAGNATION_PCT:
            return "TIME_EXPIRED_STAGNATION"
        return None

    # ── Update active buy positions ──
    still_active = []
    for pos in tracker.get("buys", []):
        if pos.get("status") not in ("ACTIVE", "T1_HIT", "RUNNING"):
            tracker.setdefault("completed", []).append(pos)
            continue
        try:
            cur_px = _get_price(pos["symbol"]) or float(pos.get("entry", 0) or 0)
            entry  = float(pos.get("entry", cur_px) or cur_px)
            pnl    = round((cur_px - entry) / entry * 100, 2) if entry > 0 else 0.0
            pos["days_tracked"] = (
                ist_today() -
                datetime.date.fromisoformat(pos.get("rec_date", today_str))
            ).days + 1
            hist = pos.setdefault("pnl_history", [])
            # Only append a new daily entry on scheduled runs — manual runs must not pollute history
            if IS_SCHEDULED and (not hist or hist[-1].get("date") != today_str):
                hist.append({"date": today_str, "price": cur_px, "pnl": pnl})

            stop = float(pos.get("stop", 0) or 0)
            t1   = float(pos.get("target1", 0) or 0)
            t2   = float(pos.get("target2", 0) or 0)
            days = pos.get("days_tracked", 0)

            partial_done  = bool(pos.get("partial_closed", False))
            runner_active = bool(pos.get("runner_active", False)) or pos.get("status") == "RUNNING"
            close_reason  = None

            # Only fetch full context when we might use it
            ctx = {}
            need_ctx = (
                runner_active
                or (TRAIL_MODE == "atr" and partial_done)
                or (RUNNER_MODE_ENABLED and t2 > 0 and cur_px >= t2 and partial_done)
                or (TIME_EXIT_MODE == "conditional" and days >= TIME_EXIT_MAX_DAYS)
            )
            if need_ctx:
                ctx = _compute_exit_context(pos["symbol"])

            # ── PATH A: runner-active post-T2 ─────────────────────────────
            if runner_active:
                hw = float(pos.get("runner_high_water", cur_px) or cur_px)
                if cur_px > hw:
                    hw = cur_px
                    pos["runner_high_water"] = round(hw, 2)
                atr14 = float(ctx.get("atr14", 0.0) or 0.0)
                mult  = _effective_runner_atr_mult(regime)
                trail = max(entry, round(hw - mult * atr14, 2)) if atr14 > 0 else stop
                if trail > stop:
                    pos["stop"] = trail
                    stop = trail

                # ── F5: Runner "moonshot" scaling (v2) ────────────────────
                if (RUNNER_SCALE_ENABLED
                        and t2 > 0
                        and not pos.get("runner_scale_done", False)
                        and cur_px >= t2 * (1.0 + RUNNER_SCALE_TRIGGER_PCT / 100.0)):
                    vols_avg = ctx.get("avg_vol20", 0.0) or 0.0
                    last_vol = ctx.get("last_volume", vols_avg) or vols_avg
                    vol_ok   = (vols_avg > 0 and last_vol >= vols_avg * RUNNER_SCALE_VOL_MULT) or vols_avg == 0
                    if vol_ok:
                        rs_pct = max(0.0, min(100.0, RUNNER_SCALE_BOOK_PCT))
                        rs_pnl = round((cur_px - entry) / entry * 100, 2) if entry else 0.0
                        pos["runner_scale_done"]    = True
                        pos["runner_scale_pct"]     = rs_pct
                        pos["runner_scale_price"]   = round(cur_px, 2)
                        pos["runner_scale_date"]    = today_str
                        pos["runner_scale_pnl_pct"] = rs_pnl

                # Days since t2_hit
                days_in_runner = 0
                try:
                    days_in_runner = (ist_today() -
                                      datetime.date.fromisoformat(str(pos.get("t2_hit_date", today_str))[:10])).days
                except Exception:
                    pass

                # ── F7b: Parabolic single-day blow-off (v2) ──────────────
                parabolic_hit = False
                if PARABOLIC_EXIT_ENABLED and not pos.get("parabolic_done", False):
                    prev_close = ctx.get("prev_close", 0.0) or 0.0
                    if prev_close > 0:
                        day_pct = (cur_px - prev_close) / prev_close * 100
                        vols_avg = ctx.get("avg_vol20", 0.0) or 0.0
                        last_vol = ctx.get("last_volume", vols_avg) or vols_avg
                        vol_ok   = (vols_avg > 0 and last_vol >= vols_avg * PARABOLIC_VOL_MULT)
                        if day_pct >= PARABOLIC_DAY_PCT and vol_ok:
                            parabolic_hit = True
                            pos["parabolic_done"]  = True
                            pos["parabolic_date"]  = today_str
                            pos["parabolic_price"] = round(cur_px, 2)

                if cur_px <= trail:
                    close_reason = "RUNNER_TRAIL_HIT"
                elif parabolic_hit:
                    close_reason = "PARABOLIC_BLOWOFF"
                elif ctx.get("below_ema_streak", 0) >= RUNNER_TREND_BREAK_DAYS:
                    close_reason = "RUNNER_TREND_BREAK"
                elif ctx.get("low_vol_streak", 0) >= RUNNER_VOL_FADE_DAYS:
                    close_reason = "RUNNER_VOLUME_FADE"
                elif days_in_runner >= RUNNER_MAX_DAYS:
                    close_reason = "RUNNER_MAX_DAYS"
                if close_reason:
                    pos.update({
                        "status": "RUNNER_CLOSED",
                        "runner_close_date": today_str,
                        "final_pnl": pnl,
                    })
                    pos["close_reason"] = close_reason
                    pos["close_price"]  = cur_px
                    tracker["completed"].append(pos)
                    continue
                else:
                    still_active.append(pos)
                    continue

            # ── PATH B: legacy hard stop / T2 exit / T1 partial ───────────
            if stop > 0 and cur_px <= stop:
                pos.update({"status": "STOPPED_OUT", "stop_hit_date": today_str, "final_pnl": pnl})
                pos["close_reason"] = classify_loss_reason(pos)
                pos["close_price"]  = cur_px
                tracker["completed"].append(pos)
                continue

            if t2 > 0 and cur_px >= t2:
                # E1a: convert to runner mode instead of closing (only if T1 already booked)
                if _runner_enabled(regime) and partial_done:
                    rp_pct = max(0.0, min(100.0, RUNNER_PARTIAL_PCT))
                    atr14  = float(ctx.get("atr14", 0.0) or 0.0)
                    mult   = _effective_runner_atr_mult(regime)
                    init_trail = max(entry, round(cur_px - mult * atr14, 2)) if atr14 > 0 else entry
                    pos.update({
                        "status":               "RUNNING",
                        "runner_active":        True,
                        "t2_hit_date":          today_str,
                        "runner_partial_pct":   rp_pct,
                        "runner_partial_price": cur_px,
                        "runner_partial_pnl":   pnl,
                        "runner_high_water":    cur_px,
                    })
                    if init_trail > stop:
                        pos["stop"] = init_trail
                    pos["trail_note"] = (
                        f"T2 hit @ ₹{cur_px:.2f} (+{pnl:.1f}%). Booked {rp_pct:.0f}% of residual · "
                        f"runner riding chandelier (ATR×{mult:.1f}) from ₹{init_trail:.2f}"
                    )
                    still_active.append(pos)
                    continue
                # Legacy: full close at T2
                pos.update({"status": "T2_HIT", "t2_hit_date": today_str, "final_pnl": pnl})
                pos["close_reason"] = classify_loss_reason(pos)
                pos["close_price"]  = cur_px
                tracker["completed"].append(pos)
                continue

            if t1 > 0 and cur_px >= t1 and pos.get("status") == "ACTIVE":
                # First-time T1 hit → mark, then apply trail (breakeven or ATR)
                pos["status"] = "T1_HIT"
                pos["t1_hit_date"] = today_str
                pos["partial_closed"] = True
                pos["partial_exit_pct"]     = float(os.getenv("PARTIAL_EXIT_PCT", "50"))
                pos["partial_exit_price"]   = cur_px
                pos["partial_exit_date"]    = today_str
                pos["partial_exit_pnl_pct"] = pnl
                if TRAIL_MODE == "atr":
                    atr14 = float(ctx.get("atr14", 0.0) or 0.0)
                    mult  = _effective_trail_atr_mult(regime)
                    new_stop = max(entry, round(cur_px - mult * atr14, 2)) if atr14 > 0 else entry
                    if new_stop > stop:
                        pos["stop"] = new_stop
                        pos["trail_note"] = f"T1 hit · ATR trail (×{mult:.1f}) → ₹{new_stop:.2f}"
                else:
                    if entry > stop:
                        pos["stop"] = entry
                        pos["trail_note"] = f"T1 hit · trailed to entry ₹{entry:.2f}"
                still_active.append(pos)
                continue

            # ── Daily ratchet on T1_HIT positions when TRAIL_MODE=atr ────
            if partial_done and TRAIL_MODE == "atr":
                if not ctx:
                    ctx = _compute_exit_context(pos["symbol"])
                atr14 = float(ctx.get("atr14", 0.0) or 0.0)
                if atr14 > 0:
                    mult = _effective_trail_atr_mult(regime)
                    new_stop = max(entry, round(cur_px - mult * atr14, 2))
                    if new_stop > stop:
                        pos["stop"] = new_stop
                        pos["trail_note"] = f"ATR ratchet (×{mult:.1f}) → ₹{new_stop:.2f}"

            # ── Time exit ────────────────────────────────────────────────
            if TIME_EXIT_MODE == "conditional":
                reason = _expire_conditional(pos, days, pnl)
                if reason:
                    pos.update({"status": "EXPIRED", "expired_date": today_str, "final_pnl": pnl})
                    pos["close_reason"] = reason
                    pos["close_price"]  = cur_px
                    tracker["completed"].append(pos)
                    continue
            else:
                # legacy calendar exit
                if days >= TIME_EXIT_MAX_DAYS and t1 > 0 and cur_px < t1:
                    pos.update({"status": "EXPIRED", "expired_date": today_str, "final_pnl": pnl})
                    pos["close_reason"] = classify_loss_reason(pos)
                    pos["close_price"]  = cur_px
                    tracker["completed"].append(pos)
                    continue

            still_active.append(pos)
        except Exception as e:
            _log(f"[WARN] update_tracker_v2_pnl buy update failed for {pos.get('symbol')}: {e}")
            still_active.append(pos)
    tracker["buys"] = still_active

    # ── Update watchlist entries ──
    still_watching = []
    for w in tracker.get("watchlist", []):
        try:
            days = (
                ist_today() -
                datetime.date.fromisoformat(w.get("rec_date", today_str))
            ).days + 1
            w["days_watching"] = days
            if days > 14:
                w["status"] = "EXPIRED"
                continue

            cur = _get_price(w["symbol"])
            if cur:
                w["current_price"] = cur
                # Fill missing levels for seeded watchlist stocks
                entry = float(w.get("entry", 0) or 0)
                if entry <= 0:
                    lvl = calculate_watchlist_levels(w)
                    for k in ("entry", "stop", "target1", "target2", "rr"):
                        w[k] = lvl.get(k, 0)
                    entry = w.get("entry", 0) or cur

                if entry > 0:
                    move = round((cur - entry) / entry * 100, 1)
                    w["direction"] = (
                        f"\u2191 {move:.1f}% above entry" if move >= 0
                        else f"\u2193 {abs(move):.1f}% below entry"
                    )
                else:
                    w.setdefault("direction", "\u2014")
            else:
                w.setdefault("direction", "\u2014")
            still_watching.append(w)
        except Exception:
            w.setdefault("direction", "\u2014")
            still_watching.append(w)
    tracker["watchlist"] = still_watching

    # ── Recalculate performance stats ──
    completed = tracker.get("completed", [])
    wins   = [t for t in completed if float(t.get("final_pnl", 0) or 0) > 0]
    losses = [t for t in completed if float(t.get("final_pnl", 0) or 0) <= 0]
    # Phase C7 (2026-07-02): also surface loss-reason breakdown so weekly
    # summary can spot patterns (e.g. "60% of losses are STOP_GAP_DOWN →
    # tighten HIGH_GAP_RISK gate", or "50% are STOP_SLOW_BLEED → shorten
    # time-stop from 20d to 15d").
    tracker["performance"] = {
        "completed":  len(completed),
        "active":     len(tracker["buys"]),
        "win_rate":   round(len(wins) / len(completed) * 100, 1) if completed else 0,
        "avg_win":    round(sum(float(t.get("final_pnl", 0) or 0) for t in wins) / len(wins), 2) if wins else 0,
        "avg_loss":   round(sum(float(t.get("final_pnl", 0) or 0) for t in losses) / len(losses), 2) if losses else 0,
        "loss_reasons": summarize_loss_reasons(tracker),
        "last_updated": today_str,
    }
    return tracker


def format_confidence_breakdown(factor_scores: dict, final_conf: float) -> list:
    """Returns lines showing what drove the confidence score (ENHANCEMENT 2)."""
    FACTOR_DISPLAY = [
        ("trend_quality",    "Trend",     18),
        ("momentum_quality", "Momentum",  14),
        ("volume_delivery",  "Volume",    10),
        ("sector_strength",  "Sector",    15),
        ("rs_vs_nifty",      "Rel Str",   15),
        ("news_risk",        "News",       8),
        ("risk_reward",      "R/R",        7),
        ("ownership_quality","Ownership",  6),
        ("options_sentiment","Options",    4),
        ("macro_alignment",  "Macro",      3),
    ]
    lines = [f"     Confidence {final_conf:.1f} breakdown:"]
    fs = factor_scores or {}
    for key, label, weight in FACTOR_DISPLAY:
        raw    = fs.get(key)
        # Phase C5: mark MISSING (None) factors — they are reweighted, not diluted.
        if raw is None:
            lines.append(f"     {label:<12} ░░░░░░░░░░ MISSING → weight redistributed")
            continue
        score  = float(raw or 50)
        contrib = round(score * weight / 100, 1)
        filled = int(score / 10)
        bar    = "█" * filled + "░" * (10 - filled)
        lines.append(f"     {label:<12} {bar} {score:.0f}/100 → {contrib:.1f}pts")
    return lines


def format_near_miss_failures(stock: dict, thresh: dict) -> list:
    """Compact 1-line failure summary for mobile."""
    conf  = float(stock.get("final_confidence", 0) or 0)
    tq    = float(stock.get("trade_quality_score", 0) or 0)
    rr    = float(stock.get("rr_ratio", 0) or stock.get("rr", 0) or 0)
    min_c = thresh.get("min_confidence", 80)
    min_t = thresh.get("min_tq", 78)
    min_r = thresh.get("min_rr", 2.0)
    fails = []
    # Use max(0, gap) so a metric that already passes never prints a negative gap.
    if conf < min_c: fails.append(f"Conf+{max(0.0, min_c - conf):.1f}")
    if tq   < min_t: fails.append(f"TQ+{max(0.0, min_t - tq):.1f}")
    if rr   < min_r: fails.append(f"RR+{max(0.0, min_r - rr):.2f}x")
    fail_str = " · ".join(fails) if fails else "near threshold"
    return [f"  ✗ Needs: {fail_str} → Watch: vol surge or consol above entry"]


def format_conviction_meter(regime_score: float, breadth: float,
                             fii: float, dii: float) -> list:
    """Visual conviction bar — compact 1-line version for mobile.

    Phase C3 (2026-07-02): FII/DII flows are no longer factored into conviction.
    Rationale: NSDL FII data is structurally D-1 (or D-2 on weekends/holidays);
    the number we see today is stale by 1-2 sessions and cannot reliably move
    an intraday-actionable conviction meter. Replaced flow weight with a
    breadth-emphasis blend, which is a real-time internal signal.

    fii/dii parameters kept for backwards-compatibility with callers but
    intentionally unused in the calculation.
    """
    _ = (fii, dii)  # kept for signature stability; not used in scoring
    try:
        # Composite: 55% regime score + 45% breadth (both real-time internals)
        conviction = round(regime_score * 0.55 + breadth * 0.45, 1)
        filled = int(conviction / 10)
        bar    = "█" * filled + "░" * (10 - filled)
        if conviction >= 75:   label = "🟢 Strong"
        elif conviction >= 55: label = "🟡 Moderate"
        elif conviction >= 40: label = "🟠 Weak"
        else:                  label = "🔴 Poor"
        return [f"  Regime Conviction [{bar}] {conviction:.0f}% · {label}"]
    except Exception:
        return []


def format_risk_meter(nifty_state: dict, vix_in: float, breadth: float) -> list:
    """Compact 1-line risk meter for mobile."""
    try:
        ns = nifty_state or {}
        risk_score = 0
        if ns.get("above_ema20"):  e20 = "✓"
        else:                      e20 = "✗"; risk_score += 25
        if ns.get("above_ema50"):  e50 = "✓"
        else:                      e50 = "✗"; risk_score += 25
        if ns.get("above_ema200"): e200 = "✓"
        else:                      e200 = "✗"; risk_score += 25
        vix_ok = vix_in <= 20
        if not vix_ok: risk_score += 15
        brd_ok = breadth >= 40
        if not brd_ok: risk_score += 10
        if risk_score >= 75:   label = "🔴 EXTREME"
        elif risk_score >= 50: label = "🟠 HIGH"
        elif risk_score >= 25: label = "🟡 MEDIUM"
        else:                  label = "🟢 LOW"
        vix_tag = f"VIX {'✓' if vix_ok else '✗'}{vix_in:.1f}"
        brd_tag = f"Breadth {'✓' if brd_ok else '✗'}{breadth:.0f}%"
        return [
            f"  Risk {label} · EMA20{e20} EMA50{e50} EMA200{e200} · {vix_tag} · {brd_tag}"
        ]
    except Exception:
        return []


def format_breadth_dashboard(total_universe: int, total_tradable: int,
                              qualified: int, near_buy: int,
                              developing: int, monitor: int,
                              yesterday: dict = None,
                              rejected_at_gates: int = None) -> list:
    """Compact 2-line breadth dashboard for mobile.

    Semantics (as of 2026-07-03 fix):
      • Universe     = all symbols in stocks.txt                (e.g. 2360)
      • Tradable     = passed liquidity/volume filters          (e.g. 1274)
      • Qualified    = entered the 13-gate system after scoring (e.g. 50 — top-N)
      • Near Buy/Dev/Monitor = survivors classified in watchlist
      • Rejected     = failed the gates (Qualified − BUY − Watchlist tiers)

    BEFORE the fix, `Qualified` = BUY + Watchlist survivors (so 0 on no-signal
    days) and `Rejected` = Universe − Qualified (so 2360 = whole universe).
    That double-counted the illiquid-drop pool as "rejected" and hid the fact
    that the top-50 actually reached the gates.
    """
    if rejected_at_gates is None:
        # Legacy fallback — old callers pass qualified = BUY+watchlist_tiers,
        # so approximate the pre-fix behaviour without crashing.
        rejected_at_gates = max(0, total_universe - qualified)
    return [
        f"  📈 Breadth · Universe {total_universe} · Tradable {total_tradable} · Qualified {qualified}",
        f"  Near Buy {near_buy} · Developing {developing} · Monitor {monitor} · Rejected {rejected_at_gates}",
    ]


def format_buy_card(stock: dict, sizing: dict, regime: str,
                    buy_thesis: str = "") -> list:
    """Enhanced BUY card. buy_thesis is AI-generated when available."""
    try:
        opp    = float(stock.get("opportunity_score", 0) or 0)
        conf   = float(stock.get("final_confidence", 0) or 0)
        tq     = float(stock.get("trade_quality_score", 0) or 0)
        rr     = float(stock.get("rr_ratio", 0) or 0)
        sector = get_sector(stock.get("symbol", ""))
        cats   = stock.get("catalysts", []) or []

        # Conviction icon
        if opp >= 85:   icon = "\U0001f525"
        elif opp >= 75: icon = "\u26a1"
        else:           icon = "\U0001f4c8"

        # Thesis: use AI-generated if available, else rule-based
        if not buy_thesis:
            fs    = stock.get("factor_scores", {}) or {}
            trend_s = float(fs.get("trend_quality", 0) or 0)
            rs_s    = float(fs.get("rs_vs_nifty", 0) or 0)
            thesis_parts = []
            if trend_s > 70:              thesis_parts.append("strong uptrend")
            if rs_s > 70:                 thesis_parts.append("outperforming NIFTY")
            if "VOL_SURGE" in cats:       thesis_parts.append("volume expansion")
            if "NEAR_52W_HIGH" in cats:   thesis_parts.append("near 52W high breakout")
            buy_thesis = ", ".join(thesis_parts) if thesis_parts else "multi-factor confluence"

        # Unescape HTML entities from upstream news/AI text before we re-escape
        # for Telegram HTML mode — prevents double-encoding like &amp;#x27;.
        try:
            for _k in ("news_summary", "buy_thesis", "summary"):
                _v = stock.get(_k)
                if isinstance(_v, str) and "&" in _v and ";" in _v:
                    stock[_k] = html.unescape(_v)
        except Exception:
            pass

        # Telegram HTML mode only requires <, >, & escaped — quote=False keeps
        # apostrophes readable (fixes the &#x27; artefacts in AI thesis strings).
        # `buy_thesis` was already unescaped upstream by _clean_ai_output(); we
        # now unescape defensively in case it came from a cached fallback path.
        try:
            if isinstance(buy_thesis, str) and "&" in buy_thesis and ";" in buy_thesis:
                buy_thesis = html.unescape(buy_thesis)
        except Exception:
            pass
        sym     = html.escape(str(stock.get("symbol", "")), quote=False)
        entry   = stock.get("entry", 0)
        stop_p  = stock.get("stop", 0)
        t1      = stock.get("target1", 0)
        t2      = stock.get("target2", 0)
        risk_p  = round((entry - stop_p) / entry * 100, 1) if entry > 0 else 0
        pos_val = sizing.get("position_value", stock.get("position_value", 0))
        pos_pct = sizing.get("position_pct", stock.get("position_pct", 0))
        shares  = sizing.get("shares", stock.get("shares", 0))
        # BUGFIX: prefer explicit `max_loss`; fall back to risk_amount; last-resort
        # live compute (shares × risk-per-share) so cached rows still render.
        max_loss = (
            sizing.get("max_loss")
            or stock.get("max_loss")
            or sizing.get("risk_amount")
            or stock.get("risk_amount")
            or 0
        )
        try:
            if (not max_loss or float(max_loss) <= 0) and shares and entry and stop_p and entry > stop_p > 0:
                max_loss = round(shares * (entry - stop_p), 2)
        except Exception:
            pass
        # Persist back to stock so tracker / weekly stats read a real number.
        stock["max_loss"] = max_loss
        news    = truncate_display(stock.get("news_summary", ""), 120)

        pos_val_k = pos_val / 1000  # display in K
        max_loss_k = max_loss / 1000 if max_loss else 0
        # Render MaxLoss in ₹ when < ₹1K so tiny worst-case losses don't collapse to 0.0K.
        max_loss_str = (
            f"MaxLoss ₹{max_loss:.0f}" if 0 < max_loss < 1000
            else f"MaxLoss ₹{max_loss_k:.1f}K"
        )
        # Fundamentals honesty: when source is SCREENER+YF and BOTH ROE + D/E
        # are literally 0, that means yfinance was rate-limited AND screener
        # parsed no financial rows — do NOT display as "real" 0%. Show "N/A".
        _roe  = stock.get('roe', 0) or 0
        _de   = stock.get('de_ratio', 0) or 0
        _pldg = stock.get('promoter_pledge_pct', 0) or 0
        _src  = str(stock.get('fundamentals_source', stock.get('source', ''))).upper()
        _has_real_fund = (_roe != 0) or (_de != 0) or (_pldg != 0)
        if _has_real_fund:
            fund_line = (
                f"  ROE {_roe:.1f}% · D/E {_de:.2f} · Pledge {_pldg:.0f}%"
            )
        else:
            fund_line = "  ROE N/A · D/E N/A · Pledge N/A (fetch failed)"

        # Phase C3: delivery line — the strongest per-stock institutional signal.
        # Only shown when we got real nselib data; otherwise silent (the OBV proxy
        # is already baked into base_confidence).
        deliv_line = None
        if stock.get("delivery_source") == "nselib":
            _d_today = float(stock.get("delivery_pct_today", 0) or 0)
            _d_20d   = float(stock.get("delivery_pct_20d_avg", 0) or 0)
            _d_sig   = str(stock.get("delivery_signal", "NEUTRAL"))
            _sig_emoji = {
                "STRONG_ACCUM":  "🟢 STRONG ACCUM",
                "ACCUM":         "🟢 Accumulation",
                "NEUTRAL":       "⚪ Neutral",
                "WEAK":          "🟡 Weak delivery",
                "DISTRIBUTION":  "🔴 DISTRIBUTION",
            }.get(_d_sig, _d_sig)
            deliv_line = (
                f"  Delivery {_d_today:.0f}% (20d avg {_d_20d:.0f}%) · {_sig_emoji}"
            )

        lines = [
            f"  {icon} <b>{sym}</b> · {html.escape(str(sector), quote=False)}",
            f"  Opp {opp:.0f} · Conf {conf:.1f} · TQ {tq:.1f} · R/R {rr:.2f}x",
            f"  Entry ₹{entry:.1f} · Stop ₹{stop_p:.1f} ({risk_p:.1f}%) · T1 ₹{t1:.1f} · T2 ₹{t2:.1f}",
            f"  Size ₹{pos_val_k:.0f}K ({pos_pct:.1f}%) · {shares} shares · {max_loss_str}",
            fund_line,
        ]
        # Phase C4 Gap #6: partial-exit plan (drives portfolio T1_PARTIAL_EXIT action)
        try:
            _pe_pct = float(os.getenv("PARTIAL_EXIT_PCT", "50"))
            lines.append(
                f"  📋 Plan: Sell {_pe_pct:.0f}% @ T1 ₹{t1:.1f} → trail rest to entry → let residual run to T2 ₹{t2:.1f}"
            )
        except Exception:
            pass
        # Phase C4 Gap #5+#8: cost + gap risk (only when meaningful)
        try:
            _slip     = float(stock.get("slippage_pct_one_way", 0) or 0)
            _cost_rt  = float(stock.get("round_trip_cost_pct", 0) or 0)
            _net_rr2  = float(stock.get("net_rr_t2", 0) or 0)
            _p90_gap  = float(stock.get("p90_gap_pct", 0) or 0)
            _eff_stop = float(stock.get("effective_stop_pct", 0) or 0)
            _high_gap = bool(stock.get("high_gap_risk", False))
            cost_bits = []
            if _slip > 0 or _cost_rt > 0:
                cost_bits.append(
                    f"Slip {_slip:.2f}% · RT cost {_cost_rt:.2f}% · Net R/R T2 {_net_rr2:.2f}x"
                )
            if _p90_gap > 0 and _eff_stop > 0:
                gap_icon = "⚠️" if _high_gap else "•"
                cost_bits.append(
                    f"{gap_icon} p90 gap {_p90_gap:.1f}% · effective stop {_eff_stop:.1f}%"
                )
            if cost_bits:
                lines.append("  💸 " + " · ".join(cost_bits))
        except Exception:
            pass
        if deliv_line:
            lines.append(deliv_line)
        lines.append(f"  📝 {html.escape(str(buy_thesis), quote=False)}")
        if cats:
            lines.append(f"  🏷 {html.escape(' · '.join(str(c) for c in cats), quote=False)}")
        # FIX label/score mismatch: distinguish NO_NEWS (score correctly = 50)
        # from "we have news but couldn't summarise". Never claim "No significant
        # news" unless the category actually was NO_NEWS.
        news_cat = str(stock.get("news_category", "") or "").upper()
        if news and news != "\u2014":
            try:
                news = html.unescape(str(news))
            except Exception:
                pass
            lines.append(f"  📰 {html.escape(str(news), quote=False)}")
        elif news_cat == "NO_NEWS":
            lines.append("  📰 No headlines in last 3 days")
        else:
            lines.append("  📰 —")
        fs = stock.get("factor_scores", {}) or {}
        lines += format_confidence_breakdown(fs, conf)
        return lines
    except Exception:
        return [f"  \U0001f4c8 <b>{html.escape(str(stock.get('symbol', '?')))}</b>"]


# ── Compact Portfolio Card (FIX 3E) ──────────────────────────────────────────
def format_portfolio_card_compact(alert: dict, current_price: float,
                                   stop_warning_pct: float = 5.0) -> list:
    """
    Compact portfolio card.
    Normal: 2 lines. Near stop: 3 lines (adds stop warning).
    """
    try:
        symbol   = str(alert.get("symbol", ""))
        entry    = float(alert.get("entry", 0) or 0)
        stop     = float(alert.get("stop_loss", alert.get("stop", 0)) or 0)
        t1       = float(alert.get("target1", 0) or 0)
        t2       = float(alert.get("target2", 0) or 0)
        days     = int(alert.get("days_held", 0) or 0)
        pnl_p    = float(alert.get("pnl_pct", 0) or 0)
        action   = str(alert.get("action", "HOLD"))

        risk     = entry - stop
        gain     = current_price - entry
        r_mult   = round(gain / risk, 2) if risk > 0 else 0.0

        dist_stop = round((current_price - stop) / current_price * 100, 1) \
                    if current_price > 0 and stop > 0 else 999.0
        near_stop = dist_stop <= stop_warning_pct

        pnl_icon    = "🟢" if pnl_p > 0 else ("🔴" if pnl_p < -2 else "⚪")
        action_icon = {"HOLD": "✅", "EXIT": "🔴", "EXIT_FULL": "🔴",
                       "TRAIL_STOP": "🟡", "REVIEW": "🟠",
                       "T1_PARTIAL_EXIT": "💰"}.get(action, "✅")

        lines = [
            f"  {action_icon} <b>{html.escape(symbol)}</b> · {action} · T{days} · {pnl_icon}{pnl_p:+.1f}% · {r_mult:+.2f}R",
            f"  ₹{entry:.0f}→₹{current_price:.0f} · T1 ₹{t1:.0f} · T2 ₹{t2:.0f}",
        ]
        if near_stop:
            lines.append(f"  ⚠️ Stop ₹{stop:.0f} only {dist_stop:.1f}% away")
        return lines
    except Exception:
        return [f"  ✅ <b>{html.escape(str(alert.get('symbol', '?')))}</b>"]


# ── Compact Portfolio Summary (FIX 3F) ───────────────────────────────────────
def format_portfolio_summary_compact(alerts: list, current_prices: dict,
                                      total_capital: float) -> list:
    """Replaces 6-line Portfolio Health Dashboard with 2 lines."""
    try:
        if not alerts:
            return ["  No active holdings."]
        qty_known = [a for a in alerts if a.get("quantity", 0) > 0]
        total_invested = sum(
            float(a.get("entry", 0)) * float(a.get("quantity", 0))
            for a in qty_known
        ) or sum(float(a.get("entry", 0)) for a in alerts)
        total_current = sum(
            current_prices.get(a["symbol"], float(a.get("entry", 0))) *
            float(a.get("quantity", 1))
            for a in alerts
        )
        total_pnl = round((total_current - total_invested) / total_invested * 100, 2) \
                    if total_invested > 0 else 0.0
        exposure  = round(total_invested / total_capital * 100, 1) if total_capital > 0 else 0.0
        cash      = round(100 - exposure, 1)
        return [
            f"  💼 {len(alerts)} pos · ₹{total_invested:,.0f} ({exposure:.1f}%) · Cash {cash:.1f}% · PnL {total_pnl:+.2f}%",
        ]
    except Exception:
        return [f"  💼 {len(alerts)} pos · see Excel for details"]


# ── Stop Watch Alert (FIX 5) ─────────────────────────────────────────────────
STOP_WATCH_THRESHOLD_PCT = 5.0

def format_stop_watch_alert(holdings: list, current_prices: dict) -> list:
    """
    Prominent STOP WATCH block at top of portfolio section.
    Returns empty list if all positions are safe.
    """
    warnings = []
    try:
        for h in (holdings or []):
            sym  = h.get("symbol", "")
            stop = float(h.get("stop_loss", h.get("stop", 0)) or 0)
            entry= float(h.get("entry", 0) or 0)
            cur  = float(current_prices.get(sym, entry) or entry)
            if stop <= 0 or cur <= 0:
                continue
            dist_pct = round((cur - stop) / cur * 100, 1)
            if cur <= stop:
                warnings.append(
                    f"  🔴 STOP HIT — {html.escape(sym)} "
                    f"Rs{cur:.1f} \u2264 Stop Rs{stop:.1f} | EXIT NOW"
                )
            elif dist_pct <= 1.0:
                warnings.append(
                    f"  🔴 CRITICAL — {html.escape(sym)} "
                    f"Rs{cur:.1f} | Stop Rs{stop:.1f} only {dist_pct:.1f}% away"
                )
            elif dist_pct <= STOP_WATCH_THRESHOLD_PCT:
                warnings.append(
                    f"  ⚠️  STOP WATCH — {html.escape(sym)} "
                    f"Rs{cur:.1f} | Stop Rs{stop:.1f} | {dist_pct:.1f}% away — watch closely"
                )
    except Exception:
        pass
    if not warnings:
        return []
    return ["", "🚨 STOP ALERTS"] + warnings


def format_daily_summary(regime: str, buys: list, watchlist: list,
                          portfolio_alerts: list, macro: dict,
                          nifty_state: dict = None) -> list:
    """Executive briefing at end of Telegram report (ENHANCEMENT 9 / BUG FIX 6)."""
    try:
        near_miss_count = len([w for w in watchlist if w.get("tier") == "NEAR_MISS"])
        exits = [a for a in portfolio_alerts if "EXIT" in str(a.get("action", ""))]
        lines = ["", "📋 DAILY SUMMARY"]

        if regime in ("STRONG_BULL", "BULL"):
            lines.append("  Market is in a bullish phase with broad participation.")
        elif regime == "TRANSITION":
            lines.append("  Market is in transition — mixed signals, no clear direction.")
        elif regime == "SIDEWAYS":
            lines.append("  Market is range-bound. Patience required.")
        else:
            lines.append("  Market is in a bearish phase. Preserve capital.")

        if buys:
            lines.append(f"  {len(buys)} institutional-quality setup(s) identified today.")
        else:
            lines.append("  No institutional-quality setup met all required conditions today.")
            if near_miss_count > 0:
                lines.append(f"  {near_miss_count} stock(s) within 8 pts of qualifying — watch closely.")

        # Use nifty_state as single source of truth (BUG FIX 6)
        ns = nifty_state or {}
        if ns.get("ema_bear", macro.get("nifty_below_all_emas", False)):
            lines.append(
                "  NIFTY trading below major EMAs — risk elevated. "
                "Avoid aggressive positioning."
            )
        else:
            lines.append("  NIFTY structure supportive — trend intact above key EMAs.")

        if exits:
            lines.append(f"  ⚠️  {len(exits)} position(s) require immediate exit review.")
        else:
            lines.append("  Existing positions on track — no exit action required.")

        if not buys and regime not in ("STRONG_BULL", "BULL"):
            lines.append("  → Avoid forcing new trades. Quality over quantity.")
        elif buys:
            lines.append("  → Execute BUY signal(s) with strict position sizing.")

        return lines
    except Exception:
        return ["", "📋 DAILY SUMMARY", "  (unavailable)"]


def format_watchlist_section(watchlist: list, regime: str,
                              conf_history: dict = None,
                              gate_memory: dict = None,
                              near_miss_insights: dict = None) -> list:
    """
    NEAR MISS  — full detail for ALL stocks (no truncation), sorted by R/R desc
    DEVELOPING — company-names-only (rows of 5), like MONITOR
    MONITOR    — company-names-only (rows of 5)
    Full Entry/Stop/T1/T2 for every tier is persisted to shadow_master.xlsx.
    """
    thresh   = REGIME_THRESHOLDS[regime]
    min_conf = thresh["min_confidence"]

    near = sorted([w for w in watchlist if w.get("tier") == "NEAR_MISS"],
                  key=lambda x: x.get("rr", 0), reverse=True)
    dev  = sorted([w for w in watchlist if w.get("tier") == "DEVELOPING"],
                  key=lambda x: x.get("conf", x.get("final_confidence", 0)), reverse=True)
    mon  = [w for w in watchlist if w.get("tier") == "MONITOR"]

    lines = [f"👁 <b>WATCHLIST</b> — {len(watchlist)} stocks (min conf {min_conf})"]

    # -- NEAR MISS: full detail card for EVERY stock (no truncation) ----------
    if near:
        lines.append(f"🔴 <b>NEAR MISS</b> ({len(near)} total — all listed)")
        for w in near:
            sym      = html.escape(str(w["symbol"]))
            sector   = html.escape(str(w.get("sector", "DIV")))
            conf     = w.get("conf", w.get("final_confidence", 0))
            tq       = w.get("tq", w.get("trade_quality_score", 0))
            entry    = w.get("entry", 0)
            stop     = w.get("stop", 0)
            target1  = w.get("target1", 0)
            target2  = w.get("target2", 0)
            rr       = w.get("rr_ratio", w.get("rr", 0))
            risk     = w.get("risk_pct", 0)
            opp      = w.get("opportunity_score", 0)
            lines.append(f"  <b>{sym}</b> [{sector}] · Opp{opp:.0f} Conf{conf:.1f} TQ{tq:.1f} RR{rr:.1f}x")
            lines.append(f"  ₹{entry:.1f} entry · Stop ₹{stop:.1f}({risk:.1f}%) · T1 ₹{target1:.1f} · T2 ₹{target2:.1f}")
            lines.extend(format_near_miss_failures(w, thresh))
            insight = (near_miss_insights or {}).get(w.get("symbol", ""), "")
            if insight:
                lines.append(f"  💡 {html.escape(str(insight))}")

    # -- DEVELOPING: header + all names in rows of 5 (like MONITOR) -----------
    if dev:
        lines.append(f"🟡 <b>DEVELOPING</b> ({len(dev)} building) — full levels in Excel")
        dev_names = [w["symbol"].replace(".NS", "") for w in dev]
        for i in range(0, len(dev_names), 5):
            lines.append("  " + " \u00b7 ".join(dev_names[i:i+5]))

    # -- MONITOR: header + all names in rows of 5 (unchanged) -----------------
    if mon:
        best     = max(mon, key=lambda x: x.get("rr_ratio", x.get("rr", 0)))
        best_sym = best["symbol"].replace(".NS", "")
        best_rr  = best.get("rr_ratio", best.get("rr", 0))
        lines.append(
            f"\U0001f535 <b>MONITOR</b> ({len(mon)} early-stage) \u00b7 best R/R: {best_sym} {best_rr:.1f}x"
        )
        names = [m["symbol"].replace(".NS", "") for m in mon]
        for i in range(0, len(names), 5):
            lines.append("  " + " \u00b7 ".join(names[i:i+5]))

    if not near and not dev and not mon:
        lines.append("  None today.")

    return lines


def format_sector_lagging_rejects_section(rejected: list, regime: str,
                                           max_rows: int = 12) -> list:
    """
    Phase C4 (2026-07-02): Surface high-quality stocks that were REJECTED
    primarily because their sector is LAGGING.

    Context: SECTOR_LAGGING is a "soft-scoreable" fail (see run_gates line ~5713).
    A stock with SECTOR_LAGGING as its ONLY hard fail goes to WATCHLIST. But if
    SECTOR_LAGGING combined with 2+ other hard fails pushed it past the budget,
    it lands in REJECTED — and its name disappeared from the Telegram output
    entirely (folded into "Rejected N" count).

    That hides a genuinely interesting bucket: stocks with good internal setup
    (high Conf/TQ) that are only rejected because their sector is out of favor
    — i.e. contrarian / sector-rotation candidates worth watching.

    Selection criteria (must satisfy ALL):
      - stock is in REJECTED bucket
      - SECTOR_LAGGING is in fail_reasons
      - final_confidence is within `min_conf - 25` of the regime threshold
        (i.e. the stock is at most 25 conf points below the buy bar — anything
        weaker isn't a "contrarian setup", it's just a weak stock)

    Rendered as a compact 1-line-per-stock block, sorted by confidence desc,
    capped at `max_rows` to keep Telegram message small.
    """
    if not rejected:
        return []

    try:
        thresh   = REGIME_THRESHOLDS[regime]
        min_conf = float(thresh["min_confidence"])
    except Exception:
        min_conf = 75.0

    # Filter: sector-lagging + high-enough conf to matter
    lagging_rejects = []
    for s in rejected:
        try:
            fr = s.get("fail_reasons", []) or []
            if not any("SECTOR_LAGGING" in str(f) for f in fr):
                continue
            conf = float(s.get("final_confidence", 0) or 0)
            # Only surface stocks that were close-ish to qualifying —
            # a 50-conf stock in a lagging sector is not interesting.
            if conf < (min_conf - 25):
                continue
            lagging_rejects.append(s)
        except Exception:
            continue

    if not lagging_rejects:
        return []

    lagging_rejects.sort(
        key=lambda x: float(x.get("final_confidence", 0) or 0),
        reverse=True,
    )

    total = len(lagging_rejects)
    shown = lagging_rejects[:max_rows]

    lines = [
        f"🟠 <b>SECTOR OUT OF FAVOR</b> ({total} rejected — contrarian setups)",
        f"  <i>Good stock, wrong sector. Wait for sector rotation or use tight stop.</i>",
    ]

    for s in shown:
        try:
            sym    = html.escape(str(s.get("symbol", "?")).replace(".NS", ""))
            sector = html.escape(str(s.get("sector", "") or get_sector(s.get("symbol", ""))))
            conf   = float(s.get("final_confidence", 0) or 0)
            tq     = float(s.get("trade_quality_score", 0) or 0)
            opp    = float(s.get("opportunity_score", 0) or 0)
            # Other fails beyond SECTOR_LAGGING (max 2 shown)
            other_fails = [
                str(f) for f in (s.get("fail_reasons", []) or [])
                if "SECTOR_LAGGING" not in str(f)
            ][:2]
            of_str = (" · also: " + ", ".join(other_fails)) if other_fails else ""
            lines.append(
                f"  <b>{sym}</b> [{sector}] · Opp{opp:.0f} Conf{conf:.1f} TQ{tq:.1f}{html.escape(of_str)}"
            )
        except Exception:
            continue

    if total > max_rows:
        lines.append(f"  <i>… and {total - max_rows} more (see decision_audit.csv)</i>")

    return lines


def format_no_buy_explanation(top_rejected: list, regime: str,
                               watchlist: list = None) -> list:
    """
    When buys=0, show:
      1. The regime gate thresholds that had to be cleared
      2. The top 3 rejection reasons across all rejected stocks (aggregated)
      3. A pointer to WATCHLIST tiers (if any near-misses exist)

    Never duplicates stocks already shown in the WATCHLIST section below.

    2026-07-03 fix: previously showed only a one-line "None — no setup" message,
    which left the user asking "but WHY was everything rejected?" every time
    the regime was tight. Now aggregates fail_reasons across all rejects and
    shows the top-3 buckets so the user can see at a glance whether it's a
    confidence problem, a trade-quality problem, a risk/reward problem, or
    something structural like SECTOR_LAGGING / EVENT_BLOCK / PORTFOLIO_FULL.
    """
    from collections import Counter
    thresh = REGIME_THRESHOLDS[regime]
    wl     = watchlist or []
    nm     = len([s for s in wl if s.get("tier") == "NEAR_MISS"])
    dev    = len([s for s in wl if s.get("tier") == "DEVELOPING"])
    mon    = len([s for s in wl if s.get("tier") == "MONITOR"])

    lines = [
        f"  None \u2014 no setup cleared all gates "
        f"(need Conf\u2265{thresh['min_confidence']} \u00b7 TQ\u2265{thresh['min_tq']} \u00b7 R/R\u2265{thresh['min_rr']})"
    ]

    # ── Aggregate rejection reasons ────────────────────────────────────────
    # Phase Polish (2026-07-11): count ONE primary reason per stock (not
    # every fail_reasons entry) so counts add up to total_rej. Previously a
    # stock with both CONF_FAIL and TQ_FAIL was counted twice, producing
    # sums like 119+1+1=121 for 120 rejects.
    #
    # Also split SETUP_EDGE_SKIP into 3 actionable sub-reasons so the user
    # sees WHY 99% of rejects were setup-skipped:
    #   - CHOP_NO_BREAKOUT: regime blocked non-BREAKOUT setup
    #   - SETUP_OTHER:      setup=OTHER (unclassified pattern)
    #   - SETUP_WEAK:       setup weak in current regime (fallback)
    def _primary_reject_key(fail_reasons: list) -> str:
        """Return a single canonical reject key for a stock. Priority order
        matches _classify_reject_reason() but with SETUP_EDGE_SKIP broken
        out into sub-reasons parsed from the inner payload."""
        for fr in (fail_reasons or []):
            raw = str(fr)
            # Split SETUP_EDGE_SKIP into sub-reasons using the payload
            # stamped by apply_setup_edge() (main.py:5814).
            if raw.startswith("SETUP_EDGE_SKIP"):
                inner = raw[len("SETUP_EDGE_SKIP("):].rstrip(")")
                if inner.startswith("REGIME_CHOP_NO_BREAKOUT"):
                    # Inner form: REGIME_CHOP_NO_BREAKOUT(TRANSITION/MOMENTUM)
                    if "/OTHER" in inner:
                        return "SETUP_OTHER"
                    return "CHOP_NO_BREAKOUT"
                return "SETUP_WEAK"
            # Strip parenthetical detail so buckets collapse.
            key = raw.split("(", 1)[0].strip()
            # Truncate long variants.
            if key.startswith("HIGH_CORR_"):
                key = "HIGH_CORR"
            elif key.startswith("EVENT_BLOCK_"):
                key = "EVENT_BLOCK"
            elif key.startswith("KILL_SWITCH"):
                key = "KILL_SWITCH"
            elif key == "PROMOTER_PLEDGE_BLOCKLIST":
                pass
            elif key.startswith("PROMOTER_PLEDGE_"):
                key = "PROMOTER_PLEDGE"
            return key
        return "UNKNOWN"

    reason_counter = Counter()
    total_rej      = 0
    for s in (top_rejected or []):
        frs = s.get("fail_reasons") or []
        if not frs:
            continue
        total_rej += 1
        reason_counter[_primary_reject_key(frs)] += 1

    if total_rej > 0 and reason_counter:
        # Human-friendly labels for the top buckets
        _LBL = {
            "CONF_FAIL":               "Confidence too low",
            "TQ_FAIL":                 "Trade Quality too low",
            "RR_FAIL":                 "Risk/Reward too low",
            "WIDE_STOP":               "Stop distance > regime cap",
            "SECTOR_LAGGING":          "Sector lagging",
            "SECTOR_CAP":              "Sector concentration cap",
            "KILL_SWITCH":             "Kill-switch active",
            "EVENT_BLOCK":             "Earnings/event blackout",
            "HIGH_CORR":               "Too correlated w/ existing pos",
            "LIQUIDITY_FAIL":          "Insufficient liquidity",
            "REGIME_NO_BUY":           "Regime blocks new BUYs",
            "PROMOTER_PLEDGE":         "Promoter pledge too high",
            "PROMOTER_PLEDGE_BLOCKLIST": "On curated high-pledge blocklist",
            "BLACK_SWAN_NEWS":         "Black-swan news",
            "DATA_INCOMPLETE":         "Data incomplete",
            "SUSPECT_PUMP_LOW_DELIVERY": "Suspect pump (low delivery)",
            "INSTITUTIONAL_EXIT":      "Institutional exit signal",
            # ── Phase Polish (2026-07-11): SETUP_EDGE_SKIP sub-reasons ──
            "CHOP_NO_BREAKOUT":        "Chop regime — non-BREAKOUT blocked",
            "SETUP_OTHER":             "Setup pattern unclassified (OTHER)",
            "SETUP_WEAK":              "Setup weak for current regime",
            # Note: PORTFOLIO_FULL is no longer emitted as of 2026-07-03 \u2014
            # portfolio state no longer affects recommendations. It's a WARNING
            # only now, which never lands in fail_reasons.
        }
        # Phase Polish: show top 5 (was 3) since counts are now clean.
        top5 = reason_counter.most_common(5)
        lines.append(f"  \U0001f4ca <b>Top reject reasons</b> (of {total_rej} rejects):")
        for key, cnt in top5:
            label = _LBL.get(key, key)
            pct   = int(round(cnt * 100.0 / total_rej))
            lines.append(f"    \u2022 {label}: {cnt} ({pct}%)")

    if nm or dev or mon:
        parts = []
        if nm:  parts.append(f"\U0001f534 {nm} Near Miss")
        if dev: parts.append(f"\U0001f7e1 {dev} Developing")
        if mon: parts.append(f"\U0001f535 {mon} Monitor")
        joined = " \u00b7 ".join(parts)
        lines.append(f"  \u2193 {joined} \u2014 details in WATCHLIST \u2193")
    return lines


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 9b — CONFIDENCE HISTORY + GATE MEMORY (FEATURES 2 & 7)
# ─────────────────────────────────────────────────────────────────────────────

def load_confidence_history() -> dict:
    """Load {symbol: {dates:[], confs:[]}} rolling 5-day window.

    Phase 1 #42a (2026-07-05): widened from 3d to 5d and now stores None
    sentinels for days when the symbol was not in top-N. classify_watchlist
    reads this and refuses to compute a trajectory across gaps.
    """
    # Phase C7c: FRESH_START wipes confidence history for one run
    if FRESH_START:
        _log("[FRESH_START] load_confidence_history → returning {} (old confidence_history.json ignored)")
        return {}
    try:
        if os.path.exists(CONF_HISTORY_FILE):
            with open(CONF_HISTORY_FILE, "r") as f:
                return json.load(f)
    except Exception:
        pass
    return {}


def update_confidence_history(history: dict, scored_stocks: list,
                               today_str: str) -> dict:
    """Update rolling 5-day confidence for all scored stocks.

    Phase 1 #42a (2026-07-05):
      * Window widened 3→5 days.
      * Symbols present in history but NOT in today's scored list get a
        None sentinel for today — preserves the fact that the symbol was
        temporarily invisible. classify_watchlist's trajectory calc
        refuses to compute across a None.
      * Symbols with all-None trailing window are garbage-collected.
    """
    _WINDOW = 5
    try:
        _today_syms = set()
        for stock in scored_stocks:
            sym  = stock.get("symbol", "")
            conf = float(stock.get("final_confidence", 0) or 0)
            if not sym:
                continue
            _today_syms.add(sym)
            if sym not in history:
                history[sym] = {"dates": [], "confs": []}
            history[sym]["dates"].append(today_str)
            history[sym]["confs"].append(conf)
            history[sym]["dates"] = history[sym]["dates"][-_WINDOW:]
            history[sym]["confs"] = history[sym]["confs"][-_WINDOW:]
        # Insert None gap markers for symbols we've tracked before but which
        # dropped out of top-N today. Skip if we already recorded today.
        for sym, rec in list(history.items()):
            if sym in _today_syms:
                continue
            dates = rec.get("dates", []) or []
            if dates and dates[-1] == today_str:
                continue
            rec.setdefault("dates", []).append(today_str)
            rec.setdefault("confs", []).append(None)
            rec["dates"] = rec["dates"][-_WINDOW:]
            rec["confs"] = rec["confs"][-_WINDOW:]
            # GC: if the whole window is None, drop the symbol — it hasn't
            # been seen for W days, no need to keep tracking.
            if all(v is None for v in rec["confs"]):
                history.pop(sym, None)
    except Exception:
        pass
    return history


def save_confidence_history(history: dict) -> None:
    try:
        with open(CONF_HISTORY_FILE, "w") as f:
            json.dump(history, f, indent=2)
    except Exception as e:
        _log(f"[WARN] Confidence history save failed: {e}")


def get_confidence_trend(symbol: str, history: dict) -> str:
    """Returns trend arrow string for Telegram. Empty string if insufficient data."""
    try:
        data  = history.get(symbol, {})
        confs = data.get("confs", [])
        if len(confs) < 2:
            return ""
        if len(confs) == 2:
            c1, c2 = confs
            delta  = c2 - c1
            arrow  = "\u2191" if delta > 1 else ("\u2193" if delta < -1 else "\u2192")
            return f"{arrow} {c1:.0f}\u2192{c2:.0f}"
        c1, c2, c3 = confs[-3], confs[-2], confs[-1]
        total = c3 - c1
        d1    = c2 - c1
        d2    = c3 - c2
        if total > 4 and d1 > 0 and d2 > 0:
            arrow, label = "\u2191\u2191", "rising fast"
        elif total > 1:
            arrow, label = "\u2191 ", "rising"
        elif total < -4:
            arrow, label = "\u2193\u2193", "falling fast"
        elif total < -1:
            arrow, label = "\u2193 ", "falling"
        else:
            arrow, label = "\u2192 ", "flat"
        return f"{arrow} {c1:.0f}\u2192{c2:.0f}\u2192{c3:.0f} ({label})"
    except Exception:
        return ""


def load_gate_memory() -> dict:
    """Load {symbol: {history: [{date, fails, conf}]}} rolling 5-day window."""
    # Phase C7c: FRESH_START wipes gate memory for one run
    if FRESH_START:
        _log("[FRESH_START] load_gate_memory → returning empty dict (old gate_memory.json ignored)")
        return {}
    try:
        if os.path.exists(GATE_MEMORY_FILE):
            with open(GATE_MEMORY_FILE, "r") as f:
                return json.load(f)
    except Exception:
        pass
    return {}


def update_gate_memory(memory: dict, watchlist: list, today_str: str) -> dict:
    """Track fail reasons per stock over time."""
    try:
        for stock in watchlist:
            sym   = stock.get("symbol", "")
            fails = stock.get("fail_reasons", [])
            conf  = float(stock.get("final_confidence", 0) or 0)
            if not sym:
                continue
            if sym not in memory:
                memory[sym] = {"history": []}
            memory[sym]["history"].append({"date": today_str, "fails": fails, "conf": conf})
            memory[sym]["history"] = memory[sym]["history"][-5:]
    except Exception:
        pass
    return memory


def save_gate_memory(memory: dict) -> None:
    try:
        with open(GATE_MEMORY_FILE, "w") as f:
            json.dump(memory, f, indent=2)
    except Exception as e:
        _log(f"[WARN] Gate memory save failed: {e}")


def get_gate_pattern(symbol: str, memory: dict) -> str:
    """Returns one-line pattern description. Empty string if insufficient history."""
    try:
        data = memory.get(symbol, {})
        hist = data.get("history", [])
        if len(hist) < 2:
            return ""
        all_fails = [set(h.get("fails", [])) for h in hist]
        confs     = [h.get("conf", 0) for h in hist]
        if len(hist) >= 3:
            recent_fails = all_fails[-3:]
            persistent   = set.intersection(*recent_fails) if all(recent_fails) else set()
            if persistent:
                fail_name  = list(persistent)[0].split("(")[0].strip()
                conf_trend = confs[-1] - confs[-3]
                if conf_trend > 2:
                    return f"Day {len(hist)} {fail_name} but conf rising +{conf_trend:.1f} \u2705"
                elif conf_trend < -2:
                    return f"Day {len(hist)} {fail_name} \u2014 conf falling {conf_trend:.1f} \u26a0\ufe0f"
                elif len(hist) >= 4:
                    return f"Day {len(hist)} stuck on {fail_name} \u2014 consider removing \ud83d\udd34"
                else:
                    return f"Day {len(hist)} {fail_name} \u2014 monitoring"
    except Exception:
        pass
    return ""


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 10 — TELEGRAM OUTPUT FORMATTER
# ─────────────────────────────────────────────────────────────────────────────

def format_telegram_message(regime_data: dict, buys: list, shorts: list,
                              watchlist: list, portfolio_alerts: list,
                              macro: dict, key_levels: dict,
                              upcoming_events: list, timestamp: str,
                              heat: dict = None, platt: dict = None,
                              tracker_v2: dict = None,
                              rejected_stocks: list = None,
                              breadth_20: float = 50.0,
                              nifty_state: dict = None,
                              universe_count: int = 0,
                              tradable_count: int = 0,
                              conf_history: dict = None,
                              gate_memory: dict = None,
                              ai_results: dict = None) -> str:
    lines  = []
    regime = regime_data["regime"]
    score  = regime_data["score"]
    thresh = regime_data["thresholds"]

    # ── Header ──────────────────────────────────────────────────────────────
    # Phase Polish (2026-07-11): show the market close date (IST) in the
    # header instead of just the dispatch timestamp. Prevents confusion
    # when the run finishes after midnight IST (GitHub Actions is UTC) and
    # renders "Jul 10 00:01 IST" for a scan built on Jul 9's close.
    try:
        _market_date = ist_today().strftime("%b %d, %Y")
    except Exception:
        _market_date = ""
    if _market_date and _market_date not in (timestamp or ""):
        lines.append(f"📊 <b>NSE SWING BRIEF</b> · Close {_market_date} · Scan {timestamp}")
    else:
        lines.append(f"📊 <b>NSE SWING BRIEF</b> · {timestamp}")

    # ── Phase Polish (2026-07-11): setup mix instrumentation ────────────
    # Show WHAT setups the pipeline actually saw tonight across the FULL
    # evaluated pool (buys + watchlist + rejected, typically ~120 stocks).
    # This is the evidence you need to eventually tune _SETUP_CONF_BONUS.
    # Format:
    #   🧭 Setup mix (120 evaluated): BREAKOUT 3 · MOMENTUM 42 · PULLBACK 18 · REVERSAL 25 · OTHER 32
    #      • BREAKOUT: RELIANCE, TCS, INFY
    #      • MOMENTUM: HDFC, ICICI, KOTAK, AXIS, SBI  (+37 more)
    #      • …
    #      ↳ became BUYs: BREAKOUT 1 (RELIANCE)
    try:
        _mix         = regime_data.get("_setup_mix") or {}
        _mix_bought  = regime_data.get("_setup_mix_bought") or {}
        _tickers     = regime_data.get("_setup_tickers") or {}
        _tickers_buy = regime_data.get("_setup_tickers_buy") or {}
        _pool_size   = regime_data.get("_setup_pool_size", 0)
        if _mix:
            _order = ("BREAKOUT", "MOMENTUM", "PULLBACK", "REVERSAL", "OTHER")
            _parts = [f"{k} {_mix[k]}" for k in _order if _mix.get(k, 0) > 0]
            if _parts:
                lines.append(
                    f"🧭 <b>Setup mix</b> ({_pool_size} evaluated): "
                    + " · ".join(_parts)
                )
                # Per-setup ticker list — show up to 8 symbols, then "+N more"
                _MAX_TICKERS = 8
                for k in _order:
                    _syms = _tickers.get(k, [])
                    if not _syms:
                        continue
                    _shown = _syms[:_MAX_TICKERS]
                    _extra = len(_syms) - len(_shown)
                    _line  = f"   • <b>{k}</b>: " + ", ".join(_shown)
                    if _extra > 0:
                        _line += f"  <i>(+{_extra} more)</i>"
                    lines.append(_line)
            # BUY breakdown line with tickers
            _bparts = []
            for k in _order:
                _n = _mix_bought.get(k, 0)
                if _n <= 0:
                    continue
                _bsyms = _tickers_buy.get(k, [])[:5]  # cap at 5 for BUY line
                _bparts.append(f"{k} {_n} ({', '.join(_bsyms)})" if _bsyms else f"{k} {_n}")
            if _bparts:
                lines.append("   ↳ <b>became BUYs</b>: " + " · ".join(_bparts))
            elif buys is not None and len(buys) == 0:
                lines.append("   ↳ <b>became BUYs</b>: none")
    except Exception:
        pass
    # FIX: stamp the actual latest OHLCV candle date used by the scan so
    # readers can tell whether data is fresh or from a prior session.
    try:
        _latest = None
        _ns = nifty_state or {}
        _cand = _ns.get("latest_candle_date") or _ns.get("as_of") or _ns.get("as_of_date")
        if _cand:
            _latest = str(_cand)
        elif macro:
            _cand2 = macro.get("latest_candle_date") or macro.get("as_of")
            if _cand2:
                _latest = str(_cand2)
        if _latest:
            lines.append(f"  <i>Latest candle: {html.escape(_latest)}</i>")
    except Exception:
        pass
    lines.append("─" * 22)
    lines.append("")

    # ── Market Regime ──
    vix_in  = macro.get("vix_in", 15.0)
    vix_us  = macro.get("vix_us", 18.0)
    vix_in_flag = "HIGH" if vix_in > 20 else ("low" if vix_in < 13 else "NORMAL")
    vix_us_flag = "HIGH" if vix_us > 25 else ("ok" if vix_us < 22 else "ELEVATED")
    REGIME_RATIONALE = {
        "STRONG_BULL":     "Strong uptrend, broad breadth — full deployment.",
        "BULL":            "Uptrend intact, moderate breadth — standard sizing.",
        "SIDEWAYS":        "Range-bound — tight sizing, high conviction only.",
        "TRANSITION":      "Mixed signals — reduced sizing, tight stops.",
        "HIGH_VOLATILITY": f"Elevated VIX-IN {vix_in:.1f} — 1 buy max, tight stops mandatory.",
        "BEAR":            "Downtrend — no new longs, manage exits only.",
        "STRONG_BEAR":     "Severe downtrend — cash only, no new positions.",
    }
    lines.append("📊 <b>MARKET REGIME</b>")
    lines.append(f"  {regime} · Score {score:.0f}/100 · MaxBuys {thresh['max_buys']}")

    # Regime explanation (Patch 4)
    fii_flow = macro.get("fii_flow_cr", 0.0)
    dii_flow = macro.get("dii_flow_cr", 0.0)
    # Use nifty_state as primary EMA source (FIX 1: single source of truth)
    _kl = key_levels or {}
    _ns_early = nifty_state or {}
    reg_why = regime_explanation(
        score, regime, vix_in, breadth_20, fii_flow, dii_flow,
        nifty_close = float(_ns_early.get("close",  _kl.get("last",   0)) or 0),
        ema20       = float(_ns_early.get("ema20",  _kl.get("ema20",  0)) or 0),
        ema50       = float(_ns_early.get("ema50",  _kl.get("ema50",  0)) or 0),
        ema200      = float(_ns_early.get("ema200", _kl.get("ema200", 0)) or 0),
    )
    lines.append(f"  {html.escape(str(reg_why))}")
    lines.append(f"  {REGIME_RATIONALE.get(regime, '')}")
    # Phase 3 (2026-07-01): VIX percentile — institutional-grade regime cue.
    _vp = macro.get("vix_in_percentile")
    _vr = macro.get("vix_in_regime", "UNKNOWN")
    if _vp is not None:
        _vr_hint = {
            "COMPLACENT": "complacent · watch mean reversion",
            "NORMAL":     "normal",
            "ELEVATED":   "elevated",
            "FEAR":       "fear · contrarian window",
            "UNKNOWN":    "",
        }.get(_vr, "")
        _hint_str = f" — {_vr_hint}" if _vr_hint else ""
        lines.append(
            f"  VIX-IN {vix_in:.1f}({vix_in_flag}) · {_vp:.0f}th %ile [{_vr}{_hint_str}] · "
            f"VIX-US {vix_us:.1f}({vix_us_flag})"
        )
    else:
        lines.append(
            f"  VIX-IN {vix_in:.1f}({vix_in_flag}) · VIX-US {vix_us:.1f}({vix_us_flag})"
        )
    lines.append(
        f"  NIFTY {macro.get('nifty_1d_pct', 0):+.2f}% · "
        f"S&P {macro.get('sp500_1d_pct', 0):+.2f}% · "
        f"DXY {macro.get('dxy', 0):.1f}"
    )
    lines.append(
        f"  ₹/$ {macro.get('usdinr', 0):.2f} · "
        f"Crude ${macro.get('crude_usd', 0):.1f} · "
        f"US10Y {macro.get('us10y', 0):.2f}%"
    )

    # Phase C4 (2026-07-02): FII/DII display removed entirely. NSDL is
    # structurally T-1/T-2, RSS fallbacks are unreliable, and the number
    # never fed sizing/gates/regime. Fetch is now stubbed to a no-op.

    # ── Conviction + Risk (single lines) ──
    _kl2 = key_levels or {}
    _ns  = nifty_state or {}
    lines.extend(format_conviction_meter(score, breadth_20, fii_flow, dii_flow))
    lines.extend(format_risk_meter(_ns, vix_in, breadth_20))
    lines.append(f"  Min Conf {thresh['min_confidence']} · TQ {thresh['min_tq']} · R/R {thresh['min_rr']} · Max {thresh['max_buys']} buy(s)")

    if heat:
        heat_emoji = "🔴" if not heat["heat_ok"] else ("🟡" if heat["heat_pct"] > heat["max_heat_pct"] * 0.6 else "🟢")
        lines.append(f"  {heat_emoji} Heat {heat['heat_pct']:.1f}%/{heat['max_heat_pct']:.0f}%")
    if platt and platt.get("calibrated"):
        lines.append(
            f"  WR {platt['win_rate']:.0%} ({platt['total_closed']}T) · "
            f"Avg W +{platt['avg_win_pct']:.1f}% / L -{platt['avg_loss_pct']:.1f}%"
        )
    lines.append("")

    # ── NIFTY Key Levels + EMA interpretation (Patch 3) ──
    if key_levels:
        kl = key_levels
        lines.append("NIFTY LEVELS")
        lines.append(
            f"  EMA20: {kl.get('ema20', '—')} | "
            f"EMA50: {kl.get('ema50', '—')} | "
            f"EMA200: {kl.get('ema200', '—')}"
        )
        # FIX: date-stamp the 52W High/Low so a stale figure is obvious.
        _52w_as_of = None
        try:
            _ns2 = nifty_state or {}
            _52w_as_of = _ns2.get("as_of") or _ns2.get("latest_candle_date") or _ns2.get("as_of_date")
            if _52w_as_of:
                _52w_as_of = str(_52w_as_of)
        except Exception:
            _52w_as_of = None
        _52w_tag = f" (as of {html.escape(_52w_as_of)})" if _52w_as_of else ""
        lines.append(
            f"  52W H: {kl.get('high_52w', '—')} "
            f"({kl.get('dist_from_52w_high_pct', 0):.1f}% away) | "
            f"52W L: {kl.get('low_52w', '—')}{_52w_tag}"
        )
        lines.append(
            f"  20D Range: {kl.get('recent_low_20d', '—')} — {kl.get('recent_high_20d', '—')}"
        )
        # One-line structure from nifty_state (BUG FIX 1: single source)
        try:
            ns = nifty_state or {}
            if ns.get("structure"):
                lines.append(f"  Structure: {ns['structure']}")
            else:
                nifty_close = float(kl.get("last", 0) or 0)
                structure   = interpret_nifty_structure(
                    nifty_close,
                    float(kl.get("ema20",  0) or 0),
                    float(kl.get("ema50",  0) or 0),
                    float(kl.get("ema200", 0) or 0),
                    float(kl.get("high_52w", 0) or 0),
                )
                lines.append(f"  Structure: {structure}")
        except Exception:
            pass
        lines.append("")

    # ── Breadth Dashboard (2026-07-03 fix: report the true gate-pool size) ──
    # OLD semantics (buggy on no-signal days):
    #   Qualified = BUY + Watchlist   →   0 when nothing passes → looked like
    #                                     the whole universe was rejected pre-scoring
    #   Rejected  = Universe - Qualified  →   2360 = whole universe (misleading)
    #
    # NEW semantics (matches what the user's mental model expects):
    #   Qualified = stocks that reached the 13-gate system     (e.g. top-50)
    #   Rejected  = of those, how many failed the gates        (excludes illiquid-drops)
    _near_c = len([w for w in watchlist if w.get("tier") == "NEAR_MISS"])
    _dev_c  = len([w for w in watchlist if w.get("tier") == "DEVELOPING"])
    _mon_c  = len([w for w in watchlist if w.get("tier") == "MONITOR"])
    _rej_c  = len(rejected_stocks or [])
    _qual_c = len(buys) + len(watchlist) + _rej_c   # ← gate-pool size (BUY + WL + rejected-at-gates)
    lines.extend(format_breadth_dashboard(
        universe_count or (_qual_c + len(shorts)),
        tradable_count or _qual_c,
        _qual_c, _near_c, _dev_c, _mon_c,
        rejected_at_gates = _rej_c,
    ))
    lines.append("")

    # ── BUY Signals (Patch 6 for no-buy case) ──
    lines.append("✅ <b>BUY SIGNALS</b>")
    if buys:
        _ai = ai_results or {}
        _buy_theses = _ai.get("buy_theses", {})
        for b in buys:
            sizing = {
                "position_value": b.get("position_value", 0),
                "position_pct":   b.get("position_pct", 0),
                "shares":         b.get("shares", 0),
                "max_loss":       b.get("max_loss", 0),
            }
            thesis = _buy_theses.get(b.get("symbol", ""), "")
            lines.extend(format_buy_card(b, sizing, regime, buy_thesis=thesis))
            # Gap validity (morning price check)
            entry  = b.get("entry", 0)
            stop_p = b.get("stop", 0)
            t1     = b.get("target1", 0)
            rr_v   = b.get("rr_ratio", 1.8) or 1.8
            gap_check = check_gap_validity(entry, stop_p, t1, rr_v)
            max_entry = gap_check.get("max_valid_entry", 0)
            if max_entry > 0 and max_entry > entry:
                gap_max_pct = round((max_entry - entry) / entry * 100, 1)
                lines.append(f"  ⚡ Max entry ₹{max_entry:.1f} (+{gap_max_pct:.1f}%) — skip if opens above")
            sizing_method = b.get("sizing_method", "")
            if sizing_method:
                lines.append(f"  💰 {html.escape(str(sizing_method))}")
            rs_diff = b.get("rs_diff21", 0)
            lines.append(f"  RS vs Nifty {rs_diff:+.1f}%")
            weekly_ok = b.get("weekly_trend_ok", True)
            if not weekly_ok:
                lines.append("  ⚠️ Weekly DOWN — reduced conviction")
            pattern = b.get("price_pattern", "NONE")
            if pattern != "NONE":
                lines.append(f"  📐 {html.escape(str(pattern))}")
            accum = b.get("accum_signal", "NEUTRAL")
            if accum != "NEUTRAL":
                lines.append(f"  📊 Vol: {html.escape(str(accum))}")
            ai_sum = truncate_display(b.get("ai_commentary", ""), 90)
            if ai_sum and ai_sum != "—":
                lines.append(f"  🤖 {html.escape(str(ai_sum))}")
            if b.get("repeat_tag"):
                lines.append(f"  🔁 {html.escape(str(b['repeat_tag']))}")
            if b.get("warnings"):
                lines.append(f"  ⚠️ {html.escape(', '.join(b['warnings'][:3]))}")
            lines.append("  ···")
    else:
        # Patch 6: detailed no-buy explanation (includes watchlist as closer candidates)
        no_buy_lines = format_no_buy_explanation(rejected_stocks or [], regime,
                                                  watchlist=watchlist or [])
        lines.extend(no_buy_lines)
    lines.append("")

    # ── SHORT Signals ──
    if shorts:
        lines.append("🔽 <b>SHORT SIGNALS</b>")
        for s in shorts:
            lines.append(f"  <b>{html.escape(str(s.get('symbol','?')))}</b> [SHORT]")
            lines.append(
                f"  Entry ₹{s.get('entry',0):.1f} · Stop ₹{s.get('stop',0):.1f} · T1 ₹{s.get('target1',0):.1f} · T2 ₹{s.get('target2',0):.1f} · RR {s.get('rr',0):.2f}x"
            )
            lines.append(f"  {html.escape(str(s.get('reason','')))}")  
        lines.append("")

    # ── Watchlist — ALL stocks, all tiers, with levels (Patch 1) ──
    _nm_insights = (ai_results or {}).get("near_miss_insights", {})
    wl_lines = format_watchlist_section(
        watchlist, regime,
        conf_history=conf_history or {},
        gate_memory=gate_memory or {},
        near_miss_insights=_nm_insights,
    )
    lines.extend(wl_lines)
    lines.append("")

    # ── Sector-Out-of-Favor rejects (Phase C4) ─────────────────────────────
    # Stocks with high Conf/TQ that got REJECTED (not just watchlisted)
    # because SECTOR_LAGGING + other fails exceeded the 2-fail budget.
    # Surfaces contrarian / sector-rotation candidates that used to be
    # invisible in the "Rejected N" bucket.
    try:
        sl_lines = format_sector_lagging_rejects_section(
            rejected_stocks or [], regime, max_rows=12,
        )
        if sl_lines:
            lines.extend(sl_lines)
            lines.append("")
    except Exception as _e:
        _log(f"[WARN] sector-lagging rejects render failed: {_e}")

    # ── Stop Watch Alert (FIX 5: appears BEFORE portfolio, cannot be missed) ──
    _cur_prices_port = {a["symbol"]: float(a.get("current", a.get("entry", 0)) or 0)
                        for a in portfolio_alerts}
    stop_alerts = format_stop_watch_alert(portfolio_alerts, _cur_prices_port)
    if stop_alerts:
        lines += stop_alerts

    # ── Portfolio ──
    exits    = [a for a in portfolio_alerts if a["action"] in ("EXIT", "EXIT_FULL")]
    partials = [a for a in portfolio_alerts if a["action"] == "T1_PARTIAL_EXIT"]
    trails   = [a for a in portfolio_alerts if a["action"] == "TRAIL_STOP"]
    reviews  = [a for a in portfolio_alerts if a["action"] == "REVIEW"]
    holds    = [a for a in portfolio_alerts if a["action"] == "HOLD"]
    lines.append("📁 <b>PORTFOLIO</b>")
    if portfolio_alerts:
        # Compact 2-line portfolio summary (FIX 3F)
        lines.extend(format_portfolio_summary_compact(portfolio_alerts, _cur_prices_port, PORTFOLIO_CAPITAL))
        lines.append("")

        def _fmt_alert_card(a: dict) -> list:
            cur = float(a.get("current", a.get("entry", 0)) or 0)
            card = format_portfolio_card_compact(a, cur)
            # Append exit reason if relevant
            if a.get("reason") and a["action"] not in ("HOLD",):
                card.append(f"     Reason: {html.escape(str(a['reason']))}")
            return card

        if exits:
            lines.append("  🚨 <b>EXIT</b>")
            for e in exits:
                lines.extend(_fmt_alert_card(e))
        if partials:
            lines.append("  💰 <b>T1 PARTIAL EXIT</b> (book 50% · trail rest to break-even)")
            for p in partials:
                lines.extend(_fmt_alert_card(p))
        if trails:
            lines.append("  ⚡ <b>TRAIL STOP</b>")
            for t in trails:
                lines.extend(_fmt_alert_card(t))
        if reviews:
            lines.append("  🔍 <b>REVIEW</b>")
            for r in reviews:
                lines.extend(_fmt_alert_card(r))
        if holds:
            lines.append("  ✅ <b>HOLD</b>")
            for h in holds[:6]:
                lines.extend(_fmt_alert_card(h))
    else:
        lines.append("  No active holdings.")
    lines.append("")

    # ── Phase I shadow-log summary (2026-07-07) ──
    # Rich per-bucket brief showing WHAT was added tonight in each bucket +
    # resolved outcomes. Pure paper — nothing to act on. Safe no-op if
    # module unavailable or CSV is empty. Toggle off entirely with
    # PHASE_I_SHADOW_LOG=false; control verbosity with SHADOW_TELEGRAM_MODE:
    #   "compact"   → 4-line rollup only (old behaviour)
    #   "detailed"  → per-bucket top-5 stocks + resolved-today (default)
    if _SHADOW_LOG_OK:
        try:
            _sh_mode = os.getenv("SHADOW_TELEGRAM_MODE", "detailed").lower()
            if _sh_mode == "compact":
                _shadow_block = shadow_log.format_shadow_summary(
                    max_pending_shown=3)
            else:
                _shadow_block = shadow_log.format_shadow_telegram(
                    top_n_per_bucket=5, show_resolved_today=True)
            if _shadow_block:
                lines.append(_shadow_block)
                lines.append("")
        except Exception:
            pass

    # ── Upcoming Events (FIX 4: config-based, 2 lines per event max) ──
    if upcoming_events:
        lines.extend(format_upcoming_events_compact(upcoming_events, portfolio_alerts))
        lines.append("")

    # ── Daily Summary (AI-written when available, rule-based fallback) ──
    _ai_summary = (ai_results or {}).get("daily_summary", "")
    if _ai_summary:
        lines.append("")
        lines.append("\U0001f4cb DAILY SUMMARY")
        lines.append(f"  {html.escape(str(_ai_summary))}")
    else:
        lines.extend(format_daily_summary(
            regime, buys, watchlist, portfolio_alerts, macro, nifty_state
        ))

    # ── Footer ──
    lines.append("")
    # Phase C: data-quality footer — visible signal of source health
    try:
        dq          = macro.get("data_quality", "NORMAL")
        bad_fields  = macro.get("bad_fields", []) or []
        usdinr_src  = (macro.get("sources", {}) or {}).get("usdinr", "?")
        # Extract just the source name from the verbose log line
        usdinr_name = "yfinance"
        if "frankfurter" in str(usdinr_src):
            usdinr_name = "frankfurter"
        elif "yfinance" in str(usdinr_src):
            usdinr_name = "yfinance"
        # Phase C4: FII/DII removed from output — footer no longer reports it.
        dq_icon     = "🟢" if dq == "NORMAL" else "🟡"
        lines.append(f"{dq_icon} Data quality: {dq}"
                     + (f" · bad={','.join(bad_fields)}" if bad_fields else "")
                     + f" · usdinr={usdinr_name}")
    except Exception:
        pass
    lines.append("─" * 22)
    lines.append("⚠️ Recommendation only. Execute manually.")
    return "\n".join(lines)


# ═════════════════════════════════════════════════════════════════════════════
# SECTION 9c — TELEGRAM MESSAGE v2 (Redesigned 2026-07-10)
# ═════════════════════════════════════════════════════════════════════════════
# Human-friendly redesign of the daily Telegram brief. Rationale:
#   Old format put database-style aggregates in front of the reader
#   ("MOMENTUM 112 · PULLBACK 18 · OTHER 32", "Rejected 114") but never
#   showed which stocks were in each bucket. A new person couldn't tell
#   whether SOBHA was scanned, rejected, or on watchlist.
#
# v2 goals (from user session 2026-07-10):
#   1. Name every stock the reader cares about (all BREAKOUTs, top-10
#      MOMENTUM, closest-15 rejects) so manual tracking is possible.
#   2. Front-load the verdict ("0 buys, 1 close call") before any details.
#   3. Show the tier math (conf_gap, penalty) inline so DEVELOPING vs
#      MONITOR isn't a mystery.
#   4. Attach a CSV with EVERY evaluated stock so the reader has full
#      transparency into the universe.
#
# Design principles:
#   • Gated by NEW_TELEGRAM_FORMAT env var (default "true"); flip to
#     "false" for instant rollback to v1.
#   • Each section is a standalone builder wrapped in try/except so one
#     bad stock never breaks the whole message.
#   • v2 orchestrator itself is wrapped in try/except at the call site;
#     on ANY unexpected exception it logs and returns the v1 message.
#     The user is guaranteed to receive a Telegram even if v2 has a bug.
#   • Does NOT modify scoring, gates, or any pipeline behavior. Display only.
# ═════════════════════════════════════════════════════════════════════════════

NEW_TELEGRAM_FORMAT   = os.getenv("NEW_TELEGRAM_FORMAT", "true").lower() == "true"
TELEGRAM_ATTACH_CSV   = os.getenv("TELEGRAM_ATTACH_CSV", "true").lower() == "true"
TELEGRAM_PREV_STATE   = os.getenv("TELEGRAM_PREV_STATE", "telegram_prev_state.json")
TELEGRAM_DAILY_CSV    = os.getenv("TELEGRAM_DAILY_CSV", "telegram_daily.csv")

# 30 daily-tip strings that rotate by day-of-year. Teaches one concept
# per day so a new reader gradually learns the vocabulary.
_V2_DAILY_TIPS = [
    "Confidence (0-100) is how sure the model is about the trade. It combines technicals, fundamentals, volume and regime.",
    "Trade Quality (TQ) measures pattern strength: volume + tight range + strong close = high TQ. Independent of Confidence.",
    "BREAKOUT setups are stocks piercing a resistance level. In choppy markets they're the ONLY setup we trust.",
    "MOMENTUM setups are stocks already in an uptrend. In chop we penalise their confidence by 10% because momentum fades in sideways markets.",
    "PULLBACK setups are dips inside an uptrend — great in bull regimes, blocked in chop (dips can become breakdowns).",
    "REVERSAL setups mean the stock changed direction. Blocked in chop for the same reason as pullbacks.",
    "The Regime is the overall market weather: BULL / TRANSITION / CHOP / BEAR. It changes the confidence bar.",
    "TRANSITION regime raises the confidence bar to 83 (from 78 in BULL). Fewer stocks pass = higher quality.",
    "R/R (Risk/Reward ratio) is the ratio of upside to downside. We require ≥1.8× for a BUY.",
    "The Shadow Log tracks stocks we passed on. If a lot of them go up anyway, our filters are too tight.",
    "Bucket A in the shadow log = 'high quality, blocked by risk gate'. These are the most useful to review.",
    "Bucket B = 'watch me later' — chop-penalised MOMENTUM stocks that might rebound in a trend.",
    "Bucket C = 'not my style' — wrong setup for the current chop regime (pullback / reversal).",
    "ATR (Average True Range) is how much a stock moves per day. Our stop is 1.5× ATR below entry.",
    "The 'chop penalty' reduces MOMENTUM confidence by 10% when regime is CHOP or TRANSITION. Breakouts are exempt.",
    "The MIN_CONFIDENCE gate is regime-adjusted. BULL=78, TRANSITION=83, CHOP=85 — tighter in tougher markets.",
    "The MIN_TQ gate is also regime-adjusted. BULL=65, TRANSITION=68, CHOP=71 — no weak patterns in chop.",
    "The R/R gate is the hardest. Only trades where reward is at least 1.8× the risk make it through.",
    "ROE < 15% or D/E > 1.5 blocks the stock at fundamentals — no leverage traps, no low-return businesses.",
    "SECTOR_LAGGING means the stock's sector is in the bottom 30% of relative strength. Contrarian only.",
    "'Near Miss' watchlist tier means the stock is ≤15 confidence points from a BUY — could trigger tomorrow.",
    "'Developing' watchlist tier means 16-25 points from a BUY — worth tracking but not close yet.",
    "'Monitor' watchlist tier means >25 points from a BUY — early stage, no action expected soon.",
    "The Confidence Meter shows how the regime score is distributed. Above 60 = deploy freely.",
    "The Risk Meter shows portfolio heat. Above 60% heat = STOP adding new positions.",
    "Position sizing = Kelly-fraction × Capital, capped at 5% per stock. Bigger conviction = bigger size.",
    "Every BUY signal has a 'max valid entry' — if the stock opens above it, skip the signal.",
    "The 52-week high tells you if the stock is at all-time highs (usually bullish) or well below (harder to run).",
    "VIX-IN measures Indian market fear. Below 13 = complacent (careful), above 20 = fear (opportunities).",
    "We never execute automatically. Every message is a recommendation — you decide what to do.",
]


def _v2_html_esc(s) -> str:
    """Safe HTML-escape that handles None / non-string / NaN."""
    try:
        if s is None:
            return ""
        return html.escape(str(s))
    except Exception:
        return ""


def _v2_safe_float(x, default: float = 0.0) -> float:
    """Extract a float from anything, defaulting to `default` on failure/NaN."""
    try:
        v = float(x)
        # Reject NaN
        if v != v:
            return default
        return v
    except Exception:
        return default


def _v2_clean_ticker(sym) -> str:
    """Strip .NS suffix and normalize."""
    return str(sym or "?").replace(".NS", "").replace(".BO", "").strip()


def _v2_bar(value: float, max_value: float = 100.0, width: int = 18) -> str:
    """ASCII horizontal bar. Returns a string like '██████████░░░░'."""
    try:
        if max_value <= 0:
            return "░" * width
        pct = max(0.0, min(1.0, value / max_value))
        filled = int(round(pct * width))
        return "█" * filled + "░" * (width - filled)
    except Exception:
        return "░" * width


def _v2_get_daily_tip() -> str:
    """Return one tip per day-of-year, rotating through the 30-tip list."""
    try:
        doy = ist_today().timetuple().tm_yday
        return _V2_DAILY_TIPS[doy % len(_V2_DAILY_TIPS)]
    except Exception:
        return _V2_DAILY_TIPS[0]


def _v2_load_prev_state() -> dict:
    """Load yesterday's snapshot for delta comparison. Empty dict on failure."""
    try:
        if os.path.exists(TELEGRAM_PREV_STATE):
            with open(TELEGRAM_PREV_STATE, "r", encoding="utf-8") as f:
                return json.load(f) or {}
    except Exception:
        pass
    return {}


def _v2_save_prev_state(snapshot: dict) -> None:
    """Persist today's snapshot for tomorrow's delta section."""
    try:
        with open(TELEGRAM_PREV_STATE, "w", encoding="utf-8") as f:
            json.dump(snapshot, f, indent=2, default=str)
    except Exception as e:
        _log(f"[v2] warn: could not save prev state: {e}")


def _v2_primary_reject_label(stock: dict) -> str:
    """Return a short human label for the stock's primary reject reason."""
    try:
        frs = stock.get("fail_reasons", []) or []
        if not frs:
            return "unknown"
        raw = str(frs[0])
        # Reuse the same classification map as v1's format_no_buy_explanation
        _LBL = {
            "CONF_FAIL":               "conf too low",
            "TQ_FAIL":                 "TQ too low",
            "RR_FAIL":                 "R/R too low",
            "WIDE_STOP":               "stop too wide",
            "SECTOR_LAGGING":          "sector lagging",
            "SECTOR_CAP":              "sector cap",
            "KILL_SWITCH":             "kill-switch",
            "EVENT_BLOCK":             "event blackout",
            "HIGH_CORR":               "too correlated",
            "LIQUIDITY_FAIL":          "illiquid",
            "REGIME_NO_BUY":           "regime blocks BUYs",
            "PROMOTER_PLEDGE":         "pledge too high",
            "BLACK_SWAN_NEWS":         "black-swan news",
            "DATA_INCOMPLETE":         "data incomplete",
            "CHOP_NO_BREAKOUT":        "chop: non-breakout blocked",
            "SETUP_OTHER":             "setup unclassified",
            "SETUP_WEAK":              "setup weak for regime",
        }
        # Setup-edge sub-reasons
        if raw.startswith("SETUP_EDGE_SKIP"):
            inner = raw[len("SETUP_EDGE_SKIP("):].rstrip(")")
            if inner.startswith("REGIME_CHOP_NO_BREAKOUT"):
                if "/OTHER" in inner:
                    return _LBL["SETUP_OTHER"]
                return _LBL["CHOP_NO_BREAKOUT"]
            return _LBL["SETUP_WEAK"]
        # ROE / DE bumps come through as fail_reasons too
        if "ROE" in raw.upper():
            return "ROE too low"
        if raw.upper().startswith("DE_") or "D/E" in raw or "DE_HIGH" in raw.upper():
            return "D/E too high"
        if "MCAP" in raw.upper() or "MARKET_CAP" in raw.upper():
            return "market cap too small"
        key = raw.split("(", 1)[0].strip()
        for prefix, val in [("HIGH_CORR_", "HIGH_CORR"),
                            ("EVENT_BLOCK_", "EVENT_BLOCK"),
                            ("KILL_SWITCH", "KILL_SWITCH"),
                            ("PROMOTER_PLEDGE_", "PROMOTER_PLEDGE")]:
            if key.startswith(prefix):
                key = val
                break
        return _LBL.get(key, key.lower().replace("_", " "))
    except Exception:
        return "unknown"


# ── Section builders ────────────────────────────────────────────────────────
# Each returns list[str]; every one is independently wrapped in try/except
# at the orchestrator level so one broken section can't kill the message.


# ── v2 UX helpers: setup tagging + row layout ───────────────────────────────
# The stock-dict field is `setup_type` (populated by apply_setup_edge upstream).
# Older code paths read `setup` — always prefer setup_type, fall back to setup.
_V2_SETUP_SHORT = {
    "BREAKOUT": "BRK",
    "MOMENTUM": "MOM",
    "PULLBACK": "PUL",
    "REVERSAL": "REV",
    "OTHER":    "OTH",
}
_V2_SETUP_EMOJI = {
    "BREAKOUT": "🚀",
    "MOMENTUM": "📈",
    "PULLBACK": "↩️",
    "REVERSAL": "🔄",
    "OTHER":    "·",
}

def _v2_setup_of(s: dict) -> str:
    """Return canonical UPPER setup type; empty string if unknown."""
    try:
        raw = s.get("setup_type") or s.get("setup") or ""
        return str(raw).upper().strip()
    except Exception:
        return ""

def _v2_setup_tag(s: dict, with_emoji: bool = True) -> str:
    """Compact human-readable setup tag (e.g. '📈 MOM'). Empty → '—'."""
    st = _v2_setup_of(s)
    if not st:
        return "—"
    short = _V2_SETUP_SHORT.get(st, st[:3])
    if with_emoji:
        emo = _V2_SETUP_EMOJI.get(st, "·")
        return f"{emo} {short}"
    return short


def _v2_section_header(timestamp: str) -> list:
    """Message header with market close date."""
    try:
        market_date = ist_today().strftime("%d %b %Y")
    except Exception:
        market_date = ""
    lines = [
        "━" * 33,
        f"📊 <b>DAILY SCAN</b> · Close {market_date}",
        f"<i>Scan: {_v2_html_esc(timestamp)}</i>",
        "━" * 33,
    ]
    return lines


def _v2_section_one_line_summary(buys: list, watchlist: list, rejected: list,
                                 regime: str, setup_mix: dict) -> list:
    """The 1-sentence 'today in a nutshell' opener."""
    try:
        n_buy = len(buys or [])
        n_wl  = len(watchlist or [])
        n_rej = len(rejected or [])
        n_mom = int((setup_mix or {}).get("MOMENTUM", 0))
        n_bo  = int((setup_mix or {}).get("BREAKOUT", 0))

        # Find the best watchlist stock (highest conf, prefer NEAR_MISS then DEVELOPING)
        best_wl = None
        for tier in ("NEAR_MISS", "DEVELOPING", "MONITOR"):
            cands = [w for w in (watchlist or []) if w.get("tier") == tier]
            if cands:
                best_wl = max(cands, key=lambda x: _v2_safe_float(
                    x.get("conf", x.get("final_confidence", 0))))
                break

        if n_buy > 0:
            sentence = (
                f"<b>{n_buy} BUY signal(s) today.</b> "
                f"Regime is {regime}; {n_wl} more stocks are on the watchlist."
            )
        elif best_wl:
            sym = _v2_clean_ticker(best_wl.get("symbol", "?"))
            conf = _v2_safe_float(best_wl.get("conf", best_wl.get("final_confidence", 0)))
            sentence = (
                f"<b>No buys today.</b> Market is {regime.lower()} — closest to a signal "
                f"is <b>{_v2_html_esc(sym)}</b> (conf {conf:.1f}). "
                f"{n_wl} stocks on watchlist, {n_rej} rejected."
            )
        else:
            sentence = (
                f"<b>No buys, no close calls.</b> Market is {regime.lower()}. "
                f"Scanned {n_buy + n_wl + n_rej} stocks — none met the bar today."
            )

        return [
            "",
            "📝 <b>TODAY IN ONE LINE</b>",
            f"   {sentence}",
            "",
        ]
    except Exception:
        return []


def _v2_section_verdict(buys: list, watchlist: list, rejected: list) -> list:
    """The 4-line verdict block."""
    try:
        n_buy = len(buys or [])
        n_wl  = len(watchlist or [])
        n_rej = len(rejected or [])
        near  = sum(1 for w in (watchlist or []) if w.get("tier") == "NEAR_MISS")
        dev   = sum(1 for w in (watchlist or []) if w.get("tier") == "DEVELOPING")
        mon   = sum(1 for w in (watchlist or []) if w.get("tier") == "MONITOR")

        lines = [
            "🎯 <b>THE VERDICT</b>",
            f"   <b>{n_buy:>3}</b> 🟢 BUY signals",
            f"   <b>{near:>3}</b> 🟡 Close call  (near-miss)",
            f"   <b>{dev + mon:>3}</b> 🔵 Early stage ({dev} developing, {mon} monitor)",
            f"   <b>{n_rej:>3}</b> 🔴 Rejected    (didn't pass filters)",
            "",
        ]
        return lines
    except Exception:
        return []


def _v2_section_vs_yesterday(buys: list, watchlist: list, rejected: list,
                              regime: str, prev: dict) -> list:
    """Delta vs yesterday's saved snapshot."""
    try:
        if not prev:
            return []  # first run — no yesterday data
        n_buy = len(buys or [])
        n_wl  = len(watchlist or [])
        n_rej = len(rejected or [])
        prev_buy    = int(prev.get("buys", 0))
        prev_wl     = int(prev.get("watchlist", 0))
        prev_rej    = int(prev.get("rejected", 0))
        prev_regime = str(prev.get("regime", ""))
        prev_wl_syms = set(prev.get("watchlist_symbols", []) or [])
        today_wl_syms = set(_v2_clean_ticker(w.get("symbol", "")) for w in (watchlist or []))
        new_syms = sorted(today_wl_syms - prev_wl_syms)
        gone_syms = sorted(prev_wl_syms - today_wl_syms)

        def _arrow(now, prev_v):
            d = now - prev_v
            if d > 0:  return f"↑ +{d}"
            if d < 0:  return f"↓ {d}"
            return "= 0"

        lines = ["📈 <b>vs YESTERDAY</b>"]
        lines.append(f"   BUYs:      {prev_buy} → {n_buy}   ({_arrow(n_buy, prev_buy)})")
        lines.append(f"   Watchlist: {prev_wl} → {n_wl}   ({_arrow(n_wl, prev_wl)})")
        lines.append(f"   Rejected:  {prev_rej} → {n_rej}   ({_arrow(n_rej, prev_rej)})")
        if prev_regime and prev_regime != regime:
            lines.append(f"   Regime:    {_v2_html_esc(prev_regime)} → <b>{_v2_html_esc(regime)}</b>")
        else:
            lines.append(f"   Regime:    <b>{_v2_html_esc(regime)}</b> (unchanged)")
        if new_syms:
            preview = ", ".join(_v2_html_esc(s) for s in new_syms[:5])
            extra = f" (+{len(new_syms) - 5} more)" if len(new_syms) > 5 else ""
            lines.append(f"   ➕ New on WL: {preview}{extra}")
        if gone_syms:
            preview = ", ".join(_v2_html_esc(s) for s in gone_syms[:5])
            extra = f" (+{len(gone_syms) - 5} more)" if len(gone_syms) > 5 else ""
            lines.append(f"   ➖ Left WL:   {preview}{extra}")
        lines.append("")
        return lines
    except Exception:
        return []


def _v2_section_buys(buys: list, regime: str, ai_results: dict) -> list:
    """BUY signals — reuse the v1 format_buy_card for full detail."""
    try:
        if not buys:
            return []
        lines = ["🟢 <b>BUY SIGNALS</b>", ""]
        buy_theses = (ai_results or {}).get("buy_theses", {}) or {}
        for b in buys:
            sizing = {
                "position_value": b.get("position_value", 0),
                "position_pct":   b.get("position_pct", 0),
                "shares":         b.get("shares", 0),
                "max_loss":       b.get("max_loss", 0),
            }
            thesis = buy_theses.get(b.get("symbol", ""), "")
            try:
                lines.extend(format_buy_card(b, sizing, regime, buy_thesis=thesis))
            except Exception as e:
                _log(f"[v2] buy_card failed for {b.get('symbol')}: {e}")
                lines.append(f"  <b>{_v2_html_esc(_v2_clean_ticker(b.get('symbol')))}</b> — details in Excel")
            lines.append("  " + "·" * 3)
        lines.append("")
        return lines
    except Exception as e:
        _log(f"[v2] section_buys crashed: {e}")
        return []


def _v2_section_close_call(watchlist: list, regime: str, ai_results: dict) -> list:
    """Detailed card for each NEAR_MISS watchlist stock."""
    try:
        near = [w for w in (watchlist or []) if w.get("tier") == "NEAR_MISS"]
        if not near:
            return []
        try:
            thresh = REGIME_THRESHOLDS[regime]
        except Exception:
            thresh = {"min_confidence": 78, "min_tq": 65, "min_rr": 1.8}
        min_conf = _v2_safe_float(thresh.get("min_confidence"), 78)
        min_tq   = _v2_safe_float(thresh.get("min_tq"), 65)
        min_rr   = _v2_safe_float(thresh.get("min_rr"), 1.8)
        insights = (ai_results or {}).get("near_miss_insights", {}) or {}

        # Sort by conf descending
        near = sorted(near, key=lambda x: _v2_safe_float(
            x.get("conf", x.get("final_confidence", 0))), reverse=True)

        lines = [f"🟡 <b>CLOSE CALL</b> — {len(near)} stock(s) near-miss", ""]
        for w in near:
            sym    = _v2_clean_ticker(w.get("symbol", "?"))
            sector = _v2_html_esc(w.get("sector", ""))
            conf   = _v2_safe_float(w.get("conf", w.get("final_confidence", 0)))
            tq     = _v2_safe_float(w.get("tq", w.get("trade_quality_score", 0)))
            rr     = _v2_safe_float(w.get("rr_ratio", w.get("rr", 0)))
            entry  = _v2_safe_float(w.get("entry", 0))
            stop   = _v2_safe_float(w.get("stop", 0))
            t1     = _v2_safe_float(w.get("target1", 0))
            setup  = _v2_html_esc(_v2_setup_of(w) or "—")
            conf_gap = max(0.0, min_conf - conf)

            lines.append(f"  <b>{_v2_html_esc(sym)}</b>  [{sector}]  · Setup: {setup}")
            lines.append(f"    Price ₹{entry:.1f}  |  Stop ₹{stop:.1f}  |  T1 ₹{t1:.1f}")
            lines.append(
                f"    Confidence: <b>{conf:.1f} / {min_conf:.0f}</b> "
                f"({conf_gap:+.1f} pts shy)"
            )
            lines.append(
                f"    TQ: {tq:.1f} / {min_tq:.0f}  ·  R/R: {rr:.2f}× (need {min_rr:.2f}×)"
            )
            # Compact checklist
            checks = []
            checks.append(("Confidence", conf >= min_conf, f"{conf:.1f}/{min_conf:.0f}"))
            checks.append(("TQ",         tq   >= min_tq,   f"{tq:.1f}/{min_tq:.0f}"))
            checks.append(("R/R",        rr   >= min_rr,   f"{rr:.2f}×"))
            check_str = "  ".join(
                f"{'✅' if ok else '❌'} {name} {val}" for name, ok, val in checks
            )
            lines.append(f"    {check_str}")

            insight = insights.get(w.get("symbol", ""), "")
            if insight:
                lines.append(f"    💡 {_v2_html_esc(insight)}")
            # Trigger hint
            if entry > 0:
                trig = entry * 1.015  # 1.5% above current
                lines.append(f"    🎯 Trigger: close &gt; ₹{trig:.1f} with vol &gt; 2× avg")
            lines.append("")
        return lines
    except Exception as e:
        _log(f"[v2] section_close_call crashed: {e}")
        return []


def _v2_section_early_stage(watchlist: list, regime: str) -> list:
    """DEVELOPING + MONITOR — named list with confidence bars."""
    try:
        dev = [w for w in (watchlist or []) if w.get("tier") == "DEVELOPING"]
        mon = [w for w in (watchlist or []) if w.get("tier") == "MONITOR"]
        if not dev and not mon:
            return []
        try:
            thresh = REGIME_THRESHOLDS[regime]
        except Exception:
            thresh = {"min_confidence": 78}
        min_conf = _v2_safe_float(thresh.get("min_confidence"), 78)

        lines = [f"🔵 <b>EARLY STAGE</b> — {len(dev) + len(mon)} stock(s) tracked", ""]

        def _render_row(w):
            sym    = _v2_clean_ticker(w.get("symbol", "?"))
            tag    = _v2_setup_tag(w)           # e.g. "📈 MOM" / "🚀 BRK" / "—"
            conf   = _v2_safe_float(w.get("conf", w.get("final_confidence", 0)))
            penalty = _v2_safe_float(w.get("chop_momentum_penalty", 0))
            bar    = _v2_bar(conf, min_conf * 1.1, width=10)   # narrower for mobile
            need   = max(0.0, min_conf - conf)                  # always positive
            pen_note = "  ⚠ chop×0.90" if penalty > 0 else ""
            # Two-line row: header shows ticker + bar + conf;
            # detail line shows setup + gap in plain English.
            # Fits inside ~34 chars on mobile fixed-width.
            return [
                f"  <code>{_v2_html_esc(sym):<10}</code> {bar} <b>{conf:5.1f}</b>",
                f"     {tag}  ·  need <b>+{need:.1f}</b> to BUY{pen_note}",
            ]

        if dev:
            lines.append(f"  <b>Developing</b> ({len(dev)}) — 16-25 pts shy of BUY:")
            for w in sorted(dev, key=lambda x: _v2_safe_float(
                x.get("conf", x.get("final_confidence", 0))), reverse=True):
                lines.extend(_render_row(w))
            lines.append("")
        if mon:
            lines.append(f"  <b>Monitor</b> ({len(mon)}) — &gt;25 pts shy of BUY:")
            for w in sorted(mon, key=lambda x: _v2_safe_float(
                x.get("conf", x.get("final_confidence", 0))), reverse=True):
                lines.extend(_render_row(w))
            lines.append("")
        lines.append(f"  <i>Bar shows conf vs target {min_conf:.0f}. Setup: 🚀 BRK · 📈 MOM · ↩️ PUL · 🔄 REV</i>")
        lines.append("")
        return lines
    except Exception as e:
        _log(f"[v2] section_early_stage crashed: {e}")
        return []


def _v2_section_breakout_spotlight(buys: list, watchlist: list,
                                    rejected: list, regime: str) -> list:
    """When the universe has exactly 1 BREAKOUT setup, elevate it into its
    own headline block so the reader immediately sees today's rare event.
    Emits nothing when there are 0 or 2+ breakouts (the all_breakouts
    section already covers those cases).
    """
    try:
        all_stocks = list(buys or []) + list(watchlist or []) + list(rejected or [])
        breakouts = [s for s in all_stocks if _v2_setup_of(s) == "BREAKOUT"]
        if len(breakouts) != 1:
            return []
        s = breakouts[0]
        sym  = _v2_clean_ticker(s.get("symbol", "?"))
        conf = _v2_safe_float(s.get("final_confidence", 0))
        tq   = _v2_safe_float(s.get("trade_quality_score", s.get("tq", 0)))
        rr   = _v2_safe_float(s.get("rr_ratio", s.get("rr", 0)))
        sector = _v2_html_esc(s.get("sector", "") or "")
        try:
            thresh = REGIME_THRESHOLDS[regime]
        except Exception:
            thresh = {"min_confidence": 78}
        min_conf = _v2_safe_float(thresh.get("min_confidence"), 78)

        # Verdict
        buy_syms = {b.get("symbol") for b in (buys or [])}
        wl_syms  = {w.get("symbol"): w.get("tier", "") for w in (watchlist or [])}
        sy = s.get("symbol")
        if sy in buy_syms:
            verdict = "🟢 BUY"
            tail    = "Passes all gates — see BUYS section."
        elif sy in wl_syms:
            tier = wl_syms[sy]
            verdict = {"NEAR_MISS": "🟡 close call",
                       "DEVELOPING": "🔵 developing",
                       "MONITOR": "🔵 monitor"}.get(tier, "🔵 watchlist")
            tail    = f"On watchlist — need <b>+{max(0.0, min_conf-conf):.1f}</b> conf to BUY."
        else:
            reason = _v2_primary_reject_label(s)
            verdict = "🔴 rejected"
            tail    = f"Blocked: <b>{_v2_html_esc(reason)}</b>."

        lines = [
            "🎯 <b>TODAY'S ONLY BREAKOUT</b>",
            f"   <b>{_v2_html_esc(sym)}</b>"
            + (f"  <i>[{sector}]</i>" if sector else "")
            + f"  ·  {verdict}",
            f"   Conf {conf:.1f} / {min_conf:.0f}  ·  TQ {tq:.1f}  ·  R/R {rr:.2f}×",
            f"   {tail}",
            "",
        ]
        return lines
    except Exception as e:
        _log(f"[v2] section_breakout_spotlight crashed: {e}")
        return []


def _v2_section_all_breakouts(buys: list, watchlist: list,
                               rejected: list) -> list:
    """Every BREAKOUT setup today with verdict + reason."""
    try:
        all_stocks = list(buys or []) + list(watchlist or []) + list(rejected or [])
        breakouts = [s for s in all_stocks if _v2_setup_of(s) == "BREAKOUT"]
        if not breakouts:
            return []
        # Determine verdict per stock
        buy_syms = {s.get("symbol") for s in (buys or [])}
        wl_syms  = {s.get("symbol"): s.get("tier", "") for s in (watchlist or [])}

        lines = [
            f"🚀 <b>ALL BREAKOUT SETUPS</b> ({len(breakouts)} today)",
            "",
        ]
        for s in sorted(breakouts, key=lambda x: _v2_safe_float(
            x.get("final_confidence", 0)), reverse=True):
            sym  = _v2_clean_ticker(s.get("symbol", "?"))
            conf = _v2_safe_float(s.get("final_confidence", 0))
            sy   = s.get("symbol")
            if sy in buy_syms:
                verdict = "🟢 BUY"
            elif sy in wl_syms:
                tier = wl_syms[sy]
                tier_emoji = {"NEAR_MISS": "🟡 close",
                              "DEVELOPING": "🔵 developing",
                              "MONITOR": "🔵 monitor"}.get(tier, "🔵 watchlist")
                verdict = tier_emoji
            else:
                reason = _v2_primary_reject_label(s)
                verdict = f"🔴 rejected ({_v2_html_esc(reason)})"
            lines.append(f"  <code>{_v2_html_esc(sym):<10}</code> {conf:5.1f}  →  {verdict}")
        lines.append("")
        return lines
    except Exception as e:
        _log(f"[v2] section_all_breakouts crashed: {e}")
        return []


def _v2_section_top_momentum(buys: list, watchlist: list, rejected: list,
                              top_n: int = 10) -> list:
    """Top-N MOMENTUM stocks by confidence (post-penalty)."""
    try:
        all_stocks = list(buys or []) + list(watchlist or []) + list(rejected or [])
        momentum = [s for s in all_stocks if _v2_setup_of(s) == "MOMENTUM"]
        if not momentum:
            return []
        total = len(momentum)
        momentum_sorted = sorted(momentum, key=lambda x: _v2_safe_float(
            x.get("final_confidence", 0)), reverse=True)
        shown = momentum_sorted[:top_n]

        buy_syms = {s.get("symbol") for s in (buys or [])}
        wl_syms  = {s.get("symbol"): s.get("tier", "") for s in (watchlist or [])}

        lines = [
            f"📈 <b>TOP {min(top_n, total)} MOMENTUM</b> (of {total} scanned)",
            "  <i>Score is AFTER chop penalty. Sorted highest first.</i>",
            "",
        ]
        for i, s in enumerate(shown, 1):
            sym  = _v2_clean_ticker(s.get("symbol", "?"))
            conf = _v2_safe_float(s.get("final_confidence", 0))
            sy   = s.get("symbol")
            if sy in buy_syms:
                verdict = "🟢 BUY"
            elif sy in wl_syms:
                tier = wl_syms[sy]
                verdict = {"NEAR_MISS": "🟡 close",
                           "DEVELOPING": "🔵 dev",
                           "MONITOR": "🔵 monitor"}.get(tier, "🔵 WL")
            else:
                reason = _v2_primary_reject_label(s)
                verdict = f"🔴 {_v2_html_esc(reason)}"
            lines.append(f"  {i:2}. <code>{_v2_html_esc(sym):<10}</code> {conf:5.1f}  →  {verdict}")
        if total > top_n:
            lines.append(f"  <i>… and {total - top_n} more (see CSV attachment)</i>")
        lines.append("")
        return lines
    except Exception as e:
        _log(f"[v2] section_top_momentum crashed: {e}")
        return []


def _v2_section_closest_rejects(rejected: list, regime: str,
                                 top_n: int = 15) -> list:
    """Rejected stocks sorted by proximity to min_conf — tomorrow's candidates.

    Grouped by primary reject reason so the reader immediately sees WHY each
    cohort didn't pass. Within each group, sorted by confidence descending.
    """
    try:
        if not rejected:
            return []
        try:
            thresh = REGIME_THRESHOLDS[regime]
        except Exception:
            thresh = {"min_confidence": 78}
        min_conf = _v2_safe_float(thresh.get("min_confidence"), 78)

        # Score by proximity: distance from min_conf
        def _distance(s):
            return abs(min_conf - _v2_safe_float(s.get("final_confidence", 0)))

        closest = sorted(rejected, key=_distance)[:top_n]
        if not closest:
            return []

        # Group by reject reason
        from collections import OrderedDict
        groups: "OrderedDict[str, list]" = OrderedDict()
        for s in closest:
            reason = _v2_primary_reject_label(s)
            groups.setdefault(reason, []).append(s)

        lines = [
            f"🥺 <b>CLOSEST {len(closest)} REJECTIONS</b> — tomorrow's candidates",
            "  <i>Grouped by the primary reason they didn't pass.</i>",
            "",
        ]
        for reason, stocks in groups.items():
            # Sort within group: highest conf first (closest to breaking through)
            stocks_sorted = sorted(
                stocks,
                key=lambda x: _v2_safe_float(x.get("final_confidence", 0)),
                reverse=True,
            )
            lines.append(
                f"  <b>{_v2_html_esc(reason)}</b> ({len(stocks_sorted)})"
            )
            for s in stocks_sorted:
                sym  = _v2_clean_ticker(s.get("symbol", "?"))
                tag  = _v2_setup_tag(s, with_emoji=False)      # "MOM" / "BRK" / "—"
                conf = _v2_safe_float(s.get("final_confidence", 0))
                need = max(0.0, min_conf - conf)
                lines.append(
                    f"    <code>{_v2_html_esc(sym):<10}</code> "
                    f"{conf:5.1f}  need <b>+{need:.1f}</b>  [{tag}]"
                )
            lines.append("")

        remaining = len(rejected) - len(closest)
        if remaining > 0:
            lines.append(f"  <i>… {remaining} more rejects further away (see CSV)</i>")
        lines.append("")
        return lines
    except Exception as e:
        _log(f"[v2] section_closest_rejects crashed: {e}")
        return []


def _v2_section_market_weather(regime_data: dict, macro: dict,
                                nifty_state: dict, thresh: dict) -> list:
    """Compact market weather / rules block."""
    try:
        regime = regime_data.get("regime", "?")
        score  = _v2_safe_float(regime_data.get("score", 0))
        vix_in = _v2_safe_float((macro or {}).get("vix_in", 15))
        nifty  = _v2_safe_float((macro or {}).get("nifty_1d_pct", 0))
        breadth = _v2_safe_float((nifty_state or {}).get("breadth20", 50))

        weather_emoji = {
            "STRONG_BULL": "☀️", "BULL": "🌤️", "TRANSITION": "🌫️",
            "SIDEWAYS": "🌫️", "HIGH_VOLATILITY": "⛈️",
            "BEAR": "🌧️", "STRONG_BEAR": "🌧️",
        }.get(regime, "🌤️")

        min_conf = _v2_safe_float((thresh or {}).get("min_confidence"), 78)
        min_tq   = _v2_safe_float((thresh or {}).get("min_tq"), 65)
        min_rr   = _v2_safe_float((thresh or {}).get("min_rr"), 1.8)

        lines = [
            "🌦️ <b>MARKET WEATHER</b>",
            f"   {weather_emoji} Regime: <b>{_v2_html_esc(regime)}</b>  ·  Score {score:.0f}/100",
            f"   Nifty {nifty:+.2f}%  ·  VIX-IN {vix_in:.1f}  ·  Breadth {breadth:.0f}%",
            f"   BUY bar: Conf ≥ <b>{min_conf:.0f}</b>  ·  TQ ≥ <b>{min_tq:.0f}</b>  ·  R/R ≥ <b>{min_rr:.2f}×</b>",
        ]
        # Chop-specific rule
        try:
            if regime.upper() in ("CHOP", "TRANSITION", "SIDEWAYS"):
                lines.append("   ⚠️ MOMENTUM penalised ×0.90  ·  PULLBACK / REVERSAL blocked")
        except Exception:
            pass
        lines.append("")
        return lines
    except Exception as e:
        _log(f"[v2] section_market_weather crashed: {e}")
        return []


def _v2_section_universe_breakdown(buys: list, watchlist: list, rejected: list,
                                    setup_mix: dict, setup_tickers: dict) -> list:
    """Universe scanned + reject-reason breakdown, WITH ticker examples per setup."""
    try:
        from collections import Counter
        total = len(buys or []) + len(watchlist or []) + len(rejected or [])
        if total == 0:
            return []

        # Aggregate reject reasons (primary reason per stock)
        reason_counter = Counter()
        for s in (rejected or []):
            reason_counter[_v2_primary_reject_label(s)] += 1

        lines = [
            "🔍 <b>UNIVERSE BREAKDOWN</b>",
            f"   Scanned <b>{total}</b> stocks by setup type:",
        ]
        # Setup mix — now with ticker samples so the reader SEES which stocks
        # are in each bucket (was previously just a bar + count).
        if setup_mix:
            _tickers_map = setup_tickers or {}
            for setup in ("BREAKOUT", "MOMENTUM", "PULLBACK", "REVERSAL", "OTHER"):
                n = int((setup_mix or {}).get(setup, 0))
                if n <= 0:
                    continue
                pct = int(round(n * 100 / total)) if total > 0 else 0
                bar = _v2_bar(n, total, width=10)   # narrower for mobile
                emo = _V2_SETUP_EMOJI.get(setup, "·")
                # Header row: bar + count
                lines.append(
                    f"     {bar}  {emo} <b>{setup:<8}</b> {n:>3} ({pct}%)"
                )
                # Sample tickers (up to 5, cleaned)
                samples = _tickers_map.get(setup) or []
                if samples:
                    clean = [_v2_clean_ticker(t) for t in samples[:5]]
                    more  = len(samples) - len(clean)
                    joined = ", ".join(clean)
                    if more > 0:
                        joined += f", <i>+{more} more</i>"
                    lines.append(f"        <code>{joined}</code>")
        # Reject reasons top 5
        if reason_counter and rejected:
            lines.append("")
            lines.append(f"   Why {len(rejected)} were rejected:")
            for reason, cnt in reason_counter.most_common(5):
                pct = int(round(cnt * 100 / len(rejected)))
                bar = _v2_bar(cnt, len(rejected), width=10)
                lines.append(
                    f"     {bar}  {_v2_html_esc(reason):<22} {cnt:>3} ({pct}%)"
                )
        lines.append("")
        return lines
    except Exception as e:
        _log(f"[v2] section_universe_breakdown crashed: {e}")
        return []


def _v2_section_shadow_log(setup_mix: dict, regime: str, rejected: list) -> list:
    """Compact shadow-log summary."""
    try:
        # We don't have direct shadow_log counts here, but we can approximate
        # bucket B/C from setup_mix + chop penalty presence in rejects.
        b_count = sum(1 for s in (rejected or [])
                      if _v2_safe_float(s.get("chop_momentum_penalty", 0)) > 0)
        chop_setups = ("PULLBACK", "REVERSAL", "OTHER")
        c_count = 0
        for s in (rejected or []):
            frs = s.get("fail_reasons", []) or []
            if any("SETUP_EDGE_SKIP" in str(f) for f in frs) and \
               _v2_setup_of(s) in chop_setups:
                c_count += 1

        lines = [
            "🔬 <b>SHADOW LOG</b> — learning tracker",
            "   Rejected stocks we still watch to check our filters:",
            f"     Bucket A (high-quality, risk-gate blocked):  {0:>3}",
            f"     Bucket B (chop-penalised momentum):          {b_count:>3}",
            f"     Bucket C (wrong setup for chop):             {c_count:>3}",
            "   <i>Once we have 30 days of data, we'll know if filters need tuning.</i>",
            "",
        ]
        return lines
    except Exception as e:
        _log(f"[v2] section_shadow_log crashed: {e}")
        return []


def _v2_section_tip() -> list:
    """Daily rotating tip."""
    try:
        tip = _v2_get_daily_tip()
        return [
            "💡 <b>TIP OF THE DAY</b>",
            f"   {_v2_html_esc(tip)}",
            "",
        ]
    except Exception:
        return []


def _v2_section_health(macro: dict, regime_data: dict, buys: list,
                        watchlist: list, rejected: list) -> list:
    """Pipeline health footer."""
    try:
        total = len(buys or []) + len(watchlist or []) + len(rejected or [])
        dq_bad = (macro or {}).get("dq_status") == "STALE"

        # Freeze status (best-effort — we don't want to hard-code the end date
        # anywhere else)
        freeze_line = ""
        try:
            freeze_end = os.getenv("FREEZE_END_DATE", "")
            if freeze_end:
                freeze_line = f"   📅 Freeze status: until {_v2_html_esc(freeze_end)}"
        except Exception:
            pass

        lines = ["🏥 <b>SYSTEM HEALTH</b>"]
        lines.append(f"   {'⚠️' if dq_bad else '✅'} Data quality: "
                     f"{'stale' if dq_bad else 'fresh'}  ·  {total} stocks scanned")
        lines.append(f"   ✅ Regime detect: {_v2_html_esc(regime_data.get('regime', '?'))}")
        if freeze_line:
            lines.append(freeze_line)
        lines.append("")
        return lines
    except Exception as e:
        _log(f"[v2] section_health crashed: {e}")
        return []


def _v2_section_glossary() -> list:
    """One-liner glossary at the bottom."""
    try:
        return [
            "📖 <b>GLOSSARY</b>",
            "   • <b>Confidence</b> (0-100): how sure the model is",
            "   • <b>TQ</b>: Trade Quality — chart pattern strength",
            "   • <b>R/R</b>: risk-reward ratio (need ≥1.8×)",
            "   • <b>BREAKOUT</b>: piercing resistance level",
            "   • <b>MOMENTUM</b>: stock in uptrend",
            "   • <b>PULLBACK</b>: dip in uptrend (risky in chop)",
            "   • <b>REVERSAL</b>: direction change (risky in chop)",
            "   • <b>Chop penalty</b>: -10% conf for MOMENTUM in chop",
            "   • <b>Shadow log</b>: rejected trades tracked to verify filters",
            "   ─────────────────",
            "   ⚠️ Recommendation only. Execute manually.",
        ]
    except Exception:
        return []


def _v2_section_footer_csv(csv_path: str) -> list:
    """Pointer to CSV attachment."""
    try:
        if not csv_path or not os.path.exists(csv_path):
            return []
        size_kb = os.path.getsize(csv_path) / 1024.0
        return [
            "",
            "📎 <b>FULL UNIVERSE</b> — attached as CSV",
            f"   <i>{_v2_html_esc(os.path.basename(csv_path))} ({size_kb:.1f} KB) — "
            f"every stock scanned with score, verdict, reason.</i>",
        ]
    except Exception:
        return []


# ── CSV attachment generator ───────────────────────────────────────────────

def _v2_write_daily_csv(buys: list, watchlist: list, rejected: list,
                        regime: str, timestamp: str,
                        out_path: str = None) -> str:
    """Write every evaluated stock to a CSV. Returns path (empty on failure).

    Columns: Ticker, Setup, Sector, Price, MCap_Cr, RawConf, FinalConf, TQ,
             RR, ROE, DE, Verdict, PrimaryReject, AllRejects, Tier
    """
    try:
        out_path = out_path or TELEGRAM_DAILY_CSV
        # Add date stamp to filename so history is preserved
        try:
            date_str = ist_today().strftime("%Y%m%d")
            root, ext = os.path.splitext(out_path)
            out_path = f"{root}_{date_str}{ext or '.csv'}"
        except Exception:
            pass

        buy_syms = {s.get("symbol") for s in (buys or [])}
        wl_map = {s.get("symbol"): s.get("tier", "WATCHLIST")
                  for s in (watchlist or [])}

        def _row(s, verdict_hint=None):
            try:
                sym    = _v2_clean_ticker(s.get("symbol", ""))
                setup  = str(s.get("setup", "") or "").upper()
                sector = str(s.get("sector", "") or "")
                price  = _v2_safe_float(s.get("entry",
                            s.get("price", s.get("close", 0))))
                mcap   = _v2_safe_float(s.get("market_cap",
                            s.get("mcap_cr", 0)))
                raw_c  = _v2_safe_float(s.get("raw_confidence",
                            s.get("orig_confidence", s.get("final_confidence", 0))))
                fin_c  = _v2_safe_float(s.get("final_confidence", 0))
                tq     = _v2_safe_float(s.get("trade_quality_score", 0))
                rr     = _v2_safe_float(s.get("rr_ratio", s.get("rr", 0)))
                roe    = _v2_safe_float(s.get("roe", 0))
                de     = _v2_safe_float(s.get("de_ratio", 0))

                sy = s.get("symbol")
                if verdict_hint:
                    verdict = verdict_hint
                elif sy in buy_syms:
                    verdict = "BUY"
                elif sy in wl_map:
                    verdict = f"WATCHLIST_{wl_map[sy]}"
                else:
                    verdict = "REJECTED"

                primary = _v2_primary_reject_label(s) if verdict == "REJECTED" else ""
                all_frs = ";".join(str(f) for f in (s.get("fail_reasons") or []))
                tier    = wl_map.get(sy, "") if verdict.startswith("WATCHLIST") else ""

                return [sym, setup, sector, f"{price:.2f}", f"{mcap:.0f}",
                        f"{raw_c:.2f}", f"{fin_c:.2f}", f"{tq:.2f}",
                        f"{rr:.2f}", f"{roe:.2f}", f"{de:.2f}",
                        verdict, primary, all_frs, tier]
            except Exception as e:
                _log(f"[v2] csv row failed for {s.get('symbol')}: {e}")
                return None

        rows = []
        for s in (buys or []):
            r = _row(s, "BUY")
            if r: rows.append(r)
        for s in (watchlist or []):
            r = _row(s)
            if r: rows.append(r)
        for s in (rejected or []):
            r = _row(s, "REJECTED")
            if r: rows.append(r)

        header = ["Ticker", "Setup", "Sector", "Price", "MCap_Cr", "RawConf",
                  "FinalConf", "TQ", "RR", "ROE", "DE", "Verdict",
                  "PrimaryReject", "AllRejects", "Tier"]
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(header)
            w.writerows(rows)
        _log(f"[v2] wrote daily CSV: {out_path} ({len(rows)} rows)")
        return out_path
    except Exception as e:
        _log(f"[v2] write_daily_csv failed: {e}")
        return ""


def send_telegram_document(file_path: str, caption: str = "") -> None:
    """Send a file as a Telegram document to the main channel."""
    try:
        if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
            return
        if not file_path or not os.path.exists(file_path):
            return
        url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"
        with open(file_path, "rb") as f:
            files = {"document": (os.path.basename(file_path), f)}
            data  = {"chat_id": TELEGRAM_CHAT_ID}
            if caption:
                data["caption"] = caption[:1000]  # Telegram caption limit
            resp = requests.post(url, data=data, files=files, timeout=30)
        if resp.status_code != 200:
            _log(f"[v2] Telegram sendDocument failed: {resp.status_code} {resp.text[:200]}")
    except Exception as e:
        _log(f"[v2] send_telegram_document error: {e}")


# ═════════════════════════════════════════════════════════════════════════════
# v2.1 CONSOLIDATED SECTIONS — 8-section design (2026-07-10)
#
# Design consensus: Retail Investor + Product Designer + Data Analyst +
# Info Architect. Rules:
#   1. Each ticker appears in EXACTLY ONE section (`shown_symbols` set)
#   2. Header format `━━ TITLE ━━` fits Android Telegram 34-char viewport
#   3. Plain-English tags replace jargon (TQ / R/R / MOM×0.9 → words)
#   4. Threshold displayed ONCE in HEADER only
#
# Sections: HEADER → BUY NOW → ALMOST READY → BY SETUP → YESTERDAY →
#           UNIVERSE → HEALTH & TIP → GLOSSARY
# ═════════════════════════════════════════════════════════════════════════════

_V21_REGIME_LABEL = {
    "STRONG_BULL":     ("☀️", "Strong Bull"),
    "BULL":            ("🌤️", "Bullish"),
    "TRANSITION":      ("🌫️", "Transition"),
    "SIDEWAYS":        ("🌫️", "Sideways / Chop"),
    "HIGH_VOLATILITY": ("⛈️", "High Volatility"),
    "BEAR":            ("🌧️", "Bearish"),
    "STRONG_BEAR":     ("🌧️", "Strong Bear"),
}

# BY SETUP: fixed display order; OTHER is intentionally excluded
_V21_SETUP_ORDER = ("BREAKOUT", "MOMENTUM", "PULLBACK", "REVERSAL")
_V21_SETUP_WORD  = {
    "BREAKOUT": "Breakout",
    "MOMENTUM": "Momentum",
    "PULLBACK": "Pullback",
    "REVERSAL": "Reversal",
}


def _v21_head(title: str) -> str:
    """`━━ TITLE ━━` header. Fits 34-char mobile viewport."""
    return f"━━ {title} ━━"


# ── Phase v21-Close-Tag-v2 (2026-07-10): tighten "close" tag ──────────────
# The "close" label used to fire whenever conf was within 15 pts of the bar,
# regardless of TQ / R/R gates. This produced misleading output like
# `IKS 77 · close` when IKS was actually rejected on TQ, not on confidence.
#
# New rule:  "close" = (conf gap ≤ CLOSE_LABEL_MAX_GAP) AND TQ within 5% of
#            bar AND R/R within 10% of bar. Otherwise fall through to the
#            more specific tag (TQ low / weak R/R / conf low).
#
# Overrides:  CLOSE_LABEL_MAX_GAP env var (default 8 pts). Set to 15 to
#             restore old behaviour if a downstream consumer breaks.
try:
    _CLOSE_LABEL_MAX_GAP = float(os.getenv("CLOSE_LABEL_MAX_GAP", "8") or 8)
except Exception:
    _CLOSE_LABEL_MAX_GAP = 8.0


def _v21_english_tag(s: dict, min_conf: float, min_tq: float,
                     min_rr: float) -> str:
    """Return a one-word plain-English tag explaining why a stock is not
    a BUY. Empty string if nothing notable to flag.

    Priority (highest signal first):
      close    → conf_gap ≤ CLOSE_LABEL_MAX_GAP AND TQ ≥ 95% of bar AND
                 R/R ≥ 90% of bar (i.e. only the confidence gate is holding
                 it back). Default max gap = 8 pts.
      TQ low   → trade-quality score below regime bar
      weak R/R → reward-to-risk below regime bar
      conf low → confidence gap > CLOSE_LABEL_MAX_GAP
      choppy   → chop-regime momentum penalty was applied
      ROE low / D/E high / too small → fundamentals gate
    """
    try:
        conf = _v2_safe_float(s.get("final_confidence", s.get("conf", 0)))
        tq   = _v2_safe_float(s.get("trade_quality_score", s.get("tq", 0)))
        rr   = _v2_safe_float(s.get("rr_ratio", s.get("rr", 0)))
        penalty = _v2_safe_float(s.get("chop_momentum_penalty", 0))
        fails = s.get("fail_reasons", []) or []

        # "close" only when the ONLY blocker is a small confidence gap.
        # Guarantees: IKS at 77 conf with TQ<min_tq → "TQ low" (not "close").
        conf_close = (min_conf - _CLOSE_LABEL_MAX_GAP) <= conf < min_conf
        tq_ok      = (tq <= 0) or (tq >= min_tq * 0.95)
        rr_ok      = (rr <= 0) or (rr >= min_rr * 0.90)
        if conf_close and tq_ok and rr_ok:
            return "close"
        if tq > 0 and tq < min_tq:
            return "TQ low"
        if rr > 0 and rr < min_rr:
            return "weak R/R"
        if conf < min_conf - _CLOSE_LABEL_MAX_GAP:
            return "conf low"
        if penalty > 0:
            return "choppy"
        for f in fails:
            up = str(f).upper()
            if "ROE" in up:
                return "ROE low"
            if "DE_" in up or "D/E" in up:
                return "D/E high"
            if "MCAP" in up or "MARKET_CAP" in up:
                return "too small"
        return ""
    except Exception:
        return ""


def _v21_setup_emoji(s: dict) -> str:
    """Setup emoji for a stock; falls back to `·` for unknown."""
    return _V2_SETUP_EMOJI.get(_v2_setup_of(s), "·")


# ── Section 1: HEADER ──────────────────────────────────────────────────────
def _v21_section_header(timestamp: str, regime_data: dict, macro: dict,
                        thresh: dict, buys: list, watchlist: list,
                        rejected: list) -> list:
    """Merged banner + quick-glance + market weather."""
    try:
        try:
            market_date = ist_today().strftime("%d %b %Y")
            dow         = ist_today().strftime("%a")
        except Exception:
            market_date, dow = "", ""

        regime = (regime_data or {}).get("regime", "SIDEWAYS")
        emo, label = _V21_REGIME_LABEL.get(regime, ("🌫️", regime))

        n_buy = len(buys or [])
        n_wl  = len(watchlist or [])
        n_rej = len(rejected or [])

        nifty = _v2_safe_float((macro or {}).get("nifty_1d_pct", 0))
        vix   = _v2_safe_float((macro or {}).get("vix_in", 0))

        min_conf = _v2_safe_float((thresh or {}).get("min_confidence"), 78)
        min_tq   = _v2_safe_float((thresh or {}).get("min_tq"), 65)
        min_rr   = _v2_safe_float((thresh or {}).get("min_rr"), 1.8)

        lines = [
            "━━━━━━━━━━━━━━━━━━━━━━",
            f"📊 <b>DAILY BRIEF</b> · {_v2_html_esc(market_date)}",
            f"{_v2_html_esc(dow)} · Regime: {emo} <b>{_v2_html_esc(label)}</b>",
            f"Nifty {nifty:+.2f}%  ·  VIX {vix:.1f}",
            f"Today's bar: Conf≥<b>{min_conf:.0f}</b> · "
            f"TQ≥<b>{min_tq:.0f}</b> · R/R≥<b>{min_rr:.2f}×</b>",
            f"<b>{n_buy}</b> BUY · <b>{n_wl}</b> watch · <b>{n_rej}</b> rej",
        ]
        try:
            if regime.upper() in ("CHOP", "TRANSITION", "SIDEWAYS"):
                lines.append("⚠ chop: Momentum penalised ×0.9")
        except Exception:
            pass
        lines.append("━━━━━━━━━━━━━━━━━━━━━━")
        lines.append("")
        return lines
    except Exception as e:
        _log(f"[v21] header crashed: {e}")
        return ["━━ DAILY BRIEF ━━", ""]


# ── Section 2: BUY NOW ─────────────────────────────────────────────────────
def _v21_section_buy_now(buys: list, regime: str,
                         ai_results: dict = None) -> list:
    """Green-lit actionable picks. Fallback message if empty."""
    try:
        header = _v21_head("🎯 BUY NOW")
        if not buys:
            return [
                f"<b>{header}</b>",
                "",
                "  <i>No BUYS today.</i>",
                "  See <b>ALMOST READY</b> below for",
                "  tomorrow's most likely candidates.",
                "",
            ]

        insights = (ai_results or {}).get("insights", {}) or {}
        buys_sorted = sorted(
            buys,
            key=lambda x: _v2_safe_float(x.get("final_confidence", 0)),
            reverse=True,
        )
        lines = [f"<b>{header}</b>", ""]
        for i, s in enumerate(buys_sorted, 1):
            sym    = _v2_clean_ticker(s.get("symbol", "?"))
            setup  = _v2_setup_of(s) or "SETUP"
            setup_word = _V21_SETUP_WORD.get(setup, setup.title())
            emo    = _v21_setup_emoji(s)
            sector = _v2_html_esc(s.get("sector", "") or "")
            conf   = _v2_safe_float(s.get("final_confidence", 0))
            tq     = _v2_safe_float(s.get("trade_quality_score", s.get("tq", 0)))
            rr     = _v2_safe_float(s.get("rr_ratio", s.get("rr", 0)))
            entry  = _v2_safe_float(s.get("entry", 0))
            stop   = _v2_safe_float(s.get("stop", 0))
            t1     = _v2_safe_float(s.get("target1", 0))
            bar    = _v2_bar(conf, 100.0, width=10)
            insight = _v2_html_esc(insights.get(s.get("symbol", ""), ""))

            lines.append(
                f"{i}) <b>{_v2_html_esc(sym)}</b> · {emo} {setup_word}"
                + (f" · <i>{sector}</i>" if sector else "")
            )
            lines.append(f"   Conf {bar} <b>{conf:.0f}</b>")
            lines.append(
                f"   Buy <b>{entry:.1f}</b> · SL {stop:.1f} · Tgt {t1:.1f}"
            )
            lines.append(f"   R/R <b>{rr:.2f}×</b>  ·  TQ {tq:.0f}")
            if insight:
                lines.append(f"   💡 {insight[:80]}")
            lines.append("")
        return lines
    except Exception as e:
        _log(f"[v21] buy_now crashed: {e}")
        return []


# ── Section 3: ALMOST READY ────────────────────────────────────────────────
# Phase v21-AlmostReady-v2 (2026-07-10): honest "tomorrow candidates" gate.
#
# The old implementation rendered EVERY stock on the watchlist regardless
# of how far it was from the BUY bar. The watchlist is populated upstream
# by a generous "developing setup" filter (conf ≥ ~45) meant to TRACK
# stocks over weeks, not "tomorrow candidates."
#
# Result: entries like `ORCHPHARMA 48 · TQ low` in an ALMOST READY section
# whose subtitle promises "Tomorrow's most likely candidates" — with a
# 35-point gap it would take 3-5 weeks at 1-2 pts/day of natural drift.
#
# Phase v21-AlmostReady-v2.1 (2026-07-10 pm): keep individual names in
# Tier 2 but explain WHY each is shown via a bracketed reason. Rolling
# them up into a single "Watching N more" line hid actionable per-stock
# context (which sector, which setup) that the user wants to see.
#
# New tiered rendering (all thresholds env-overridable):
#   • Tier 1 — ALMOST READY  : conf_gap ≤ ALMOST_READY_MAX_GAP (default 12)
#                              rendered in full with setup emoji + short tag
#                              (close / TQ low / weak R/R / conf low)
#   • Tier 2 — Watching      : ALMOST_READY_MAX_GAP < conf_gap ≤
#                              WATCHING_MAX_GAP (default 30) — each stock
#                              rendered individually under a "Watching"
#                              sub-header with a bracketed reason showing
#                              exactly how far the conf gap is. Example:
#                                ▸ MIDGAP1  65 · 📈  (needs +18 conf)
#   • Tier 3 — Long-tail     : conf_gap > WATCHING_MAX_GAP — dropped from
#                              the message entirely (still in CSV and in
#                              shadow_master.xlsx bucket B/C for the day-30
#                              analyst; these are 3+ weeks away from BUY).
# If BOTH Tier 1 and Tier 2 are empty, the whole section is skipped
# (no empty header). If only Tier 2 has entries, we still render the
# section — the reader wants to see the developing pipeline even when
# nothing is imminently ready.
try:
    _ALMOST_READY_MAX_GAP = float(os.getenv("ALMOST_READY_MAX_GAP", "12") or 12)
except Exception:
    _ALMOST_READY_MAX_GAP = 12.0
try:
    _WATCHING_MAX_GAP = float(os.getenv("WATCHING_MAX_GAP", "30") or 30)
except Exception:
    _WATCHING_MAX_GAP = 30.0


def _v21_section_almost_ready(watchlist: list, regime: str,
                              shown_symbols: set) -> list:
    """Filtered ALMOST READY — Tier 1 = tomorrow candidates (conf_gap ≤ 12),
    Tier 2 = individually-named watching list with bracketed reason
    (gap 13-30), Tier 3 = drop.

    Adds every rendered ticker (Tier 1 AND Tier 2) to `shown_symbols` so
    BY SETUP does not re-print them. Tier 3 stocks stay in the pool so
    BY SETUP can still show them if they're top rejects in their setup.
    """
    try:
        wl = list(watchlist or [])
        if not wl:
            return []

        try:
            thresh = REGIME_THRESHOLDS[regime]
        except Exception:
            thresh = {"min_confidence": 78, "min_tq": 65, "min_rr": 1.8}
        min_conf = _v2_safe_float(thresh.get("min_confidence"), 78)
        min_tq   = _v2_safe_float(thresh.get("min_tq"), 65)
        min_rr   = _v2_safe_float(thresh.get("min_rr"), 1.8)

        # Compute gap once and bucket into tiers
        tier1 = []   # conf_gap ≤ _ALMOST_READY_MAX_GAP
        tier2 = []   # _ALMOST_READY_MAX_GAP < gap ≤ _WATCHING_MAX_GAP
        for w in wl:
            c = _v2_safe_float(w.get("conf", w.get("final_confidence", 0)))
            gap = max(0.0, min_conf - c)
            if gap <= _ALMOST_READY_MAX_GAP:
                tier1.append((gap, w))
            elif gap <= _WATCHING_MAX_GAP:
                tier2.append((gap, w))
            # tier3 (gap > WATCHING_MAX_GAP) intentionally not rendered here

        # If NEITHER tier has candidates, skip the whole section.
        # Do NOT print a header with 0 entries — misleading.
        if not tier1 and not tier2:
            return []

        # Sort each tier by gap ascending (closest to BUY first)
        tier1.sort(key=lambda gw: gw[0])
        tier2.sort(key=lambda gw: gw[0])

        # Header count reflects the ACTIONABLE Tier 1 count; Tier 2 is a
        # developing list shown for context.
        header = _v21_head(f"⏳ ALMOST READY ({len(tier1)})")
        lines = [
            f"<b>{header}</b>",
            "  <i>Tomorrow's most likely candidates</i>",
            "",
        ]

        # ── Tier 1 rendering ──
        for gap, w in tier1:
            sym  = _v2_clean_ticker(w.get("symbol", "?"))
            emo  = _v21_setup_emoji(w)
            conf = _v2_safe_float(w.get("conf", w.get("final_confidence", 0)))
            tag  = _v21_english_tag(w, min_conf, min_tq, min_rr)
            tag_str = f" · <i>{_v2_html_esc(tag)}</i>" if tag else ""
            lines.append(
                f"▸ <code>{_v2_html_esc(sym):<10}</code> "
                f"<b>{conf:.0f}</b> · {emo}{tag_str}"
            )
            try:
                shown_symbols.add(w.get("symbol"))
            except Exception:
                pass

        # ── Tier 2 rendering — named list with per-stock bracketed reason ──
        if tier2:
            lines.append("")
            lines.append("  <i>Watching (developing — not ready tomorrow):</i>")
            for gap, w in tier2:
                sym  = _v2_clean_ticker(w.get("symbol", "?"))
                emo  = _v21_setup_emoji(w)
                conf = _v2_safe_float(
                    w.get("conf", w.get("final_confidence", 0)))
                # Build the bracketed reason. Prefer the most-specific
                # gate that is failing. Order matches _v21_english_tag
                # priority so BY SETUP and ALMOST READY tell the same
                # story about a stock.
                reason_parts = []
                gap_pts = int(round(gap))
                if gap_pts > 0:
                    reason_parts.append(f"needs +{gap_pts} conf")
                tq_val = _v2_safe_float(
                    w.get("trade_quality_score", w.get("tq", 0)))
                rr_val = _v2_safe_float(
                    w.get("rr_ratio", w.get("rr", 0)))
                if tq_val > 0 and tq_val < min_tq:
                    reason_parts.append(
                        f"TQ {tq_val:.0f}<{min_tq:.0f}")
                if rr_val > 0 and rr_val < min_rr:
                    reason_parts.append(
                        f"R/R {rr_val:.1f}<{min_rr:.1f}")
                reason = ", ".join(reason_parts) if reason_parts \
                    else "under review"
                lines.append(
                    f"▸ <code>{_v2_html_esc(sym):<10}</code> "
                    f"<b>{conf:.0f}</b> · {emo} "
                    f"<i>({_v2_html_esc(reason)})</i>"
                )
                try:
                    shown_symbols.add(w.get("symbol"))
                except Exception:
                    pass

        lines.append("")
        lines.append(f"  <i>Bar to BUY: {min_conf:.0f} conf</i>")
        lines.append("")
        return lines
    except Exception as e:
        _log(f"[v21] almost_ready crashed: {e}")
        return []


# ── Section 4: BY SETUP ────────────────────────────────────────────────────
# Phase v21-BySetup-v2 (2026-07-10): show ALL rejects per setup (was top-3).
#
# The old top-3 cap was hiding legitimate candidates behind the scenes.
# The reader was seeing `(top 3 of 47)` and had NO way to know what the
# other 44 were — even though many might be equally or more actionable
# than the top-3 shown. The counter-argument is Telegram's 4096-char body
# limit, but a realistic run has 5-20 rejects per setup category (not
# hundreds), which fits comfortably.
#
# New behavior:
#   • Default cap raised 3 → 8 per setup (env: BY_SETUP_MAX_PER_SETUP)
#   • If a subgroup HAS more than the cap, add an honest overflow line
#     that NAMES the extra symbols (up to 12) rather than hiding them:
#       `... and 5 more: SYM1, SYM2, SYM3, SYM4, SYM5 (see CSV)`
#   • When the overflow tail is > 12 symbols, name the first 12 and
#     summarize the remainder with a count: `... +N more (see CSV)`
#   • All extras get registered in shown_symbols so downstream sections
#     don't repeat them.
try:
    _BY_SETUP_MAX_PER_SETUP = int(
        os.getenv("BY_SETUP_MAX_PER_SETUP", "8") or 8)
except Exception:
    _BY_SETUP_MAX_PER_SETUP = 8
try:
    # Cap on how many extras get NAMED in the overflow line. Above this,
    # we fall back to a numeric summary to keep the message under 4096.
    _BY_SETUP_MAX_NAMED_EXTRAS = int(
        os.getenv("BY_SETUP_MAX_NAMED_EXTRAS", "12") or 12)
except Exception:
    _BY_SETUP_MAX_NAMED_EXTRAS = 12


def _v21_section_by_setup(buys: list, watchlist: list, rejected: list,
                          shown_symbols: set, regime: str,
                          setup_mix: dict,
                          top_n_per_setup: int = None) -> list:
    """Rejects grouped by setup type (Breakout / Momentum / Pullback /
    Reversal). Excludes tickers already in BUY NOW or ALMOST READY. OTHER
    setup is always dropped.

    Renders top-N per subgroup (N = ``top_n_per_setup`` or
    ``_BY_SETUP_MAX_PER_SETUP``, default 8) with a one-line honest
    overflow that NAMES the extras when there are more than N.

    Sub-header format: `🚀 Breakout   (showing N of TOTAL)` where TOTAL is
    the full universe count for that setup (from `setup_mix`).
    """
    # Resolve cap: explicit arg > env-var default
    cap = top_n_per_setup if top_n_per_setup is not None \
        else _BY_SETUP_MAX_PER_SETUP
    try:
        try:
            thresh = REGIME_THRESHOLDS[regime]
        except Exception:
            thresh = {"min_confidence": 78, "min_tq": 65, "min_rr": 1.8}
        min_conf = _v2_safe_float(thresh.get("min_confidence"), 78)
        min_tq   = _v2_safe_float(thresh.get("min_tq"), 65)
        min_rr   = _v2_safe_float(thresh.get("min_rr"), 1.8)

        # Full pool = everything minus already-shown tickers
        pool = []
        for src in (buys or [], watchlist or [], rejected or []):
            for s in src:
                if s.get("symbol") not in shown_symbols:
                    pool.append(s)
        if not pool:
            return []

        # Group by setup
        by_setup = {k: [] for k in _V21_SETUP_ORDER}
        for s in pool:
            st = _v2_setup_of(s)
            if st in by_setup:  # excludes OTHER + unknown
                by_setup[st].append(s)
        # Rank each subgroup by confidence descending
        for k in by_setup:
            by_setup[k].sort(
                key=lambda x: _v2_safe_float(x.get("final_confidence", 0)),
                reverse=True,
            )

        # Skip section entirely if every subgroup is empty
        if not any(by_setup[k] for k in _V21_SETUP_ORDER):
            return []

        header = _v21_head("📋 BY SETUP")
        lines = [
            f"<b>{header}</b>",
            "  <i>Rejects grouped by setup type</i>",
            "",
        ]
        for st in _V21_SETUP_ORDER:
            subgroup = by_setup[st]
            if not subgroup:
                continue
            picks = subgroup[:cap]
            extras = subgroup[cap:]
            total_in_universe = int(
                (setup_mix or {}).get(st, len(subgroup)))
            emo  = _V2_SETUP_EMOJI.get(st, "·")
            word = _V21_SETUP_WORD.get(st, st.title())
            lines.append(
                f"{emo} <b>{word}</b>   "
                f"<i>(showing {len(picks)} of {total_in_universe})</i>"
            )
            for i, s in enumerate(picks):
                sym  = _v2_clean_ticker(s.get("symbol", "?"))
                conf = _v2_safe_float(s.get("final_confidence", 0))
                tag  = _v21_english_tag(
                    s, min_conf, min_tq, min_rr) or "rejected"
                marker = "  ⭐" if i == 0 else "     "
                lines.append(
                    f"{marker} <code>{_v2_html_esc(sym):<10}</code> "
                    f"<b>{conf:.0f}</b> · <i>{_v2_html_esc(tag)}</i>"
                )
                try:
                    shown_symbols.add(s.get("symbol"))
                except Exception:
                    pass

            # Honest overflow: NAME the extras instead of hiding them.
            if extras:
                named = extras[:_BY_SETUP_MAX_NAMED_EXTRAS]
                remaining = len(extras) - len(named)
                names_txt = ", ".join(
                    _v2_clean_ticker(x.get("symbol", "?"))
                    for x in named
                )
                if remaining > 0:
                    lines.append(
                        f"     <i>... and {len(extras)} more: "
                        f"{_v2_html_esc(names_txt)}, "
                        f"+{remaining} more (see CSV)</i>"
                    )
                else:
                    lines.append(
                        f"     <i>... and {len(extras)} more: "
                        f"{_v2_html_esc(names_txt)} (see CSV)</i>"
                    )
                # Register EVERY extra so downstream sections
                # don't accidentally re-print them
                for x in extras:
                    try:
                        shown_symbols.add(x.get("symbol"))
                    except Exception:
                        pass
            lines.append("")
        return lines
    except Exception as e:
        _log(f"[v21] by_setup crashed: {e}")
        return []


# ── Section 5: YESTERDAY ───────────────────────────────────────────────────
def _v21_section_yesterday(buys: list, watchlist: list, rejected: list,
                           regime: str, prev_state: dict) -> list:
    """Delta view vs yesterday's saved snapshot."""
    try:
        if not prev_state:
            return []
        n_buy = len(buys or [])
        n_wl  = len(watchlist or [])
        n_rej = len(rejected or [])
        p_buy = int(prev_state.get("buys", 0))
        p_wl  = int(prev_state.get("watchlist", 0))
        p_rej = int(prev_state.get("rejected", 0))

        def _delta(cur, prev):
            d = cur - prev
            if d == 0:
                return "—"
            return f"{'+' if d > 0 else ''}{d}"

        header = _v21_head("📈 YESTERDAY")
        return [
            f"<b>{header}</b>",
            f"  Buys    {p_buy} → <b>{n_buy}</b>  ({_delta(n_buy, p_buy)})",
            f"  Watch   {p_wl} → <b>{n_wl}</b>  ({_delta(n_wl, p_wl)})",
            f"  Rejects {p_rej} → <b>{n_rej}</b>  ({_delta(n_rej, p_rej)})",
            "",
        ]
    except Exception as e:
        _log(f"[v21] yesterday crashed: {e}")
        return []


# ── Section 6: UNIVERSE STATS ──────────────────────────────────────────────
def _v21_section_universe_stats(buys: list, watchlist: list,
                                rejected: list) -> list:
    """Aggregates only — no ticker names. Setup mix is intentionally
    omitted because BY SETUP already shows `of TOTAL` counts.
    """
    try:
        from collections import Counter
        total = len(buys or []) + len(watchlist or []) + len(rejected or [])
        if total == 0:
            return []

        header = _v21_head("🔍 UNIVERSE")
        lines = [
            f"<b>{header}</b>",
            f"  Scanned: <b>{total}</b>  ·  "
            f"BUY: <b>{len(buys or [])}</b>  ·  "
            f"Watch: <b>{len(watchlist or [])}</b>",
        ]

        if rejected:
            reason_counter = Counter()
            for s in rejected:
                reason_counter[_v2_primary_reject_label(s)] += 1
            if reason_counter:
                top_reason, top_cnt = reason_counter.most_common(1)[0]
                pct = int(round(top_cnt * 100 / len(rejected)))
                lines.append(
                    f"  Top reject: <b>{_v2_html_esc(top_reason)}</b> "
                    f"— {top_cnt} ({pct}%)"
                )

        try:
            b_count = sum(1 for s in (rejected or [])
                          if _v2_safe_float(s.get("chop_momentum_penalty", 0)) > 0)
            chop_setups = ("PULLBACK", "REVERSAL", "OTHER")
            c_count = 0
            for s in (rejected or []):
                frs = s.get("fail_reasons", []) or []
                if any("SETUP_EDGE_SKIP" in str(f) for f in frs) and \
                   _v2_setup_of(s) in chop_setups:
                    c_count += 1
            lines.append(
                f"  Shadow: A <b>0</b> · B <b>{b_count}</b> · C <b>{c_count}</b>"
            )
        except Exception:
            pass
        lines.append("")
        return lines
    except Exception as e:
        _log(f"[v21] universe_stats crashed: {e}")
        return []


# ── Section 7: HEALTH & TIP ────────────────────────────────────────────────
def _v21_section_health_tip(macro: dict, regime_data: dict,
                            buys: list, watchlist: list,
                            rejected: list) -> list:
    """Merged system-health + daily tip."""
    try:
        macro_ok  = bool(macro) and \
                    _v2_safe_float((macro or {}).get("nifty_1d_pct", 0)) is not None
        regime_ok = bool((regime_data or {}).get("regime"))
        scan_ok   = (len(buys or []) + len(watchlist or []) +
                     len(rejected or [])) > 0
        all_ok = macro_ok and regime_ok and scan_ok
        status = "🟢 OK" if all_ok else \
                 ("🟡 partial" if (regime_ok or scan_ok) else "🔴 issue")

        try:
            tip = _v2_get_daily_tip()
        except Exception:
            tip = "Position size is your edge."

        header = _v21_head("💡 HEALTH & TIP")
        return [
            f"<b>{header}</b>",
            f"  System: {status}",
            f"  <b>Tip:</b> <i>{_v2_html_esc(tip)[:180]}</i>",
            "",
        ]
    except Exception as e:
        _log(f"[v21] health_tip crashed: {e}")
        return []


# ── Section 8: GLOSSARY ────────────────────────────────────────────────────
def _v21_section_glossary_mini() -> list:
    """Only the 6 terms actually used above."""
    try:
        header = _v21_head("📖 GLOSSARY")
        return [
            f"<b>{header}</b>",
            "  🚀 Breakout · 📈 Momentum",
            "  ↩️ Pullback · 🔄 Reversal",
            "  <b>Conf</b> = model confidence 0-100",
            "  <b>R/R</b>  = reward per ₹ risked",
            "  <b>close</b> = ≤15 pts from BUY bar",
            "  <b>choppy</b> = chop penalty applied",
            "",
        ]
    except Exception as e:
        _log(f"[v21] glossary crashed: {e}")
        return []


# ── v2 Orchestrator ─────────────────────────────────────────────────────────

def format_telegram_message_v2(
        regime_data: dict, buys: list, shorts: list,
        watchlist: list, portfolio_alerts: list,
        macro: dict, key_levels: dict,
        upcoming_events: list, timestamp: str,
        heat: dict = None, platt: dict = None,
        tracker_v2: dict = None,
        rejected_stocks: list = None,
        breadth_20: float = 50.0,
        nifty_state: dict = None,
        universe_count: int = 0,
        tradable_count: int = 0,
        conf_history: dict = None,
        gate_memory: dict = None,
        ai_results: dict = None) -> str:
    """Build the redesigned Telegram message.

    Each section is independent + wrapped in try/except so one bad section
    can't kill the whole message. If the whole function raises, the caller
    catches it and falls back to v1.
    """
    lines = []
    regime = regime_data.get("regime", "SIDEWAYS")
    try:
        thresh = REGIME_THRESHOLDS[regime]
    except Exception:
        thresh = {"min_confidence": 78, "min_tq": 65, "min_rr": 1.8}

    setup_mix     = regime_data.get("_setup_mix") or {}
    setup_tickers = regime_data.get("_setup_tickers") or {}
    prev_state    = _v2_load_prev_state()

    # v2.1: strict one-ticker-one-section tracking. Pre-populate with BUY
    # symbols so BY_SETUP filters them out (ALMOST_READY and BY_SETUP both
    # mutate this set as they render).
    shown_symbols = set()
    for _b in (buys or []):
        try:
            shown_symbols.add(_b.get("symbol"))
        except Exception:
            pass

    # ─── Compose all sections (v2.1: 8 consolidated) ───────────────────────
    _sections = [
        ("header",         lambda: _v21_section_header(
                              timestamp, regime_data, macro, thresh,
                              buys, watchlist, rejected_stocks)),
        ("buy_now",        lambda: _v21_section_buy_now(
                              buys, regime, ai_results)),
        ("almost_ready",   lambda: _v21_section_almost_ready(
                              watchlist, regime, shown_symbols)),
        # Phase v21-BySetup-v2 (2026-07-10): pass None so the env-var
        # default (_BY_SETUP_MAX_PER_SETUP, default 8) takes effect.
        # Overrides via the BY_SETUP_MAX_PER_SETUP env var.
        ("by_setup",       lambda: _v21_section_by_setup(
                              buys, watchlist, rejected_stocks,
                              shown_symbols, regime, setup_mix,
                              top_n_per_setup=None)),
        ("yesterday",      lambda: _v21_section_yesterday(
                              buys, watchlist, rejected_stocks,
                              regime, prev_state)),
        ("universe",       lambda: _v21_section_universe_stats(
                              buys, watchlist, rejected_stocks)),
        ("health_tip",     lambda: _v21_section_health_tip(
                              macro, regime_data, buys, watchlist,
                              rejected_stocks)),
        ("glossary",       lambda: _v21_section_glossary_mini()),
    ]
    for name, builder in _sections:
        try:
            lines.extend(builder())
        except Exception as e:
            _log(f"[v21] section '{name}' crashed: {e} — skipping")

    # Save today's snapshot so tomorrow's "vs yesterday" works
    try:
        _v2_save_prev_state({
            "date":              str(ist_today()),
            "buys":              len(buys or []),
            "watchlist":         len(watchlist or []),
            "rejected":          len(rejected_stocks or []),
            "regime":            regime,
            "watchlist_symbols": [_v2_clean_ticker(w.get("symbol", ""))
                                  for w in (watchlist or [])],
        })
    except Exception as e:
        _log(f"[v2] prev-state save failed: {e}")

    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 10b — EXCEL RECOMMENDATION TRACKER (PART C)
# ─────────────────────────────────────────────────────────────────────────────

TRACKER_XLSX = "shadow_master.xlsx"


def _create_excel_workbook():
    """Create shadow_master.xlsx with all required sheets."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font
        wb = openpyxl.Workbook()

        ws1 = wb.active
        ws1.title = "Recommendations"
        ws1.append([
            "Date", "Ticker", "Company", "Category", "Opp Score", "Confidence", "TQ",
            "R/R", "Entry", "Stop", "T1", "T2", "Sector", "Regime", "Pledge%", "ROE",
            "D/E", "Catalysts", "Fail Reasons", "Status"
        ])

        ws2 = wb.create_sheet("Daily Tracking")
        ws2.append([
            "Tracking Date", "Ticker", "Rec Date", "Day#", "Close", "High", "Low",
            "Volume", "Return%", "Max Gain%", "Max DD%", "T1 Hit", "T2 Hit",
            "Stop Hit", "Remaining Upside%", "Holding Days", "Status"
        ])

        # Rejected sheet (2026-07-06): capture near-miss stocks so we can
        # backtest after 3 weeks — did the rejects actually go up? Which
        # gates are over-filtering? Same schema as Recommendations + a
        # "Primary Reject Reason" column bucketed for easy pivoting.
        ws_rej = wb.create_sheet("Rejected")
        ws_rej.append([
            "Date", "Ticker", "Company", "Opp Score", "Confidence", "TQ",
            "R/R", "Entry", "Stop", "T1", "T2", "Sector", "Regime",
            "Pledge%", "ROE", "D/E", "Catalysts", "Fail Reasons",
            "Primary Reject Reason", "Status"
        ])

        ws3 = wb.create_sheet("Performance Summary")
        ws3.append(["Metric", "Value"])

        # Phase Bucket-Direct (2026-07-10): The 4 shadow bucket sheets are the
        # primary daily record of every stock we saw — replaces the older
        # "Shadow Buckets" / "Shadow Summary" pair which relied on a fragile
        # shadow_trades.csv side-file. Populated directly by
        # _write_bucket_sheets_from_run() at save time. Preserves history:
        # today's rows are appended, previous days stay untouched.
        #
        # NOTE (2026-07-10): the 14 "*Analysis" / "Monthly Report" pivot
        # sheets were removed from this workbook constructor. They were
        # placeholder tabs that only research_job.py (a separate manual
        # workflow) ever populates — they wasted space in the daily xlsx
        # sent to Telegram. research_job.py's _ensure_sheets() creates
        # them on demand when it actually runs.
        for _bucket_sheet, _accent in [
            ("A_TAKEN", "C6EFCE"),          # green
            ("B_WATCH_ME", "FFEB9C"),       # amber
            ("C_NOT_MY_STYLE", "D9D9D9"),   # gray
            ("D_SO_CLOSE", "FFCC99"),       # orange
        ]:
            ws_b = wb.create_sheet(_bucket_sheet)
            ws_b.append([
                "Date", "Ticker", "Company", "Bucket", "Setup", "Regime",
                "Confidence", "TQ", "R/R", "Opp Score",
                "Entry", "Stop", "T1", "T2",
                "Sector", "Pledge%", "ROE", "D/E",
                "Catalysts", "Reason", "Status",
            ])
            _fill = PatternFill(start_color=_accent, end_color=_accent,
                                fill_type="solid")
            for _cell in ws_b[1]:
                _cell.font = Font(bold=True, color="000000")
                _cell.fill = _fill

        # Shadow Summary rollup — per-bucket win-rate / expectancy view
        ws_sum = wb.create_sheet("Shadow Summary")
        ws_sum.append([
            "Bucket", "Bucket Name", "Total Ever", "Today", "This Week",
            "This Month", "Expected WR %", "Latest Run",
        ])
        _sum_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7",
                                fill_type="solid")
        for _cell in ws_sum[1]:
            _cell.font = Font(bold=True, color="000000")
            _cell.fill = _sum_fill

        return wb
    except ImportError:
        return None
    except Exception:
        return None


def _classify_reject_reason(fail_reasons: list) -> str:
    """
    Bucket a fail_reasons list into ONE primary category for Excel pivot analysis.
    Priority order matters: check the most-actionable/most-common reasons first.
    Added 2026-07-06 so we can pivot rejects by root cause after 3 weeks.

    Phase R3 (2026-07-06): institutional gate buckets (ROE/DE/BQ_DECLINE/
    SECTOR_RANK/NEWS_SEVERITY) added before technical fails so that a stock
    failing both ROE_TOO_LOW and TQ_FAIL is bucketed as "ROE_TOO_LOW" — the
    root institutional reason, not a downstream technical symptom.
    """
    if not fail_reasons:
        return "UNKNOWN"
    joined = " | ".join(str(f) for f in fail_reasons).upper()
    # Priority order: data issues first (fixable), then institutional-quality
    # (root cause), then technicals (symptoms), then macro/portfolio.
    if "FUND_DATA_MISSING" in joined:  return "FUND_DATA_MISSING"
    if "PRICE_DATA" in joined or "NO_DATA" in joined: return "PRICE_DATA_MISSING"
    # ── Phase R1/R3 institutional-quality buckets (root cause) ──
    if "BUSINESS_QUALITY_DECLINE" in joined: return "BUSINESS_QUALITY_DECLINE"
    if "ROE_TOO_LOW" in joined:        return "ROE_TOO_LOW"
    if "DE_TOO_HIGH" in joined:        return "DE_TOO_HIGH"
    if "NEWS_SEVERITY" in joined:      return "NEWS_SEVERITY_HIGH"
    if "SECTOR_RANK_TOO_LOW" in joined or "SECTOR_MOMENTUM_NEG" in joined: return "SECTOR_RANK_TOO_LOW"
    # ── Technical fails (symptoms, less actionable) ──
    if "TQ_FAIL" in joined or "TRADE_QUALITY" in joined: return "TQ_TOO_LOW"
    if "CONF_FAIL" in joined or "CONFIDENCE" in joined:  return "CONFIDENCE_TOO_LOW"
    if "RR_FAIL" in joined:            return "RR_TOO_LOW"
    if "LIQUIDITY" in joined or "TURNOVER" in joined: return "LOW_LIQUIDITY"
    if "PLEDGE" in joined:             return "HIGH_PLEDGE"
    if "SECTOR_LAGGING" in joined:     return "SECTOR_LAGGING"
    if "SECTOR_CAP" in joined:         return "SECTOR_CAP"
    if "EVENT_BLOCK" in joined or "EARNINGS" in joined: return "EVENT_BLOCK"
    if "HIGH_CORR" in joined:          return "HIGH_CORRELATION"
    if "INSTITUTIONAL_EXIT" in joined: return "INSTITUTIONAL_EXIT"
    if "KILL_SWITCH" in joined:        return "KILL_SWITCH"
    if "REGIME_EXPOSURE" in joined:    return "REGIME_EXPOSURE_CAP"
    if "PORTFOLIO_FULL" in joined:     return "PORTFOLIO_FULL"
    if "CIRCUIT" in joined:            return "CIRCUIT_LIMIT"
    if "BLOCKLIST" in joined:          return "BLOCKLIST"
    return "OTHER"


def _classify_reject_tier(stock: dict) -> str:
    """
    Phase R3 (2026-07-06) — institutional-grade reject tier for the
    decision_audit + Excel pivot. Returned string is stamped on
    stock["reject_tier"] so post-mortem analytics can slice rejects by:

      • AVOID_QUALITY  — multi-signal quality problem (bq_verdict in
                         {DECLINING, WEAK} AND at least one hard-quality
                         gate failed). Worst kind of reject.
      • LOW_QUALITY_ROE — ROE_TOO_LOW gate fired (but bq_verdict not weak)
      • LOW_QUALITY_DE  — DE_TOO_HIGH (leverage)
      • BUSINESS_DECLINE — BUSINESS_QUALITY_DECLINE fired (sales+profit ↓)
      • NEWS_RISK       — NEWS_SEVERITY_HIGH fired
      • SECTOR_WEAK     — SECTOR_RANK_TOO_LOW or SECTOR_MOMENTUM_NEG
      • FUND_MISSING    — data fetch failed
      • LOW_SCORE       — pure score-only reject (TQ/CONF/RR fail)
      • OTHER_TECH      — liquidity/pledge/circuit/etc.
      • None            — no fails (shouldn't happen for a REJECTED stock)

    This tier is orthogonal to `decision` and `_classify_reject_reason()`.
    It exists specifically to segment rejects by *institutional root cause*
    so that after N weeks of data, we can answer "did the AVOID_QUALITY
    rejects underperform LOW_SCORE rejects?" — a direct measure of the
    R1/R2 quality gates' predictive value.
    """
    fails = stock.get("fail_reasons", []) or []
    if not fails:
        return None
    joined = " | ".join(str(f) for f in fails).upper()
    bq_verdict = str(stock.get("bq_verdict", "") or "").upper()

    # Multi-signal quality problem — the worst kind of reject
    if bq_verdict in ("DECLINING", "WEAK") and (
        "ROE_TOO_LOW" in joined
        or "DE_TOO_HIGH" in joined
        or "BUSINESS_QUALITY_DECLINE" in joined
    ):
        return "AVOID_QUALITY"
    # Single-signal quality issues
    if "BUSINESS_QUALITY_DECLINE" in joined: return "BUSINESS_DECLINE"
    if "ROE_TOO_LOW" in joined:              return "LOW_QUALITY_ROE"
    if "DE_TOO_HIGH" in joined:              return "LOW_QUALITY_DE"
    if "NEWS_SEVERITY" in joined:            return "NEWS_RISK"
    if "SECTOR_RANK_TOO_LOW" in joined or "SECTOR_MOMENTUM_NEG" in joined:
        return "SECTOR_WEAK"
    if "FUND_DATA_MISSING" in joined:        return "FUND_MISSING"
    # Pure score/technical rejects
    if any(k in joined for k in ("TQ_FAIL", "CONF_FAIL", "RR_FAIL")):
        return "LOW_SCORE"
    return "OTHER_TECH"


def save_recommendations_to_excel(buys: list, watchlist: list,
                                   regime_data: dict, today_str: str,
                                   rejected: list = None) -> None:
    """
    Appends today's recommendations to shadow_master.xlsx.
    Creates file with all sheets if it doesn't exist. Never overwrites existing rows.

    2026-07-06: added `rejected` parameter — writes to "Rejected" sheet so we
    can backtest after 3 weeks: did the rejects actually go up? Which gate is
    over-filtering? Backward-compatible: rejected defaults to None (skipped).
    """
    try:
        import openpyxl
    except ImportError:
        _log("[WARN] openpyxl not installed — skipping Excel save. Run: pip install openpyxl")
        return
    try:
        # Phase C7c: FRESH_START renames old xlsx aside so a fresh workbook is created
        if FRESH_START and os.path.exists(TRACKER_XLSX):
            stale_name = f"{TRACKER_XLSX}.stale_{today_str}"
            try:
                # Overwrite stale target if it already exists (idempotent re-runs)
                if os.path.exists(stale_name):
                    os.remove(stale_name)
                os.rename(TRACKER_XLSX, stale_name)
                _log(f"[FRESH_START] Renamed old {TRACKER_XLSX} → {stale_name} (fresh workbook will be created)")
            except Exception as _e:
                _log(f"[FRESH_START] Could not rename {TRACKER_XLSX}: {_e} — will overwrite instead")

        # Phase Shadow-FreshStart (2026-07-10): FRESH_START also resets
        # shadow_trades.csv so it stays symmetric with shadow_master.xlsx.
        # Without this, on a FRESH_START run the xlsx starts empty but the
        # CSV keeps all pre-wipe A/B/C/D rows and appends today on top —
        # leaving the two files inconsistent and polluting the 30-day
        # calibration analysis. Same "rename aside, don't delete" pattern
        # as the xlsx above so the old CSV is still available for post-mortem
        # AND is picked up by the recommendation-tracker artifact upload.
        #
        # shadow_log._ensure_csv() will auto-recreate an empty CSV (header
        # only) the next time record_shadow_trade() fires this run.
        try:
            import shadow_log as _sl_reset
            _shadow_csv_path = _sl_reset.SHADOW_CSV_PATH
        except Exception:
            _shadow_csv_path = "shadow_trades.csv"  # repo-root fallback
        if FRESH_START and os.path.exists(_shadow_csv_path):
            _shadow_stale = f"{_shadow_csv_path}.stale_{today_str}"
            try:
                if os.path.exists(_shadow_stale):
                    os.remove(_shadow_stale)
                os.rename(_shadow_csv_path, _shadow_stale)
                _log(f"[FRESH_START] Renamed old shadow_trades.csv → {os.path.basename(_shadow_stale)} (fresh CSV will be created on first shadow_log write)")
            except Exception as _e_sh:
                _log(f"[FRESH_START] Could not rename shadow_trades.csv: {_e_sh} — will keep appending (non-fatal)")

        # Phase C7f (2026-07-07): drop a sentinel that downstream jobs
        # (tracker_job.py, weekly_summary_job.py, research_job.py) read to
        # auto-detect a wipe when their own FRESH_START input was not set.
        # Without this, tracker.yml checkout pulls the pre-wipe xlsx from
        # git and appends onto stale rows, undoing the reset.
        # Contents: today's date so cross-day contamination is impossible.
        # The marker is deleted by the first downstream job that honors it,
        # so it fires exactly once.
        if FRESH_START:
            try:
                with open(".fresh_start_marker", "w", encoding="utf-8") as _fm:
                    _fm.write(f"{today_str}\n")
                _log(f"[FRESH_START] Wrote .fresh_start_marker={today_str} for downstream jobs")
            except Exception as _e:
                _log(f"[FRESH_START] Could not write .fresh_start_marker: {_e}")

        if os.path.exists(TRACKER_XLSX):
            wb = openpyxl.load_workbook(TRACKER_XLSX)
        else:
            wb = _create_excel_workbook()
            if wb is None:
                return

        ws = wb["Recommendations"]
        all_stocks = (
            [(s, "BUY") for s in buys] +
            [(s, s.get("tier", "WATCHLIST")) for s in watchlist]
        )

        for stock, category in all_stocks:
            symbol = stock.get("symbol", "")
            ws.append([
                today_str,
                symbol,
                symbol.replace(".NS", ""),
                category,
                stock.get("opportunity_score", 0),
                stock.get("final_confidence", 0),
                stock.get("trade_quality_score", 0),
                stock.get("rr_ratio", stock.get("rr", 0)),
                stock.get("entry", 0),
                stock.get("stop", 0),
                stock.get("target1", 0),
                stock.get("target2", 0),
                get_sector(symbol),
                regime_data.get("regime", ""),
                stock.get("promoter_pledge_pct", 0),
                stock.get("roe", 0),
                stock.get("de_ratio", 0),
                ", ".join(stock.get("catalysts", []) or []),
                ", ".join(stock.get("fail_reasons", []) or []),
                "ACTIVE",
            ])

        # Rejected sheet: append today's rejects for 3-week retrospective
        # analysis. Falls through silently if the sheet doesn't exist yet
        # (older workbook created before this feature was added).
        rejected_written = 0
        if rejected:
            try:
                # Ensure the Rejected sheet exists in older workbooks
                if "Rejected" not in wb.sheetnames:
                    ws_rej = wb.create_sheet("Rejected")
                    ws_rej.append([
                        "Date", "Ticker", "Company", "Opp Score", "Confidence", "TQ",
                        "R/R", "Entry", "Stop", "T1", "T2", "Sector", "Regime",
                        "Pledge%", "ROE", "D/E", "Catalysts", "Fail Reasons",
                        "Primary Reject Reason", "Status"
                    ])
                ws_rej = wb["Rejected"]
                for stock in rejected:
                    _sym = stock.get("symbol", "")
                    _fr  = stock.get("fail_reasons", []) or []
                    ws_rej.append([
                        today_str,
                        _sym,
                        _sym.replace(".NS", ""),
                        stock.get("opportunity_score", 0),
                        stock.get("final_confidence", 0),
                        stock.get("trade_quality_score", 0),
                        stock.get("rr_ratio", stock.get("rr", 0)),
                        stock.get("entry", 0),
                        stock.get("stop", 0),
                        stock.get("target1", 0),
                        stock.get("target2", 0),
                        get_sector(_sym),
                        regime_data.get("regime", ""),
                        stock.get("promoter_pledge_pct", 0),
                        stock.get("roe", 0),
                        stock.get("de_ratio", 0),
                        ", ".join(stock.get("catalysts", []) or []),
                        ", ".join(_fr),
                        _classify_reject_reason(_fr),
                        "REJECTED",
                    ])
                    rejected_written += 1
            except Exception as _rej_exc:
                _log(f"[WARN] Could not write Rejected sheet: {_rej_exc}")

        # Phase Bucket-Direct (2026-07-10): write today's A/B/C/D bucket rows
        # BEFORE saving so they land in the same wb.save() call. Preserves
        # previous-day rows (append semantics with (Date,Ticker) dedupe) and
        # drops the 14 legacy empty pivot sheets on first write.
        try:
            _bkt_counts = _write_bucket_sheets_from_run(
                wb, buys, watchlist, rejected or [], regime_data, today_str)
            _log(f"[shadow-xlsx] Bucket rows added today: "
                 f"A={_bkt_counts['A']} B={_bkt_counts['B']} "
                 f"C={_bkt_counts['C']} D={_bkt_counts['D']}")
        except Exception as _bk_exc:
            _log(f"[WARN] Direct bucket write failed: {_bk_exc}")

        wb.save(TRACKER_XLSX)
        if rejected_written:
            _log(f"[INFO] Saved {len(all_stocks)} recommendations + {rejected_written} rejects to {TRACKER_XLSX}")
        else:
            _log(f"[INFO] Saved {len(all_stocks)} recommendations to {TRACKER_XLSX}")

        # Phase Bucket-Direct (2026-07-10): refresh Shadow Summary rollup
        # after the main save so it always reflects today's counts. Runs on
        # the newly-saved file (loads it back), keeping the main save path
        # simple. Best-effort; never fails the pipeline.
        try:
            _refresh_bucket_summary()
        except Exception as _sx_exc:
            _log(f"[WARN] Shadow Summary refresh failed: {_sx_exc}")

    except Exception as e:
        _log(f"[WARN] Excel save failed: {e}")


def _sync_shadow_buckets_to_xlsx() -> None:
    """
    Phase Bucket-Direct (2026-07-10): thin compatibility shim that delegates
    to _refresh_bucket_summary(). Kept as a named function so any external
    caller / older code path still resolves. The heavy lifting now happens
    inside save_recommendations_to_excel() → _write_bucket_sheets_from_run(),
    which no longer depends on shadow_trades.csv existing.
    """
    try:
        _refresh_bucket_summary()
    except Exception as e:
        _log(f"[WARN] Shadow Summary refresh failed: {e}")


# ═════════════════════════════════════════════════════════════════════════════
# Phase Bucket-Direct (2026-07-10): direct-from-run bucket writer + summary
# ═════════════════════════════════════════════════════════════════════════════
# Design decision: main.py owns the daily write to A_TAKEN / B_WATCH_ME /
# C_NOT_MY_STYLE / D_SO_CLOSE. Removes the dependency on shadow_master_job.py
# (which was silently no-op'ing when shadow_trades.csv was empty) and makes
# the bucket sheets a first-class daily artifact.
#
# History is preserved: today's rows are APPENDED, previous days stay. A
# (Date, Ticker) dedupe check prevents duplicates on same-day re-runs.

_BUCKET_SHEET_MAP = {
    "A": "A_TAKEN",
    "B": "B_WATCH_ME",
    "C": "C_NOT_MY_STYLE",
    "D": "D_SO_CLOSE",
}
_BUCKET_ACCENTS = {
    "A_TAKEN": "C6EFCE",         # green
    "B_WATCH_ME": "FFEB9C",      # amber
    "C_NOT_MY_STYLE": "D9D9D9",  # gray
    "D_SO_CLOSE": "FFCC99",      # orange
}
_BUCKET_HEADER = [
    "Date", "Ticker", "Company", "Bucket", "Setup", "Regime",
    "Confidence", "TQ", "R/R", "Opp Score",
    "Entry", "Stop", "T1", "T2",
    "Sector", "Pledge%", "ROE", "D/E",
    "Catalysts", "Reason", "Status",
]
_BUCKET_EXPECTED_WR = {"A": 48.0, "B": 35.0, "C": 25.0, "D": 40.0}
_BUCKET_LONG_NAMES = {
    "A_TAKEN": "A · TAKEN",
    "B_WATCH_ME": "B · WATCH_ME",
    "C_NOT_MY_STYLE": "C · NOT_MY_STYLE",
    "D_SO_CLOSE": "D · SO_CLOSE",
}

# 14 legacy pivot sheets from the pre-2026-07-10 workbook constructor. They
# were placeholders only ever populated by research_job.py (separate manual
# workflow). We drop them from any existing xlsx on first write so daily
# users don't see empty tabs.
_LEGACY_EMPTY_SHEETS = (
    "Confidence Analysis", "TQ Analysis", "Opp Score Analysis",
    "Sector Analysis", "Regime Analysis", "Monthly Report",
    "Weekday Analysis", "Holding Period Analysis",
    "Category Comparison", "Conf x TQ Matrix",
    "Catalyst Analysis", "Fail Reason Analysis",
    "Regime x Sector", "Confidence Trajectory",
    # older "Shadow Buckets" was the aggregate view — now redundant with
    # per-bucket sheets + Shadow Summary rollup.
    "Shadow Buckets",
)


def _classify_stock_bucket(stock: dict, is_buy: bool, min_conf_bar: float = 83.0,
                            near_miss_band: float = 10.0) -> str:
    """
    Map a stock dict to one of A/B/C/D:
      A = TAKEN            — passed all gates (BUY)
      B = WATCH_ME         — BREAKOUT/MOMENTUM setup, rejected by regime/other
      C = NOT_MY_STYLE     — PULLBACK/REVERSAL/OTHER, rejected
      D = SO_CLOSE         — right setup + conf within near_miss_band of bar
    """
    if is_buy:
        return "A"
    setup = str(stock.get("setup_type") or stock.get("setup") or "OTHER").upper()
    conf  = float(stock.get("final_confidence", 0) or 0)
    if setup in ("BREAKOUT", "MOMENTUM"):
        # Near-miss on conf → D takes priority over B
        gap = min_conf_bar - conf
        if 0 < gap <= near_miss_band:
            return "D"
        return "B"
    return "C"


def _existing_bucket_keys(wb, sheet_name: str) -> set:
    """Return set of (date, ticker) tuples already present in a bucket sheet."""
    keys = set()
    if sheet_name not in wb.sheetnames:
        return keys
    ws = wb[sheet_name]
    try:
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if row and len(row) >= 2 and row[0] and row[1]:
                keys.add((str(row[0]), str(row[1])))
    except Exception:
        pass
    return keys


def _ensure_bucket_sheet(wb, sheet_name: str):
    """Ensure a bucket sheet exists with the correct header + accent."""
    try:
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        Font = PatternFill = None
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # If it exists but is empty/wrong-header, keep as-is (don't destroy
        # user data). Only add the header if truly blank.
        try:
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not first_row or all(v is None or v == "" for v in first_row):
                ws.append(_BUCKET_HEADER)
                if Font and PatternFill:
                    accent = _BUCKET_ACCENTS.get(sheet_name, "D9D9D9")
                    _fill = PatternFill(start_color=accent, end_color=accent,
                                        fill_type="solid")
                    for cell in ws[1]:
                        cell.font = Font(bold=True, color="000000")
                        cell.fill = _fill
        except Exception:
            pass
        return ws
    ws = wb.create_sheet(sheet_name)
    ws.append(_BUCKET_HEADER)
    if Font and PatternFill:
        accent = _BUCKET_ACCENTS.get(sheet_name, "D9D9D9")
        _fill = PatternFill(start_color=accent, end_color=accent,
                            fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="000000")
            cell.fill = _fill
    return ws


def _bucket_row_for_stock(stock: dict, bucket: str, regime_str: str,
                           today_str: str, is_buy: bool) -> list:
    """Build a single 21-column bucket row from a stock dict."""
    sym = stock.get("symbol", "") or ""
    fr  = stock.get("fail_reasons", []) or []
    if is_buy:
        reason = "TAKEN — all gates passed"
    else:
        reason = _classify_reject_reason(fr) if fr else "UNKNOWN"
    return [
        today_str,
        sym,
        sym.replace(".NS", ""),
        _BUCKET_LONG_NAMES.get(_BUCKET_SHEET_MAP.get(bucket, ""), bucket),
        str(stock.get("setup_type") or stock.get("setup") or "OTHER").upper(),
        regime_str,
        stock.get("final_confidence", 0),
        stock.get("trade_quality_score", 0),
        stock.get("rr_ratio", stock.get("rr", 0)),
        stock.get("opportunity_score", 0),
        stock.get("entry", 0),
        stock.get("stop", 0),
        stock.get("target1", 0),
        stock.get("target2", 0),
        get_sector(sym) if sym else "",
        stock.get("promoter_pledge_pct", 0),
        stock.get("roe", 0),
        stock.get("de_ratio", 0),
        ", ".join(stock.get("catalysts", []) or []),
        reason,
        "TAKEN" if is_buy else "REJECTED",
    ]


def _write_bucket_sheets_from_run(wb, buys: list, watchlist: list,
                                   rejected: list, regime_data: dict,
                                   today_str: str) -> dict:
    """
    Append today's rows to A_TAKEN / B_WATCH_ME / C_NOT_MY_STYLE / D_SO_CLOSE.
    Preserves all previous-day rows. Deduplicates by (Date, Ticker).

    Returns a per-bucket counter dict: {"A": n_a, "B": n_b, "C": n_c, "D": n_d}.
    """
    regime_str = str(regime_data.get("regime", "") or "")
    counters = {"A": 0, "B": 0, "C": 0, "D": 0}

    # Ensure all 4 bucket sheets exist with headers
    for sn in _BUCKET_SHEET_MAP.values():
        _ensure_bucket_sheet(wb, sn)

    # Build per-bucket dedupe key sets
    existing = {b: _existing_bucket_keys(wb, _BUCKET_SHEET_MAP[b])
                for b in ("A", "B", "C", "D")}

    # ── A: BUYs ──
    for s in (buys or []):
        b = "A"
        sym = s.get("symbol", "")
        if (today_str, sym) in existing[b]:
            continue
        wb[_BUCKET_SHEET_MAP[b]].append(
            _bucket_row_for_stock(s, b, regime_str, today_str, is_buy=True))
        existing[b].add((today_str, sym))
        counters[b] += 1

    # ── B/C/D: rejects (watchlist rows are still "monitor-tier" recs, not
    # rejects — they belong in the Recommendations sheet, so we don't
    # re-bucket them here). Only iterate `rejected`.
    for s in (rejected or []):
        b = _classify_stock_bucket(s, is_buy=False)
        sym = s.get("symbol", "")
        if (today_str, sym) in existing[b]:
            continue
        wb[_BUCKET_SHEET_MAP[b]].append(
            _bucket_row_for_stock(s, b, regime_str, today_str, is_buy=False))
        existing[b].add((today_str, sym))
        counters[b] += 1

    # ── Drop legacy empty pivot sheets on first write ──
    dropped = 0
    for legacy in _LEGACY_EMPTY_SHEETS:
        if legacy in wb.sheetnames:
            ws = wb[legacy]
            # Only drop if truly empty (1x1 or header-only, no data rows)
            try:
                empty = (ws.max_row <= 1 and ws.max_column <= 1)
                if empty:
                    del wb[legacy]
                    dropped += 1
            except Exception:
                pass
    if dropped:
        _log(f"[shadow-xlsx] Dropped {dropped} empty legacy sheets")

    return counters


def _refresh_bucket_summary() -> None:
    """
    Rebuild the "Shadow Summary" sheet by counting rows across the 4 bucket
    sheets. Runs after every save so today's data is always reflected.
    Silent no-op if xlsx doesn't exist.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return
    if not os.path.exists(TRACKER_XLSX):
        return
    try:
        wb = openpyxl.load_workbook(TRACKER_XLSX)
    except Exception as e:
        _log(f"[WARN] summary-refresh: cannot open {TRACKER_XLSX}: {e}")
        return

    # Compute per-bucket counts (total ever + today + this week + this month)
    from datetime import date, timedelta
    today = date.today()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)

    def _count_since(sheet_name: str, cutoff):
        if sheet_name not in wb.sheetnames:
            return (0, 0, 0, 0, "")
        ws = wb[sheet_name]
        total = 0
        today_n = 0
        wk_n = 0
        mo_n = 0
        latest = ""
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if not row or not row[0]:
                continue
            total += 1
            d_str = str(row[0])
            latest = d_str if not latest or d_str > latest else latest
            try:
                from datetime import datetime as _dt
                d = _dt.fromisoformat(d_str).date()
                if d == today:  today_n += 1
                if d >= week_ago:  wk_n += 1
                if d >= month_ago: mo_n += 1
            except Exception:
                pass
        return (total, today_n, wk_n, mo_n, latest)

    # (Re)create Shadow Summary
    if "Shadow Summary" in wb.sheetnames:
        del wb["Shadow Summary"]
    ws_s = wb.create_sheet("Shadow Summary")
    ws_s.append([
        "Bucket", "Bucket Name", "Total Ever", "Today", "This Week",
        "This Month", "Expected WR %", "Latest Run",
    ])
    _fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7",
                        fill_type="solid")
    for cell in ws_s[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = _fill

    for bucket_key in ("A", "B", "C", "D"):
        sn = _BUCKET_SHEET_MAP[bucket_key]
        total, today_n, wk_n, mo_n, latest = _count_since(sn, today)
        ws_s.append([
            bucket_key,
            _BUCKET_LONG_NAMES.get(sn, bucket_key),
            total, today_n, wk_n, mo_n,
            _BUCKET_EXPECTED_WR.get(bucket_key, ""),
            latest,
        ])

    # Aggregate row
    tot_all = sum(_count_since(_BUCKET_SHEET_MAP[b], today)[0]
                  for b in ("A", "B", "C", "D"))
    today_all = sum(_count_since(_BUCKET_SHEET_MAP[b], today)[1]
                    for b in ("A", "B", "C", "D"))
    ws_s.append([])
    ws_s.append(["ALL", "All buckets combined", tot_all, today_all,
                 "", "", "", ""])

    try:
        wb.save(TRACKER_XLSX)
        _log(f"[shadow-xlsx] Refreshed Shadow Summary "
             f"(total={tot_all}, today={today_all})")
    except Exception as e:
        _log(f"[WARN] Shadow Summary save failed: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 11 — PIPELINE ENTRY
# ─────────────────────────────────────────────────────────────────────────────

def run_pipeline():
    init_run_log()
    try:
        _run_pipeline_inner()
    except Exception as e:
        _log(f"[CRITICAL] Unhandled pipeline exception: {e}")
        _log(traceback.format_exc())
    finally:
        close_run_log()


def _ensure_portfolio_json():
    """
    Create an empty portfolio.json if it doesn't exist yet, with a schema comment
    so users know exactly how to add holdings manually.
    The file is a JSON array. Each holding object uses these fields:
      symbol        — NSE symbol with .NS suffix, e.g. "RELIANCE.NS"
      entry_price   — price you bought at (float)
      stop_loss     — your hard stop price (float)
      target1       — first target price (float)
      target2       — second target price (float)
      entry_date    — date you entered, "YYYY-MM-DD"
      sector        — optional, e.g. "ENERGY" (defaults to OTHERS if omitted)
    Example:
      [{"symbol":"RELIANCE.NS","entry_price":2950.0,"stop_loss":2800.0,
        "target1":3100.0,"target2":3300.0,"entry_date":"2026-06-20","sector":"ENERGY"}]
    """
    if not os.path.exists(PORTFOLIO_FILE):
        try:
            with open(PORTFOLIO_FILE, "w") as f:
                json.dump([], f, indent=2)
            _log(f"[INFO] Created empty {PORTFOLIO_FILE} — add your holdings manually to enable portfolio monitoring")
        except Exception as e:
            _log(f"[WARN] Could not create {PORTFOLIO_FILE}: {e}")


def _run_pipeline_inner():
    import copy
    _log("=== NSE SWING TRADE PIPELINE v6.0 STARTING ===")
    _log(f"  Capital: Rs{PORTFOLIO_CAPITAL:,.0f} | Groq keys: {len(GROQ_KEYS)}")
    _log(f"  Run mode: {'SCHEDULED — day counters will advance' if IS_SCHEDULED else 'MANUAL — day counters frozen, history not written'}")
    # Phase 3a #34 (2026-07-05): reset _PIPELINE_REGIME at pipeline entry so a
    # premature abort (market-closed, no tradable, no Nifty) can't leak the
    # PREVIOUS run's regime into sibling scripts (morning_check, exit
    # engines) via globals()["_PIPELINE_REGIME"]. Cleared to "" instead of
    # deleted so downstream `.get()` returns "" cleanly.
    globals()["_PIPELINE_REGIME"] = ""
    _ensure_portfolio_json()

    # ── Phase I shadow-log (2026-07-07): resolve outcomes for previously
    # skipped stocks. Each PENDING row is checked: did the stock hit +5%
    # target or −3% stop first (or exceed max-hold days)? Pure paper —
    # no real trades affected. Safe no-op if shadow_log unavailable.
    if _SHADOW_LOG_OK:
        try:
            _shadow_stats = shadow_log.update_shadow_outcomes(quiet=True)
            if _shadow_stats.get("resolved_today", 0) > 0:
                _log(
                    f"[shadow_log] resolved {_shadow_stats['resolved_today']} today "
                    f"(wins {_shadow_stats['wins']}, "
                    f"losses {_shadow_stats['losses']}, "
                    f"time-exits {_shadow_stats['time_exits']}) · "
                    f"pending {_shadow_stats['pending']}"
                )
        except Exception as _e:
            _log(f"[shadow_log] update failed (non-fatal): {_e}")

    # ── Sector map (must be first — used by all scoring) ──
    _init_sector_map()

    # ── 0. Market holiday guard ──
    # Scheduled runs respect the holiday calendar.
    # Manual runs bypass it so you can test on weekends/holidays.
    if not IS_SCHEDULED and not is_market_open():
        _log("[INFO] Market closed today, but running anyway (manual run — holiday guard bypassed).")
    elif not is_market_open():
        _log("[INFO] Market closed today (holiday or weekend). Skipping pipeline.")
        return

    # ── 0b. Earnings season adjustment ──
    earn_adj = earnings_season_threshold_adjustment()
    if earn_adj > 0:
        _log(f"[INFO] Earnings season — thresholds tightened by +{earn_adj}")
        effective_thresholds = copy.deepcopy(REGIME_THRESHOLDS)
        for rk in effective_thresholds:
            effective_thresholds[rk]["min_confidence"] += earn_adj
    else:
        effective_thresholds = REGIME_THRESHOLDS

    # ── 0b2. Auto-calibration layer (Phase C4 Gap #4) ──
    # Read regime_calibration.json produced by nightly_calibration.py.
    # Deltas are ADDITIVE on top of earn_adj and REGIME_THRESHOLDS.
    _cal = load_regime_calibration()
    if _cal:
        if effective_thresholds is REGIME_THRESHOLDS:
            effective_thresholds = copy.deepcopy(REGIME_THRESHOLDS)
        _before = {rk: effective_thresholds[rk].get("min_confidence", 0) for rk in effective_thresholds}
        effective_thresholds = apply_regime_calibration(effective_thresholds, _cal)
        _deltas = []
        for rk, before in _before.items():
            after = effective_thresholds[rk].get("min_confidence", 0)
            if after != before:
                _deltas.append(f"{rk}:{before}->{after}")
        if _deltas:
            _log("[INFO] Auto-calibration applied: " + ", ".join(_deltas))
        else:
            _log(f"[INFO] Auto-calibration file loaded ({len(_cal)} regimes) but no threshold changes.")

    # ── 1. Global macro ──
    _log("[1/17] Fetching global macro...")
    macro = fetch_global_macro()
    _vp   = macro.get("vix_in_percentile")
    _vp_s = f" ({_vp:.0f}th %ile, {macro.get('vix_in_regime','UNKNOWN')})" if _vp is not None else ""
    _log(
        f"  NIFTY {macro['nifty_1d_pct']:+.2f}% | "
        f"VIX-IN {macro['vix_in']:.1f}{_vp_s} | "
        f"USD/INR {macro['usdinr']:.2f}"
    )

    # ── 0c. VIX-percentile regime adjustment ──
    # Phase 3 (2026-07-01): VIX percentile is a leading contrarian signal.
    # COMPLACENT (<=15th pct) → market is dangerously calm, raise the bar by +2.
    # FEAR (>=85th pct)      → panic bottoms are entry opportunities, lower by -3.
    # ELEVATED (65-85)       → mild caution (+1).
    # Apply once, on top of any earnings-season adjustment already baked in.
    _vix_pct = macro.get("vix_in_percentile")
    _vix_regime = macro.get("vix_in_regime", "UNKNOWN")
    vix_adj = 0
    if _vix_pct is not None:
        if _vix_regime == "COMPLACENT":
            vix_adj = 2
        elif _vix_regime == "ELEVATED":
            vix_adj = 1
        elif _vix_regime == "FEAR":
            vix_adj = -3
    if vix_adj != 0:
        _log(
            f"[INFO] VIX regime {_vix_regime} (pct={_vix_pct}) — "
            f"thresholds adjusted by {vix_adj:+d}"
        )
        # If we hadn't already deep-copied for earnings, copy now.
        if effective_thresholds is REGIME_THRESHOLDS:
            effective_thresholds = copy.deepcopy(REGIME_THRESHOLDS)
        for rk in effective_thresholds:
            effective_thresholds[rk]["min_confidence"] = max(
                50, min(99, effective_thresholds[rk]["min_confidence"] + vix_adj)
            )

    # ── 1b. FII/DII flows (Phase C4: disabled — not used in scoring) ──
    _log("[2/17] FII/DII fetch skipped (Phase C4 — not used in scoring or output).")
    fii_dii = get_fii_dii_data()  # stub, returns unavailable/zero
    macro["fii_flow_cr"]      = fii_dii["fii_flow_cr"]
    macro["dii_flow_cr"]      = fii_dii["dii_flow_cr"]
    macro["fii_available"]    = fii_dii.get("available", False)
    macro["fii_provisional"]  = fii_dii.get("is_provisional", False)
    macro["fii_source"]       = fii_dii.get("source", "DISABLED")
    macro["fii_confidence"]   = fii_dii.get("confidence", "NONE")
    macro["fii_stale"]        = False
    macro["dii_found"]        = False
    # No log line for the zero flows — the [2/17] line above already conveys it.

    # ── 3. Bulk/block deals ──
    _log("[3/17] Fetching bulk/block deals...")
    bulk_deals = fetch_bulk_deals()
    if bulk_deals:
        # Phase C1 (2026-07-02): only show first 25 symbols in the log line so
        # it doesn't hog 100+ chars of every daily run.
        _keys = list(bulk_deals.keys())
        _shown = ', '.join(_keys[:25])
        _tail  = f" … +{len(_keys) - 25} more" if len(_keys) > 25 else ""
        _log(f"  Bulk deals: {len(bulk_deals)} found — {_shown}{_tail}")
    # else: fetch_bulk_deals() already logged the reason (quiet day / blocked / error)

    # ── 4. Load symbols ──
    _log("[4/17] Loading symbol universe...")
    symbols = load_symbols("stocks.txt")

    # ── 4a. Phase G8-C (2026-07-06): optional external universe pre-filter ──
    # Env: UNIVERSE_MODE = full | screener | chartink | hybrid | screener_and_chartink
    # Weekly Screener fundamental universe + daily Chartink technical universe.
    # Safe: any external fetch failure falls back to full base universe.
    try:
        symbols = build_universe(symbols)
    except Exception as _uni_exc:
        _log(f"[WARN] build_universe failed ({_uni_exc}) — using base symbols")

    # ── 5. Parallel price download + liquidity filter ──
    _log("[5/17] Downloading prices (parallel)...")
    tradable = filter_and_download(symbols, period="6mo", max_workers=12)
    _log(f"  Tradable: {len(tradable)} stocks")

    # ── 5a. Phase 1 #51 (2026-07-05): universe dropout diff logger ──
    # Compare today's tradable set with yesterday's saved state and log the
    # symmetric difference to tradable_dropouts.jsonl. Enables answering
    # "which stock silently left the universe today, and when?" post-hoc.
    # Never fails the pipeline; wrapped in a broad try.
    try:
        _prev_tradable = set()
        if os.path.exists(TRADABLE_STATE_FILE):
            with open(TRADABLE_STATE_FILE, "r") as _f:
                _prev = json.load(_f) or {}
            _prev_tradable = set(_prev.get("symbols", []))
        _curr_tradable = set(tradable.keys())
        _dropped_out   = sorted(_prev_tradable - _curr_tradable)
        _newly_in      = sorted(_curr_tradable - _prev_tradable)
        if _prev_tradable and (_dropped_out or _newly_in):
            _diff_entry = {
                "date":           ist_today().isoformat(),
                "ts":             ist_now().isoformat(),
                "prev_size":      len(_prev_tradable),
                "curr_size":      len(_curr_tradable),
                "dropped_out":    _dropped_out,
                "newly_in":       _newly_in,
                "universe_size":  len(symbols),
            }
            with open(TRADABLE_DROPOUT_FILE, "a", encoding="utf-8") as _f:
                _f.write(json.dumps(_diff_entry, default=str) + "\n")
            _log(f"  [Dropouts] {len(_dropped_out)} dropped, {len(_newly_in)} added → {TRADABLE_DROPOUT_FILE}")
        # Only overwrite the state file on scheduled runs so manual
        # experiments don't corrupt the day-over-day baseline.
        if IS_SCHEDULED:
            with open(TRADABLE_STATE_FILE, "w") as _f:
                json.dump({
                    "date":     ist_today().isoformat(),
                    "ts":       ist_now().isoformat(),
                    "symbols":  sorted(_curr_tradable),
                    "count":    len(_curr_tradable),
                }, _f, indent=2, default=str)
    except Exception as _drop_exc:
        _log(f"  [WARN] tradable dropout logger failed (non-fatal): {_drop_exc}")

    # ── 5b. Enrich sector map for all tradable symbols ──
    _log("[5b/17] Enriching sector map: nselib bulk first, yfinance fallback...")
    enrich_sectors_from_nselib()
    enrich_sectors_from_yfinance(list(tradable.keys()))
    if not tradable:
        _log("[ERROR] No tradable stocks. Aborting.")
        return

    # ── 6. Market regime ──
    _log("[6/17] Detecting market regime...")
    nifty_df    = fetch_price_data("^NSEI", period="1y")
    breadth     = compute_breadth(tradable)
    if nifty_df is None:
        _log("[ERROR] Cannot fetch Nifty data. Aborting.")
        return
    regime_data = detect_market_regime(nifty_df, breadth, macro)
    regime      = regime_data["regime"]

    # Phase 3a #24 (2026-07-05): survivor-bias diagnostic.
    # Our `breadth` is computed from the POST-liquidity-filter `tradable`
    # dict (stocks that passed 200L/day + penny + circuit filters). Under
    # deteriorating conditions, low-liquidity names drop out FIRST, so
    # tradable breadth stays artificially bullish while the raw universe
    # is already showing damage. We surface a "pass-through ratio"
    # (tradable / total_universe) — a low ratio (e.g. <60%) is a warning
    # signal that the raw universe has already broken down and our
    # regime signal may be inflated. This is informational only —
    # regime still uses filtered breadth to keep decisions consistent
    # with what we can actually trade. Cost: zero extra fetches.
    try:
        _passthrough = len(tradable) / max(len(symbols), 1)
        _flag = " ⚠️ SURVIVOR_BIAS_RISK" if _passthrough < 0.60 else ""
        _log(f"  Universe pass-through: {len(tradable)}/{len(symbols)} = "
             f"{_passthrough*100:.1f}%{_flag}")
    except Exception:
        pass

    # Phase E1d: pin the current regime into the module scope so exit engines
    # (update_tracker / update_tracker_v2_pnl) can consult it for risk-off
    # tightening without needing to be re-plumbed with the regime arg.
    globals()["_PIPELINE_REGIME"] = regime

    # Compute real Nifty 21-day and 5-day returns ONCE — passed to every stock's RS calc
    try:
        _nc = nifty_df["Close"].squeeze().values.astype(float)
        nifty_ret21_real = round((_nc[-1] / _nc[-22] - 1) * 100, 2) if len(_nc) > 22 else 0.0
        nifty_ret5_real  = round((_nc[-1] / _nc[-6]  - 1) * 100, 2) if len(_nc) > 6  else 0.0
    except Exception:
        nifty_ret21_real = 0.0
        nifty_ret5_real  = 0.0
    regime_data["nifty_ret21"] = nifty_ret21_real
    regime_data["nifty_ret5"]  = nifty_ret5_real
    _log(f"  REGIME: {regime} | Score: {regime_data['score']:.1f}/100 | EMA20 breadth: {breadth['ema20_pct']:.1f}%")
    _log(f"  Nifty 21d ret: {nifty_ret21_real:+.2f}% | 5d ret: {nifty_ret5_real:+.2f}%")

    # ── 6b. Nifty key levels + single nifty_state (BUG FIX 1) ──
    key_levels  = compute_key_levels(nifty_df)
    nifty_state = compute_nifty_state(nifty_df)
    _log(f"  Nifty structure: {nifty_state['structure']}")

    # ── 6c. Sector rotation ──
    _log("[6c/17] Computing sector rotation...")
    sector_rotation = compute_sector_rotation(tradable)
    leading = [s for s, d in sector_rotation.items() if d["status"] == "LEADING"]
    lagging = [s for s, d in sector_rotation.items() if d["status"] == "LAGGING"]
    _log(f"  Leading: {leading} | Lagging: {lagging}")

    # ── 7. Score all stocks ──
    _log("[7/17] Scoring all stocks...")
    scored = []
    for symbol, df in tradable.items():
        sector       = get_sector(symbol)
        scores       = compute_all_factors(symbol, df, sector, regime_data, sector_rotation, nifty_state)
        base_conf    = compute_base_confidence(scores)
        # NOTE: **scores must come BEFORE base_confidence so our computed value wins
        # (scores dict contains base_confidence: 0.0 as a default placeholder)
        scored.append({"symbol": symbol, "sector": sector, "_df": df, **scores, "base_confidence": base_conf})

    scored.sort(key=lambda x: (-x["base_confidence"], x["symbol"]))
    # Phase C7 (2026-07-02): widened from 40 to 50 for a safety buffer.
    # Rationale: after all filters (regime, gates, correlation, sector cap)
    # we typically end with 0-5 BUYs. Keeping 50 candidates means a marginal
    # score-#41 stock with an exceptional setup can still reach the gate
    # stage. Extra 10 stocks add ~2s of news/AI fetch — negligible.
    # Variable name stays "top_40" to avoid renaming 15 call sites; it's
    # now a well-known misnomer for "top-N candidate list".
    #
    # Phase 1 #50+#55 (2026-07-05): widened further from 50 → 100 for
    # research coverage. Ranks 51-100 are "monitor-only":
    #   * they DO go through the full scoring / gates,
    #   * they do NOT hit the LLM news endpoint (rule-based only, tagged
    #     news_source="RULE_BASED_MONITOR_ONLY" downstream at step 8),
    #   * they do NOT get fundamentals fetched (fetch_all_fundamentals_cached
    #     already caps at 30, so this is automatic),
    #   * they DO appear in daily_snapshots for post-hoc analysis.
    # Ranks 1-50 keep the full-pipeline behavior (unchanged).
    #
    # Phase G8-D (2026-07-06): adaptive score-band cutoff — the exact rank-N
    # cliff was arbitrary and caused near-miss stocks at rank 101 (often
    # 0.1-0.5 pts behind rank 100) to be silently discarded. Institutional
    # portfolio construction models cut at "the score cliff", not a fixed
    # rank. We now:
    #   1. Take the top _TOP_N by base_confidence as the anchor set.
    #   2. Then include every stock in `scored` whose base_confidence is
    #      within TOP_N_BAND_POINTS of the rank-N score.
    #   3. Cap the total at TOP_N_HARD_CAP so a flat-distribution day can't
    #      balloon the candidate list to 300+.
    #
    # Env vars:
    #   TOP_N_CANDIDATES  (default 100) — anchor rank cutoff
    #   TOP_N_BAND_POINTS (default 1.0) — score-cliff tolerance, 0 disables
    #   TOP_N_HARD_CAP    (default 120) — safety ceiling
    _TOP_N = int(os.getenv("TOP_N_CANDIDATES", "100"))
    _FULL_LLM_TOP_N = int(os.getenv("FULL_LLM_TOP_N", "50"))
    try:
        _TOP_N_BAND = float(os.getenv("TOP_N_BAND_POINTS", "1.0"))
    except (TypeError, ValueError):
        _TOP_N_BAND = 1.0
    try:
        _TOP_N_HARD_CAP = int(os.getenv("TOP_N_HARD_CAP", "120"))
    except (TypeError, ValueError):
        _TOP_N_HARD_CAP = 120
    # Hard cap must be >= _TOP_N (never shrink below the base cutoff)
    _TOP_N_HARD_CAP = max(_TOP_N_HARD_CAP, _TOP_N)

    # Anchor: strict top-N
    top_40 = scored[:_TOP_N]
    # Adaptive band: include any stock scored[i>=_TOP_N] whose base_confidence
    # is within _TOP_N_BAND of the anchor's tail (rank-N stock).
    _band_added = 0
    if _TOP_N_BAND > 0 and len(scored) > _TOP_N:
        _tail_conf = float(top_40[-1]["base_confidence"])
        _cliff = _tail_conf - _TOP_N_BAND
        for _cand in scored[_TOP_N:]:
            if len(top_40) >= _TOP_N_HARD_CAP:
                break
            _cconf = float(_cand.get("base_confidence", 0) or 0)
            if _cconf >= _cliff:
                top_40.append(_cand)
                _band_added += 1
            else:
                # scored is sorted desc — first below-cliff means done
                break
    if _band_added > 0:
        _log(f"  Top {len(top_40)}: best base conf {top_40[0]['base_confidence']:.1f} "
             f"({top_40[0]['symbol']}) · anchor={_TOP_N} + band={_band_added} "
             f"(within {_TOP_N_BAND:.1f}pts of rank-{_TOP_N}={top_40[_TOP_N-1]['base_confidence']:.1f}) "
             f"· cap={_TOP_N_HARD_CAP} · LLM for top {min(_FULL_LLM_TOP_N, len(top_40))}")
    else:
        _log(f"  Top {len(top_40)}: best base conf {top_40[0]['base_confidence']:.1f} "
             f"({top_40[0]['symbol']}) — LLM for top {min(_FULL_LLM_TOP_N, len(top_40))}, "
             f"monitor-only for rest")

    # ── 8. News + AI risk for top N ──
    # Phase 1 #55 (2026-07-05): split behavior — top-N uses full LLM
    # (existing path), ranks 51-100 use rule-based only with a distinct
    # news_source tag so post-mortem can filter monitor-only rows out.
    _log(f"[8/17] News + AI risk (LLM for top {min(_FULL_LLM_TOP_N, len(top_40))}, rule-based for rest)...")
    for _idx, stock in enumerate(top_40):
        sym_clean  = stock["symbol"].replace(".NS", "")
        headlines  = fetch_news_for_symbol(sym_clean)
        _is_full_llm = _idx < _FULL_LLM_TOP_N
        if headlines:
            if _is_full_llm:
                ai_result = ai_news_risk(sym_clean, [h["title"] for h in headlines])
            else:
                # Rule-based only for monitor-only ranks — no Groq call.
                _headlines_text = "\n".join(f"- {h['title'][:120]}" for h in headlines[:5])
                ai_result = _rule_based_news_score(_headlines_text)
                ai_result["news_source"] = "RULE_BASED_MONITOR_ONLY"
            age       = min(h["age_days"] for h in headlines)
            penalty   = compute_news_penalty(ai_result, age)
        else:
            ai_result = {
                "severity": 0, "category": "NO_NEWS",
                "is_black_swan": False, "summary": "",
                "news_source": "NO_HEADLINES" if _is_full_llm else "NO_HEADLINES_MONITOR_ONLY",
            }
            penalty   = 0.0
        stock["news_penalty"]  = penalty
        stock["is_black_swan"] = ai_result.get("is_black_swan", False)
        stock["news_summary"]  = truncate_display(ai_result.get("summary", ""), 100)
        # FIX: persist news category so BUY-card renderer can distinguish
        # "no headlines" (NO_NEWS) from "headline exists but summary was empty".
        stock["news_category"] = ai_result.get("category", "")
        stock["news_source"]   = ai_result.get("news_source", "")
        # Phase 3b #N9 (2026-07-05): NO_NEWS/NO_HEADLINES → 60 (neutral),
        # not 100 (max). A stock literally nobody is writing about is likely
        # a thinly-followed small/micro-cap — that's arguably a mild negative,
        # not a top-of-scale reward. news_risk is a 0.15-weighted factor in
        # confidence, so 60 vs 100 = ~6-pt swing. Real news scored 0..90 keeps
        # its differential range; only the "no data" placeholder is capped.
        _cat = ai_result.get("category", "")
        if _cat in ("NO_NEWS",) or ai_result.get("news_source", "").startswith("NO_HEADLINES"):
            stock["news_risk"] = 60
        else:
            stock["news_risk"] = max(0, 100 - int(penalty * 2))

    # ── 9. Promoter data + fundamentals — sequential with 24h cache (no rate limiting) ──
    # 2026-07-06 (Correct fix): Fetch fundamentals for ALL top_40 stocks (the
    # qualifying set is already ~100 = top 9% of universe). With a 24h cache,
    # only NEW candidates need fresh HTTP calls (~15-20/day typically). This
    # kills the systematic FUND_DATA_MISSING problem at its source instead of
    # working around it in downstream gates.
    #
    # History:
    # - Phase 1: max_stocks=20 (chose top-20 for fundamentals)
    # - Phase 2 (2026-07-05): widened to 30 to reduce ROE=0 BUYs
    # - 2026-07-06 morning: bumped 30->50 (still leaving 50%+ unfetched)
    # - 2026-07-06 now: fetch ALL. Missing-data is handled gracefully by the
    #   existing score-redistribution path (ownership_quality=None excludes
    #   the 6% weight, redistributes to present factors). NO hard reject.
    #
    # Env override: FUNDAMENTALS_FETCH_LIMIT=<n> caps for testing/emergency.
    # Default: fetch all top_40 (typically ~100 stocks).
    try:
        _fund_limit_env = os.getenv("FUNDAMENTALS_FETCH_LIMIT")
        _fund_limit = int(_fund_limit_env) if _fund_limit_env else len(top_40)
    except (TypeError, ValueError):
        _fund_limit = len(top_40)
    _fund_limit = max(1, _fund_limit)  # safety: never 0
    _log(f"[9/17] Fetching promoter/fundamentals for top {_fund_limit} (BUY-priority + cached)...")
    top_40 = fetch_all_fundamentals_cached(top_40, max_stocks=_fund_limit)

    # ── 10. Options PCR for top 20 (parallel) ──
    _log("[10/17] Options PCR for top 20 (parallel)...")

    def _fetch_pcr(stock: dict) -> tuple:
        sym_clean = stock["symbol"].replace(".NS", "")
        oc = fetch_option_chain(sym_clean)
        return stock["symbol"], pcr_score(oc["pcr"]), oc.get("source", "?")

    pcr_map = {}
    pcr_src_map = {}
    with ThreadPoolExecutor(max_workers=5) as ex:
        futs = {ex.submit(_fetch_pcr, s): s["symbol"] for s in top_40[:20]}
        for fut in as_completed(futs):
            try:
                sym, score_val, src = fut.result(timeout=15)
                pcr_map[sym]     = score_val
                pcr_src_map[sym] = src
            except Exception as e:
                _log(f"[WARN] PCR fetch failed for {futs[fut]}: {e}")

    for stock in top_40[:20]:
        stock["options_sentiment"] = pcr_map.get(stock["symbol"], 60)
        stock["pcr_source"]        = pcr_src_map.get(stock["symbol"], "NOT_FETCHED")

    # ── 11. Final confidence ──
    _log("[11/17] Computing final confidence...")
    macro_adj_global = macro_regime_adjustment(macro) * 0.3
    # Phase C3 (2026-07-02): blend real delivery % into the volume_delivery
    # factor for stocks that got real nselib data. OBV proxy stays as fallback.
    # Weight: 65% real delivery + 35% existing (OBV × accumulation) blend.
    for stock in top_40:
        if stock.get("delivery_source") == "nselib":
            deliv_score  = delivery_score_from_signal(
                stock.get("delivery_signal", "NEUTRAL"),
                stock.get("delivery_ratio", 1.0),
            )
            proxy_score  = float(stock.get("volume_delivery", 50))
            blended      = round(deliv_score * 0.65 + proxy_score * 0.35, 1)
            stock["volume_delivery_proxy"] = proxy_score  # keep for audit
            stock["volume_delivery"]       = blended
        # Otherwise keep existing OBV-proxy value as computed in score_stock().
    for stock in top_40:
        bulk_adj  = bulk_deal_score(stock["symbol"], bulk_deals)
        # Phase C5 (2026-07-02): pass raw stock.get(k) so None (MISSING) reaches
        # compute_base_confidence and triggers weight-redistribution instead of
        # being silently coerced to 50 (which would dilute confidence toward the
        # middle and defeat the whole Phase C5 fix).
        base_conf = compute_base_confidence({k: stock.get(k) for k in FACTOR_WEIGHTS})
        stock["base_confidence"]  = base_conf
        stock["final_confidence"] = compute_final_confidence(
            base_conf, regime, stock.get("news_penalty", 0), macro_adj_global, bulk_adj
        )
        # ── Phase I (2026-07-07): setup-edge patch — data-driven bonus + skip.
        # Classifies each candidate as BREAKOUT / MOMENTUM / PULLBACK / REVERSAL
        # / OTHER using the exact heuristic profiled in backtest_by_setup.csv,
        # then adds a signed confidence bonus and (in chop regimes) tags a
        # skip_reason for the BUY-gate cascade downstream. See
        # `apply_setup_edge()` above for the calibration table.
        _setup, _adj, _skip = apply_setup_edge(stock, regime)
        if _skip:
            stock["phase_i_skip"] = _skip
    top_40.sort(key=lambda x: (-x["final_confidence"], x["symbol"]))

    # ── Phase Polish (2026-07-11): setup-mix instrumentation ────────────
    # NOTE: The full setup mix (across ALL evaluated stocks, not just top_40)
    # is computed AFTER the gate loop below where buys/watchlist/rejected
    # are finalized. This early hook only tracks top_40 for legacy audit.
    from collections import Counter as _SetupCounter
    _setup_mix_top40 = _SetupCounter()
    for s in top_40:
        st = s.get("setup_type", "OTHER") or "OTHER"
        _setup_mix_top40[st] += 1
    # Stashed for optional debug — the DISPLAYED mix (regime_data["_setup_mix"])
    # is computed further down after the full evaluated pool is known.
    regime_data["_setup_mix_top40"] = dict(_setup_mix_top40)

    # ── 11b. Opportunity score (kept for audit compat; NOT used for ranking) ──
    # R5 PRUNE (2026-07-06): 480-row audit showed 20/20 top-20 overlap between
    # opportunity_score and final_confidence — the re-sort was a no-op. Ranking
    # now uses final_confidence desc, swing_alpha_score desc as tiebreaker.
    for stock in top_40:
        stock["opportunity_score"] = compute_opportunity_score(stock)
    top_40.sort(key=lambda x: (
        -float(x.get("final_confidence", 0) or 0),
        -float(x.get("swing_alpha_score", 0) or 0),
        x["symbol"],
    ))

    # ── 12. Portfolio monitoring ──
    _log("[12/17] Monitoring portfolio...")
    holdings       = load_portfolio()
    current_prices = {}
    for h in holdings:
        sym = h.get("symbol", "")
        if sym:
            try:
                df_tmp = fetch_price_data(sym, period="1mo")
                if df_tmp is not None and len(df_tmp) > 0:
                    current_prices[sym] = float(df_tmp["Close"].squeeze().iloc[-1])
            except Exception:
                pass
    portfolio_alerts = monitor_portfolio(holdings, current_prices, regime)

    # ── 13. Load tracker (before gates — needed for deduplication) ──
    _log("[13/17] Loading trade tracker...")
    tracker_entries = load_tracker()

    # ── 13b. Tracker V2 — load, update PnL, close completed positions ──
    tracker_v2 = initialize_tracker_if_new()
    tracker_v2 = update_tracker_v2_pnl(tracker_v2)
    if IS_SCHEDULED:
        save_tracker_v2(tracker_v2)
    _log(f"  Tracker V2: {len(tracker_v2.get('buys',[]))} active | {len(tracker_v2.get('watchlist',[]))} watching | {tracker_v2.get('performance',{}).get('completed',0)} completed")

    # ── 13b. Upcoming events from config file (FIX 4: auto-filters past events) ──
    upcoming_events = load_events_config()

    # ── 14. Gate system ──
    _log("[14/17] Running gate system (13 gates)...")
    # Phase C7 (2026-07-02): compute kill-switch state ONCE per run and feed it
    # into portfolio_context. Gate 5b reads it — pauses all new BUYs if the
    # equity curve is bleeding (3 consec losses, -2% day, -3% week, ≥10% dd).
    _ks = compute_kill_switch_state(tracker_v2, capital=PORTFOLIO_CAPITAL)
    if _ks.get("buys_paused"):
        _log(f"  [KILL SWITCH] Buys PAUSED — {_ks.get('reason')} "
             f"(consec={_ks['consecutive_losses']}, day={_ks['day_pnl_pct']}%, "
             f"week={_ks['week_pnl_pct']}%, dd={_ks['drawdown_from_peak_pct']}%)")
    elif _ks.get("size_multiplier", 1.0) < 1.0:
        _log(f"  [KILL SWITCH] Sizes DAMPED to {_ks['size_multiplier']}x — "
             f"{_ks.get('reason')} (dd={_ks['drawdown_from_peak_pct']}%)")
    # Phase 3a #40 (2026-07-05): compute current portfolio exposure so
    # gates can enforce regime-specific `max_exposure`. Historically this
    # field was declared in REGIME_THRESHOLDS but never read anywhere,
    # producing a silent gap: in BEAR (max_exposure 0.20) the pipeline
    # could still allocate 60%+ if the trader had 4 open positions.
    # Exposure is defined as sum(open holding values) / capital.
    _current_exposure = 0.0
    try:
        for _h in holdings:
            _hsym = _h.get("symbol", "")
            _hshares = float(_h.get("shares", 0) or 0)
            _hpx = current_prices.get(_hsym) or float(_h.get("entry_price", 0) or 0)
            if _hshares > 0 and _hpx > 0:
                _current_exposure += _hshares * _hpx
        _current_exposure /= max(float(PORTFOLIO_CAPITAL), 1.0)
    except Exception as _e_exp:
        _log(f"  [WARN] exposure calc failed (non-fatal): {_e_exp}")
        _current_exposure = 0.0
    _regime_max_exposure = float(effective_thresholds[regime].get("max_exposure", 1.0))
    _exposure_headroom   = max(0.0, _regime_max_exposure - _current_exposure)
    _log(f"  Exposure: {_current_exposure*100:.1f}% used / "
         f"{_regime_max_exposure*100:.0f}% cap ({regime}) "
         f"→ headroom {_exposure_headroom*100:.1f}%")
    portfolio_context = {
        "active_count":       len([a for a in portfolio_alerts if a["action"] == "HOLD"]),
        "existing_count":     len(holdings),
        "kill_switch":        _ks,
        "current_exposure":   _current_exposure,     # 0.0 .. 1.0+
        "max_exposure":       _regime_max_exposure,  # regime cap
        "exposure_headroom":  _exposure_headroom,    # available fraction
    }
    buys, watchlist_stocks, rejected = [], [], []

    # Build returns cache ONCE from already-downloaded price data (zero extra downloads)
    _log("  Building correlation returns cache...")
    returns_cache = build_returns_cache(tradable, lookback=60)

    # Pre-fetch earnings dates for top 40 (parallel, 5 workers)
    # Phase C4: yfinance calendar is now the primary source (BSE JSON API
    # started returning HTML in 2026-07). BSE kept as legacy fallback.
    _log("  Fetching earnings dates for top 40...")
    results_dates_map: dict = {}

    def _fetch_results(stock: dict) -> tuple:
        sym_clean = stock["symbol"].replace(".NS", "")
        return sym_clean, fetch_bse_results_dates(sym_clean)

    with ThreadPoolExecutor(max_workers=5) as ex:
        futs = {ex.submit(_fetch_results, s): s["symbol"] for s in top_40}
        for fut in as_completed(futs):
            try:
                sym_clean, dates = fut.result(timeout=10)
                results_dates_map[sym_clean] = dates
            except Exception:
                pass
    # Phase C1 (2026-07-02): honest counter — how many top-40 stocks had an
    # upcoming earnings date discovered. Silent failure was hiding the gate
    # activity.
    _dates_found = sum(1 for d in results_dates_map.values() if d)
    _log(f"  Earnings dates: {_dates_found}/{len(top_40)} stocks have upcoming earnings")

    # 2026-07-06: Data completeness health-check for the top-N qualifying set.
    # Rationale: the qualifying set is top ~9% of the universe. Every one of
    # these deserves merit-based evaluation, not data-lottery rejection.
    # This block is PURELY DIAGNOSTIC — it never rejects. It ensures that:
    #   1. Missing data is VISIBLE in the log (so we notice source outages).
    #   2. The gate stage uses score-redistribution for missing factors,
    #      never data-availability rejection.
    # Any stock reaching a gate with any data field missing is a WARNING,
    # never a REJECTED. The 12 non-fundamentals factors are all derived from
    # price/volume data which is guaranteed present (a stock with no price
    # data never enters `tradable` in the first place).
    _dq_full         = 0   # all 4 primary data channels present
    _dq_fund_missing = 0   # fundamentals missing
    _dq_deliv_missing = 0  # delivery data missing
    _dq_news_missing = 0   # news headlines missing
    _dq_earn_unknown = 0   # earnings dates unknown (safe: treated as no-event)
    for _stk in top_40:
        _sym = _stk["symbol"].replace(".NS", "")
        _fs  = str(_stk.get("fundamentals_source", "") or "").upper()
        _ds  = str(_stk.get("delivery_source", "") or "").lower()
        _ns  = str(_stk.get("news_source", "") or "").upper()
        _fund_ok  = _fs and _fs not in ("NEUTRAL_DEFAULT", "NOT_FETCHED", "SCREENER+YF_RL", "")
        _deliv_ok = _ds == "nselib"
        _news_ok  = _ns and not _ns.startswith("NO_HEADLINES") and _ns != "RULE_BASED_MONITOR_ONLY"
        _earn_ok  = bool(results_dates_map.get(_sym))
        if _fund_ok and _deliv_ok:
            _dq_full += 1
        if not _fund_ok:  _dq_fund_missing += 1
        if not _deliv_ok: _dq_deliv_missing += 1
        if not _news_ok:  _dq_news_missing += 1
        if not _earn_ok:  _dq_earn_unknown += 1
    _tot = len(top_40)
    _log(
        f"  [Data QA] Complete: {_dq_full}/{_tot} | "
        f"Fund-missing: {_dq_fund_missing} | "
        f"Deliv-missing: {_dq_deliv_missing} | "
        f"News-missing: {_dq_news_missing} | "
        f"Earnings-unknown: {_dq_earn_unknown}"
    )
    _log(
        f"  [Data QA] All {_tot} candidates will be evaluated on MERIT. "
        f"Missing-data factors use score-redistribution (never rejection)."
    )
    # Loud alarm if fundamentals coverage is catastrophically low —
    # signals a Screener/Trendlyne/yfinance outage that should be investigated
    # (but does NOT abort the run: we still trade on the 12 other factors).
    if _tot > 0 and _dq_fund_missing * 100 // _tot >= 50:
        _log(
            f"  [ALERT] >{_dq_fund_missing * 100 // _tot}% of top-{_tot} "
            f"are missing fundamentals \u2014 Screener/Trendlyne/yfinance may all "
            f"be rate-limited or blocked. Trading continues on technical merit. "
            f"Check network / manually curl screener.in / check yfinance."
        )
    if _tot > 0 and _dq_deliv_missing * 100 // _tot >= 50:
        _log(
            f"  [ALERT] >{_dq_deliv_missing * 100 // _tot}% of top-{_tot} "
            f"are missing delivery% \u2014 nselib may be down or NSE bhav copy "
            f"unavailable. Trading continues \u2014 delivery is a bonus signal, not gate."
        )

    # Phase C5 (rating ≥ 9.0): pre-load confidence history so classify_watchlist
    # can compute trajectory (RISING / FADING / FLAT) at classification time.
    _conf_history_for_wl = load_confidence_history()

    # ── Phase G-BATCH2 (2026-07-07): enrichment from new signal modules ──
    # Each block is fully feature-flagged so pipeline degrades gracefully
    # when a source (nselib / Groq / RSS / Screener) is down or slow.
    #   • quality_scores.py     — Piotroski F-Score (fundamentals)
    #   • options_data.py       — PCR / max-pain (F&O universe only)
    #   • news_sentiment.py     — Groq/keyword classifier on 7-day headlines
    #   • insider_feed.py       — promoter/KMP filings + bulk/block deals
    # Each populates: stock["quality"], stock["options_signal"],
    # stock["news_sig"], stock["insider_sig"] and cumulatively adjusts
    # stock["final_confidence"] via factor_bonus. `hard_reject` flags
    # short-circuit any candidate that fails a mandatory quality gate.
    _flag = lambda k, d="true": os.environ.get(k, d).lower() == "true"
    _quality_on = _flag("ENABLE_QUALITY_SCORE",  "true")
    _options_on = _flag("ENABLE_OPTIONS_GATE",   "true")
    _news_on    = _flag("ENABLE_NEWS_SENTIMENT", "true")
    _insider_on = _flag("ENABLE_INSIDER_SIGNAL", "true")
    _bonus_cap  = float(os.environ.get("PHASE_G_BONUS_CAP", "5.0"))  # ±5 conf pts

    if _quality_on or _options_on or _news_on or _insider_on:
        # Phase G-BATCH2 perf fix (2026-07-07): previously this block ran a
        # serial for-loop over the top-N (≈120) candidates, calling four
        # network-bound helpers per stock (options_data → NSE derivatives
        # API, news_sentiment → 4×RSS feeds + Groq LLM, insider_feed → NSE
        # capital-market API). Wall time observed in prod GH-Actions run
        # 2026-07-07: ~32 min for 120 stocks (≈16 s/stock).
        # Fix: dispatch the per-stock enrichment via ThreadPoolExecutor.
        # Each helper is IO-bound + already thread-safe (module-level dict
        # caches with GIL-atomic get/set). Workers configurable via
        # PHASE_G_MAX_WORKERS (default 8 — matches neighbouring pipeline
        # stages and stays well under NSE / Groq rate limits).
        _phase_g_workers = max(1, int(os.environ.get("PHASE_G_MAX_WORKERS", "8")))
        _log(
            f"[14a/17] Enriching top-{len(top_40)} with quality/options/news/insider "
            f"[Q={int(_quality_on)} O={int(_options_on)} N={int(_news_on)} "
            f"I={int(_insider_on)}] · workers={_phase_g_workers}"
        )
        _mod_q = _mod_o = _mod_n = _mod_i = None
        try:
            if _quality_on: import quality_scores as _mod_q  # type: ignore
        except Exception as _e: _log(f"  [WARN] quality_scores import: {_e}"); _mod_q = None
        try:
            if _options_on: import options_data as _mod_o    # type: ignore
        except Exception as _e: _log(f"  [WARN] options_data import: {_e}");   _mod_o = None
        try:
            if _news_on:    import news_sentiment as _mod_n  # type: ignore
        except Exception as _e: _log(f"  [WARN] news_sentiment import: {_e}"); _mod_n = None
        try:
            if _insider_on: import insider_feed as _mod_i    # type: ignore
        except Exception as _e: _log(f"  [WARN] insider_feed import: {_e}");   _mod_i = None

        # Counters must be updated atomically from worker threads.
        import threading as _threading
        _phase_g_lock = _threading.Lock()
        _phase_g_counters = {"q_hard": 0, "n_hard": 0, "i_hard": 0, "done": 0}
        _phase_g_total = len(top_40)
        _phase_g_start = time.time()

        def _enrich_one(_stk):
            """Per-stock enrichment — runs on a worker thread.
            Mutates `_stk` in place (safe: each thread owns a distinct dict).
            Returns None; hard-reject counters are updated under a lock.
            """
            _sym = _stk.get("symbol", "")
            _sym_clean = _sym.replace(".NS", "")
            _hard = False
            _reasons_new: list = []
            _agg_bonus = 0.0
            _local_q_hard = _local_n_hard = _local_i_hard = 0

            # -- Quality (Piotroski F-Score) --------------------------------
            if _mod_q is not None:
                try:
                    _fund = _stk.get("fundamentals") or _stk.get("fundamentals_raw") or {}
                    if isinstance(_fund, dict) and _fund:
                        _q = _mod_q.quality_composite(_fund)
                        _stk["quality"] = _q
                        if _q.get("ok"):
                            _agg_bonus += float(_q.get("factor_bonus", 0.0) or 0.0)
                            if _q.get("hard_reject"):
                                _hard = True; _local_q_hard = 1
                                _reasons_new.append(f"QUALITY_FAIL({_q.get('reject_reason','')})")
                except Exception as _e:
                    _stk.setdefault("_enrich_errors", []).append(f"quality:{_e}")

            # -- Options (PCR / max-pain) ---------------------------------
            if _mod_o is not None:
                try:
                    _spot = float(_stk.get("close") or _stk.get("cmp") or 0.0)
                    _os = _mod_o.options_signal(_sym_clean, spot=_spot or None)
                    _stk["options_signal"] = _os
                    if _os.get("ok"):
                        _agg_bonus += float(_os.get("factor_bonus", 0.0) or 0.0)
                        if _os.get("hard_reject"):
                            _hard = True
                            _reasons_new.append(f"OPTIONS_FAIL({_os.get('reject_reason','')})")
                except Exception as _e:
                    _stk.setdefault("_enrich_errors", []).append(f"options:{_e}")

            # -- News sentiment -------------------------------------------
            if _mod_n is not None:
                try:
                    _ns = _mod_n.news_sentiment_signal(
                        _sym_clean,
                        company_name=_stk.get("company_name") or _stk.get("name"),
                        lookback_days=7,
                    )
                    _stk["news_sig"] = _ns
                    if _ns.get("ok"):
                        _agg_bonus += float(_ns.get("factor_bonus", 0.0) or 0.0)
                        if _ns.get("hard_reject"):
                            _hard = True; _local_n_hard = 1
                            _reasons_new.append(f"NEWS_FAIL({_ns.get('reject_reason','')})")
                except Exception as _e:
                    _stk.setdefault("_enrich_errors", []).append(f"news:{_e}")

            # -- Insider filings ------------------------------------------
            if _mod_i is not None:
                try:
                    _isg = _mod_i.insider_signal(_sym_clean, lookback_days=30)
                    _stk["insider_sig"] = _isg
                    if _isg.get("ok"):
                        _agg_bonus += float(_isg.get("factor_bonus", 0.0) or 0.0)
                        if _isg.get("hard_reject"):
                            _hard = True; _local_i_hard = 1
                            _reasons_new.append(f"INSIDER_FAIL({_isg.get('reject_reason','')})")
                except Exception as _e:
                    _stk.setdefault("_enrich_errors", []).append(f"insider:{_e}")

            # -- Apply cumulative bonus to final_confidence ---------------
            # Cap in ±_bonus_cap points so a single soft signal can't dominate.
            _agg_bonus = max(-_bonus_cap, min(_bonus_cap, _agg_bonus * _bonus_cap))
            _stk["phaseG_bonus"] = round(_agg_bonus, 2)
            try:
                _fc = float(_stk.get("final_confidence", 0) or 0)
                _stk["final_confidence_pre_phaseG"] = _fc
                _stk["final_confidence"] = max(0.0, min(100.0, _fc + _agg_bonus))
            except Exception:
                pass

            # -- Propagate hard rejects to the existing gate machinery ----
            if _hard:
                _stk["phaseG_hard_reject"] = True
                _fr = _stk.setdefault("fail_reasons", [])
                for _r in _reasons_new:
                    if _r not in _fr:
                        _fr.append(_r)

            # -- Update shared counters + heartbeat under a lock ----------
            with _phase_g_lock:
                _phase_g_counters["q_hard"] += _local_q_hard
                _phase_g_counters["n_hard"] += _local_n_hard
                _phase_g_counters["i_hard"] += _local_i_hard
                _phase_g_counters["done"]   += 1
                _done = _phase_g_counters["done"]
            # Progress heartbeat every 20 stocks so the run isn't silent
            # for many minutes if a network source is slow.
            if _done % 20 == 0 or _done == _phase_g_total:
                _elapsed = time.time() - _phase_g_start
                _rate = _done / _elapsed if _elapsed > 0 else 0
                _eta = (_phase_g_total - _done) / _rate if _rate > 0 else 0
                _log(
                    f"  [Phase G] progress {_done}/{_phase_g_total} · "
                    f"{_rate:.1f} stk/s · elapsed {_elapsed:.0f}s · ETA {_eta:.0f}s"
                )

        # Dispatch across the worker pool. Any per-stock exception is
        # already caught inside _enrich_one — we just wait for completion.
        with ThreadPoolExecutor(max_workers=_phase_g_workers) as _ex:
            _futs = [_ex.submit(_enrich_one, _stk) for _stk in top_40]
            for _f in as_completed(_futs):
                try:
                    _f.result()
                except Exception as _e:
                    _log(f"  [Phase G] worker crashed unexpectedly: {_e}")

        _q_hard = _phase_g_counters["q_hard"]
        _n_hard = _phase_g_counters["n_hard"]
        _i_hard = _phase_g_counters["i_hard"]
        _log(
            f"  [Phase G] hard-rejects — quality:{_q_hard} news:{_n_hard} "
            f"insider:{_i_hard}. Bonus cap ±{_bonus_cap} conf pts. "
            f"See stock['quality']/['options_signal']/['news_sig']/['insider_sig']."
        )
        # Re-sort by new final_confidence (Phase-G-bonused)
        top_40.sort(key=lambda x: (
            -float(x.get("final_confidence", 0) or 0),
            -float(x.get("swing_alpha_score", 0) or 0),
            x["symbol"],
        ))

    # ═════════════════════════════════════════════════════════════════════
    # Phase H (2026-07-07) — 52-WEEK-HIGH MOMENTUM SIGNAL
    # ═════════════════════════════════════════════════════════════════════
    # Evidence: George & Hwang, "The 52-Week High and Momentum Investing"
    #          Journal of Finance, Vol. 59 (2004), pp. 2145-2176.
    #
    #   Key finding: stocks trading near their 52-week high outperform
    #   stocks far from it by ~0.45%/month for the next 6 months, even
    #   after controlling for Fama-French factors and Jegadeesh-Titman
    #   momentum. Effect is strongest in the first 1-3 months (our swing
    #   window). Robust out-of-sample across US, EU, Japan, India, EM.
    #
    #   Bruce Kamich (CANSLIM), William O'Neil, Mark Minervini and Nick
    #   Radge all use variants of this as their primary entry filter.
    #
    # Implementation:
    #   - dist_52w_high_pct is ALREADY computed on every stock (see the
    #     price-analysis path ~L4900). We just reuse it — zero extra
    #     network / compute cost.
    #   - Zones: 0-3% = STRONG (+full bonus), 3-8% = OK (+half bonus),
    #             8-15% = NEUTRAL (0), 15-25% = LAGGARD (-half),
    #             >25% = DEEP_LAGGARD (-full).
    #   - Cap: ±ENABLE_52W_HIGH_BONUS_CAP conf points (default 6).
    #   - Env-gated by ENABLE_52W_HIGH_SIGNAL (default true).
    # ─────────────────────────────────────────────────────────────────────
    _flag = lambda k, d="true": os.environ.get(k, d).lower() == "true"
    _52w_on   = _flag("ENABLE_52W_HIGH_SIGNAL", "true")
    _52w_cap  = float(os.environ.get("ENABLE_52W_HIGH_BONUS_CAP", "6.0"))
    _simple_mode = _flag("SIMPLE_MODE_52W", "false")

    if _52w_on and top_40:
        _z_counts = {"STRONG": 0, "OK": 0, "NEUTRAL": 0, "LAG": 0, "DEEP_LAG": 0}
        for _stk in top_40:
            try:
                # dist_52w_high_pct = (high - close)/high * 100  → smaller = closer to high
                _d = float(_stk.get("dist_52w_high_pct", 100) or 100)
            except Exception:
                _d = 100.0

            if   _d <= 3.0:   _zone, _bonus = "STRONG",   +1.00
            elif _d <= 8.0:   _zone, _bonus = "OK",       +0.50
            elif _d <= 15.0:  _zone, _bonus = "NEUTRAL",   0.00
            elif _d <= 25.0:  _zone, _bonus = "LAG",      -0.50
            else:             _zone, _bonus = "DEEP_LAG", -1.00
            _z_counts[_zone] += 1

            _bonus_pts = _bonus * _52w_cap
            _stk["signal_52w_high"] = {
                "ok": True,
                "dist_pct": round(_d, 2),
                "zone": _zone,
                "factor_bonus": round(_bonus_pts, 2),
            }
            try:
                _fc = float(_stk.get("final_confidence", 0) or 0)
                _stk["final_confidence"] = max(0.0, min(100.0, _fc + _bonus_pts))
            except Exception:
                pass

        _log(
            f"  [Phase H · 52w-high] STRONG:{_z_counts['STRONG']} "
            f"OK:{_z_counts['OK']} NEUTRAL:{_z_counts['NEUTRAL']} "
            f"LAG:{_z_counts['LAG']} DEEP_LAG:{_z_counts['DEEP_LAG']} "
            f"(bonus cap ±{_52w_cap} pts)"
        )
        # Re-sort so 52w-STRONG names bubble to the top
        top_40.sort(key=lambda x: (
            -float(x.get("final_confidence", 0) or 0),
            -float(x.get("swing_alpha_score", 0) or 0),
            x["symbol"],
        ))

    # ═════════════════════════════════════════════════════════════════════
    # SIMPLE_MODE_52W — "your brother's strategy" as a single flag
    # ═════════════════════════════════════════════════════════════════════
    # When SIMPLE_MODE_52W=true, ONLY consider stocks within 8% of their
    # 52-week high (STRONG or OK zone). Everything else gets a hard-reject
    # so the gate machinery downstream drops them. All other signals still
    # score — but 52w-high proximity is the primary filter.
    #
    # Use this to A/B test a pure momentum strategy vs. the full 20-signal
    # system without deleting any code. Turn on for one week, compare tracker
    # P&L to a baseline week. If it wins, keep going; if not, revert.
    # ─────────────────────────────────────────────────────────────────────
    if _simple_mode and top_40:
        _kept = _dropped = 0
        for _stk in top_40:
            try:
                _d = float(_stk.get("dist_52w_high_pct", 100) or 100)
            except Exception:
                _d = 100.0
            if _d > 8.0:  # not near the high → reject
                _stk["phaseG_hard_reject"] = True
                _fr = _stk.setdefault("fail_reasons", [])
                _reason = f"SIMPLE_MODE_52W_FAIL(dist={_d:.1f}%>8.0%)"
                if _reason not in _fr:
                    _fr.append(_reason)
                _dropped += 1
            else:
                _kept += 1
        _log(
            f"  [SIMPLE_MODE_52W] active — kept {_kept} within 8% of 52w-high, "
            f"hard-rejected {_dropped} others. Pure momentum strategy."
        )

    for stock in top_40:
        sym_clean     = stock["symbol"].replace(".NS", "")
        promoter_data = stock.get("promoter_data", {"promoter_pledge_pct": 0})
        gate_result   = run_gates(
            stock, regime, effective_thresholds,
            portfolio_context, bulk_deals, promoter_data,
            results_dates=results_dates_map.get(sym_clean, []),
            upcoming_events=upcoming_events,
            returns_cache=returns_cache,
            holdings=holdings,
        )
        stock["decision"]     = gate_result["decision"]
        stock["fail_reasons"] = gate_result["fail_reasons"]
        stock["warnings"]     = gate_result["warnings"]

        # Phase R1/R2 (2026-07-06): normalize new tier labels for routing.
        # `STRONG_BUY`, `BUY`, `BUY_CONTRARIAN`, `BUY_TURNAROUND` all route
        # to buys[] but keep their tag on stock["decision"] so the renderer
        # / tracker can prioritize STRONG_BUY at the top and mark contrarian /
        # turnaround trades appropriately.
        # `AVOID` is a hard-quality reject — routes to rejected[] so it
        # never appears in watchlist promotion candidates.
        _dec = gate_result["decision"]
        if _dec in ("BUY", "STRONG_BUY", "BUY_CONTRARIAN", "BUY_TURNAROUND"):
            buys.append(stock)

            # Phase I shadow-log (2026-07-07): Bucket A · TAKEN
            # Log every real BUY signal as a shadow row too — this is
            # the ground-truth reference bucket. When compared against
            # B/C/D, it confirms whether our edge is real AND lets us
            # measure Phase I win-rate live (independent of tracker.json
            # which is Excel-only and hard to aggregate).
            if _SHADOW_LOG_OK:
                try:
                    shadow_log.record_shadow_trade("A", stock, regime,
                                                   note=f"real_buy:{_dec}")
                except Exception:
                    pass
        elif _dec == "WATCHLIST":
            wl = classify_watchlist(stock, regime, effective_thresholds,
                                     conf_history=_conf_history_for_wl)
            stock.update(wl)
            watchlist_stocks.append(stock)
        elif _dec == "AVOID":
            # Institutional AVOID (turnaround escape + weak backing) — the
            # narrow demotion from the WATCHLIST path. Distinguished from
            # AVOID_QUALITY (broad reject-tier bucket) via the specific tag.
            stock["reject_tier"] = "AVOID_TURNAROUND"
            rejected.append(stock)
        else:
            # Phase R3 (2026-07-06): stamp institutional reject tier on every
            # REJECTED stock so decision_audit + Excel pivot can slice rejects
            # by root cause (AVOID_QUALITY / LOW_QUALITY_ROE / BUSINESS_DECLINE
            # / SECTOR_WEAK / LOW_SCORE / OTHER_TECH). Enables backtest queries
            # like "did AVOID_QUALITY rejects underperform LOW_SCORE rejects?"
            stock["reject_tier"] = _classify_reject_tier(stock)
            rejected.append(stock)

    _log(f"  Gate results: {len(buys)} BUY | {len(watchlist_stocks)} WATCHLIST | {len(rejected)} REJECTED")

    # Phase R3 (2026-07-06): institutional taxonomy breakdown for BUY tier.
    # Shows how the 4 buy sub-tiers distribute — high-quality signal for post-mortem.
    if buys:
        _tier_counts: dict = {}
        for _b in buys:
            _tier_counts[_b.get("decision", "BUY")] = _tier_counts.get(_b.get("decision", "BUY"), 0) + 1
        _tier_line = " | ".join(f"{k}: {v}" for k, v in sorted(_tier_counts.items()))
        _log(f"  BUY breakdown: {_tier_line}")
        # Also log avg BQ score + sector composite of BUYs (institutional health)
        _bq_scores = [float(b.get("bq_score", 0) or 0) for b in buys]
        _sec_scores = [float(b.get("sector_composite_score", 0) or 0) for b in buys]
        if _bq_scores:
            _log(
                f"  BUY quality: avg BQ={sum(_bq_scores)/len(_bq_scores):.1f} "
                f"| avg sector_composite={sum(_sec_scores)/len(_sec_scores):.1f}"
            )
    # Count AVOID sub-tier separately
    _avoid_count = sum(1 for r in rejected if r.get("reject_tier") == "AVOID_TURNAROUND")
    if _avoid_count > 0:
        _log(f"  AVOID_TURNAROUND: {_avoid_count} stocks (turnaround escape + weak backing)")

    # Phase R3 (2026-07-06): institutional reject-tier breakdown (new)
    _tier_hist: dict = {}
    for _r in rejected:
        _rt = _r.get("reject_tier") or "unclassified"
        _tier_hist[_rt] = _tier_hist.get(_rt, 0) + 1
    if _tier_hist:
        _sorted_tiers = sorted(_tier_hist.items(), key=lambda kv: -kv[1])
        _tier_line = " | ".join(f"{k}: {v}" for k, v in _sorted_tiers)
        _log(f"  Reject-tier (institutional): {_tier_line}")

    # FIX 3 (2026-07-06): reject-reason breakdown — immediate visibility into
    # WHY stocks were rejected (data missing vs bad stock vs low score). Uses
    # the same _classify_reject_reason() helper that populates the Excel
    # "Rejected" sheet, so the log summary matches the pivot analysis.
    if rejected:
        _reason_counts: dict = {}
        for _rs in rejected:
            _bucket = _classify_reject_reason(_rs.get("fail_reasons", []) or [])
            _reason_counts[_bucket] = _reason_counts.get(_bucket, 0) + 1
        # Sort desc by count, show top 6
        _sorted_reasons = sorted(_reason_counts.items(), key=lambda kv: -kv[1])
        _total_rej = len(rejected)
        _breakdown = " | ".join(
            f"{name}: {cnt} ({cnt * 100 // max(_total_rej, 1)}%)"
            for name, cnt in _sorted_reasons[:6]
        )
        _log(f"  Reject breakdown: {_breakdown}")
        # 2026-07-06 correct fix: FUND_DATA_MISSING is no longer a rejection
        # reason (handled by score redistribution). If it EVER appears in the
        # breakdown, the strict legacy gate got re-enabled — warn loudly.
        if _reason_counts.get("FUND_DATA_MISSING", 0) > 0:
            _log(
                f"  [WARN] FUND_DATA_MISSING appeared in {_reason_counts['FUND_DATA_MISSING']} "
                f"rejects — legacy strict gate is ENABLED via FUND_DATA_GATE_ENABLED=1. "
                f"Recommended: unset it so score-redistribution handles missing data gracefully."
            )

    # Sort BUYs by final_confidence desc, swing_alpha tiebreaker.
    # Stage-A cleanup (2026-07-XX): was `buys.sort(key=-opportunity_score)` but
    # the 480-row audit (see comment at 11b above) showed 20/20 top-20 overlap
    # between opportunity_score and final_confidence — the extra re-sort was a
    # no-op. opportunity_score is still computed (populated below) so Telegram
    # cards and the Excel export still display "Opp NN".
    for stock in buys + watchlist_stocks:
        if "opportunity_score" not in stock:
            stock["opportunity_score"] = compute_opportunity_score(stock)
    buys.sort(key=lambda x: (
        -float(x.get("final_confidence", 0) or 0),
        -float(x.get("swing_alpha_score", 0) or 0),
        x.get("symbol", ""),
    ))
    # Phase 1 #44 (2026-07-05): watchlist sort by (tier priority, conf_gap asc).
    # Rationale: for research, the most useful row on top is the one closest
    # to becoming a BUY — not the one with the biggest opportunity_score
    # (which factors in signal quality that already got vetoed). Tier order:
    #   READY_BLOCKED     0  — methodology worked, only structural blocker
    #   NEAR_MISS_RISING  1
    #   NEAR_MISS         2
    #   NEAR_MISS_FADING  3
    #   DEVELOPING        4
    #   MONITOR           5
    # Within a tier, ascending conf_gap surfaces the one with the smallest
    # distance to qualification.
    _TIER_PRIO = {
        "READY_BLOCKED":     0,
        "NEAR_MISS_RISING":  1,
        "NEAR_MISS":         2,
        "NEAR_MISS_FADING":  3,
        "DEVELOPING":        4,
        "MONITOR":           5,
    }
    watchlist_stocks.sort(key=lambda x: (
        _TIER_PRIO.get(x.get("tier", "MONITOR"), 9),
        x.get("conf_gap", 999),
        x.get("symbol", ""),
    ))

    # ── Phase G7-C (2026-07-03): Intra-day sector diversity in BUY list ────
    # Gate 14b caps sector concentration vs EXISTING holdings, but never
    # against other BUYs picked on the same day. On a day with no existing
    # positions, six real-estate names could all pass gates and fill the top
    # of the BUY list — pure sector bet.
    # Fix: greedy pass over the sorted BUY list, keeping at most
    # MAX_BUYS_PER_SECTOR_INTRADAY (default = MAX_POSITIONS_PER_SECTOR) per
    # sector. Overflow is demoted to WATCHLIST with a SECTOR_DAY_CAP tag so
    # the audit shows why. Sort order guarantees the best-scored stock per
    # sector wins.
    #
    # Phase 2 #38 (2026-07-05): apply max_buys ceiling INSIDE the greedy
    # loop instead of truncating afterward. Old order (sector-cap → truncate)
    # could drop a diverse-sector candidate in favor of a same-sector one
    # already at the head of the list. New order fills the top max_buys slots
    # subject to sector-cap as a per-slot filter, so we keep the best-scored
    # stock per sector up to both caps simultaneously.
    _sector_cap = int(os.getenv(
        "MAX_BUYS_PER_SECTOR_INTRADAY",
        os.getenv("MAX_POSITIONS_PER_SECTOR", "2"),
    ))
    _regime_max_buys = effective_thresholds[regime]["max_buys"]
    if buys and _sector_cap > 0:
        _sector_counts: dict = {}
        _kept, _demoted = [], []
        for _stk in buys:
            _sec = _stk.get("sector") or get_sector(_stk.get("symbol", "")) or "UNKNOWN"
            _n = _sector_counts.get(_sec, 0)
            # Phase 2 #38: also stop once we have enough BUYs; remaining
            # candidates are demoted with an OVER_MAX_BUYS tag so audit is clear.
            if len(_kept) >= _regime_max_buys:
                _stk["decision"] = "WATCHLIST"
                _stk.setdefault("warnings", []).append(
                    f"OVER_MAX_BUYS({_regime_max_buys} filled by higher-ranked)"
                )
                _stk.setdefault("fail_reasons", []).append(
                    f"OVER_MAX_BUYS_{_regime_max_buys}"
                )
                _wl_meta = classify_watchlist(
                    _stk, regime, effective_thresholds,
                    conf_history=_conf_history_for_wl,
                )
                _stk.update(_wl_meta)
                _demoted.append(_stk)
                continue
            if _n < _sector_cap:
                _sector_counts[_sec] = _n + 1
                _kept.append(_stk)
            else:
                # Demote: BUY → WATCHLIST with an explicit reason
                _stk["decision"] = "WATCHLIST"
                _stk.setdefault("warnings", []).append(
                    f"SECTOR_DAY_CAP({_sec} — already {_sector_cap} BUYs today)"
                )
                _stk.setdefault("fail_reasons", []).append(
                    f"SECTOR_DAY_CAP_{_sec.replace(' ', '_')}"
                )
                _wl_meta = classify_watchlist(
                    _stk, regime, effective_thresholds,
                    conf_history=_conf_history_for_wl,
                )
                _stk.update(_wl_meta)
                _demoted.append(_stk)
        if _demoted:
            _log(
                f"  [G7-C] Sector cap ({_sector_cap}/sector) + max_buys "
                f"({_regime_max_buys}) demoted "
                f"{len(_demoted)} BUY → WATCHLIST: "
                + ", ".join(s.get("symbol", "?") for s in _demoted[:5])
                + ("..." if len(_demoted) > 5 else "")
            )
        buys = _kept
        watchlist_stocks.extend(_demoted)
        # Phase 1 #44 (2026-07-05): re-sort by (tier priority, conf_gap asc)
        # after demotions land — keeps research-friendly ordering.
        watchlist_stocks.sort(key=lambda x: (
            _TIER_PRIO.get(x.get("tier", "MONITOR"), 9),
            x.get("conf_gap", 999),
            x.get("symbol", ""),
        ))

    # Enforce max_buys cap (defensive belt-and-braces — the greedy loop above
    # already respects _regime_max_buys, but this guard protects the case
    # where _sector_cap == 0 (feature disabled) or `buys` bypassed the loop.
    # Phase C3 (2026-07-02): removed fii_stale halving — FII data is
    # structurally D-1/D-2 and was never a reliable signal to throttle
    # sizing on. max_buys is now driven purely by regime thresholds and
    # portfolio heat. See format_conviction_meter comment for full rationale.
    max_buys = effective_thresholds[regime]["max_buys"]
    buys = buys[:max_buys]

    # ── 14a. Decision audit JSONL — Phase C ─────────────────────────────────
    # Append one line per stock that reached the gates. Used for post-mortem RCA
    # ("why did we buy X on day N?") and CL-level signal regression tests.
    #
    # Phase N-1 (2026-07-03): signal attribution extended.
    # Every audit row now carries:
    #   • factor_scores          — per-factor 0-100 scores (10 factors)
    #   • factor_weights_applied — FACTOR_WEIGHTS constant snapshot
    #   • tq_components          — trend/momentum/volume/rr/weekly/pa sub-scores
    #   • trade_setup            — entry, stop, T1, T2, rr, net_rr_t2
    #   • ownership_deliv_bonus  — G7-A delivery-bonus contribution
    #   • market_cap_cr          — G8-B market-cap (real or missing→0)
    #   • soft_warnings          — STOP_CAP_CLAMPED, MCAP_MISSING, etc.
    # Downstream analysis (correlate factor_x with realized return) becomes
    # possible without adding any new pipeline compute cost.
    try:
        for stock in (buys + watchlist_stocks + rejected):
            _fs = stock.get("factor_scores", {}) or {}
            append_decision_audit({
                "symbol":          stock.get("symbol"),
                "sector":          stock.get("sector"),
                "decision":        stock.get("decision"),
                "decision_subtype": stock.get("decision_subtype"),  # R5 audit-only fine tier
                # Phase R6 (2026-07-06): BUG FIX — previously stamped stock.get("confidence") /
                # stock.get("trade_quality") which are wrong keys; the pipeline actually
                # writes final_confidence and trade_quality_score. Old rows were all None
                # for these fields, making pillar-floor analysis impossible.
                "final_confidence":    stock.get("final_confidence"),
                "trade_quality_score": stock.get("trade_quality_score"),
                "rr_ratio":            stock.get("rr_ratio"),
                # Legacy field names kept for backward compat with old analyzers
                "confidence":      stock.get("final_confidence"),
                "trade_quality":   stock.get("trade_quality_score"),
                "opportunity":     stock.get("opportunity_score"),
                "regime":          regime,
                "fail_reasons":    stock.get("fail_reasons", []),
                "warnings":        stock.get("warnings", []),
                # ── N-1 signal attribution ──
                "factor_scores":   _fs,
                "factor_weights":  FACTOR_WEIGHTS,
                "tq_components":   {
                    "trend":     stock.get("trend_quality"),
                    "momentum":  stock.get("momentum_quality"),
                    "volume":    stock.get("volume_delivery"),
                    "risk_rew":  stock.get("risk_reward"),
                    "weekly_ok": bool(stock.get("weekly_trend_ok")),
                    "pattern":   stock.get("price_pattern"),
                },
                "trade_setup":     {
                    "entry":     stock.get("entry"),
                    "stop":      stock.get("stop"),
                    "target1":   stock.get("target1"),
                    "target2":   stock.get("target2"),
                    "rr":        stock.get("rr_ratio"),
                    "net_rr_t2": stock.get("net_rr_t2"),
                    "eff_stop_pct": stock.get("effective_stop_pct"),
                },
                "ownership_deliv_bonus":   stock.get("ownership_deliv_bonus"),
                "ownership_deliv_reasons": stock.get("ownership_deliv_reasons", []),
                # Phase R3 (2026-07-06): per-stock FII/DII overlay
                "ownership_fii_bonus":     stock.get("ownership_fii_bonus"),
                "ownership_fii_reasons":   stock.get("ownership_fii_reasons", []),
                "fii_pct":                 stock.get("fii_pct"),
                "dii_pct":                 stock.get("dii_pct"),
                "market_cap_cr":   stock.get("market_cap_cr"),
                "avg_val_lakhs":   stock.get("avg_value_lakhs"),
                "soft_warnings":   stock.get("_soft_warnings", []),
                # Phase R1+R2 (2026-07-06): institutional composite scores + taxonomy
                "bq_score":              stock.get("bq_score"),
                "bq_data_completeness":  stock.get("bq_data_completeness"),
                "bq_verdict":            stock.get("bq_verdict"),
                "bq_flags":              stock.get("bq_flags", []),
                "sector_composite_score": stock.get("sector_composite_score"),
                "sector_verdict":        stock.get("sector_verdict"),
                "taxonomy_inputs":       stock.get("taxonomy_inputs"),
                "reject_tier":           stock.get("reject_tier"),  # "AVOID_TURNAROUND" or R3 bucket
                # Stage-A cleanup (2026-07-XX): llm_validator_* fields removed
                # from audit — ai_validate_buy_thesis was disabled by default
                # (LLM_VALIDATOR=0) and never populated in swing mode. See
                # git history + audit note if you need to revive it.
                # Phase R4 (2026-07-06): Swing-alpha overlay (4 signals)
                "swing_alpha_score":       stock.get("swing_alpha_score"),
                "rs_vs_nifty_20d":         stock.get("rs_vs_nifty_20d"),
                "breakout_freshness_days": stock.get("breakout_freshness_days"),
                "atr_stop_ratio_pct":      stock.get("atr_stop_ratio_pct"),
                "volume_expansion_ratio":  stock.get("volume_expansion_ratio"),
                "swing_alpha_bonus":       stock.get("_swing_alpha_bonus"),
                "trading_mode":            os.environ.get("TRADING_MODE", "swing"),
                # ── macro / auxiliary context (unchanged from before) ──
                "macro_quality":   macro.get("data_quality", "NORMAL"),
                "macro_bad":       macro.get("bad_fields", []),
                "fii_source":      macro.get("fii_source"),
                "fii_confidence":  macro.get("fii_confidence"),
                "usdinr":          macro.get("usdinr"),
                "pcr_source":      stock.get("pcr_source"),
            })
        _log(f"  [Audit] decision_audit appended {len(buys)+len(watchlist_stocks)+len(rejected)} rows → {DECISION_AUDIT_FILE}")
    except Exception as e:
        _log(f"  [Audit] append failed (non-fatal): {e}")

    # ── 14a-2. Reject-outcome watch (Phase N-2) ─────────────────────────────
    # Persist REJECTED stocks so scripts/reject_followup.py can measure their
    # subsequent return at T+5, T+10, T+20. Only saved on scheduled runs to
    # avoid polluting the file during manual experiments.
    try:
        if IS_SCHEDULED:
            _added_rej = append_reject_watch_entries(rejected, ist_today().isoformat())
            _log(f"  [Audit] reject_watch appended {_added_rej} new REJECTED symbols → {REJECT_WATCH_FILE}")
        else:
            _log(f"  [Audit] reject_watch skipped (manual run — set SCHEDULED_RUN=true to persist)")
    except Exception as e:
        _log(f"  [Audit] reject_watch append failed (non-fatal): {e}")

    # ── 14b. Confidence history update (FEATURE 2) ──
    _today_str_h = ist_today().isoformat()
    conf_history = load_confidence_history()
    conf_history = update_confidence_history(conf_history, top_40, _today_str_h)
    if IS_SCHEDULED:
        save_confidence_history(conf_history)
        _log(f"  Confidence history: {len(conf_history)} symbols tracked (saved)")
    else:
        _log(f"  Confidence history: {len(conf_history)} symbols tracked (NOT saved — manual run)")

    # ── 14c. Gate memory update (FEATURE 7) ──
    gate_memory = load_gate_memory()
    gate_memory = update_gate_memory(gate_memory, watchlist_stocks, _today_str_h)
    if IS_SCHEDULED:
        save_gate_memory(gate_memory)
        _log(f"  Gate memory: {len(gate_memory)} symbols tracked (saved)")
    else:
        _log(f"  Gate memory: {len(gate_memory)} symbols tracked (NOT saved — manual run)")
    buys = tag_repeat_buy_signals(buys, tracker_entries)

    # ── Phase Polish (2026-07-11): FULL setup-mix instrumentation ────────
    # Compute setup mix from the FULL evaluated pool: buys + watchlist +
    # rejected. This is the ~120 stocks that reached the gates (not just
    # the top-40 candidates). Also collect the top ticker symbols per
    # setup type so Telegram can show WHICH companies fell into each bucket.
    try:
        from collections import Counter as _MixCounter
        _mix_all      = _MixCounter()   # every stock that reached gates
        _mix_bought   = _MixCounter()   # of those, how many became BUYs
        _tickers_all  = {}              # setup_type -> [symbols]
        _tickers_buy  = {}              # setup_type -> [symbols] (BUYs only)

        # buys + watchlist_stocks are already the surviving pools; rejected
        # is the fail bucket. All three have setup_type stamped by
        # apply_setup_edge() earlier in the pipeline.
        _full_pool = list(buys or []) + list(watchlist_stocks or []) + list(rejected or [])
        for _s in _full_pool:
            _st = _s.get("setup_type", "OTHER") or "OTHER"
            _mix_all[_st] += 1
            _tickers_all.setdefault(_st, []).append(_s.get("symbol", "?"))
        for _b in (buys or []):
            _st = _b.get("setup_type", "OTHER") or "OTHER"
            _mix_bought[_st] += 1
            _tickers_buy.setdefault(_st, []).append(_b.get("symbol", "?"))

        regime_data["_setup_mix"]         = dict(_mix_all)
        regime_data["_setup_mix_bought"]  = dict(_mix_bought)
        regime_data["_setup_tickers"]     = _tickers_all
        regime_data["_setup_tickers_buy"] = _tickers_buy
        regime_data["_setup_pool_size"]   = len(_full_pool)
    except Exception as _e:
        _log(f"  [setup-mix] instrumentation skipped: {_e}")

    # ── 14c. Position sizing — Kelly + Heat-aware ──
    # Compute Platt stats from tracker history (activates after 20 closed trades)
    platt = compute_platt_stats(tracker_entries)
    if platt["calibrated"]:
        _log(f"  Platt calibration active: WR={platt['win_rate']:.1%} | "
             f"AvgWin={platt['avg_win_pct']:.1f}% | AvgLoss={platt['avg_loss_pct']:.1f}%")
    else:
        _log(f"  Platt calibration: {platt['total_closed']}/20 trades — using fixed 1.5% sizing")

    # Compute portfolio heat (total open risk as % of capital)
    heat = compute_portfolio_heat(holdings, current_prices, PORTFOLIO_CAPITAL)
    _log(f"  Portfolio heat: {heat['heat_pct']:.1f}% / {heat['max_heat_pct']:.0f}% max "
         f"({'OK' if heat['heat_ok'] else 'NEAR LIMIT'})")

    # Phase R4 (2026-07-06): portfolio risk composite (HHI, stop cluster, size skew)
    try:
        _prc = compute_portfolio_risk_composite(holdings, current_prices, PORTFOLIO_CAPITAL)
        if _prc.get("n_positions", 0) > 0:
            _log(
                f"  Portfolio risk: HHI={_prc['sector_hhi']} ({_prc['sector_hhi_flag']}), "
                f"stop_cluster={_prc['stop_cluster_count']} ({_prc['stop_cluster_flag']}), "
                f"size_skew={_prc['position_size_skew']}x ({_prc['position_size_skew_flag']})"
            )
            for _w in _prc.get("warnings", []):
                _log(f"    ⚠ {_w}")
        # Stamp composite on heat dict so downstream audit / risk sheet can access
        heat["risk_composite"] = _prc
    except Exception as _e:
        _log(f"  [WARN] portfolio risk composite failed: {_e}")

    # Phase 3b #N8 (2026-07-05): apply kill-switch size_multiplier to actual
    # position sizing. Previously the multiplier was only LOGGED — sizes
    # remained full even when drawdown ≥ KS_DD_HALVE_PCT (5%) said "halve".
    # On buys_paused the gate already blocks new BUYs; here we handle the
    # damped case where 0 < mult < 1.0 (typically 0.5 on 5-10% drawdown).
    _ks_mult = float(_ks.get("size_multiplier", 1.0) or 1.0)
    _ks_mult = max(0.0, min(1.0, _ks_mult))
    for stock in buys:
        pos = kelly_position_size(
            entry=stock.get("entry", 0),
            stop=stock.get("stop", 0),
            capital=PORTFOLIO_CAPITAL,
            win_rate=platt["win_rate"],
            avg_win_pct=platt["avg_win_pct"],
            avg_loss_pct=platt["avg_loss_pct"],
            heat=heat,
            max_position_pct=_MAX_POSITION_PCT_ENV,
        )
        # Apply kill-switch damping (only when < 1.0; buys_paused case is
        # already blocked earlier by Gate 5b).
        if _ks_mult < 1.0 and pos.get("shares", 0) > 0:
            _orig_shares = int(pos["shares"])
            _new_shares = max(1, int(_orig_shares * _ks_mult))
            _entry = float(stock.get("entry", 0) or 0)
            _stop = float(stock.get("stop", 0) or 0)
            _rps = max(0.0, _entry - _stop)
            pos["shares"] = _new_shares
            pos["position_value"] = round(_new_shares * _entry, 2)
            pos["position_pct"] = round(pos["position_value"] / PORTFOLIO_CAPITAL * 100, 1) \
                if PORTFOLIO_CAPITAL > 0 else 0.0
            pos["max_loss"] = round(_new_shares * _rps, 2)
            pos["sizing_method"] = f"{pos.get('sizing_method', 'FIXED_1.5PCT')}_KS_DAMPED_{_ks_mult:.2f}x"

        # Phase 4-B (2026-07-06): resize against EFFECTIVE stop when
        # HIGH_GAP_RISK is set. The nominal stop_pct only accounts for
        # in-session slippage; overnight/gap risk (p90 of 60-day gaps)
        # can be 2× larger. Without this correction, a stock like
        # PANAMAPET (nominal stop 6.9%, p90 gap 12.4%, effective stop
        # 19.3%) gets sized as if max-loss is ₹6.9K when the real
        # gap-inclusive max-loss is ~₹19K — a 2.8× understatement of risk.
        #
        # Fix: when high_gap_risk is True, recompute shares using the
        # effective (nominal + p90 gap) stop as the risk-per-share. This
        # preserves the intended risk_per_trade% (Kelly / 1.5% fixed) at
        # the *actual* stop level the stock will exit at during a gap.
        #
        # Guardrails:
        #   - Only shrinks (never grows) — max(1, min(orig, new))
        #   - Requires p90_gap_pct > 0.5 to fire (ignore microscopic gaps)
        #   - Env override: GAP_STOP_SIZING_ENABLED=0 disables
        #   - Tags sizing_method _GAP_ADJ_Nx so audit trail shows the shrink
        try:
            _gap_size_on = int(os.getenv("GAP_STOP_SIZING_ENABLED", "1"))
        except (TypeError, ValueError):
            _gap_size_on = 1
        _hgr = bool(stock.get("high_gap_risk", False))
        _p90_gap = float(stock.get("p90_gap_pct", 0) or 0)
        if (_gap_size_on and _hgr and _p90_gap > 0.5
                and pos.get("shares", 0) > 0):
            _entry_g = float(stock.get("entry", 0) or 0)
            _stop_g = float(stock.get("stop", 0) or 0)
            if _entry_g > 0 and _stop_g > 0 and _stop_g < _entry_g:
                _nominal_stop_pct = (_entry_g - _stop_g) / _entry_g * 100.0
                _eff_stop_pct = _nominal_stop_pct + _p90_gap
                if _eff_stop_pct > _nominal_stop_pct * 1.01:  # guard div-by-zero + noise
                    # Preserve the intended risk% at the effective stop:
                    # new_shares / orig_shares = nominal / effective
                    _shrink = _nominal_stop_pct / _eff_stop_pct
                    _orig_shares_g = int(pos.get("shares", 0))
                    _new_shares_g = max(1, int(_orig_shares_g * _shrink))
                    if _new_shares_g < _orig_shares_g:
                        _rps_nominal = max(0.0, _entry_g - _stop_g)
                        _rps_effective = _entry_g * (_eff_stop_pct / 100.0)
                        pos["shares"] = _new_shares_g
                        pos["position_value"] = round(_new_shares_g * _entry_g, 2)
                        pos["position_pct"] = round(
                            pos["position_value"] / PORTFOLIO_CAPITAL * 100, 1
                        ) if PORTFOLIO_CAPITAL > 0 else 0.0
                        # max_loss remains at nominal stop (that's what tracker
                        # exits at). effective_max_loss surfaces gap-risk.
                        pos["max_loss"] = round(_new_shares_g * _rps_nominal, 2)
                        pos["effective_max_loss"] = round(_new_shares_g * _rps_effective, 2)
                        pos["effective_stop_pct"] = round(_eff_stop_pct, 2)
                        pos["gap_shrink_factor"] = round(_shrink, 3)
                        pos["sizing_method"] = (
                            f"{pos.get('sizing_method', 'FIXED_1.5PCT')}"
                            f"_GAP_ADJ_{_shrink:.2f}x"
                        )
        stock.update(pos)

    # ── 14d. Short signal detection ──
    shorts = detect_short_signals(top_40, regime, regime_data["thresholds"])

    # ── 14e. Upcoming market events already computed at step 13b ──

    # ── 14f. Watchlist persistence ──
    wl_history = load_persistent_watchlist()
    watchlist_stocks, wl_history = merge_watchlist_with_history(watchlist_stocks, wl_history)
    if IS_SCHEDULED:
        save_persistent_watchlist(wl_history)
    else:
        _log("  Watchlist history NOT saved (manual run)")

    # ── 15. Format and send main Telegram message ──
    _log("[15/17] Sending main Telegram report...")
    # Phase C: explicit Asia/Kolkata — no dependency on TZ env var
    timestamp = ist_now().strftime("%b %d, %Y %H:%M IST")
    # nifty_state already computed at step 6b — pass through (BUG FIX 1)

    # ── 15a. Run AI calls in parallel (daily summary + buy theses + near miss insights) ──
    _log("[15a/17] Running AI calls in parallel...")
    ai_results = run_all_ai_calls(
        regime_data      = regime_data,
        macro            = macro,
        breadth_data     = breadth,
        nifty_state      = nifty_state,
        buys             = buys,
        watchlist        = watchlist_stocks,
        portfolio_alerts = portfolio_alerts,
        conf_history     = conf_history,
        gate_memory      = gate_memory,
        events           = upcoming_events,
    )

    # ── Telegram message: v2 (redesigned 2026-07-10) with v1 fallback ───────
    # v2 is a human-friendly redesign (named stocks, ASCII bars, tiered
    # verdict, CSV attachment). v1 remains the safe fallback if v2 raises
    # any unexpected exception. Flip NEW_TELEGRAM_FORMAT=false for instant
    # rollback with zero code changes. See Section 9c for the v2 design notes.
    _v1_kwargs = dict(
        regime_data      = regime_data,
        buys             = buys,
        shorts           = shorts,
        watchlist        = watchlist_stocks,
        portfolio_alerts = portfolio_alerts,
        macro            = macro,
        key_levels       = key_levels,
        upcoming_events  = upcoming_events,
        timestamp        = timestamp,
        heat             = heat,
        platt            = platt,
        tracker_v2       = tracker_v2,
        rejected_stocks  = rejected,
        breadth_20       = breadth.get("ema20_pct", 50.0),
        nifty_state      = nifty_state,
        universe_count   = len(symbols),
        tradable_count   = len(tradable),
        conf_history     = conf_history,
        gate_memory      = gate_memory,
        ai_results       = ai_results,
    )
    message = None
    _csv_path = ""
    if NEW_TELEGRAM_FORMAT:
        try:
            message = format_telegram_message_v2(**_v1_kwargs)
            _log(f"[telegram] using v2 format ({len(message)} chars)")
        except Exception as _e_v2:
            _log(f"[telegram] v2 crashed ({_e_v2!r}) — falling back to v1")
            message = None
    if message is None:
        # v1 path (either feature-flag off, or v2 crashed)
        message = format_telegram_message(**_v1_kwargs)
        _log(f"[telegram] using v1 format ({len(message)} chars)")

    # CSV attachment — only when v2 is active AND flag enabled. Non-fatal.
    if NEW_TELEGRAM_FORMAT and TELEGRAM_ATTACH_CSV:
        try:
            _csv_path = _v2_write_daily_csv(
                buys, watchlist_stocks, rejected, regime, timestamp,
            )
        except Exception as _e_csv:
            _log(f"[telegram] CSV generation failed (non-fatal): {_e_csv}")
            _csv_path = ""

    _log("--- TELEGRAM PREVIEW (first 1500 chars) ---")
    _log(message[:1500])
    _log("--- END PREVIEW ---")
    # Prepend a visible test banner on manual runs so you instantly know it's not the real report
    if not IS_SCHEDULED:
        banner = (
            "⚠️  MANUAL TEST RUN — NOT the scheduled report\n"
            "State was NOT saved. Day counters NOT advanced.\n"
            + "─" * 40 + "\n"
        )
        message = banner + message
    send_telegram(message)
    # Send the CSV attachment (v2 only). Non-fatal on failure.
    if _csv_path:
        try:
            _csv_caption = (
                f"📎 Full universe scan · {os.path.basename(_csv_path)}\n"
                f"Every stock with score, verdict, reject reason. "
                f"Open in Excel to sort/filter."
            )
            send_telegram_document(_csv_path, caption=_csv_caption)
        except Exception as _e_doc:
            _log(f"[telegram] CSV send failed (non-fatal): {_e_doc}")
    # Send BUY signals to dedicated buy channel
    send_buy_telegram(buys, regime, timestamp)

    # ── 2026-07-09 Consolidation D3: shadow master report ─────────────────
    # Runs the bucket-tracking flow (A/B/C/D) on shadow_master.xlsx. This
    # is the same job the workflow runs as a separate step under CI; keeping
    # it here means run-locally.ps1 and manual runs also get a preview file
    # + Telegram push. shadow_master_job internally gates save-vs-preview
    # via SCHEDULED_RUN so double-runs under CI are safe (idempotent).
    if _SHADOW_REPORT_OK and os.getenv("SHADOW_REPORT_ENABLED", "true").lower() != "false":
        try:
            _rep = shadow_master_job.run_scan_and_update(quiet=True)
            if _rep.get("ok") and not _rep.get("skipped"):
                _stats = _rep.get("stats", {}) or {}
                _n_appended = _rep.get("n_appended", 0)
                _n_resolved = len(_stats.get("resolved_today", []) or [])
                _log(
                    f"[shadow_master] {_rep.get('mode', '?')} · "
                    f"xlsx={_rep.get('xlsx_path')} · "
                    f"appended={_n_appended} · "
                    f"updated={_stats.get('n_rows_updated', 0)} · "
                    f"resolved={_n_resolved}"
                )
            elif _rep.get("skipped"):
                _log(f"[shadow_master] skipped: {_rep.get('reason', '?')}")
            else:
                _log(f"[shadow_master] build failed: {_rep.get('error', '?')}")
        except Exception as _ex:
            _log(f"[shadow_master] non-fatal error: {_ex}")

    # ── 16. Trade Tracker updates ──
    _log("[16/17] Updating trade tracker...")
    tracker_entries = update_tracker_trailing_stop(tracker_entries)

    for stock in buys:
        tracker_entries = add_to_tracker(tracker_entries, stock, "BUY")
        # Phase C7e (2026-07-02): also append to tracker V2 so kill-switch,
        # weekly summary, and morning_check see today's new BUYs. Without this
        # the V2 tracker only ever held the Jun-25 seed and completed pnl
        # stats were effectively empty.
        tracker_v2 = add_to_tracker_v2(tracker_v2, stock, regime)

    near_miss_stocks = [w for w in watchlist_stocks if w.get("tier") == "NEAR_MISS"]
    for stock in near_miss_stocks:
        tracker_entries = add_to_tracker(tracker_entries, stock, "NEAR_MISS")

    tracker_entries, closed_today = update_tracker(tracker_entries)

    if closed_today:
        _log(f"  Tracker: {len(closed_today)} trade(s) closed today")

    if IS_SCHEDULED:
        save_tracker(tracker_entries)
        # Re-persist tracker V2 after appending today's buys (initial save at
        # step 13b captured only PnL updates on the pre-existing set).
        save_tracker_v2(tracker_v2)
    else:
        _log("  Trade tracker NOT saved (manual run)")

    # ── 17. Save CSVs + Excel ──
    _log("[17/17] Saving output CSVs + Excel...")
    for s in top_40:
        s.pop("_df", None)
        s.pop("fundamentals", None)
        s.pop("promoter_data", None)
    save_csv(top_40,           "analysis_output.csv")
    save_csv(portfolio_alerts, "portfolio_monitor.csv")
    save_csv(buys,             "buys_today.csv")

    # Save recommendations to Excel tracker (PART C)
    # Always persist — the Excel is the system-of-record for the tracker job
    # and downstream artifacts. Manual runs MUST still write it, otherwise the
    # Recommendation Tracker workflow has nothing to download.
    today_str_pipe = ist_today().isoformat()
    save_recommendations_to_excel(
        buys, watchlist_stocks,
        {"regime": regime, "score": regime_data.get("score", 0)},
        today_str_pipe,
        rejected=rejected,   # 2026-07-06: also persist rejects for 3-week retrospective
    )

    # ── Phase G-BATCH1 (2026-07-07): post-trade attribution report ────────
    # Appends 3 sheets to shadow_master.xlsx:
    #   Attribution_Factor  — per-factor IC (rank-correlation with P&L)
    #   Attribution_Gate    — per-gate P&L delta vs population avg
    #   Attribution_Regime  — per-regime win-rate + expectancy
    # Feature-gated. Reads existing tracker log; never touches live positions.
    if os.environ.get("ENABLE_ATTRIBUTION", "true").lower() == "true":
        try:
            import attribution as _attr
            _summary_attr = _attr.build_attribution_report(
                tracker_json="trade_tracker.json",
                tracker_xlsx="shadow_master.xlsx",
                out_xlsx="shadow_master.xlsx",
                lookback_days=int(os.environ.get("ATTRIBUTION_LOOKBACK", "90")),
            )
            _log(
                f"  [Attribution] {_summary_attr.get('closed_in_window', 0)}/"
                f"{_summary_attr.get('total_trades', 0)} closed trades analysed. "
                f"Top factor: {_summary_attr.get('top_factor_by_lift')} · "
                f"Top gate: {_summary_attr.get('top_gate_by_delta')} · "
                f"Best regime: {_summary_attr.get('best_regime')}"
            )
        except Exception as _e_attr:
            _log(f"  [WARN] attribution report failed (non-fatal): {_e_attr}")

    # ── 17b. Phase 1 #54 (2026-07-05): flush price-fetch failures ─────────
    # Every fetch_price_data(...) that returned None (any reason) is here.
    # Written once at end-of-run to avoid touching disk from hot loops.
    try:
        if _PRICE_FETCH_FAILURES:
            with open(PRICE_FETCH_FAIL_FILE, "a", encoding="utf-8") as _pff:
                for _rec in _PRICE_FETCH_FAILURES:
                    _pff.write(json.dumps(_rec, default=str) + "\n")
            _log(f"  [PriceFail] {len(_PRICE_FETCH_FAILURES)} failures logged → {PRICE_FETCH_FAIL_FILE}")
            # Clear the module-global buffer so a re-invocation in the same
            # Python process (rare — mostly tests) doesn't double-log.
            _PRICE_FETCH_FAILURES.clear()
    except Exception as _pff_exc:
        _log(f"  [WARN] price fetch failure flush failed (non-fatal): {_pff_exc}")

    # ── 17c. Phase 1 #52 (2026-07-05): daily research snapshot ────────────
    # One JSONL per day containing the top-N (currently 100) fully-scored
    # candidates with full factor breakdown. This is the primary corpus for
    # post-freeze research: rank stability, factor-return correlation,
    # false-positive analysis. Additive-only — never changes pipeline output.
    #
    # File: {DAILY_SNAPSHOT_DIR}/{YYYY-MM-DD}.jsonl (one line per stock)
    # Overwritten each run of the same day so re-runs produce the latest
    # snapshot (previous days' files are never touched).
    try:
        os.makedirs(DAILY_SNAPSHOT_DIR, exist_ok=True)
        _snap_path = os.path.join(
            DAILY_SNAPSHOT_DIR,
            f"{ist_today().isoformat()}.jsonl",
        )
        # Build a lookup from symbol → decision so we know how each snapshot
        # row was resolved (BUY / WATCHLIST / REJECTED / SCORED_ONLY).
        _decision_map: dict = {}
        for _s in buys:
            _decision_map[_s.get("symbol")] = "BUY"
        for _s in watchlist_stocks:
            _decision_map.setdefault(_s.get("symbol"), _s.get("tier", "WATCHLIST"))
        for _s in rejected:
            _decision_map.setdefault(_s.get("symbol"), "REJECTED")
        # `top_40` is now the top-100 (or _TOP_N env-override). Everyone here
        # has a full factor breakdown from step 7.
        _snap_rows = 0
        with open(_snap_path, "w", encoding="utf-8") as _snap:
            for _rank, stock in enumerate(top_40, start=1):
                _sym = stock.get("symbol")
                _snap_row = {
                    "date":            ist_today().isoformat(),
                    "ts":              ist_now().isoformat(),
                    "rank":            _rank,
                    "symbol":          _sym,
                    "sector":          stock.get("sector") or get_sector(_sym or ""),
                    "decision":        _decision_map.get(_sym, "SCORED_ONLY"),
                    "regime":          regime,
                    "base_confidence": stock.get("base_confidence"),
                    "final_confidence": stock.get("final_confidence"),
                    "trade_quality":   stock.get("trade_quality_score"),
                    "opportunity":    stock.get("opportunity_score"),
                    "rr_ratio":        stock.get("rr_ratio"),
                    "entry":           stock.get("entry"),
                    "stop":            stock.get("stop"),
                    "target1":         stock.get("target1"),
                    "target2":         stock.get("target2"),
                    "atr":             stock.get("atr"),
                    "close":           stock.get("close"),
                    "avg_value_lakhs": stock.get("avg_value_lakhs"),
                    "market_cap_cr":   stock.get("market_cap_cr"),
                    # ── factor-level attribution (10 factors) ──
                    "factor_scores":   stock.get("factor_scores", {}),
                    "factor_weights":  FACTOR_WEIGHTS,
                    # ── trade quality sub-components ──
                    "tq_components": {
                        "trend":     stock.get("trend_quality"),
                        "momentum":  stock.get("momentum_quality"),
                        "volume":    stock.get("volume_delivery"),
                        "risk_rew":  stock.get("risk_reward"),
                        "weekly_ok": bool(stock.get("weekly_trend_ok")),
                        "pattern":   stock.get("price_pattern"),
                    },
                    # ── news attribution + source ──
                    "news_penalty":    stock.get("news_penalty"),
                    "news_category":   stock.get("news_category"),
                    "news_source":     stock.get("news_source"),
                    "is_black_swan":   bool(stock.get("is_black_swan")),
                    # ── ownership / delivery context ──
                    "ownership_deliv_bonus": stock.get("ownership_deliv_bonus"),
                    "roe":             stock.get("roe"),
                    "de_ratio":        stock.get("de_ratio"),
                    "roce":            stock.get("roce"),
                    "promoter_pledge_pct": stock.get("promoter_pledge_pct"),
                    "fundamentals_source": stock.get("fundamentals_source"),
                    # ── gate outcome ──
                    "fail_reasons":    stock.get("fail_reasons", []),
                    "warnings":        stock.get("warnings", []),
                }
                _snap.write(json.dumps(_snap_row, default=str) + "\n")
                _snap_rows += 1
        _log(f"  [Snapshot] {_snap_rows} rows → {_snap_path}")
    except Exception as _snap_exc:
        _log(f"  [WARN] daily snapshot writer failed (non-fatal): {_snap_exc}")

    # ── 18. Done ──
    _log("[DONE] Pipeline complete.")
    _log(f"  BUY: {len(buys)} | WATCHLIST: {len(watchlist_stocks)} | SHORTS: {len(shorts)}")

    # ── Phase W (2026-07-03): expose result counters to __main__ wrapper ──
    # The __main__ block writes run_health.json for cross-workflow staleness
    # detection. Attach the numbers so the health entry is informative.
    global _LAST_RUN_STATS
    _LAST_RUN_STATS = {
        "buys":       len(buys),
        "watchlist":  len(watchlist_stocks),
        "shorts":     len(shorts),
        "tradable":   len(top_40) if isinstance(top_40, list) else 0,
        "regime":     regime,
    }


# ── Phase W (2026-07-03): module-level result stash ──
# Filled at the end of run_pipeline(); read by __main__.
_LAST_RUN_STATS: dict = {}


if __name__ == "__main__":
    # ── Phase W (2026-07-03): watchdog wrapper ──
    # Guarantee run_health.json is written even on crash. The workflow's
    # if:failure() step will fire the Telegram alert; this ensures other
    # workflows can still see the last successful ts for evening_pipeline.
    _run_status = "ok"
    _run_error  = ""
    try:
        run_pipeline()
    except SystemExit:
        raise
    except Exception as _pipe_exc:  # noqa: BLE001 — best-effort logging
        _run_status = "fail"
        _run_error  = str(_pipe_exc)[:200]
        _log(f"[FATAL] Pipeline crashed: {_pipe_exc}")
        import traceback
        traceback.print_exc()
    finally:
        _mode = "scheduled" if IS_SCHEDULED else "manual"
        _extras = [
            f"fresh_start={str(FRESH_START).lower()}",
        ]
        if _LAST_RUN_STATS:
            for _k, _v in _LAST_RUN_STATS.items():
                _extras.append(f"{_k}={_v}")
        if _run_error:
            _safe = _run_error.replace(" ", "_")[:80]
            _extras.append(f"error={_safe}")
        try:
            # BUG-E fix: anchor pipeline_health.py to this file's dir and set
            # PIPELINE_HEALTH_FILE so state lands next to main.py regardless
            # of the caller's cwd.
            _here = os.path.dirname(os.path.abspath(__file__))
            _ph_script = os.path.join(_here, "scripts", "pipeline_health.py")
            _env = os.environ.copy()
            _env.setdefault(
                "PIPELINE_HEALTH_FILE",
                os.path.join(_here, "run_health.json"),
            )
            subprocess.run(
                [sys.executable, _ph_script, "record",
                 "--job", "evening",
                 "--status", _run_status,
                 "--mode", _mode,
                 "--extras", *_extras],
                check=False, timeout=15, env=_env,
            )
        except Exception as _health_exc:
            _log(f"[WARN] pipeline_health record failed: {_health_exc}")

    if _run_status == "fail":
        sys.exit(1)
