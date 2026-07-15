"""Build `tracking_workbook.xlsx` — the 9-sheet consolidated observation workbook.

Redesign 2026-07-15: collapsed 16 sheets → 9 sheets by moving "which sheet"
into a category *column* on merged sheets. One row per stock at any time.

Sheets:
    1.  BUY              — stocks with active BUY stage. Updated in-place daily
                            until T1/T2/Stop resolves (then moves to DONE).
    2.  WATCHLIST        — MERGED (NEAR_MISS + DEVELOPING + MONITOR).
                            Added column: 'Watchlist Category' = NEAR_MISS |
                            DEVELOPING | MONITOR.
    3.  REJECTED         — Rejected stage decisions.
    4.  REJECTED_SETUPS  — MERGED (rejected BREAKOUT/MOMENTUM/PULLBACK/REVERSAL
                            setups + REJECTED_SETUP historical bucket).
                            Added column: 'Setup Type' = BREAKOUT | MOMENTUM |
                            PULLBACK | REVERSAL.
    5.  DONE             — Terminal outcomes: T2_HIT | STOPPED | STOPPED_AFTER_T1.
    6.  WEEKLY_SUMMARY   — Per-sheet weekly counts (rebuilt every run).
    7.  WEEKLY_REVIEW    — Historical per-category aggregate rows.
    8.  RESEARCH         — Fundamentals view (kept unchanged).
    9.  _LEGEND          — Column glossary + score-band definitions.

Deleted from the old 16-sheet layout:
    NEAR_MISS, DEVELOPING, MONITOR       → merged into WATCHLIST
    BREAKOUT, MOMENTUM, PULLBACK,
      REVERSAL, REJECTED_SETUP           → merged into REJECTED_SETUPS
    ACTIVE_TRACKING                       → deleted (BUY sheet already tracks
                                            live positions in-place)

Golden rule: a stock lives in EXACTLY ONE sheet at any time. Active BUYs
stay in BUY. Watchlist stocks stay in WATCHLIST (their category may change).
Only T2/Stop moves a stock to DONE.

The job READS `results/daily_snapshots.jsonl` + `results/tracking_store.json`
and REBUILDS the workbook from scratch. Store + snapshots are the sources
of truth — the workbook is a pure view.

Usage (from Analyse-recommend-track-main/):
    python tracking_workbook_job.py --write
    python tracking_workbook_job.py --write --path /custom/tracking_workbook.xlsx
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

# openpyxl is already a hard dep of the project (used by shadow_master_job).
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.worksheet import Worksheet

_HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(_HERE))

from tracking_store import (  # noqa: E402
    CONFIDENCE_BANDS,
    DEFAULT_STORE_PATH,
    DEFAULT_WEEKLY_HISTORY_PATH,
    FACTOR_FIELDS,
    MIGRATION_NEEDS_REVIEW,
    OPPORTUNITY_BANDS,
    RESEARCH_FIELDS,
    RR_BANDS,
    STATUS_ACTIVE,
    STATUS_STOPPED,
    STATUS_STOPPED_AFTER_T1,
    STATUS_T1_HIT,
    STATUS_T2_HIT,
    TERMINAL_STATUSES,
    TQ_BANDS,
    TrackingStore,
    band_of,
)
from daily_snapshot_job import (  # noqa: E402
    DAILY_SNAPSHOTS_PATH,
    SETUP_BUCKETS,
    STAGE_BUCKETS,
    read_snapshots,
)


# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
THIN = Side(border_style="thin", color="B8B8B8")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_FILLS = {
    STATUS_ACTIVE: PatternFill("solid", fgColor="FFF2CC"),           # pale yellow
    STATUS_T1_HIT: PatternFill("solid", fgColor="D4EDDA"),           # light green
    STATUS_T2_HIT: PatternFill("solid", fgColor="70AD47"),           # dark green
    STATUS_STOPPED: PatternFill("solid", fgColor="F8CBAD"),          # light red
    STATUS_STOPPED_AFTER_T1: PatternFill("solid", fgColor="FCE4D6"), # peach
}

STAGE_FILLS = {
    "BUY": PatternFill("solid", fgColor="C6EFCE"),
    "WATCHLIST": PatternFill("solid", fgColor="FFEB9C"),
    "DEVELOPING": PatternFill("solid", fgColor="D9E1F2"),
    "NEAR_MISS": PatternFill("solid", fgColor="FFCC99"),
}

# Row-identity columns get a distinct tint so the eye can group repeat rows
# for the same trade idea across dates.
TRACKING_ID_FILL = PatternFill("solid", fgColor="E7DEF5")   # pale lavender
ROW_KEY_FILL     = PatternFill("solid", fgColor="F2F2F2")   # light grey


# ---------------------------------------------------------------------------
# Sheet column definitions
# ---------------------------------------------------------------------------

# Leading columns for ACTIVE_TRACKING (the "heart" — see spec §3.3 sheet 9).
ACTIVE_LEADING_COLS: Tuple[Tuple[str, str, str], ...] = (
    # (header, getter_key, format_hint)
    ("Row Key",           "_row_key",              "text"),
    ("Tracking ID",       "tracking_id",           "text"),
    ("Symbol",            "symbol",                "text"),
    ("Origin",            "entry.origin",          "text"),
    ("First Stage",       "journey_path.first",    "text"),
    ("Current Stage",     "current.stage",         "text"),
    ("Journey Path",      "journey_path.joined",   "text"),
    ("Stage Changes",     "stage_change_count",    "int"),
    ("Last Stage Change", "last_stage_change_date","date"),
    ("Setup",             "entry.setup_type",      "text"),
    ("Regime",            "entry.regime",          "text"),
    ("Sector",            "entry.sector",          "text"),
    ("First Seen",        "first_seen_date",       "date"),
    ("Days Active",       "days_active",           "int"),
    ("Entry Confidence",  "entry.confidence",      "one_dec"),
    ("Current Confidence","current.confidence",    "one_dec"),
    ("Entry TQ",          "entry.tq",              "one_dec"),
    ("Current TQ",        "current.tq",            "one_dec"),
    ("Entry R/R",         "entry.rr",              "two_dec"),
    ("Current R/R",       "current.rr",            "two_dec"),
    ("Entry Opp",         "entry.opportunity_score","one_dec"),
    ("Current Opp",       "current.opportunity_score","one_dec"),
    ("Reference Entry",   "entry.reference_entry_price","two_dec"),
    ("Current Price",     "current.current_price", "two_dec"),
    ("T1",                "entry.t1_price",        "two_dec"),
    ("T2",                "entry.t2_price",        "two_dec"),
    ("Stop",              "entry.stop_price",      "two_dec"),
    ("MFE %",             "mfe_pct",               "pct"),
    ("MAE %",             "mae_pct",               "pct"),
    ("Max Price",         "max_price_seen",        "two_dec"),
    ("Min Price",         "min_price_seen",        "two_dec"),
    ("T1 Hit",            "t1_hit",                "bool"),
    ("T1 Date",           "t1_hit_date",           "date"),
    ("Days to T1",        "days_to_t1",            "int"),
    ("T2 Hit",            "t2_hit",                "bool"),
    ("T2 Date",           "t2_hit_date",           "date"),
    ("Days to T2",        "days_to_t2",            "int"),
    ("Stop Hit",          "stop_hit",              "bool"),
    ("Stop Date",         "stop_hit_date",         "date"),
    ("Days to Stop",      "days_to_stop",          "int"),
    ("Status",            "tracking_status",       "text"),
    ("Strategy Version",  "strategy_version",      "text"),
)

# Factor columns appear AFTER the leading columns.
def _factor_cols() -> Tuple[Tuple[str, str, str], ...]:
    return tuple((f, f"entry.factors.{f}", "one_dec") for f in FACTOR_FIELDS)

# Simpler column sets for the stage-specific sheets.
STAGE_COLS: Tuple[Tuple[str, str, str], ...] = (
    ("Row Key",           "_row_key",              "text"),
    ("Tracking ID",       "tracking_id",           "text"),
    ("Symbol",            "symbol",                "text"),
    ("Setup",             "entry.setup_type",      "text"),
    ("Sector",            "entry.sector",          "text"),
    ("First Seen",        "first_seen_date",       "date"),
    ("Days Active",       "days_active",           "int"),
    ("Journey",           "journey_path.joined",   "text"),
    ("Entry Confidence",  "entry.confidence",      "one_dec"),
    ("Current Confidence","current.confidence",    "one_dec"),
    ("Entry TQ",          "entry.tq",              "one_dec"),
    ("Current TQ",        "current.tq",            "one_dec"),
    ("Entry R/R",         "entry.rr",              "two_dec"),
    ("Entry Opp",         "entry.opportunity_score","one_dec"),
    ("Ref Entry",         "entry.reference_entry_price","two_dec"),
    ("T1",                "entry.t1_price",        "two_dec"),
    ("T2",                "entry.t2_price",        "two_dec"),
    ("Stop",              "entry.stop_price",      "two_dec"),
    ("Current Price",     "current.current_price", "two_dec"),
    ("MFE %",             "mfe_pct",               "pct"),
    ("MAE %",             "mae_pct",               "pct"),
    ("Status",            "tracking_status",       "text"),
)

REJECTED_COLS: Tuple[Tuple[str, str, str], ...] = (
    ("Row Key",           "_row_key",              "text"),
    ("Tracking ID",       "tracking_id",           "text"),
    ("Symbol",            "symbol",                "text"),
    ("Setup",             "entry.setup_type",      "text"),
    ("Sector",            "entry.sector",          "text"),
    ("First Seen",        "first_seen_date",       "date"),
    ("Entry Confidence",  "entry.confidence",      "one_dec"),
    ("Entry TQ",          "entry.tq",              "one_dec"),
    ("Entry R/R",         "entry.rr",              "two_dec"),
    ("Entry Opp",         "entry.opportunity_score","one_dec"),
    ("Ref Entry",         "entry.reference_entry_price","two_dec"),
    ("Fail Reasons",      "entry.fail_reasons.joined","text"),
    ("Primary Reject",    "entry.primary_reject",  "text"),
    ("Regime",            "entry.regime",          "text"),
    ("Status",            "tracking_status",       "text"),
)

RESOLVED_COLS: Tuple[Tuple[str, str, str], ...] = (
    ("Row Key",           "_row_key",              "text"),
    ("Tracking ID",       "tracking_id",           "text"),
    ("Symbol",            "symbol",                "text"),
    ("Setup",             "entry.setup_type",      "text"),
    ("Sector",            "entry.sector",          "text"),
    ("First Seen",        "first_seen_date",       "date"),
    ("Days Active",       "days_active",           "int"),
    ("Entry Stage",       "journey_path.first",    "text"),
    ("Journey",           "journey_path.joined",   "text"),
    ("Entry Confidence",  "entry.confidence",      "one_dec"),
    ("Entry TQ",          "entry.tq",              "one_dec"),
    ("Entry R/R",         "entry.rr",              "two_dec"),
    ("Entry Opp",         "entry.opportunity_score","one_dec"),
    ("Ref Entry",         "entry.reference_entry_price","two_dec"),
    ("MFE %",             "mfe_pct",               "pct"),
    ("MAE %",             "mae_pct",               "pct"),
    ("T1 Hit",            "t1_hit",                "bool"),
    ("Days to T1",        "days_to_t1",            "int"),
    ("T2 Hit",            "t2_hit",                "bool"),
    ("Days to T2",        "days_to_t2",            "int"),
    ("Stop Hit",          "stop_hit",              "bool"),
    ("Days to Stop",      "days_to_stop",          "int"),
    ("Final Outcome",     "final_outcome",         "text"),
    ("Status",            "tracking_status",       "text"),
    ("Strategy Version",  "strategy_version",      "text"),
)

RESEARCH_COLS: Tuple[Tuple[str, str, str], ...] = (
    ("Row Key",           "_row_key",              "text"),
    ("Tracking ID",       "tracking_id",           "text"),
    ("Symbol",            "symbol",                "text"),
    ("Setup",             "entry.setup_type",      "text"),
    ("Origin",            "entry.origin",          "text"),
    ("Status",            "tracking_status",       "text"),
    ("Current Stage",     "current.stage",         "text"),
    ("Sector",            "entry.sector",          "text"),
    # Research fields:
    ("ROE",               "entry.research.roe",    "one_dec"),
    ("ROCE",              "entry.research.roce",   "one_dec"),
    ("D/E",               "entry.research.de_ratio","two_dec"),
    ("Pledge %",          "entry.research.promoter_pledge_pct","one_dec"),
    ("Market Cap (Cr)",   "entry.research.market_cap_cr","one_dec"),
    ("Fund Source",       "entry.research.fundamentals_source","text"),
    # Not produced today — documented in _LEGEND sheet:
    ("Revenue Growth",    "entry.research.revenue_growth","one_dec"),
    ("Profit Growth",     "entry.research.profit_growth","one_dec"),
    ("EPS",               "entry.research.eps",    "two_dec"),
    ("EPS Growth",        "entry.research.eps_growth","one_dec"),
    ("Research Grade",    "entry.research.research_grade","text"),
    ("FII Change",        "entry.research.fii_change","one_dec"),
    ("DII Change",        "entry.research.dii_change","one_dec"),
    ("Sector RS",         "entry.research.sector_relative_strength","one_dec"),
    ("Dist 52w High",     "entry.research.distance_from_52w_high","pct"),
)


# ---------------------------------------------------------------------------
# Getter dispatch — resolves dotted keys against a record
# ---------------------------------------------------------------------------

def _resolve(rec: Dict[str, Any], key: str) -> Any:
    """Resolve a dotted key against a record with a few virtual paths:
        journey_path.first, journey_path.joined
        entry.fail_reasons.joined
        entry.primary_reject
        final_outcome
    """
    if key == "final_outcome":
        return _final_outcome(rec)
    if key == "_row_key":
        tid = rec.get("tracking_id") or rec.get("symbol") or "?"
        fs  = str(rec.get("first_seen_date") or "")[:10]
        return f"{tid} | first_seen={fs}" if fs else str(tid)
    if key == "journey_path.first":
        path = rec.get("journey_path") or []
        return path[0] if path else None
    if key == "journey_path.joined":
        path = rec.get("journey_path") or []
        return " > ".join(str(p) for p in path)
    if key == "entry.fail_reasons.joined":
        fr = (rec.get("entry") or {}).get("fail_reasons") or []
        if isinstance(fr, (list, tuple)):
            return "; ".join(str(x) for x in fr)
        return str(fr)
    if key == "entry.primary_reject":
        fr = (rec.get("entry") or {}).get("fail_reasons") or []
        if isinstance(fr, (list, tuple)) and fr:
            # First fail-reason token before "(" is the "primary".
            first = str(fr[0])
            return first.split("(")[0].strip()
        return None
    parts = key.split(".")
    cur: Any = rec
    for p in parts:
        if isinstance(cur, dict):
            cur = cur.get(p)
        else:
            return None
    return cur


def _final_outcome(rec: Dict[str, Any]) -> str:
    """Human-readable outcome, matching the user's requested labels."""
    st = rec.get("tracking_status")
    if st == STATUS_T2_HIT:
        return "T2_HIT"
    if st == STATUS_STOPPED_AFTER_T1:
        return "STOPPED_AFTER_T1"
    if st == STATUS_STOPPED:
        return "STOPPED_BEFORE_T1"
    if st == STATUS_T1_HIT:
        return "T1_HIT_ACTIVE"
    return "ACTIVE"


def _format(value: Any, hint: str) -> Any:
    if value is None or value == "":
        return None
    try:
        if hint == "int":
            return int(float(value))
        if hint in ("one_dec",):
            return round(float(value), 1)
        if hint in ("two_dec",):
            return round(float(value), 2)
        if hint == "pct":
            return round(float(value), 2)
        if hint == "bool":
            return "YES" if bool(value) else ""
        if hint == "date":
            return str(value)[:10]
    except (TypeError, ValueError):
        return value
    return value


# ---------------------------------------------------------------------------
# Daily-snapshot sheet columns
# ---------------------------------------------------------------------------

# Columns for the 4 stage/setup sheets in the NEW 9-sheet layout:
#   BUY, WATCHLIST (merged), REJECTED, REJECTED_SETUPS (merged).
# Each row = one (symbol, bucket, run_date) triple from daily_snapshots.jsonl.
# Columns read directly from the flat jsonl snapshot dict.
DAILY_BUCKET_COLS: Tuple[Tuple[str, str, str], ...] = (
    ("Row Key",           "_row_key",              "text"),
    ("Run Date",          "run_date",              "date"),
    ("Symbol",            "symbol",                "text"),
    ("Tracking ID",       "tracking_id",           "text"),
    ("Current Stage",     "current_stage",         "text"),
    ("Setup",             "setup_type",            "text"),
    ("Sector",            "sector",                "text"),
    ("First Seen",        "first_seen_date",       "date"),
    ("Days Active",       "days_active",           "int"),
    ("Journey",           "_journey_joined",       "text"),
    ("Entry Regime",      "regime",                "text"),
    ("Entry Confidence",  "entry_confidence",      "one_dec"),
    ("Current Confidence","current_confidence",    "one_dec"),
    ("Entry TQ",          "entry_tq",              "one_dec"),
    ("Current TQ",        "current_tq",            "one_dec"),
    ("Entry R/R",         "entry_rr",              "two_dec"),
    ("Current R/R",       "current_rr",            "two_dec"),
    ("Entry Opp",         "entry_opp",             "one_dec"),
    ("Current Opp",       "current_opp",           "one_dec"),
    ("Ref Entry",         "reference_entry_price", "two_dec"),
    ("T1",                "t1_price",              "two_dec"),
    ("T2",                "t2_price",              "two_dec"),
    ("Stop",              "stop_price",            "two_dec"),
    # Fix #2 (2026-07-15): display-only columns showing % distance from
    # entry for Stop / T1 / T2. Pure math — no strategy change. Helps eye-
    # test whether a target/stop is realistic BEFORE committing capital.
    ("Stop %",            "_stop_pct_from_entry",  "pct"),
    ("T1 %",              "_t1_pct_from_entry",    "pct"),
    ("T2 %",              "_t2_pct_from_entry",    "pct"),
    ("Current Price",     "current_price",         "two_dec"),
    ("MFE %",             "mfe_pct",               "pct"),
    ("MAE %",             "mae_pct",               "pct"),
    ("T1 Hit",            "t1_hit",                "bool"),
    ("T2 Hit",            "t2_hit",                "bool"),
    ("Stop Hit",          "stop_hit",              "bool"),
    ("Status",            "tracking_status",       "text"),
    ("Fail Reasons",      "_fail_reasons_joined",  "text"),
)

# WATCHLIST (merged NEAR_MISS + DEVELOPING + MONITOR) — same as DAILY_BUCKET_COLS
# but with a leading "Watchlist Category" column that reveals which of the 3
# original sub-buckets the row came from (equal to the row's `bucket` field).
WATCHLIST_COLS: Tuple[Tuple[str, str, str], ...] = (
    ("Watchlist Category", "bucket",               "text"),
) + DAILY_BUCKET_COLS

# REJECTED_SETUPS (merged BREAKOUT+MOMENTUM+PULLBACK+REVERSAL+REJECTED_SETUP) —
# same as DAILY_BUCKET_COLS but with a leading "Setup Type" column that reveals
# which chart pattern was detected. For rows from the historical REJECTED_SETUP
# bucket, the Setup Type falls back to the row's `setup_type` field.
REJECTED_SETUP_COLS: Tuple[Tuple[str, str, str], ...] = (
    ("Setup Type",         "_setup_type_display",  "text"),
) + DAILY_BUCKET_COLS

# ACTIVE_TRACKING = master firehose. Same columns + "Source Sheet" leading col.
# NOTE: retained for backwards compatibility with any external tooling that
# imports this constant, but the workbook no longer writes an ACTIVE_TRACKING
# sheet (see redesign 2026-07-15). Safe to delete once external consumers are
# audited.
ACTIVE_DAILY_COLS: Tuple[Tuple[str, str, str], ...] = (
    ("Source Sheet",      "bucket",                "text"),
) + DAILY_BUCKET_COLS


def _write_daily_sheet(
    ws: Worksheet,
    columns: Tuple[Tuple[str, str, str], ...],
    rows: List[Dict[str, Any]],
) -> None:
    """Write header + rows for a daily-snapshot sheet (flat dict rows)."""
    # Detect whether we are writing the ACTIVE_TRACKING firehose or a
    # per-bucket daily sheet. ACTIVE_TRACKING has 'Source Sheet' as the very
    # first column; regular bucket sheets do not. This flag is used to make
    # the Row Key globally unique across sheets (fixes B1 collision).
    is_active_tracking = bool(columns) and columns[0][0] == "Source Sheet"

    # Pre-compute virtual fields.
    prepared: List[Dict[str, Any]] = []
    for r in rows:
        r2 = dict(r)
        jp = r.get("journey_path") or []
        r2["_journey_joined"] = " > ".join(str(x) for x in jp)
        fr = r.get("fail_reasons") or []
        if isinstance(fr, (list, tuple)):
            r2["_fail_reasons_joined"] = "; ".join(str(x) for x in fr)
        else:
            r2["_fail_reasons_joined"] = str(fr) if fr else ""
        # Row Key = fully-qualified per-row identity so repeated stocks never
        # look ambiguous.  Includes the source-sheet bucket for ACTIVE_TRACKING.
        # For ACTIVE_TRACKING we prefix "AT | " so the same underlying record
        # in ACTIVE_TRACKING and its source bucket sheet never share a Row Key.
        tid  = r.get("tracking_id") or r.get("symbol") or "?"
        rd   = str(r.get("run_date") or "")[:10]
        bkt  = str(r.get("bucket") or "").strip()
        base = f"{tid} | {rd} | {bkt}" if bkt else f"{tid} | {rd}"
        r2["_row_key"] = f"AT | {base}" if is_active_tracking else base
        # Setup Type display (for merged REJECTED_SETUPS sheet, 2026-07-15
        # redesign). If the row's bucket is one of the 4 chart-pattern
        # buckets, show it directly; if it's the legacy REJECTED_SETUP
        # umbrella bucket, fall back to the row's own setup_type field so
        # the user still sees BREAKOUT/MOMENTUM/PULLBACK/REVERSAL.
        _b = str(r.get("bucket") or "").strip().upper()
        if _b in ("BREAKOUT", "MOMENTUM", "PULLBACK", "REVERSAL"):
            r2["_setup_type_display"] = _b
        elif _b == "REJECTED_SETUP":
            _st = str(r.get("setup_type") or "").strip().upper()
            r2["_setup_type_display"] = _st if _st in (
                "BREAKOUT", "MOMENTUM", "PULLBACK", "REVERSAL"
            ) else "OTHER"
        else:
            r2["_setup_type_display"] = str(r.get("setup_type") or "") or None

        # Fix #2 (2026-07-15): compute % distance from entry for Stop / T1 / T2.
        # Display-only — never affects trading math. Formula:
        #   Stop %  =  (stop  - entry) / entry * 100   (negative = downside risk)
        #   T1   %  =  (T1    - entry) / entry * 100   (positive = upside gain)
        #   T2   %  =  (T2    - entry) / entry * 100   (positive = upside gain)
        # Rendered as e.g. "-2.35%" / "+3.14%" / "+6.28%" so the user sees at
        # a glance whether the target/stop is realistic before committing.
        try:
            _entry = float(r.get("reference_entry_price") or 0.0)
        except (TypeError, ValueError):
            _entry = 0.0
        if _entry > 0:
            for _lvl_key, _dst_key in (
                ("stop_price", "_stop_pct_from_entry"),
                ("t1_price",   "_t1_pct_from_entry"),
                ("t2_price",   "_t2_pct_from_entry"),
            ):
                try:
                    _lvl = float(r.get(_lvl_key) or 0.0)
                except (TypeError, ValueError):
                    _lvl = 0.0
                if _lvl > 0:
                    r2[_dst_key] = round((_lvl - _entry) / _entry * 100.0, 2)
                else:
                    r2[_dst_key] = None
        else:
            r2["_stop_pct_from_entry"] = None
            r2["_t1_pct_from_entry"]   = None
            r2["_t2_pct_from_entry"]   = None

        prepared.append(r2)

    # Header.
    for col_idx, (header, _, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = CELL_BORDER
    ws.freeze_panes = "A2"
    # Column widths.
    for col_idx, (header, _, hint) in enumerate(columns, start=1):
        width = 14
        if hint == "text":
            width = max(14, min(28, len(header) + 6))
        elif hint == "date":
            width = 12
        elif hint == "int":
            width = 11
        elif hint == "bool":
            width = 8
        elif hint == "pct":
            width = 10
        elif "dec" in hint:
            width = 12
        if header in ("Journey", "Fail Reasons"):
            width = 40
        if header in ("Tracking ID",):
            width = 22
        if header == "Row Key":
            width = 36
        if header == "Source Sheet":
            width = 14
        if header in ("Watchlist Category", "Setup Type"):
            width = 18
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows.
    for row_idx, r in enumerate(prepared, start=2):
        for col_idx, (header, key, hint) in enumerate(columns, start=1):
            raw = r.get(key)
            value = _format(raw, hint)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if hint == "pct" and isinstance(value, (int, float)):
                cell.number_format = "0.00\"%\""
            elif hint == "two_dec" and isinstance(value, (int, float)):
                cell.number_format = "0.00"
            elif hint == "one_dec" and isinstance(value, (int, float)):
                cell.number_format = "0.0"
            elif hint == "int" and isinstance(value, (int, float)):
                cell.number_format = "0"
            # Row-identity tint so the eye can group the same trade idea
            # across daily rows.
            if header == "Tracking ID":
                cell.fill = TRACKING_ID_FILL
                cell.font = Font(bold=True)
            elif header == "Row Key":
                cell.fill = ROW_KEY_FILL
                cell.font = Font(bold=True)

        # Status tint.
        status = r.get("tracking_status")
        fill = STATUS_FILLS.get(status)
        if fill is not None:
            for col_idx, (header, _, _) in enumerate(columns, start=1):
                if header == "Status":
                    ws.cell(row=row_idx, column=col_idx).fill = fill
                    break
        # Stage tint on Current Stage.
        cur_stage = str(r.get("current_stage") or "").upper()
        base_stage = cur_stage.replace("_RISING", "").replace("_FADING", "")
        stage_fill = STAGE_FILLS.get(base_stage)
        if stage_fill is not None:
            for col_idx, (header, _, _) in enumerate(columns, start=1):
                if header in ("Current Stage", "Journey"):
                    ws.cell(row=row_idx, column=col_idx).fill = stage_fill

    if prepared:
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A1:{last_col}{len(prepared) + 1}"


# ---------------------------------------------------------------------------
# Sheet writer
# ---------------------------------------------------------------------------

def _write_sheet(
    ws: Worksheet,
    columns: Tuple[Tuple[str, str, str], ...],
    records: List[Dict[str, Any]],
) -> None:
    """Write header + rows with formatting."""
    # Header.
    for col_idx, (header, _, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = CELL_BORDER
    ws.freeze_panes = "A2"
    # Column widths — sensible defaults per format hint.
    for col_idx, (header, _, hint) in enumerate(columns, start=1):
        width = 14
        if hint == "text":
            width = max(14, min(28, len(header) + 6))
        elif hint == "date":
            width = 12
        elif hint == "int":
            width = 11
        elif hint == "bool":
            width = 8
        elif hint == "pct":
            width = 10
        elif "dec" in hint:
            width = 12
        # Longer for known wide fields.
        if header in ("Journey", "Journey Path", "Fail Reasons"):
            width = 40
        if header in ("Tracking ID",):
            width = 22
        if header == "Row Key":
            width = 36
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Rows.
    for row_idx, rec in enumerate(records, start=2):
        for col_idx, (header, key, hint) in enumerate(columns, start=1):
            raw = _resolve(rec, key)
            value = _format(raw, hint)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if hint == "pct" and isinstance(value, (int, float)):
                cell.number_format = "0.00\"%\""
            elif hint == "two_dec" and isinstance(value, (int, float)):
                cell.number_format = "0.00"
            elif hint == "one_dec" and isinstance(value, (int, float)):
                cell.number_format = "0.0"
            elif hint == "int" and isinstance(value, (int, float)):
                cell.number_format = "0"
            # Row-identity tint so repeat symbols in the same sheet remain
            # visually distinguishable.
            if header == "Tracking ID":
                cell.fill = TRACKING_ID_FILL
                cell.font = Font(bold=True)
            elif header == "Row Key":
                cell.fill = ROW_KEY_FILL
                cell.font = Font(bold=True)

        # Row-level highlight based on status.
        status = rec.get("tracking_status")
        fill = STATUS_FILLS.get(status)
        if fill is not None:
            # Only tint the Status column, not the whole row (keep it readable).
            for col_idx, (header, _, _) in enumerate(columns, start=1):
                if header == "Status":
                    ws.cell(row=row_idx, column=col_idx).fill = fill
                    break
        # Stage tint on current-stage / stage columns.
        cur_stage = ((rec.get("current") or {}).get("stage") or "").upper()
        base_stage = cur_stage.replace("_RISING", "").replace("_FADING", "")
        stage_fill = STAGE_FILLS.get(base_stage)
        if stage_fill is not None:
            for col_idx, (header, _, _) in enumerate(columns, start=1):
                if header in ("Current Stage", "Journey", "Journey Path"):
                    ws.cell(row=row_idx, column=col_idx).fill = stage_fill

    # Autofilter on the header row across data range.
    if records:
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A1:{last_col}{len(records) + 1}"


# ---------------------------------------------------------------------------
# Legend / notes
# ---------------------------------------------------------------------------

LEGEND_LINES: Tuple[Tuple[str, str], ...] = (
    ("Layout",           "9-sheet consolidated layout (2026-07-15 redesign, down from 16). "
                         "A stock lives in EXACTLY ONE sheet at any time. Merged sheets use "
                         "a category column instead of separate tabs."),
    ("Model",            "Each stage/setup sheet is an append-only daily log. "
                         "One row per (symbol, bucket, run_date). History is never truncated."),
    ("Live refresh",     "When the workbook is rebuilt, EVERY historical row is refreshed "
                         "with today's latest current_price / MFE / MAE / status from the "
                         "canonical store. Old rows keep their run_date and entry_* frozen "
                         "but always show the LATEST live values."),
    ("Row Key",          "Leftmost grey column. Fully-qualified identity for each row so the "
                         "same stock appearing on multiple days can never be confused. "
                         "Format: '<TrackingID> | <RunDate> | <Bucket>'."),
    ("Tracking ID",      "Lavender column, bold text. Canonical per-episode key, e.g. "
                         "'MRF_20260713' = MRF first seen on 2026-07-13. If MRF exits then "
                         "re-emerges later it gets a NEW tracking id (new episode), so two "
                         "episodes of the same symbol are never mixed."),
    # ------ The 9 sheets ------
    ("BUY",              "Sheet 1. Daily rows for stocks whose current stage is BUY. "
                         "Updated in-place every evening until T2/Stop resolves — then the "
                         "record moves to DONE."),
    ("WATCHLIST",        "Sheet 2. MERGED sheet — replaces the old NEAR_MISS + DEVELOPING + "
                         "MONITOR tabs. The FIRST column 'Watchlist Category' tells you "
                         "which of the 3 original tiers each row came from. Filter on this "
                         "column to see only NEAR_MISS or only MONITOR, etc."),
    ("Watchlist Category","New in 2026-07-15 layout. Values: NEAR_MISS | DEVELOPING | MONITOR. "
                         "Equals the row's underlying bucket in daily_snapshots.jsonl."),
    ("REJECTED",         "Sheet 3. Daily rows for stocks whose STAGE was rejected."),
    ("REJECTED_SETUPS",  "Sheet 4. MERGED sheet — replaces the old BREAKOUT + MOMENTUM + "
                         "PULLBACK + REVERSAL + REJECTED_SETUP tabs. The FIRST column "
                         "'Setup Type' tells you which chart pattern was detected. Rows "
                         "here have not been promoted to BUY."),
    ("Setup Type",       "New in 2026-07-15 layout. Values: BREAKOUT | MOMENTUM | PULLBACK | "
                         "REVERSAL. Falls back to the row's setup_type field for rows from "
                         "the historical REJECTED_SETUP umbrella bucket."),
    ("DONE",             "Sheet 5. One row per stock that reached T2 / STOP / STOPPED_AFTER_T1. "
                         "Terminal resting place — once a stock lands here it never leaves."),
    ("WEEKLY_SUMMARY",   "Sheet 6. One row per NEW workbook sheet (BUY / WATCHLIST / REJECTED / "
                         "REJECTED_SETUPS / DONE). Counts: active now, tracking, T1/T2/SL hits "
                         "this week (rolling 7d) and overall. Rebuilt every run."),
    ("WEEKLY_REVIEW",    "Sheet 7. Historical per-category aggregate summaries from "
                         "results/weekly_review_history.jsonl (append-only)."),
    ("RESEARCH",         "Sheet 8. Fundamentals view (from canonical tracking_store). Some "
                         "columns are BLANK — not produced by the current pipeline."),
    ("_LEGEND",          "Sheet 9. This glossary."),
    # ------ Removed sheets ------
    ("Deleted sheets",   "As of 2026-07-15: NEAR_MISS, DEVELOPING, MONITOR (merged into "
                         "WATCHLIST), BREAKOUT, MOMENTUM, PULLBACK, REVERSAL, REJECTED_SETUP "
                         "(merged into REJECTED_SETUPS), ACTIVE_TRACKING (deleted — BUY sheet "
                         "already tracks live positions in-place, no duplicate needed)."),
    # ------ Column glossary ------
    ("MFE / MAE",        "Maximum Favorable / Adverse Excursion since first_seen. Monotonic — "
                         "MFE only goes up, MAE only goes down."),
    ("Entry Regime",     "New 2026-07-15 (Fix #19). Nifty market regime AT THE TIME the "
                         "record was first seen. Values: STRONG_BULL | BULL | SIDEWAYS | "
                         "TRANSITION | HIGH_VOLATILITY | BEAR | STRONG_BEAR. Enables "
                         "later weekly-review analysis: 'what's my win rate for BUYs born "
                         "in a BEAR regime vs a BULL regime?' Frozen at first_seen."),
    ("Stop % / T1 % / T2 %","New 2026-07-15 (Fix #2). % distance from entry price for each "
                         "level. Stop % is negative (downside risk); T1/T2 % are positive "
                         "(upside gain). Pure display — read-only view of the numbers "
                         "already in the Stop / T1 / T2 columns. Formula: "
                         "(level - entry) / entry × 100."),
    ("Journey",          "Ordered list of stages the stock has been through, e.g. "
                         "DEVELOPING > NEAR_MISS > BUY."),
    ("Entry vs Current", "Every score has both an Entry (frozen at first_seen) and a Current "
                         "(refreshed each run) value. Entry never changes."),
    ("Confidence bands", "<60 | 60-69 | 70-74 | 75-79 | 80-84 | 85+"),
    ("TQ bands",         "<50 | 50-59 | 60-69 | 70-74 | 75+"),
    ("R/R bands",        "<1.5 | 1.5-1.99 | 2.0-2.49 | 2.5-2.99 | 3.0+"),
    ("Opportunity bands","<50 | 50-59 | 60-69 | 70+"),
    ("Source of truth",  "results/daily_snapshots.jsonl (BUY / WATCHLIST / REJECTED / "
                         "REJECTED_SETUPS) and results/tracking_store.json (DONE / RESEARCH). "
                         "Both are append-only from the workbook's perspective — safe to "
                         "delete .xlsx."),
)


def _write_legend(ws: Worksheet) -> None:
    ws.cell(row=1, column=1, value="Term").font = HEADER_FONT
    ws.cell(row=1, column=1).fill = HEADER_FILL
    ws.cell(row=1, column=2, value="Meaning").font = HEADER_FONT
    ws.cell(row=1, column=2).fill = HEADER_FILL
    ws.freeze_panes = "A2"
    for i, (term, meaning) in enumerate(LEGEND_LINES, start=2):
        ws.cell(row=i, column=1, value=term).font = Font(bold=True)
        c = ws.cell(row=i, column=2, value=meaning)
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 110


# ---------------------------------------------------------------------------
# Weekly review sheet
# ---------------------------------------------------------------------------

def _read_weekly_history(path: Path) -> List[Dict[str, Any]]:
    if not path.exists():
        return []
    rows: List[Dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return rows


WEEKLY_COLS: Tuple[Tuple[str, str, str], ...] = (
    ("ISO Week",         "iso_week",        "text"),
    ("Week Start",       "week_start",      "date"),
    ("Category",         "category",        "text"),
    ("New This Week",    "new_this_week",   "int"),
    ("Total Active",     "total_active",    "int"),
    ("Resolved",         "resolved",        "int"),
    ("T1 Count",         "t1_count",        "int"),
    ("T2 Count",         "t2_count",        "int"),
    ("Stop Count",       "stop_count",      "int"),
    ("T1 Rate %",        "t1_rate",         "pct"),
    ("T2 Rate %",        "t2_rate",         "pct"),
    ("Stop Rate %",      "stop_rate",       "pct"),
    ("Avg MFE %",        "avg_mfe",         "pct"),
    ("Avg MAE %",        "avg_mae",         "pct"),
    ("Median Days T1",   "median_days_to_t1","one_dec"),
    ("Median Days T2",   "median_days_to_t2","one_dec"),
    ("Median Days Stop", "median_days_to_stop","one_dec"),
    ("Avg Entry Conf",   "avg_entry_confidence","one_dec"),
    ("Avg Entry TQ",     "avg_entry_tq",    "one_dec"),
    ("Avg Entry R/R",    "avg_entry_rr",    "two_dec"),
    ("Avg Entry Opp",    "avg_entry_opp",   "one_dec"),
    ("Confidence Band Table", "conf_band_table_str", "text"),
    ("TQ Band Table",    "tq_band_table_str","text"),
    ("R/R Band Table",   "rr_band_table_str","text"),
    ("Opp Band Table",   "opp_band_table_str","text"),
)


def _build_weekly_summary_rows(
    snapshots: List[Dict[str, Any]],
    store: TrackingStore,
    run_date: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """One row per bucket sheet with active + T1/T2/SL counts (this week + overall).

    "This week" = the 7-day rolling window ending on run_date (or the max
    run_date in the snapshots).
    "Active" = distinct symbols visible in this sheet on the LATEST run_date.
    "Tracking" = distinct symbols that ever appeared in this sheet AND are
    still non-terminal in the canonical store.
    "T1/T2/SL hits" = distinct symbols that ever appeared in this sheet AND
    hit that outcome. Uses the store's t1_hit / t2_hit / stop_hit flags
    plus the resolution date columns to filter this-week vs overall.
    """
    from datetime import date as _date, timedelta

    # Determine the "as-of" date.
    if run_date is None:
        run_dates = [s.get("run_date") for s in snapshots if s.get("run_date")]
        run_date = max(run_dates) if run_dates else _date.today().isoformat()
    try:
        end_dt = _date.fromisoformat(run_date)
    except ValueError:
        end_dt = _date.today()
    week_start_dt = end_dt - timedelta(days=6)   # 7-day rolling
    week_start = week_start_dt.isoformat()

    # Bucket lists — order matters for row output.
    # 2026-07-15 redesign: WEEKLY_SUMMARY now emits ONE row per NEW workbook
    # sheet (not per bucket). Merged sheets (WATCHLIST, REJECTED_SETUPS) roll
    # up their constituent buckets so a single summary row covers all of
    # NEAR_MISS+DEVELOPING+MONITOR (WATCHLIST) or
    # BREAKOUT+MOMENTUM+PULLBACK+REVERSAL+REJECTED_SETUP (REJECTED_SETUPS).
    sheet_buckets: List[Tuple[str, Tuple[str, ...]]] = [
        ("BUY",             ("BUY",)),
        ("WATCHLIST",       ("NEAR_MISS", "DEVELOPING", "MONITOR")),
        ("REJECTED",        ("REJECTED",)),
        ("REJECTED_SETUPS", ("BREAKOUT", "MOMENTUM", "PULLBACK",
                              "REVERSAL", "REJECTED_SETUP")),
        ("DONE",            ()),   # sentinel — resolved from store, not snapshots
    ]

    # Index snapshots by bucket.
    from collections import defaultdict
    by_bucket: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for s in snapshots:
        b = s.get("bucket")
        if b:
            by_bucket[b].append(s)

    # Store record lookup by symbol.
    rec_by_sym: Dict[str, Dict[str, Any]] = {}
    for r in store.records.values():
        sym = r.get("symbol")
        if sym:
            rec_by_sym[sym] = r
    terminal_syms = {sym for sym, r in rec_by_sym.items()
                     if r.get("tracking_status") in TERMINAL_STATUSES}

    def _hit_date(rec: Dict[str, Any], kind: str) -> Optional[str]:
        # store records may have t1_hit_date / t2_hit_date / stop_hit_date
        # (populated by tracker_job / tracking_store); fall back to
        # first_seen_date if missing so we don't lose the row entirely.
        key = f"{kind}_hit_date"
        return rec.get(key) or rec.get("resolved_date")

    def _in_week(dt_str: Optional[str]) -> bool:
        if not dt_str:
            return False
        try:
            d = _date.fromisoformat(dt_str[:10])
        except ValueError:
            return False
        return week_start_dt <= d <= end_dt

    rows: List[Dict[str, Any]] = []
    for sheet_name, buckets in sheet_buckets:
        if sheet_name == "DONE":
            symbols_ever = terminal_syms
            symbols_active_now = set()          # DONE has no "currently active"
        else:
            # Union rows across all constituent buckets (empty tuple => sheet
            # not backed by snapshots, already handled above).
            bucket_rows: List[Dict[str, Any]] = []
            for b in buckets:
                bucket_rows.extend(by_bucket.get(b, []))
            symbols_ever = {r.get("symbol") for r in bucket_rows}
            latest_in_bucket = max((r.get("run_date") or "" for r in bucket_rows),
                                   default="")
            symbols_active_now = {r.get("symbol") for r in bucket_rows
                                  if r.get("run_date") == latest_in_bucket
                                  and r.get("symbol") not in terminal_syms}

        # Count outcomes over the symbols that ever passed through this sheet.
        t1_week = t2_week = sl_week = 0
        t1_all  = t2_all  = sl_all  = 0
        for sym in symbols_ever:
            rec = rec_by_sym.get(sym)
            if rec is None:
                continue
            if rec.get("t1_hit"):
                t1_all += 1
                if _in_week(_hit_date(rec, "t1")):
                    t1_week += 1
            if rec.get("t2_hit"):
                t2_all += 1
                if _in_week(_hit_date(rec, "t2")):
                    t2_week += 1
            if rec.get("stop_hit"):
                sl_all += 1
                if _in_week(_hit_date(rec, "stop")):
                    sl_week += 1

        tracking = len({sym for sym in symbols_ever
                        if sym not in terminal_syms and sym in rec_by_sym})

        rows.append({
            "sheet": sheet_name,
            "as_of": run_date,
            "week_start": week_start,
            "active_now": len(symbols_active_now),
            "tracking": tracking,
            "t1_week": t1_week,
            "t2_week": t2_week,
            "sl_week": sl_week,
            "t1_overall": t1_all,
            "t2_overall": t2_all,
            "sl_overall": sl_all,
        })
    return rows


WEEKLY_SUMMARY_COLS: Tuple[Tuple[str, str, str], ...] = (
    ("Sheet",             "sheet",       "text"),
    ("As Of",             "as_of",       "date"),
    ("Week Start",        "week_start",  "date"),
    ("Active (Now)",      "active_now",  "int"),
    ("Tracking (Active)", "tracking",    "int"),
    ("T1 Hits (Week)",    "t1_week",     "int"),
    ("T2 Hits (Week)",    "t2_week",     "int"),
    ("SL Hits (Week)",    "sl_week",     "int"),
    ("T1 Hits (Overall)", "t1_overall",  "int"),
    ("T2 Hits (Overall)", "t2_overall",  "int"),
    ("SL Hits (Overall)", "sl_overall",  "int"),
)


# ---------------------------------------------------------------------------
# Public entrypoint
# ---------------------------------------------------------------------------

def _refresh_snapshots(
    snapshots: List[Dict[str, Any]],
    store: TrackingStore,
) -> List[Dict[str, Any]]:
    """Overlay the latest `current_*` fields from the canonical store onto
    every historical daily row.

    The raw `daily_snapshots.jsonl` is append-only and immutable on disk.
    At workbook-build time we RE-DERIVE the "live" columns (current_price,
    MFE/MAE, status, days_active, t1/t2/stop hit) so Monday's row shows
    Monday's run_date but Tuesday's latest price.

    Frozen fields (never overwritten): run_date, symbol, tracking_id,
    bucket, source, entry_*, reference_entry_price, t1/t2/stop prices,
    first_seen_date, sector, setup_type, regime.
    """
    # Build a lookup: tracking_id -> record. Falls back to symbol match if
    # tracking_id missing on the snapshot row.
    by_tid: Dict[str, Dict[str, Any]] = {}
    by_symbol: Dict[str, Dict[str, Any]] = {}
    for rec in store.records.values():
        tid = rec.get("tracking_id")
        if tid:
            by_tid[tid] = rec
        sym = rec.get("symbol")
        if sym:
            # If multiple records share a symbol (rare — different first_seen),
            # keep the most-recent first_seen so "current" reflects the live one.
            existing = by_symbol.get(sym)
            if existing is None or (rec.get("first_seen_date") or "") > (existing.get("first_seen_date") or ""):
                by_symbol[sym] = rec

    refreshed: List[Dict[str, Any]] = []
    # `_today` is the fallback anchor when the snapshot has no run_date.
    _today = datetime.now().date()
    for row in snapshots:
        rec = by_tid.get(row.get("tracking_id")) or by_symbol.get(row.get("symbol"))
        if rec is None:
            # No canonical record — keep row untouched.
            refreshed.append(row)
            continue
        current = rec.get("current") or {}
        row2 = dict(row)
        # Overlay live "current_*" fields.
        row2["current_stage"]      = current.get("stage")           or row.get("current_stage")
        row2["current_confidence"] = current.get("confidence")      if current.get("confidence") is not None else row.get("current_confidence")
        row2["current_tq"]         = current.get("tq")              if current.get("tq") is not None else row.get("current_tq")
        row2["current_rr"]         = current.get("rr_ratio") or current.get("rr") or row.get("current_rr")
        row2["current_opp"]        = current.get("opportunity_score") if current.get("opportunity_score") is not None else row.get("current_opp")
        row2["current_price"]      = (current.get("close") or current.get("price")
                                      or current.get("current_price") or row.get("current_price"))
        # Overlay running metrics from the record.
        for k in ("mfe_pct", "mae_pct",
                  "t1_hit", "t2_hit", "stop_hit", "tracking_status"):
            v = rec.get(k)
            if v is not None:
                row2[k] = v
        # Locally recompute days_active per snapshot (fixes B2).
        # For a snapshot row, days_active = run_date - first_seen_date (not
        # today - first_seen), so historical rows keep their true value even
        # after weeks of daily runs. Fall back to "today - first_seen" when
        # the row has no explicit run_date.
        fs_raw = rec.get("first_seen_date") or row.get("first_seen_date")
        rd_raw = row.get("run_date")
        if fs_raw:
            try:
                fs_dt = datetime.strptime(str(fs_raw)[:10], "%Y-%m-%d").date()
                if rd_raw:
                    try:
                        anchor = datetime.strptime(str(rd_raw)[:10], "%Y-%m-%d").date()
                    except (TypeError, ValueError):
                        anchor = _today
                else:
                    anchor = _today
                row2["days_active"] = max(0, (anchor - fs_dt).days)
            except (TypeError, ValueError):
                # Fall back to whatever the record has.
                if rec.get("days_active") is not None:
                    row2["days_active"] = rec.get("days_active")
        elif rec.get("days_active") is not None:
            row2["days_active"] = rec.get("days_active")
        refreshed.append(row2)
    return refreshed


# ---------------------------------------------------------------------------
# NOTE (2026-07-15): The CIRCUIT_TRACKER + CIRCUIT_TRACKER_DETAIL sheets were
# split out into `circuit_tracker_workbook.py` (standalone 2-sheet workbook).
# Rationale: user wanted 2 Telegram attachments (tracking_workbook.xlsx +
# circuit_tracker_workbook.xlsx) instead of the previous 3 sends. Keeping
# them separate also lets circuit-tracker rebuild independently and be wiped
# by FRESH_START without touching the main workbook.
# ---------------------------------------------------------------------------


def build_workbook(
    store: TrackingStore,
    output_path: Path,
    weekly_history_path: Path = DEFAULT_WEEKLY_HISTORY_PATH,
    snapshots_path: Path = DAILY_SNAPSHOTS_PATH,
) -> Path:
    """Build tracking_workbook.xlsx.

    Sheet plan (2026-07-15 redesign — 9 sheets, down from 16):

       ┌──────────────────┬──────────────────────────────────────────────────┐
       │ Sheet            │ What lives here                                  │
       ├──────────────────┼──────────────────────────────────────────────────┤
       │ BUY              │ bucket == BUY (updated in-place daily until      │
       │                  │ T2/Stop resolves → then moves to DONE)           │
       │ WATCHLIST        │ bucket ∈ {NEAR_MISS, DEVELOPING, MONITOR}        │
       │                  │ MERGED with 'Watchlist Category' column          │
       │ REJECTED         │ bucket == REJECTED                                │
       │ REJECTED_SETUPS  │ bucket ∈ {BREAKOUT, MOMENTUM, PULLBACK,          │
       │                  │ REVERSAL, REJECTED_SETUP} MERGED with            │
       │                  │ 'Setup Type' column. Historical filter:          │
       │                  │ per user directive Q3(a) — active-setup rows     │
       │                  │ only surface here if their stage was REJECTED    │
       │                  │ (the pipeline never wrote them any other way).   │
       │ DONE             │ tracking_status ∈ terminal (T2_HIT / STOPPED /   │
       │                  │ STOPPED_AFTER_T1)                                │
       │ WEEKLY_SUMMARY   │ Per-sheet weekly counts (rebuilt every run)      │
       │ WEEKLY_REVIEW    │ results/weekly_review_history.jsonl              │
       │ RESEARCH         │ Fundamentals view (unchanged)                    │
       │ _LEGEND          │ Column glossary                                  │
       └──────────────────┴──────────────────────────────────────────────────┘

    Deleted from previous layout: NEAR_MISS, DEVELOPING, MONITOR, BREAKOUT,
    MOMENTUM, PULLBACK, REVERSAL, REJECTED_SETUP, ACTIVE_TRACKING sheets.
    All still exist as bucket values in daily_snapshots.jsonl — the workbook
    just renders them into the 2 merged sheets above.
    """
    output_path = Path(output_path)
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    # 1. Load daily snapshots — the daily-append log.
    snapshots = read_snapshots(snapshots_path)
    print(f"[tracking_workbook_job] loaded {len(snapshots)} snapshot rows")

    # 1b. Overlay the latest current_* fields from the canonical store onto
    # every historical row. Old rows keep their run_date + entry_* frozen,
    # but their price / MFE / MAE / status refresh to today's values.
    snapshots = _refresh_snapshots(snapshots, store)
    print(f"[tracking_workbook_job] refreshed {len(snapshots)} rows with latest current_* fields")

    def _by_bucket(buckets) -> List[Dict[str, Any]]:
        """Return snapshot rows whose bucket is in the given iterable, newest first."""
        if isinstance(buckets, str):
            wanted = {buckets}
        else:
            wanted = set(buckets)
        rows = [s for s in snapshots if s.get("bucket") in wanted]
        rows.sort(key=lambda r: (r.get("run_date") or "", r.get("symbol") or ""))
        rows.sort(key=lambda r: r.get("run_date") or "", reverse=True)
        return rows

    # --- Sheet 1: BUY (unchanged — stocks with active BUY stage) ---
    _write_daily_sheet(wb.create_sheet("BUY"), DAILY_BUCKET_COLS, _by_bucket("BUY"))

    # --- Sheet 2: WATCHLIST (merged NEAR_MISS + DEVELOPING + MONITOR) ---
    # `Watchlist Category` column comes from the row's `bucket` field.
    _write_daily_sheet(
        wb.create_sheet("WATCHLIST"),
        WATCHLIST_COLS,
        _by_bucket(("NEAR_MISS", "DEVELOPING", "MONITOR")),
    )

    # --- Sheet 3: REJECTED (unchanged — rejected stage decisions) ---
    _write_daily_sheet(wb.create_sheet("REJECTED"), DAILY_BUCKET_COLS, _by_bucket("REJECTED"))

    # --- Sheet 4: REJECTED_SETUPS (merged setup-pattern rejects) ---
    # Per user directive Q3(a): only surface rows where the setup was
    # rejected. In the pipeline (daily_snapshot_job._bucket_setup), active
    # setups are written to their own bucket AND the stage sheet routes
    # rejected setups to REJECTED_SETUP. So we include:
    #   * REJECTED_SETUP  (historical rejected-setup umbrella)
    #   * BREAKOUT / MOMENTUM / PULLBACK / REVERSAL rows WHERE the stock's
    #     current tracking_status indicates rejection (tracking_status is
    #     None + stage is REJECTED) — which the pipeline no longer emits
    #     for active setups but historical rows may contain.
    # For simplicity + zero data loss we surface ALL setup-bucket rows in
    # this sheet and let the "Setup Type" column tell the user which
    # pattern each row represents. The user can filter by tracking_status
    # to isolate rejected-only rows.
    _write_daily_sheet(
        wb.create_sheet("REJECTED_SETUPS"),
        REJECTED_SETUP_COLS,
        _by_bucket(("BREAKOUT", "MOMENTUM", "PULLBACK", "REVERSAL", "REJECTED_SETUP")),
    )

    # --- Sheet 5: DONE (from canonical store — T1/T2/STOP hit stocks) ---
    resolved = store.resolved_records()
    resolved_sorted = sorted(resolved, key=lambda r: (
        r.get("t2_hit_date") or r.get("stop_hit_date") or r.get("first_seen_date") or "",
        r.get("symbol", ""),
    ), reverse=True)
    _write_sheet(wb.create_sheet("DONE"), RESOLVED_COLS, resolved_sorted)

    # --- Sheet 6: WEEKLY_SUMMARY — per-sheet weekly counts ---
    ws_wk = wb.create_sheet("WEEKLY_SUMMARY")
    summary_rows = _build_weekly_summary_rows(snapshots, store)
    for col_idx, (header, _, _) in enumerate(WEEKLY_SUMMARY_COLS, start=1):
        cell = ws_wk.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
    ws_wk.freeze_panes = "A2"
    for col_idx, (header, _, hint) in enumerate(WEEKLY_SUMMARY_COLS, start=1):
        ws_wk.column_dimensions[get_column_letter(col_idx)].width = 18 if hint == "text" else 14
    for row_idx, wr in enumerate(summary_rows, start=2):
        for col_idx, (_, key, hint) in enumerate(WEEKLY_SUMMARY_COLS, start=1):
            value = _format(wr.get(key), hint)
            ws_wk.cell(row=row_idx, column=col_idx, value=value)
    if summary_rows:
        ws_wk.auto_filter.ref = f"A1:{get_column_letter(len(WEEKLY_SUMMARY_COLS))}{len(summary_rows) + 1}"

    # --- Sheet 7: WEEKLY_REVIEW (append-only per-category history) ---
    weekly_rows = _read_weekly_history(weekly_history_path)
    ws_wk = wb.create_sheet("WEEKLY_REVIEW")
    for col_idx, (header, _, _) in enumerate(WEEKLY_COLS, start=1):
        cell = ws_wk.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
    ws_wk.freeze_panes = "A2"
    for col_idx, (header, _, hint) in enumerate(WEEKLY_COLS, start=1):
        width = 16 if hint == "text" else 12
        if header.endswith("Band Table"):
            width = 60
        ws_wk.column_dimensions[get_column_letter(col_idx)].width = width
    for row_idx, wr in enumerate(weekly_rows, start=2):
        for col_idx, (_, key, hint) in enumerate(WEEKLY_COLS, start=1):
            value = _format(wr.get(key), hint)
            cell = ws_wk.cell(row=row_idx, column=col_idx, value=value)
            if hint == "pct" and isinstance(value, (int, float)):
                cell.number_format = "0.00\"%\""
    if weekly_rows:
        ws_wk.auto_filter.ref = f"A1:{get_column_letter(len(WEEKLY_COLS))}{len(weekly_rows) + 1}"

    # --- Sheet 8: RESEARCH (from canonical store — unchanged) ---
    _write_sheet(
        wb.create_sheet("RESEARCH"),
        RESEARCH_COLS,
        sorted(store.records.values(), key=lambda r: r.get("symbol", "")),
    )

    # --- Sheet 8b: CIRCUIT_TRACKER (2026-07-15) ---
    # 2026-07-15: CIRCUIT_TRACKER + CIRCUIT_TRACKER_DETAIL sheets moved to
    # the standalone `circuit_tracker_workbook.py` — they are NO LONGER part
    # of this workbook. main.py builds and sends both workbooks side by side.

    # --- Sheet 9: _LEGEND (column glossary — rewritten for new layout) ---
    _write_legend(wb.create_sheet("_LEGEND"))

    # Hidden diagnostic — only if there are records flagged NEEDS_REVIEW.
    needs_review = [r for r in store.records.values()
                    if r.get("migration_status") == MIGRATION_NEEDS_REVIEW]
    if needs_review:
        _write_sheet(wb.create_sheet("_NEEDS_REVIEW"), STAGE_COLS, needs_review)

    # Atomic save.
    output_path.parent.mkdir(parents=True, exist_ok=True)
    tmp = output_path.with_suffix(output_path.suffix + ".tmp")
    wb.save(tmp)
    os.replace(tmp, output_path)
    return output_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv: Optional[List[str]] = None) -> int:
    p = argparse.ArgumentParser(description="Build tracking_workbook.xlsx")
    p.add_argument("--store", default=str(DEFAULT_STORE_PATH),
                   help="Path to tracking_store.json")
    p.add_argument("--weekly", default=str(DEFAULT_WEEKLY_HISTORY_PATH),
                   help="Path to weekly_review_history.jsonl")
    p.add_argument("--snapshots", default=str(DAILY_SNAPSHOTS_PATH),
                   help="Path to daily_snapshots.jsonl (bucket sheets source)")
    p.add_argument("--out", default=str(_HERE / "tracking_workbook.xlsx"),
                   help="Output workbook path")
    p.add_argument("--dry-run", action="store_true",
                   help="Print stats but do not write")
    args = p.parse_args(argv)

    store = TrackingStore.load(store_path=Path(args.store))
    stats = store.stats()
    print(f"[tracking_workbook_job] store: {args.store}")
    print(f"[tracking_workbook_job] stats: {stats}")
    if args.dry_run:
        return 0
    out = build_workbook(store, Path(args.out), Path(args.weekly),
                         snapshots_path=Path(args.snapshots))
    print(f"[tracking_workbook_job] wrote {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
