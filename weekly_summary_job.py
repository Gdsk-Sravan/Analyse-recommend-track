"""
weekly_summary_job.py — Saturday Morning Weekly Recap (FEATURE 6)
GitHub Actions: runs every Saturday at 9:30 AM IST (4:00 UTC)
Sends a concise weekly summary to Telegram.
"""
import os
import json
import requests
import yfinance as yf
import numpy as np
from datetime import datetime, timedelta

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "")
TELEGRAM_CHAT_ID   = os.getenv("TELEGRAM_CHAT_ID", "")
TRACKER_FILE       = os.getenv("TRADE_TRACKER_V2_FILE", "trade_tracker.json")
CONF_HISTORY_FILE  = os.getenv("CONF_HISTORY_FILE", "confidence_history.json")
TRACKER_XLSX       = os.getenv("TRACKER_XLSX", "recommendation_tracker.xlsx")

SECTOR_PROXIES = {
    "PHARMA":        "SUNPHARMA.NS",
    "BANKING":       "HDFCBANK.NS",
    "IT":            "INFY.NS",
    "AUTO":          "MARUTI.NS",
    "CAPITAL_GOODS": "BHEL.NS",
    "ENERGY":        "RELIANCE.NS",
    "METALS":        "TATASTEEL.NS",
    "DEFENCE":       "HAL.NS",
}


def _weekly_pct(ticker: str) -> float:
    try:
        df = yf.download(ticker, period="5d", interval="1d",
                         progress=False, auto_adjust=True, multi_level_index=False)
        if df is None or len(df) < 2:
            return 0.0
        return round(
            (float(df["Close"].iloc[-1]) - float(df["Close"].iloc[0])) /
            float(df["Close"].iloc[0]) * 100, 2
        )
    except Exception:
        return 0.0


# ─────────────────────────────────────────────────────────────────────────────
# WEEKLY FACTOR SUMMARY SHEET — appends one row per factor to the xlsx
# so you can see at-a-glance which factors correlate with winners vs losers.
# ─────────────────────────────────────────────────────────────────────────────

# Factors read from the Recommendations sheet (numeric only)
_REC_FACTORS = [
    ("Opp Score",   "Opp Score"),
    ("Confidence",  "Confidence"),
    ("TQ",          "TQ"),
    ("R/R",         "R/R"),
    ("Pledge%",     "Pledge%"),
    ("ROE",         "ROE"),
    ("D/E",         "D/E"),
]

# Factors read from Daily Tracking (outcome side — one row per rec/day, we take last)
_TRK_FACTORS = [
    ("Return%",           "Return%"),
    ("Max Gain%",         "Max Gain%"),
    ("Max DD%",           "Max DD%"),
    ("Remaining Upside%", "Remaining Upside%"),
    ("Holding Days",      "Holding Days"),
]


def _safe_float(v):
    try:
        if v is None or v == "":
            return None
        return float(v)
    except (ValueError, TypeError):
        return None


def _avg(vals):
    """Average of a list of numbers, ignoring None. Returns None if empty."""
    nums = [v for v in vals if v is not None]
    if not nums:
        return None
    return round(sum(nums) / len(nums), 2)


def _fmt(v):
    """Format a cell value: '—' if None else the number."""
    return "—" if v is None else v


def write_weekly_factor_summary(week_start: str, week_end: str) -> None:
    """
    Reads Recommendations + Daily Tracking sheets from recommendation_tracker.xlsx
    and writes a new 'Weekly Factor Summary' sheet with one row per factor.

    Columns per factor:
      All Picks | Winners | Losers | Still Active | BUY | WATCHLIST | Count

    Winners  = final Status in {T1_HIT, T2_HIT, T1_HIT_ACTIVE}
    Losers   = final Status in {STOPPED, EXPIRED with Return% < 0}
    Active   = final Status in {ACTIVE, T1_HIT_ACTIVE}

    Only rows with Rec Date within the past 7 days are counted.
    """
    if not os.path.exists(TRACKER_XLSX):
        print(f"[WARN] {TRACKER_XLSX} not found — skipping factor summary")
        return

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[WARN] openpyxl not installed — skipping factor summary")
        return

    try:
        wb = openpyxl.load_workbook(TRACKER_XLSX)
    except Exception as e:
        print(f"[WARN] Cannot open {TRACKER_XLSX}: {e}")
        return

    if "Recommendations" not in wb.sheetnames:
        print("[WARN] Recommendations sheet missing — skipping factor summary")
        return

    # ── Load Recommendations for the past 7 days ──
    ws_rec = wb["Recommendations"]
    rec_headers = [c.value for c in ws_rec[1]]
    week_ago = (datetime.now() - timedelta(days=7)).date()

    recs = []  # list of dict rows within the week
    for row in ws_rec.iter_rows(min_row=2, values_only=True):
        d = dict(zip(rec_headers, row))
        date_val = d.get("Date")
        # Date may be a datetime or a "YYYY-MM-DD" string
        try:
            if isinstance(date_val, datetime):
                rec_dt = date_val.date()
            elif isinstance(date_val, str) and date_val:
                rec_dt = datetime.strptime(date_val[:10], "%Y-%m-%d").date()
            else:
                continue
        except Exception:
            continue
        if rec_dt < week_ago:
            continue
        recs.append(d)

    if not recs:
        print("[INFO] No recommendations in the past 7 days — factor summary skipped")
        return

    # ── Build final-status map from Daily Tracking (last row per Ticker+Rec Date) ──
    outcomes = {}  # key = f"{Ticker}_{Rec Date}" → dict with Status + Return% + factors
    if "Daily Tracking" in wb.sheetnames:
        ws_trk = wb["Daily Tracking"]
        trk_headers = [c.value for c in ws_trk[1]]
        for row in ws_trk.iter_rows(min_row=2, values_only=True):
            d = dict(zip(trk_headers, row))
            tkr = d.get("Ticker")
            rec_date = d.get("Rec Date")
            if not tkr or not rec_date:
                continue
            key = f"{tkr}_{rec_date}"
            day_n = _safe_float(d.get("Day#")) or 0
            prev = outcomes.get(key)
            if (prev is None) or (day_n > (_safe_float(prev.get("Day#")) or 0)):
                outcomes[key] = d

    # ── Bucket recs into winner/loser/active/BUY/WATCHLIST ──
    def _bucket(rec):
        tkr = rec.get("Ticker")
        rd  = rec.get("Date")
        rd_str = rd.strftime("%Y-%m-%d") if isinstance(rd, datetime) else str(rd or "")[:10]
        # try both key shapes since tracker_job stores whatever Excel stored
        outcome = outcomes.get(f"{tkr}_{rd_str}") or outcomes.get(f"{tkr}_{rd}")
        status = (outcome or {}).get("Status", rec.get("Status", "ACTIVE"))
        ret_pct = _safe_float((outcome or {}).get("Return%"))
        return status, ret_pct

    buckets = {"winners": [], "losers": [], "active": [], "buy": [], "watchlist": [], "all": []}
    outcome_by_rec = {}  # tkr+date → outcome row (for reading TRK factors)

    for rec in recs:
        status, ret_pct = _bucket(rec)
        buckets["all"].append(rec)
        cat = str(rec.get("Category", "") or "").upper()
        if cat == "BUY":
            buckets["buy"].append(rec)
        else:
            buckets["watchlist"].append(rec)

        if status in ("T1_HIT", "T2_HIT"):
            buckets["winners"].append(rec)
        elif status == "STOPPED" or (status == "EXPIRED" and (ret_pct or 0) < 0):
            buckets["losers"].append(rec)
        else:
            buckets["active"].append(rec)

        # remember matching outcome row for TRK factors
        tkr = rec.get("Ticker")
        rd  = rec.get("Date")
        rd_str = rd.strftime("%Y-%m-%d") if isinstance(rd, datetime) else str(rd or "")[:10]
        outcome_by_rec[id(rec)] = outcomes.get(f"{tkr}_{rd_str}") or outcomes.get(f"{tkr}_{rd}")

    # ── (Re)create the summary sheet ──
    sheet_name = "Weekly Factor Summary"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # Title rows
    ws.append([f"Weekly Factor Summary  ({week_start} — {week_end})"])
    ws.append([f"Recommendations in window: {len(recs)}  |  "
               f"Winners: {len(buckets['winners'])}  |  "
               f"Losers: {len(buckets['losers'])}  |  "
               f"Still Active: {len(buckets['active'])}"])
    ws.append([])  # blank row

    header = [
        "Factor",
        "All Picks (avg)",
        "Winners (avg)",
        "Losers (avg)",
        "Still Active (avg)",
        "BUY (avg)",
        "WATCHLIST (avg)",
        "N (All)",
    ]
    ws.append(header)

    def _col_vals(rec_key, source_dict_list):
        """Return numeric values for a factor across a list of rec-dicts."""
        return [_safe_float(r.get(rec_key)) for r in source_dict_list]

    def _trk_vals(rec_key, source_recs):
        """Return numeric values for a Daily-Tracking factor from matching outcomes."""
        out = []
        for r in source_recs:
            oc = outcome_by_rec.get(id(r))
            if oc:
                out.append(_safe_float(oc.get(rec_key)))
        return out

    # Recommendations-side factors
    for label, key in _REC_FACTORS:
        ws.append([
            label,
            _fmt(_avg(_col_vals(key, buckets["all"]))),
            _fmt(_avg(_col_vals(key, buckets["winners"]))),
            _fmt(_avg(_col_vals(key, buckets["losers"]))),
            _fmt(_avg(_col_vals(key, buckets["active"]))),
            _fmt(_avg(_col_vals(key, buckets["buy"]))),
            _fmt(_avg(_col_vals(key, buckets["watchlist"]))),
            len(buckets["all"]),
        ])

    # Blank separator
    ws.append([])
    ws.append(["── Outcome factors (from Daily Tracking) ──"])

    # Daily-Tracking-side factors
    for label, key in _TRK_FACTORS:
        ws.append([
            label,
            _fmt(_avg(_trk_vals(key, buckets["all"]))),
            _fmt(_avg(_trk_vals(key, buckets["winners"]))),
            _fmt(_avg(_trk_vals(key, buckets["losers"]))),
            _fmt(_avg(_trk_vals(key, buckets["active"]))),
            _fmt(_avg(_trk_vals(key, buckets["buy"]))),
            _fmt(_avg(_trk_vals(key, buckets["watchlist"]))),
            len(buckets["all"]),
        ])

    # Hit rates row
    ws.append([])
    total = len(recs)
    settled = len(buckets["winners"]) + len(buckets["losers"])
    hit_rate = round(len(buckets["winners"]) / settled * 100, 1) if settled else None
    ws.append([
        "Hit Rate % (settled only)",
        _fmt(hit_rate),
        "—", "—", "—",
        _fmt(round(len([r for r in buckets["buy"] if r in buckets["winners"]]) /
                   max(1, len([r for r in buckets["buy"]
                               if r in buckets["winners"] or r in buckets["losers"]])) * 100, 1)),
        _fmt(round(len([r for r in buckets["watchlist"] if r in buckets["winners"]]) /
                   max(1, len([r for r in buckets["watchlist"]
                               if r in buckets["winners"] or r in buckets["losers"]])) * 100, 1)),
        total,
    ])

    # ─────────────────────────────────────────────────────────────────────
    # SECOND TABLE — Factor → Return
    # For each factor, split picks into HIGH / MID / LOW terciles by factor
    # value, then show the average Return% + Hit Rate for each tercile.
    # Answers: "When Confidence was high, what did I actually earn?"
    # ─────────────────────────────────────────────────────────────────────
    ws.append([])
    ws.append([])
    ws.append(["── Factor → Return  (what does each factor level actually pay?) ──"])
    ws.append([
        "Factor",
        "HIGH bucket",
        "HIGH avg Return%",
        "HIGH Hit Rate%",
        "MID bucket",
        "MID avg Return%",
        "LOW bucket",
        "LOW avg Return%",
        "LOW Hit Rate%",
        "Edge (HIGH−LOW Return%)",
    ])

    def _return_for(rec):
        """Latest Return% from the matching Daily Tracking outcome row."""
        oc = outcome_by_rec.get(id(rec))
        if not oc:
            return None
        return _safe_float(oc.get("Return%"))

    def _is_winner(rec):
        oc = outcome_by_rec.get(id(rec)) or {}
        st = oc.get("Status", rec.get("Status", "ACTIVE"))
        return st in ("T1_HIT", "T2_HIT")

    def _is_loser(rec):
        oc = outcome_by_rec.get(id(rec)) or {}
        st = oc.get("Status", rec.get("Status", "ACTIVE"))
        rp = _safe_float(oc.get("Return%")) or 0
        return st == "STOPPED" or (st == "EXPIRED" and rp < 0)

    def _pct_range(vals):
        """Return (low_cut, high_cut) at the 33rd and 66th percentile."""
        clean = sorted([v for v in vals if v is not None])
        if len(clean) < 3:
            return None, None
        lo = clean[len(clean) // 3]
        hi = clean[(len(clean) * 2) // 3]
        return lo, hi

    # For each REC-side factor, tercile the picks by that factor value
    for label, key in _REC_FACTORS:
        vals_with_rec = [(_safe_float(r.get(key)), r) for r in buckets["all"]]
        just_vals = [v for v, _ in vals_with_rec if v is not None]
        lo_cut, hi_cut = _pct_range(just_vals)
        if lo_cut is None:
            # not enough data — write a placeholder row
            ws.append([label, "—", "—", "—", "—", "—", "—", "—", "—", "—"])
            continue

        high_picks, mid_picks, low_picks = [], [], []
        for v, r in vals_with_rec:
            if v is None:
                continue
            if v >= hi_cut:
                high_picks.append(r)
            elif v <= lo_cut:
                low_picks.append(r)
            else:
                mid_picks.append(r)

        def _stats(picks):
            rets = [_return_for(r) for r in picks]
            rets = [x for x in rets if x is not None]
            settled = [r for r in picks if _is_winner(r) or _is_loser(r)]
            wins    = [r for r in settled if _is_winner(r)]
            avg_ret = round(sum(rets) / len(rets), 2) if rets else None
            hit_rt  = round(len(wins) / len(settled) * 100, 1) if settled else None
            return avg_ret, hit_rt

        hi_ret, hi_hr = _stats(high_picks)
        mid_ret, _    = _stats(mid_picks)
        lo_ret, lo_hr = _stats(low_picks)

        edge = None
        if hi_ret is not None and lo_ret is not None:
            edge = round(hi_ret - lo_ret, 2)

        ws.append([
            label,
            f"≥ {hi_cut:.2f}  (n={len(high_picks)})",
            _fmt(hi_ret),
            _fmt(hi_hr),
            f"{lo_cut:.2f} – {hi_cut:.2f}  (n={len(mid_picks)})",
            _fmt(mid_ret),
            f"≤ {lo_cut:.2f}  (n={len(low_picks)})",
            _fmt(lo_ret),
            _fmt(lo_hr),
            _fmt(edge),
        ])

    # ─────────────────────────────────────────────────────────────────────
    # Formatting
    # ─────────────────────────────────────────────────────────────────────
    try:
        ws["A1"].font = Font(bold=True, size=13)
        ws["A2"].font = Font(italic=True, size=10)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        for col in range(1, len(header) + 1):
            cell = ws.cell(row=4, column=col)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        # widen columns for BOTH tables (max of both — first table uses cols A-H,
        # second Factor→Return table uses cols A-J)
        widths = [26, 22, 18, 16, 22, 18, 22, 18, 18, 24]
        for i, w in enumerate(widths, start=1):
            col_letter = chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)
            ws.column_dimensions[col_letter].width = w
    except Exception:
        pass

    try:
        wb.save(TRACKER_XLSX)
        print(f"[INFO] Weekly Factor Summary written to {TRACKER_XLSX}")
    except Exception as e:
        print(f"[WARN] Could not save {TRACKER_XLSX}: {e}")


def run_weekly_summary():
    today      = datetime.now()
    week_start = (today - timedelta(days=6)).strftime("%b %d")
    week_end   = today.strftime("%b %d, %Y")
    print(f"=== WEEKLY SUMMARY: {week_start} \u2014 {week_end} ===")

    # Write the factor-summary sheet into the xlsx first (best-effort)
    try:
        write_weekly_factor_summary(week_start, week_end)
    except Exception as e:
        print(f"[WARN] Factor summary failed: {e}")

    # Load tracker
    tracker = {}
    try:
        if os.path.exists(TRACKER_FILE):
            with open(TRACKER_FILE, "r") as f:
                tracker = json.load(f)
    except Exception as e:
        print(f"[WARN] Tracker load failed: {e}")

    active    = tracker.get("buys", [])
    watching  = tracker.get("watchlist", [])
    completed = tracker.get("completed", [])
    perf      = tracker.get("performance", {})

    # NIFTY weekly
    nifty_pct = _weekly_pct("^NSEI")

    # Sector performance
    sector_perf = {}
    for sector, proxy in SECTOR_PROXIES.items():
        pct = _weekly_pct(proxy)
        if pct != 0.0:
            sector_perf[sector] = pct

    best_sector  = max(sector_perf, key=sector_perf.get) if sector_perf else "\u2014"
    worst_sector = min(sector_perf, key=sector_perf.get) if sector_perf else "\u2014"

    # Confidence history near miss rising stocks
    rising_stocks = []
    try:
        if os.path.exists(CONF_HISTORY_FILE):
            with open(CONF_HISTORY_FILE, "r") as f:
                conf_hist = json.load(f)
            for sym, data in conf_hist.items():
                confs = data.get("confs", [])
                if len(confs) >= 3 and confs[-1] > confs[-3] + 3:
                    rising_stocks.append((sym, confs[-3], confs[-1]))
            rising_stocks.sort(key=lambda x: x[2] - x[1], reverse=True)
    except Exception:
        pass

    # Build message
    lines = []
    lines.append("=" * 40)
    lines.append("\U0001f4c5 WEEKLY SUMMARY")
    lines.append(f"   {week_start} \u2014 {week_end}")
    lines.append("=" * 40)
    lines.append("")

    # Market
    nifty_icon = "\U0001f7e2" if nifty_pct > 0 else "\U0001f534"
    lines.append("\U0001f4ca MARKET THIS WEEK")
    lines.append(f"  NIFTY: {nifty_icon} {nifty_pct:+.2f}%")
    if sector_perf:
        lines.append(
            f"  Best Sector:  {best_sector} {sector_perf.get(best_sector, 0):+.2f}%"
        )
        lines.append(
            f"  Worst Sector: {worst_sector} {sector_perf.get(worst_sector, 0):+.2f}%"
        )
    lines.append("")

    # Signals summary
    lines.append("\U0001f4cb SIGNALS THIS WEEK")
    lines.append(f"  Active Positions:  {len(active)}")
    lines.append(f"  Watching:          {len(watching)} stocks")
    lines.append(f"  Completed (total): {len(completed)}")
    if completed:
        lines.append(
            f"  Win Rate: {perf.get('win_rate', 0):.0f}% | "
            f"Avg W {perf.get('avg_win', 0):+.1f}% | "
            f"Avg L {perf.get('avg_loss', 0):+.1f}%"
        )
    lines.append("")

    # Active position summary
    if active:
        lines.append("\U0001f4bc ACTIVE POSITIONS")
        for pos in active:
            try:
                hist     = pos.get("pnl_history", [])
                pnl      = hist[-1]["pnl"] if hist else 0
                day_n    = pos.get("days_tracked", 0)
                pnl_icon = "\U0001f7e2" if pnl > 0 else ("\U0001f534" if pnl < 0 else "\u26aa")
                lines.append(
                    f"  {pnl_icon} {pos['symbol']} | Day {day_n}/15 | PnL {pnl:+.1f}%"
                )
            except Exception:
                lines.append(f"  {pos.get('symbol', '?')}")
        lines.append("")

    # Near miss with rising confidence
    if rising_stocks:
        lines.append("\U0001f4c8 RISING CONFIDENCE THIS WEEK")
        for sym, c_old, c_new in rising_stocks[:5]:
            lines.append(f"  \u2191\u2191 {sym}: {c_old:.0f} \u2192 {c_new:.0f} (+{c_new-c_old:.0f})")
        lines.append("")

    # Near miss status
    near_miss = [w for w in watching if w.get("tier") == "NEAR_MISS"]
    if near_miss:
        lines.append("\U0001f441 NEAR MISS STATUS")
        for w in near_miss[:5]:
            try:
                entry     = float(w.get("entry", 0) or 0)
                conf_gap  = float(w.get("conf_gap", 0) or 0)
                direction = w.get("direction", "\u2014")
                lines.append(
                    f"  {w['symbol']} | Gap {conf_gap:.1f} | {direction}"
                )
            except Exception:
                lines.append(f"  {w.get('symbol', '?')}")
        lines.append("")

    # Next week outlook
    lines.append("\U0001f52d NEXT WEEK OUTLOOK")
    if nifty_pct < -1:
        lines.append("  Market pulled back \u2014 watch for recovery or further weakness")
    elif nifty_pct > 1:
        lines.append("  Market up this week \u2014 momentum may continue, watch for overextension")
    else:
        lines.append("  Market flat \u2014 waiting for directional catalyst")
    lines.append("  Check evening reports for daily signals")
    lines.append("")
    lines.append("\u2500" * 40)
    lines.append("\u26a0\ufe0f  Weekly recap only. Not a trading signal.")
    lines.append("=" * 40)

    message = "\n".join(lines)

    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        print("[WARN] Telegram not configured \u2014 printing to stdout")
        print(message)
        return

    try:
        r = requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage",
            json={"chat_id": TELEGRAM_CHAT_ID, "text": message},
            timeout=15
        )
        if r.status_code == 200:
            print("[INFO] Weekly summary sent")
        else:
            print(f"[WARN] Send failed: {r.text[:100]}")
    except Exception as e:
        print(f"[WARN] Send error: {e}")


if __name__ == "__main__":
    run_weekly_summary()
