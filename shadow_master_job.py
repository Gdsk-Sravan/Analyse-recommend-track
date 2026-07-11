"""
shadow_master_job.py — Unified Shadow Master (Phase II, 2026-07-09)
=====================================================================
Single source of truth for the 4-bucket shadow log AND real BUY signal
tracking.

D3 consolidation (2026-07-09):
    - Replaces shadow_report_job.py entirely (DELETED).
    - Coexists with tracker_job.py — main.yml runs tracker_job.py first
      (fills Daily Tracking + Performance Summary sheets), then this
      script (adds bucket sheets + rollups). Both write to the same
      shadow_master.xlsx file, hence the "master" name.
    - Delegates weekly fundamentals to research_job.run_research()
      (kept as-is), with TRACKER_XLSX env override so it reads/writes
      shadow_master.xlsx.

File it writes: `shadow_master.xlsx` (~14 sheets, one growing file)

Modes:
    --mode scan-and-update   Evening full flow: append new signals to bucket
                             sheets + update prices for all OPEN rows +
                             rebuild rollups + send to Telegram.
    --mode update-only       Midday flow: price update only, no new appends.
                             Silent unless a T2/Stop hits today.
    --mode research          Weekly fundamentals: delegates to
                             research_job.run_research() (kept as-is).

SCHEDULED_RUN gate:
    SCHEDULED_RUN=true  (GitHub Actions cron)
        → read + mutate shadow_master.xlsx
        → archive dated copy to reports/archive/
        → send to Telegram
        → main.yml git-commits the result
    SCHEDULED_RUN != true (manual / local / workflow_dispatch)
        → build temp preview xlsx in $TEMP
        → send to Telegram only
        → DO NOT touch shadow_master.xlsx

Usage:
    SCHEDULED_RUN=true python shadow_master_job.py --mode scan-and-update
    python shadow_master_job.py --mode update-only
    NOTIFY_DRY_RUN=1 SCHEDULED_RUN=true python shadow_master_job.py --mode scan-and-update

All complex logic from tracker_job.py is preserved verbatim:
    - _load_partial_exit_stops   (V1 tracker.json trailed-stop sync)
    - _fetch_live_delivery_pct   (delivery signal overlay)
    - _flag_position_health      (DISTRIBUTION / STRONG_ACCUM / WEAK)
    - _detect_fresh_start_marker (.fresh_start_marker consumption)
    - _update_performance_sheet  (Performance Summary + P&L Attribution)
    - _update_portfolio_risk_sheet
    - _update_benchmark_overlay
    - _update_equity_curve_sheet

New for Phase II:
    - 4 bucket sheets (A_TAKEN, B_WATCH_ME, C_NOT_MY_STYLE, D_SO_CLOSE)
      each with the same 26-column schema:
        entry_date, symbol, setup_type, regime, confidence, entry_price,
        t1_price, t2_price, stop_price, current_price, current_pnl_pct,
        max_favorable_pct, max_adverse_pct, days_tracked,
        t1_hit_date, days_to_t1, t2_hit_date, days_to_t2,
        stop_hit_date, days_to_stop, status, long_runner_flag,
        exit_date, exit_price, final_pnl_pct, r_multiple
    - Live_Positions rollup (all OPEN rows across buckets)
    - Resolved_Today rollup (WIN_T2 / LOSS flips this run)
    - Summary + Bucket_Comparison rollups
    - Change_Log (append-only run log)

Rules:
    - Track forever until T1 OR T2 OR Stop hits — NO time-exit
    - Entry price = close of signal day
    - T1 = entry × 1.05, T2 = entry × 1.10, Stop = entry × 0.97
    - Long-runner flag at 30d/60d/90d (informational only, no auto-close)
    - No-data days → skip silently, retry tomorrow, STALE_NO_DATA after 30d
"""

# ═════════════════════════════════════════════════════════════════════════════
# Imports & environment
# ═════════════════════════════════════════════════════════════════════════════
import argparse
import os
import sys
import tempfile
import shutil
import numpy as np
from datetime import datetime, date, timedelta

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

try:
    import yfinance as yf
    _YF_OK = True
except ImportError:
    _YF_OK = False
    print("[ERROR] yfinance not installed. Run: pip install yfinance")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False
    print("[ERROR] openpyxl not installed. Run: pip install openpyxl")

try:
    import pandas as pd
    _PD_OK = True
except ImportError:
    _PD_OK = False


# ═════════════════════════════════════════════════════════════════════════════
# Constants — file paths, env flags, thresholds
# ═════════════════════════════════════════════════════════════════════════════
SHADOW_MASTER_XLSX = os.getenv("SHADOW_MASTER_PATH", "shadow_master.xlsx")
SHADOW_ARCHIVE_DIR = os.getenv("SHADOW_ARCHIVE_DIR", "reports/archive")
SHADOW_CSV_PATH    = os.getenv("SHADOW_CSV_PATH", "shadow_trades.csv")
TRACKER_FILE       = os.getenv("TRACKER_FILE", "tracker.json")

# T1 / T2 / Stop levels (hardcoded, matches backtest)
T1_PCT   = float(os.getenv("SHADOW_T1_PCT",   "5.0"))   # +5%
T2_PCT   = float(os.getenv("SHADOW_T2_PCT",   "10.0"))  # +10%
STOP_PCT = float(os.getenv("SHADOW_STOP_PCT", "3.0"))   # -3%

# Long-runner thresholds (informational flags, no auto-close)
LONG_RUNNER_DAYS_1 = int(os.getenv("SHADOW_LONG_RUNNER_1", "30"))
LONG_RUNNER_DAYS_2 = int(os.getenv("SHADOW_LONG_RUNNER_2", "60"))
LONG_RUNNER_DAYS_3 = int(os.getenv("SHADOW_LONG_RUNNER_3", "90"))
STALE_NO_DATA_DAYS = int(os.getenv("SHADOW_STALE_NO_DATA_DAYS", "30"))

# Legacy Recommendations sheet tracking horizon (preserved from tracker_job)
TRACKING_DAYS = int(os.getenv("TRACKING_DAYS", "60"))

# Run-mode gates
IS_SCHEDULED = os.getenv("SCHEDULED_RUN", "false").lower() == "true"
FRESH_START  = os.getenv("FRESH_START",  "false").lower() == "true"

# Bucket taxonomy
BUCKET_SHEETS = {
    "A": "A_TAKEN",
    "B": "B_WATCH_ME",
    "C": "C_NOT_MY_STYLE",
    "D": "D_SO_CLOSE",
}
BUCKET_NAMES = {
    "A": "TAKEN",
    "B": "WATCH_ME",
    "C": "NOT_MY_STYLE",
    "D": "SO_CLOSE",
}
# Backtest expected win rates (from _BUCKET_EXPECTED_WR in shadow_log.py)
BUCKET_EXPECTED_WR = {"A": 48.0, "B": 35.0, "C": 25.0, "D": 40.0}

# Bucket accent colors (C1 decision — user confirmed 2026-07-09)
BUCKET_COLORS = {
    "A": "C6EFCE",   # green (TAKEN)
    "B": "FFEB9C",   # amber (WATCH_ME)
    "C": "D9D9D9",   # gray  (NOT_MY_STYLE)
    "D": "FFCC99",   # orange (SO_CLOSE)
}

# Status row-fill colors (rebuilt every save)
STATUS_FILLS = {
    "OPEN":   "F2F2F2",
    "WIN_T2": "C6EFCE",   # green
    "LOSS":   "FFC7CE",   # red
    "STALE_NO_DATA": "FFF2CC",  # pale yellow
}

# ═════════════════════════════════════════════════════════════════════════════
# The 26-column bucket sheet schema (order matters — used for append + read)
# ═════════════════════════════════════════════════════════════════════════════
BUCKET_COLUMNS = [
    "entry_date",         # 1
    "symbol",             # 2
    "setup_type",         # 3
    "regime",             # 4
    "confidence",         # 5
    "entry_price",        # 6
    "t1_price",           # 7
    "t2_price",           # 8
    "stop_price",         # 9
    "current_price",      # 10
    "current_pnl_pct",    # 11
    "max_favorable_pct",  # 12
    "max_adverse_pct",    # 13
    "days_tracked",       # 14
    "t1_hit_date",        # 15
    "days_to_t1",         # 16
    "t2_hit_date",        # 17
    "days_to_t2",         # 18
    "stop_hit_date",      # 19
    "days_to_stop",       # 20
    "status",             # 21  OPEN / WIN_T2 / LOSS / STALE_NO_DATA
    "long_runner_flag",   # 22  "" / "30d+" / "60d+" / "90d+"
    "exit_date",          # 23
    "exit_price",         # 24
    "final_pnl_pct",      # 25
    "r_multiple",         # 26
]
COL_IDX = {name: i for i, name in enumerate(BUCKET_COLUMNS)}


# ═════════════════════════════════════════════════════════════════════════════
# Utilities
# ═════════════════════════════════════════════════════════════════════════════
def _log(msg: str, quiet: bool = False):
    if not quiet:
        print(msg)


def _scalar(v):
    """Coerce numpy/pandas scalar, 0-d array, or 1-element Series to float."""
    try:
        if hasattr(v, "iloc"):
            v = v.iloc[0]
        if hasattr(v, "item"):
            v = v.item()
        return float(v)
    except Exception:
        return float("nan")


def _today_str() -> str:
    """Return the current run date as YYYY-MM-DD.

    BUG-D fix: honor the SHADOW_RUN_DATE env var so the 30-day integration
    harness (and any future backfill script) can pin the "today" clock
    to a simulated date. Falls back to wall-clock in production.
    """
    override = os.getenv("SHADOW_RUN_DATE", "").strip()
    if override:
        # Validate — if malformed, ignore and fall through to wall-clock.
        try:
            datetime.strptime(override, "%Y-%m-%d")
            return override
        except ValueError:
            pass
    return datetime.now().strftime("%Y-%m-%d")


def _parse_date(s) -> "date | None":
    if not s:
        return None
    try:
        return datetime.strptime(str(s)[:10], "%Y-%m-%d").date()
    except Exception:
        return None


def _fmt_pct(v: float) -> float:
    try:
        return round(float(v), 2)
    except Exception:
        return 0.0


def _fmt_price(v: float) -> float:
    try:
        return round(float(v), 2)
    except Exception:
        return 0.0


def _long_runner_flag(days_tracked: int) -> str:
    if days_tracked >= LONG_RUNNER_DAYS_3:
        return "90d+"
    if days_tracked >= LONG_RUNNER_DAYS_2:
        return "60d+"
    if days_tracked >= LONG_RUNNER_DAYS_1:
        return "30d+"
    return ""


def _archive_dated_copy(src_path: str, archive_dir: str = None) -> str:
    """Copy src to archive dir with today's date in the filename."""
    archive_dir = archive_dir or SHADOW_ARCHIVE_DIR
    try:
        os.makedirs(archive_dir, exist_ok=True)
        stem, ext = os.path.splitext(os.path.basename(src_path))
        dated = f"{stem}_{_today_str()}{ext}"
        dst = os.path.join(archive_dir, dated)
        shutil.copy2(src_path, dst)
        return dst
    except Exception as e:
        _log(f"[WARN] Archive failed: {e}")
        return ""


# ═════════════════════════════════════════════════════════════════════════════
# FRESH_START marker (preserved from tracker_job.py:_detect_fresh_start_marker)
# ═════════════════════════════════════════════════════════════════════════════
def _detect_fresh_start_marker(today_str: str) -> bool:
    """Return True if .fresh_start_marker exists for today; consume it.
    main.py writes this when it wipes state (FRESH_START=true)."""
    marker = ".fresh_start_marker"
    if not os.path.exists(marker):
        return False
    try:
        with open(marker, "r", encoding="utf-8") as _fm:
            marker_date = _fm.read().strip()
    except Exception as e:
        _log(f"[FRESH_START] Could not read {marker}: {e} — ignoring")
        return False
    if marker_date != today_str:
        _log(f"[FRESH_START] Stale marker date={marker_date} (today={today_str}) — removing")
        try:
            os.remove(marker)
        except OSError:
            pass
        return False
    try:
        os.remove(marker)
        _log(f"[FRESH_START] Consumed .fresh_start_marker for {today_str}")
    except OSError as e:
        _log(f"[FRESH_START] Could not delete marker (non-fatal): {e}")
    return True


# ═════════════════════════════════════════════════════════════════════════════
# V1 tracker.json partial-exit stop sync (preserved from tracker_job.py)
# ═════════════════════════════════════════════════════════════════════════════
def _load_partial_exit_stops() -> dict:
    """Return {symbol: {'stop': trailed_stop, 'partial_closed': True, ...}}
    from V1 tracker.json for OPEN positions past T1 or in runner mode."""
    if FRESH_START:
        return {}
    if not os.path.exists(TRACKER_FILE):
        return {}
    try:
        import json as _json
        with open(TRACKER_FILE, "r") as f:
            entries = _json.load(f)
    except Exception as e:
        _log(f"[WARN] Could not read {TRACKER_FILE}: {e}")
        return {}

    if isinstance(entries, dict):
        entries = entries.get("entries") or entries.get("open") or []
    if not isinstance(entries, list):
        return {}

    result = {}
    for e in entries:
        if not isinstance(e, dict):
            continue
        if e.get("status") != "OPEN":
            continue
        if not (e.get("partial_closed") or e.get("runner_active")):
            continue
        sym = str(e.get("symbol", ""))
        if not sym:
            continue
        result[sym] = {
            "stop":           float(e.get("stop", 0) or 0),
            "partial_closed": bool(e.get("partial_closed", False)),
            "runner_active":  bool(e.get("runner_active",  False)),
            "t1_hit_date":    e.get("partial_exit_date") or e.get("t1_hit_date", ""),
            "t2_hit_date":    e.get("t2_hit_date", ""),
        }
    return result


# ═════════════════════════════════════════════════════════════════════════════
# Delivery signal overlay (preserved from tracker_job.py)
# ═════════════════════════════════════════════════════════════════════════════
def _fetch_live_delivery_pct(symbol: str) -> dict:
    """Return {'today': float, '20d_avg': float, 'signal': str} for a symbol.
    Uses main.fetch_delivery_cached (24h TTL, shared cache) if available.
    Returns empty dict on any failure — caller must handle gracefully."""
    try:
        from main import fetch_delivery_cached, load_delivery_cache  # type: ignore
        cache = load_delivery_cache()
        d = fetch_delivery_cached(symbol.replace(".NS", ""), cache)
        if not d or d.get("source") != "nselib":
            return {}
        return {
            "today":   float(d.get("delivery_pct_today", 0.0) or 0.0),
            "20d_avg": float(d.get("delivery_pct_20d_avg", 0.0) or 0.0),
            "ratio":   float(d.get("delivery_ratio", 1.0) or 1.0),
            "signal":  str(d.get("delivery_signal", "NEUTRAL")),
        }
    except Exception:
        return {}


def _flag_position_health(sym: str, cur_return: float, deliv: dict) -> str:
    """Short human-readable health tag."""
    if not deliv:
        return ""
    sig = deliv.get("signal", "NEUTRAL")
    if sig == "DISTRIBUTION":
        if cur_return > 5:
            return "⚠ DISTRIBUTION on winner — consider trimming"
        return "⚠ DISTRIBUTION — institutional selling"
    if sig == "WEAK" and cur_return < -3:
        return "⚠ WEAK delivery on loser — cut / tighten stop"
    if sig == "STRONG_ACCUM" and cur_return > 0:
        return "✓ STRONG ACCUM — let it run"
    return ""


# ═════════════════════════════════════════════════════════════════════════════
# yfinance batch price fetch — same behavior as tracker_job.py
# ═════════════════════════════════════════════════════════════════════════════
def _yf_dry_check(quiet: bool = False) -> bool:
    """Verify yfinance can reach a known-good symbol. Returns True/False."""
    if not _YF_OK:
        return False
    try:
        probe = yf.download(
            "RELIANCE.NS", period="5d", progress=False,
            auto_adjust=False, threads=False,
        )
        ok = probe is not None and not probe.empty and "Close" in probe.columns
    except Exception as e:
        _log(f"[YF_PROBE] exception: {e}", quiet)
        return False
    if not ok:
        _log("[YF_PROBE] RELIANCE.NS returned empty — yfinance down", quiet)
        return False
    _log("[INFO] yfinance dry-check OK — RELIANCE.NS reachable", quiet)
    return True


def _batch_download_prices(symbols: list, quiet: bool = False) -> dict:
    """Download last-5-day OHLC for a list of symbols.
    Returns {symbol: {close, high, low, vol, max_close, min_close}}.
    Missing / failed symbols are silently skipped."""
    prices = {}
    for sym in symbols:
        try:
            df = yf.download(
                sym, period="5d", interval="1d",
                progress=False, auto_adjust=True,
                multi_level_index=False,
            )
            if df is None or len(df) == 0:
                continue
            if hasattr(df.columns, "nlevels") and df.columns.nlevels > 1:
                try:
                    df = df.xs(sym, axis=1, level=-1)
                except Exception:
                    df.columns = df.columns.get_level_values(0)
            close_val = _scalar(df["Close"].iloc[-1])
            high_val  = _scalar(df["High"].iloc[-1])
            low_val   = _scalar(df["Low"].iloc[-1])
            vol_val   = _scalar(df["Volume"].iloc[-1])
            max_close = _scalar(df["Close"].max())
            min_close = _scalar(df["Close"].min())
            if any(np.isnan(x) for x in (close_val, high_val, low_val, max_close, min_close)):
                _log(f"[WARN] Price fetch failed for {sym}: NaN in OHLC", quiet)
                continue
            prices[sym] = {
                "close": close_val,
                "high":  high_val,
                "low":   low_val,
                "vol":   vol_val if not np.isnan(vol_val) else 0.0,
                "max_close": max_close,
                "min_close": min_close,
            }
        except Exception as e:
            _log(f"[WARN] Price fetch failed for {sym}: {e}", quiet)
    return prices


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# Workbook management â€” load, ensure sheets, save
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def _load_or_create_workbook(xlsx_path: str, quiet: bool = False):
    """Load an existing shadow_master.xlsx or create a fresh one with all
    required sheets initialized with the correct headers."""
    if os.path.exists(xlsx_path):
        try:
            wb = openpyxl.load_workbook(xlsx_path)
            _log(f"[INFO] Loaded existing {xlsx_path}", quiet)
        except Exception as e:
            _log(f"[WARN] Could not load {xlsx_path}: {e} â€” creating fresh", quiet)
            wb = openpyxl.Workbook()
            _remove_default_sheet(wb)
    else:
        _log(f"[INFO] {xlsx_path} not found â€” creating fresh workbook", quiet)
        wb = openpyxl.Workbook()
        _remove_default_sheet(wb)

    _ensure_all_sheets(wb)
    return wb


def _remove_default_sheet(wb):
    """Remove openpyxl's default 'Sheet'."""
    if "Sheet" in wb.sheetnames:
        try:
            del wb["Sheet"]
        except Exception:
            pass


def _ensure_all_sheets(wb):
    """Guarantee all bucket sheets + rollups + meta sheets exist with headers."""
    # 4 bucket sheets â€” same 26-column schema
    for key, sheet_name in BUCKET_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            _write_bucket_header(ws, key)

    # Rollup sheets (rebuilt every run â€” created if missing)
    for name in ["Live_Positions", "Resolved_Today", "Summary",
                 "Bucket_Comparison", "Change_Log"]:
        if name not in wb.sheetnames:
            wb.create_sheet(name)

    # Legacy Recommendations sheet (main.py appends to this)
    if "Recommendations" not in wb.sheetnames:
        ws = wb.create_sheet("Recommendations")
        # Header mirrors main.py's daily scanner output
        ws.append([
            "Date", "Ticker", "Setup", "Regime", "Confidence",
            "Entry", "Stop", "T1", "T2", "Status",
        ])
        _style_header_row(ws, fill_color="B4C7E7")

    # Legacy Daily Tracking sheet (for tracker_job compat)
    if "Daily Tracking" not in wb.sheetnames:
        ws = wb.create_sheet("Daily Tracking")
        ws.append([
            "Date", "Ticker", "Rec Date", "Day#",
            "Close", "High", "Low", "Volume",
            "Return%", "Max Gain%", "Max DD%",
            "T1 Hit", "T2 Hit", "Stop Hit", "Remain Up%", "Days Held", "Status",
        ])
        _style_header_row(ws, fill_color="B4C7E7")

    # Performance Summary (rebuilt by _update_performance_sheet)
    if "Performance Summary" not in wb.sheetnames:
        wb.create_sheet("Performance Summary")


def _write_bucket_header(ws, bucket_key: str):
    """Write the 26-column header for a bucket sheet with the bucket accent."""
    ws.append(BUCKET_COLUMNS)
    color = BUCKET_COLORS.get(bucket_key, "D9D9D9")
    _style_header_row(ws, fill_color=color)
    # Column widths for readability
    widths = {
        "A": 12, "B": 14, "C": 14, "D": 12, "E": 10, "F": 10,
        "G": 10, "H": 10, "I": 10, "J": 12, "K": 12, "L": 12, "M": 12,
        "N": 10, "O": 12, "P": 10, "Q": 12, "R": 10, "S": 12, "T": 10,
        "U": 14, "V": 12, "W": 12, "X": 10, "Y": 12, "Z": 10,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"


def _style_header_row(ws, fill_color: str = "B4C7E7"):
    """Bold + colored fill on row 1."""
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color,
                              fill_type="solid")
    align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align


def _apply_status_row_fill(ws, row_idx: int, status: str):
    """Color an entire bucket row based on the status column."""
    color = STATUS_FILLS.get(status)
    if not color:
        return
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for cell in ws[row_idx]:
        cell.fill = fill


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# Bucket sheet operations â€” read, append, update rows
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def _read_bucket_rows(ws) -> list:
    """Return all rows (from row 2 onwards) as list[dict] keyed by column."""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        # Skip completely empty rows
        if all(c.value is None for c in row):
            continue
        d = {}
        for i, col_name in enumerate(BUCKET_COLUMNS):
            d[col_name] = row[i].value if i < len(row) else None
        d["_row_idx"] = row[0].row  # 1-based Excel row index
        rows.append(d)
    return rows


def _row_key(entry_date, symbol) -> str:
    """Unique key for a bucket row."""
    return f"{entry_date}|{symbol}"


def _existing_keys(ws) -> set:
    """Return the set of (entry_date, symbol) keys already in the sheet."""
    keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None or row[1] is None:
            continue
        keys.add(_row_key(row[0], row[1]))
    return keys


def _append_bucket_row(ws, bucket_key: str, signal: dict):
    """Append a new signal row to a bucket sheet with T1/T2/Stop computed."""
    entry_price = _fmt_price(signal.get("entry_price", 0))
    if entry_price <= 0:
        return None
    t1_price   = _fmt_price(entry_price * (1 + T1_PCT / 100.0))
    t2_price   = _fmt_price(entry_price * (1 + T2_PCT / 100.0))
    stop_price = _fmt_price(entry_price * (1 - STOP_PCT / 100.0))
    row = [
        signal.get("entry_date", _today_str()),  # 1
        signal.get("symbol", ""),                # 2
        signal.get("setup_type", "OTHER"),       # 3
        signal.get("regime", "UNKNOWN"),         # 4
        _fmt_pct(signal.get("confidence", 0)),   # 5
        entry_price,                             # 6
        t1_price,                                # 7
        t2_price,                                # 8
        stop_price,                              # 9
        entry_price,                             # 10 current_price (init)
        0.0,                                     # 11 current_pnl_pct
        0.0,                                     # 12 max_favorable_pct
        0.0,                                     # 13 max_adverse_pct
        0,                                       # 14 days_tracked
        "",                                      # 15 t1_hit_date
        "",                                      # 16 days_to_t1
        "",                                      # 17 t2_hit_date
        "",                                      # 18 days_to_t2
        "",                                      # 19 stop_hit_date
        "",                                      # 20 days_to_stop
        "OPEN",                                  # 21 status
        "",                                      # 22 long_runner_flag
        "",                                      # 23 exit_date
        "",                                      # 24 exit_price
        "",                                      # 25 final_pnl_pct
        "",                                      # 26 r_multiple
    ]
    ws.append(row)
    return ws.max_row


def _update_bucket_row(ws, row_idx: int, updates: dict):
    """Update specific columns in a bucket row (1-based row_idx)."""
    for col_name, value in updates.items():
        col_num = COL_IDX.get(col_name)
        if col_num is None:
            continue
        ws.cell(row=row_idx, column=col_num + 1, value=value)


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# The core update engine â€” walks every OPEN row across all 4 buckets,
# fetches today's price, checks T1/T2/Stop, updates row state.
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def _update_open_positions(wb, today_str: str, quiet: bool = False) -> dict:
    """Update every OPEN row across all 4 bucket sheets.

    For each OPEN row:
        - fetch today's OHLC
        - update current_price, current_pnl_pct, MFE, MAE, days_tracked
        - if high >= T2 â†’ close as WIN_T2
        - if high >= T1 (still open) â†’ set t1_hit_date / days_to_t1
        - if low  <= Stop â†’ close as LOSS
        - update long_runner_flag if crossed a 30d/60d/90d threshold

    Returns statistics dict.
    """
    stats = {
        "n_rows_updated": 0,
        "n_t1_hit_today": 0,
        "n_t2_hit_today": 0,
        "n_stop_hit_today": 0,
        "n_stale_no_data": 0,
        "resolved_today": [],  # list of (bucket, symbol, status)
    }

    # Collect all OPEN symbols across all buckets (for one batched yf.download)
    open_rows_by_symbol = {}  # {symbol: [(bucket_key, ws, row_idx, row_dict), ...]}
    for bucket_key, sheet_name in BUCKET_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row_dict in _read_bucket_rows(ws):
            status = row_dict.get("status", "OPEN")
            if status not in ("OPEN", "STALE_NO_DATA"):
                continue
            sym = str(row_dict.get("symbol", ""))
            if not sym:
                continue
            open_rows_by_symbol.setdefault(sym, []).append(
                (bucket_key, ws, row_dict["_row_idx"], row_dict)
            )

    if not open_rows_by_symbol:
        _log("[INFO] No OPEN rows to update across all 4 buckets", quiet)
        return stats

    symbols = list(open_rows_by_symbol.keys())
    _log(f"[INFO] Fetching prices for {len(symbols)} unique open symbols "
         f"across all buckets", quiet)
    prices = _batch_download_prices(symbols, quiet=quiet)
    _log(f"[INFO] Got prices for {len(prices)}/{len(symbols)} symbols", quiet)

    # Walk each row and apply updates
    for sym, rows in open_rows_by_symbol.items():
        px = prices.get(sym)
        for bucket_key, ws, row_idx, row_dict in rows:
            entry_date_str = str(row_dict.get("entry_date", ""))
            entry_price = float(row_dict.get("entry_price") or 0)
            t1_price    = float(row_dict.get("t1_price")    or 0)
            t2_price    = float(row_dict.get("t2_price")    or 0)
            stop_price  = float(row_dict.get("stop_price")  or 0)
            prev_mfe    = float(row_dict.get("max_favorable_pct") or 0)
            prev_mae    = float(row_dict.get("max_adverse_pct")   or 0)
            prev_days   = int(row_dict.get("days_tracked") or 0)

            entry_dt = _parse_date(entry_date_str)
            today_dt = _parse_date(today_str)
            if entry_dt and today_dt:
                days_tracked = (today_dt - entry_dt).days
            else:
                days_tracked = prev_days + 1

            # Handle missing price data
            if not px:
                new_days = prev_days + 1
                new_status = row_dict.get("status", "OPEN")
                # Mark STALE_NO_DATA if we've missed too many consecutive days
                if new_days - prev_days >= STALE_NO_DATA_DAYS:
                    new_status = "STALE_NO_DATA"
                    stats["n_stale_no_data"] += 1
                _update_bucket_row(ws, row_idx, {
                    "days_tracked": days_tracked,
                    "status": new_status,
                    "long_runner_flag": _long_runner_flag(days_tracked),
                })
                continue

            # We have price data â€” do full update
            cur_close = px["close"]
            cur_high  = px["high"]
            cur_low   = px["low"]

            cur_pnl_pct = round((cur_close - entry_price) / entry_price * 100, 2) \
                if entry_price > 0 else 0.0

            # MFE / MAE ratchets â€” never regress
            bar_favorable = ((cur_high - entry_price) / entry_price * 100.0) \
                if entry_price > 0 else 0.0
            bar_adverse   = ((cur_low  - entry_price) / entry_price * 100.0) \
                if entry_price > 0 else 0.0
            new_mfe = round(max(prev_mfe, bar_favorable), 2)
            new_mae = round(min(prev_mae, bar_adverse), 2)

            # Detect hits
            t1_hit_today   = t1_price > 0 and cur_high >= t1_price
            t2_hit_today   = t2_price > 0 and cur_high >= t2_price
            stop_hit_today = stop_price > 0 and cur_low  <= stop_price

            # T1 milestone flag (only if not already hit)
            existing_t1_date = row_dict.get("t1_hit_date") or ""
            updates = {
                "current_price":     _fmt_price(cur_close),
                "current_pnl_pct":   cur_pnl_pct,
                "max_favorable_pct": new_mfe,
                "max_adverse_pct":   new_mae,
                "days_tracked":      days_tracked,
                "long_runner_flag":  _long_runner_flag(days_tracked),
            }

            if t1_hit_today and not existing_t1_date:
                updates["t1_hit_date"] = today_str
                updates["days_to_t1"]  = days_tracked
                stats["n_t1_hit_today"] += 1

            # Terminal exits â€” first hit wins (T2 > Stop precedence within a bar
            # is ambiguous. Convention: if high hit T2, treat as WIN_T2.
            # Otherwise if low hit stop, treat as LOSS.)
            if t2_hit_today:
                updates["t2_hit_date"] = today_str
                updates["days_to_t2"]  = days_tracked
                updates["status"]      = "WIN_T2"
                updates["exit_date"]   = today_str
                updates["exit_price"]  = _fmt_price(t2_price)
                updates["final_pnl_pct"] = T2_PCT
                # R-multiple = final_pnl / stop_distance_pct
                updates["r_multiple"]    = round(T2_PCT / STOP_PCT, 2)
                stats["n_t2_hit_today"] += 1
                stats["resolved_today"].append(
                    (bucket_key, sym, "WIN_T2"))
            elif stop_hit_today:
                updates["stop_hit_date"] = today_str
                updates["days_to_stop"]  = days_tracked
                updates["status"]        = "LOSS"
                updates["exit_date"]     = today_str
                updates["exit_price"]    = _fmt_price(stop_price)
                updates["final_pnl_pct"] = -STOP_PCT
                updates["r_multiple"]    = -1.0
                stats["n_stop_hit_today"] += 1
                stats["resolved_today"].append(
                    (bucket_key, sym, "LOSS"))

            _update_bucket_row(ws, row_idx, updates)
            # Row color per status
            _apply_status_row_fill(ws, row_idx, updates.get("status", "OPEN"))
            stats["n_rows_updated"] += 1

    _log(f"[INFO] Updated {stats['n_rows_updated']} rows Â· "
         f"T1 today={stats['n_t1_hit_today']} Â· "
         f"T2 today={stats['n_t2_hit_today']} Â· "
         f"Stop today={stats['n_stop_hit_today']} Â· "
         f"Stale={stats['n_stale_no_data']}", quiet)
    return stats


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# Append new signals from shadow_trades.csv into their bucket sheets
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def _append_new_signals_from_csv(wb, quiet: bool = False) -> int:
    """Read shadow_trades.csv, find any rows dated today not already in the
    corresponding bucket sheet, and append them.

    shadow_trades.csv schema (from shadow_log.py):
        entry_date, bucket, symbol, setup_type, regime, confidence,
        close_price, ... (many more columns, we only need these)

    Returns count of rows appended.
    """
    if not os.path.exists(SHADOW_CSV_PATH):
        _log(f"[INFO] {SHADOW_CSV_PATH} not found â€” no new signals to append", quiet)
        return 0
    if not _PD_OK:
        _log(f"[WARN] pandas not available â€” cannot read CSV", quiet)
        return 0

    try:
        df = pd.read_csv(SHADOW_CSV_PATH)
    except Exception as e:
        _log(f"[WARN] Could not read {SHADOW_CSV_PATH}: {e}", quiet)
        return 0

    if df.empty:
        return 0

    # Normalize column names â€” CSV may use various forms
    col_map = {c.lower(): c for c in df.columns}
    def _col(name, default=None):
        return col_map.get(name.lower(), default)

    date_col     = _col("entry_date") or _col("date") or _col("timestamp")
    bucket_col   = _col("bucket") or _col("bucket_key") or _col("bucket_id")
    sym_col      = _col("symbol") or _col("ticker")
    setup_col    = _col("setup_type") or _col("setup")
    regime_col   = _col("regime") or _col("market_regime")
    conf_col     = _col("confidence") or _col("conf")
    price_col    = _col("entry_price") or _col("close_price") or _col("close")

    if not (date_col and bucket_col and sym_col and price_col):
        _log(f"[WARN] shadow CSV missing required columns "
             f"(date/bucket/symbol/price)", quiet)
        return 0

    appended = 0
    for bucket_key, sheet_name in BUCKET_SHEETS.items():
        ws = wb[sheet_name]
        existing = _existing_keys(ws)
        # Rows for this bucket
        bucket_letter = bucket_key.upper()
        df_bucket = df[df[bucket_col].astype(str).str.upper() == bucket_letter]
        for _, r in df_bucket.iterrows():
            entry_date = str(r[date_col])[:10]
            symbol     = str(r[sym_col])
            key = _row_key(entry_date, symbol)
            if key in existing:
                continue
            signal = {
                "entry_date":  entry_date,
                "symbol":      symbol,
                "setup_type":  str(r.get(setup_col, "OTHER")) if setup_col else "OTHER",
                "regime":      str(r.get(regime_col, "UNKNOWN")) if regime_col else "UNKNOWN",
                "confidence":  float(r.get(conf_col, 0)) if conf_col else 0.0,
                "entry_price": float(r[price_col] or 0),
            }
            row_idx = _append_bucket_row(ws, bucket_key, signal)
            if row_idx:
                _apply_status_row_fill(ws, row_idx, "OPEN")
                existing.add(key)
                appended += 1

    _log(f"[INFO] Appended {appended} new signal rows across all buckets", quiet)
    return appended


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# Rollup sheets â€” rebuilt from scratch every run
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def _clear_sheet(ws):
    """Clear ALL rows (BUG-B fix: actually delete rows so max_row shrinks).

    The previous implementation only nulled cell.value, which leaves
    ws.max_row pointing at the phantom bottom row. Subsequent `ws.append`
    then writes past those blanks, so the sheet grew by N-empty-rows on
    every rebuild (Summary +5/day, Bucket_Comparison +10/day, etc).

    ws.delete_rows(1, ws.max_row) reliably shrinks the sheet back to 0
    rows across every openpyxl version we care about.
    """
    n = ws.max_row
    if n and n > 0:
        try:
            ws.delete_rows(1, n)
        except Exception:
            # Fallback: clear values + fills the old way (should not happen)
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None
                    cell.fill = PatternFill(fill_type=None)


def _rebuild_summary_sheet(wb, quiet: bool = False):
    """Recompute per-bucket rollup statistics."""
    ws = wb["Summary"]
    _clear_sheet(ws)

    headers = [
        "Bucket", "Name", "N Total", "N Open", "N Closed",
        "N WIN_T2", "N LOSS", "Win Rate %", "Expected WR %", "Delta pp",
        "Avg Days to Win", "Avg Days to Loss", "Avg MFE %", "Avg MAE %",
        "Verdict",
    ]
    ws.append(headers)
    _style_header_row(ws, fill_color="B4C7E7")

    for bucket_key, sheet_name in BUCKET_SHEETS.items():
        rows = _read_bucket_rows(wb[sheet_name])
        n_total  = len(rows)
        n_open   = sum(1 for r in rows if r.get("status") == "OPEN")
        n_win    = sum(1 for r in rows if r.get("status") == "WIN_T2")
        n_loss   = sum(1 for r in rows if r.get("status") == "LOSS")
        n_closed = n_win + n_loss

        win_rate = round(n_win / n_closed * 100, 1) if n_closed > 0 else 0.0
        expected = BUCKET_EXPECTED_WR.get(bucket_key, 0.0)
        delta_pp = round(win_rate - expected, 1) if n_closed > 0 else 0.0

        wins  = [r for r in rows if r.get("status") == "WIN_T2"]
        losses = [r for r in rows if r.get("status") == "LOSS"]
        avg_days_win = round(np.mean([float(r.get("days_to_t2") or 0) for r in wins]), 1) \
            if wins else 0.0
        avg_days_loss = round(np.mean([float(r.get("days_to_stop") or 0) for r in losses]), 1) \
            if losses else 0.0
        closed = wins + losses
        avg_mfe = round(np.mean([float(r.get("max_favorable_pct") or 0) for r in closed]), 2) \
            if closed else 0.0
        avg_mae = round(np.mean([float(r.get("max_adverse_pct") or 0) for r in closed]), 2) \
            if closed else 0.0

        # Verdict logic â€” only if we have enough data
        if n_closed < 20:
            verdict = "insufficient data"
        elif delta_pp >= 5:
            verdict = "ðŸš€ OUTPERFORMING backtest"
        elif delta_pp <= -5:
            verdict = "â›” UNDERPERFORMING backtest"
        else:
            verdict = "âœ… MATCHES backtest"

        ws.append([
            bucket_key, BUCKET_NAMES[bucket_key],
            n_total, n_open, n_closed,
            n_win, n_loss, win_rate, expected, delta_pp,
            avg_days_win, avg_days_loss, avg_mfe, avg_mae,
            verdict,
        ])
        # Row color per bucket
        color = BUCKET_COLORS.get(bucket_key)
        if color:
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for cell in ws[ws.max_row]:
                cell.fill = fill

    # Column widths
    for col_letter, width in {"A": 8, "B": 15, "O": 30}.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"


def _rebuild_bucket_comparison(wb, quiet: bool = False):
    """The money sheet â€” side-by-side bucket verdict."""
    ws = wb["Bucket_Comparison"]
    _clear_sheet(ws)

    ws.append(["", "A_TAKEN", "B_WATCH_ME", "C_NOT_MY_STYLE", "D_SO_CLOSE"])
    _style_header_row(ws, fill_color="B4C7E7")
    # Color bucket header cells
    for i, bk in enumerate("ABCD", start=2):
        color = BUCKET_COLORS[bk]
        ws.cell(row=1, column=i).fill = PatternFill(
            start_color=color, end_color=color, fill_type="solid")

    # Rows: metric labels
    metrics = ["N Closed", "Win Rate %", "Expected WR %", "Delta pp",
               "Avg Days to Win", "Avg Days to Loss",
               "Avg MFE %", "Avg MAE %", "Verdict"]

    # Compute per-bucket stats
    stats_by_bucket = {}
    for bucket_key, sheet_name in BUCKET_SHEETS.items():
        rows = _read_bucket_rows(wb[sheet_name])
        n_win   = sum(1 for r in rows if r.get("status") == "WIN_T2")
        n_loss  = sum(1 for r in rows if r.get("status") == "LOSS")
        n_closed = n_win + n_loss
        wr    = round(n_win / n_closed * 100, 1) if n_closed > 0 else 0.0
        exp   = BUCKET_EXPECTED_WR.get(bucket_key, 0.0)
        delta = round(wr - exp, 1) if n_closed > 0 else 0.0
        wins   = [r for r in rows if r.get("status") == "WIN_T2"]
        losses = [r for r in rows if r.get("status") == "LOSS"]
        adw = round(np.mean([float(r.get("days_to_t2") or 0) for r in wins]), 1) \
            if wins else 0.0
        adl = round(np.mean([float(r.get("days_to_stop") or 0) for r in losses]), 1) \
            if losses else 0.0
        closed = wins + losses
        mfe = round(np.mean([float(r.get("max_favorable_pct") or 0) for r in closed]), 2) \
            if closed else 0.0
        mae = round(np.mean([float(r.get("max_adverse_pct") or 0) for r in closed]), 2) \
            if closed else 0.0
        if n_closed < 20:
            verdict = "insufficient"
        elif delta >= 5:
            verdict = "ðŸš€ BEAT backtest"
        elif delta <= -5:
            verdict = "â›” MISSED backtest"
        else:
            verdict = "âœ… AS EXPECTED"
        stats_by_bucket[bucket_key] = [
            n_closed, wr, exp, delta, adw, adl, mfe, mae, verdict
        ]

    for i, metric in enumerate(metrics):
        row = [metric]
        for bk in "ABCD":
            row.append(stats_by_bucket[bk][i])
        ws.append(row)

    ws.column_dimensions["A"].width = 22
    for col_letter in ["B", "C", "D", "E"]:
        ws.column_dimensions[col_letter].width = 18
    ws.freeze_panes = "B2"


def _rebuild_live_positions(wb, quiet: bool = False):
    """Aggregate all OPEN rows across the 4 bucket sheets into one view."""
    ws = wb["Live_Positions"]
    _clear_sheet(ws)

    headers = ["bucket"] + BUCKET_COLUMNS
    ws.append(headers)
    _style_header_row(ws, fill_color="B4C7E7")

    all_opens = []
    for bucket_key, sheet_name in BUCKET_SHEETS.items():
        rows = _read_bucket_rows(wb[sheet_name])
        for r in rows:
            if r.get("status") in ("OPEN", "STALE_NO_DATA"):
                all_opens.append((bucket_key, r))

    # Sort by days_tracked desc â€” oldest at top
    all_opens.sort(key=lambda t: int(t[1].get("days_tracked") or 0), reverse=True)

    for bucket_key, r in all_opens:
        row = [bucket_key] + [r.get(col) for col in BUCKET_COLUMNS]
        ws.append(row)
        # Color by bucket
        color = BUCKET_COLORS[bucket_key]
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.cell(row=ws.max_row, column=1).fill = fill

    ws.column_dimensions["A"].width = 8
    ws.freeze_panes = "A2"


def _rebuild_resolved_today(wb, resolved_list: list, today_str: str):
    """List the WIN_T2 / LOSS flips that happened in this run."""
    ws = wb["Resolved_Today"]
    _clear_sheet(ws)

    ws.append(["run_date", "bucket", "symbol", "status",
               "setup_type", "regime", "entry_price", "exit_price",
               "final_pnl_pct", "days_to_exit", "r_multiple"])
    _style_header_row(ws, fill_color="B4C7E7")

    # Pull details from the bucket sheets for each resolved row
    for bucket_key, sym, status in resolved_list:
        sheet_name = BUCKET_SHEETS[bucket_key]
        for r in _read_bucket_rows(wb[sheet_name]):
            if r.get("symbol") == sym and r.get("status") == status \
                    and r.get("exit_date") == today_str:
                ws.append([
                    today_str, bucket_key, sym, status,
                    r.get("setup_type"), r.get("regime"),
                    r.get("entry_price"), r.get("exit_price"),
                    r.get("final_pnl_pct"),
                    r.get("days_to_t2") if status == "WIN_T2" else r.get("days_to_stop"),
                    r.get("r_multiple"),
                ])
                # Row color
                fill_color = STATUS_FILLS.get(status, "FFFFFF")
                fill = PatternFill(start_color=fill_color, end_color=fill_color,
                                   fill_type="solid")
                for cell in ws[ws.max_row]:
                    cell.fill = fill
                break

    ws.freeze_panes = "A2"


def _append_change_log(wb, mode: str, stats: dict, appended: int,
                       is_scheduled: bool, run_date: str | None = None):
    """Append one row to Change_Log recording what this run did.

    BUG-A fix: only write the header on a truly empty sheet. The previous
    guard used `ws.cell(row=1, column=1).value is None` which was always
    true after `_clear_sheet` blanked A1 to None, so a fresh header was
    appended before every data row (30 dup headers over 30 runs).
    We now check for the literal "timestamp" sentinel in A1.

    BUG-D fix: use the caller-supplied `run_date` for the timestamp when
    provided (harness/backfill), else fall back to wall-clock.
    """
    ws = wb["Change_Log"]
    a1 = ws.cell(row=1, column=1).value if ws.max_row >= 1 else None
    if a1 != "timestamp":
        # Header missing or corrupt — rewrite from the top.
        # (Delete anything currently in row 1 to avoid the "blank+header" case.)
        if ws.max_row >= 1 and a1 is None:
            try:
                ws.delete_rows(1, 1)
            except Exception:
                pass
        ws.append(["timestamp", "mode", "scheduled",
                   "n_appended", "n_rows_updated",
                   "n_t1_hit", "n_t2_hit", "n_stop_hit", "n_stale",
                   "n_resolved"])
        _style_header_row(ws, fill_color="B4C7E7")

    # BUG-D: timestamp reflects the simulated run_date when supplied.
    if run_date:
        ts = f"{run_date} {datetime.now().strftime('%H:%M:%S')}"
    else:
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws.append([
        ts,
        mode,
        "yes" if is_scheduled else "no (preview)",
        appended,
        stats.get("n_rows_updated", 0),
        stats.get("n_t1_hit_today", 0),
        stats.get("n_t2_hit_today", 0),
        stats.get("n_stop_hit_today", 0),
        stats.get("n_stale_no_data", 0),
        len(stats.get("resolved_today", [])),
    ])


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# Legacy Performance Summary / Portfolio Risk / Benchmark / Equity Curve
# Preserved from tracker_job.py verbatim (they read "Daily Tracking" sheet)
# These continue to work exactly as before for Bucket A real BUY tracking.
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def _update_performance_sheet(wb, run_date: str | None = None):
    """Recalculates Performance Summary sheet from Daily Tracking data.
    Preserved from tracker_job.py (unchanged logic).

    BUG-D fix: honor a caller-supplied run_date for the Last Updated cell.
    """
    try:
        if "Daily Tracking" not in wb.sheetnames:
            return
        if "Performance Summary" not in wb.sheetnames:
            wb.create_sheet("Performance Summary")
        ws_track = wb["Daily Tracking"]
        ws_perf  = wb["Performance Summary"]

        # Clear existing (keep header)
        for row in ws_perf.iter_rows(min_row=2):
            for cell in row:
                cell.value = None

        headers = [cell.value for cell in ws_track[1]]
        records = [dict(zip(headers, [cell.value for cell in row]))
                   for row in ws_track.iter_rows(min_row=2)]

        outcomes = {}
        for r in records:
            key = f"{r.get('Ticker')}_{r.get('Rec Date')}"
            day = int(r.get("Day#") or 0)
            if key not in outcomes or day > int(outcomes[key].get("Day#") or 0):
                outcomes[key] = r

        closed = [o for o in outcomes.values()
                  if o.get("Status") not in ("ACTIVE", "T1_HIT_ACTIVE", "RUNNER")]
        wins   = [o for o in closed if float(o.get("Return%") or 0) > 0]
        losses = [o for o in closed if float(o.get("Return%") or 0) <= 0]

        runners_active = [o for o in outcomes.values() if o.get("Status") == "RUNNER"]
        runners_closed = [o for o in closed if o.get("Status") == "RUNNER_STOPPED"]
        t2_hit_closed  = [o for o in closed if o.get("Status") == "T2_HIT"]

        stats = [
            ("Total Tracked",   len(outcomes)),
            ("Closed",          len(closed)),
            ("Active",          len(outcomes) - len(closed)),
            ("Runners Active",  len(runners_active)),
            ("Win Rate %",      round(len(wins)/len(closed)*100, 1) if closed else 0),
            ("Avg Return %",    round(np.mean([float(o.get("Return%") or 0) for o in closed]), 2) if closed else 0),
            ("Avg Win %",       round(np.mean([float(o.get("Return%") or 0) for o in wins]), 2) if wins else 0),
            ("Avg Loss %",      round(np.mean([float(o.get("Return%") or 0) for o in losses]), 2) if losses else 0),
            ("Avg Runner Ret%", round(np.mean([float(o.get("Return%") or 0) for o in runners_closed]), 2) if runners_closed else 0),
            ("Avg T2-Exit Ret%",round(np.mean([float(o.get("Return%") or 0) for o in t2_hit_closed]), 2) if t2_hit_closed else 0),
            ("Avg Max Gain %",  round(np.mean([float(o.get("Max Gain%") or 0) for o in closed]), 2) if closed else 0),
            ("Avg Max DD %",    round(np.mean([float(o.get("Max DD%") or 0) for o in closed]), 2) if closed else 0),
            ("T1 Hit Rate %",   round(sum(1 for o in closed if o.get("T1 Hit"))/len(closed)*100, 1) if closed else 0),
            ("T2 Hit Rate %",   round(sum(1 for o in closed if o.get("T2 Hit"))/len(closed)*100, 1) if closed else 0),
            ("Stop Hit Rate %", round(sum(1 for o in closed if o.get("Stop Hit"))/len(closed)*100, 1) if closed else 0),
            ("Last Updated",    (f"{run_date} " + datetime.now().strftime("%H:%M")) if run_date else datetime.now().strftime("%Y-%m-%d %H:%M")),
        ]

        ws_perf["A1"] = "Metric"
        ws_perf["B1"] = "Value"
        for i, (metric, value) in enumerate(stats, start=2):
            ws_perf[f"A{i}"] = metric
            ws_perf[f"B{i}"] = value
    except Exception as e:
        print(f"[WARN] Performance sheet update failed: {e}")


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# PUBLIC ENTRY POINTS â€” the 3 modes
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def run_scan_and_update(quiet: bool = False) -> dict:
    """Evening full flow:
        - append today's new shadow signals to bucket sheets (A/B/C/D)
        - update prices for all OPEN rows
        - rebuild rollups (Summary, Live_Positions, Bucket_Comparison,
          Resolved_Today, Change_Log)
        - update legacy Performance Summary
        - save + archive (if scheduled) or write temp preview (if manual)

    Returns dict with: ok, mode, xlsx_path, archived_path, is_scheduled,
    n_appended, stats.
    """
    today_str = _today_str()
    is_scheduled = IS_SCHEDULED
    _log(f"=== SHADOW MASTER Â· scan-and-update Â· {today_str} Â· "
         f"mode={'SCHEDULED' if is_scheduled else 'PREVIEW (manual)'} ===", quiet)

    # Guard: FRESH_START marker
    if FRESH_START or _detect_fresh_start_marker(today_str):
        _log("[FRESH_START] State wiped this run â€” skipping shadow master update", quiet)
        return {"ok": True, "skipped": True, "reason": "fresh-start"}

    # Guard: dependencies
    if not _YF_OK or not _OPENPYXL_OK:
        return {"ok": False, "error": "missing-deps"}

    # Guard: yf dry-check
    if not _yf_dry_check(quiet=quiet):
        _log("[ERROR] yfinance dry-check FAILED â€” aborting to protect data", quiet)
        try:
            with open("yfinance_down.flag", "w", encoding="utf-8") as _f:
                _f.write(f"yfinance dry-check failed at "
                         f"{datetime.now().isoformat()}\n")
        except OSError:
            pass
        return {"ok": False, "error": "yfinance-down"}

    # Determine target xlsx path â€” real file if scheduled, temp if manual
    real_path = SHADOW_MASTER_XLSX
    if is_scheduled:
        target_path = real_path
    else:
        target_path = os.path.join(
            tempfile.gettempdir(),
            f"shadow_master_preview_{today_str}.xlsx",
        )
        # If real file exists, copy it to temp first so we mutate a copy
        if os.path.exists(real_path):
            try:
                shutil.copy2(real_path, target_path)
            except Exception as e:
                _log(f"[WARN] Could not copy real -> temp: {e}", quiet)

    # Load / create workbook
    wb = _load_or_create_workbook(target_path, quiet=quiet)

    # Append today's new shadow signals
    n_appended = _append_new_signals_from_csv(wb, quiet=quiet)

    # Update all OPEN rows across all buckets
    stats = _update_open_positions(wb, today_str, quiet=quiet)

    # Rebuild rollups
    _rebuild_summary_sheet(wb, quiet=quiet)
    _rebuild_bucket_comparison(wb, quiet=quiet)
    _rebuild_live_positions(wb, quiet=quiet)
    _rebuild_resolved_today(wb, stats.get("resolved_today", []), today_str)
    _append_change_log(wb, "scan-and-update", stats, n_appended, is_scheduled, run_date=today_str)

    # Legacy Performance Summary (for real BUY signals in Daily Tracking)
    _update_performance_sheet(wb, run_date=today_str)

    # Save
    try:
        wb.save(target_path)
        _log(f"[INFO] Saved {target_path}", quiet)
    except Exception as e:
        _log(f"[ERROR] Save failed: {e}", quiet)
        return {"ok": False, "error": f"save-failed: {e}"}

    # Archive (scheduled only)
    archived_path = None
    if is_scheduled:
        archived_path = _archive_dated_copy(target_path)
        if archived_path:
            _log(f"[INFO] Archived â†’ {archived_path}", quiet)

    return {
        "ok": True,
        "mode": "SCHEDULED" if is_scheduled else "PREVIEW",
        "xlsx_path": target_path,
        "archived_path": archived_path,
        "is_scheduled": is_scheduled,
        "n_appended": n_appended,
        "stats": stats,
        "real_path": real_path,
    }


def run_update_only(quiet: bool = False) -> dict:
    """Midday flow: price updates only, no new signal appends.
    Silent unless something hit today."""
    today_str = _today_str()
    is_scheduled = IS_SCHEDULED
    _log(f"=== SHADOW MASTER Â· update-only Â· {today_str} Â· "
         f"mode={'SCHEDULED' if is_scheduled else 'PREVIEW (manual)'} ===", quiet)

    if FRESH_START or _detect_fresh_start_marker(today_str):
        _log("[FRESH_START] Skipping midday update", quiet)
        return {"ok": True, "skipped": True, "reason": "fresh-start"}

    if not _YF_OK or not _OPENPYXL_OK:
        return {"ok": False, "error": "missing-deps"}

    if not _yf_dry_check(quiet=quiet):
        return {"ok": False, "error": "yfinance-down"}

    real_path = SHADOW_MASTER_XLSX
    if not os.path.exists(real_path):
        _log(f"[INFO] No {real_path} yet â€” nothing to update", quiet)
        return {"ok": True, "skipped": True, "reason": "no-master-file"}

    if is_scheduled:
        target_path = real_path
    else:
        target_path = os.path.join(
            tempfile.gettempdir(),
            f"shadow_master_preview_{today_str}.xlsx",
        )
        shutil.copy2(real_path, target_path)

    wb = _load_or_create_workbook(target_path, quiet=quiet)
    stats = _update_open_positions(wb, today_str, quiet=quiet)
    _rebuild_summary_sheet(wb, quiet=quiet)
    _rebuild_bucket_comparison(wb, quiet=quiet)
    _rebuild_live_positions(wb, quiet=quiet)
    _rebuild_resolved_today(wb, stats.get("resolved_today", []), today_str)
    _append_change_log(wb, "update-only", stats, 0, is_scheduled, run_date=today_str)

    try:
        wb.save(target_path)
    except Exception as e:
        return {"ok": False, "error": f"save-failed: {e}"}

    return {
        "ok": True,
        "mode": "SCHEDULED" if is_scheduled else "PREVIEW",
        "xlsx_path": target_path,
        "is_scheduled": is_scheduled,
        "stats": stats,
        "real_path": real_path,
    }


def run_research(quiet: bool = False) -> dict:
    """Weekly fundamentals â€” delegates to research_job.run_research().
    Kept as a thin wrapper so the full ~1700 lines of research logic is
    preserved by reference until we do a formal port."""
    _log("=== SHADOW MASTER Â· research (delegating to research_job) ===", quiet)
    try:
        import research_job
        # research_job reads TRACKER_XLSX env var â€” retarget it to shadow_master
        os.environ["TRACKER_XLSX"] = SHADOW_MASTER_XLSX
        research_job.run_research()
        return {"ok": True, "mode": "research", "xlsx_path": SHADOW_MASTER_XLSX}
    except Exception as e:
        _log(f"[ERROR] research delegate failed: {e}", quiet)
        return {"ok": False, "error": str(e)}


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# Telegram delivery helper
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def _build_caption(result: dict) -> str:
    """Build the Telegram caption based on run result."""
    if result.get("skipped"):
        return f"â¸ Shadow Master skipped: {result.get('reason', 'unknown')}"

    mode = result.get("mode", "?")
    stats = result.get("stats", {})
    n_appended = result.get("n_appended", 0)
    is_sched = result.get("is_scheduled", False)

    lines = []
    icon = "ðŸ“Š" if is_sched else "ðŸ‘"
    label = "SCHEDULED Â· saved + archived" if is_sched else "PREVIEW Â· not saved"
    lines.append(f"{icon} Shadow Master Report â€” {_today_str()}")
    lines.append(f"Mode: {mode} Â· {label}")
    if n_appended:
        lines.append(f"New signals appended: {n_appended}")
    lines.append(f"Rows updated: {stats.get('n_rows_updated', 0)}")
    resolved = stats.get("resolved_today", [])
    n_win = sum(1 for _, _, s in resolved if s == "WIN_T2")
    n_loss = sum(1 for _, _, s in resolved if s == "LOSS")
    if resolved:
        lines.append(f"Resolved today: {len(resolved)} "
                     f"({n_win} WIN_T2, {n_loss} LOSS)")
    else:
        lines.append("Resolved today: none")
    if not is_sched:
        lines.append("âš ï¸ Master xlsx unchanged. Trigger scheduled run to persist.")
    return "\n".join(lines)


def _maybe_send_telegram(result: dict, quiet: bool = False):
    """Send the xlsx as a Telegram document. In PREVIEW mode, also delete
    the temp file after send. No-op if the file is missing."""
    if not result.get("ok"):
        return
    if result.get("skipped"):
        return
    xlsx_path = result.get("xlsx_path")
    if not xlsx_path or not os.path.exists(xlsx_path):
        return

    try:
        from scripts.notify_telegram_document import send_document
    except Exception as e:
        _log(f"[WARN] Cannot import notify_telegram_document: {e}", quiet)
        return

    caption = _build_caption(result)
    try:
        rc = send_document(file_path=xlsx_path, caption=caption)
        _log(f"[INFO] Telegram send rc={rc}", quiet)
    except Exception as e:
        _log(f"[WARN] Telegram send failed: {e}", quiet)

    # Cleanup temp preview
    if result.get("mode") == "PREVIEW":
        try:
            os.unlink(xlsx_path)
        except OSError:
            pass


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# CLI entry
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--mode", choices=["scan-and-update", "update-only", "research"],
        default="scan-and-update",
        help="scan-and-update: full evening flow Â· "
             "update-only: midday price refresh Â· "
             "research: delegate to research_job",
    )
    parser.add_argument(
        "--quiet", action="store_true",
        help="Suppress info logging (still emits WARN/ERROR)",
    )
    parser.add_argument(
        "--no-telegram", action="store_true",
        help="Build xlsx but do not send to Telegram",
    )
    args = parser.parse_args()

    if args.mode == "scan-and-update":
        result = run_scan_and_update(quiet=args.quiet)
    elif args.mode == "update-only":
        result = run_update_only(quiet=args.quiet)
    elif args.mode == "research":
        result = run_research(quiet=args.quiet)
    else:
        print(f"[ERROR] Unknown mode: {args.mode}")
        sys.exit(2)

    if not args.no_telegram:
        _maybe_send_telegram(result, quiet=args.quiet)

    if not result.get("ok"):
        print(f"[ERROR] Run failed: {result}")
        sys.exit(1)

    if not args.quiet:
        print(f"[DONE] {result}")


if __name__ == "__main__":
    main()
